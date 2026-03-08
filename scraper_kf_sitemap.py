#!/usr/bin/env python3
"""Scrape Knight Frank from sitemap + Rightmove API redirect"""
import re, json, time, requests
from datetime import date
from openpyxl import load_workbook

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'

def save_to_excel(new_objects, source_name):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
    new_count = 0
    for obj in new_objects:
        url = obj.get('url', '—')
        if url != '—' and url in existing_urls: continue
        ws.append([obj.get('titel',''), obj.get('quelle', source_name), url,
                   obj.get('preis'), obj.get('zimmer'), obj.get('grundstueck'),
                   obj.get('wohnflaeche'), obj.get('ort',''), str(date.today()), 'Neu'])
        if url != '—': existing_urls.add(url)
        new_count += 1
    wb.save(EXCEL_PATH)
    print(f"  ✅ {new_count} neue Objekte aus {source_name} gespeichert")
    return new_count

def parse_price(text):
    if not text: return None
    nums = re.findall(r'\d[\d\.,]+', str(text))
    for n in nums:
        try:
            v = int(float(n.replace('.', '').replace(',', '')))
            if v > 10000: return v
        except: pass
    return None

# ===== KNIGHT FRANK via SITEMAP =====
def scrape_knight_frank_sitemap():
    print("\n🏠 Knight Frank - Sitemap")
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120',
        'Accept': 'text/html,application/xhtml+xml',
    }
    
    # Collect all Mallorca URLs from sitemaps
    mallorca_urls = set()
    
    for i in range(1, 6):
        try:
            resp = requests.get(f'https://www.knightfrank.com/sitemaps/sitemap-pdp-{i}.xml', 
                               headers=headers, timeout=15)
            if resp.status_code != 200:
                print(f"  Sitemap {i}: {resp.status_code}")
                break
            
            urls = re.findall(r'https://www\.knightfrank\.com/properties/[^\s<"]+', resp.text)
            mallorca = [u for u in urls if 'mallorca' in u.lower()]
            print(f"  Sitemap {i}: {len(urls)} total, {len(mallorca)} Mallorca")
            mallorca_urls.update(mallorca)
            time.sleep(0.5)
        except Exception as e:
            print(f"  Sitemap {i} error: {e}")
    
    print(f"  Total Mallorca URLs: {len(mallorca_urls)}")
    
    # Now fetch each property page for details
    objects = []
    session = requests.Session()
    session.headers.update(headers)
    
    for url in list(mallorca_urls)[:50]:  # Limit to first 50 to be quick
        try:
            resp = session.get(url, timeout=10)
            if resp.status_code != 200:
                # Still add with URL only
                location = re.search(r'for-sale/([^/]+)-mallorca', url)
                ort = location.group(1).replace('-', ' ').title() if location else 'Mallorca'
                objects.append({'titel': ort + ' - Mallorca', 'quelle': 'Knight Frank', 'url': url,
                               'preis': None, 'zimmer': None, 'grundstueck': None, 'wohnflaeche': None, 'ort': ort})
                continue
            
            html = resp.text
            
            # Extract structured data / JSON-LD
            json_ld = re.findall(r'<script[^>]*type="application/ld\+json"[^>]*>(.*?)</script>', html, re.DOTALL)
            price = None
            rooms = None
            area = None
            title = ''
            
            for jld in json_ld:
                try:
                    data = json.loads(jld)
                    if isinstance(data, dict):
                        if 'offers' in data:
                            price = data['offers'].get('price', data['offers'].get('lowPrice'))
                        if 'name' in data: title = data['name']
                        if 'numberOfRooms' in data: rooms = data['numberOfRooms']
                        if 'floorSize' in data: area = data['floorSize'].get('value')
                except: pass
            
            # Also try meta tags and data attributes
            if not price:
                price_match = re.search(r'"price"\s*:\s*"?([\d,\.]+)"?', html)
                if price_match:
                    price = parse_price(price_match.group(1))
            
            if not title:
                title_match = re.search(r'<h1[^>]*>([^<]+)</h1>', html)
                if title_match: title = title_match.group(1).strip()
            
            # Location from URL
            loc_match = re.search(r'for-sale/(.+?)/rsi', url)
            ort = loc_match.group(1).replace('-', ' ').title() if loc_match else 'Mallorca'
            
            # Bedroom count
            beds_match = re.search(r'(\d+)\s*bed', html, re.I)
            if beds_match and not rooms: rooms = int(beds_match.group(1))
            
            # Area
            area_match = re.search(r'(\d+)\s*(?:sq\.?\s*m|m²|sqm)', html, re.I)
            if area_match and not area: area = int(area_match.group(1))
            
            objects.append({
                'titel': title[:100] or ort,
                'quelle': 'Knight Frank',
                'url': url,
                'preis': price,
                'zimmer': rooms,
                'grundstueck': None,
                'wohnflaeche': area,
                'ort': ort,
            })
            
            print(f"  ✓ {url[-40:]}: {title[:30]} | {price} | {rooms}Zi | {area}m²")
            time.sleep(0.3)
        
        except Exception as e:
            print(f"  Error {url[-40:]}: {e}")
    
    if objects:
        return save_to_excel(objects, 'Knight Frank')
    return 0


# ===== RIGHTMOVE - API with correct endpoint =====
def scrape_rightmove_api():
    print("\n🏠 Rightmove - API")
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Referer': 'https://www.rightmove.co.uk/overseas-property/in-Mallorca.html',
        'X-Requested-With': 'XMLHttpRequest',
    }
    
    session = requests.Session()
    session.headers.update(headers)
    
    # First visit the page to get session cookies
    try:
        page_resp = session.get('https://www.rightmove.co.uk/overseas-property/in-Mallorca.html', timeout=15)
        print(f"  Page status: {page_resp.status_code}")
        
        # Look for the locationIdentifier in the page
        loc_match = re.search(r'locationIdentifier["\s:=]+([A-Z_\^%\d]+)', page_resp.text)
        if loc_match:
            loc_id = loc_match.group(1)
            print(f"  Location ID: {loc_id}")
        else:
            loc_id = 'OVERSEAS%5E916'  # Fallback
        
        objects = []
        
        for index in range(0, 300, 24):
            api_url = f'https://www.rightmove.co.uk/api/_search?locationIdentifier={loc_id}&numberOfPropertiesPerPage=24&radius=0.0&sortType=2&index={index}&propertyTypes=&includeSSTC=false&viewType=LIST&channel=BUY&areaSizeUnit=sqm&currencyCode=EUR'
            resp = session.get(api_url, timeout=15)
            print(f"  index={index}: {resp.status_code}")
            
            if resp.status_code == 200:
                try:
                    data = resp.json()
                    props = data.get('properties', [])
                    print(f"  Properties: {len(props)}")
                    
                    if not props:
                        print(f"  Keys: {list(data.keys())}")
                        break
                    
                    for p in props:
                        url_val = p.get('propertyUrl', '')
                        if url_val and not url_val.startswith('http'):
                            url_val = 'https://www.rightmove.co.uk' + url_val
                        
                        price_data = p.get('price', {})
                        price = price_data.get('amount') if isinstance(price_data, dict) else price_data
                        
                        objects.append({
                            'titel': p.get('displayAddress', p.get('summary', '')),
                            'quelle': 'Rightmove',
                            'url': url_val or '—',
                            'preis': price,
                            'zimmer': p.get('bedrooms'),
                            'grundstueck': None,
                            'wohnflaeche': None,
                            'ort': p.get('displayAddress', 'Mallorca').split(',')[-1].strip(),
                        })
                    
                    total = data.get('resultCount', 0)
                    if isinstance(total, str): total = int(total.replace(',', ''))
                    print(f"  Total: {total}, Got so far: {len(objects)}")
                    if len(objects) >= total or not props:
                        break
                    
                    time.sleep(1)
                except Exception as e:
                    print(f"  JSON error: {e} | {resp.text[:200]}")
                    break
            else:
                print(f"  Response: {resp.text[:200]}")
                break
    
    except Exception as e:
        print(f"  Rightmove error: {e}")
        import traceback; traceback.print_exc()
        return 0
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Rightmove')
    return 0


# ===== FOTOCASA - Direct API call =====
def scrape_fotocasa_direct():
    print("\n🏠 Fotocasa - Direct API")
    
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'es-ES,es;q=0.9',
        'Origin': 'https://www.fotocasa.es',
        'Referer': 'https://www.fotocasa.es/',
    })
    
    # First get cookies
    try:
        session.get('https://www.fotocasa.es/', timeout=10)
    except: pass
    
    objects = []
    
    try:
        # Try the real search endpoint
        for page in range(1, 15):
            payload = {
                "culture": "es-ES",
                "locationIds": [724],
                "maxItems": 40,
                "page": page,
                "propertyTypeIds": [2],  # Houses
                "purchaseTypeIds": [1],  # Sale
                "sortBy": "score",
                "transactionTypeIds": [1]
            }
            
            resp = session.post(
                'https://search.gw.fotocasa.es/v2/propertysearch/search',
                json=payload, timeout=15
            )
            
            if resp.status_code != 200:
                print(f"  Page {page}: {resp.status_code} - {resp.text[:100]}")
                break
            
            data = resp.json()
            items = data.get('realEstates', [])
            if not items:
                print(f"  Keine Items auf Seite {page}")
                break
            
            print(f"  Page {page}: {len(items)} Items")
            
            for item in items:
                # Price
                price = None
                txs = item.get('transactions', [])
                if txs:
                    vals = txs[0].get('value', [])
                    price = vals[0] if vals else txs[0].get('price')
                
                # Features
                features = {f.get('key',''):(f.get('value',[None])[0] if isinstance(f.get('value'),list) else f.get('value')) for f in item.get('features',[])}
                
                # URL
                detail = item.get('detail', {})
                url_val = (detail.get('es','') if isinstance(detail,dict) else '')
                if url_val and not url_val.startswith('http'):
                    url_val = 'https://www.fotocasa.es' + url_val
                
                addr = item.get('address', {})
                
                objects.append({
                    'titel': addr.get('ubication', '') if isinstance(addr,dict) else '',
                    'quelle': 'Fotocasa',
                    'url': url_val or '—',
                    'preis': price,
                    'zimmer': features.get('roomsNumber'),
                    'grundstueck': features.get('plotArea'),
                    'wohnflaeche': features.get('constructedArea', features.get('surface')),
                    'ort': addr.get('municipality', '') if isinstance(addr,dict) else '',
                })
            
            total = data.get('totalAdCount', 999)
            if len(objects) >= total: break
            time.sleep(0.5)
    
    except Exception as e:
        print(f"  Error: {e}")
        import traceback; traceback.print_exc()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Fotocasa')
    return 0


if __name__ == '__main__':
    results = {}
    results['Knight Frank'] = scrape_knight_frank_sitemap()
    results['Rightmove'] = scrape_rightmove_api()
    results['Fotocasa'] = scrape_fotocasa_direct()
    
    print("\n=== ERGEBNIS ===")
    for src, count in results.items():
        print(f"  {'✅' if count else '❌'} {src}: {count}")
