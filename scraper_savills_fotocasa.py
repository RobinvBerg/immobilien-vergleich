#!/usr/bin/env python3
"""Savills SearchByUrl API + Fotocasa + Vives Pons + Von Poll"""
import re, json, time, requests
from datetime import date
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'

def save_to_excel(new_objects, source_name):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
    new_count = 0
    for obj in new_objects:
        url = obj.get('url', '—')
        if url != '—' and url in existing_urls: continue
        ws.append([obj.get('titel',''), obj.get('quelle', source_name), url,
                   obj.get('preis'), obj.get('zimmer'), obj.get('grundstueck'),
                   obj.get('wohnflaeche'), obj.get('ort',''), str(date.today()), 'Neu'])
        if url != '—': existing_urls.add(url)
        new_count += 1
    wb.save(EXCEL_PATH)
    print(f"  ✅ {new_count} neue Objekte aus {source_name} gespeichert")
    return new_count

def parse_price(text):
    if not text: return None
    nums = re.findall(r'\d[\d\.]+', str(text).replace(',',''))
    for n in nums:
        try:
            v = int(float(n.replace('.', '')))
            if v > 10000: return v
        except: pass
    return None

def parse_int(text):
    if not text: return None
    nums = re.findall(r'\d+', str(text))
    return int(nums[0]) if nums else None

# ===== SAVILLS - Direct search API =====
def scrape_savills():
    print("\n🏠 Savills - livev6-searchapi.savills.com")
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120',
        'Accept': 'application/json, text/plain, */*',
        'Origin': 'https://search.savills.com',
        'Referer': 'https://search.savills.com/',
        'Content-Type': 'application/json',
    }
    
    session = requests.Session()
    session.headers.update(headers)
    
    # Try the search endpoint directly
    try:
        # The URL from the Playwright capture: SearchByUrl
        payload_url = {
            "Url": "/es/en/list?SearchList=IsInRegion_EUR5002&SaleType=ForSale&PropertyType=RES"
        }
        resp = session.post('https://livev6-searchapi.savills.com/Data/SearchByUrl', 
                           json=payload_url, timeout=15)
        print(f"  SearchByUrl status: {resp.status_code}")
        if resp.status_code == 200:
            data = resp.json()
            print(f"  Keys: {list(data.keys()) if isinstance(data, dict) else 'list'}")
            if isinstance(data, dict):
                for k, v in data.items():
                    if isinstance(v, (dict, list)):
                        print(f"  {k}: {type(v).__name__}({len(v) if isinstance(v, list) else len(v.keys()) if isinstance(v, dict) else '?'})")
                    else:
                        print(f"  {k}: {v}")
    except Exception as e:
        print(f"  SearchByUrl error: {e}")
    
    # Try Map/Search endpoint  
    objects = []
    try:
        # Need to figure out the right params
        # Try with SearchList for Balearics (EUR5002)
        payload_search = {
            "SearchList": "IsInRegion_EUR5002",
            "SaleType": "ForSale",
            "PropertyType": "RES",
            "Currency": "EUR",
            "Take": 50,
            "Skip": 0
        }
        resp = session.post('https://livev6-searchapi.savills.com/Map/Search', 
                           json=payload_search, timeout=15)
        print(f"  Map/Search status: {resp.status_code}")
        if resp.status_code == 200:
            data = resp.json()
            print(f"  Response keys: {list(data.keys()) if isinstance(data, dict) else 'list'}")
            
            items = None
            if isinstance(data, dict):
                for k in ['Results', 'results', 'Listings', 'listings', 'Properties', 'properties', 'Data', 'data']:
                    if k in data and isinstance(data[k], list):
                        items = data[k]
                        print(f"  Items under '{k}': {len(items)}")
                        if items: print(f"  Sample: {json.dumps(items[0], ensure_ascii=False)[:400]}")
                        break
            
            if items:
                for item in items:
                    url_v = item.get('Url', item.get('url', item.get('Link', item.get('link', '—'))))
                    if url_v and not url_v.startswith('http') and url_v != '—':
                        url_v = 'https://search.savills.com' + url_v
                    
                    price_raw = item.get('Price', item.get('price', {}))
                    if isinstance(price_raw, dict):
                        price = price_raw.get('Value', price_raw.get('value', price_raw.get('Amount', price_raw.get('amount'))))
                    else:
                        price = price_raw
                    
                    objects.append({
                        'titel': item.get('Title', item.get('title', item.get('Address', item.get('address', '')))),
                        'quelle': 'Savills',
                        'url': url_v,
                        'preis': price,
                        'zimmer': item.get('Bedrooms', item.get('bedrooms', item.get('Beds', item.get('beds')))),
                        'grundstueck': item.get('PlotSize', item.get('plotSize', item.get('LandArea'))),
                        'wohnflaeche': item.get('FloorArea', item.get('floorArea', item.get('Area', item.get('area')))),
                        'ort': item.get('Location', item.get('location', item.get('City', item.get('city', 'Mallorca')))),
                    })
    except Exception as e:
        print(f"  Map/Search error: {e}")
        import traceback; traceback.print_exc()
    
    # Try the actual Savills search API used by their website
    if not objects:
        try:
            # Try with GET
            resp = session.get(
                'https://livev6-searchapi.savills.com/Map/Search?SearchList=IsInRegion_EUR5002&SaleType=ForSale&PropertyType=RES',
                timeout=15
            )
            print(f"  GET Map/Search: {resp.status_code}")
            if resp.status_code == 200:
                data = resp.json()
                print(f"  GET Response: {json.dumps(data, ensure_ascii=False)[:500]}")
        except Exception as e:
            print(f"  GET error: {e}")
    
    # Playwright - wait for the Map/Search response to get the actual data
    if not objects:
        print("  Using Playwright to intercept Savills search...")
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=['--no-sandbox'])
            context = browser.new_context(
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120',
                viewport={'width': 1280, 'height': 800}
            )
            pg = context.new_page()
            Stealth().apply_stealth_sync(pg)
            
            search_results = {}
            
            def on_response(resp):
                if 'livev6-searchapi.savills.com' in resp.url and resp.status == 200:
                    try:
                        body = resp.body()
                        data = json.loads(body)
                        search_results[resp.url] = data
                        print(f"  Savills API: {resp.url}")
                    except: pass
            
            pg.on('response', on_response)
            
            pg.goto('https://search.savills.com/es/en/list?SearchList=IsInRegion_EUR5002&SaleType=ForSale&PropertyType=RES',
                   wait_until='networkidle', timeout=30000)
            time.sleep(5)
            
            # Try to accept cookie
            try:
                pg.click('#onetrust-accept-btn-handler', timeout=3000)
                time.sleep(2)
            except: pass
            
            # Scroll to trigger more loads
            for _ in range(3):
                pg.evaluate('window.scrollBy(0, 500)')
                time.sleep(1)
            
            print(f"  Captured Savills API calls: {len(search_results)}")
            
            for url, data in search_results.items():
                print(f"  API {url}: type={type(data)}")
                if isinstance(data, dict):
                    for k, v in data.items():
                        if isinstance(v, list) and len(v) > 0:
                            print(f"    {k}: {len(v)} items")
                            # Sample first item
                            print(f"    Sample: {json.dumps(v[0], ensure_ascii=False)[:300]}")
                            for item in v:
                                url_v = item.get('Url', item.get('url', item.get('Link', '—')))
                                if url_v and not url_v.startswith('http') and url_v != '—':
                                    url_v = 'https://search.savills.com' + url_v
                                
                                price_raw = item.get('Price', item.get('price', {}))
                                if isinstance(price_raw, dict):
                                    price = price_raw.get('Value', price_raw.get('value', price_raw.get('Amount')))
                                else:
                                    price = price_raw
                                
                                objects.append({
                                    'titel': item.get('Title', item.get('title', item.get('Address', ''))),
                                    'quelle': 'Savills',
                                    'url': url_v,
                                    'preis': price,
                                    'zimmer': item.get('Bedrooms', item.get('bedrooms')),
                                    'grundstueck': item.get('PlotSize', item.get('LandArea')),
                                    'wohnflaeche': item.get('FloorArea', item.get('Area')),
                                    'ort': item.get('Location', item.get('City', 'Mallorca')),
                                })
            
            browser.close()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Savills')
    return 0


# ===== FOTOCASA - Try correct API endpoint =====
def scrape_fotocasa():
    print("\n🏠 Fotocasa - Different Endpoints")
    objects = []
    
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'es-ES,es;q=0.9',
        'Origin': 'https://www.fotocasa.es',
        'Referer': 'https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l',
    })
    
    # Try getting session cookies first
    try:
        resp = session.get('https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l', timeout=10)
        print(f"  Page status: {resp.status_code}, cookies: {list(session.cookies.keys())}")
    except Exception as e:
        print(f"  Page error: {e}")
    
    # Try different API variants
    endpoints = [
        ('GET', 'https://search.gw.fotocasa.es/v2/propertysearch/search?locationIds=724&purchaseTypeIds=1&propertyTypeIds=2&maxItems=40&page=1'),
        ('GET', 'https://search.gw.fotocasa.es/v2/propertysearch/search?location=mallorca&type=sale&page=1'),
        ('POST', 'https://search.gw.fotocasa.es/v2/propertysearch/search'),
        ('POST', 'https://search.gw.fotocasa.es/v2/propertysearch/search?locationIds=724'),
        ('GET', 'https://api.fotocasa.es/v2/propertysearch/search?locationIds=724&purchaseTypeIds=1&page=1'),
    ]
    
    for method, url in endpoints:
        try:
            if method == 'GET':
                resp = session.get(url, timeout=10)
            else:
                payload = {
                    "culture": "es-ES",
                    "locationIds": [724],
                    "maxItems": 40,
                    "page": 1,
                    "propertyTypeIds": [2],
                    "purchaseTypeIds": [1],
                    "sortBy": "score",
                    "transactionTypeIds": [1]
                }
                resp = session.post(url, json=payload, timeout=10)
            
            print(f"  {method} {url[-60:]}: {resp.status_code}")
            if resp.status_code == 200:
                data = resp.json()
                items = data.get('realEstates', data.get('results', []))
                print(f"  ✓ Items: {len(items)}")
                if items:
                    for item in items:
                        price = None
                        txs = item.get('transactions', [])
                        if txs:
                            vals = txs[0].get('value', [])
                            price = vals[0] if vals else None
                        features = {f.get('key',''):(f.get('value',[None])[0] if isinstance(f.get('value'),list) else f.get('value')) for f in item.get('features',[])}
                        detail = item.get('detail', {})
                        url_v = detail.get('es','') if isinstance(detail,dict) else ''
                        if url_v and not url_v.startswith('http'): url_v = 'https://www.fotocasa.es' + url_v
                        addr = item.get('address', {})
                        objects.append({
                            'titel': addr.get('ubication','') if isinstance(addr,dict) else '',
                            'quelle': 'Fotocasa', 'url': url_v or '—',
                            'preis': price, 'zimmer': features.get('roomsNumber'),
                            'grundstueck': features.get('plotArea'),
                            'wohnflaeche': features.get('constructedArea', features.get('surface')),
                            'ort': addr.get('municipality','') if isinstance(addr,dict) else '',
                        })
                    break
        except Exception as e:
            print(f"  {url[-50:]}: error: {e}")
    
    # If direct API works, get more pages
    if objects:
        working_url = None
        for method, url in endpoints:
            # Recheck which one worked
            pass
        
        # Try more pages of the working endpoint
        for pg in range(2, 20):
            try:
                payload = {
                    "culture": "es-ES",
                    "locationIds": [724],
                    "maxItems": 40,
                    "page": pg,
                    "propertyTypeIds": [2],
                    "purchaseTypeIds": [1],
                    "sortBy": "score",
                    "transactionTypeIds": [1]
                }
                resp = session.post('https://search.gw.fotocasa.es/v2/propertysearch/search', 
                                   json=payload, timeout=10)
                if resp.status_code != 200: break
                data = resp.json()
                items = data.get('realEstates', [])
                if not items: break
                for item in items:
                    price = None
                    txs = item.get('transactions', [])
                    if txs:
                        vals = txs[0].get('value', [])
                        price = vals[0] if vals else None
                    features = {f.get('key',''):(f.get('value',[None])[0] if isinstance(f.get('value'),list) else f.get('value')) for f in item.get('features',[])}
                    detail = item.get('detail', {})
                    url_v = detail.get('es','') if isinstance(detail,dict) else ''
                    if url_v and not url_v.startswith('http'): url_v = 'https://www.fotocasa.es' + url_v
                    addr = item.get('address', {})
                    objects.append({
                        'titel': addr.get('ubication','') if isinstance(addr,dict) else '',
                        'quelle': 'Fotocasa', 'url': url_v or '—',
                        'preis': price, 'zimmer': features.get('roomsNumber'),
                        'grundstueck': features.get('plotArea'),
                        'wohnflaeche': features.get('constructedArea', features.get('surface')),
                        'ort': addr.get('municipality','') if isinstance(addr,dict) else '',
                    })
                print(f"  Seite {pg}: +{len(items)}, gesamt: {len(objects)}")
                time.sleep(0.5)
            except: break
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Fotocasa')
    return 0


# ===== VIVES PONS - Different approach =====
def scrape_vives_pons():
    print("\n🏠 Vives Pons - XML Sitemap + Direct")
    objects = []
    
    session = requests.Session()
    session.headers.update({'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'})
    
    # Try sitemap
    try:
        resp = session.get('https://www.vivespons.com/sitemap.xml', timeout=10)
        print(f"  Sitemap: {resp.status_code}")
        if resp.status_code == 200:
            property_urls = re.findall(r'<loc>(https://www\.vivespons\.com/en/properties/[^<]+)</loc>', resp.text)
            print(f"  Property URLs from sitemap: {len(property_urls)}")
            for url in property_urls[:50]:
                objects.append({
                    'titel': url.split('/')[-1].replace('-', ' ').title(),
                    'quelle': 'Vives Pons',
                    'url': url,
                    'preis': None, 'zimmer': None, 'grundstueck': None,
                    'wohnflaeche': None, 'ort': 'Mallorca',
                })
    except Exception as e:
        print(f"  Sitemap error: {e}")
    
    # Try their API
    if not objects:
        for endpoint in [
            'https://www.vivespons.com/api/properties?lang=en&type=sale',
            'https://www.vivespons.com/api/en/properties',
            'https://www.vivespons.com/wp-json/wp/v2/properties',
            'https://www.vivespons.com/wp-json/wpresidence/v1/listings',
        ]:
            try:
                resp = session.get(endpoint, timeout=8)
                print(f"  {endpoint[-50:]}: {resp.status_code}")
                if resp.status_code == 200:
                    data = resp.json()
                    print(f"  Keys/type: {list(data.keys()) if isinstance(data, dict) else type(data)}")
            except: pass
    
    # Playwright as last resort
    if not objects:
        print("  Playwright fallback...")
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=['--no-sandbox'])
            context = browser.new_context(user_agent='Mozilla/5.0 Chrome/120', viewport={'width': 1280, 'height': 800})
            pg = context.new_page()
            Stealth().apply_stealth_sync(pg)
            
            captured = []
            def on_resp(r):
                if 'vivespons' in r.url and r.status == 200:
                    ct = r.headers.get('content-type', '')
                    if 'json' in ct:
                        try:
                            b = r.body()
                            captured.append({'url': r.url, 'data': json.loads(b)})
                            print(f"  API: {r.url}")
                        except: pass
            pg.on('response', on_resp)
            
            pg.goto('https://www.vivespons.com/en/properties', wait_until='networkidle', timeout=25000)
            time.sleep(3)
            
            html = pg.content()
            print(f"  HTML: {len(html)}b")
            
            # Try to find properties in the page
            # Look for JSON data in script tags
            json_scripts = re.findall(r'<script[^>]*>(.*?)</script>', html, re.DOTALL)
            for script in json_scripts:
                if 'properties' in script.lower() and len(script) > 100:
                    try:
                        # Try to find JSON arrays
                        matches = re.findall(r'\{[^{}]*"title"[^{}]*"url"[^{}]*\}', script)
                        if matches:
                            print(f"  Found {len(matches)} property objects in script")
                    except: pass
            
            # Get all links
            all_links = pg.query_selector_all('a')
            seen = set()
            for link in all_links:
                try:
                    href = link.get_attribute('href') or ''
                    if not href or href in seen: continue
                    seen.add(href)
                    if not href.startswith('http'): href = 'https://www.vivespons.com' + href
                    text = link.inner_text().strip()
                    if '/properties/' in href and href != 'https://www.vivespons.com/en/properties':
                        objects.append({
                            'titel': text[:100] or href.split('/')[-1].replace('-', ' ').title(),
                            'quelle': 'Vives Pons', 'url': href,
                            'preis': None, 'zimmer': None, 'grundstueck': None,
                            'wohnflaeche': None, 'ort': 'Mallorca',
                        })
                except: pass
            
            print(f"  Links found: {len(objects)}")
            
            # Try scrolling to load more
            if not objects:
                for _ in range(5):
                    pg.evaluate('window.scrollBy(0, 800)')
                    time.sleep(1)
                
                # Try to find property items
                items = pg.query_selector_all('[class*="property"], [class*="listing"], article, .card')
                print(f"  Items after scroll: {len(items)}")
                
                for item in items[:30]:
                    try:
                        link = item.query_selector('a')
                        href = link.get_attribute('href') if link else None
                        if href and not href.startswith('http'): href = 'https://www.vivespons.com' + href
                        title_el = item.query_selector('h2, h3, .title')
                        price_el = item.query_selector('[class*="price"]')
                        if href:
                            objects.append({
                                'titel': title_el.inner_text()[:100] if title_el else '',
                                'quelle': 'Vives Pons', 'url': href or '—',
                                'preis': parse_price(price_el.inner_text() if price_el else ''),
                                'zimmer': None, 'grundstueck': None, 'wohnflaeche': None, 'ort': 'Mallorca',
                            })
                    except: pass
            
            browser.close()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Vives Pons')
    return 0


# ===== VON POLL - Try with domcontentloaded =====
def scrape_von_poll():
    print("\n🏠 Von Poll - domcontentloaded")
    objects = []
    
    # First try to get their API via curl
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'de-DE,de;q=0.9',
    })
    
    try:
        resp = session.get('https://www.von-poll.com/de/immobilien/spanien/balearen/mallorca', 
                          timeout=20, allow_redirects=True)
        print(f"  Page status: {resp.status_code}, len: {len(resp.text)}")
        
        if resp.status_code == 200:
            html = resp.text
            
            # Look for property data in page
            links = re.findall(r'href="(/de/expose/[^"]+)"', html)
            print(f"  Expose links: {len(links)}")
            
            for href in links[:50]:
                full_url = 'https://www.von-poll.com' + href
                objects.append({
                    'titel': 'Von Poll Objekt',
                    'quelle': 'Von Poll Real Estate',
                    'url': full_url,
                    'preis': None, 'zimmer': None, 'grundstueck': None,
                    'wohnflaeche': None, 'ort': 'Mallorca',
                })
            
            # Also look for JSON data in page
            json_data = re.findall(r'window\.__INITIAL_STATE__\s*=\s*JSON\.parse\(["\'](.+?)["\']\)', html)
            json_data += re.findall(r'"objects"\s*:\s*(\[.{100,}\])', html[:500000], re.DOTALL)
            
            for jd in json_data[:2]:
                try:
                    data = json.loads(jd.replace('\\"', '"').replace('\\\\', '\\'))
                    print(f"  JSON data type: {type(data)}, keys: {list(data.keys())[:5] if isinstance(data, dict) else 'list'}")
                except Exception as e:
                    print(f"  JSON parse: {e}")
        
    except Exception as e:
        print(f"  Requests error: {e}")
    
    if not objects:
        print("  Playwright with domcontentloaded...")
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=['--no-sandbox'])
            context = browser.new_context(
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120',
                viewport={'width': 1280, 'height': 800}
            )
            pg = context.new_page()
            Stealth().apply_stealth_sync(pg)
            
            api_captured = []
            def on_resp(r):
                if 'von-poll' in r.url and r.status == 200:
                    ct = r.headers.get('content-type', '')
                    if 'json' in ct:
                        try:
                            b = r.body()
                            if len(b) > 200:
                                api_captured.append({'url': r.url, 'body': b})
                                print(f"  VP API: {r.url[:100]}")
                        except: pass
            pg.on('response', on_resp)
            
            try:
                pg.goto('https://www.von-poll.com/de/immobilien/spanien/balearen/mallorca',
                       wait_until='domcontentloaded', timeout=25000)
                time.sleep(5)
                
                html = pg.content()
                print(f"  HTML: {len(html)}b, APIs: {len(api_captured)}")
                print(f"  Title: {pg.title()}")
                
                # Extract links
                expose_links = re.findall(r'/de/expose/[^"\'?\s]+', html)
                print(f"  Expose links in HTML: {len(expose_links)}")
                
                seen = set()
                for href in expose_links:
                    if href in seen: continue
                    seen.add(href)
                    full_url = 'https://www.von-poll.com' + href
                    objects.append({
                        'titel': 'Von Poll Mallorca',
                        'quelle': 'Von Poll Real Estate',
                        'url': full_url,
                        'preis': None, 'zimmer': None, 'grundstueck': None,
                        'wohnflaeche': None, 'ort': 'Mallorca',
                    })
                
                # Process any API data
                for api in api_captured:
                    try:
                        data = json.loads(api['body'])
                        for k in ['items', 'results', 'properties', 'objects', 'exposees', 'data']:
                            if k in data and isinstance(data[k], list) and data[k]:
                                print(f"  VP data '{k}': {len(data[k])}")
                                for item in data[k]:
                                    url_v = item.get('url', item.get('link', '—'))
                                    if url_v and not url_v.startswith('http') and url_v != '—':
                                        url_v = 'https://www.von-poll.com' + url_v
                                    objects.append({
                                        'titel': item.get('title', item.get('headline', '')),
                                        'quelle': 'Von Poll Real Estate',
                                        'url': url_v,
                                        'preis': item.get('price', item.get('priceValue')),
                                        'zimmer': item.get('rooms', item.get('bedrooms')),
                                        'grundstueck': item.get('plotArea'),
                                        'wohnflaeche': item.get('area', item.get('livingArea')),
                                        'ort': item.get('location', item.get('city', 'Mallorca')),
                                    })
                                break
                    except: pass
            
            except Exception as e:
                print(f"  Playwright error: {e}")
            finally:
                browser.close()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Von Poll Real Estate')
    return 0


# ===== PROPERSTAR - Try REST API =====
def scrape_properstar():
    print("\n🏠 Properstar - Verschiedene APIs")
    objects = []
    
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120',
        'Accept': 'application/json',
    })
    
    # Try different Properstar API endpoints
    endpoints = [
        ('GET', 'https://www.properstar.com/api/listing/search?country=ES&region=Balearic+Islands&type=sale&page=1&limit=48'),
        ('GET', 'https://api.properstar.com/v1/listings?country=ES&region=mallorca&type=sale&page=1'),
        ('GET', 'https://www.properstar.com/api/v2/search?country=ES&city=mallorca&type=sale'),
        ('POST', 'https://listing-api.properstar.com/api/Listing/search'),
    ]
    
    for method, url in endpoints:
        try:
            if method == 'GET':
                resp = session.get(url, timeout=8)
            else:
                resp = session.post(url, json={
                    "countryCode": "ES",
                    "city": "Mallorca",
                    "listingType": "sale",
                    "pageNumber": 1,
                    "pageSize": 48
                }, timeout=8)
            print(f"  {method} {url[-60:]}: {resp.status_code}")
            if resp.status_code == 200:
                try:
                    data = resp.json()
                    print(f"  Success! {list(data.keys()) if isinstance(data, dict) else type(data)}")
                except: pass
        except Exception as e:
            print(f"  {e}")
    
    print(f"  Gesammelt: {len(objects)}")
    return 0


# ===== RIGHTMOVE - Via Playwright with cookies =====
def scrape_rightmove():
    print("\n🏠 Rightmove - Playwright mit Session")
    objects = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=['--no-sandbox'])
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120',
            viewport={'width': 1280, 'height': 800},
        )
        pg = context.new_page()
        Stealth().apply_stealth_sync(pg)
        
        api_calls = []
        def on_resp(r):
            if 'rightmove' in r.url and r.status == 200:
                if 'json' in r.headers.get('content-type', ''):
                    try:
                        b = r.body()
                        data = json.loads(b)
                        if 'properties' in data or 'searchResults' in data:
                            api_calls.append(data)
                            print(f"  RM JSON: {r.url[:80]}")
                    except: pass
        pg.on('response', on_resp)
        
        try:
            pg.goto('https://www.rightmove.co.uk/overseas-property/in-Mallorca.html',
                   wait_until='domcontentloaded', timeout=25000)
            time.sleep(4)
            
            html = pg.content()
            print(f"  RM HTML: {len(html)}b, API calls: {len(api_calls)}")
            
            # Process any API data
            for data in api_calls:
                props = data.get('properties', data.get('searchResults', []))
                for p_item in props:
                    url_v = p_item.get('propertyUrl', p_item.get('url', '—'))
                    if url_v and not url_v.startswith('http') and url_v != '—':
                        url_v = 'https://www.rightmove.co.uk' + url_v
                    price_d = p_item.get('price', {})
                    price = price_d.get('amount') if isinstance(price_d, dict) else price_d
                    objects.append({
                        'titel': p_item.get('displayAddress', p_item.get('summary', '')),
                        'quelle': 'Rightmove', 'url': url_v,
                        'preis': price, 'zimmer': p_item.get('bedrooms'),
                        'grundstueck': None, 'wohnflaeche': None,
                        'ort': p_item.get('displayAddress', 'Mallorca').split(',')[-1].strip(),
                    })
            
            # HTML scraping for cards
            cards = pg.query_selector_all('[class*="propertyCard"], [data-test*="property-result"]')
            links = pg.query_selector_all('a[href*="overseas-property/property-"]')
            print(f"  Cards: {len(cards)}, Links: {len(links)}")
            
            seen = set()
            for link in links:
                try:
                    href = link.get_attribute('href') or ''
                    if href in seen: continue
                    seen.add(href)
                    if not href.startswith('http'): href = 'https://www.rightmove.co.uk' + href
                    parent = link
                    # Get text
                    text = link.inner_text().strip()
                    objects.append({
                        'titel': text[:100] or 'Rightmove Mallorca',
                        'quelle': 'Rightmove', 'url': href,
                        'preis': None, 'zimmer': None, 'grundstueck': None,
                        'wohnflaeche': None, 'ort': 'Mallorca',
                    })
                except: pass
        
        except Exception as e:
            print(f"  Error: {e}")
        
        browser.close()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Rightmove')
    return 0


if __name__ == '__main__':
    results = {}
    results['Savills'] = scrape_savills()
    results['Fotocasa'] = scrape_fotocasa()
    results['Vives Pons'] = scrape_vives_pons()
    results['Von Poll'] = scrape_von_poll()
    results['Properstar'] = scrape_properstar()
    results['Rightmove'] = scrape_rightmove()
    
    print("\n=== ERGEBNIS ===")
    total = 0
    for src, count in results.items():
        print(f"  {'✅' if count else '❌'} {src}: {count}")
        total += count
    print(f"  Total: {total}")
