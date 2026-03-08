#!/usr/bin/env python3
"""
Apify Idealista Run6 Processor
Polls until SUCCEEDED, downloads dataset, filters, appends to Excel.
"""

import json
import time
import urllib.request
import urllib.error
import os
import sys
from datetime import date

# ── Config ─────────────────────────────────────────────────────────────────
RUN_ID     = "Zr1L4mCKAvW0o5DRW"
DATASET_ID = "ZCNiYHfQktY2m6qo2"
TOKEN      = "apify_api_feD2KhARHjtuV9CrSwOReYgoePFSF44nsDL6"

BASE_DIR   = "/Users/robin/.openclaw/workspace/mallorca-projekt"
RAW_FILE   = os.path.join(BASE_DIR, "idealista_run6_raw.json")
EXCEL_FILE = os.path.join(BASE_DIR, "Mallorca_Markt_Gesamt.xlsx")
SEEN_FILE  = os.path.join(BASE_DIR, "mallorca_seen_ids.json")

POLL_INTERVAL = 120   # seconds
MAX_POLLS     = 15    # 15 × 2 min = 30 min

ALLOWED_TYPES = {"chalet", "countryHouse"}
MIN_PRICE  = 2_000_000
MAX_PRICE  = 6_500_000
MIN_ROOMS  = 5

TODAY = "2026-02-22"

# ── Helpers ─────────────────────────────────────────────────────────────────
def api_get(url):
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode())

def check_status():
    url = f"https://api.apify.com/v2/actor-runs/{RUN_ID}?token={TOKEN}"
    data = api_get(url)
    return data["data"]["status"]

def download_dataset():
    print("⬇️  Downloading dataset …")
    url = f"https://api.apify.com/v2/datasets/{DATASET_ID}/items?token={TOKEN}&format=json&limit=100000&clean=true"
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=120) as r:
        raw = r.read()
    with open(RAW_FILE, "wb") as f:
        f.write(raw)
    items = json.loads(raw)
    print(f"✅ Downloaded {len(items)} items → {RAW_FILE}")
    return items

def analyse(items):
    from collections import Counter
    types = Counter(i.get("propertyType","?") for i in items)
    provinces = Counter(i.get("province","?") for i in items)
    print(f"\n📊 ANALYSE ({len(items)} total items)")
    print("  Property Types:")
    for t, n in types.most_common():
        print(f"    {t}: {n}")
    print("  Provinces:")
    for p, n in provinces.most_common():
        print(f"    {p}: {n}")
    return types, provinces

def load_seen():
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE) as f:
            return set(json.load(f))
    return set()

def save_seen(seen_set):
    with open(SEEN_FILE, "w") as f:
        json.dump(sorted(seen_set), f, indent=2)

def get_title(item):
    st = item.get("suggestedTexts", {})
    if isinstance(st, dict):
        t = st.get("title", "")
        if t:
            return t[:100]
    desc = item.get("description", "")
    return desc[:60] if desc else "Kein Titel"

def get_location(item):
    district     = (item.get("district") or "").strip()
    municipality = (item.get("municipality") or "").strip()
    if district and district != municipality:
        return f"{district}, {municipality}"
    return municipality

def filter_items(items, seen):
    new_rows = []
    new_seen = []
    skipped_type = skipped_price = skipped_rooms = skipped_dup = 0

    for item in items:
        pt = item.get("propertyType", "")
        if pt not in ALLOWED_TYPES:
            skipped_type += 1
            continue

        price = item.get("price", 0) or 0
        if not (MIN_PRICE <= price <= MAX_PRICE):
            skipped_price += 1
            continue

        rooms = item.get("rooms", 0) or 0
        if rooms < MIN_ROOMS:
            skipped_rooms += 1
            continue

        pc = item.get("propertyCode", "")
        seen_id = f"idealista_{pc}"
        if seen_id in seen:
            skipped_dup += 1
            continue

        title    = get_title(item)
        url      = item.get("url", "")
        size     = item.get("size") or ""
        location = get_location(item)

        new_rows.append([
            title,
            "Idealista",
            url,
            int(price),
            int(rooms),
            "",          # Grundstück – not available
            int(size) if size else "",
            location,
            TODAY,
            "Neu"
        ])
        new_seen.append(seen_id)

    print(f"\n🔍 Filter-Ergebnis:")
    print(f"  Übersprungen (Typ):      {skipped_type}")
    print(f"  Übersprungen (Preis):    {skipped_price}")
    print(f"  Übersprungen (Zimmer):   {skipped_rooms}")
    print(f"  Übersprungen (Duplikat): {skipped_dup}")
    print(f"  ✅ Neue Einträge:         {len(new_rows)}")

    return new_rows, new_seen

def update_excel(new_rows):
    import openpyxl

    HEADERS = ["Titel","Quelle","URL","Preis (€)","Zimmer",
               "Grundstück (m²)","Wohnfläche (m²)","Ort / Gemeinde",
               "Gefunden am","Status"]

    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        existing_rows = ws.max_row
        print(f"📂 Excel geöffnet: {existing_rows} Zeilen (inkl. Header)")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mallorca Markt"
        ws.append(HEADERS)
        existing_rows = 1
        print("📂 Excel neu erstellt mit Header")

    for row in new_rows:
        ws.append(row)

    wb.save(EXCEL_FILE)
    total_data_rows = ws.max_row - 1  # minus header
    print(f"💾 Excel gespeichert: {total_data_rows} Daten-Zeilen total")
    return total_data_rows

# ── Main ────────────────────────────────────────────────────────────────────
def main():
    print(f"🚀 Run6 Monitor gestartet | Run ID: {RUN_ID}")
    print(f"⏳ Polling alle {POLL_INTERVAL}s, max {MAX_POLLS} Versuche\n")

    status = None
    for attempt in range(1, MAX_POLLS + 1):
        try:
            status = check_status()
        except Exception as e:
            print(f"[{attempt}/{MAX_POLLS}] Fehler beim Status-Check: {e}")
            time.sleep(POLL_INTERVAL)
            continue

        print(f"[{attempt}/{MAX_POLLS}] Status: {status}")

        if status == "SUCCEEDED":
            print("✅ Run SUCCEEDED!\n")
            break
        elif status == "FAILED":
            print("❌ Run FAILED! Abbruch.")
            sys.exit(1)
        elif status in ("ABORTED", "TIMED-OUT", "TIMING-OUT"):
            print(f"❌ Run hat Status '{status}'. Abbruch.")
            sys.exit(1)
        else:
            if attempt < MAX_POLLS:
                print(f"   → Warte {POLL_INTERVAL}s …")
                time.sleep(POLL_INTERVAL)
    else:
        print("⏰ Timeout: Run hat nach 30 Minuten noch nicht SUCCEEDED.")
        sys.exit(1)

    # Download
    items = download_dataset()

    # Analyse
    analyse(items)

    # Load seen IDs
    seen = load_seen()
    print(f"\n📋 Bekannte IDs geladen: {len(seen)}")

    # Filter
    new_rows, new_seen_ids = filter_items(items, seen)

    if not new_rows:
        print("\nℹ️  Keine neuen Einträge nach Filter – Excel und seen_ids unverändert.")
        return

    # Update Excel
    total_rows = update_excel(new_rows)

    # Update seen IDs
    seen.update(new_seen_ids)
    save_seen(seen)
    print(f"💾 seen_ids.json aktualisiert: {len(seen)} IDs total")

    # Summary
    print(f"\n{'='*50}")
    print(f"✅ FERTIG")
    print(f"   Neue Einträge hinzugefügt: {len(new_rows)}")
    print(f"   Gesamt in Excel:           {total_rows}")
    print(f"   Gesamt seen IDs:           {len(seen)}")
    print(f"{'='*50}")

if __name__ == "__main__":
    main()
