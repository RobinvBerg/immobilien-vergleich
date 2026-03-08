import requests
from bs4 import BeautifulSoup
import re, time
from openpyxl import load_workbook
from datetime import date

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'de-DE,de;q=0.9,en;q=0.8',
}

wb = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
ws = wb['Mallorca Objekte']
existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')

# === Versuch 1: Sitemaps ===
print("=== Von Poll Sitemaps ===")
sitemap_urls = [
    'https://www.von-poll.com/sitemap.xml',
    'https://www.von-poll.com/sitemap_index.xml',
    'https://www.von-poll.com/de/sitemap.xml',
    'https://www.von-poll.com/robots.txt',
]

property_urls = []
for surl in sitemap_urls:
    try:
        r = requests.get(surl, headers=headers, timeout=15)
        print(f"{surl}: {r.status_code} | {len(r.text)} chars")
        if r.status_code == 200:
            # Suche nach Mallorca/Balearen URLs
            mallorca_links = [line.strip() for line in r.text.split('\n') 
                             if ('mallorca' in line.lower() or 'balearen' in line.lower() or 'spanien' in line.lower())
                             and ('http' in line.lower())]
            print(f"  Mallorca-Links: {len(mallorca_links)}")
            for link in mallorca_links[:5]:
                print(f"  {link}")
            # Auch aus XML parsen
            if '<url>' in r.text or '<loc>' in r.text:
                soup = BeautifulSoup(r.text, 'xml')
                locs = soup.find_all('loc')
                for loc in locs:
                    url_text = loc.get_text()
                    if 'mallorca' in url_text.lower() or 'balearen' in url_text.lower():
                        property_urls.append(url_text)
            # Sub-sitemaps entdecken
            if '<sitemap>' in r.text or 'sitemap' in r.text.lower():
                sub_sitemaps = re.findall(r'https?://[^\s<>"]+sitemap[^\s<>"]*', r.text)
                print(f"  Sub-Sitemaps gefunden: {sub_sitemaps[:5]}")
    except Exception as e:
        print(f"Error {surl}: {e}")
    time.sleep(0.5)

# Versuch Sub-Sitemaps zu finden (Expose-Sitemap)
print("\n=== Sub-Sitemap-Suche ===")
extra_sitemaps = [
    'https://www.von-poll.com/sitemap-expose.xml',
    'https://www.von-poll.com/sitemap-properties.xml',
    'https://www.von-poll.com/sitemap-immobilien.xml',
    'https://www.von-poll.com/de/sitemap-expose.xml',
    'https://www.von-poll.com/sitemap1.xml',
    'https://www.von-poll.com/sitemap2.xml',
]
for surl in extra_sitemaps:
    try:
        r = requests.get(surl, headers=headers, timeout=10)
        print(f"{surl}: {r.status_code} | {len(r.text)} chars")
        if r.status_code == 200 and ('<loc>' in r.text or '<url>' in r.text):
            soup = BeautifulSoup(r.text, 'xml')
            locs = soup.find_all('loc')
            print(f"  Total URLs: {len(locs)}")
            for loc in locs:
                url_text = loc.get_text()
                if 'mallorca' in url_text.lower() or 'balearen' in url_text.lower() or 'expose' in url_text.lower():
                    property_urls.append(url_text)
    except Exception as e:
        print(f"Error {surl}: {e}")
    time.sleep(0.3)

print(f"\nGefundene Mallorca-URLs in Sitemap: {len(property_urls)}")
for u in property_urls[:10]:
    print(f"  {u}")

# === Versuch 2: Direkte Objekt-URLs testen (CF-Challenge nur auf Suchseite?) ===
print("\n=== Von Poll direkte Objekt-Seiten ===")
test_urls = [
    'https://www.von-poll.com/de/immobilien/spanien/balearen/mallorca',
    'https://www.von-poll.com/de/expose/mallorca',
    'https://www.von-poll.com/de/immobilien/spanien',
    'https://www.von-poll.com/de/suche/mallorca',
]
for turl in test_urls:
    try:
        r = requests.get(turl, headers=headers, timeout=15)
        title_tag = BeautifulSoup(r.text,'html.parser').title
        title_str = title_tag.string if title_tag else 'N/A'
        print(f"{turl}: {r.status_code} | Title: {title_str}")
        if r.status_code == 200 and 'cloudflare' not in r.text.lower() and 'Moment' not in r.text:
            soup = BeautifulSoup(r.text, 'html.parser')
            print(f"  SUCCESS! {len(r.text)} chars")
            # Listings suchen
            for sel in ['.property', '.listing', 'article', '.expose', '[class*="expose"]', '[class*="object"]', '[class*="immo"]']:
                items = soup.select(sel)
                if items:
                    print(f"  Selector '{sel}': {len(items)} items")
            # Links zu Exposes
            expose_links = [a['href'] for a in soup.find_all('a', href=True) if 'expose' in a['href'].lower()]
            print(f"  Expose-Links: {len(expose_links)}")
            for el in expose_links[:5]:
                print(f"    {el}")
        else:
            # Cloudflare check
            if 'challenge' in r.text.lower() or 'turnstile' in r.text.lower():
                print(f"  ⚠️ Cloudflare Challenge erkannt")
            elif 'Moment' in r.text:
                print(f"  ⚠️ Cloudflare 'Just a moment' erkannt")
    except Exception as e:
        print(f"Error: {e}")
    time.sleep(1)

# === Versuch 3: API-Endpoint direkt ===
print("\n=== Von Poll API ===")
api_attempts = [
    'https://www.von-poll.com/api/properties?location=mallorca&country=es',
    'https://www.von-poll.com/de/api/search?location=mallorca',
    'https://api.von-poll.com/properties?region=mallorca',
    'https://www.von-poll.com/de/expose/search?region=mallorca',
    'https://www.von-poll.com/api/expose?region=balearen',
]
for api_url in api_attempts:
    try:
        r = requests.get(api_url, headers={**headers, 'Accept': 'application/json'}, timeout=10)
        print(f"{api_url}: {r.status_code} | {r.text[:200]}")
    except Exception as e:
        print(f"Error: {e}")
    time.sleep(0.3)

# === Versuch 4: Google-Cache / alternative Quellen ===
print("\n=== Alternative: Von Poll Expose-URLs via bekannte Muster ===")
# Manche Seiten haben numerische IDs in URLs wie /de/expose/12345
# Testen ob ein bekanntes Format existiert
expose_test_urls = [
    'https://www.von-poll.com/de/expose/3-zimmer-villa-mallorca-1234567',
    'https://www.von-poll.com/de/immobilien/1234567',
]
for eurl in expose_test_urls:
    try:
        r = requests.get(eurl, headers=headers, timeout=10)
        print(f"{eurl}: {r.status_code}")
    except Exception as e:
        print(f"Error: {e}")

# Wenn Property-URLs gefunden: einzeln abrufen
count = 0
print(f"\n=== Verarbeite {len(property_urls)} Mallorca-URLs ===")
for prop_url in property_urls[:100]:
    try:
        r = requests.get(prop_url, headers=headers, timeout=15)
        if r.status_code == 200 and 'Moment' not in r.text and 'challenge' not in r.text.lower():
            soup = BeautifulSoup(r.text, 'html.parser')
            titel = soup.find('h1')
            titel_text = titel.get_text(strip=True) if titel else 'Von Poll Objekt'
            
            # Preis
            preis = None
            preis_match = re.search(r'([\d\.]+)\s*€', r.text)
            if preis_match:
                try: preis = int(preis_match.group(1).replace('.',''))
                except: pass
            
            if prop_url not in existing_urls:
                ws.append([titel_text[:100], 'Von Poll Real Estate', prop_url, preis, 
                          None, None, None, 'Mallorca', str(date.today()), 'Neu'])
                existing_urls.add(prop_url)
                count += 1
                print(f"  ✅ {titel_text[:60]} | {preis}€")
        else:
            print(f"  ❌ {prop_url}: {r.status_code} / CF-Block")
        time.sleep(0.3)
    except Exception as e:
        print(f"  Error {prop_url}: {e}")

wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
print(f"\n✅ Von Poll: {count} Objekte gespeichert")
