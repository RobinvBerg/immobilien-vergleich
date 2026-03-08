#!/usr/bin/env python3
"""Scraper for Pollentia Properties - all 85 pages"""
import requests, re, time
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import date

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
BASE_URL = 'https://www.pollentiaproperties.com'
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
}
SOURCE = 'Pollentia Properties'

session = requests.Session()
session.headers.update(HEADERS)
# First get a cookie
session.get(f'{BASE_URL}/')

def parse_price(text):
    m = re.search(r'Price:\s*([\d,\.]+)\s*€', text)
    if m:
        p = m.group(1).replace('.', '').replace(',', '')
        try: return int(float(p))
        except: pass
    return None

def parse_int(text, label):
    m = re.search(label + r'[:\s]+(\d+)', text, re.IGNORECASE)
    if m:
        try: return int(m.group(1))
        except: pass
    return None

def parse_sqm(text, label):
    m = re.search(label + r'[:\s]+([\d,\.]+)\s*m', text, re.IGNORECASE)
    if m:
        try: return float(m.group(1).replace(',', '.'))
        except: pass
    return None

def scrape_all():
    all_objects = []
    max_page = 85
    
    for page in range(1, max_page + 1):
        url = f'{BASE_URL}/en/properties/type/0?page={page}'
        if page == 1:
            url = f'{BASE_URL}/en/results?sf=true&lang=en&op=buy&propertyTypes=0&page=1'
        
        try:
            r = session.get(url, timeout=20)
            if r.status_code != 200:
                print(f"  Page {page}: HTTP {r.status_code}")
                if r.status_code == 404:
                    break
                time.sleep(2)
                continue
            
            soup = BeautifulSoup(r.text, 'html.parser')
            articles = soup.find_all('article', class_='ic-hilite')
            
            if not articles:
                print(f"  Page {page}: No articles found")
                # Check for end of pagination
                if page > 10:
                    break
                continue
            
            page_objects = []
            for art in articles:
                link = art.find('a', href=re.compile(r'/property/id/'))
                if not link:
                    continue
                
                href = link.get('href', '')
                full_url = BASE_URL + href if not href.startswith('http') else href
                text = art.get_text(separator=' | ', strip=True)
                
                # Extract location from URL slug
                ort = 'Mallorca'
                loc_match = re.search(r'/id/\d+-(?:\w+-for-sale-in-|[a-z]+-for-sale-)([a-z-]+)', href)
                if loc_match:
                    ort = loc_match.group(1).replace('-', ' ').title()
                
                # Also check the location in text (after URL title, there's often a location)
                # Text format: "Title | Location | Bedrooms: X | Price: X €"
                parts = [p.strip() for p in text.split('|')]
                title = ''
                for p in parts:
                    if 'for sale' in p.lower() or 'for rent' in p.lower():
                        title = p
                        break
                if not title and parts:
                    title = parts[0]
                
                # Find location - usually 2nd or 3rd part after title
                title_idx = parts.index(title) if title in parts else -1
                if title_idx >= 0 and title_idx + 1 < len(parts):
                    potential_loc = parts[title_idx + 1]
                    if potential_loc and len(potential_loc) < 50 and not any(k in potential_loc.lower() for k in ['bedroom', 'bathroom', 'price', 'built', 'plot', 'm²', 'new', 'id:']):
                        ort = potential_loc.strip()
                
                price = parse_price(text)
                zimmer = parse_int(text, 'Bedrooms')
                wohnflaeche = parse_sqm(text, 'Built space')
                grundstueck = parse_sqm(text, 'Plot size')
                
                page_objects.append({
                    'titel': title[:200] if title else href.split('/')[-1].replace('-', ' ').title()[:100],
                    'quelle': SOURCE,
                    'url': full_url,
                    'preis': price,
                    'zimmer': zimmer,
                    'grundstueck': int(grundstueck) if grundstueck else None,
                    'wohnflaeche': int(wohnflaeche) if wohnflaeche else None,
                    'ort': ort,
                })
            
            all_objects.extend(page_objects)
            print(f"  Page {page}/{max_page}: {len(page_objects)} props (total: {len(all_objects)})")
            
            time.sleep(0.5)
            
        except Exception as e:
            print(f"  Page {page} error: {e}")
            time.sleep(2)
    
    return all_objects

def save(objects):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
    new_count = 0
    for obj in objects:
        url = obj.get('url', '—')
        if url != '—' and url in existing:
            continue
        ws.append([
            obj.get('titel', ''), obj.get('quelle', SOURCE), url,
            obj.get('preis'), obj.get('zimmer'), obj.get('grundstueck'),
            obj.get('wohnflaeche'), obj.get('ort', ''), str(date.today()), 'Neu'
        ])
        if url != '—': existing.add(url)
        new_count += 1
    wb.save(EXCEL_PATH)
    print(f"✅ {SOURCE}: {new_count} neue Objekte gespeichert (von {len(objects)} gefunden)")
    return new_count

if __name__ == '__main__':
    print(f"=== {SOURCE} Scraper ===")
    objects = scrape_all()
    save(objects)
