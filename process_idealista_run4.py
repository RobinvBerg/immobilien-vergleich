#!/usr/bin/env python3
"""Process Apify Idealista Run 4 - Wait for completion, download, filter, update Excel."""

import time
import json
import requests
from datetime import date
from pathlib import Path
import openpyxl

API_TOKEN = "apify_api_feD2KhARHjtuV9CrSwOReYgoePFSF44nsDL6"
RUN_ID = "DxhXYXmhNlUWSo5k5"
DATASET_ID = "Wb9Ip3yCbN3lEwAHY"
ACT_ID = "igolaizola~idealista-scraper"

RUN_URL = f"https://api.apify.com/v2/acts/{ACT_ID}/runs/{RUN_ID}"
DATASET_URL = f"https://api.apify.com/v2/datasets/{DATASET_ID}/items?limit=500&format=json"

EXCEL_PATH = Path("/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx")
RAW_JSON_PATH = Path("/Users/robin/.openclaw/workspace/mallorca-projekt/idealista_run3_raw.json")

HEADERS = {"Authorization": f"Bearer {API_TOKEN}"}

EXCLUDED_TYPES = {"flat", "penthouse", "studio", "duplex"}
HEADERS_ROW = ["Titel", "Quelle", "URL", "Preis (€)", "Zimmer", "Grundstück (m²)", "Wohnfläche (m²)", "Ort / Gemeinde", "Gefunden am", "Status"]

# ── Step 1: Wait for SUCCEEDED ──────────────────────────────────────────────
print("⏳ Waiting for run to SUCCEED (max 15 min)...")
max_wait = 15 * 60
interval = 30
waited = 0
status = None

while waited <= max_wait:
    resp = requests.get(RUN_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    status = data.get("data", {}).get("status", "")
    print(f"  [{waited}s] Status: {status}")
    if status == "SUCCEEDED":
        print("✅ Run SUCCEEDED!")
        break
    elif status in ("FAILED", "ABORTED", "TIMED-OUT"):
        print(f"❌ Run ended with status: {status}")
        exit(1)
    time.sleep(interval)
    waited += interval
else:
    print("❌ Timeout after 15 minutes")
    exit(1)

# ── Step 2: Download items ───────────────────────────────────────────────────
print(f"\n📥 Downloading dataset items from {DATASET_ID}...")
resp = requests.get(DATASET_URL, headers=HEADERS, timeout=60)
resp.raise_for_status()
items = resp.json()
print(f"  Total items received: {len(items)}")

# Save raw JSON
RAW_JSON_PATH.write_text(json.dumps(items, ensure_ascii=False, indent=2))
print(f"  Raw JSON saved to: {RAW_JSON_PATH}")

# ── Step 3: Verify location ──────────────────────────────────────────────────
location_ok = False
for item in items:
    province = str(item.get("province", "")).lower()
    municipality = str(item.get("municipality", "")).lower()
    if "balears" in province or "mallorca" in municipality:
        location_ok = True
        break

if not location_ok:
    print("\n⚠️  NO items match Mallorca/Balears location check → Excel NOT modified")
    print(f"Provinces found: {list(set(i.get('province','') for i in items[:20]))}")
    exit(0)

print(f"✅ Location check PASSED (province=Balears or municipality=Mallorca found)")

# ── Step 4: Filter ───────────────────────────────────────────────────────────
def get_price(item):
    p = item.get("price")
    if p is None:
        return None
    try:
        return int(p)
    except (ValueError, TypeError):
        return None

def get_rooms(item):
    r = item.get("rooms")
    if r is None:
        return None
    try:
        return int(r)
    except (ValueError, TypeError):
        return None

filtered = []
for item in items:
    price = get_price(item)
    rooms = get_rooms(item)
    ptype = str(item.get("propertyType", "")).lower().strip()

    if price is None or rooms is None:
        continue
    if price < 2_000_000 or price > 6_500_000:
        continue
    if rooms < 5:
        continue
    if ptype in EXCLUDED_TYPES:
        continue
    filtered.append(item)

print(f"\n🔍 After filter: {len(filtered)} items (from {len(items)} total)")

# ── Step 5: Load Excel and insert new rows ───────────────────────────────────
if EXCEL_PATH.exists():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    print(f"  Opened existing Excel: {EXCEL_PATH}")
    # Check/fix header
    if ws.max_row == 0 or ws.cell(1, 1).value != "Titel":
        ws.insert_rows(1)
        for col, header in enumerate(HEADERS_ROW, start=1):
            ws.cell(1, col, header)
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mallorca Markt"
    for col, header in enumerate(HEADERS_ROW, start=1):
        ws.cell(1, col, header)
    print(f"  Created new Excel: {EXCEL_PATH}")

# Collect existing URLs to avoid duplicates
existing_urls = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[2]:  # URL column (index 2)
        existing_urls.add(str(row[2]).strip())

print(f"  Existing rows in Excel: {ws.max_row - 1}, unique URLs: {len(existing_urls)}")

today_str = date.today().strftime("%Y-%m-%d")
added = 0

for item in filtered:
    prop_code = item.get("propertyCode", "")
    url = f"https://www.idealista.com/inmueble/{prop_code}/"

    if url in existing_urls:
        continue

    address = item.get("address", "")
    ptype = item.get("propertyType", "")
    title = f"{address} – {ptype}"
    price = get_price(item)
    rooms = get_rooms(item)

    # Try various field names for land/living area
    land_area = item.get("plotArea") or item.get("landArea") or item.get("plot") or item.get("plotAreaInSquareMeters") or ""
    living_area = item.get("size") or item.get("livingArea") or item.get("constructedArea") or item.get("sizeInSquareMeters") or ""

    municipality = item.get("municipality") or item.get("district") or item.get("neighborhood") or ""

    ws.append([
        title,          # Titel
        "Idealista",    # Quelle
        url,            # URL
        price,          # Preis (€)
        rooms,          # Zimmer
        land_area,      # Grundstück (m²)
        living_area,    # Wohnfläche (m²)
        municipality,   # Ort / Gemeinde
        today_str,      # Gefunden am
        "Neu – Idealista",  # Status
    ])
    existing_urls.add(url)
    added += 1

wb.save(EXCEL_PATH)
print(f"\n✅ Excel updated: {added} new rows added → {EXCEL_PATH}")

# ── Summary ──────────────────────────────────────────────────────────────────
print(f"""
╔════════════════════════════════════════╗
║  ZUSAMMENFASSUNG                       ║
╠════════════════════════════════════════╣
║  Standort OK?   : {'✅ JA' if location_ok else '❌ NEIN'}
║  Items total    : {len(items)}
║  Nach Filter    : {len(filtered)}
║  In Excel neu   : {added}
╚════════════════════════════════════════╝
""")
