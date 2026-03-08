#!/usr/bin/env python3
"""Phase 4 Scraper - Die letzten 12 Quellen"""

import sys
import json
import time
import requests
from datetime import date
from openpyxl import load_workbook

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
APIFY_TOKEN = 'apify_api_feD2KhARHjtuV9CrSwOReYgoePFSF44nsDL6'

def load_existing_urls():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    return set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')

def save_to_excel(new_objects, source_name):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
    
    new_count = 0
    for obj in new_objects:
        url = obj.get('url', '—')
        if url != '—' and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', ''),
            obj.get('quelle', source_name),
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            str(date.today()),
            'Neu'
        ])
        if url != '—':
            existing_urls.add(url)
        new_count += 1
    
    wb.save(EXCEL_PATH)
    print(f"  ✅ {new_count} neue Objekte aus {source_name} gespeichert")
    return new_count


def run_apify_actor(actor_id, run_input, timeout=120):
    """Run Apify actor and return results"""
    url = f"https://api.apify.com/v2/acts/{actor_id}/runs?token={APIFY_TOKEN}"
    resp = requests.post(url, json=run_input, timeout=30)
    if resp.status_code not in (200, 201):
        print(f"  ❌ Apify run failed: {resp.status_code} - {resp.text[:200]}")
        return None
    
    run_data = resp.json()
    run_id = run_data.get('data', {}).get('id')
    if not run_id:
        print(f"  ❌ No run ID: {run_data}")
        return None
    
    print(f"  ⏳ Apify run {run_id} gestartet...")
    
    # Poll for completion
    deadline = time.time() + timeout
    while time.time() < deadline:
        time.sleep(5)
        status_url = f"https://api.apify.com/v2/actor-runs/{run_id}?token={APIFY_TOKEN}"
        sr = requests.get(status_url, timeout=10)
        status = sr.json().get('data', {}).get('status', '')
        print(f"  Status: {status}")
        if status == 'SUCCEEDED':
            # Get dataset
            ds_id = sr.json()['data']['defaultDatasetId']
            items_url = f"https://api.apify.com/v2/datasets/{ds_id}/items?token={APIFY_TOKEN}&limit=500"
            ir = requests.get(items_url, timeout=30)
            return ir.json()
        elif status in ('FAILED', 'ABORTED', 'TIMED-OUT'):
            print(f"  ❌ Run {status}")
            return None
    
    print(f"  ❌ Timeout nach {timeout}s")
    return None


# ============================================================
# SOURCE 1: Rightmove
# ============================================================
def scrape_rightmove():
    print("\n🏠 SOURCE 1: Rightmove")
    
    # Try Apify first
    search_url = f"https://api.apify.com/v2/store?token={APIFY_TOKEN}&search=rightmove&limit=5"
    try:
        resp = requests.get(search_url, timeout=10)
        actors = resp.json().get('data', {}).get('items', [])
        for a in actors:
            print(f"  Actor: {a.get('username')}/{a.get('name')} - {a.get('title')}")
    except Exception as e:
        print(f"  Search failed: {e}")
    
    # Direct API scraping
    # Rightmove international property search
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': 'application/json',
        'Referer': 'https://www.rightmove.co.uk/',
    }
    
    objects = []
    try:
        # Rightmove overseas property API
        # Location identifier for Mallorca
        url = 'https://www.rightmove.co.uk/api/_search?locationIdentifier=OVERSEAS%5E916&numberOfPropertiesPerPage=48&radius=0.0&sortType=2&index=0&propertyTypes=&includeSSTC=false&viewType=LIST&channel=BUY&areaSizeUnit=sqm&currencyCode=EUR&isFetching=false'
        resp = requests.get(url, headers=headers, timeout=15)
        print(f"  Rightmove API status: {resp.status_code}")
        if resp.status_code == 200:
            data = resp.json()
            props = data.get('properties', [])
            print(f"  Gefunden: {len(props)} Objekte")
            for p in props:
                objects.append({
                    'titel': p.get('displayAddress', ''),
                    'quelle': 'Rightmove',
                    'url': 'https://www.rightmove.co.uk' + p.get('propertyUrl', ''),
                    'preis': p.get('price', {}).get('amount'),
                    'zimmer': p.get('bedrooms'),
                    'grundstueck': None,
                    'wohnflaeche': None,
                    'ort': p.get('displayAddress', '').split(',')[-1].strip(),
                })
    except Exception as e:
        print(f"  ❌ Rightmove API error: {e}")
    
    if objects:
        # Get more pages
        try:
            for idx in range(48, 240, 48):
                url = f'https://www.rightmove.co.uk/api/_search?locationIdentifier=OVERSEAS%5E916&numberOfPropertiesPerPage=48&radius=0.0&sortType=2&index={idx}&propertyTypes=&includeSSTC=false&viewType=LIST&channel=BUY&areaSizeUnit=sqm&currencyCode=EUR&isFetching=false'
                resp = requests.get(url, headers=headers, timeout=15)
                if resp.status_code == 200:
                    data = resp.json()
                    props = data.get('properties', [])
                    if not props:
                        break
                    for p in props:
                        objects.append({
                            'titel': p.get('displayAddress', ''),
                            'quelle': 'Rightmove',
                            'url': 'https://www.rightmove.co.uk' + p.get('propertyUrl', ''),
                            'preis': p.get('price', {}).get('amount'),
                            'zimmer': p.get('bedrooms'),
                            'grundstueck': None,
                            'wohnflaeche': None,
                            'ort': p.get('displayAddress', '').split(',')[-1].strip(),
                        })
                    print(f"  Seite index={idx}: +{len(props)}")
                time.sleep(1)
        except Exception as e:
            print(f"  Paging error: {e}")
        
        return save_to_excel(objects, 'Rightmove')
    
    return 0


# ============================================================
# SOURCE 2: Fotocasa
# ============================================================
def scrape_fotocasa():
    print("\n🏠 SOURCE 2: Fotocasa")
    
    headers = {
        'Content-Type': 'application/json',
        'Origin': 'https://www.fotocasa.es',
        'Referer': 'https://www.fotocasa.es/',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'es-ES,es;q=0.9',
    }
    
    objects = []
    
    try:
        # Try fotocasa API
        for page in range(1, 10):
            payload = {
                "culture": "es-ES",
                "includePurchaseTypeFacets": True,
                "isNeedMapCoordinates": False,
                "isNeedOpenHouseAds": False,
                "locationIds": [724],  # Mallorca
                "maxItems": 40,
                "page": page,
                "periodicity": 0,
                "propertyTypeIds": [2],  # Houses
                "purchaseTypeIds": [1],  # Sale
                "sortBy": "score",
                "transactionTypeIds": [1]
            }
            
            resp = requests.post(
                'https://search.gw.fotocasa.es/v2/propertysearch/search',
                json=payload,
                headers=headers,
                timeout=15
            )
            print(f"  Page {page}: status {resp.status_code}")
            
            if resp.status_code != 200:
                print(f"  Response: {resp.text[:300]}")
                break
            
            data = resp.json()
            items = data.get('realEstates', [])
            if not items:
                print(f"  Keine weiteren Items auf Seite {page}")
                break
            
            for item in items:
                price_info = item.get('transactions', [{}])[0] if item.get('transactions') else {}
                price = price_info.get('value', [None])[0] if price_info.get('value') else None
                
                features = {f.get('key'): f.get('value') for f in item.get('features', [])}
                
                url_base = 'https://www.fotocasa.es'
                prop_url = item.get('detail', {}).get('es', '')
                
                objects.append({
                    'titel': item.get('address', {}).get('ubication', ''),
                    'quelle': 'Fotocasa',
                    'url': url_base + prop_url if prop_url else '—',
                    'preis': price,
                    'zimmer': features.get('roomsNumber'),
                    'grundstueck': features.get('surface'),
                    'wohnflaeche': features.get('constructedArea') or features.get('surface'),
                    'ort': item.get('address', {}).get('municipality', ''),
                })
            
            print(f"  +{len(items)} Items, gesamt: {len(objects)}")
            time.sleep(1)
            
            total = data.get('totalAdCount', 0)
            if len(objects) >= total or len(objects) >= 400:
                break
    
    except Exception as e:
        print(f"  ❌ Fotocasa error: {e}")
        import traceback; traceback.print_exc()
    
    if objects:
        return save_to_excel(objects, 'Fotocasa')
    return 0


# ============================================================
# SOURCE 3: Properstar
# ============================================================
def scrape_properstar():
    print("\n🏠 SOURCE 3: Properstar")
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': 'application/json',
        'Content-Type': 'application/json',
        'Origin': 'https://www.properstar.com',
        'Referer': 'https://www.properstar.com/',
    }
    
    objects = []
    
    try:
        # Try the Properstar listing API
        for page in range(1, 8):
            payload = {
                "countryCode": "ES",
                "regionName": "Balearic Islands",
                "listingType": "sale",
                "pageNumber": page,
                "pageSize": 48,
                "sortBy": "relevance"
            }
            
            resp = requests.post(
                'https://listing-api.properstar.com/api/Listing/search',
                json=payload,
                headers=headers,
                timeout=15
            )
            print(f"  Page {page}: status {resp.status_code}")
            
            if resp.status_code != 200:
                print(f"  Response: {resp.text[:300]}")
                break
            
            data = resp.json()
            items = data.get('listings', data.get('items', data.get('results', [])))
            if not items and isinstance(data, list):
                items = data
            
            if not items:
                print(f"  Keys: {list(data.keys()) if isinstance(data, dict) else 'list'}")
                break
            
            for item in items:
                objects.append({
                    'titel': item.get('title', item.get('address', '')),
                    'quelle': 'Properstar',
                    'url': item.get('url', item.get('link', '—')),
                    'preis': item.get('price', item.get('priceValue')),
                    'zimmer': item.get('bedrooms', item.get('rooms')),
                    'grundstueck': item.get('landArea', item.get('plotSize')),
                    'wohnflaeche': item.get('area', item.get('livingArea', item.get('surface'))),
                    'ort': item.get('city', item.get('location', '')),
                })
            
            print(f"  +{len(items)} Items")
            time.sleep(1)
    
    except Exception as e:
        print(f"  ❌ Properstar error: {e}")
    
    if objects:
        return save_to_excel(objects, 'Properstar')
    return 0


# ============================================================
# Run all sources
# ============================================================
if __name__ == '__main__':
    results = {}
    
    results['Rightmove'] = scrape_rightmove()
    results['Fotocasa'] = scrape_fotocasa()
    results['Properstar'] = scrape_properstar()
    
    print("\n=== ZWISCHENERGEBNIS API-QUELLEN ===")
    for src, count in results.items():
        print(f"  {src}: {count} Objekte")
