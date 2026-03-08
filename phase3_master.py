#!/usr/bin/env python3
"""
Mallorca Scraper Phase 3 - Master Script
Banken, Auktionen, Developer, Privat
"""

import sys
import os
sys.path.insert(0, '/Users/robin/.openclaw/workspace/mallorca-projekt')

from openpyxl import load_workbook
from datetime import date
import traceback

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'

def save_to_excel(new_objects, source_name=""):
    """Speichere neue Objekte in Excel mit De-Duplizierung"""
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    
    existing_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            existing_urls.add(str(row[2]).strip())
    
    existing_titles = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing_titles.add(str(row[0]).strip().lower())
    
    new_count = 0
    skip_count = 0
    
    for obj in new_objects:
        url = obj.get('url', '')
        titel = obj.get('titel', '')
        
        # URL-Duplikat
        if url and url in existing_urls:
            skip_count += 1
            continue
        
        # Titel-Duplikat (Fuzzy)
        if titel and titel.lower() in existing_titles:
            skip_count += 1
            continue
        
        ws.append([
            titel or '—',
            obj.get('quelle', source_name or '—'),
            url or '—',
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', '—'),
            str(date.today()),
            'Neu',
        ])
        
        if url:
            existing_urls.add(url)
        if titel:
            existing_titles.add(titel.lower())
        new_count += 1
    
    wb.save(EXCEL_PATH)
    print(f"  ✅ Gespeichert: {new_count} neue | Übersprungen: {skip_count} Duplikate")
    return new_count

def main():
    results = {}
    total_new = 0
    
    # === 1. BOE SUBASTAS ===
    print("\n" + "="*60)
    print("1. BOE SUBASTAS (Zwangsversteigerungen Baleares)")
    print("="*60)
    try:
        from scraper_boe_final import main as boe_main
        boe_objects = boe_main()
        results['BOE Subastas'] = {'found': len(boe_objects), 'saved': 0, 'error': None}
        if boe_objects:
            saved = save_to_excel(boe_objects, 'BOE Subastas')
            results['BOE Subastas']['saved'] = saved
            total_new += saved
        print(f"  -> Gefunden: {len(boe_objects)}")
    except Exception as e:
        print(f"  ❌ Fehler: {e}")
        traceback.print_exc()
        results['BOE Subastas'] = {'found': 0, 'saved': 0, 'error': str(e)}
    
    # === 2. SAREB ===
    print("\n" + "="*60)
    print("2. SAREB (Spanische Bad Bank)")
    print("="*60)
    try:
        from scraper_sareb import main as sareb_main
        sareb_objects = sareb_main()
        results['SAREB'] = {'found': len(sareb_objects), 'saved': 0, 'error': None}
        if sareb_objects:
            saved = save_to_excel(sareb_objects, 'SAREB')
            results['SAREB']['saved'] = saved
            total_new += saved
        print(f"  -> Gefunden: {len(sareb_objects)}")
    except Exception as e:
        print(f"  ❌ Fehler: {e}")
        results['SAREB'] = {'found': 0, 'saved': 0, 'error': str(e)}
    
    # === 3. BANKPORTALE ===
    print("\n" + "="*60)
    print("3. BANKPORTALE (Servihabitat, Haya, Solvia, Sabadell, imaginedge)")
    print("="*60)
    try:
        from scraper_bankportale import main as bank_main
        bank_objects = bank_main()
        results['Bankportale'] = {'found': len(bank_objects), 'saved': 0, 'error': None}
        if bank_objects:
            saved = save_to_excel(bank_objects)
            results['Bankportale']['saved'] = saved
            total_new += saved
        print(f"  -> Gefunden: {len(bank_objects)}")
    except Exception as e:
        print(f"  ❌ Fehler: {e}")
        traceback.print_exc()
        results['Bankportale'] = {'found': 0, 'saved': 0, 'error': str(e)}
    
    # === 4. DEVELOPER ===
    print("\n" + "="*60)
    print("4. DEVELOPER (Taylor Wimpey, Vives Pons, Barrau)")
    print("="*60)
    try:
        from scraper_developers import main as dev_main
        dev_objects = dev_main()
        results['Developer'] = {'found': len(dev_objects), 'saved': 0, 'error': None}
        if dev_objects:
            saved = save_to_excel(dev_objects)
            results['Developer']['saved'] = saved
            total_new += saved
        print(f"  -> Gefunden: {len(dev_objects)}")
    except Exception as e:
        print(f"  ❌ Fehler: {e}")
        traceback.print_exc()
        results['Developer'] = {'found': 0, 'saved': 0, 'error': str(e)}
    
    # === 5. PRIVAT ===
    print("\n" + "="*60)
    print("5. PRIVAT (Wallapop, Milanuncios)")
    print("="*60)
    try:
        from scraper_privat import main as privat_main
        privat_objects = privat_main()
        results['Privat'] = {'found': len(privat_objects), 'saved': 0, 'error': None}
        if privat_objects:
            saved = save_to_excel(privat_objects)
            results['Privat']['saved'] = saved
            total_new += saved
        print(f"  -> Gefunden: {len(privat_objects)}")
    except Exception as e:
        print(f"  ❌ Fehler: {e}")
        results['Privat'] = {'found': 0, 'saved': 0, 'error': str(e)}
    
    # === ZUSAMMENFASSUNG ===
    print("\n" + "="*60)
    print("📊 ZUSAMMENFASSUNG PHASE 3")
    print("="*60)
    
    for src, data in results.items():
        status = "✅" if not data['error'] else "❌"
        print(f"{status} {src}: {data['found']} gefunden, {data['saved']} neu gespeichert")
        if data['error']:
            print(f"    Fehler: {data['error'][:80]}")
    
    print(f"\n🎯 Gesamt neue Objekte: {total_new}")
    
    # Aktueller Excel-Stand
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    total_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in _))
    print(f"📋 Excel-Stand: {total_rows} Objekte gesamt")
    
    return results, total_new

if __name__ == '__main__':
    main()
