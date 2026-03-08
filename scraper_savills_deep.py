#!/usr/bin/env python3
"""Savills deep scraper - intercept Map/Search response"""
import json, time, re
from datetime import date
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'

def save_to_excel(new_objects, source_name):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
    new_count = 0
    for obj in new_objects:
        url = obj.get('url', '—')
        if url != '—' and url in existing_urls: continue
        ws.append([obj.get('titel',''), obj.get('quelle', source_name), url,
                   obj.get('preis'), obj.get('zimmer'), obj.get('grundstueck'),
                   obj.get('wohnflaeche'), obj.get('ort',''), str(date.today()), 'Neu'])
        if url != '—': existing_urls.add(url)
        new_count += 1
    wb.save(EXCEL_PATH)
    print(f"  ✅ {new_count} neue Objekte aus {source_name} gespeichert")
    return new_count

# ===== SAVILLS =====
def scrape_savills():
    print("\n🏠 Savills - Playwright mit langer Wartezeit")
    objects = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=['--no-sandbox'])
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120',
            viewport={'width': 1280, 'height': 900},
            locale='en-US',
        )
        page = context.new_page()
        Stealth().apply_stealth_sync(page)
        
        map_search_data = []
        search_by_url_data = []
        
        def on_response(resp):
            url = resp.url
            if 'livev6-searchapi.savills.com' in url:
                try:
                    if resp.status == 200:
                        body = resp.body()
                        print(f"  Savills API captured: {url} ({len(body)}b)")
                        if 'Map/Search' in url:
                            map_search_data.append(json.loads(body))
                        elif 'SearchByUrl' in url:
                            search_by_url_data.append(json.loads(body))
                except Exception as e:
                    print(f"  Response error: {e}")
        
        def on_request(req):
            url = req.url
            if 'livev6-searchapi.savills.com' in url:
                try:
                    body = req.post_data
                    if body:
                        print(f"  Request to: {url} | Body: {body[:300]}")
                except: pass
        
        page.on('response', on_response)
        page.on('request', on_request)
        
        try:
            print("  Loading Savills Mallorca search...")
            page.goto('https://search.savills.com/es/en/list?SearchList=IsInRegion_EUR5002&SaleType=ForSale&PropertyType=RES',
                     wait_until='domcontentloaded', timeout=30000)
            
            # Wait for the page to fully render
            time.sleep(8)
            
            # Cookie consent
            try:
                page.click('#onetrust-accept-btn-handler', timeout=4000)
                time.sleep(2)
            except: pass
            
            # Wait more for JS
            time.sleep(5)
            
            html = page.content()
            print(f"  HTML: {len(html)}b, Map/Search calls: {len(map_search_data)}, SearchByUrl: {len(search_by_url_data)}")
            print(f"  Title: {page.title()}")
            
            # Look for properties in the DOM
            items = page.query_selector_all('[class*="ResultsGridcard"], [class*="PropertyCard"], [class*="listing-card"]')
            print(f"  DOM items: {len(items)}")
            
            # Process Map/Search data
            for data in map_search_data:
                try:
                    results = data.get('Results', {})
                    inside = results.get('Inside', {})
                    outside = results.get('Outside', {})
                    
                    all_props = []
                    for region_data in [inside, outside]:
                        for prop in region_data.get('Properties', []):
                            all_props.append(prop)
                        for stack in region_data.get('StackedProperties', []):
                            for prop in stack.get('Properties', []):
                                all_props.append(prop)
                    
                    print(f"  Found {len(all_props)} properties in Map/Search")
                    
                    for prop in all_props:
                        detail_url = prop.get('DetailPageUrl', '')
                        canonical = prop.get('MetaInformation', {}).get('CanonicalUrl', '')
                        if canonical and not canonical.startswith('http'):
                            full_url = f'https://search.savills.com/{canonical}'
                        elif detail_url:
                            full_url = 'https://search.savills.com' + detail_url
                        else:
                            full_url = '—'
                        
                        desc = prop.get('MetaInformation', {}).get('Description', '')
                        
                        # Extract price from description
                        price_match = re.search(r'€[\s]*([\d,\.]+)', desc)
                        price = None
                        if price_match:
                            price = int(price_match.group(1).replace(',', '').replace('.', ''))
                        
                        # Get beds from description
                        beds_match = re.search(r'(\d+)\s*bed', desc, re.I)
                        beds = int(beds_match.group(1)) if beds_match else None
                        
                        objects.append({
                            'titel': desc[:100] if desc else 'Savills Property',
                            'quelle': 'Savills',
                            'url': full_url,
                            'preis': price or prop.get('Price'),
                            'zimmer': beds or prop.get('Bedrooms'),
                            'grundstueck': None,
                            'wohnflaeche': None,
                            'ort': 'Mallorca',
                        })
                except Exception as e:
                    print(f"  Parse error: {e}")
                    import traceback; traceback.print_exc()
            
            # Process SearchByUrl
            for data in search_by_url_data:
                print(f"  SearchByUrl data: {list(data.keys()) if isinstance(data, dict) else type(data)}")
                if isinstance(data, dict):
                    for k in ['SearchParameters', 'searchParameters', 'Filters', 'filters']:
                        if k in data:
                            print(f"  {k}: {json.dumps(data[k], ensure_ascii=False)[:200]}")
            
            # HTML-based extraction if no API data
            if not objects:
                # Try to scrape from loaded HTML
                cards = page.query_selector_all('article, [class*="Card"], [class*="listing"]')
                print(f"  Cards: {len(cards)}")
                
                # Find links
                all_links = page.query_selector_all('a[href*="property-detail"]')
                print(f"  Property links: {len(all_links)}")
                
                seen = set()
                for link in all_links:
                    try:
                        href = link.get_attribute('href') or ''
                        if href in seen: continue
                        seen.add(href)
                        if not href.startswith('http'): href = 'https://search.savills.com' + href
                        text = link.inner_text()[:100]
                        objects.append({
                            'titel': text, 'quelle': 'Savills', 'url': href,
                            'preis': None, 'zimmer': None, 'grundstueck': None,
                            'wohnflaeche': None, 'ort': 'Mallorca',
                        })
                    except: pass
        
        except Exception as e:
            print(f"  Error: {e}")
            import traceback; traceback.print_exc()
        
        browser.close()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Savills')
    return 0


# ===== FOTOCASA - Wait for the search API =====
def scrape_fotocasa():
    print("\n🏠 Fotocasa - Long wait for search API")
    objects = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=['--no-sandbox'])
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120',
            viewport={'width': 1280, 'height': 900},
            locale='es-ES',
        )
        page = context.new_page()
        Stealth().apply_stealth_sync(page)
        
        api_calls = {}
        
        def on_response(resp):
            url = resp.url
            if 'gw.fotocasa.es' in url or 'fotocasa.es' in url:
                try:
                    if resp.status == 200:
                        ct = resp.headers.get('content-type', '')
                        body = resp.body()
                        if len(body) > 200:
                            print(f"  FC API: {url[:100]} ({len(body)}b) ct={ct[:30]}")
                            if 'json' in ct:
                                try:
                                    api_calls[url] = json.loads(body)
                                except: pass
                except Exception as e:
                    print(f"  Resp error: {e}")
        
        def on_request(req):
            url = req.url
            if 'gw.fotocasa.es' in url:
                body = req.post_data
                print(f"  FC Request: {req.method} {url[:100]}{(' | body: ' + str(body)[:100]) if body else ''}")
        
        page.on('response', on_response)
        page.on('request', on_request)
        
        try:
            page.goto('https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l',
                     wait_until='domcontentloaded', timeout=30000)
            
            # Cookie consent - try multiple approaches
            for cookie_sel in ['#didomi-notice-agree-button', '[class*="didomi"] button', 
                               'button[aria-label*="consentement"]', '.fc-button-label']:
                try:
                    page.wait_for_selector(cookie_sel, timeout=3000)
                    page.click(cookie_sel)
                    print(f"  Cookie via {cookie_sel}")
                    break
                except: pass
            
            time.sleep(8)  # Long wait for JS
            
            print(f"  FC API calls captured: {len(api_calls)}")
            
            # Process all captured API calls
            for url, data in api_calls.items():
                print(f"  Processing {url[:60]}: keys={list(data.keys()) if isinstance(data, dict) else 'list'}")
                
                listings = None
                if isinstance(data, dict):
                    for k in ['realEstates', 'results', 'items', 'listings']:
                        if k in data and isinstance(data[k], list) and data[k]:
                            listings = data[k]
                            print(f"  Listings under '{k}': {len(listings)}")
                            break
                
                if listings:
                    for item in listings:
                        price = None
                        txs = item.get('transactions', [])
                        if txs:
                            vals = txs[0].get('value', [])
                            price = vals[0] if vals else None
                        
                        features = {}
                        for f in item.get('features', []):
                            key = f.get('key', '')
                            val = f.get('value', [None])
                            features[key] = val[0] if isinstance(val, list) else val
                        
                        detail = item.get('detail', {})
                        url_v = detail.get('es', '') if isinstance(detail, dict) else ''
                        if url_v and not url_v.startswith('http'):
                            url_v = 'https://www.fotocasa.es' + url_v
                        
                        addr = item.get('address', {})
                        objects.append({
                            'titel': addr.get('ubication', '') if isinstance(addr, dict) else '',
                            'quelle': 'Fotocasa',
                            'url': url_v or '—',
                            'preis': price,
                            'zimmer': features.get('roomsNumber'),
                            'grundstueck': features.get('plotArea'),
                            'wohnflaeche': features.get('constructedArea', features.get('surface')),
                            'ort': addr.get('municipality', '') if isinstance(addr, dict) else '',
                        })
            
            # HTML fallback
            if not objects:
                html = page.content()
                print(f"  HTML: {len(html)}b")
                cards = page.query_selector_all('[class*="re-Card"], [class*="CardList"], article[class*="re-"]')
                print(f"  Cards: {len(cards)}")
                
                # Try to find any links
                links = page.query_selector_all('a[href*="/es/vivienda/"]')
                print(f"  Property links: {len(links)}")
                
                seen = set()
                for link in links[:100]:
                    try:
                        href = link.get_attribute('href') or ''
                        if href in seen: continue
                        seen.add(href)
                        if not href.startswith('http'): href = 'https://www.fotocasa.es' + href
                        objects.append({
                            'titel': link.inner_text()[:100],
                            'quelle': 'Fotocasa', 'url': href,
                            'preis': None, 'zimmer': None, 'grundstueck': None,
                            'wohnflaeche': None, 'ort': 'Mallorca',
                        })
                    except: pass
        
        except Exception as e:
            print(f"  Error: {e}")
            import traceback; traceback.print_exc()
        
        browser.close()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Fotocasa')
    return 0


# ===== RIGHTMOVE - HTML-based =====
def scrape_rightmove():
    print("\n🏠 Rightmove - HTML Analyse")
    objects = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=['--no-sandbox'])
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120',
            viewport={'width': 1280, 'height': 900},
        )
        page = context.new_page()
        Stealth().apply_stealth_sync(page)
        
        api_data = []
        def on_resp(r):
            if 'rightmove' in r.url and r.status == 200:
                ct = r.headers.get('content-type', '')
                if 'json' in ct:
                    try:
                        b = r.body()
                        d = json.loads(b)
                        if any(k in d for k in ['properties', 'searchResults', 'result']):
                            api_data.append(d)
                            print(f"  RM JSON: {r.url[:80]} ({len(b)}b)")
                    except: pass
        page.on('response', on_resp)
        
        try:
            page.goto('https://www.rightmove.co.uk/overseas-property/in-Mallorca.html',
                     wait_until='domcontentloaded', timeout=25000)
            time.sleep(5)
            
            html = page.content()
            print(f"  HTML: {len(html)}b, API: {len(api_data)}")
            
            # Process API data
            for d in api_data:
                props = d.get('properties', d.get('searchResults', []))
                for p in props:
                    url_v = p.get('propertyUrl', p.get('url', '—'))
                    if url_v and not url_v.startswith('http') and url_v != '—':
                        url_v = 'https://www.rightmove.co.uk' + url_v
                    price_d = p.get('price', {})
                    price = price_d.get('amount') if isinstance(price_d, dict) else price_d
                    objects.append({
                        'titel': p.get('displayAddress', ''),
                        'quelle': 'Rightmove', 'url': url_v,
                        'preis': price, 'zimmer': p.get('bedrooms'),
                        'grundstueck': None, 'wohnflaeche': None,
                        'ort': p.get('displayAddress', '').split(',')[-1].strip(),
                    })
            
            # HTML: Look for property cards and links
            if not objects:
                # RM structure: property cards in l-searchResult
                cards = page.query_selector_all('.l-searchResult, [class*="propertyCard"], [data-test="propertyCard"]')
                print(f"  Cards: {len(cards)}")
                
                # Find all property links from HTML
                prop_links = re.findall(r'href=["\'](/overseas-property/property-\d+/type-[^"\']+)["\']', html)
                prop_links2 = re.findall(r'href=["\'](/overseas-property/[^"\']+/[^"\']+\.html)["\']', html)
                print(f"  Property links in HTML: {len(prop_links)} + {len(prop_links2)}")
                
                seen = set()
                for href in prop_links + prop_links2:
                    if href in seen: continue
                    seen.add(href)
                    full_url = 'https://www.rightmove.co.uk' + href
                    objects.append({
                        'titel': 'Rightmove Mallorca',
                        'quelle': 'Rightmove', 'url': full_url,
                        'preis': None, 'zimmer': None, 'grundstueck': None,
                        'wohnflaeche': None, 'ort': 'Mallorca',
                    })
        
        except Exception as e:
            print(f"  Error: {e}")
        
        browser.close()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Rightmove')
    return 0


if __name__ == '__main__':
    results = {}
    results['Savills'] = scrape_savills()
    results['Fotocasa'] = scrape_fotocasa()
    results['Rightmove'] = scrape_rightmove()
    
    print("\n=== ERGEBNIS ===")
    total = 0
    for src, count in results.items():
        print(f"  {'✅' if count else '❌'} {src}: {count}")
        total += count
    print(f"  Total: {total}")
