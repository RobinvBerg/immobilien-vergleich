#!/usr/bin/env python3
"""
Mallorca Bilder-Analyse
- Lädt 3 Bilder pro Objekt (main + 2 details) von den Makler-Seiten
- Analysiert via Anthropic Vision (Haiku)
- Schreibt Charme/Renovierung/Gebäudestruktur/Gästehaus zurück ins Excel
- Speichert Bilder lokal als bilder/ROW_main.jpg, bilder/ROW_detail1.jpg etc.
"""

import json, os, re, time, requests, base64, shutil
from pathlib import Path

# Setup
BASE = Path(__file__).parent
BILDER_DIR = BASE / "bilder"
BILDER_DIR.mkdir(exist_ok=True)
LOG = BASE / "bilder_analyse_log.txt"
CHECKPOINT = BASE / "bilder_analyse_checkpoint.json"
EXCEL = BASE / "mallorca-kandidaten-v2.xlsx"
BILDER_JSON = BASE / "mallorca-bilder.json"

import anthropic
import openpyxl

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

def log(msg):
    print(msg, flush=True)
    with open(LOG, "a") as f:
        f.write(msg + "\n")

def load_checkpoint():
    if CHECKPOINT.exists():
        return json.load(open(CHECKPOINT))
    return {}

def save_checkpoint(data):
    json.dump(data, open(CHECKPOINT, "w"))

def download_image(url, path, timeout=15):
    """Download image, return True if success"""
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
        r = requests.get(url, headers=headers, timeout=timeout, stream=True)
        if r.status_code == 200 and len(r.content) > 1000:
            with open(path, "wb") as f:
                f.write(r.content)
            return True
    except Exception as e:
        log(f"  Download fehler {url[:60]}: {e}")
    return False

def get_extra_images(obj_url, source_img_url, row):
    """Versuche 2-3 Bilder von der Detailseite zu holen"""
    imgs = []
    if source_img_url:
        imgs.append(source_img_url)
    
    # Versuche weitere Bilder zu finden basierend auf der Quelle
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
        
        if "balearic-properties.com" in obj_url:
            r = requests.get(obj_url, headers=headers, timeout=15)
            if r.status_code == 200:
                matches = re.findall(r'https://www\.balearic-properties\.com/property-image/xlarge_[^"\']+', r.text)
                for m in matches[:4]:
                    if m not in imgs:
                        imgs.append(m)
        
        elif "engelvoelkers.com" in obj_url:
            r = requests.get(obj_url, headers=headers, timeout=15)
            if r.status_code == 200:
                matches = re.findall(r'https://[^"\']*cdnev[^"\']*\.jpg[^"\']*', r.text)
                matches += re.findall(r'https://[^"\']*ev-cdn[^"\']*\.jpg[^"\']*', r.text)
                for m in matches[:4]:
                    if m not in imgs:
                        imgs.append(m)
        
        elif "livingblue-mallorca.com" in obj_url or "luxury-estates-mallorca.com" in obj_url:
            r = requests.get(obj_url, headers=headers, timeout=15)
            if r.status_code == 200:
                matches = re.findall(r'https://[^"\']*\.(jpg|jpeg|webp)[^"\']*', r.text, re.I)
                for m, _ in matches[:4]:
                    if m not in imgs and "placeholder" not in m.lower():
                        imgs.append(m)
                        
    except Exception as e:
        log(f"  Extra-Bilder Fehler ({obj_url[:40]}): {e}")
    
    return imgs[:4]  # max 4 Bilder

def encode_image(path):
    """Encode image as base64"""
    with open(path, "rb") as f:
        data = f.read()
    # Detect media type
    if path.suffix.lower() in [".jpg", ".jpeg"]:
        mt = "image/jpeg"
    elif path.suffix.lower() == ".png":
        mt = "image/png"
    elif path.suffix.lower() == ".webp":
        mt = "image/webp"
    else:
        mt = "image/jpeg"
    return base64.standard_b64encode(data).decode("utf-8"), mt

def analyze_images(image_paths, beschreibung=""):
    """Analysiere Bilder via Claude Haiku Vision"""
    content = []
    
    for p in image_paths:
        if os.path.exists(p) and os.path.getsize(p) > 1000:
            try:
                b64, mt = encode_image(Path(p))
                content.append({
                    "type": "image",
                    "source": {"type": "base64", "media_type": mt, "data": b64}
                })
            except Exception as e:
                log(f"  Encode-Fehler {p}: {e}")
    
    if not content:
        return None
    
    prompt = f"""Du bewertest Mallorca Immobilien für einen deutschen Käufer (Familie, 3 Kinder, sucht Finca/Landgut mit großem Grundstück).

Beschreibung: {beschreibung[:300] if beschreibung else 'keine'}

Bewerte auf Basis der Bilder:

1. CHARME (1-5): Ästhetischer Gesamteindruck, Authentizität, Charakter
   1=hässlich/banal, 2=durchschnittlich, 3=ansprechend, 4=schön, 5=außergewöhnlich

2. RENOVIERUNG (0-100): Renovierungsbedarf (höher = mehr Renovierung nötig)
   0=einzugsbereit perfekt, 40=kleinere Arbeiten, 60=mittlere Renovierung, 80=große Renovierung, 100=Rohbau/Abriss

3. GEBÄUDESTRUKTUR: Kurze Beschreibung (z.B. "Rustikale Finca, Haupthaus + Gästehaus, Pool", "Moderne Villa, 2-stöckig", "Herrenhaus, historisch")

4. GÄSTEHAUS: ja/nein/unklar

5. RENO_BEGRUENDUNG: 1 Satz warum dieser Reno-Score

Antworte NUR in diesem Format (keine andere Erklärung):
CHARME: X
RENOVIERUNG: XX
GEBÄUDESTRUKTUR: ...
GÄSTEHAUS: ja/nein/unklar
RENO_BEGRUENDUNG: ..."""

    content.append({"type": "text", "text": prompt})
    
    try:
        response = client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=300,
            messages=[{"role": "user", "content": content}]
        )
        return parse_response(response.content[0].text)
    except Exception as e:
        log(f"  API Fehler: {e}")
        return None

def parse_response(text):
    result = {}
    for line in text.strip().split("\n"):
        if line.startswith("CHARME:"):
            try: result["charme"] = int(line.split(":")[1].strip())
            except: pass
        elif line.startswith("RENOVIERUNG:"):
            try: result["renovierung"] = int(line.split(":")[1].strip())
            except: pass
        elif line.startswith("GEBÄUDESTRUKTUR:"):
            result["gebaeude"] = line.split(":", 1)[1].strip()
        elif line.startswith("GÄSTEHAUS:"):
            result["gaestehaus"] = line.split(":", 1)[1].strip()
        elif line.startswith("RENO_BEGRUENDUNG:"):
            result["reno_begruendung"] = line.split(":", 1)[1].strip()
    return result if result else None

def main():
    log("=== Mallorca Bilder-Analyse gestartet ===")
    
    # Lade Daten
    bilder_data = json.load(open(BILDER_JSON))
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    checkpoint = load_checkpoint()
    
    rows = sorted([int(k) for k in bilder_data.keys()])
    total = len(rows)
    done = len(checkpoint)
    log(f"Gesamt: {total} Objekte, bereits erledigt: {done}")
    
    for i, row in enumerate(rows):
        row_str = str(row)
        if row_str in checkpoint:
            continue
        
        entry = bilder_data[row_str]
        obj_url = entry.get("url", "")
        main_img = entry.get("img_url", "")
        beschreibung = entry.get("desc", "")
        
        log(f"\n[{i+1}/{total}] Zeile {row}: {obj_url[:60]}")
        
        # Bilder holen
        img_urls = get_extra_images(obj_url, main_img, row)
        
        # Bilder herunterladen
        local_paths = []
        labels = ["main", "detail1", "detail2", "detail3"]
        for j, img_url in enumerate(img_urls[:4]):
            ext = ".jpg"
            if ".png" in img_url.lower(): ext = ".png"
            elif ".webp" in img_url.lower(): ext = ".webp"
            local_path = BILDER_DIR / f"{row}_{labels[j]}{ext}"
            if download_image(img_url, str(local_path)):
                local_paths.append(str(local_path))
                log(f"  ✅ {labels[j]}: {img_url[:60]}")
            else:
                log(f"  ❌ {labels[j]} fehlgeschlagen")
        
        if not local_paths:
            log(f"  ⚠️ Keine Bilder — überspringe")
            checkpoint[row_str] = {"skipped": True}
            save_checkpoint(checkpoint)
            continue
        
        # Vision-Analyse
        result = analyze_images(local_paths, beschreibung)
        
        if result:
            # Excel updaten
            if "charme" in result:
                ws.cell(row, 4).value = result["charme"]  # D
            if "renovierung" in result:
                ws.cell(row, 22).value = result["renovierung"]  # V
                ws.cell(row, 32).value = result["renovierung"]  # AF Reno-Score
            if "gebaeude" in result:
                gebaeude = result["gebaeude"]
                if "gaestehaus" in result and result["gaestehaus"] == "ja":
                    if "gästehaus" not in gebaeude.lower() and "gaestehaus" not in gebaeude.lower():
                        gebaeude += " + Gästehaus"
                ws.cell(row, 29).value = gebaeude  # AC
            if "reno_begruendung" in result:
                ws.cell(row, 33).value = result["reno_begruendung"]  # AG
            
            n_bilder = len(local_paths)
            ws.cell(row, 34).value = f"Vision-Analyse ({n_bilder} Bild{'er' if n_bilder>1 else ''})"  # AH
            
            log(f"  → Charme={result.get('charme','?')} Reno={result.get('renovierung','?')} Gebäude={result.get('gebaeude','?')[:40]}")
            checkpoint[row_str] = result
        else:
            log(f"  ⚠️ Keine Analyse-Ergebnis")
            checkpoint[row_str] = {"failed": True}
        
        save_checkpoint(checkpoint)
        
        # Alle 20 Zeilen speichern
        if (i+1) % 20 == 0:
            wb.save(EXCEL)
            log(f"  💾 Excel gespeichert ({i+1}/{total})")
        
        time.sleep(0.5)  # Rate limiting
    
    # Final save
    wb.save(EXCEL)
    done = sum(1 for v in checkpoint.values() if not v.get("skipped") and not v.get("failed"))
    skipped = sum(1 for v in checkpoint.values() if v.get("skipped"))
    failed = sum(1 for v in checkpoint.values() if v.get("failed"))
    log(f"\n=== FERTIG ===")
    log(f"Analysiert: {done}, Übersprungen: {skipped}, Fehler: {failed}")
    log(f"Bilder gespeichert in: {BILDER_DIR}")
    
    # Notify
    os.system('openclaw system event --text "Bilder-Analyse fertig: ' + str(done) + ' Objekte bewertet" --mode now')

if __name__ == "__main__":
    main()
