#!/usr/bin/env python3
"""
Engel & Völkers Mallorca Scraper (BFF API)
Scrapes all Mallorca residential sale properties with ≥5 rooms.
"""
import requests
import time
import random
from datetime import date
from pathlib import Path
import openpyxl

XLSX_PATH = Path("/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx")
SOURCE = "Engel & Völkers"
DETAIL_URL_BASE = "https://www.engelvoelkers.com/de/de/exposes/{id}"
API_URL = "https://search-bff.engelvoelkers.com/api/v2/listing/search"

HEADERS = ["Titel", "Quelle", "URL", "Preis (€)", "Zimmer", "Grundstück (m²)", "Wohnfläche (m²)", "Ort / Gemeinde", "Gefunden am", "Status"]

API_HEADERS = {
    "Content-Type": "application/json",
    "Accept": "application/json",
    "Origin": "https://www.engelvoelkers.com",
    "Referer": "https://www.engelvoelkers.com/",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
}

API_PARAMS_BASE = {
    "language": "de",
    "measurementSystem": "metric",
    "currency": "EUR",
    "marketCountryCode": "DE",
    "limit": 24,
}

SEARCH_BODY = {
    "businessArea": ["residential"],
    "propertyMarketingType": ["sale"],
    "searchRadius": 0,
    "placeId": "ChIJKcEGZna4lxIRwOzSAv-b67c",  # Mallorca, Baleares, Spanien
    "sortingOptions": ["SALES_PRICE_DESC"],
    "rooms": {"min": 5},
}


def get_ort(listing):
    """Extract municipality/city from addressComponents"""
    ac = listing.get("addressComponents", [])
    # Prefer level 4 (municipality), then level 3, then level 2, then city
    priority = {
        "administrative_area_level_4": 0,
        "administrative_area_level_3": 1,
        "administrative_area_level_2": 2,
        "locality": 3,
        "administrative_area_level_1": 4,
    }
    best = None
    best_prio = 99
    for comp in ac:
        ptype = comp.get("placeType", "")
        if ptype in priority and priority[ptype] < best_prio:
            best = comp.get("text", "")
            best_prio = priority[ptype]
    return best or ""


def load_existing_urls():
    if not XLSX_PATH.exists():
        return set()
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    try:
        url_col = header.index("URL")
        src_col = header.index("Quelle")
    except ValueError:
        return set()
    urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[src_col] == SOURCE and row[url_col]:
            urls.add(row[url_col])
    return urls


def append_to_xlsx(new_rows):
    if not XLSX_PATH.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)
    else:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
    for row in new_rows:
        ws.append([
            row.get("titel", ""),
            SOURCE,
            row.get("url", ""),
            row.get("preis", ""),
            row.get("zimmer", ""),
            row.get("grundstueck", ""),
            row.get("wohnflaeche", ""),
            row.get("ort", ""),
            date.today().isoformat(),
            "Neu",
        ])
    wb.save(XLSX_PATH)


def fetch_page(offset):
    params = {**API_PARAMS_BASE, "offset": offset}
    resp = requests.post(API_URL, json=SEARCH_BODY, headers=API_HEADERS, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


def parse_listing(listing):
    uid = listing.get("id", "")
    url = DETAIL_URL_BASE.format(id=uid)
    
    titel = listing.get("profile", {}).get("title", "")
    
    # Price
    preis = listing.get("price", {}).get("salesPrice", {}).get("min")
    if preis and preis > 0:
        preis = int(preis)
    elif listing.get("hasPriceOnRequest"):
        preis = "Preis auf Anfrage"
    else:
        preis = ""
    
    # Rooms
    rooms_val = listing.get("rooms", {})
    zimmer = rooms_val.get("min") if rooms_val else None
    if zimmer:
        zimmer = int(zimmer)
    
    # Areas
    area = listing.get("area", {})
    
    grundstueck = None
    plot = area.get("plotSurface", {})
    if plot and plot.get("min"):
        grundstueck = int(round(plot["min"]))
    
    wohnflaeche = None
    living = area.get("livingSurface", {})
    if living and living.get("min"):
        wohnflaeche = int(round(living["min"]))
    
    ort = get_ort(listing)
    
    return {
        "url": url,
        "titel": titel,
        "preis": preis,
        "zimmer": zimmer,
        "grundstueck": grundstueck,
        "wohnflaeche": wohnflaeche,
        "ort": ort,
    }


def main():
    print("=== Engel & Völkers Mallorca Scraper (BFF API) ===")
    
    existing_urls = load_existing_urls()
    print(f"Bestehende E&V Einträge: {len(existing_urls)}")
    
    # Get total
    first_page = fetch_page(0)
    total = first_page.get("listingsTotal", 0)
    print(f"Gesamte Objekte (≥5 Zimmer, Mallorca): {total}")
    
    all_listings = []
    
    # Process first page
    for item in first_page.get("listings", []):
        listing = item.get("listing", {})
        if listing:
            all_listings.append(parse_listing(listing))
    
    print(f"Seite 1 (offset=0): {len(first_page.get('listings', []))} Objekte")
    
    # Paginate
    offset = 24
    page_size = 24
    
    while offset < total:
        time.sleep(random.uniform(1.5, 3.0))
        print(f"  Fetching offset={offset}...")
        try:
            page_data = fetch_page(offset)
            listings = page_data.get("listings", [])
            if not listings:
                print(f"  Leere Seite bei offset={offset}, Abbruch")
                break
            for item in listings:
                listing = item.get("listing", {})
                if listing:
                    all_listings.append(parse_listing(listing))
            print(f"  offset={offset}: {len(listings)} Objekte (kumulativ: {len(all_listings)})")
            offset += page_size
        except Exception as e:
            print(f"  Fehler bei offset={offset}: {e}")
            break
    
    print(f"\nGesamt gescraped: {len(all_listings)} Objekte")
    
    # Filter duplicates
    new_rows = []
    seen = set()
    
    for p in all_listings:
        url = p["url"]
        if url in existing_urls or url in seen:
            continue
        seen.add(url)
        new_rows.append(p)
    
    print(f"Davon neu (nicht Duplikat): {len(new_rows)}")
    
    if new_rows:
        append_to_xlsx(new_rows)
        print(f"\n✓ {len(new_rows)} neue Objekte in Mallorca_Markt_Gesamt.xlsx eingetragen")
    else:
        print("Keine neuen Objekte zum Eintragen.")
    
    print("\n=== FERTIG ===")
    return len(new_rows)


if __name__ == "__main__":
    main()
