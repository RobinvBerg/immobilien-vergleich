#!/usr/bin/env python3
"""
Scraper for: Properstar, Savills, Balearic Properties, The Agency RE, Pollensa Properties
"""
import json, time, re, sys
from datetime import date
from openpyxl import load_workbook

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'

def load_existing_urls():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
    return wb, ws, existing

def save_objects(new_objects, source_name):
    wb, ws, existing_urls = load_existing_urls()
    new_count = 0
    for obj in new_objects:
        url = obj.get('url', '—')
        if url != '—' and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', ''),
            obj.get('quelle', source_name),
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            str(date.today()),
            'Neu'
        ])
        if url != '—':
            existing_urls.add(url)
        new_count += 1
    wb.save(EXCEL_PATH)
    print(f"✅ {source_name}: {new_count} neue Objekte gespeichert (von {len(new_objects)} gefunden)")
    return new_count

# ─────────────────────────────────────────────────────────────
# 1. BALEARIC PROPERTIES (BeautifulSoup, direct HTML)
# ─────────────────────────────────────────────────────────────
def scrape_balearic_properties():
    print("\n=== Balearic Properties ===")
    import requests
    from bs4 import BeautifulSoup
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
    }
    
    objects = []
    base_url = 'https://www.balearicproperties.com'
    page = 1
    
    while True:
        url = f'{base_url}/properties-for-sale/mallorca?page={page}' if page > 1 else f'{base_url}/properties-for-sale/mallorca'
        print(f"  Fetching page {page}: {url}")
        
        try:
            r = requests.get(url, headers=headers, timeout=20)
            print(f"  Status: {r.status_code}, len={len(r.text)}")
            if r.status_code != 200:
                break
                
            soup = BeautifulSoup(r.text, 'html.parser')
            
            # Try different selectors
            cards = soup.select('.property-card, .property-item, .listing-item, article.property, .property, [class*="property"]')
            print(f"  Cards found with broad selector: {len(cards)}")
            
            # Try links with /property/ pattern
            prop_links = soup.find_all('a', href=re.compile(r'/property|/properties/'))
            print(f"  Property links: {len(prop_links)}")
            
            # Print some HTML for debugging
            if page == 1:
                print("  Title:", soup.title.string if soup.title else 'N/A')
                # Look for price patterns
                prices = soup.find_all(string=re.compile(r'€[\d,\.]+|[\d,]+\s*€'))
                print(f"  Price strings found: {len(prices)}")
                if prices:
                    print(f"  Sample prices: {prices[:3]}")
            
            # Try to find any listing containers
            containers = soup.select('.listing, .property-listing, .search-result, .results-list li, .properties-list li')
            print(f"  Listing containers: {len(containers)}")
            
            # Broad approach: find all links
            all_links = soup.find_all('a', href=True)
            prop_urls = [l['href'] for l in all_links if '/property/' in l['href'] or '/sale/' in l['href'] or '/listing/' in l['href']]
            prop_urls = list(set(prop_urls))
            print(f"  Unique property-like URLs: {len(prop_urls)}")
            
            if not prop_links and not prop_urls and len(objects) == 0 and page == 1:
                # Save raw HTML snippet for debugging
                print("  Raw HTML snippet (first 2000 chars):")
                print(r.text[:2000])
                break
            
            found_this_page = 0
            # Process what we found
            for link in all_links:
                href = link.get('href', '')
                if not href:
                    continue
                full_url = href if href.startswith('http') else base_url + href
                
                # Check if it looks like a property URL
                if not any(p in href for p in ['/property/', '/properties/', '/sale/', '/listing/', '/villa/', '/house/', '/finca/']):
                    continue
                
                # Get title
                title = link.get_text(strip=True) or link.get('title', '') or href
                if len(title) < 5:
                    # Look at parent
                    parent = link.parent
                    if parent:
                        title = parent.get_text(strip=True)[:100]
                
                if not title or full_url in [o['url'] for o in objects]:
                    continue
                
                # Try to find price near this link
                parent = link.parent
                text = parent.get_text() if parent else ''
                price_match = re.search(r'([\d,\.]+)\s*€|€\s*([\d,\.]+)', text)
                price = None
                if price_match:
                    p_str = (price_match.group(1) or price_match.group(2)).replace('.', '').replace(',', '')
                    try:
                        price = int(p_str)
                    except:
                        pass
                
                objects.append({
                    'titel': title[:200],
                    'quelle': 'Balearic Properties',
                    'url': full_url,
                    'preis': price,
                    'ort': 'Mallorca',
                })
                found_this_page += 1
            
            print(f"  Objects this page: {found_this_page}")
            
            if found_this_page == 0:
                break
            
            # Check for next page
            next_page = soup.find('a', string=re.compile(r'next|Next|›|»|\d+')) 
            next_btn = soup.select_one('.pagination .next, .next-page, [rel="next"]')
            if not next_btn:
                break
                
            page += 1
            if page > 20:
                break
            time.sleep(1)
            
        except Exception as e:
            print(f"  Error: {e}")
            import traceback
            traceback.print_exc()
            break
    
    print(f"Total found: {len(objects)}")
    return objects


# ─────────────────────────────────────────────────────────────
# 2. POLLENSA PROPERTIES (BeautifulSoup, direct HTML)
# ─────────────────────────────────────────────────────────────
def scrape_pollensa_properties():
    print("\n=== Pollensa Properties ===")
    import requests
    from bs4 import BeautifulSoup
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
    }
    
    objects = []
    base_url = 'https://www.pollensaproperties.com'
    page = 1
    
    while True:
        url = f'{base_url}/properties-for-sale' if page == 1 else f'{base_url}/properties-for-sale?page={page}'
        print(f"  Fetching page {page}: {url}")
        
        try:
            r = requests.get(url, headers=headers, timeout=20)
            print(f"  Status: {r.status_code}, len={len(r.text)}")
            if r.status_code != 200:
                break
            
            soup = BeautifulSoup(r.text, 'html.parser')
            
            if page == 1:
                print("  Title:", soup.title.string if soup.title else 'N/A')
            
            # Find property listings
            found_this_page = 0
            seen_urls = set(o['url'] for o in objects)
            
            # Look for property cards/links
            all_links = soup.find_all('a', href=True)
            for link in all_links:
                href = link.get('href', '')
                if not href:
                    continue
                
                # Property URL patterns for Pollensa Properties
                if not any(p in href.lower() for p in ['/property/', '/villa', '/finca', '/house', '/apartment', '/penthouse', '/townhouse', '/estate']):
                    continue
                
                full_url = href if href.startswith('http') else base_url + href
                if full_url in seen_urls:
                    continue
                
                # Get title from link or surroundings
                title = link.get_text(strip=True)
                if len(title) < 5:
                    parent = link.parent
                    if parent:
                        # Look for h2/h3 nearby
                        heading = parent.find(['h1','h2','h3','h4'])
                        if heading:
                            title = heading.get_text(strip=True)
                        else:
                            title = parent.get_text(strip=True)[:100]
                
                # Find price
                container = link.parent
                for _ in range(4):  # Go up a few levels
                    if container:
                        text = container.get_text()
                        price_match = re.search(r'€\s*([\d,\.]+)|([\d,\.]+)\s*€', text)
                        if price_match:
                            p_str = (price_match.group(1) or price_match.group(2)).replace('.', '').replace(',', '')
                            try:
                                price = int(float(p_str))
                                break
                            except:
                                pass
                        container = container.parent
                else:
                    price = None
                
                # Find bedrooms
                container = link.parent
                zimmer = None
                for _ in range(4):
                    if container:
                        text = container.get_text()
                        bed_match = re.search(r'(\d+)\s*(?:bed|bedroom|hab|zimmer)', text, re.IGNORECASE)
                        if bed_match:
                            zimmer = int(bed_match.group(1))
                            break
                        container = container.parent
                
                # Extract location from URL or title
                ort = 'Mallorca'
                for loc in ['Pollensa', 'Pollença', 'Puerto Pollensa', 'Alcudia', 'Formentor', 'Cala San Vicente']:
                    if loc.lower() in title.lower() or loc.lower() in href.lower():
                        ort = loc
                        break
                
                objects.append({
                    'titel': title[:200] if title else href,
                    'quelle': 'Pollensa Properties',
                    'url': full_url,
                    'preis': price,
                    'zimmer': zimmer,
                    'ort': ort,
                })
                seen_urls.add(full_url)
                found_this_page += 1
            
            print(f"  Objects this page: {found_this_page}")
            
            if found_this_page == 0 and page == 1:
                print("  Debug - raw HTML snippet:")
                print(r.text[:2000])
                break
            
            if found_this_page == 0:
                break
            
            # Check for pagination
            next_btn = soup.select_one('a[rel="next"], .pagination .next, .next-page-link')
            if not next_btn:
                # Check for numbered pagination
                pag = soup.select('.pagination a, .pager a')
                max_page = 1
                for a in pag:
                    try:
                        n = int(a.get_text(strip=True))
                        max_page = max(max_page, n)
                    except:
                        pass
                if page >= max_page:
                    break
            
            page += 1
            if page > 30:
                break
            time.sleep(1)
            
        except Exception as e:
            print(f"  Error: {e}")
            import traceback
            traceback.print_exc()
            break
    
    print(f"Total found: {len(objects)}")
    return objects


if __name__ == '__main__':
    # Run Balearic Properties first (simplest)
    bp_objects = scrape_balearic_properties()
    save_objects(bp_objects, 'Balearic Properties')
    
    # Then Pollensa Properties
    pp_objects = scrape_pollensa_properties()
    save_objects(pp_objects, 'Pollensa Properties')
