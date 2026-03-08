import openpyxl, requests, time

API_KEY = "AIzaSyAzplaFOGUrCygk3mrzS--wrHNIz0arKok"

# Referenzpunkte (lat,lng)
REFS = {
    "flughafen": "39.5517,2.7388",
    "daia":      "39.4667,2.7833",  # Daia Referenzpunkt
    "andratx":   "39.5731,2.3846",
    "ses_salines":"39.3347,3.0471",
}

# Die 22 Einträge mit Ort
entries = {
    16:  "Escorca, Mallorca",
    26:  "Llucmajor, Mallorca",
    39:  "Llucmajor, Mallorca",
    47:  "Llucmajor, Mallorca",
    53:  "Calvià, Mallorca",
    56:  "Santa Margalida, Mallorca",
    66:  "Artà, Mallorca",
    67:  "Alcudia, Mallorca",
    71:  "Son Macià, Mallorca",
    101: "Es Carritxó, Mallorca",
    126: "Campos, Mallorca",
    146: "Bunyola, Mallorca",
    152: "Felanitx, Mallorca",
    178: "Marratxí, Mallorca",
    193: "Calvià, Mallorca",
    195: "Ses Salines, Mallorca",
    244: "Montuïri, Mallorca",
    248: "Montuïri, Mallorca",
    265: "Binissalem, Mallorca",
    293: "Campos, Mallorca",
    303: "Santa Maria del Camí, Mallorca",
    310: "Gènova, Palma, Mallorca",
}

def get_distances(origin, destinations):
    """Returns dict: dest_key -> (km, min)"""
    dest_str = "|".join(destinations.values())
    url = "https://maps.googleapis.com/maps/api/distancematrix/json"
    params = {
        "origins": origin,
        "destinations": dest_str,
        "mode": "driving",
        "key": API_KEY,
    }
    r = requests.get(url, params=params, timeout=10)
    data = r.json()
    if data["status"] != "OK":
        raise Exception(f"API error: {data['status']}")
    
    result = {}
    row = data["rows"][0]["elements"]
    for i, key in enumerate(destinations.keys()):
        el = row[i]
        if el["status"] == "OK":
            km = round(el["distance"]["value"] / 1000, 1)
            mins = round(el["duration"]["value"] / 60)
            result[key] = (km, mins)
        else:
            result[key] = (None, None)
    return result

# Load Excel
wb = openpyxl.load_workbook("mallorca-kandidaten-v2.xlsx")
ws = wb.active

# Col mapping: K=11,L=12,M=13,N=14,O=15,P=16,Q=17,R=18
# K=Flugh km, L=Flugh min, M=Daia km, N=Daia min
# O=Andratx km, P=Andratx min, Q=SesSalines km, R=SesSalines min

updated = []
for r in range(2, ws.max_row+1):
    nr = ws.cell(r, 1).value
    if nr not in entries:
        continue
    
    origin = entries[nr]
    print(f"Nr.{nr} ({origin})...", end=" ", flush=True)
    
    try:
        dists = get_distances(origin, REFS)
        
        ws.cell(r, 11).value = dists["flughafen"][0]  # K km
        ws.cell(r, 12).value = dists["flughafen"][1]  # L min
        ws.cell(r, 13).value = dists["daia"][0]        # M km
        ws.cell(r, 14).value = dists["daia"][1]        # N min
        ws.cell(r, 15).value = dists["andratx"][0]     # O km
        ws.cell(r, 16).value = dists["andratx"][1]     # P min
        ws.cell(r, 17).value = dists["ses_salines"][0] # Q km
        ws.cell(r, 18).value = dists["ses_salines"][1] # R min
        
        # Location befüllen wenn leer
        if not ws.cell(r, 10).value or ws.cell(r, 10).value == "Mallorca":
            ws.cell(r, 10).value = origin.split(",")[0]
        
        print(f"✓ Flugh={dists['flughafen'][1]}min Daia={dists['daia'][1]}min")
        updated.append(nr)
    except Exception as e:
        print(f"ERROR: {e}")
    
    time.sleep(0.2)

wb.save("mallorca-kandidaten-v2.xlsx")
print(f"\nFertig. {len(updated)} Einträge aktualisiert: {updated}")
