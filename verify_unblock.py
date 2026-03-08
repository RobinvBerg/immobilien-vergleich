#!/usr/bin/env python3
"""
Verify/download images for bot-blocked entries using Decodo Site Unblocker.
"""
import asyncio, json, hashlib, sys
from pathlib import Path
from playwright.async_api import async_playwright
import openpyxl

BASE = Path(__file__).parent

PROXY = {
    'server': 'http://unblock.decodo.com:60000',
    'username': 'U0000364062',
    'password': 'PW_1047072161848b0d67b68ff1b160986e6'
}

# Load no-live entries from previous run
prev = json.load(open(BASE / 'verify_nonlb_results.json'))
no_live_nrs = {r['nr'] for r in prev if r['status'] == 'NO_LIVE'}

# Load Excel for URLs
wb = openpyxl.load_workbook(BASE / 'mallorca-kandidaten-v2.xlsx')
ws = wb.active
headers = [c.value for c in ws[1]]

entries = []
for row in ws.iter_rows(min_row=2):
    nr = row[headers.index('Ordnungsnummer')].value
    if not nr or int(nr) not in no_live_nrs: continue
    entries.append({
        'nr': int(nr),
        'url': row[headers.index('Link Objekt (URL)')].value or '',
        'makler': row[headers.index('Makler')].value or '',
        'name': row[headers.index('Name')].value or '',
    })

entries.sort(key=lambda x: x['nr'])
print(f"Entries to unblock: {len(entries)}", flush=True)

async def get_image(page, url):
    await page.goto(url, timeout=40000, wait_until='domcontentloaded')
    await asyncio.sleep(3)

    img_url = None
    selectors = [
        'img[src*="photo"]','img[src*="image"]','img[src*="foto"]',
        'img[src*="cdn"]','img[src*="media"]','img[src*="property"]',
        '.gallery img', '.slider img', 'picture img',
        '[class*="gallery"] img','[class*="photo"] img',
        '[class*="main"] img','[class*="hero"] img',
    ]
    for sel in selectors:
        try:
            els = await page.query_selector_all(sel)
            for el in els[:5]:
                for attr in ['src','data-src','data-lazy']:
                    src = await el.get_attribute(attr) or ''
                    src = src.split(',')[0].split(' ')[0]
                    if src and any(x in src.lower() for x in ['jpg','jpeg','webp','png']) and len(src)>20 and 'logo' not in src.lower():
                        img_url = src; break
                if img_url: break
        except: pass
        if img_url: break

    if not img_url:
        imgs = await page.query_selector_all('img')
        best = (0, None)
        for img in imgs[:40]:
            try:
                w = await page.evaluate('(el) => el.naturalWidth', img)
                src = await img.get_attribute('src') or ''
                if w > best[0] and w > 300 and src and 'logo' not in src.lower():
                    best = (w, src)
            except: pass
        img_url = best[1]

    if not img_url: return None
    if img_url.startswith('//'): img_url = 'https:' + img_url
    elif img_url.startswith('/'):
        from urllib.parse import urlparse
        p = urlparse(url)
        img_url = f"{p.scheme}://{p.netloc}{img_url}"

    try:
        resp = await page.request.get(img_url, timeout=20000)
        if resp.status == 200:
            return await resp.body()
    except: pass
    return None

async def main():
    results = []
    stats = {'fixed': 0, 'match': 0, 'failed': 0}

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, proxy=PROXY)
        ctx = await browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
            viewport={'width': 1280, 'height': 900},
            extra_http_headers={'X-Decodo-JS-Render': 'true'},
            ignore_https_errors=True
        )

        for i, entry in enumerate(entries):
            nr, url = entry['nr'], entry['url']
            stored = BASE / f'bilder/{nr}_main.jpg'
            sys.stdout.write(f"[{i+1}/{len(entries)}] Nr.{nr:3d} {entry['makler'][:15]:<15} ")
            sys.stdout.flush()

            page = await ctx.new_page()
            try:
                live_bytes = await get_image(page, url)
                if not live_bytes:
                    print("❓ kein Bild")
                    stats['failed'] += 1
                    results.append({'nr': nr, 'status': 'STILL_BLOCKED'})
                    continue

                live_hash = hashlib.md5(live_bytes).hexdigest()
                stored_hash = hashlib.md5(stored.read_bytes()).hexdigest() if stored.exists() else None

                if stored_hash == live_hash:
                    print("✅ match")
                    stats['match'] += 1
                    results.append({'nr': nr, 'status': 'MATCH'})
                else:
                    stored.write_bytes(live_bytes)
                    action = 'fixed' if stored_hash else 'new'
                    print(f"🔄 {action} ({live_hash[:8]})")
                    stats['fixed'] += 1
                    results.append({'nr': nr, 'status': 'FIXED', 'action': action})

            except Exception as e:
                print(f"⚠️  {str(e)[:50]}")
                stats['failed'] += 1
                results.append({'nr': nr, 'status': 'ERR', 'error': str(e)[:80]})
            finally:
                await page.close()

        await browser.close()

    json.dump(results, open(BASE / 'verify_unblock_results.json', 'w'), indent=2)
    print(f"\n{'='*60}")
    print(f"FERTIG: {stats}")
    still = [r['nr'] for r in results if r['status'] in ('STILL_BLOCKED','ERR')]
    if still: print(f"Noch geblockt: {still}")

asyncio.run(main())
