#!/usr/bin/env python3
"""Scrape 6 boutique Mallorca real estate agencies and update xlsx files."""

import sys
import time
import re
import datetime
import traceback
import requests
from bs4 import BeautifulSoup
import openpyxl

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9,de;q=0.8',
}

QUELLEN_FILE = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Immobilien_Quellen.xlsx'
GESAMT_FILE = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
TODAY = '2026-03-03'

NEW_SOURCES = [
    (56, 'Luxury Estates Mallorca', 'Makler (Boutique)', 'luxury-estates-mallorca.com', 'EN/DE', 'Luxury/Off-Market', "Christie's International Vertreter auf Mallorca"),
    (57, 'Equus Mallorca', 'Makler (Boutique)', 'equusmallorca.com', 'DE/EN', 'Luxury', 'Alex & Britt Jolig, sehr persönlich'),
    (58, 'Luxury on Mallorca', 'Makler (Boutique)', 'luxuryonmallorca.com', 'EN/DE', 'Luxury', 'Christoph Kornschober, kuratiert'),
    (59, 'Mallorca Realtors', 'Makler (Boutique)', 'mallorcarealtors.com', 'EN/DE/ES', 'Luxury/Off-Market', 'Spezialist Top-End & Luxury'),
    (60, 'Marcel Remus Real Estate', 'Makler (Boutique)', 'marcelremusrealestate.com', 'DE/EN', 'Ultra-Luxury', 'Sehr bekannt, HNWI-Klientel, Off-Market'),
    (61, 'Rossitza Hantelmann', 'Makler (Boutique)', 'rossitzahantelmann.com', 'DE/EN', 'Luxury/Off-Market', 'Diskret, Luxus-Spezialist'),
]

def get_soup(url, timeout=20):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        r.raise_for_status()
        return BeautifulSoup(r.text, 'html.parser'), r.text, r.url
    except Exception as e:
        print(f"  ERROR fetching {url}: {e}")
        return None, None, None

def parse_price(text):
    if not text:
        return None
    text = str(text).replace('\xa0', '').replace(' ', '').replace('.', '').replace(',', '')
    m = re.search(r'\d+', text)
    if m:
        try:
            v = int(m.group())
            # Handle Mio
            if 'mio' in str(text).lower() or 'mill' in str(text).lower():
                v = v * 1000000
            return v
        except:
            return None
    return None

def parse_num(text):
    if not text:
        return None
    m = re.search(r'\d+', str(text).replace('.', '').replace(',', ''))
    if m:
        try:
            return int(m.group())
        except:
            return None
    return None

def scrape_generic(base_urls, domain_base):
    """Generic scraper that tries multiple URL patterns."""
    props = []
    seen_hrefs = set()
    
    for base_url in base_urls:
        soup, html, final_url = get_soup(base_url)
        if not soup:
            time.sleep(2)
            continue
        
        title_str = soup.title.string if soup.title else 'N/A'
        print(f"  URL: {base_url} -> {final_url}")
        print(f"  Title: {title_str}")
        
        # Try property cards
        cards = soup.find_all(['div','article','li','section'], 
                               class_=re.compile(r'property|listing|card|estate|item|result|propert', re.I))
        print(f"  Card elements: {len(cards)}")
        
        for card in cards[:50]:
            a_tags = card.find_all('a', href=True)
            main_a = None
            for a in a_tags:
                href = a['href']
                if re.search(r'/propert|/immob|/villa|/finca|/house|/haus|/objekt|/buy|/sale|/ref', href, re.I):
                    main_a = a
                    break
            if not main_a and a_tags:
                main_a = a_tags[0]
            if not main_a:
                continue
            
            href = main_a['href']
            if not href.startswith('http'):
                href = domain_base.rstrip('/') + '/' + href.lstrip('/')
            if href in seen_hrefs:
                continue
            seen_hrefs.add(href)
            
            title_el = card.find(['h1','h2','h3','h4'])
            title = title_el.get_text(strip=True) if title_el else main_a.get_text(strip=True)
            if not title or len(title) < 3:
                continue
            
            text = card.get_text(' ', strip=True)
            
            # Price
            price_m = re.search(r'([\d][.\d]*)\s*(?:Mio\.?|mio\.?)\s*(?:€|EUR)', text) or \
                      re.search(r'(?:€|EUR)\s*([\d][,\.\d]+)', text) or \
                      re.search(r'([\d][,\.\d]+)\s*(?:€|EUR)', text)
            price = None
            if price_m:
                raw = price_m.group(1).replace('.', '').replace(',', '')
                try:
                    price = int(raw)
                    if 'mio' in price_m.group(0).lower():
                        price *= 1000000
                    elif price < 1000:
                        price *= 1000  # likely in thousands
                except:
                    pass
            
            rooms_m = re.search(r'(\d+)\s*(?:Bed|bed|SZ|Schlaf|zimmer|BR|rooms?)\b', text, re.I)
            rooms = int(rooms_m.group(1)) if rooms_m else None
            
            living_m = re.search(r'(\d+)\s*m[²2]?\s*(?:living|Wohn|built|construc)', text, re.I)
            living = int(living_m.group(1)) if living_m else None
            
            plot_m = re.search(r'(\d[\d\.]*)\s*m[²2]?\s*(?:plot|Grund|land|terreno)', text, re.I)
            plot = parse_num(plot_m.group(1)) if plot_m else None
            
            # Location
            loc_m = re.search(r'(?:in|@|location:?)\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)', text)
            location = loc_m.group(1) if loc_m else None
            
            props.append({
                'title': title[:200],
                'url': href,
                'price': price,
                'rooms': rooms,
                'plot': plot,
                'living': living,
                'location': location,
            })
        
        if props:
            break
        
        # Fallback: find any property-like links
        links = soup.find_all('a', href=re.compile(
            r'/propert|/immob|/villa|/finca|/haus|/house|/objekt|/ref-|/for-sale', re.I))
        print(f"  Property links (fallback): {len(links)}")
        for a in links[:30]:
            href = a['href']
            if not href.startswith('http'):
                href = domain_base.rstrip('/') + '/' + href.lstrip('/')
            if href in seen_hrefs:
                continue
            seen_hrefs.add(href)
            title = a.get_text(strip=True)
            if title and len(title) > 3:
                props.append({'title': title[:200], 'url': href, 'price': None, 'rooms': None,
                              'plot': None, 'living': None, 'location': None})
        if props:
            break
        
        time.sleep(2)
    
    return props

results = {}

# ─── 1. Luxury Estates Mallorca ───────────────────────────────────────────────
print("\n=== 1. Luxury Estates Mallorca ===")
props = scrape_generic([
    'https://www.luxury-estates-mallorca.com/en/properties',
    'https://www.luxury-estates-mallorca.com/properties',
    'https://www.luxury-estates-mallorca.com',
], 'https://www.luxury-estates-mallorca.com')
print(f"  Found: {len(props)}")
results['Luxury Estates Mallorca'] = props

# ─── 2. Equus Mallorca ────────────────────────────────────────────────────────
print("\n=== 2. Equus Mallorca ===")
props = scrape_generic([
    'https://www.equusmallorca.com/en/properties',
    'https://www.equusmallorca.com/properties',
    'https://www.equusmallorca.com/de/immobilien',
    'https://www.equusmallorca.com',
], 'https://www.equusmallorca.com')
print(f"  Found: {len(props)}")
results['Equus Mallorca'] = props

# ─── 3. Luxury on Mallorca ────────────────────────────────────────────────────
print("\n=== 3. Luxury on Mallorca ===")
props = scrape_generic([
    'https://www.luxuryonmallorca.com/properties',
    'https://www.luxuryonmallorca.com/buy',
    'https://www.luxuryonmallorca.com/for-sale',
    'https://www.luxuryonmallorca.com',
], 'https://www.luxuryonmallorca.com')
print(f"  Found: {len(props)}")
results['Luxury on Mallorca'] = props

# ─── 4. Mallorca Realtors ─────────────────────────────────────────────────────
print("\n=== 4. Mallorca Realtors ===")
props = scrape_generic([
    'https://www.mallorcarealtors.com/properties',
    'https://www.mallorcarealtors.com/buy',
    'https://www.mallorcarealtors.com',
], 'https://www.mallorcarealtors.com')
print(f"  Found: {len(props)}")
results['Mallorca Realtors'] = props

# ─── 5. Marcel Remus Real Estate ──────────────────────────────────────────────
print("\n=== 5. Marcel Remus Real Estate ===")
props = scrape_generic([
    'https://www.marcelremusrealestate.com/en/properties',
    'https://www.marcelremusrealestate.com/properties',
    'https://www.marcelremusrealestate.com/de/immobilien',
    'https://www.marcelremusrealestate.com',
], 'https://www.marcelremusrealestate.com')
print(f"  Found: {len(props)}")
results['Marcel Remus Real Estate'] = props

# ─── 6. Rossitza Hantelmann ───────────────────────────────────────────────────
print("\n=== 6. Rossitza Hantelmann ===")
props = scrape_generic([
    'https://www.rossitzahantelmann.com/properties',
    'https://www.rossitzahantelmann.com/en/properties',
    'https://www.rossitzahantelmann.com/de/immobilien',
    'https://www.rossitzahantelmann.com',
], 'https://www.rossitzahantelmann.com')
print(f"  Found: {len(props)}")
results['Rossitza Hantelmann'] = props

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: Update Quellenliste
# ─────────────────────────────────────────────────────────────────────────────
print("\n=== Updating Quellenliste ===")
wb_q = openpyxl.load_workbook(QUELLEN_FILE)
ws_q = wb_q.active

existing_nums = set()
for row in ws_q.iter_rows(min_row=2, values_only=True):
    if row[0]:
        existing_nums.add(row[0])

added_sources = 0
for src in NEW_SOURCES:
    num = src[0]
    if num in existing_nums:
        print(f"  #{num} already exists, skipping")
        continue
    ws_q.append([num, src[1], src[2], src[3], src[4], src[5], None, src[6], None, None, None, None])
    added_sources += 1
    print(f"  Added #{num}: {src[1]}")

wb_q.save(QUELLEN_FILE)
print(f"  Saved: {added_sources} new sources")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: Update Gesamtliste
# ─────────────────────────────────────────────────────────────────────────────
print("\n=== Updating Gesamtliste ===")
wb_g = openpyxl.load_workbook(GESAMT_FILE)
ws_g = wb_g.active

existing_urls = set()
for row in ws_g.iter_rows(min_row=2, values_only=True):
    if row[2]:
        existing_urls.add(str(row[2]).strip())

total_added = 0
source_counts = {}

for src in NEW_SOURCES:
    name = src[1]
    props_list = results.get(name, [])
    added = 0
    for p in props_list:
        url = str(p['url']).strip()
        if url in existing_urls:
            continue
        existing_urls.add(url)
        ws_g.append([
            p['title'], name, url, p['price'], p['rooms'],
            p['plot'], p['living'], p['location'], TODAY, 'Neu',
        ])
        added += 1
        total_added += 1
    source_counts[name] = added
    print(f"  {name}: +{added} (from {len(props_list)} scraped)")

wb_g.save(GESAMT_FILE)
print(f"  Total new: {total_added}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: Update Quellenliste Status
# ─────────────────────────────────────────────────────────────────────────────
print("\n=== Updating Status ===")
wb_q2 = openpyxl.load_workbook(QUELLEN_FILE)
ws_q2 = wb_q2.active

source_name_map = {src[1]: src for src in NEW_SOURCES}

for row_idx in range(2, ws_q2.max_row + 1):
    name_val = ws_q2.cell(row_idx, 2).value
    if name_val in source_name_map:
        count = source_counts.get(name_val, 0)
        status = 'Ja' if count > 0 else 'Teilweise'
        ws_q2.cell(row_idx, 6).value = status
        ws_q2.cell(row_idx, 7).value = count
        ws_q2.cell(row_idx, 8).value = f"Gescrapt {TODAY}: {count} Objekte. " + source_name_map[name_val][6]
        print(f"  {name_val}: Status={status}, count={count}")

wb_q2.save(QUELLEN_FILE)

# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "="*60)
print("FINAL REPORT")
print("="*60)
for src in NEW_SOURCES:
    name = src[1]
    count = source_counts.get(name, 0)
    scraped = len(results.get(name, []))
    print(f"  #{src[0]} {name}: scraped={scraped}, added={count}")
print(f"\n  GESAMT neue Einträge: {total_added}")
print(f"  Quellen hinzugefügt: {added_sources}")
