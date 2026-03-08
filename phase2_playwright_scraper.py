#!/usr/bin/env python3
"""
Phase 2: Playwright Scraper for Mallorca real estate agencies
"""
import sys
import json
import time
import re
import traceback
from datetime import date
from openpyxl import load_workbook

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
all_results = []
summary = {}

def clean_price(txt):
    if not txt:
        return None
    txt = str(txt).replace('\xa0', '').replace('\u202f', '').replace(' ', '').replace('.', '').replace(',', '').replace('€','').replace('EUR','')
    m = re.search(r'(\d{4,})', txt)
    if m:
        v = int(m.group(1))
        # Some prices need correction (e.g. "1500" should be 1500000 if in thousands)
        return v
    return None

def clean_int(txt):
    if not txt:
        return None
    txt = str(txt).replace('.', '').replace(',', '').replace('\xa0', '').replace(' ','').replace('m²','').replace('m2','')
    m = re.search(r'(\d+)', txt)
    if m:
        return int(m.group(1))
    return None

def save_to_excel(new_objects):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set(str(row[2]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[2])
    new_count = 0
    for obj in new_objects:
        url = obj.get('url', '—')
        if url and url != '—' and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', 'Unbekannt'),
            obj.get('quelle', ''),
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            str(date.today()),
            'Neu'
        ])
        if url and url != '—':
            existing_urls.add(url)
        new_count += 1
    wb.save(EXCEL_PATH)
    return new_count

# ─────────────────────────────────────────────────────────────
# Helper: use playwright with stealth
# ─────────────────────────────────────────────────────────────
def get_page_content(url, wait_for=None, timeout=30000):
    """Returns (html_content, final_url) or (None, None) on error"""
    try:
        from playwright.sync_api import sync_playwright
        try:
            from playwright_stealth import Stealth
            use_stealth = True
        except:
            use_stealth = False
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                viewport={'width': 1280, 'height': 900}
            )
            page = context.new_page()
            if use_stealth:
                try:
                    Stealth().use_sync(page)
                except:
                    pass
            
            try:
                page.goto(url, timeout=timeout, wait_until='domcontentloaded')
                if wait_for:
                    try:
                        page.wait_for_selector(wait_for, timeout=8000)
                    except:
                        pass
                else:
                    page.wait_for_timeout(2000)
                
                content = page.content()
                final_url = page.url
                browser.close()
                return content, final_url
            except Exception as e:
                print(f"    Page error: {e}")
                try:
                    content = page.content()
                    final_url = page.url
                    browser.close()
                    return content, final_url
                except:
                    browser.close()
                    return None, None
    except Exception as e:
        print(f"    Playwright error: {e}")
        return None, None

def extract_listings(html, source, base_url, title_sel=None, price_sel=None, link_sel=None, card_sel=None, rooms_sel=None, plot_sel=None, area_sel=None, loc_sel=None):
    """Generic listing extractor"""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, 'html.parser')
    results = []
    
    if not card_sel:
        card_sel = '.property, .property-card, .property-item, article.property, .listing-item, .estate-item, [class*="property-card"], [class*="listing-card"]'
    
    cards = soup.select(card_sel)
    
    if not cards:
        # Broader search
        cards = soup.select('article, .card')
        cards = [c for c in cards if c.select_one('a[href*="property"], a[href*="immobilie"], a[href*="finca"], a[href*="villa"], a[href*="ref-"]')]
    
    for card in cards[:80]:
        title_el = card.select_one(title_sel or 'h2, h3, h4, .title, [class*="title"], [class*="name"]')
        price_el = card.select_one(price_sel or '[class*="price"], [class*="preis"], [class*="Price"]')
        link_el = card.select_one(link_sel or 'a[href]')
        rooms_el = card.select_one(rooms_sel or '[class*="bed"], [class*="room"], [class*="zimmer"]') if rooms_sel or not rooms_sel else None
        plot_el = card.select_one(plot_sel or '[class*="plot"], [class*="land"], [class*="grundst"]') if plot_sel or not plot_sel else None
        area_el = card.select_one(area_sel or '[class*="area"], [class*="flaeche"], [class*="sqm"]') if area_sel or not area_sel else None
        loc_el = card.select_one(loc_sel or '[class*="location"], [class*="ort"], [class*="region"], [class*="area"]') if loc_sel or not loc_sel else None
        
        if not title_el and not price_el:
            continue
        
        href = ''
        if link_el:
            href = link_el.get('href', '')
            if href and not href.startswith('http'):
                href = base_url.rstrip('/') + '/' + href.lstrip('/')
        
        obj = {
            'titel': title_el.get_text(strip=True) if title_el else source + ' Objekt',
            'quelle': source,
            'url': href or base_url,
            'preis': clean_price(price_el.get_text() if price_el else None),
            'zimmer': clean_int(rooms_el.get_text() if rooms_el else None),
            'grundstueck': clean_int(plot_el.get_text() if plot_el else None),
            'wohnflaeche': clean_int(area_el.get_text() if area_el else None),
            'ort': loc_el.get_text(strip=True)[:50] if loc_el else '',
        }
        results.append(obj)
    
    return results

# ─────────────────────────────────────────────────────────────
# INDIVIDUAL SCRAPERS
# ─────────────────────────────────────────────────────────────

def scrape_portamallorquina():
    """portamallorquina.com"""
    print("\n[1] Porta Mallorquina (portamallorquina.com)...")
    results = []
    try:
        # Try different URL patterns
        urls = [
            "https://www.portamallorquina.com/kaufen/?zimmer=5",
            "https://www.portamallorquina.com/kaufen/",
            "https://www.portamallorquina.com/objekte/",
        ]
        for url in urls:
            html, final = get_page_content(url)
            if not html:
                continue
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            
            # Check page title
            title = soup.select_one('title')
            print(f"  Page: {title.text[:60] if title else 'no title'} at {final}")
            
            # Try various card selectors
            for sel in ['.property-item', '.listing-item', '.object-item', 'article', '.property', '.expose-item', '.estate-card']:
                cards = soup.select(sel)
                if cards:
                    print(f"  Found {len(cards)} cards with selector '{sel}'")
                    for card in cards[:50]:
                        t = card.select_one('h2, h3, h4, .title')
                        p = card.select_one('[class*="price"], [class*="preis"]')
                        a = card.select_one('a[href]')
                        if t or p:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.portamallorquina.com' + href
                            results.append({
                                'titel': t.get_text(strip=True) if t else 'Porta Mallorquina',
                                'quelle': 'Porta Mallorquina',
                                'url': href or url,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    if results:
                        break
            
            if not results:
                # Look for any links with property indicators
                links = soup.select('a[href*="expose"], a[href*="objekt"], a[href*="property"], a[href*="kaufen"]')
                print(f"  Found {len(links)} property links")
                for link in links[:30]:
                    title_text = link.get_text(strip=True)
                    if title_text and len(title_text) > 5:
                        href = link['href']
                        if not href.startswith('http'):
                            href = 'https://www.portamallorquina.com' + href
                        results.append({
                            'titel': title_text[:100],
                            'quelle': 'Porta Mallorquina',
                            'url': href,
                        })
            
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
        traceback.print_exc()
    print(f"  → {len(results)} Objekte")
    return results

def scrape_unique_mallorca():
    """uniquemallorca.com"""
    print("\n[2] Unique Mallorca...")
    results = []
    try:
        urls = [
            "https://www.uniquemallorca.com/properties-for-sale/?bedrooms=5&",
            "https://www.uniquemallorca.com/for-sale/",
            "https://www.uniquemallorca.com/properties/",
            "https://www.uniquemallorca.com/",
        ]
        for url in urls:
            html, final = get_page_content(url)
            if not html:
                continue
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  Page: {title.text[:60] if title else 'no title'} at {final}")
            
            for sel in ['.property-item', '.listing', 'article.property', '.property-card', '.listing-card', '.property', '[class*="property-card"]']:
                cards = soup.select(sel)
                if cards:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:50]:
                        t = card.select_one('h2, h3, .title, [class*="title"]')
                        p = card.select_one('[class*="price"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.uniquemallorca.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Unique Mallorca',
                                'url': href or url,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_sandberg():
    """sandberg-estates.com"""
    print("\n[3] Sandberg Estates...")
    results = []
    try:
        html, final = get_page_content("https://sandberg-estates.com/properties/for-sale/")
        if html:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  Page: {title.text[:60] if title else 'no title'}")
            
            for sel in ['.property-item', '.listing-item', 'article', '.property', '.property-card', '[class*="property"]']:
                cards = soup.select(sel)
                if len(cards) > 2:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:60]:
                        t = card.select_one('h2, h3, h4, .title, [class*="title"]')
                        p = card.select_one('[class*="price"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://sandberg-estates.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Sandberg Estates',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
            
            if not results:
                # Dump some page structure
                print(f"  Page length: {len(html)}")
                # Check for JSON data
                import json as json_mod
                scripts = soup.select('script')
                for s in scripts:
                    text = s.string or ''
                    if 'property' in text.lower() and len(text) > 200:
                        try:
                            # Try to find JSON array
                            m = re.search(r'(\[{.*?}\])', text, re.DOTALL)
                            if m:
                                data = json_mod.loads(m.group(1))
                                print(f"  Found JSON array with {len(data)} items")
                        except:
                            pass
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_private_property():
    """privatepropertymallorca.com"""
    print("\n[4] Private Property Mallorca...")
    results = []
    try:
        # Try different property listing paths
        for url in [
            "https://www.privatepropertymallorca.com/?s=&bedrooms=5",
            "https://www.privatepropertymallorca.com/properties/",
            "https://www.privatepropertymallorca.com/",
        ]:
            html, final = get_page_content(url)
            if not html:
                continue
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'} at {final}")
            
            for sel in ['.property-item', '.listing-item', 'article', '.property-card', '.listing', '.property']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:50]:
                        t = card.select_one('h2, h3, h4, .title')
                        p = card.select_one('[class*="price"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.privatepropertymallorca.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Private Property Mallorca',
                                'url': href or url,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_minkner():
    """minkner.com"""
    print("\n[5] Minkner & Bonitz...")
    results = []
    try:
        html, final = get_page_content("https://www.minkner.com/immobilien/kaufen/?zimmer_min=5", wait_for='.property, article, .listing-item')
        if not html:
            html, final = get_page_content("https://www.minkner.com/immobilien/kaufen/")
        if html:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'}")
            
            for sel in ['.property-item', '.immobilien-item', 'article.immobilien', '.expose-item', 'article', '.property', '.listing-item', '[class*="expose"]', '[class*="property"]']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:60]:
                        t = card.select_one('h2, h3, h4, .title, .expose-title')
                        p = card.select_one('[class*="price"], [class*="preis"], .kaufpreis')
                        a = card.select_one('a[href]')
                        rooms = card.select_one('[class*="zimmer"], [class*="room"]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.minkner.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Minkner & Bonitz',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                                'zimmer': clean_int(rooms.get_text() if rooms else None),
                            })
                    break
    except Exception as e:
        print(f"  Error: {e}")
        traceback.print_exc()
    print(f"  → {len(results)} Objekte")
    return results

def scrape_dahler():
    """dahlercompany.com"""
    print("\n[6] DAHLER Company...")
    results = []
    try:
        html, final = get_page_content("https://www.dahlercompany.com/de/mallorca/immobilie-kaufen")
        if html:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'}")
            
            for sel in ['.property-card', '.expose', '.listing-item', 'article', '.property', '[class*="expose"]']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:50]:
                        t = card.select_one('h2, h3, h4, .title, [class*="title"]')
                        p = card.select_one('[class*="price"], [class*="preis"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.dahlercompany.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'DAHLER Company',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_lucas_fox():
    """lucasfox.com"""
    print("\n[7] Lucas Fox...")
    results = []
    try:
        # Lucas Fox has a proper search
        urls = [
            "https://www.lucasfox.com/properties-for-sale/spain/mallorca/?bedrooms=5",
            "https://www.lucasfox.com/properties-for-sale/spain/mallorca/",
            "https://www.lucasfox.com/mallorca/properties-for-sale/",
        ]
        for url in urls:
            html, final = get_page_content(url, wait_for='.property-card, article, .listing')
            if not html:
                continue
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'} at {final[:60]}")
            
            for sel in ['.property-card', '.property-item', 'article.property', '.listing-item', 'article', '[class*="property-card"]']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:60]:
                        t = card.select_one('h2, h3, h4, .title, [class*="title"]')
                        p = card.select_one('[class*="price"]')
                        a = card.select_one('a[href]')
                        beds = card.select_one('[class*="bed"], [class*="room"]')
                        area = card.select_one('[class*="area"], [class*="built"]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.lucasfox.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Lucas Fox',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                                'zimmer': clean_int(beds.get_text() if beds else None),
                                'wohnflaeche': clean_int(area.get_text() if area else None),
                            })
                    break
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_knight_frank():
    """knightfrank.com"""
    print("\n[8] Knight Frank...")
    results = []
    try:
        urls = [
            "https://www.knightfrank.com/buy/spain/balearic-islands/mallorca/?beds=5",
            "https://www.knightfrank.com/buy/spain/balearic-islands/mallorca/",
            "https://www.knightfrank.com/international-residential/for-sale/?Location=mallorca&Bedrooms=5",
        ]
        for url in urls:
            html, final = get_page_content(url)
            if not html:
                continue
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'} at {final[:60]}")
            
            for sel in ['.property-listing', '.listing-card', 'article.property', '.property-card', '.property', '[class*="listing-card"]']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:50]:
                        t = card.select_one('h2, h3, h4, .title')
                        p = card.select_one('[class*="price"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.knightfrank.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Knight Frank',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_savills():
    """savills.es"""
    print("\n[9] Savills...")
    results = []
    try:
        urls = [
            "https://www.savills.es/find-a-property/residential-property-for-sale/spain/mallorca/?bedrooms=5",
            "https://www.savills.es/find-a-property/residential-property-for-sale/spain/mallorca/",
            "https://www.savills.com/buy/international-property-for-sale/spain/balearic-islands/mallorca.html",
        ]
        for url in urls:
            html, final = get_page_content(url)
            if not html:
                continue
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'} at {final[:60]}")
            
            for sel in ['.property-card', '.listing-result', 'article', '.property', '.search-result', '[class*="property-card"]', '[class*="listing"]']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:50]:
                        t = card.select_one('h2, h3, h4, .title')
                        p = card.select_one('[class*="price"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.savills.es' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Savills',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_coldwell():
    """coldwellbanker.es"""
    print("\n[10] Coldwell Banker...")
    results = []
    try:
        html, final = get_page_content("https://www.coldwellbanker.es/propiedades/?habitaciones=5&zona=mallorca")
        if not html:
            html, final = get_page_content("https://www.coldwellbanker.es/propiedades/?zona=mallorca")
        if html:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'}")
            
            for sel in ['.property-card', '.listing', 'article', '.property', '.propiedad', '[class*="property"]']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:50]:
                        t = card.select_one('h2, h3, h4, .title')
                        p = card.select_one('[class*="price"], [class*="precio"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.coldwellbanker.es' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Coldwell Banker',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_vonpoll():
    """von-poll.com"""
    print("\n[11] Von Poll Real Estate...")
    results = []
    try:
        html, final = get_page_content("https://www.von-poll.com/de/immobilien/spanien/balearen/mallorca")
        if html:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'}")
            
            for sel in ['.expose-item', '.expose-card', '.property', 'article', '[class*="expose"]', '.listing']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:60]:
                        t = card.select_one('h2, h3, h4, [class*="title"]')
                        p = card.select_one('[class*="price"], [class*="preis"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.von-poll.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Von Poll Real Estate',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_ev():
    """engelvoelkers.com"""
    print("\n[12] Engel & Völkers...")
    results = []
    try:
        urls = [
            "https://www.engelvoelkers.com/de/search/?q=&domainId=21&businessArea=residential&mode=buy&country=ESP&categories=villa,finca&rooms=5",
            "https://www.engelvoelkers.com/de/spanien/balearen/mallorca/",
            "https://www.engelvoelkers.com/de/search/?domainId=21&mode=buy&country=ESP&rooms=5",
        ]
        for url in urls:
            html, final = get_page_content(url, wait_for='.ev-property-card, [class*="property-card"]')
            if not html:
                continue
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'} at {final[:60]}")
            
            for sel in ['.ev-property-card', '[class*="property-card"]', '.property-item', 'article', '.property']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:60]:
                        t = card.select_one('h2, h3, h4, [class*="title"]')
                        p = card.select_one('[class*="price"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.engelvoelkers.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Engel & Völkers',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_sothebys():
    """sothebysrealty.com"""
    print("\n[13] Mallorca Sotheby's...")
    results = []
    try:
        urls = [
            "https://www.sothebysrealty.com/eng/sales/mallorca-sp",
            "https://www.sothebysrealty.com/eng/sales/mallorca-sp?bedrooms=5",
        ]
        for url in urls:
            html, final = get_page_content(url, wait_for='.listing-card, .property, article')
            if not html:
                continue
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'}")
            
            for sel in ['.listing-card', '.property-card', 'article', '.property', '.listing', '[class*="listing-card"]']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:50]:
                        t = card.select_one('h2, h3, h4, .title, [class*="title"]')
                        p = card.select_one('[class*="price"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.sothebysrealty.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': "Mallorca Sotheby's",
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_remax():
    """remax.es"""
    print("\n[14] Re/Max...")
    results = []
    try:
        urls = [
            "https://www.remax.es/Comprar?habitaciones=5&comunidad=Islas+Baleares",
            "https://www.remax.es/Comprar?comunidad=Islas+Baleares",
            "https://www.remax.es/Comprar?provincia=Mallorca",
        ]
        for url in urls:
            html, final = get_page_content(url, wait_for='.property-item, article, .listing')
            if not html:
                continue
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'} at {final[:60]}")
            
            for sel in ['.property-item', '.listing-item', 'article', '.property', '.propiedad']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:60]:
                        t = card.select_one('h2, h3, h4, .title')
                        p = card.select_one('[class*="price"], [class*="precio"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.remax.es' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Re/Max Mallorca',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_mallorca_dream():
    """mallorcadreamhomes.com"""
    print("\n[15] Mallorca Dream Homes...")
    results = []
    try:
        urls = [
            "https://www.mallorcadreamhomes.com/search/?transaction=sale&bedrooms=5",
            "https://www.mallorcadreamhomes.com/search/?transaction=sale",
            "https://www.mallorcadreamhomes.com/",
        ]
        for url in urls:
            html, final = get_page_content(url)
            if not html:
                continue
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'} at {final[:60]}")
            
            for sel in ['.property-card', '.listing', 'article', '.property', '.property-item']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:50]:
                        t = card.select_one('h2, h3, h4, .title')
                        p = card.select_one('[class*="price"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.mallorcadreamhomes.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Mallorca Dream Homes',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

def scrape_fincallorca():
    """fincallorca.com"""
    print("\n[16] Fincallorca...")
    results = []
    try:
        # Site returns 202 - Cloudflare challenge
        urls = [
            "https://www.fincallorca.com/mallorca/buy/",
            "https://www.fincallorca.com/buy-property-mallorca/",
            "https://www.fincallorca.com/search/?action=kaufen&location=mallorca",
        ]
        for url in urls:
            html, final = get_page_content(url, timeout=20000)
            if not html:
                continue
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.select_one('title')
            print(f"  {title.text[:60] if title else 'no title'} at {final[:60]}")
            
            for sel in ['.property-item', '.listing-item', 'article', '.property', '.property-card']:
                cards = soup.select(sel)
                if len(cards) > 1:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:60]:
                        t = card.select_one('h2, h3, h4, .title')
                        p = card.select_one('[class*="price"]')
                        a = card.select_one('a[href]')
                        if t:
                            href = a['href'] if a else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.fincallorca.com' + href
                            results.append({
                                'titel': t.get_text(strip=True),
                                'quelle': 'Fincallorca',
                                'url': href or final,
                                'preis': clean_price(p.get_text() if p else None),
                            })
                    break
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# RUN ALL
# ─────────────────────────────────────────────────────────────
scrapers = [
    ("portamallorquina", scrape_portamallorquina),
    ("unique_mallorca", scrape_unique_mallorca),
    ("sandberg", scrape_sandberg),
    ("private_property", scrape_private_property),
    ("minkner", scrape_minkner),
    ("dahler", scrape_dahler),
    ("lucas_fox", scrape_lucas_fox),
    ("knight_frank", scrape_knight_frank),
    ("savills", scrape_savills),
    ("coldwell", scrape_coldwell),
    ("vonpoll", scrape_vonpoll),
    ("ev", scrape_ev),
    ("sothebys", scrape_sothebys),
    ("remax", scrape_remax),
    ("mallorca_dream", scrape_mallorca_dream),
    ("fincallorca", scrape_fincallorca),
]

for name, scraper in scrapers:
    try:
        items = scraper()
        all_results.extend(items)
        summary[name] = len(items)
    except Exception as e:
        print(f"  FATAL {name}: {e}")
        summary[name] = 0

print(f"\n\n=== TOTAL RAW: {len(all_results)} ===")

# Save raw
with open('/Users/robin/.openclaw/workspace/mallorca-projekt/phase2_playwright_raw.json', 'w') as f:
    json.dump(all_results, f, ensure_ascii=False, indent=2)
print("Saved to phase2_playwright_raw.json")

# Save to Excel
saved = save_to_excel(all_results)
print(f"\n=== SAVED TO EXCEL: {saved} neue Objekte ===")

print("\n--- Summary ---")
for name, count in sorted(summary.items(), key=lambda x: -x[1]):
    print(f"  {name}: {count}")
