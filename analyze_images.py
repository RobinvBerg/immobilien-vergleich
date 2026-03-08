#!/usr/bin/env python3
"""
Analyze Mallorca property images using GPT-4o Vision
and fill Excel columns D, V, AF, AG, AC
"""

import json
import os
import sys
import time
import openpyxl
from openai import OpenAI

# Config
XLSX = "mallorca-kandidaten-v2.xlsx"
JSON = "mallorca-bilder.json"
CHECKPOINT_INTERVAL = 25
START_ROW = 15
END_ROW = 333

# Column indices (1-based)
COL_D = 4   # Charme/Ästhetik (1-5)
COL_V = 22  # Renovierung (0-100)
COL_AC = 29 # Gebäudestruktur
COL_AF = 32 # Reno-Score (0-100)
COL_AG = 33 # Reno-Begründung

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
client = OpenAI(api_key=OPENAI_API_KEY)

def analyze_image(img_url: str, desc: str = "", makler: str = "") -> dict:
    """Analyze a property image using GPT-4o Vision"""
    
    prompt = f"""Du analysierst ein Bild einer Mallorca-Immobilie (Finca/Landgut/Villa).
Beschreibung: {desc[:500] if desc else 'N/A'}

Bewerte auf Basis des Bildes:

1. **Charme/Ästhetik (1-5):**
   - 5 = außergewöhnlich, einzigartig, sofort verliebt
   - 4 = sehr schön, hochwertig, klarer Charakter
   - 3 = solide, ansprechend aber nichts Besonderes
   - 2 = funktional, wenig Charme
   - 1 = wenig ansprechend, stark renovierungsbedürftig

2. **Renovierungszustand (0-100):**
   - 90-100 = einzugsbereit, perfekter Zustand
   - 70-85 = guter Zustand, kleine Arbeiten nötig
   - 40-65 = sichtbarer Renovierungsbedarf
   - 10-35 = starker Renovierungsbedarf / Rohling

3. **Gebäudestruktur** (kurze Beschreibung):
   - Typ: Finca / Villa / Herrenhaus / Landhaus / Neubau-Villa
   - Stil: Rustikal / Modern / Historisch / Mediterran
   - Sichtbares: Pool / Gästehaus / Ställe / Weinberg / Panoramablick / Meerblick

4. **Reno-Begründung** (1 kurzer Satz auf Deutsch, erklärt Renovierungsscore)

Antworte NUR als JSON:
{{
  "charme": <int 1-5>,
  "renovierung": <int 0-100>,
  "gebaeude": "<Typ - Stil - Sichtbares>",
  "reno_begruendung": "<1 Satz>"
}}"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {"url": img_url, "detail": "low"}
                        },
                        {
                            "type": "text",
                            "text": prompt
                        }
                    ]
                }
            ],
            max_tokens=300,
            temperature=0.3
        )
        
        text = response.choices[0].message.content.strip()
        # Extract JSON from response
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0].strip()
        elif "```" in text:
            text = text.split("```")[1].split("```")[0].strip()
        
        result = json.loads(text)
        return {
            "charme": int(result.get("charme", 3)),
            "renovierung": int(result.get("renovierung", 60)),
            "gebaeude": str(result.get("gebaeude", "Nicht verfügbar")),
            "reno_begruendung": str(result.get("reno_begruendung", "Keine Begründung verfügbar"))
        }
    except Exception as e:
        print(f"  ERROR analyzing: {e}")
        return {
            "charme": 3,
            "renovierung": 60,
            "gebaeude": "Nicht verfügbar",
            "reno_begruendung": f"Bildanalyse fehlgeschlagen: {str(e)[:80]}"
        }

def main():
    print("Loading data...")
    with open(JSON) as f:
        bilder = json.load(f)
    
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    
    rows = sorted([int(k) for k in bilder.keys() if START_ROW <= int(k) <= END_ROW])
    print(f"Total rows to process: {len(rows)}")
    
    stats = {"processed": 0, "errors": 0, "skipped": 0}
    processed_since_checkpoint = 0
    
    for i, row_num in enumerate(rows):
        row_data = bilder[str(row_num)]
        img_url = row_data.get("img_url", "")
        desc = row_data.get("desc", "")
        makler = row_data.get("makler", "")
        
        print(f"[{i+1}/{len(rows)}] Row {row_num}: {img_url[:80]}...")
        
        if not img_url:
            print(f"  No image URL, using defaults")
            result = {"charme": 3, "renovierung": 60, "gebaeude": "Nicht verfügbar", "reno_begruendung": "Kein Bild verfügbar"}
            stats["skipped"] += 1
        else:
            result = analyze_image(img_url, desc, makler)
            stats["processed"] += 1
        
        # Write to Excel
        ws.cell(row=row_num, column=COL_D).value = result["charme"]
        ws.cell(row=row_num, column=COL_V).value = result["renovierung"]
        ws.cell(row=row_num, column=COL_AC).value = result["gebaeude"]
        ws.cell(row=row_num, column=COL_AF).value = result["renovierung"]  # Same as V
        ws.cell(row=row_num, column=COL_AG).value = result["reno_begruendung"]
        
        print(f"  Charme={result['charme']}, Reno={result['renovierung']}, Gebäude={result['gebaeude'][:50]}")
        
        processed_since_checkpoint += 1
        
        # Checkpoint every 25 rows
        if processed_since_checkpoint >= CHECKPOINT_INTERVAL:
            wb.save(XLSX)
            print(f"  💾 CHECKPOINT saved at row {row_num}")
            processed_since_checkpoint = 0
        
        # Small delay to avoid rate limiting
        time.sleep(0.3)
    
    # Final save
    wb.save(XLSX)
    print(f"\n✅ DONE! Final save.")
    print(f"Stats: processed={stats['processed']}, skipped={stats['skipped']}, errors={stats['errors']}")
    print(f"Total rows handled: {len(rows)}")

if __name__ == "__main__":
    main()
