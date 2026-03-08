from openpyxl import load_workbook
from datetime import date

wb = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
ws = wb['Mallorca Objekte']
existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
print(f"Existing URLs: {len(existing_urls)}")

# Alle gesammelten Von Poll Expose-URLs aus den Brave-Suchen
listings = [
    ("Spektakuläre Wohnung am Meer in Portixol, Mallorca", "https://www.von-poll.com/de/expose/mallorca-santa-maria/4160036805"),
    ("Neubauprojekt Luxusvilla Son Vida Mallorca", "https://www.von-poll.com/de/expose/mallorca-santa-maria/neubauprojekt-fr-eine-luxusvilla-mit-pool-und-traumblick-in-son-vida-mallorca-4160034665"),
    ("Einmalige Luxus-Finca mit weitläufigem Grundstück", "https://www.von-poll.com/de/expose/mallorca-santa-maria/einmalige-luxus-finca-und-weitlaufigem-grundstuck-4160045157"),
    ("Einmalige Luxus-Finca mit Privatgolfplatz", "https://www.von-poll.com/de/expose/mallorca-santa-maria/einmalige-luxus-finca-mit-privatgolfplatz-und-weitlaufigem-grundstuck-4160045157"),
    ("Authentische rustikale Finca Artá Mallorca", "https://www.von-poll.com/de/expose/mallorca-pollensa/authentische-rustikale-finca-auf-einem-groen-grundstck-zum-renovieren-in-art-mallorca-3520033555"),
    ("Attraktives Dorfhaus in Bunyola mit Panoramablick und Pool", "https://www.von-poll.com/de/expose/mallorca-santa-maria/attraktives-dorfhaus-in-bunyola-mit-panoramablick-und-pool-4160041027"),
    ("Traditionelle Finca Santa Maria Bunyola Meerblick", "https://www.von-poll.com/de/expose/mallorca-santa-maria/traumhafte-traditionelle-finca-zwischen-santa-maria-und-bunyola-mit-herrlichem-weitblick-bis-zur-bucht-von-palma-4160046911"),
    ("Exklusives Penthouse Paseo Marítimo Palma", "https://www.von-poll.com/de/expose/mallorca-santa-maria/exklusives-penthouse-am-paseo-maritimo-wohnen-auf-hochstem-niveau-in-palma-4160046539"),
    ("Exklusive Luxusresidenzen Paseo Marítimo Palma (1)", "https://www.von-poll.com/de/expose/mallorca-santa-maria/exclusive-luxury-residences-on-the-paseo-maritimo-living-at-the-highest-level-in-palma-4160047615"),
    ("Exklusive Luxusresidenzen Paseo Marítimo Palma (2)", "https://www.von-poll.com/de/expose/mallorca-santa-maria/exclusive-luxury-residences-on-the-paseo-maritimo-living-at-the-highest-level-in-palma-4160047801"),
    ("Exklusive Luxusresidenzen Paseo Marítimo Palma (3)", "https://www.von-poll.com/de/expose/mallorca-santa-maria/exclusive-luxury-residences-on-the-paseo-maritimo-living-at-the-highest-level-in-palma-4160047621"),
    ("Exklusives Penthouse Paseo Marítimo (2)", "https://www.von-poll.com/de/expose/mallorca-santa-maria/exklusive-luxusresidenzen-am-paseo-maritimo-wohnen-auf-hochstem-niveau-in-palma-4160046539"),
    ("Schöne Finca mit Garten und Pool in Pollensa", "https://www.von-poll.com/de/expose/mallorca-pollensa/schne-finca-mit-groem-garten-und-schwimmbad-in-pollensa-mallorca-3520033375"),
    ("Spektakuläre Finca mit Pool in Pollensa", "https://www.von-poll.com/de/expose/mallorca-pollensa/spettacolare-finca-con-piscina-in-una-zona-privilegiata-di-maiorca-a-pollensa-3520043985007"),
    ("Finca im mallorquinischen Stil zwischen Alaró und Santa Maria", "https://www.von-poll.com/de/expose/mallorca-santa-maria/finca-im-mallorquinischen-stil-mit-pool-in-einzigartiger-lage-zwischen-alar-und-santa-maria-del-cam-4160043161"),
    ("Elegante Finca mit Weitblick zum Meer", "https://www.von-poll.com/de/expose/mallorca-portals/elegante-finca-privat-gelegen-mit-weitblick-ueber-die-landschaft-bis-zum-meer-290005169"),
    ("Wunderschöne Finca mit Pool nahe Pollensa Zentrum", "https://www.von-poll.com/de/expose/mallorca-pollensa/wunderschone-finca-mit-pool-in-der-nahe-des-zentrums-von-pollensa-3520040755"),
    ("Neubau Finca Es-Trenc-Strand Campos Pool", "https://www.von-poll.com/de/expose/mallorca-santa-maria/grozgige-neubaufinca-in-der-nhe-des-estrencstrandes-bei-campos-mit-pool-und-eigenem-weinanbau-4160043131"),
    ("Fantastische Finca Felanitx mit Infinity-Pool", "https://www.von-poll.com/de/expose/mallorca-pollensa/fantasztikus-finca-festoi-helyen-felanitxban-3520047423006"),
    ("Neu gebaute Finca mit Pool in Alaró", "https://www.von-poll.com/de/expose/mallorca-santa-maria/wunderschne-neu-gebaute-finca-mit-schwimmbad-in-santa-mara-del-cam-mallorca-4160032129"),
    ("Villa mit Pool und Meerblick Bucht Pollensa Bonaire", "https://www.von-poll.com/de/expose/mallorca-pollensa/3520047733"),
    ("Spektakuläre Villa am Meer mit Pool Barcarés Alcudia", "https://www.von-poll.com/de/expose/mallorca-pollensa/3520046907"),
    ("Fantastisch renovierte Villa Pollensa", "https://www.von-poll.com/de/expose/mallorca-pollensa/fantasztikusan-felujitott-villa-latvanyos-kilatassal-pollensa-ban-3520045095006"),
    ("Traumhafte Villa Cala Tuent mit Meerblick", "https://www.von-poll.com/de/expose/mallorca-pollensa/traumhafte-villa-in-cala-tuent-mit-spektakularem-meerblick-3520037685"),
    ("Neues Landhaus Pollensa Alcudia Pool ETV-Lizenz (1)", "https://www.von-poll.com/de/expose/mallorca-pollensa/wunderschnes-neu-erbautes-landhaus-zwischen-pollensa-und-alcudia-mit-pool-und-etv-vermietlizenz-3520037587"),
    ("Neues Landhaus Pollensa Alcudia Pool ETV-Lizenz (2)", "https://www.von-poll.com/de/expose/mallorca-pollensa/hermosa-casa-de-campo-de-nueva-construccin-situada-entre-pollensa-y-alcudia-con-piscina-3520037589"),
    ("Gemütliches Haus Cala San Vicente Strand", "https://www.von-poll.com/de/expose/mallorca-pollensa/gemtliches-haus-mit-charme-und-groem-potenzial-in-der-nhe-des-herrlichen-strandes-in-cala-san-vicente-3520038699"),
    ("Luxusvilla erste Meereslinie Sa Torre Llucmajor", "https://www.von-poll.com/de/expose/mallorca-pollensa/luxusvilla-in-erster-meereslinie-an-der-kuste-von-sa-torre-llucmajor-3520047335"),
    ("Traumvilla Cala Tuent nahe Sa Calobra Pool Meerblick", "https://www.von-poll.com/de/expose/mallorca-pollensa/traumvilla-in-cala-tuent-in-der-nahe-von-sa-calobra-mit-pool-und-atemberaubendem-meerblick-3520020397001"),
    ("SECRET MARKETING Villa Meerblick Bucht Pollensa", "https://www.von-poll.com/de/expose/mallorca-pollensa/3520044185"),
    ("Exklusives Grundstück mit Projekt Alcanada", "https://www.von-poll.com/de/expose/mallorca-pollensa/3520044041"),
    ("Wohnung mit Meerblick Playa de Palma", "https://www.von-poll.com/de/expose/mallorca-llucmajor/wohnung-mit-meerblick-direkt-an-der-playa-de-palma-299007921"),
    ("Spektakuläre Wohnung am Meer in Portixol (2)", "https://www.von-poll.com/de/expose/mallorca-santa-maria/spektakulare-wohnung-am-meer-in-portixol-mallorca-4160036797"),
    ("Spektakuläre Wohnung am Meer Portixol (3)", "https://www.von-poll.com/de/expose/mallorca-santa-maria/spektakulre-wohnung-am-meer-in-portixol-mallorca-4160036797"),
    ("Neubau-Luxus-Penthouse Meerblick Paguera", "https://www.von-poll.com/de/expose/mallorca-paguera/neubauluxuspenthouse-mit-traumhaftem-meerblick-391005255"),
    ("Hervorragendes Haus mit Pool Pollensa Stadt", "https://www.von-poll.com/de/expose/mallorca-pollensa/3520034463"),
    ("Traditionelles Stadthaus mit Garten Pollensa", "https://www.von-poll.com/de/expose/mallorca-pollensa/3520043167"),
    ("Historischer Stadtpalast Alaró Garten", "https://www.von-poll.com/de/expose/mallorca-santa-maria/historischer-vollstandig-renovierter-stadtpalast-in-alaro-mit-grossem-garten-und-weitblick-4160039777"),
    ("Neu gebautes Dorfhaus Pool Traumblick Felanitx", "https://www.von-poll.com/de/expose/mallorca-santa-maria/neu-gebautes-dorfhaus-mit-pool-und-traumblick-in-felanitx-4160043777"),
    ("Historisches Herrenhaus Sineu", "https://www.von-poll.com/de/expose/mallorca-santa-maria/historic-manor-house-in-sineu-4160043151"),
    ("Charmantes Apartment im Herzen von Bunyola", "https://www.von-poll.com/de/expose/mallorca-santa-maria/piso-con-encanto-en-el-corazon-de-bunyola-4160045705"),
    ("Stilvoll renoviertes Apartment in Meernähe", "https://www.von-poll.com/de/expose/mallorca-santa-maria/stylishly-renovated-apartment-near-the-sea-4160046931"),
    ("Herrliche Wohnung mit Terrasse Meerblick Paseo Marítimo", "https://www.von-poll.com/de/expose/mallorca-santa-maria/4160035591"),
    ("Außergewöhnliche Finca Santa Maria del Camí Pool", "https://www.von-poll.com/de/expose/mallorca-santa-maria/auergewhnliche-finca-in-santa-maria-del-cam-mit-pool-und-mediterranem-garten-4160042177"),
    ("Moderne Neubau-Finca Santa Maria del Camí", "https://www.von-poll.com/de/expose/mallorca-santa-maria/moderna-finca-di-nuova-costruzione-in-una-posizione-privilegiata-con-assoluta-privacy-a-santa-maria-del-cami-4160045047007"),
    ("Santa Ponsa Living Neubau-Luxus-Penthouse", "https://www.von-poll.com/de/expose/mallorca-andratx/santa-ponsa-living-neubauluxuspenthouse-mit-meer-und-panoramablick-391004839"),
    ("Ebenerdige Neubau-Villa Marratxí", "https://www.von-poll.com/de/expose/mallorca-santa-maria/ebenerdige-neubauvilla-am-stadtrand-von-marratx-4160041741"),
    ("Neubau-Luxus-Apartment Panoramablick Santa Ponsa", "https://www.von-poll.com/de/expose/mallorca-paguera/391004863"),
    ("Modernes Luxus-Landhaus Santa María del Camí (1)", "https://www.von-poll.com/de/expose/mallorca-santa-maria/modernes-und-komfortables-luxuslandhaus-in-santa-mara-del-cam-4160034369"),
    ("Modernes Luxus-Landhaus Santa María del Camí (2)", "https://www.von-poll.com/de/expose/mallorca-santa-maria/moderne-und-komfortable-luxus-landhaus-in-santa-mara-del-cam-4160034369"),
    ("Exklusiv renovierte Luxuswohnung mit Meerblick Paguera (1)", "https://www.von-poll.com/de/expose/mallorca-paguera/exklusiv-perfekt-renovierte-moderne-luxuswohnung-mit-meerblick-391005303"),
    ("Exklusiv renovierte Luxuswohnung mit Meerblick Paguera (2)", "https://www.von-poll.com/de/expose/mallorca-paguera/391005303"),
    ("Meerblick-Grundstück mit Projekt Portals Nous", "https://www.von-poll.com/de/expose/mallorca-santa-maria/meerblickgrundstck-mit-projekt-in-portals-nous-4160040935"),
    ("Erstklassiges Grundstück Portals Nous mit Projekt (1)", "https://www.von-poll.com/de/expose/mallorca-santa-maria/erstklassiges-grundstck-in-portals-nous-mit-projekt-4160040915"),
    ("Erstklassiges Grundstück Portals Nous mit Projekt (2)", "https://www.von-poll.com/de/expose/mallorca-santa-maria/prime-plot-for-sale-in-portals-nous-with-project-4160040919"),
    ("Landhaus Hanglage Fernblick Establiments", "https://www.von-poll.com/de/expose/mallorca-portals/landhaus-in-ruhiger-unverbaubarer-hanglage-mit-fernblick-in-establiments-29000211"),
    ("Neubau Wohnungen Gemeinschaftspool Cala d'Or", "https://www.von-poll.com/de/expose/mallorca-llucmajor/neubau-wohnungen-mit-gemeinschaftspool-in-hafennhe-von-cala-dor-299007967"),
    ("Luxuriöse Familienvilla Cala Corda", "https://www.von-poll.com/de/expose/mallorca-paguera/luxurise-familienvilla-in-lauflage-zur-cala-corda-391003745"),
    ("Fantastische Finca Panoramablick Consell Alaró Santa Maria", "https://www.von-poll.com/de/expose/mallorca-santa-maria/fantastische-finca-mit-panoramablick-bei-consell-zwischen-alar-und-santa-maria-del-cam-4160036693"),
    ("Zwei Baugrundstücke für Villen in Marratxí", "https://www.von-poll.com/de/expose/mallorca-santa-maria/groes-stdtisches-grundstck-zum-bau-einer-villa-mit-pool-in-marratxi-4160044011"),
    ("Traditionelles Stadthaus mit Charakter Andratx", "https://www.von-poll.com/de/expose/mallorca-paguera/traditionelles-stadthaus-mit-charakter-391002569"),
    ("Großes Grundstück Bonaire Alcúdia", "https://www.von-poll.com/de/expose/mallorca-pollensa/3520043929"),
    ("Spektakuläres Anwesen 18. Jahrhundert Sóller", "https://www.von-poll.com/de/expose/mallorca-santa-maria/spektakulares-renoviertes-anwesen-aus-dem-18-jahrhundert-mit-historischem-charme-und-beeindruckendem-grundstuck-in-soller-4160035677"),
    ("Schöne Finca im Orangen Tal bei Sóller", "https://www.von-poll.com/de/expose/mallorca-santa-maria/schone-finca-im-orangen-tal-bei-soller-4160037769"),
    ("Historischer Stadtpalast Sóller mit Garten", "https://www.von-poll.com/de/expose/mallorca-santa-maria/historischer-stadtpalast-im-herzen-von-soller-mit-einmaligem-garten-4160043833"),
    ("Herrenhausvilla 1925 Strandlage Puerto de Sóller", "https://www.von-poll.com/de/expose/mallorca-pollensa/fantastic-1925-s-manor-villa-for-sale-on-the-beachfront-in-puerto-de-soller-3520031211"),
    ("Exklusive Finca Sóller-Tal mediterran", "https://www.von-poll.com/de/expose/mallorca-santa-maria/exklusive-finca-im-herzen-des-soller-tals-ruhe-natur-und-mediterrane-lebensart-vereint-4160047859"),
    ("Design trifft Gemütlichkeit Rückzugsort Sóller", "https://www.von-poll.com/de/expose/mallorca-santa-maria/il-design-incontra-il-comfort-un-rifugio-a-soller-4160044925007"),
    ("Stadthaus mit Renovierungsprojekt Pool Pollensa", "https://www.von-poll.com/de/expose/mallorca-pollensa/3520034505"),
]

today = str(date.today())
count = 0

for title, url in listings:
    url = url.strip()
    if url not in existing_urls:
        ws.append([title, 'Von Poll Real Estate', url, None, None, None, None, 'Mallorca', today, 'Neu'])
        existing_urls.add(url)
        count += 1
        print(f"  ✅ {title[:60]}")
    else:
        print(f"  ⏭️  SKIP (exists): {url[-50:]}")

wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
print(f"\n✅ Von Poll: {count} neue Objekte gespeichert (von {len(listings)} gesammelt)")
