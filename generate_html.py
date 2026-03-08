#!/usr/bin/env python3
"""Generate mallorca-ranking-live.html from mallorca-kandidaten-v2.xlsx"""

import openpyxl
import json
import os
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, 'mallorca-kandidaten-v2.xlsx')
OUT  = os.path.join(BASE, 'mallorca-ranking-live.html')
BILDER = os.path.join(BASE, 'bilder')

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['Mallorca Kandidaten']
ws_e = wb['Einstellungen']

# --- Read settings ---
settings = {}
reach_weights = {}
for row in ws_e.iter_rows(values_only=True):
    if not row[0]: continue
    settings[str(row[0]).strip()] = row[1]
    if str(row[0]).strip() in ('Flughafen','Daia (Hotel Corazón)','Ses Salines','Andratx'):
        reach_weights[str(row[0]).strip()] = {'ideal': row[1], 'ok': row[2], 'max': row[3], 'w': row[4]}

# --- Read all objects ---
objects = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 2).value
    if not name: continue

    nr       = ws.cell(r, 1).value
    url      = ws.cell(r, 3).value or ''
    charme   = ws.cell(r, 4).value
    zimmer   = ws.cell(r, 5).value
    baeder   = ws.cell(r, 6).value
    grund    = ws.cell(r, 7).value
    flaeche  = ws.cell(r, 8).value
    garten   = ws.cell(r, 9).value
    ort      = ws.cell(r,10).value or ''
    f_km     = ws.cell(r,11).value
    f_min    = ws.cell(r,12).value
    d_km     = ws.cell(r,13).value
    d_min    = ws.cell(r,14).value
    a_km     = ws.cell(r,15).value
    a_min    = ws.cell(r,16).value
    s_km     = ws.cell(r,17).value
    s_min    = ws.cell(r,18).value
    preis    = ws.cell(r,19).value
    epm2_b   = ws.cell(r,20).value
    epm2_g   = ws.cell(r,21).value
    reno     = ws.cell(r,22).value
    bewirt   = ws.cell(r,23).value
    vermiet  = ws.cell(r,24).value
    erreich  = ws.cell(r,25).value
    score    = ws.cell(r,26).value
    rang     = ws.cell(r,28).value
    baujahr  = ws.cell(r,30).value
    reno_score = ws.cell(r,32).value
    reno_text  = ws.cell(r,33).value or ''
    kommentar  = ws.cell(r,34).value or ''
    beschreib  = ws.cell(r,35).value or ''
    makler     = ws.cell(r,36).value or ''
    makler_ref = ws.cell(r,37).value or ''
    link_status= ws.cell(r,38).value or ''
    anzeige    = ws.cell(r,39).value or name
    gaestehaus = ws.cell(r,40).value

    # Find image
    img_path = None
    for suffix in [f'{nr}_main.jpg', f'{nr}_main.png', f'{nr}.jpg', f'{nr}.jpeg', f'{nr}.png']:
        p = os.path.join(BILDER, suffix)
        if os.path.exists(p):
            img_path = f'bilder/{suffix}'
            break

    sold = str(link_status).lower() in ('sold','delisted','offline','verkauft','nicht mehr verfügbar')
    offline = str(link_status).lower() in ('offline','nicht mehr verfügbar')

    objects.append({
        'nr': nr, 'name': str(name), 'anzeige': str(anzeige),
        'url': str(url), 'charme': charme, 'zimmer': zimmer,
        'baeder': baeder, 'grund': grund, 'flaeche': flaeche,
        'garten': garten, 'ort': str(ort),
        'f_min': f_min, 'd_min': d_min, 'a_min': a_min, 's_min': s_min,
        'preis': preis, 'epm2_b': epm2_b, 'epm2_g': epm2_g,
        'reno': reno, 'reno_score': reno_score, 'reno_text': str(reno_text),
        'bewirt': bewirt, 'vermiet': vermiet, 'erreich': erreich,
        'score': score, 'rang': rang,
        'baujahr': baujahr, 'gaestehaus': gaestehaus,
        'kommentar': str(kommentar), 'beschreib': str(beschreib),
        'makler': str(makler), 'makler_ref': str(makler_ref),
        'img': img_path, 'sold': sold, 'offline': offline,
        'link_status': str(link_status),
    })

# Sort by rang
objects.sort(key=lambda x: (x['rang'] or 9999))

def fmt_preis(v):
    if not v: return '–'
    m = v / 1_000_000
    if m >= 1: return f'€{m:.2f}M'
    return f'€{int(v):,}'

def fmt_num(v, suffix=''):
    if v is None: return '–'
    return f'{int(v):,}{suffix}'.replace(',','.')

def score_color(s):
    if s is None: return '#666'
    if s >= 70: return '#4caf50'
    if s >= 55: return '#f0a500'
    return '#e53935'

# Build JS data
js_data = json.dumps(objects, ensure_ascii=False, default=str)

ts = datetime.now().strftime('%d.%m.%Y %H:%M')

html = f'''<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Mallorca Ranking — Live {ts}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0d0d14;color:#e8e8e8;min-height:100vh}}
header{{background:linear-gradient(135deg,#1a1a2e,#16213e);padding:20px 32px;border-bottom:1px solid #2a2a3e;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px}}
header h1{{font-size:1.5rem;font-weight:700;color:#fff}}
header h1 span{{color:#f0a500}}
.meta{{font-size:0.8rem;color:#666}}
.controls{{background:#13131e;padding:16px 32px;border-bottom:1px solid #22223a;display:flex;flex-wrap:wrap;gap:12px;align-items:flex-end}}
.cg{{display:flex;flex-direction:column;gap:4px}}
.cg label{{font-size:0.7rem;color:#888;text-transform:uppercase;letter-spacing:.05em}}
input,select{{background:#1c1c2e;border:1px solid #333;color:#e8e8e8;padding:6px 10px;border-radius:6px;font-size:0.83rem;outline:none}}
input:focus,select:focus{{border-color:#f0a500}}
input[type=range]{{width:110px;padding:4px 0;accent-color:#f0a500}}
.rv{{font-size:0.72rem;color:#f0a500;text-align:center}}
.btn{{background:#f0a500;color:#000;border:none;padding:7px 14px;border-radius:6px;font-weight:600;cursor:pointer;font-size:0.83rem}}
.btn:hover{{background:#ffc107}}
.btn.sec{{background:#22223a;color:#e8e8e8}}
.btn.sec:hover{{background:#2e2e4a}}
.sortbar{{background:#10101a;padding:8px 32px;display:flex;gap:8px;flex-wrap:wrap;align-items:center;border-bottom:1px solid #1a1a2a}}
.sortbar span{{font-size:0.72rem;color:#555;margin-right:2px}}
.sb{{background:none;border:1px solid #252535;color:#999;padding:3px 10px;border-radius:4px;cursor:pointer;font-size:0.78rem}}
.sb.active,.sb:hover{{background:#f0a500;color:#000;border-color:#f0a500;font-weight:600}}
#grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(370px,1fr));gap:18px;padding:22px 32px}}
.card{{background:#181828;border-radius:10px;overflow:hidden;border:1px solid #252535;transition:transform .15s,border-color .15s;cursor:pointer}}
.card:hover{{transform:translateY(-2px);border-color:#f0a500}}
.card.hidden{{display:none}}
.card.sold{{opacity:.65}}
.ci{{position:relative;height:210px;overflow:hidden;background:#0e0e1c}}
.ci img{{width:100%;height:100%;object-fit:cover;transition:transform .3s}}
.card:hover .ci img{{transform:scale(1.04)}}
.noimg{{display:flex;align-items:center;justify-content:center;height:100%;color:#333;font-size:.82rem}}
.ribbon{{position:absolute;top:14px;right:-28px;width:120px;background:#e53935;color:#fff;font-size:.68rem;font-weight:700;text-align:center;padding:4px 0;transform:rotate(45deg);text-transform:uppercase;z-index:10}}
.ribbon.off{{background:#1565c0}}
.b-rang{{position:absolute;top:8px;left:8px;background:rgba(0,0,0,.75);color:#f0a500;font-weight:700;font-size:.82rem;padding:2px 9px;border-radius:20px;backdrop-filter:blur(4px)}}
.b-score{{position:absolute;top:8px;right:8px;color:#000;font-weight:700;font-size:.88rem;padding:2px 9px;border-radius:20px}}
.b-makler{{position:absolute;bottom:6px;left:8px;background:rgba(0,0,0,.7);color:#bbb;font-size:.68rem;padding:2px 7px;border-radius:10px;max-width:180px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.cb{{padding:12px 14px}}
.card-title{{font-size:.9rem;font-weight:600;color:#f0f0f0;line-height:1.3;margin-bottom:6px;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden}}
.card-ort{{font-size:.75rem;color:#888;margin-bottom:8px}}
.stats{{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:8px}}
.stat{{background:#0d0d1a;border:1px solid #252535;border-radius:5px;padding:3px 7px;font-size:.72rem;color:#ccc}}
.stat span{{color:#f0a500;font-weight:600}}
.preis-row{{display:flex;justify-content:space-between;align-items:center}}
.preis{{font-size:1.05rem;font-weight:700;color:#fff}}
.epm2{{font-size:.72rem;color:#666}}
/* Modal */
.overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.8);z-index:100;align-items:center;justify-content:center;padding:20px}}
.overlay.open{{display:flex}}
.modal{{background:#181828;border-radius:12px;max-width:760px;width:100%;max-height:90vh;overflow-y:auto;border:1px solid #2a2a3e}}
.mh{{position:relative;height:320px;overflow:hidden;background:#0e0e1c}}
.mh img{{width:100%;height:100%;object-fit:cover}}
.mh-close{{position:absolute;top:12px;right:12px;background:rgba(0,0,0,.7);color:#fff;border:none;border-radius:50%;width:34px;height:34px;font-size:1.1rem;cursor:pointer;display:flex;align-items:center;justify-content:center}}
.mb{{padding:22px 24px}}
.mb h2{{font-size:1.15rem;font-weight:700;color:#fff;margin-bottom:4px}}
.mb .ort{{color:#888;font-size:.82rem;margin-bottom:16px}}
.grid2{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:16px}}
.kv{{background:#0d0d1a;border-radius:7px;padding:8px 12px}}
.kv .k{{font-size:.68rem;color:#666;text-transform:uppercase;letter-spacing:.05em}}
.kv .v{{font-size:.9rem;font-weight:600;color:#f0f0f0;margin-top:2px}}
.desc{{font-size:.82rem;color:#aaa;line-height:1.5;margin-bottom:14px}}
.komm{{background:#1a1a2e;border-left:3px solid #f0a500;padding:8px 12px;font-size:.8rem;color:#ccc;border-radius:0 6px 6px 0;margin-bottom:14px}}
.modal-btns{{display:flex;gap:10px}}
.modal-btns a{{padding:8px 18px;border-radius:6px;text-decoration:none;font-size:.83rem;font-weight:600}}
.modal-btns a.primary{{background:#f0a500;color:#000}}
.modal-btns a.sec{{background:#22223a;color:#e8e8e8}}
.reach-grid{{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:14px}}
.reach-item{{background:#0d0d1a;border-radius:6px;padding:6px 10px;font-size:.75rem}}
.reach-item .label{{color:#666;margin-bottom:2px}}
.reach-item .val{{font-weight:600}}
.reach-item .ok{{color:#4caf50}}
.reach-item .warn{{color:#f0a500}}
.reach-item .bad{{color:#e53935}}
#count{{font-size:.82rem;color:#666}}
.score-bar-wrap{{margin-top:8px}}
.score-bar{{height:3px;border-radius:2px;background:#f0a500;transition:width .3s}}
</style>
</head>
<body>
<header>
  <h1>🏝️ Mallorca <span>Ranking</span></h1>
  <div>
    <div id="count">{len(objects)} Objekte</div>
    <div class="meta">Stand: {ts}</div>
  </div>
</header>
<div class="controls">
  <div class="cg"><label>Suche</label><input type="text" id="search" placeholder="Name, Ort, Makler…" style="width:200px"></div>
  <div class="cg"><label>Ort</label><select id="filt-ort"><option value="">Alle Orte</option></select></div>
  <div class="cg"><label>Zimmer min.</label><select id="filt-zimmer"><option value="0">Alle</option><option value="5">5+</option><option value="6">6+</option><option value="7">7+</option><option value="8">8+</option></select></div>
  <div class="cg"><label>Vermietlizenz</label><select id="filt-lizenz"><option value="">Alle</option><option value="100">Ja (100%)</option><option value="50">Teilw. (50%)</option><option value="0">Keine</option></select></div>
  <div class="cg"><label>Status</label><select id="filt-sold"><option value="">Alle</option><option value="active">Aktiv</option><option value="sold">Sold/Offline</option></select></div>
  <div class="cg"><label>Preis max (Mio €)</label><input type="range" id="max-preis" min="1" max="30" value="30" step="0.5" oninput="document.getElementById('pv').textContent=this.value+'M'"><div class="rv" id="pv">30M</div></div>
  <button class="btn" onclick="applyFilters()">Filter</button>
  <button class="btn sec" onclick="resetFilters()">Reset</button>
</div>
<div class="sortbar">
  <span>Sortierung:</span>
  <button class="sb active" data-sort="rang" onclick="setSort(this)">Rang</button>
  <button class="sb" data-sort="score" onclick="setSort(this)">Score</button>
  <button class="sb" data-sort="preis" onclick="setSort(this)">Preis ↑</button>
  <button class="sb" data-sort="preis-desc" onclick="setSort(this)">Preis ↓</button>
  <button class="sb" data-sort="zimmer" onclick="setSort(this)">Zimmer</button>
  <button class="sb" data-sort="grund" onclick="setSort(this)">Grundstück</button>
  <button class="sb" data-sort="daia" onclick="setSort(this)">Nähe Daia</button>
  <button class="sb" data-sort="flughafen" onclick="setSort(this)">Nähe Flughafen</button>
</div>
<div id="grid"></div>

<div class="overlay" id="overlay" onclick="closeModal(event)">
  <div class="modal" id="modal">
    <div class="mh" id="m-img"><button class="mh-close" onclick="closeModalDirect()">✕</button></div>
    <div class="mb">
      <h2 id="m-title"></h2>
      <div class="ort" id="m-ort"></div>
      <div class="grid2" id="m-grid"></div>
      <div class="reach-grid" id="m-reach"></div>
      <div class="komm" id="m-komm" style="display:none"></div>
      <div class="desc" id="m-desc"></div>
      <div class="modal-btns" id="m-btns"></div>
    </div>
  </div>
</div>

<script>
const DATA = {js_data};

function fmtPreis(v) {{
  if (!v) return '–';
  const m = v / 1e6;
  return m >= 1 ? '€' + m.toFixed(2) + 'M' : '€' + Math.round(v).toLocaleString('de');
}}
function fmtNum(v, suf='') {{
  if (v == null) return '–';
  return Math.round(v).toLocaleString('de') + suf;
}}
function scoreColor(s) {{
  if (!s) return '#666';
  if (s >= 70) return '#4caf50';
  if (s >= 55) return '#f0a500';
  return '#e53935';
}}
function reachClass(min, ideal, ok, max) {{
  if (!min) return '';
  if (min <= ideal) return 'ok';
  if (min <= ok) return 'warn';
  if (min > max) return 'bad';
  return 'warn';
}}

// Populate ort filter
const orte = [...new Set(DATA.map(o => o.ort).filter(Boolean))].sort();
const ortSel = document.getElementById('filt-ort');
orte.forEach(o => {{ const op = document.createElement('option'); op.value = o; op.textContent = o; ortSel.appendChild(op); }});

let currentSort = 'rang';

function setSort(btn) {{
  document.querySelectorAll('.sb').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  currentSort = btn.dataset.sort;
  renderGrid(getFiltered());
}}

function getFiltered() {{
  const q = document.getElementById('search').value.toLowerCase();
  const minS = 0;
  const ort = document.getElementById('filt-ort').value;
  const zimMin = +document.getElementById('filt-zimmer').value;
  const liz = document.getElementById('filt-lizenz').value;
  const sold = document.getElementById('filt-sold').value;
  const maxPreis = +document.getElementById('max-preis').value * 1e6;

  return DATA.filter(o => {{
    if (q && !o.name.toLowerCase().includes(q) && !o.ort.toLowerCase().includes(q) && !o.makler.toLowerCase().includes(q)) return false;
    if (minS && (o.score || 0) < minS) return false;
    if (ort && o.ort !== ort) return false;
    if (zimMin && (o.zimmer || 0) < zimMin) return false;
    if (liz !== '' && String(o.vermiet) !== liz) return false;
    if (sold === 'active' && o.sold) return false;
    if (sold === 'sold' && !o.sold) return false;
    if (maxPreis < 30e6 && o.preis && o.preis > maxPreis) return false;
    return true;
  }});
}}

function sortData(data) {{
  const d = [...data];
  if (currentSort === 'rang') d.sort((a,b) => (a.rang||9999)-(b.rang||9999));
  else if (currentSort === 'score') d.sort((a,b) => (b.score||0)-(a.score||0));
  else if (currentSort === 'preis') d.sort((a,b) => (a.preis||1e9)-(b.preis||1e9));
  else if (currentSort === 'preis-desc') d.sort((a,b) => (b.preis||0)-(a.preis||0));
  else if (currentSort === 'zimmer') d.sort((a,b) => (b.zimmer||0)-(a.zimmer||0));
  else if (currentSort === 'grund') d.sort((a,b) => (b.grund||0)-(a.grund||0));
  else if (currentSort === 'daia') d.sort((a,b) => (a.d_min||99)-(b.d_min||99));
  else if (currentSort === 'flughafen') d.sort((a,b) => (a.f_min||99)-(b.f_min||99));
  return d;
}}

function renderGrid(data) {{
  data = sortData(data);
  document.getElementById('count').textContent = data.length + ' Objekte';
  const grid = document.getElementById('grid');
  grid.innerHTML = '';
  data.forEach(o => {{
    const sc = o.score || 0;
    const col = scoreColor(sc);
    const imgHtml = o.img
      ? `<img src="${{o.img}}" alt="" loading="lazy">`
      : `<div class="noimg">Kein Bild</div>`;
    const ribbon = o.sold
      ? `<div class="ribbon${{o.offline?' off':''}}">Sold</div>`
      : '';

    const card = document.createElement('div');
    card.className = 'card' + (o.sold ? ' sold' : '');
    card.onclick = () => openModal(o);
    card.innerHTML = `
      <div class="ci">
        ${{imgHtml}}
        ${{ribbon}}
        <div class="b-rang">#${{o.rang||'?'}}</div>
        <div class="b-score" style="background:${{col}}">${{sc.toFixed(1)}}</div>
        ${{o.makler ? `<div class="b-makler">${{o.makler}}</div>` : ''}}
      </div>
      <div class="cb">
        <div class="card-title">${{o.anzeige||o.name}}</div>
        <div class="card-ort">📍 ${{o.ort||'–'}}</div>
        <div class="stats">
          ${{o.zimmer ? `<div class="stat">🛏 <span>${{o.zimmer}}</span></div>` : ''}}
          ${{o.baeder ? `<div class="stat">🚿 <span>${{o.baeder}}</span></div>` : ''}}
          ${{o.grund ? `<div class="stat">🌿 <span>${{fmtNum(o.grund)}} m²</span></div>` : ''}}
          ${{o.flaeche ? `<div class="stat">🏠 <span>${{fmtNum(o.flaeche)}} m²</span></div>` : ''}}
          ${{o.gaestehaus ? `<div class="stat">🏡 <span>${{o.gaestehaus}} GH</span></div>` : ''}}
          ${{o.vermiet == 100 ? `<div class="stat">✅ ETV</div>` : o.vermiet == 50 ? `<div class="stat">⚠️ Teilw.</div>` : ''}}
          ${{o.charme ? `<div class="stat">✨ <span>${{'⭐'.repeat(o.charme)}}</span></div>` : ''}}
          ${{o.reno_score != null ? `<div class="stat">🔨 <span>${{o.reno_score}}%</span></div>` : ''}}
        </div>
        <div class="preis-row">
          <div class="preis">${{fmtPreis(o.preis)}}</div>
          <div class="epm2">${{o.epm2_b ? '€'+Math.round(o.epm2_b).toLocaleString('de')+'/m²' : ''}}</div>
        </div>
        <div class="score-bar-wrap">
          <div class="score-bar" style="width:${{sc}}%;background:${{col}}"></div>
        </div>
      </div>`;
    grid.appendChild(card);
  }});
}}

function applyFilters() {{ renderGrid(getFiltered()); }}
function resetFilters() {{
  document.getElementById('search').value = '';
  document.getElementById('filt-ort').value = '';
  document.getElementById('filt-zimmer').value = '0';
  document.getElementById('filt-lizenz').value = '';
  document.getElementById('filt-sold').value = '';
  document.getElementById('max-preis').value = 30;
  document.getElementById('pv').textContent = '30M';
  renderGrid(DATA);
}}

document.getElementById('search').addEventListener('input', () => renderGrid(getFiltered()));

document.getElementById('max-preis').addEventListener('input', () => renderGrid(getFiltered()));

function openModal(o) {{
  const sc = o.score || 0;
  const col = scoreColor(sc);
  // img
  const mImg = document.getElementById('m-img');
  mImg.innerHTML = `<button class="mh-close" onclick="closeModalDirect()">✕</button>`;
  if (o.img) {{
    const img = document.createElement('img');
    img.src = o.img;
    img.style.cssText = 'width:100%;height:100%;object-fit:cover';
    mImg.insertBefore(img, mImg.firstChild);
  }}
  document.getElementById('m-title').textContent = o.anzeige || o.name;
  document.getElementById('m-ort').textContent = '📍 ' + (o.ort || '–') + (o.makler ? '  ·  ' + o.makler : '');

  // KV grid
  const kvData = [
    ['Score', `<span style="color:${{col}};font-weight:700">${{sc.toFixed(1)}}</span>`],
    ['Rang', `#${{o.rang||'?'}}`],
    ['Preis', fmtPreis(o.preis)],
    ['€/m² bebaut', o.epm2_b ? '€'+Math.round(o.epm2_b).toLocaleString('de') : '–'],
    ['Zimmer', o.zimmer || '–'],
    ['Bäder', o.baeder || '–'],
    ['Grundstück', o.grund ? fmtNum(o.grund)+' m²' : '–'],
    ['Wohnfläche', o.flaeche ? fmtNum(o.flaeche)+' m²' : '–'],
    ['Garten', o.garten ? fmtNum(o.garten)+' m²' : '–'],
    ['Gästehäuser', o.gaestehaus ?? '–'],
    ['Vermietlizenz', o.vermiet == 100 ? '✅ Ja' : o.vermiet == 50 ? '⚠️ Teilw.' : '❌ Nein'],
    ['Charme', o.charme ? '⭐'.repeat(o.charme) : '–'],
    ['Renovierung', o.reno_score != null ? o.reno_score + '%' : '–'],
    ['Baujahr', o.baujahr || '–'],
  ];
  document.getElementById('m-grid').innerHTML = kvData.map(([k,v]) =>
    `<div class="kv"><div class="k">${{k}}</div><div class="v">${{v}}</div></div>`).join('');

  // Reach
  const reachData = [
    ['✈️ Flughafen', o.f_min, 15, 25, 40],
    ['🏨 Daia', o.d_min, 20, 40, 70],
    ['🌊 Ses Salines', o.s_min, 15, 30, 45],
    ['⛵ Andratx', o.a_min, 25, 40, 60],
  ];
  document.getElementById('m-reach').innerHTML = reachData.map(([label, min, ideal, ok, max]) => {{
    const cls = reachClass(min, ideal, ok, max);
    const val = min ? min + ' min' : '–';
    return `<div class="reach-item"><div class="label">${{label}}</div><div class="val ${{cls}}">${{val}}</div></div>`;
  }}).join('');

  // Kommentar
  const komEl = document.getElementById('m-komm');
  if (o.kommentar && o.kommentar !== 'None') {{
    komEl.textContent = '💬 ' + o.kommentar;
    komEl.style.display = 'block';
  }} else {{
    komEl.style.display = 'none';
  }}

  // Beschreibung
  document.getElementById('m-desc').textContent = (o.beschreib && o.beschreib !== 'None') ? o.beschreib : '';

  // Buttons
  const btns = document.getElementById('m-btns');
  btns.innerHTML = '';
  if (o.url && o.url !== 'None') {{
    const a = document.createElement('a');
    a.href = o.url; a.target = '_blank'; a.className = 'primary';
    a.textContent = '🔗 Zum Objekt';
    btns.appendChild(a);
  }}

  document.getElementById('overlay').classList.add('open');
  document.body.style.overflow = 'hidden';
}}

function closeModal(e) {{
  if (e.target === document.getElementById('overlay')) closeModalDirect();
}}
function closeModalDirect() {{
  document.getElementById('overlay').classList.remove('open');
  document.body.style.overflow = '';
}}
document.addEventListener('keydown', e => {{ if (e.key === 'Escape') closeModalDirect(); }});

// Initial render
renderGrid(DATA);
</script>
</body>
</html>'''

with open(OUT, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"✅ Generiert: {OUT}")
print(f"   {len(objects)} Objekte")
print(f"   Top 5: {[o['rang'] for o in objects[:5]]}")
