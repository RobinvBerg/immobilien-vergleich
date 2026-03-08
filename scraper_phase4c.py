#!/usr/bin/env python3
"""Phase 4c - Rightmove Apify + Fotocasa/Properstar Playwright"""

import sys, json, time, re, requests
from datetime import date
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
APIFY_TOKEN = 'apify_api_feD2KhARHjtuV9CrSwOReYgoePFSF44nsDL6'

def save_to_excel(new_objects, source_name):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
    
    new_count = 0
    for obj in new_objects:
        url = obj.get('url', '—')
        if url != '—' and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', ''),
            obj.get('quelle', source_name),
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            str(date.today()),
            'Neu'
        ])
        if url != '—':
            existing_urls.add(url)
        new_count += 1
    
    wb.save(EXCEL_PATH)
    print(f"  ✅ {new_count} neue Objekte aus {source_name} gespeichert")
    return new_count

def parse_price(text):
    if not text: return None
    nums = re.findall(r'[\d\.]+', text.replace(',', '.'))
    for n in nums:
        try:
            val = float(n.replace('.', ''))
            if val > 1000: return int(val)
        except: pass
    return None

def parse_int(text):
    if not text: return None
    nums = re.findall(r'\d+', str(text))
    return int(nums[0]) if nums else None

# ============================================================
# Rightmove via Apify
# ============================================================
def scrape_rightmove_apify():
    print("\n🏠 SOURCE 1: Rightmove (Apify)")
    
    # Use epctex/rightmove-scraper
    run_input = {
        "startUrls": [{"url": "https://www.rightmove.co.uk/overseas-property/in-Mallorca.html"}],
        "maxItems": 200,
        "extendOutputFunction": "() => ({})",
    }
    
    try:
        resp = requests.post(
            f"https://api.apify.com/v2/acts/epctex~rightmove-scraper/runs?token={APIFY_TOKEN}",
            json=run_input, timeout=30
        )
        print(f"  Run start: {resp.status_code}")
        if resp.status_code not in (200, 201):
            print(f"  Error: {resp.text[:300]}")
            return 0
        
        run_id = resp.json().get('data', {}).get('id')
        print(f"  Run ID: {run_id}")
        
        deadline = time.time() + 180  # 3 min
        while time.time() < deadline:
            time.sleep(8)
            sr = requests.get(f"https://api.apify.com/v2/actor-runs/{run_id}?token={APIFY_TOKEN}", timeout=10)
            status = sr.json().get('data', {}).get('status', '')
            print(f"  Status: {status}")
            if status == 'SUCCEEDED':
                ds_id = sr.json()['data']['defaultDatasetId']
                items = requests.get(f"https://api.apify.com/v2/datasets/{ds_id}/items?token={APIFY_TOKEN}&limit=500", timeout=30).json()
                print(f"  Items: {len(items)}")
                if items:
                    print(f"  Sample keys: {list(items[0].keys()) if items else 'none'}")
                
                objects = []
                for item in items:
                    objects.append({
                        'titel': item.get('title', item.get('address', item.get('displayAddress', ''))),
                        'quelle': 'Rightmove',
                        'url': item.get('url', item.get('propertyUrl', '—')),
                        'preis': item.get('price', item.get('priceValue')),
                        'zimmer': item.get('bedrooms', item.get('beds')),
                        'grundstueck': item.get('plotArea', item.get('landArea')),
                        'wohnflaeche': item.get('floorArea', item.get('area', item.get('livingArea'))),
                        'ort': item.get('location', item.get('city', item.get('displayAddress', 'Mallorca'))),
                    })
                return save_to_excel(objects, 'Rightmove')
            elif status in ('FAILED', 'ABORTED', 'TIMED-OUT'):
                print(f"  Run {status}")
                return 0
    except Exception as e:
        print(f"  Error: {e}")
        import traceback; traceback.print_exc()
    
    return 0


# ============================================================
# Fotocasa with Playwright (find the real API endpoint)
# ============================================================
def scrape_fotocasa_playwright():
    print("\n🏠 SOURCE 2: Fotocasa (Playwright)")
    objects = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=['--no-sandbox'])
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            viewport={'width': 1280, 'height': 800},
            locale='es-ES',
        )
        page = context.new_page()
        Stealth().apply_stealth_sync(page)
        
        # Capture API calls
        api_responses = []
        
        def handle_response(response):
            url = response.url
            if 'fotocasa' in url and ('search' in url or 'property' in url or 'propert' in url):
                try:
                    if response.status == 200:
                        body = response.body()
                        if body and len(body) > 100:
                            api_responses.append({'url': url, 'body': body})
                            print(f"  API call captured: {url[:100]} ({len(body)} bytes)")
                except:
                    pass
        
        page.on('response', handle_response)
        
        try:
            url = 'https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l'
            print(f"  Loading: {url}")
            page.goto(url, wait_until='domcontentloaded', timeout=30000)
            time.sleep(5)
            
            # Accept cookies
            try:
                page.click('[id*="didomi-notice-agree-button"], button:has-text("Aceptar"), button:has-text("Acepto")', timeout=4000)
                time.sleep(2)
            except:
                pass
            
            # Wait for listings to load
            try:
                page.wait_for_selector('[class*="re-Card"], [class*="PropertyCard"], article[class*="re-"]', timeout=10000)
            except:
                pass
            
            time.sleep(3)
            
            print(f"  API responses captured: {len(api_responses)}")
            
            # Process captured API data
            for api_resp in api_responses:
                try:
                    data = json.loads(api_resp['body'])
                    # Look for listings
                    listings = None
                    if isinstance(data, dict):
                        for key in ['realEstates', 'results', 'items', 'listings', 'data']:
                            if key in data and isinstance(data[key], list):
                                listings = data[key]
                                print(f"  Found listings under '{key}': {len(listings)}")
                                break
                    if listings:
                        for item in listings:
                            price = None
                            if 'transactions' in item and item['transactions']:
                                price_vals = item['transactions'][0].get('value', [])
                                price = price_vals[0] if price_vals else None
                            elif 'price' in item:
                                price = item['price']
                            
                            features = {}
                            for f in item.get('features', []):
                                features[f.get('key', '')] = f.get('value')
                            
                            url_val = item.get('detail', {}).get('es', item.get('url', ''))
                            if url_val and not url_val.startswith('http'):
                                url_val = 'https://www.fotocasa.es' + url_val
                            
                            objects.append({
                                'titel': item.get('address', {}).get('ubication', item.get('title', '')),
                                'quelle': 'Fotocasa',
                                'url': url_val or '—',
                                'preis': price,
                                'zimmer': features.get('roomsNumber', features.get('rooms')),
                                'grundstueck': features.get('plotArea'),
                                'wohnflaeche': features.get('constructedArea', features.get('surface')),
                                'ort': item.get('address', {}).get('municipality', ''),
                            })
                except Exception as e:
                    print(f"  Parse error: {e}")
            
            # Fallback: scrape HTML
            if not objects:
                print("  Fallback: HTML scraping")
                cards = page.query_selector_all('[class*="re-Card"], article[class*="re-"], [class*="PropertyCard"]')
                print(f"  Cards found: {len(cards)}")
                
                seen = set()
                for card in cards[:100]:
                    try:
                        link = card.query_selector('a')
                        href = link.get_attribute('href') if link else None
                        if not href or href in seen:
                            continue
                        seen.add(href)
                        if not href.startswith('http'):
                            href = 'https://www.fotocasa.es' + href
                        
                        title_el = card.query_selector('[class*="Title"], [class*="title"], h2, h3')
                        price_el = card.query_selector('[class*="Price"], [class*="price"]')
                        beds_el = card.query_selector('[class*="Bedroom"], [class*="bedroom"], [class*="room"]')
                        area_el = card.query_selector('[class*="Surface"], [class*="area"], [class*="sqm"]')
                        loc_el = card.query_selector('[class*="Location"], [class*="location"], [class*="municipality"]')
                        
                        objects.append({
                            'titel': title_el.inner_text()[:100] if title_el else '',
                            'quelle': 'Fotocasa',
                            'url': href,
                            'preis': parse_price(price_el.inner_text() if price_el else ''),
                            'zimmer': parse_int(beds_el.inner_text() if beds_el else ''),
                            'grundstueck': None,
                            'wohnflaeche': parse_int(area_el.inner_text() if area_el else ''),
                            'ort': loc_el.inner_text()[:50] if loc_el else '',
                        })
                    except:
                        pass
                
                # Try to paginate
                if objects:
                    for pg in range(2, 10):
                        try:
                            next_url = f'https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l?page={pg}'
                            page.goto(next_url, wait_until='domcontentloaded', timeout=20000)
                            time.sleep(3)
                            before = len(objects)
                            new_cards = page.query_selector_all('[class*="re-Card"], article[class*="re-"], [class*="PropertyCard"]')
                            for card in new_cards[:100]:
                                try:
                                    link = card.query_selector('a')
                                    href = link.get_attribute('href') if link else None
                                    if not href or href in seen:
                                        continue
                                    seen.add(href)
                                    if not href.startswith('http'):
                                        href = 'https://www.fotocasa.es' + href
                                    title_el = card.query_selector('[class*="Title"], h2, h3')
                                    price_el = card.query_selector('[class*="Price"]')
                                    objects.append({
                                        'titel': title_el.inner_text()[:100] if title_el else '',
                                        'quelle': 'Fotocasa',
                                        'url': href,
                                        'preis': parse_price(price_el.inner_text() if price_el else ''),
                                        'zimmer': None, 'grundstueck': None, 'wohnflaeche': None, 'ort': 'Mallorca',
                                    })
                                except:
                                    pass
                            print(f"  Page {pg}: +{len(objects)-before}")
                            if len(objects) == before:
                                break
                        except Exception as e:
                            print(f"  Page {pg} error: {e}")
                            break
        
        except Exception as e:
            print(f"  Error: {e}")
            import traceback; traceback.print_exc()
        
        browser.close()
    
    print(f"  Gesammelt: {len(objects)} Objekte")
    if objects:
        return save_to_excel(objects, 'Fotocasa')
    return 0


# ============================================================
# Properstar with Playwright
# ============================================================
def scrape_properstar_playwright():
    print("\n🏠 SOURCE 3: Properstar (Playwright)")
    objects = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=['--no-sandbox'])
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            viewport={'width': 1280, 'height': 800},
        )
        page = context.new_page()
        Stealth().apply_stealth_sync(page)
        
        # Intercept API calls
        api_responses = []
        
        def handle_response(response):
            url = response.url
            if 'properstar' in url and any(kw in url for kw in ['search', 'listing', 'property', 'api']):
                try:
                    if response.status == 200:
                        body = response.body()
                        if body and len(body) > 200:
                            api_responses.append({'url': url, 'body': body})
                            print(f"  API: {url[:100]} ({len(body)}b)")
                except:
                    pass
        
        page.on('response', handle_response)
        
        try:
            url = 'https://www.properstar.com/es/venta?country=ES&region=Baleares&currency=EUR'
            print(f"  Loading: {url}")
            page.goto(url, wait_until='domcontentloaded', timeout=30000)
            time.sleep(5)
            
            # Cookie consent
            try:
                page.click('button:has-text("Accept"), button:has-text("Aceptar"), [id*="accept"]', timeout=4000)
                time.sleep(2)
            except:
                pass
            
            time.sleep(3)
            print(f"  API calls captured: {len(api_responses)}")
            
            # Process API data
            for api_resp in api_responses:
                try:
                    data = json.loads(api_resp['body'])
                    listings = None
                    if isinstance(data, dict):
                        for key in ['listings', 'results', 'items', 'data', 'properties']:
                            if key in data and isinstance(data[key], list) and data[key]:
                                listings = data[key]
                                print(f"  Found under '{key}': {len(listings)}")
                                break
                    elif isinstance(data, list) and data:
                        listings = data
                    
                    if listings:
                        print(f"  Sample: {json.dumps(listings[0], ensure_ascii=False)[:300]}")
                        for item in listings:
                            url_val = item.get('url', item.get('link', item.get('detailUrl', '—')))
                            if url_val and not url_val.startswith('http') and url_val != '—':
                                url_val = 'https://www.properstar.com' + url_val
                            
                            objects.append({
                                'titel': item.get('title', item.get('address', item.get('name', ''))),
                                'quelle': 'Properstar',
                                'url': url_val or '—',
                                'preis': item.get('price', item.get('priceValue', item.get('salePrice'))),
                                'zimmer': item.get('bedrooms', item.get('rooms', item.get('numberOfRooms'))),
                                'grundstueck': item.get('landArea', item.get('plotSize')),
                                'wohnflaeche': item.get('livingArea', item.get('area', item.get('surface'))),
                                'ort': item.get('city', item.get('location', item.get('municipality', ''))),
                            })
                except Exception as e:
                    print(f"  Parse error: {e}")
            
            # Fallback: HTML
            if not objects:
                print("  Fallback: HTML")
                cards = page.query_selector_all('[class*="PropertyCard"], [class*="property-card"], [class*="listing-card"], article')
                links = page.query_selector_all('a[href*="/property/"], a[href*="/listing/"], a[href*="/es/venta/"]')
                print(f"  Cards: {len(cards)}, Links: {len(links)}")
                
                seen = set()
                for card in cards[:50]:
                    try:
                        link = card.query_selector('a')
                        href = link.get_attribute('href') if link else None
                        if not href or href in seen: continue
                        seen.add(href)
                        if not href.startswith('http'): href = 'https://www.properstar.com' + href
                        title_el = card.query_selector('h2, h3, [class*="title"]')
                        price_el = card.query_selector('[class*="price"]')
                        objects.append({
                            'titel': title_el.inner_text()[:100] if title_el else '',
                            'quelle': 'Properstar',
                            'url': href,
                            'preis': parse_price(price_el.inner_text() if price_el else ''),
                            'zimmer': None, 'grundstueck': None, 'wohnflaeche': None, 'ort': 'Mallorca',
                        })
                    except: pass
        
        except Exception as e:
            print(f"  Error: {e}")
            import traceback; traceback.print_exc()
        
        browser.close()
    
    print(f"  Gesammelt: {len(objects)} Objekte")
    if objects:
        return save_to_excel(objects, 'Properstar')
    return 0


if __name__ == '__main__':
    results = {}
    results['Rightmove (Apify)'] = scrape_rightmove_apify()
    results['Fotocasa'] = scrape_fotocasa_playwright()
    results['Properstar'] = scrape_properstar_playwright()
    
    print("\n=== ERGEBNIS ===")
    for src, count in results.items():
        status = "✅" if count > 0 else "❌"
        print(f"  {status} {src}: {count} Objekte")
