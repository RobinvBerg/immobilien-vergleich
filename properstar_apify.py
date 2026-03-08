import requests, json, time
from openpyxl import load_workbook
from datetime import date

APIFY_TOKEN = "apify_api_feD2KhARHjtuV9CrSwOReYgoePFSF44nsDL6"

# Actor starten
run_input = {
    "startUrls": [
        {"url": "https://www.properstar.com/es/venta?country=ES&region=Baleares&type=house&bedrooms=5"},
        {"url": "https://www.properstar.com/es/venta?country=ES&region=Baleares&type=villa"},
        {"url": "https://www.properstar.com/es/venta?country=ES&region=Baleares&type=finca"},
    ],
    "maxItems": 1000,
}

r = requests.post(
    'https://api.apify.com/v2/acts/memo23~properstar-leads-scraper/runs',
    json=run_input,
    headers={'Authorization': f'Bearer {APIFY_TOKEN}', 'Content-Type': 'application/json'}
)
print(f"Start: {r.status_code} | {r.text[:500]}")

if r.status_code not in [200, 201]:
    print("Actor start failed, trying alternative input format")
    for input_variant in [
        {"search": "Mallorca", "country": "ES", "maxItems": 500},
        {"location": "Mallorca", "propertyType": "house", "maxItems": 500},
        {"url": "https://www.properstar.com/es/venta?country=ES&region=Baleares"},
    ]:
        r = requests.post(
            'https://api.apify.com/v2/acts/memo23~properstar-leads-scraper/runs',
            json=input_variant,
            headers={'Authorization': f'Bearer {APIFY_TOKEN}', 'Content-Type': 'application/json'}
        )
        print(f"  Variant: {r.status_code} | {r.text[:300]}")
        if r.status_code in [200, 201]:
            break

run_data = r.json().get('data', {})
run_id = run_data.get('id')
dataset_id = run_data.get('defaultDatasetId')
actor_id = 'memo23~properstar-leads-scraper'
print(f"Run ID: {run_id} | Dataset: {dataset_id}")

if not run_id:
    print("ERROR: No run ID obtained. Full response:")
    print(r.text)
    exit(1)

# Warte auf Completion (max 8 Minuten)
for i in range(48):
    time.sleep(10)
    sr = requests.get(
        f'https://api.apify.com/v2/acts/{actor_id}/runs/{run_id}',
        headers={'Authorization': f'Bearer {APIFY_TOKEN}'}
    )
    status = sr.json().get('data', {}).get('status', '')
    print(f"[{i*10}s] Status: {status}")
    if status in ['SUCCEEDED', 'FAILED', 'ABORTED', 'TIMED-OUT']:
        break

# Ergebnisse holen
results_r = requests.get(
    f'https://api.apify.com/v2/datasets/{dataset_id}/items?limit=2000',
    headers={'Authorization': f'Bearer {APIFY_TOKEN}'}
)
items = results_r.json()
print(f"\nErgebnisse: {len(items)} items")
if items:
    print(f"Sample: {json.dumps(items[0], indent=2)[:500]}")

# In Excel speichern
wb = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
ws = wb['Mallorca Objekte']
existing_urls = set(str(row[2]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[2] and row[2] != '—')

count = 0
for item in items:
    url = item.get('url', item.get('link', item.get('detailUrl', '—')))
    if not url or url == 'null': url = '—'
    if url != '—' and url in existing_urls: continue

    preis = item.get('price', item.get('Price', item.get('listingPrice')))
    if isinstance(preis, dict): preis = preis.get('amount', preis.get('value'))

    zimmer = item.get('bedrooms', item.get('rooms', item.get('Bedrooms')))
    flaeche = item.get('surface', item.get('area', item.get('livingArea', item.get('size'))))
    ort = item.get('city', item.get('municipality', item.get('location', item.get('area', ''))))
    if isinstance(ort, dict): ort = ort.get('name', '')
    titel_raw = item.get('title', item.get('name', item.get('description', f'Properstar {ort}')))
    titel = str(titel_raw)[:100]

    ws.append([titel, 'Properstar', url, preis, zimmer, None, flaeche,
               str(ort), str(date.today()), 'Neu'])
    if url != '—': existing_urls.add(url)
    count += 1

wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
print(f"✅ Properstar: {count} Objekte gespeichert")
