#!/usr/bin/env python3
"""
Balearic Properties Scraper — Mallorca
Scrapes all listings from balearic-properties.com
"""

import requests
from bs4 import BeautifulSoup
import time
import re
from openpyxl import load_workbook
from datetime import date

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
BASE_URL = 'https://www.balearic-properties.com/property-for-sale/mallorca.html'
TOTAL_PAGES = 88

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Referer': 'https://www.balearic-properties.com/',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
}

# ── Step 1: Probe page 1 and analyze HTML structure ──────────────────────────

def probe_structure():
    url = BASE_URL
    print(f"Probing: {url}")
    r = requests.get(url, headers=headers, timeout=30)
    print(f"Status: {r.status_code}  |  Content-length: {len(r.text)}")

    if r.status_code != 200:
        print("BLOCKED — switching to Playwright")
        return None

    soup = BeautifulSoup(r.text, 'html.parser')

    print("\n=== First 3000 chars of HTML ===")
    print(r.text[:3000])

    print("\n=== Selector probe ===")
    candidates = [
        '.property', '.listing', 'article', '.card',
        '[class*="prop"]', '[class*="listing"]', '[class*="result"]',
        '.item', '.estate', '.realty',
    ]
    for sel in candidates:
        items = soup.select(sel)
        if items:
            print(f"  '{sel}': {len(items)} items")
            print("    First:", str(items[0])[:400])
            print()

    return r.text

# ── Step 2: Parse listings from a soup object ─────────────────────────────────

def parse_price(text):
    """Extract numeric price from string like '€ 1,250,000'"""
    if not text:
        return None
    digits = re.sub(r'[^\d]', '', text)
    return int(digits) if digits else None

def parse_m2(text):
    """Extract m² value"""
    if not text:
        return None
    m = re.search(r'([\d,\.]+)\s*m', text)
    if m:
        return float(m.group(1).replace(',', '').replace('.', ''))
    return None

def parse_listings(soup):
    listings = []

    # Try multiple container patterns
    containers = (
        soup.select('.property-item') or
        soup.select('.listing-item') or
        soup.select('article.property') or
        soup.select('.col-xs-12.col-sm-6.col-md-4') or  # bootstrap grid cards
        soup.select('[class*="property-card"]') or
        soup.select('[class*="PropertyCard"]') or
        soup.select('.result') or
        []
    )

    # If nothing specific, try generic article or li
    if not containers:
        containers = soup.select('article') or soup.select('li.property')

    print(f"  Found {len(containers)} containers")

    for c in containers:
        obj = {}

        # Title
        title_el = c.select_one('h2, h3, h4, .title, [class*="title"], [class*="name"]')
        obj['titel'] = title_el.get_text(strip=True) if title_el else ''

        # URL
        link_el = c.select_one('a[href]')
        if link_el:
            href = link_el['href']
            if href.startswith('http'):
                obj['url'] = href
            else:
                obj['url'] = 'https://www.balearic-properties.com' + href
        else:
            obj['url'] = '—'

        # Price
        price_el = c.select_one('[class*="price"], .price, [class*="Price"]')
        if not price_el:
            # look for € sign in text
            for el in c.find_all(text=re.compile(r'€|\d{3,}')):
                if '€' in el or re.search(r'\d{6,}', el):
                    price_el = el.parent
                    break
        obj['preis'] = parse_price(price_el.get_text() if price_el else '')

        # Bedrooms
        bed_el = c.select_one('[class*="bed"], [class*="Bed"], .beds, .bedrooms, [class*="room"]')
        if bed_el:
            m = re.search(r'\d+', bed_el.get_text())
            obj['zimmer'] = int(m.group()) if m else None
        else:
            # Try icon+number pattern (common: <i class="icon-bed">3)
            for el in c.find_all(text=re.compile(r'^\d+$')):
                parent = el.parent
                if parent and any(k in (parent.get('class') or []) or
                                  k in str(parent.get('class','')) 
                                  for k in ['bed','room','sleep']):
                    obj['zimmer'] = int(el.strip())
                    break
            else:
                obj['zimmer'] = None

        # Area / Size
        size_el = c.select_one('[class*="size"], [class*="area"], [class*="m2"], [class*="sqm"]')
        if size_el:
            obj['wohnflaeche'] = parse_m2(size_el.get_text())
        else:
            obj['wohnflaeche'] = None

        # Plot size
        plot_el = c.select_one('[class*="plot"], [class*="land"], [class*="grundstueck"]')
        obj['grundstueck'] = parse_m2(plot_el.get_text()) if plot_el else None

        # Location
        loc_el = c.select_one('[class*="location"], [class*="city"], [class*="area"], [class*="region"], .location, .place')
        obj['ort'] = loc_el.get_text(strip=True) if loc_el else ''

        listings.append(obj)

    return listings

# ── Step 3: Save batch to Excel ───────────────────────────────────────────────

def save_batch(batch, existing_urls):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    saved = 0
    for obj in batch:
        url = obj.get('url', '—')
        if url != '—' and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', ''),
            'Balearic Properties',
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            str(date.today()),
            'Neu'
        ])
        if url != '—':
            existing_urls.add(url)
        saved += 1
    wb.save(EXCEL_PATH)
    return saved

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # Load existing URLs to avoid duplicates
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set(
        str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True)
        if r[2] and str(r[2]).strip() not in ('—', 'None', '')
    )
    print(f"Existing URLs in Excel: {len(existing_urls)}")

    session = requests.Session()
    session.headers.update(headers)

    all_listings = []
    total_saved = 0
    batch = []
    blocked_count = 0

    for page_num in range(1, TOTAL_PAGES + 1):
        url = BASE_URL if page_num == 1 else f'{BASE_URL}?page={page_num}'
        print(f"\n[Page {page_num}/{TOTAL_PAGES}] {url}")

        try:
            r = session.get(url, timeout=30)
            print(f"  HTTP {r.status_code}, {len(r.text)} chars")

            if r.status_code != 200:
                print(f"  !! Non-200 status. Stopping.")
                blocked_count += 1
                if blocked_count >= 3:
                    print("3 consecutive failures — aborting requests mode.")
                    break
                continue
            else:
                blocked_count = 0

            soup = BeautifulSoup(r.text, 'html.parser')

            if page_num == 1:
                # Save raw HTML for analysis
                with open('/Users/robin/.openclaw/workspace/mallorca-projekt/balearic_page1.html', 'w', encoding='utf-8') as f:
                    f.write(r.text)
                print("  Saved page 1 HTML to balearic_page1.html")

            listings = parse_listings(soup)
            print(f"  Parsed {len(listings)} listings")

            if page_num == 1 and len(listings) == 0:
                print("  WARN: No listings on page 1 — check HTML structure!")
                # Print a snippet for debugging
                print("  Body snippet:", soup.body.get_text()[:1000] if soup.body else "NO BODY")

            batch.extend(listings)

            # Save every 10 pages
            if page_num % 10 == 0 or page_num == TOTAL_PAGES:
                saved = save_batch(batch, existing_urls)
                total_saved += saved
                print(f"  >> Batch saved: {saved} new | Total so far: {total_saved}")
                batch = []

        except Exception as e:
            print(f"  ERROR: {e}")

        time.sleep(0.8)

    # Save any remaining
    if batch:
        saved = save_batch(batch, existing_urls)
        total_saved += saved
        print(f"\nFinal batch saved: {saved}")

    print(f"\n{'='*50}")
    print(f"DONE. Total new listings saved: {total_saved}")

if __name__ == '__main__':
    # Quick structure probe first
    html = probe_structure()
    if html:
        print("\n\nStructure looks OK. Starting full scrape...\n")
        main()
    else:
        print("\nrequests blocked — need Playwright fallback")
