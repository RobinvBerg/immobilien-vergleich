#!/usr/bin/env python3
"""
Verify images for all non-LivingBlue entries via Playwright hash comparison.
Downloads live main image, compares hash with stored bilder/X_main.jpg.
Saves corrected images automatically.
"""
import asyncio, json, hashlib, shutil, sys
from pathlib import Path
from playwright.async_api import async_playwright
import openpyxl

BASE = Path(__file__).parent
RESULTS_FILE = BASE / 'verify_nonlb_results.json'

# Resume support
if RESULTS_FILE.exists():
    done = {r['nr']: r for r in json.load(open(RESULTS_FILE))}
else:
    done = {}

async def get_live_image_bytes(page, url):
    await page.goto(url, timeout=25000, wait_until='domcontentloaded')
    await asyncio.sleep(2)

    img_url = None
    selectors = [
        'img[src*="photo"]','img[src*="image"]','img[src*="foto"]',
        'img[src*="cdn"]','img[src*="media"]','img[src*="property"]',
        '.gallery img','.slider img','picture img',
        '[class*="gallery"] img','[class*="photo"] img',
        '[class*="main"] img','[class*="hero"] img',
    ]
    for sel in selectors:
        try:
            els = await page.query_selector_all(sel)
            for el in els[:5]:
                for attr in ['src','data-src','data-lazy','data-original']:
                    src = await el.get_attribute(attr) or ''
                    src = src.split(',')[0].split(' ')[0]
                    if src and any(x in src.lower() for x in ['jpg','jpeg','webp','png']) and len(src) > 20 and 'logo' not in src.lower() and 'icon' not in src.lower():
                        img_url = src
                        break
                if img_url: break
        except: pass
        if img_url: break

    if not img_url:
        imgs = await page.query_selector_all('img')
        best = (0, None)
        for img in imgs[:40]:
            try:
                w = await page.evaluate('(el) => el.naturalWidth', img)
                src = await img.get_attribute('src') or ''
                if w > best[0] and w > 300 and src and 'logo' not in src.lower():
                    best = (w, src)
            except: pass
        img_url = best[1]

    if not img_url: return None, None
    if img_url.startswith('//'): img_url = 'https:' + img_url
    elif img_url.startswith('/'):
        from urllib.parse import urlparse
        parsed = urlparse(url)
        img_url = f"{parsed.scheme}://{parsed.netloc}{img_url}"

    try:
        resp = await page.request.get(img_url, timeout=15000)
        if resp.status == 200:
            return await resp.body(), img_url
    except: pass
    return None, img_url

async def main():
    wb = openpyxl.load_workbook(BASE / 'mallorca-kandidaten-v2.xlsx')
    ws = wb.active
    headers = [c.value for c in ws[1]]

    entries = []
    for row in ws.iter_rows(min_row=2):
        nr = row[headers.index('Ordnungsnummer')].value
        if not nr: continue
        makler = row[headers.index('Makler')].value or ''
        if 'Living Blue' in makler: continue
        komm = str(row[headers.index('Kommentar')].value or '')
        url = row[headers.index('Link Objekt (URL)')].value or ''
        name = row[headers.index('Name')].value or ''
        entries.append({
            'nr': int(nr), 'url': url, 'name': name,
            'makler': makler, 'delisted': '⚠️ delisted' in komm
        })

    todo = [e for e in entries if e['nr'] not in done]
    results = list(done.values())
    print(f"Non-LB total: {len(entries)} | Done: {len(done)} | Todo: {len(todo)}", flush=True)

    stats = {'match': 0, 'fixed': 0, 'no_live': 0, 'no_url': 0, 'delisted': 0, 'err': 0}

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
            viewport={'width': 1280, 'height': 900}
        )

        for i, entry in enumerate(todo):
            nr = entry['nr']
            url = entry['url']
            stored = BASE / f'bilder/{nr}_main.jpg'

            sys.stdout.write(f"[{i+1}/{len(todo)}] Nr.{nr:3d} {entry['makler'][:12]:<12} ")
            sys.stdout.flush()

            if entry['delisted']:
                r = {'nr': nr, 'status': 'DELISTED'}
                stats['delisted'] += 1
                print("⚠️  delisted"); done[nr] = r; results.append(r); continue

            if not url or url == '—':
                r = {'nr': nr, 'status': 'NO_URL', 'name': entry['name']}
                stats['no_url'] += 1
                print("⬜ kein URL"); done[nr] = r; results.append(r); continue

            page = await ctx.new_page()
            try:
                live_bytes, live_url = await get_live_image_bytes(page, url)

                if not live_bytes:
                    r = {'nr': nr, 'status': 'NO_LIVE', 'makler': entry['makler']}
                    stats['no_live'] += 1
                    print("❓ kein Live-Bild")
                    done[nr] = r; results.append(r)
                    continue

                live_hash = hashlib.md5(live_bytes).hexdigest()

                if stored.exists():
                    stored_hash = hashlib.md5(stored.read_bytes()).hexdigest()
                else:
                    stored_hash = None

                if stored_hash == live_hash:
                    r = {'nr': nr, 'status': 'MATCH'}
                    stats['match'] += 1
                    print("✅ match")
                else:
                    stored.write_bytes(live_bytes)
                    action = 'fixed' if stored_hash else 'new'
                    r = {'nr': nr, 'status': 'FIXED', 'action': action}
                    stats['fixed'] += 1
                    print(f"🔄 {action} (live={live_hash[:8]})")

                done[nr] = r; results.append(r)

            except Exception as e:
                r = {'nr': nr, 'status': 'ERR', 'error': str(e)[:80]}
                stats['err'] += 1
                print(f"⚠️  {str(e)[:50]}")
                done[nr] = r; results.append(r)
            finally:
                await page.close()

            if (i+1) % 20 == 0:
                json.dump(results, open(RESULTS_FILE,'w'), indent=2)
                print(f"  → Zwischenstand: {stats}", flush=True)

        await browser.close()

    json.dump(results, open(RESULTS_FILE,'w'), indent=2)
    print(f"\n{'='*60}")
    print(f"FERTIG: {stats}")
    no_live = [r['nr'] for r in results if r['status']=='NO_LIVE']
    print(f"Nicht erreichbar ({len(no_live)}): {no_live}")

asyncio.run(main())
