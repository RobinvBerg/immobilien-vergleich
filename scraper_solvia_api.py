#!/usr/bin/env python3
"""Solvia / Haya API Scraper - alle Baleares Objekte"""

import requests
import json
import time
from datetime import date
from openpyxl import load_workbook

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Accept': 'application/json',
    'Content-Type': 'application/json',
    'Origin': 'https://www.haya.es',
    'Referer': 'https://www.haya.es/comprar/viviendas/baleares/',
}

API_URL = 'https://www.solvia.es/api/inmuebles/v2/buscarInmuebles'

def fetch_all():
    all_objects = []
    seen_ids = set()
    
    # Alle Kategorien
    for categoria in ['1', '2', '4', '5']:  # 1=Viviendas, 2=Garajes, 4=Suelos, 5=Comercial
        cat_name = {'1': 'Viviendas', '2': 'Garajes', '4': 'Suelos', '5': 'Comercial'}.get(categoria, categoria)
        
        for page in range(1, 10):
            payload = {
                'provincia': '7',
                'categoria': categoria,
                'pagina': page,
                'elementosPorPagina': 50,
                'ordenamiento': 'RELEVANCIA',
            }
            
            try:
                r = requests.post(API_URL, headers=HEADERS, json=payload, timeout=30)
                data = r.json()
                
                inmuebles = data.get('inmuebles', [])
                total = data.get('total', 0)
                
                print(f"  {cat_name} Seite {page}: {len(inmuebles)} Objekte (Total: {total})")
                
                if not inmuebles:
                    break
                
                for item in inmuebles:
                    item_id = str(item.get('id', ''))
                    if item_id in seen_ids:
                        continue
                    seen_ids.add(item_id)
                    
                    # Ort
                    prov = item.get('provincia', {})
                    pobl = item.get('poblacion', {})
                    ort = pobl.get('nombre', 'Baleares') if isinstance(pobl, dict) and pobl.get('nombre') else \
                          prov.get('nombre', 'Baleares') if isinstance(prov, dict) else 'Baleares'
                    
                    # Preis
                    preis = item.get('precio')
                    if isinstance(preis, dict):
                        preis = preis.get('valor')
                    
                    # Zimmer
                    zimmer = item.get('dormitorios')
                    
                    # Fläche
                    flaeche = item.get('totalM2') or item.get('m2Construidos')
                    
                    # Typ
                    tipo = item.get('tipoVivienda', {})
                    tipo_name = tipo.get('nombre', '') if isinstance(tipo, dict) else ''
                    
                    # URL
                    url_id = item.get('idVivienda', item_id)
                    url = 'https://www.haya.es/es/comprar/vivienda/' + str(url_id)
                    
                    # Titel
                    if tipo_name and ort:
                        titel = tipo_name + ' en ' + ort
                    elif ort:
                        titel = 'Inmueble en ' + ort
                    else:
                        titel = 'Solvia/Haya ' + str(item_id)
                    
                    all_objects.append({
                        'titel': titel[:80],
                        'quelle': 'Solvia / Haya Real Estate',
                        'url': url,
                        'preis': float(preis) if preis else None,
                        'zimmer': int(zimmer) if zimmer else None,
                        'grundstueck': None,
                        'wohnflaeche': float(flaeche) if flaeche else None,
                        'ort': str(ort)[:60],
                    })
                
                total_pages = (total + 49) // 50 if total else 1
                if page >= total_pages or page >= 9:
                    break
                
                time.sleep(0.5)
                
            except Exception as e:
                print(f"  Fehler Seite {page}: {e}")
                break
    
    return all_objects

def save_to_excel(objects):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    
    existing_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            existing_urls.add(str(row[2]).strip())
    
    new_count = 0
    for obj in objects:
        url = obj.get('url', '')
        if url and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', ''), obj.get('quelle', ''), url,
            obj.get('preis'), obj.get('zimmer'), obj.get('grundstueck'),
            obj.get('wohnflaeche'), obj.get('ort', ''), str(date.today()), 'Neu'
        ])
        if url:
            existing_urls.add(url)
        new_count += 1
    
    wb.save(EXCEL_PATH)
    return new_count

if __name__ == '__main__':
    print("=== Solvia/Haya API ===")
    objects = fetch_all()
    print(f"\nGefunden: {len(objects)}")
    
    saved = save_to_excel(objects)
    print(f"Gespeichert: {saved} neue Objekte")
    
    # Sample
    for obj in objects[:5]:
        print(f"  {obj['titel'][:50]} | {obj['preis']} € | {obj['zimmer']} Z | {obj['ort']}")
