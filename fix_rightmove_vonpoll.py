import requests
from bs4 import BeautifulSoup
import json, time, re
from openpyxl import load_workbook
from datetime import date

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.5',
    'Referer': 'https://www.rightmove.co.uk/',
}

wb = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
ws = wb['Mallorca Objekte']
existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
print(f"Start: {len(existing_urls)} URLs bekannt, {ws.max_row} Zeilen")

rm_count = 0
location_id = 'WORLD_REGION%5E219140'  # Mallorca

for index in range(0, 1050, 25):
    url = f'https://www.rightmove.co.uk/overseas-property/find.html?channel=OVERSEAS&locationIdentifier={location_id}&index={index}&sortType=6&propertyTypes=&includeSSTC=false'
    try:
        r = requests.get(url, headers=headers, timeout=30)
        if r.status_code != 200:
            print(f"index {index}: HTTP {r.status_code}")
            break
        
        soup = BeautifulSoup(r.text, 'html.parser')
        next_data = soup.find('script', {'id': '__NEXT_DATA__'})
        if not next_data:
            print(f"index {index}: no __NEXT_DATA__")
            break
        
        data = json.loads(next_data.string)
        
        # Properties aus pageProps extrahieren
        props = []
        page_props = data.get('props', {}).get('pageProps', {})
        
        # Suche nach searchResults oder properties
        for key in ['searchResults', 'properties', 'results', 'listings']:
            if key in page_props:
                val = page_props[key]
                if isinstance(val, list):
                    props = val
                elif isinstance(val, dict):
                    props = val.get('properties', val.get('results', []))
                break
        
        # Tiefer suchen wenn nicht gefunden
        if not props:
            def find_list(obj, depth=0):
                if depth > 8: return []
                if isinstance(obj, list) and len(obj) > 3:
                    if isinstance(obj[0], dict) and any(k in obj[0] for k in ['price', 'bedrooms', 'propertyUrl', 'id']):
                        return obj
                if isinstance(obj, dict):
                    for v in obj.values():
                        result = find_list(v, depth+1)
                        if result: return result
                return []
            props = find_list(page_props)
        
        if not props:
            print(f"index {index}: no properties found, stopping")
            break
        
        page_new = 0
        for prop in props:
            prop_url = prop.get('propertyUrl', prop.get('url', prop.get('detailUrl', '')))
            if prop_url and not prop_url.startswith('http'):
                prop_url = 'https://www.rightmove.co.uk' + prop_url
            if not prop_url: prop_url = '—'
            if prop_url != '—' and prop_url in existing_urls: continue
            
            preis = prop.get('price', {})
            if isinstance(preis, dict):
                preis_val = preis.get('amount', None)
                if preis_val is None:
                    disp = preis.get('displayPrices', [{}])
                    if disp:
                        preis_val = disp[0].get('displayPrice', '')
                preis = preis_val
            if isinstance(preis, str):
                preis = re.sub(r'[^\d]', '', preis)
                try: preis = int(preis)
                except: preis = None
            
            zimmer = prop.get('bedrooms', prop.get('numberOfBedrooms'))
            flaeche = prop.get('floorplanSqm', prop.get('displaySize'))
            ort = prop.get('location', {})
            if isinstance(ort, dict): ort = ort.get('displayAddress', ort.get('town', ''))
            titel = prop.get('summary', prop.get('displayAddress', prop.get('title', 'Rightmove Mallorca')))
            if isinstance(titel, dict): titel = str(titel)
            
            ws.append([str(titel)[:100], 'Rightmove', prop_url, preis, zimmer, None, flaeche,
                       str(ort)[:100], str(date.today()), 'Neu'])
            if prop_url != '—': existing_urls.add(prop_url)
            rm_count += 1
            page_new += 1
        
        print(f"index {index}: {page_new} neu, gesamt: {rm_count}")
        time.sleep(0.5)
        
    except Exception as e:
        print(f"index {index}: Error — {e}")
        time.sleep(2)

print(f"\nRightmove total: {rm_count}")

# SOFORT speichern nach Rightmove
wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
print(f"Gespeichert nach Rightmove. Gesamt Zeilen: {ws.max_row}")

# ============================================================
# Teil 2: Von Poll via Brave Search
# ============================================================

BRAVE_KEY = "BSA5KWR8-QpVt9-S1fSCZEjAY2bdC0M"

vp_count = 0
queries = [
    'site:von-poll.com mallorca villa finca',
    'site:von-poll.com mallorca immobilien kaufen',
    'site:von-poll.com balearen villa',
    'site:von-poll.com mallorca luxury',
    'site:von-poll.com mallorca finca pool',
    '"von-poll.com" mallorca expose',
    'von-poll.com/de/expose mallorca',
    '"von-poll" mallorca villa site:von-poll.com',
]

for query in queries:
    try:
        r = requests.get(
            'https://api.search.brave.com/res/v1/web/search',
            params={'q': query, 'count': 20},
            headers={'Accept': 'application/json', 'X-Subscription-Token': BRAVE_KEY},
            timeout=15
        )
        if r.status_code == 200:
            results = r.json().get('web', {}).get('results', [])
            page_new = 0
            for res in results:
                url = res.get('url', '')
                if 'von-poll.com' in url and url not in existing_urls:
                    titel = res.get('title', 'Von Poll Objekt')[:100]
                    ws.append([titel, 'Von Poll Real Estate', url, None, None, None, None,
                               'Mallorca', str(date.today()), 'Neu'])
                    existing_urls.add(url)
                    vp_count += 1
                    page_new += 1
            print(f"Query '{query[:40]}': {page_new} neu")
        else:
            print(f"Query '{query[:40]}': HTTP {r.status_code}")
    except Exception as e:
        print(f"Query error: {e}")
    time.sleep(0.5)

print(f"Von Poll: {vp_count} gespeichert")

# FINAL SAVE
wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
print(f"\n✅ FERTIG. Rightmove: {rm_count} | Von Poll: {vp_count}")
print(f"Finale Zeilenzahl: {ws.max_row}")
