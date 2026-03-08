#!/usr/bin/env python3
"""
fetch_fotocasa_worker.py  -- single worker for one URL slice
Usage: python3 fetch_fotocasa_worker.py <worker_id> <start> <end>
  worker_id : 0-4
  start/end : row indices (0-based) into fotocasa_rows list
"""

import re, time, json, sys
from pathlib import Path
import openpyxl
from playwright.sync_api import sync_playwright

EXCEL_PATH  = Path("/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx")
BASE_PROG   = Path("/Users/robin/.openclaw/workspace/mallorca-projekt")

PATTERNS = [
    r'[Tt]erreno[:\s]+([0-9][0-9.]+)',
    r'superficieParcela[":\s]+([0-9]+)',
    r'[Pp]arcela de ([0-9][0-9.]+)',
    r'[Ss]olar[:\s]+([0-9][0-9.]+)',
]
UA = ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
      'AppleWebKit/537.36 (KHTML, like Gecko) '
      'Chrome/122.0.0.0 Safari/537.36')

def extract_plot(html):
    for pat in PATTERNS:
        m = re.search(pat, html)
        if m:
            raw = m.group(1).replace('.','').replace(',','')
            try:
                val = int(raw)
                if val >= 50:
                    return val
            except ValueError:
                continue
    return None

def main():
    worker_id = int(sys.argv[1])
    start     = int(sys.argv[2])
    end       = int(sys.argv[3])
    prog_path = BASE_PROG / f"fetchdetails_progress_{worker_id}.json"

    print(f"[W{worker_id}] Loading Excel …", flush=True)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Identify columns
    url_col = plot_col = None
    for cell in ws[1]:
        v = str(cell.value).lower() if cell.value else ''
        if 'url' in v:     url_col  = cell.column
        if 'grundst' in v or 'plot' in v or 'terreno' in v:
            plot_col = cell.column

    # Collect ALL fotocasa rows (same order every worker)
    fotocasa_rows = []
    for row in ws.iter_rows(min_row=2):
        url = row[url_col - 1].value
        if url and 'fotocasa' in str(url).lower():
            fotocasa_rows.append(row)

    my_rows = fotocasa_rows[start:end]
    total   = len(my_rows)
    print(f"[W{worker_id}] Slice {start}-{end} → {total} URLs", flush=True)

    # Load existing progress for this worker
    progress = {}
    if prog_path.exists():
        with open(prog_path) as f:
            progress = json.load(f)

    save_counter = 0

    with sync_playwright() as p:
        for idx, row in enumerate(my_rows, 1):
            url = row[url_col - 1].value

            # Skip if already done with a valid plot
            if url in progress and progress[url].get('plot'):
                plot = progress[url]['plot']
                print(f"[W{worker_id}] [{idx}/{total}] CACHED {url} → plot={plot}m²", flush=True)
                save_counter += 1
                continue

            print(f"[W{worker_id}] [{idx}/{total}] {url}", end=" → ", flush=True)
            plot = None
            browser = None
            try:
                browser = p.chromium.launch(headless=True)
                ctx  = browser.new_context(locale='es-ES', user_agent=UA)
                page = ctx.new_page()
                page.goto(url, timeout=30000)
                time.sleep(1)
                html = page.content()
                browser.close()
                browser = None
                plot = extract_plot(html)
            except Exception as e:
                print(f"ERROR: {e}", flush=True)
                if browser:
                    try: browser.close()
                    except: pass
                progress[url] = {"plot": None, "error": str(e)}
                time.sleep(2)
                continue

            if plot:
                print(f"plot={plot}m²", flush=True)
            else:
                print("plot=None", flush=True)

            progress[url] = {"plot": plot}
            # NOTE: No Excel write here — workers only write JSON.
            # Merge script writes Excel at the end to avoid race conditions.

            save_counter += 1
            if save_counter % 50 == 0:
                with open(prog_path, 'w') as f:
                    json.dump(progress, f, indent=2)
                print(f"[W{worker_id}] [checkpoint @ {idx}]", flush=True)

            time.sleep(2)

    # Final save (JSON only)
    with open(prog_path, 'w') as f:
        json.dump(progress, f, indent=2)

    found = sum(1 for v in progress.values() if v.get('plot'))
    print(f"[W{worker_id}] DONE — {found}/{total} plots found. Progress: {prog_path}", flush=True)

if __name__ == "__main__":
    main()
