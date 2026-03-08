#!/usr/bin/env python3
"""
Mallorca Scraper - Pass 2
Fixes for ThinkSpain, Rightmove, APTS, Fotocasa, Yaencontre
"""

import time
import re
import json
from datetime import date
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
TODAY = str(date.today())

def parse_price(text):
    if not text: return None
    text = str(text).replace('\xa0', '').replace(' ', '').replace('.', '').replace(',', '')
    nums = re.findall(r'\d+', text)
    if nums:
        try:
            val = float(nums[0])
            if val > 10000: return val
        except: pass
    return None

def load_existing_data():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]: existing_urls.add(str(row[2]).strip())
    return wb, ws, existing_urls

def save_objects(wb, ws, new_objects, existing_urls):
    added = 0
    for obj in new_objects:
        url = str(obj.get('url', '')).strip()
        if url and url in existing_urls: continue
        ws.append([obj.get('titel',''), obj.get('quelle',''), url, obj.get('preis'),
                   obj.get('zimmer'), obj.get('grundstueck'), obj.get('wohnflaeche'),
                   obj.get('ort',''), TODAY, 'Neu'])
        if url: existing_urls.add(url)
        added += 1
    wb.save(EXCEL_PATH)
    return added


# =============================================
# ThinkSpain - Fixed
# =============================================
def scrape_thinkspain(page):
    print("\n=== ThinkSpain (Pass 2) ===")
    results = []
    base = "https://www.thinkspain.com"
    
    urls = [
        "https://www.thinkspain.com/property-for-sale/majorca/fincas-country-houses",
        "https://www.thinkspain.com/property-for-sale/majorca/villas",
        "https://www.thinkspain.com/property-for-sale/majorca/houses",
    ]
    seen = set()
    
    for base_url in urls:
        for pg in range(1, 15):
            url = base_url + (f"?page={pg}" if pg > 1 else "")
            try:
                page.goto(url, wait_until='domcontentloaded', timeout=20000)
                time.sleep(4)
                
                all_links = page.query_selector_all('article a[href]')
                found = 0
                
                for a in all_links:
                    href = a.get_attribute('href') or ''
                    if not re.search(r'/property-for-sale/\d+', href):
                        continue
                    full_url = base + href if href.startswith('/') else href
                    if full_url in seen: continue
                    
                    # Get article parent for details
                    try:
                        article_el = page.evaluate_handle("""(el) => {
                            let p = el;
                            while (p && p.tagName !== 'ARTICLE') p = p.parentElement;
                            return p;
                        }""", a)
                        text = article_el.inner_text() if article_el else ''
                    except:
                        text = ''
                    
                    price = parse_price(re.search(r'€\s*([\d\.,]+)', text).group(1) if re.search(r'€\s*([\d\.,]+)', text) else '')
                    bed_m = re.search(r'(\d+)\s*bed', text, re.I)
                    beds = int(bed_m.group(1)) if bed_m else None
                    
                    title_lines = [l.strip() for l in text.split('\n') if l.strip() and len(l.strip()) > 10 and not l.strip().startswith('€')]
                    title = title_lines[0][:100] if title_lines else 'ThinkSpain Property'
                    
                    # Extract location
                    loc_m = re.search(r'(?:in|near)\s+([A-Z][a-z]+(?:\s+[a-z]+)?)', text)
                    ort = loc_m.group(1) if loc_m else 'Mallorca'
                    
                    seen.add(full_url)
                    found += 1
                    results.append({'titel': title, 'quelle': 'ThinkSpain', 'url': full_url,
                                   'preis': price, 'zimmer': beds, 'grundstueck': None,
                                   'wohnflaeche': None, 'ort': ort})
                
                print(f"  {base_url.split('/')[-1]} page {pg}: {found} new, {len(results)} total")
                if found == 0: break
                    
            except Exception as e:
                print(f"  Error page {pg}: {e}")
                break
    
    print(f"  ThinkSpain total: {len(results)}")
    return results


# =============================================
# Rightmove - Intercept XHR API
# =============================================
def scrape_rightmove(page):
    print("\n=== Rightmove (Pass 2) ===")
    results = []
    
    # Intercept API responses
    api_data = []
    
    def on_response(response):
        url = response.url
        if 'api' in url.lower() and ('property' in url.lower() or 'search' in url.lower()) and response.status == 200:
            try:
                data = response.json()
                if data:
                    api_data.append(data)
            except:
                pass
    
    page.on('response', on_response)
    
    try:
        page.goto("https://www.rightmove.co.uk/overseas-property/in-Mallorca.html",
                  wait_until='domcontentloaded', timeout=30000)
        time.sleep(5)
        
        # Accept cookies
        try:
            btn = page.query_selector('button:has-text("Accept all"), #onetrust-accept-btn-handler')
            if btn: btn.click(); time.sleep(1)
        except: pass
        
        # Scroll to trigger lazy loading
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(3)
        
        # Check what we got from API interception
        print(f"  API responses captured: {len(api_data)}")
        
        # Try to find property data in the page source
        page_source = page.content()
        
        # Rightmove embeds property data in JSON
        json_match = re.search(r'window\.jsonModel\s*=\s*({.+?})\s*(?:;|</script)', page_source, re.DOTALL)
        if json_match:
            try:
                model = json.loads(json_match.group(1))
                props = model.get('properties', [])
                print(f"  Found {len(props)} properties in window.jsonModel")
                
                for prop in props:
                    url = f"https://www.rightmove.co.uk{prop.get('propertyUrl', '')}"
                    price = prop.get('price', {}).get('displayPrices', [{}])[0].get('displayPrice', '')
                    price_val = parse_price(price.replace('€','').replace('£','').replace(',',''))
                    beds = prop.get('bedrooms')
                    title = prop.get('summary', '') or prop.get('displayAddress', '')
                    location = prop.get('location', {})
                    ort = prop.get('displayAddress', 'Mallorca')
                    
                    results.append({
                        'titel': title[:100] or 'Rightmove Property',
                        'quelle': 'Rightmove',
                        'url': url,
                        'preis': price_val,
                        'zimmer': beds,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': ort
                    })
            except Exception as e:
                print(f"  JSON parse error: {e}")
        
        # Also try extracting from JSON-LD or other embedded data
        if not results:
            json_ld = re.findall(r'<script type="application/ld\+json">(.+?)</script>', page_source, re.DOTALL)
            for jld in json_ld:
                try:
                    data = json.loads(jld)
                    if isinstance(data, dict) and data.get('@type') == 'ItemList':
                        items = data.get('itemListElement', [])
                        for item in items:
                            thing = item.get('item', {})
                            if 'url' in thing:
                                results.append({
                                    'titel': thing.get('name', 'Rightmove Property')[:100],
                                    'quelle': 'Rightmove',
                                    'url': thing['url'],
                                    'preis': None,
                                    'zimmer': None,
                                    'grundstueck': None,
                                    'wohnflaeche': None,
                                    'ort': 'Mallorca'
                                })
                except: pass
        
        # Try looking at current page DOM for property cards
        if not results:
            cards_data = page.evaluate("""() => {
                const cards = Array.from(document.querySelectorAll('[class*="SearchResult"], [class*="propertyCard"], li.l-searchResult'));
                return cards.map(card => {
                    const a = card.querySelector('a[href*="/properties/"]');
                    const title = card.querySelector('h2, [class*="title"]');
                    const price = card.querySelector('[class*="price"]');
                    return {
                        url: a ? a.href : '',
                        title: title ? title.innerText.trim() : '',
                        price: price ? price.innerText.trim() : ''
                    };
                }).filter(d => d.url);
            }""")
            print(f"  DOM cards: {len(cards_data)}")
            for c in cards_data:
                if c['url']:
                    results.append({
                        'titel': c['title'][:100] or 'Rightmove Property',
                        'quelle': 'Rightmove',
                        'url': c['url'],
                        'preis': parse_price(c['price']),
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca'
                    })
        
        # Multi-page: try pagination via URL
        total_pages = 1
        count_m = re.search(r'(\d+)\s+(?:result|propert)', page.title(), re.I)
        if count_m:
            count = int(count_m.group(1))
            total_pages = min(20, (count // 24) + 1)
            print(f"  Estimated {total_pages} pages")
        
        for pg in range(1, total_pages):
            purl = f"https://www.rightmove.co.uk/overseas-property/in-Mallorca.html?index={pg * 24}"
            try:
                page.goto(purl, wait_until='domcontentloaded', timeout=20000)
                time.sleep(3)
                
                src = page.content()
                json_m = re.search(r'window\.jsonModel\s*=\s*({.+?})\s*(?:;|</script)', src, re.DOTALL)
                if json_m:
                    model = json.loads(json_m.group(1))
                    props = model.get('properties', [])
                    for prop in props:
                        u = f"https://www.rightmove.co.uk{prop.get('propertyUrl', '')}"
                        if any(r['url'] == u for r in results): continue
                        price = prop.get('price', {}).get('displayPrices', [{}])[0].get('displayPrice', '')
                        results.append({
                            'titel': (prop.get('summary', '') or prop.get('displayAddress', ''))[:100] or 'Rightmove Property',
                            'quelle': 'Rightmove',
                            'url': u,
                            'preis': parse_price(price.replace('€','').replace('£','').replace(',','')),
                            'zimmer': prop.get('bedrooms'),
                            'grundstueck': None, 'wohnflaeche': None,
                            'ort': prop.get('displayAddress', 'Mallorca')
                        })
                    print(f"  Page {pg+1}: {len(props)} props, total {len(results)}")
                    if len(props) == 0: break
            except Exception as e:
                print(f"  Error page {pg}: {e}")
                break
    
    except Exception as e:
        print(f"  Error: {e}")
    
    print(f"  Rightmove total: {len(results)}")
    return results


# =============================================
# A Place in the Sun - Fixed
# =============================================
def scrape_aplaceinthesun(page):
    print("\n=== A Place in the Sun (Pass 2) ===")
    results = []
    base = "https://www.aplaceinthesun.com"
    
    # Intercept API
    api_data = []
    def on_response(r):
        if 'aplaceinthesun.com' in r.url and r.status == 200:
            if 'api' in r.url or 'json' in r.headers.get('content-type', ''):
                try:
                    api_data.append(r.json())
                except: pass
    page.on('response', on_response)
    
    try:
        page.goto("https://www.aplaceinthesun.com/property/spain/balearic-islands/mallorca",
                  wait_until='domcontentloaded', timeout=30000)
        time.sleep(4)
        
        seen = set()
        
        for attempt in range(10):
            # Find property cards
            cards_data = page.evaluate("""() => {
                const results = [];
                const seen = new Set();
                
                // Try multiple selectors
                const cards = Array.from(document.querySelectorAll('[class*="PropertyCard"], [class*="propertyCard"], article, [data-testid*="property"]'));
                
                cards.forEach(card => {
                    const links = Array.from(card.querySelectorAll('a[href]'));
                    links.forEach(a => {
                        const href = a.href;
                        if (href.includes('/property/spain/balearic-islands/mallorca/') && href.length > 70 && !seen.has(href)) {
                            seen.add(href);
                            const titleEl = card.querySelector('h2, h3, [class*="title"]');
                            const priceEl = card.querySelector('[class*="price"], [class*="Price"]');
                            const bedsEl = card.querySelector('[class*="bed"], [class*="room"]');
                            results.push({
                                url: href,
                                title: titleEl ? titleEl.innerText.trim() : '',
                                price: priceEl ? priceEl.innerText.trim() : '',
                                beds: bedsEl ? bedsEl.innerText.trim() : ''
                            });
                        }
                    });
                });
                return results;
            }""")
            
            new_found = 0
            for c in cards_data:
                if c['url'] not in seen:
                    seen.add(c['url'])
                    new_found += 1
                    results.append({
                        'titel': c['title'][:100] or 'APTS Property',
                        'quelle': 'A Place in the Sun',
                        'url': c['url'],
                        'preis': parse_price(c['price']),
                        'zimmer': int(re.search(r'(\d+)', c['beds']).group(1)) if re.search(r'(\d+)', c['beds']) else None,
                        'grundstueck': None, 'wohnflaeche': None, 'ort': 'Mallorca'
                    })
            
            print(f"  Attempt {attempt+1}: {new_found} new, total {len(results)}")
            
            # Try to load more
            load_more = page.query_selector('button:has-text("Load more"), button:has-text("Show more"), [class*="load-more"]')
            if load_more and new_found > 0:
                load_more.click()
                time.sleep(3)
            else:
                # Try next page link
                next_btn = page.query_selector('a[aria-label*="next" i], a[aria-label*="Next"], [class*="next"] a')
                if next_btn:
                    next_btn.click()
                    time.sleep(3)
                else:
                    break
    
    except Exception as e:
        print(f"  Error: {e}")
    
    print(f"  A Place in the Sun total: {len(results)}")
    return results


# =============================================
# Fotocasa - Intercept API
# =============================================
def scrape_fotocasa(page):
    print("\n=== Fotocasa (Pass 2) ===")
    results = []
    
    api_results = []
    
    def on_response(r):
        url = r.url
        if 'fotocasa' in url and r.status == 200:
            ct = r.headers.get('content-type', '')
            if 'json' in ct:
                try:
                    data = r.json()
                    if isinstance(data, dict) and ('realEstates' in data or 'items' in data or 'content' in data):
                        api_results.append({'url': url, 'data': data})
                        print(f"  API hit: {url[:100]}")
                except: pass
    
    page.on('response', on_response)
    
    try:
        page.goto("https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l",
                  wait_until='networkidle', timeout=30000)
        time.sleep(4)
        
        # Accept cookies
        try:
            btn = page.query_selector('button:has-text("Aceptar"), button[id*="accept"]')
            if btn: btn.click(); time.sleep(1)
        except: pass
        
        # Scroll to load more
        for _ in range(3):
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(2)
        
        print(f"  API responses captured: {len(api_results)}")
        
        # Parse API data
        for api_r in api_results:
            data = api_r['data']
            items = data.get('realEstates', data.get('items', data.get('content', [])))
            if isinstance(items, list):
                for item in items:
                    if isinstance(item, dict):
                        url = item.get('url', '') or item.get('link', '') or item.get('detail_url', '')
                        if url and not url.startswith('http'):
                            url = 'https://www.fotocasa.es' + url
                        price = item.get('price', {})
                        if isinstance(price, dict):
                            price = price.get('value', price.get('amount', 0))
                        
                        title = item.get('title', '') or item.get('description', '') or 'Fotocasa Property'
                        rooms = item.get('rooms', item.get('bedrooms', None))
                        surface = item.get('surface', item.get('squareMeters', None))
                        
                        address = item.get('address', {}) or {}
                        if isinstance(address, dict):
                            ort = address.get('municipalityName', address.get('city', 'Mallorca'))
                        else:
                            ort = 'Mallorca'
                        
                        if url:
                            results.append({
                                'titel': str(title)[:100],
                                'quelle': 'Fotocasa',
                                'url': url,
                                'preis': float(price) if price else None,
                                'zimmer': int(rooms) if rooms else None,
                                'grundstueck': None,
                                'wohnflaeche': float(surface) if surface else None,
                                'ort': str(ort)
                            })
        
        # Also try DOM-based extraction
        if not results:
            dom_results = page.evaluate("""() => {
                const items = Array.from(document.querySelectorAll('[class*="PropertyCard"], [class*="re-Card"], article[class*="card"]'));
                return items.map(item => {
                    const a = item.querySelector('a[href*="/comprar/"]');
                    const price = item.querySelector('[class*="Price"], [class*="price"]');
                    const rooms = item.querySelector('[class*="rooms"], [class*="bedrooms"]');
                    const title = item.querySelector('h2, h3, [class*="Title"]');
                    return {
                        url: a ? a.href : '',
                        price: price ? price.innerText : '',
                        rooms: rooms ? rooms.innerText : '',
                        title: title ? title.innerText.trim() : ''
                    };
                }).filter(d => d.url);
            }""")
            
            print(f"  DOM results: {len(dom_results)}")
            for item in dom_results:
                results.append({
                    'titel': item['title'][:100] or 'Fotocasa Property',
                    'quelle': 'Fotocasa',
                    'url': item['url'],
                    'preis': parse_price(item['price']),
                    'zimmer': int(re.search(r'(\d+)', item['rooms']).group(1)) if re.search(r'(\d+)', item['rooms']) else None,
                    'grundstueck': None, 'wohnflaeche': None, 'ort': 'Mallorca'
                })
    
    except Exception as e:
        print(f"  Error: {e}")
    
    print(f"  Fotocasa total: {len(results)}")
    return results


# =============================================
# Yaencontre - Direct requests approach
# =============================================
def scrape_yaencontre_requests():
    print("\n=== Yaencontré (requests) ===")
    import requests
    results = []
    
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'es-ES,es;q=0.9',
        'Referer': 'https://www.yaencontre.com/',
    })
    
    # Try their API
    api_urls = [
        "https://www.yaencontre.com/api/search/properties?type=HOUSE&operation=SALE&zone=mallorca&page=1&pageSize=50",
        "https://www.yaencontre.com/api/search?type=home&operation=sale&zone=costa-mallorca&page=1",
    ]
    
    for api_url in api_urls:
        try:
            r = session.get(api_url, timeout=15)
            print(f"  API {api_url[:60]}: {r.status_code}")
            if r.status_code == 200:
                data = r.json()
                print(f"  Data keys: {list(data.keys())[:10] if isinstance(data, dict) else 'list'}")
        except Exception as e:
            print(f"  Error: {e}")
    
    # Direct page scraping
    from bs4 import BeautifulSoup
    
    urls = [
        "https://www.yaencontre.com/venta/casas/costa-mallorca",
        "https://www.yaencontre.com/venta/chalets/costa-mallorca",
    ]
    
    for url in urls:
        try:
            r = session.get(url, timeout=15)
            print(f"  GET {url}: {r.status_code}")
            if r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')
                # Find JSON in page
                scripts = soup.find_all('script')
                for s in scripts:
                    if s.string and ('properties' in s.string.lower() or 'results' in s.string.lower()):
                        # Try to find JSON data
                        json_m = re.search(r'\{.{100,}\}', s.string)
                        if json_m:
                            try:
                                data = json.loads(json_m.group(0))
                                print(f"  Found JSON in script with keys: {list(data.keys())[:5]}")
                            except: pass
                
                # Try next.js data
                next_data = soup.find('script', {'id': '__NEXT_DATA__'})
                if next_data and next_data.string:
                    try:
                        data = json.loads(next_data.string)
                        props = data.get('props', {}).get('pageProps', {})
                        print(f"  __NEXT_DATA__ pageProps keys: {list(props.keys())[:10]}")
                        
                        # Navigate the data structure to find listings
                        def find_listings(d, depth=0):
                            if depth > 5: return []
                            if isinstance(d, list) and len(d) > 0:
                                if isinstance(d[0], dict) and ('price' in d[0] or 'url' in d[0] or 'id' in d[0]):
                                    return d
                            if isinstance(d, dict):
                                for key, val in d.items():
                                    found = find_listings(val, depth+1)
                                    if found:
                                        return found
                            return []
                        
                        listings = find_listings(props)
                        print(f"  Found listings array: {len(listings)}")
                        
                        for item in listings[:5]:
                            print(f"    Item keys: {list(item.keys())[:10]}")
                            
                    except Exception as e:
                        print(f"  Error parsing NEXT_DATA: {e}")
        except Exception as e:
            print(f"  Error: {e}")
    
    print(f"  Yaencontré total: {len(results)}")
    return results


# =============================================
# Green-Acres - Better extraction
# =============================================
def scrape_greenacres(page):
    print("\n=== Green-Acres (Pass 2) ===")
    results = []
    base = "https://www.green-acres.es"
    
    api_data = []
    def on_response(r):
        if 'green-acres' in r.url and r.status == 200:
            ct = r.headers.get('content-type', '')
            if 'json' in ct:
                try:
                    data = r.json()
                    if isinstance(data, (dict, list)) and str(data)[:100]:
                        api_data.append({'url': r.url, 'data': data})
                except: pass
    page.on('response', on_response)
    
    search_urls = [
        "https://www.green-acres.es/property-for-sale/majorca",
        "https://www.green-acres.es/house/majorca",
        "https://www.green-acres.es/villa/majorca",
        "https://www.green-acres.es/country-house/majorca",
    ]
    
    seen = set()
    
    for base_url in search_urls:
        for pg in range(1, 10):
            url = base_url + (f"?page={pg}" if pg > 1 else "")
            try:
                page.goto(url, wait_until='networkidle', timeout=25000)
                time.sleep(3)
                
                # Extract property links
                links_data = page.evaluate("""() => {
                    const links = Array.from(document.querySelectorAll('a[href]'));
                    const results = [];
                    const seen = new Set();
                    links.forEach(a => {
                        const href = a.href;
                        // Property pages have long URLs
                        if (href.includes('green-acres.es') && href.length > 60 && 
                            !href.includes('?') && !href.includes('#') &&
                            (href.includes('/house/') || href.includes('/villa/') || 
                             href.includes('/property/') || href.includes('/country-house/') ||
                             href.includes('/apartment/') || href.includes('/annonce'))) {
                            if (!seen.has(href)) {
                                seen.add(href);
                                const card = a.closest('article') || a.closest('[class*="card"]') || a.closest('[class*="property"]') || a.parentElement;
                                const priceEl = card ? card.querySelector('[class*="price"], [class*="Price"]') : null;
                                const titleEl = card ? card.querySelector('h2, h3, [class*="title"]') : null;
                                const bedsEl = card ? card.querySelector('[class*="bed"], [class*="room"]') : null;
                                results.push({
                                    url: href,
                                    price: priceEl ? priceEl.innerText.trim() : '',
                                    title: titleEl ? titleEl.innerText.trim() : a.innerText.trim(),
                                    beds: bedsEl ? bedsEl.innerText.trim() : ''
                                });
                            }
                        }
                    });
                    return results;
                }""")
                
                found = 0
                for item in links_data:
                    if item['url'] not in seen:
                        seen.add(item['url'])
                        found += 1
                        results.append({
                            'titel': item['title'][:100] or 'Green-Acres Property',
                            'quelle': 'Green-Acres',
                            'url': item['url'],
                            'preis': parse_price(item['price']),
                            'zimmer': int(re.search(r'(\d+)', item['beds']).group(1)) if re.search(r'(\d+)', item['beds']) else None,
                            'grundstueck': None, 'wohnflaeche': None, 'ort': 'Mallorca'
                        })
                
                print(f"  {base_url.split('/')[-1]} page {pg}: {found} new, {len(results)} total")
                if found == 0: break
            except Exception as e:
                print(f"  Error: {e}")
                break
    
    print(f"  Green-Acres total: {len(results)}")
    return results


# =============================================
# MAIN
# =============================================
def main():
    print("Loading Excel...")
    wb, ws, existing_urls = load_existing_data()
    print(f"Existing URLs: {len(existing_urls)}")
    
    summary = {}
    total_added = 0
    
    # First run requests-based scrapers (no browser needed)
    try:
        results = scrape_yaencontre_requests()
        if results:
            added = save_objects(wb, ws, results, existing_urls)
            summary['Yaencontré'] = {'scraped': len(results), 'added': added}
            total_added += added
            print(f"  ✓ Yaencontré: {added} added")
        else:
            summary['Yaencontré'] = {'scraped': 0, 'added': 0}
    except Exception as e:
        print(f"  ✗ Yaencontré: {e}")
        summary['Yaencontré'] = {'scraped': 0, 'added': 0, 'error': str(e)}
    
    # Playwright scrapers
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-dev-shm-usage'])
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
            locale='en-GB'
        )
        page = context.new_page()
        
        scrapers = [
            ('ThinkSpain', scrape_thinkspain),
            ('Rightmove', scrape_rightmove),
            ('A Place in the Sun', scrape_aplaceinthesun),
            ('Fotocasa', scrape_fotocasa),
            ('Green-Acres', scrape_greenacres),
        ]
        
        for name, fn in scrapers:
            try:
                results = fn(page)
                if results:
                    added = save_objects(wb, ws, results, existing_urls)
                    summary[name] = {'scraped': len(results), 'added': added}
                    total_added += added
                    print(f"\n  ✓ {name}: {len(results)} scraped, {added} added")
                else:
                    summary[name] = {'scraped': 0, 'added': 0}
                    print(f"\n  ✗ {name}: 0 results")
            except Exception as e:
                print(f"\n  ✗ {name}: {e}")
                summary[name] = {'scraped': 0, 'added': 0, 'error': str(e)[:100]}
        
        browser.close()
    
    print(f"\n{'='*50}")
    print("PASS 2 SUMMARY")
    for name, s in summary.items():
        print(f"  {name}: {s.get('scraped',0)} scraped → {s.get('added',0)} added")
        if 'error' in s: print(f"    ERR: {s['error'][:80]}")
    print(f"\nTotal new: {total_added}")
    
    with open('/Users/robin/.openclaw/workspace/mallorca-projekt/phase1_pass2_results.json', 'w') as f:
        json.dump({'summary': summary, 'total_added': total_added}, f, indent=2)

if __name__ == '__main__':
    main()
