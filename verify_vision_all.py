#!/usr/bin/env python3
"""
Vollständige Vision-Verifikation aller 332 Einträge.
Playwright öffnet URL, extrahiert Hauptbild, Claude Vision vergleicht mit gespeichertem.
"""
import asyncio, json, base64, os, sys
from pathlib import Path
from playwright.async_api import async_playwright
import openpyxl

BASE = Path(__file__).parent
RESULTS_FILE = BASE / 'vision_verify_results.json'

# Load progress if exists
if RESULTS_FILE.exists():
    done = {r['nr']: r for r in json.load(open(RESULTS_FILE))}
else:
    done = {}

def compare_vision(stored_path, live_path):
    """Use oracle CLI to compare two images."""
    import subprocess
    result = subprocess.run([
        'oracle', '-e', 'haiku',
        '--prompt', 'Zeigen diese zwei Bilder dasselbe Haus/Objekt? Antworte nur: JA oder NEIN',
        '--attach', str(stored_path),
        '--attach', str(live_path),
    ], capture_output=True, text=True, timeout=30)
    out = result.stdout.strip()
    return out

async def get_live_image(page, url, nr):
    try:
        await page.goto(url, timeout=25000, wait_until='domcontentloaded')
        await asyncio.sleep(2)
        
        img_url = None
        selectors = [
            'img[src*="photo"]','img[src*="image"]','img[src*="foto"]',
            'img[src*="cdn"]','img[src*="media"]',
            '.gallery img','.slider img','picture img',
            '[class*="gallery"] img','[class*="photo"] img',
        ]
        for sel in selectors:
            try:
                els = await page.query_selector_all(sel)
                for el in els[:5]:
                    for attr in ['src','data-src','data-lazy']:
                        src = await el.get_attribute(attr) or ''
                        src = src.split(',')[0].split(' ')[0]
                        if src and any(x in src.lower() for x in ['jpg','jpeg','webp','png']) and len(src)>20:
                            img_url = src
                            break
                    if img_url: break
            except: pass
            if img_url: break
        
        if not img_url:
            imgs = await page.query_selector_all('img')
            best = (0, None)
            for img in imgs[:40]:
                try:
                    w = await page.evaluate('(el) => el.naturalWidth', img)
                    src = await img.get_attribute('src') or ''
                    if w > best[0] and w > 200 and src and 'logo' not in src.lower():
                        best = (w, src)
                except: pass
            img_url = best[1]
        
        if not img_url: return None
        if img_url.startswith('//'): img_url = 'https:' + img_url
        elif img_url.startswith('/'):
            from urllib.parse import urlparse
            parsed = urlparse(url)
            img_url = f"{parsed.scheme}://{parsed.netloc}{img_url}"
        
        try:
            resp = await page.request.get(img_url, timeout=15000)
            if resp.status == 200:
                out = BASE / f'tmp_live_{nr}.jpg'
                out.write_bytes(await resp.body())
                return str(out)
        except: pass
        return None
    except Exception as e:
        return None

async def main():
    # Load Excel
    wb = openpyxl.load_workbook(BASE / 'mallorca-kandidaten-v2.xlsx')
    ws = wb.active
    headers = [c.value for c in ws[1]]
    
    entries = []
    for row in ws.iter_rows(min_row=2):
        nr = row[headers.index('Ordnungsnummer')].value
        if not nr: continue
        url = row[headers.index('Link Objekt (URL)')].value or ''
        komm = str(row[headers.index('Kommentar')].value or '')
        name = row[headers.index('Name')].value or ''
        makler = row[headers.index('Makler')].value or ''
        delisted = '⚠️ delisted' in komm
        entries.append({'nr': int(nr), 'url': url, 'name': name, 'makler': makler, 'delisted': delisted})
    
    results = list(done.values())
    todo = [e for e in entries if e['nr'] not in done]
    print(f"Total: {len(entries)} | Already done: {len(done)} | Todo: {len(todo)}", flush=True)
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
            viewport={'width': 1280, 'height': 900}
        )
        
        for i, entry in enumerate(todo):
            nr = entry['nr']
            url = entry['url']
            stored = BASE / f'bilder/{nr}_main.jpg'
            
            sys.stdout.write(f"[{i+1}/{len(todo)}] Nr.{nr:3d} {entry['makler'][:15]:<15} ")
            sys.stdout.flush()
            
            # Already delisted
            if entry['delisted']:
                r = {'nr': nr, 'status': 'DELISTED', 'name': entry['name']}
                results.append(r)
                print("⚠️  delisted")
                done[nr] = r
                continue
            
            # No URL
            if not url or url == '—':
                r = {'nr': nr, 'status': 'NO_URL', 'name': entry['name']}
                results.append(r)
                print("⬜ kein URL")
                done[nr] = r
                continue
            
            # No stored image
            if not stored.exists():
                r = {'nr': nr, 'status': 'NO_STORED', 'name': entry['name'], 'url': url}
                results.append(r)
                print("⬜ kein Bild gespeichert")
                done[nr] = r
                continue
            
            page = await ctx.new_page()
            try:
                live_path = await get_live_image(page, url, nr)
                
                if not live_path:
                    r = {'nr': nr, 'status': 'NO_LIVE', 'name': entry['name'], 'url': url, 'makler': entry['makler']}
                    results.append(r)
                    print("❓ kein Live-Bild (Bot-Block?)")
                    done[nr] = r
                    continue
                
                # Vision compare
                verdict = compare_vision(stored, live_path)
                ok = 'JA' in verdict.upper()[:5]
                status = 'MATCH' if ok else 'MISMATCH'
                icon = '✅' if ok else '❌'
                print(f"{icon} {verdict[:60]}")
                
                r = {'nr': nr, 'status': status, 'name': entry['name'], 'url': url, 'makler': entry['makler'], 'verdict': verdict}
                results.append(r)
                done[nr] = r
                
                # Cleanup live tmp
                Path(live_path).unlink(missing_ok=True)
                
            except Exception as e:
                r = {'nr': nr, 'status': 'ERR', 'name': entry['name'], 'error': str(e)}
                results.append(r)
                print(f"⚠️  {str(e)[:50]}")
                done[nr] = r
            finally:
                await page.close()
            
            # Save progress every 10
            if (i+1) % 10 == 0:
                json.dump(results, open(RESULTS_FILE, 'w'), indent=2)
        
        await browser.close()
    
    json.dump(results, open(RESULTS_FILE, 'w'), indent=2)
    
    # Summary
    from collections import Counter
    counts = Counter(r['status'] for r in results)
    print(f"\n{'='*60}")
    print("ERGEBNIS:")
    for k, v in counts.most_common():
        print(f"  {k:15s} {v}")
    
    mismatches = [r for r in results if r['status'] == 'MISMATCH']
    if mismatches:
        print(f"\n❌ MISMATCHES ({len(mismatches)}):")
        for m in mismatches:
            print(f"  Nr.{m['nr']:3d} {m['makler']:<15} {m['name'][:50]}")

asyncio.run(main())
