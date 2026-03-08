import requests
from bs4 import BeautifulSoup
import json, time, re
from openpyxl import load_workbook
from datetime import date

wb = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
ws = wb['Mallorca Objekte']
existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
print(f"Start: {len(existing_urls)} URLs bekannt, {ws.max_row} Zeilen")

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'de-DE,de;q=0.9,en;q=0.8',
}

vp_count = 0

# Methode 1: Von Poll direkt scrapen - Mallorca Suchergebnisse
# Von Poll hat eine eigene Suche unter /de/search-results
base_urls = [
    'https://www.von-poll.com/de/search-results?location=Mallorca&type=buy&property_type=&price_from=&price_to=&rooms_from=&rooms_to=',
    'https://www.von-poll.com/de/search-results?location=Mallorca&type=buy',
    'https://www.von-poll.com/de/immobilien-kaufen/mallorca',
    'https://www.von-poll.com/de/immobilien/balearen/mallorca',
]

for base_url in base_urls:
    try:
        r = requests.get(base_url, headers=headers, timeout=30)
        print(f"URL: {base_url[:60]} -> HTTP {r.status_code}")
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')
            # Suche nach Expose-Links
            links = soup.find_all('a', href=re.compile(r'/de/expose/'))
            print(f"  Expose-Links gefunden: {len(links)}")
            for link in links:
                url = link.get('href', '')
                if url and not url.startswith('http'):
                    url = 'https://www.von-poll.com' + url
                if url not in existing_urls:
                    titel = link.get_text(strip=True)[:100] or 'Von Poll Objekt'
                    ws.append([titel, 'Von Poll Real Estate', url, None, None, None, None,
                               'Mallorca', str(date.today()), 'Neu'])
                    existing_urls.add(url)
                    vp_count += 1
        time.sleep(1)
    except Exception as e:
        print(f"Error: {e}")

# Methode 2: Von Poll API/JSON Endpunkte probieren
api_urls = [
    'https://www.von-poll.com/api/properties?location=mallorca&limit=100',
    'https://www.von-poll.com/de/suche?ort=Mallorca&kaufen=1',
    'https://www.von-poll.com/de/immobilien-kaufen?region=mallorca',
    'https://www.von-poll.com/de/immobilien/spanien/mallorca/kaufen',
    'https://www.von-poll.com/de/immobilien/spanien/balearen/mallorca/kaufen',
]

for api_url in api_urls:
    try:
        r = requests.get(api_url, headers=headers, timeout=30)
        print(f"URL: {api_url[:70]} -> HTTP {r.status_code}")
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')
            links = soup.find_all('a', href=re.compile(r'/de/expose/|/expose/'))
            print(f"  Expose-Links: {len(links)}")
            for link in links:
                url = link.get('href', '')
                if url and not url.startswith('http'):
                    url = 'https://www.von-poll.com' + url
                if 'von-poll.com' in url and url not in existing_urls:
                    titel = link.get_text(strip=True)[:100] or 'Von Poll Objekt'
                    ws.append([titel, 'Von Poll Real Estate', url, None, None, None, None,
                               'Mallorca', str(date.today()), 'Neu'])
                    existing_urls.add(url)
                    vp_count += 1
        time.sleep(1)
    except Exception as e:
        print(f"Error: {e}")

# Methode 3: Von Poll Sitemap
sitemap_urls = [
    'https://www.von-poll.com/sitemap.xml',
    'https://www.von-poll.com/sitemap-properties.xml',
    'https://www.von-poll.com/sitemap_index.xml',
]

for sitemap_url in sitemap_urls:
    try:
        r = requests.get(sitemap_url, headers=headers, timeout=30)
        print(f"Sitemap: {sitemap_url} -> HTTP {r.status_code}")
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'xml')
            locs = soup.find_all('loc')
            print(f"  Einträge in Sitemap: {len(locs)}")
            # Suche nach Mallorca Expose URLs
            mallorca_urls = [l.text for l in locs if 'expose' in l.text.lower() or 'mallorca' in l.text.lower() or 'balearen' in l.text.lower()]
            print(f"  Mallorca/Expose URLs: {len(mallorca_urls)}")
            for url in mallorca_urls[:50]:
                if url not in existing_urls:
                    ws.append(['Von Poll Objekt', 'Von Poll Real Estate', url, None, None, None, None,
                               'Mallorca', str(date.today()), 'Neu'])
                    existing_urls.add(url)
                    vp_count += 1
        time.sleep(1)
    except Exception as e:
        print(f"Sitemap error: {e}")

print(f"\nVon Poll gesamt: {vp_count}")

if vp_count > 0:
    wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
    print(f"✅ Gespeichert. Finale Zeilenzahl: {ws.max_row}")
else:
    print("⚠️ Keine Von Poll Daten gefunden — kein Speichern nötig")
    print("Aktuelle Zeilenzahl:", ws.max_row)
