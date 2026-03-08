#!/usr/bin/env python3
"""
Mallorca Scraper - Final Pass
APTS, Yaencontre, Rightmove, Fotocasa with correct approaches
"""

import time
import re
import json
import base64
from datetime import date
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import requests
from bs4 import BeautifulSoup

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
TODAY = str(date.today())

def parse_price(text):
    if not text: return None
    text = str(text).replace('\xa0', '').replace(' ', '')
    # Remove currency symbols
    text = re.sub(r'[€£$]', '', text)
    # Remove thousand separators
    text = re.sub(r'[,\.](?=\d{3})', '', text)
    nums = re.findall(r'\d+', text)
    if nums:
        try:
            val = float(nums[0])
            if val > 10000: return val
        except: pass
    return None

def load_existing_data():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]: existing_urls.add(str(row[2]).strip())
    return wb, ws, existing_urls

def save_objects(wb, ws, new_objects, existing_urls):
    added = 0
    for obj in new_objects:
        url = str(obj.get('url', '')).strip()
        if url and url in existing_urls: continue
        ws.append([obj.get('titel',''), obj.get('quelle',''), url, obj.get('preis'),
                   obj.get('zimmer'), obj.get('grundstueck'), obj.get('wohnflaeche'),
                   obj.get('ort',''), TODAY, 'Neu'])
        if url: existing_urls.add(url)
        added += 1
    wb.save(EXCEL_PATH)
    return added


# =============================================
# A Place in the Sun - Fixed with correct URL pattern
# =============================================
def scrape_aplaceinthesun(page):
    print("\n=== A Place in the Sun (Final) ===")
    results = []
    base = "https://www.aplaceinthesun.com"
    seen = set()
    
    urls_to_scrape = [
        "https://www.aplaceinthesun.com/property/spain/balearic-islands/mallorca",
        "https://www.aplaceinthesun.com/property/spain/balearic-islands/mallorca/houses",
        "https://www.aplaceinthesun.com/property/spain/balearic-islands/mallorca/villa",
        "https://www.aplaceinthesun.com/property/spain/balearic-islands/mallorca/country-house",
    ]
    
    for base_url in urls_to_scrape:
        for pg in range(1, 15):
            # APTS uses /page/N/ in the URL
            url = base_url + (f"/page/{pg}/" if pg > 1 else "")
            try:
                page.goto(url, wait_until='domcontentloaded', timeout=20000)
                time.sleep(3)
                
                # Dismiss cookie popup via JS
                page.evaluate("""() => {
                    const overlays = document.querySelectorAll('[id*="consent"], [id*="cookie"], [class*="consent"], .cl-consent');
                    overlays.forEach(el => { el.style.display = 'none'; try { el.remove(); } catch(e) {} });
                    document.body.style.overflow = 'auto';
                    document.documentElement.style.overflow = 'auto';
                }""")
                time.sleep(0.5)
                
                # Get property detail links
                links_data = page.evaluate("""() => {
                    const seen = new Set();
                    const results = [];
                    const links = Array.from(document.querySelectorAll('a[href]'));
                    links.forEach(a => {
                        const href = a.href;
                        // Match property detail URLs
                        if (/\\/property\\/details\\/ap\\d+/.test(href) && !seen.has(href)) {
                            seen.add(href);
                            const card = a.closest('[class*="card"]') || a.closest('article') || a.closest('li') || a.parentElement;
                            const priceEl = card ? card.querySelector('[class*="price"], [class*="Price"]') : null;
                            const titleEl = card ? card.querySelector('h2, h3, h4, [class*="title"]') : null;
                            results.push({
                                url: href,
                                title: titleEl ? titleEl.innerText.trim() : a.innerText.trim(),
                                price: priceEl ? priceEl.innerText.trim() : ''
                            });
                        }
                    });
                    return results;
                }""")
                
                found = 0
                for item in links_data:
                    if item['url'] not in seen:
                        seen.add(item['url'])
                        found += 1
                        
                        # Extract beds from URL
                        bed_m = re.search(r'(\d+)-bed', item['url'])
                        beds = int(bed_m.group(1)) if bed_m else None
                        
                        # Extract location from URL
                        loc_m = re.search(r'for-sale-in-([^/]+)$', item['url'])
                        ort = loc_m.group(1).replace('-', ' ').title() if loc_m else 'Mallorca'
                        if 'Mallorca' not in ort: ort = f"{ort}, Mallorca"
                        
                        results.append({
                            'titel': item['title'][:100] or 'APTS Property',
                            'quelle': 'A Place in the Sun',
                            'url': item['url'],
                            'preis': parse_price(item['price']),
                            'zimmer': beds,
                            'grundstueck': None, 'wohnflaeche': None,
                            'ort': ort
                        })
                
                print(f"  {base_url.split('/')[-1] or 'mallorca'} page {pg}: {found} new, {len(results)} total")
                if found == 0: break
                    
            except Exception as e:
                print(f"  Error page {pg}: {e}")
                break
    
    print(f"  APTS total: {len(results)}")
    return results


# =============================================
# Yaencontre - Parse __INITIAL_STATE__ base64
# =============================================
def scrape_yaencontre():
    print("\n=== Yaencontré (INITIAL_STATE) ===")
    results = []
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml',
        'Accept-Language': 'es-ES,es;q=0.9',
    }
    
    search_urls = [
        "https://www.yaencontre.com/venta/casas/costa-mallorca",
        "https://www.yaencontre.com/venta/chalets/mallorca",
        "https://www.yaencontre.com/venta/fincas/mallorca",
        "https://www.yaencontre.com/venta/casas/palma-de-mallorca",
    ]
    
    seen = set()
    
    for base_url in search_urls:
        for pg in range(1, 10):
            url = base_url + (f"/{pg}" if pg > 1 else "")
            try:
                r = requests.get(url, headers=headers, timeout=20)
                if r.status_code != 200:
                    print(f"  {url}: {r.status_code}")
                    break
                
                # Find the base64 encoded state
                b64_m = re.search(r'window\.__INITIAL_STATE__\s*=\s*JSON\.parse\(atob\("([^"]+)"\)', r.text)
                if not b64_m:
                    print(f"  No __INITIAL_STATE__ at {url}")
                    break
                
                encoded = b64_m.group(1)
                decoded = base64.b64decode(encoded).decode('utf-8')
                state = json.loads(decoded)
                
                # Navigate to the results
                results_data = state.get('results', {})
                
                # Try to find the items list
                items = None
                if isinstance(results_data, dict):
                    # Try common keys
                    for key in ['results', 'items', 'list', 'properties', 'data']:
                        if key in results_data:
                            candidate = results_data[key]
                            if isinstance(candidate, list) and len(candidate) > 0:
                                items = candidate
                                break
                
                if not items:
                    # Do a deep search
                    def find_items_deep(d, depth=0):
                        if depth > 6: return None
                        if isinstance(d, list) and len(d) > 0 and isinstance(d[0], dict):
                            first = d[0]
                            if any(k in first for k in ['priceValue', 'price', 'urlDetail', 'url']):
                                return d
                        if isinstance(d, dict):
                            for k, v in d.items():
                                found = find_items_deep(v, depth+1)
                                if found: return found
                        return None
                    items = find_items_deep(results_data)
                
                if not items:
                    # Try the full state
                    items = None
                    def find_items_full(d, depth=0):
                        if depth > 8: return None
                        if isinstance(d, list) and len(d) > 0 and isinstance(d[0], dict):
                            first = d[0]
                            if any(k in first for k in ['priceValue', 'price', 'urlDetail', 'url', 'propertyUrl']):
                                return d
                        if isinstance(d, dict):
                            for k, v in d.items():
                                found = find_items_full(v, depth+1)
                                if found: return found
                        return None
                    items = find_items_full(state)
                
                if not items:
                    print(f"  No items found in state for {url}")
                    print(f"  results keys: {list(results_data.keys())[:10] if isinstance(results_data, dict) else type(results_data)}")
                    break
                
                found = 0
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    
                    # Extract URL
                    item_url = (item.get('urlDetail', '') or item.get('url', '') or 
                                item.get('propertyUrl', '') or item.get('link', ''))
                    if not item_url:
                        continue
                    if item_url.startswith('/'):
                        item_url = 'https://www.yaencontre.com' + item_url
                    
                    if item_url in seen:
                        continue
                    seen.add(item_url)
                    found += 1
                    
                    # Extract details
                    price = item.get('priceValue', item.get('price', item.get('amount', None)))
                    if isinstance(price, dict):
                        price = price.get('value', price.get('amount', None))
                    
                    rooms = item.get('rooms', item.get('bedrooms', item.get('bathrooms', None)))
                    surface = item.get('surface', item.get('squareMeters', item.get('area', None)))
                    
                    title = (item.get('title', '') or item.get('description', '') or 
                             item.get('name', '') or 'Yaencontre Property')
                    
                    address = item.get('address', {}) or {}
                    if isinstance(address, dict):
                        ort = (address.get('municipalityName', '') or address.get('city', '') or 
                               address.get('district', '') or 'Mallorca')
                    else:
                        ort = str(address) or 'Mallorca'
                    
                    results.append({
                        'titel': str(title)[:100],
                        'quelle': 'Yaencontré',
                        'url': item_url,
                        'preis': float(price) if price and str(price).replace('.','').isdigit() else None,
                        'zimmer': int(rooms) if rooms and str(rooms).isdigit() else None,
                        'grundstueck': None,
                        'wohnflaeche': float(surface) if surface and str(surface).replace('.','').isdigit() else None,
                        'ort': str(ort)
                    })
                
                print(f"  {base_url.split('/')[-1]} page {pg}: {found} new, {len(results)} total")
                
                # Check pagination
                pagination = state.get('pagination', {})
                if isinstance(pagination, dict):
                    total_pages = pagination.get('totalPages', pagination.get('pages', 1))
                    current_page = pagination.get('currentPage', pagination.get('page', 1))
                    if pg >= total_pages:
                        break
                elif found == 0:
                    break
                    
            except Exception as e:
                print(f"  Error on {url}: {e}")
                break
        
        time.sleep(1.5)  # Be polite
    
    print(f"  Yaencontré total: {len(results)}")
    return results


# =============================================
# Rightmove - Intercept XHR with Playwright
# =============================================
def scrape_rightmove(page):
    print("\n=== Rightmove (Final) ===")
    results = []
    seen = set()
    
    # Capture API/XHR responses
    xhr_data = []
    def on_response(r):
        if r.status == 200 and 'rightmove' in r.url:
            ct = r.headers.get('content-type', '')
            if 'json' in ct:
                try:
                    data = r.json()
                    if isinstance(data, dict) and ('properties' in data or 'propertyListings' in data or 'results' in data):
                        xhr_data.append({'url': r.url, 'data': data})
                        print(f"  XHR: {r.url[:80]}")
                except: pass
    
    page.on('response', on_response)
    
    try:
        page.goto("https://www.rightmove.co.uk/overseas-property/in-Mallorca.html",
                  wait_until='domcontentloaded', timeout=30000)
        time.sleep(5)
        
        # Accept cookies
        try:
            btn = page.query_selector('#onetrust-accept-btn-handler, button:has-text("Accept all")')
            if btn: btn.click(); time.sleep(1)
        except: pass
        
        # Scroll to trigger loading
        for _ in range(3):
            page.evaluate("window.scrollBy(0, 600)")
            time.sleep(1)
        
        print(f"  XHR responses: {len(xhr_data)}")
        
        # Parse XHR data
        for xd in xhr_data:
            data = xd['data']
            props = data.get('properties', data.get('propertyListings', data.get('results', [])))
            if isinstance(props, list):
                for prop in props:
                    if not isinstance(prop, dict): continue
                    u = prop.get('propertyUrl', prop.get('url', ''))
                    if u and not u.startswith('http'):
                        u = 'https://www.rightmove.co.uk' + u
                    if not u or u in seen: continue
                    seen.add(u)
                    
                    price_info = prop.get('price', {})
                    if isinstance(price_info, dict):
                        price = price_info.get('amount', price_info.get('value', price_info.get('displayPrice', 0)))
                    else:
                        price = price_info
                    
                    results.append({
                        'titel': (prop.get('displayAddress', '') or prop.get('summary', ''))[:100] or 'Rightmove Property',
                        'quelle': 'Rightmove',
                        'url': u,
                        'preis': float(price) if price and str(price).replace('.','').isdigit() else None,
                        'zimmer': prop.get('bedrooms'),
                        'grundstueck': None, 'wohnflaeche': None,
                        'ort': prop.get('displayAddress', 'Mallorca')
                    })
        
        # Try page source extraction
        if not results:
            src = page.content()
            
            # Try to find property data in React state
            json_matches = re.findall(r'\{"properties":\[.+?\](?:,"[^"]+":)', src, re.DOTALL)
            if json_matches:
                for m in json_matches[:1]:
                    try:
                        # Try to parse the JSON fragment
                        data = json.loads(m + '{}')
                        print(f"  Found properties JSON: {len(data.get('properties', []))}")
                    except: pass
            
            # Try DOM
            dom_links = page.evaluate("""() => {
                const results = [];
                const seen = new Set();
                document.querySelectorAll('a[href*="/properties/"]').forEach(a => {
                    const href = a.href;
                    if (!seen.has(href) && /\\/properties\\/\\d+/.test(href)) {
                        seen.add(href);
                        const card = a.closest('[class*="card"]') || a.closest('li') || a.closest('article');
                        const priceEl = card ? card.querySelector('[class*="price"]') : null;
                        const titleEl = card ? card.querySelector('h2, h3, [class*="title"], [class*="address"]') : null;
                        results.push({
                            url: href,
                            price: priceEl ? priceEl.innerText : '',
                            title: titleEl ? titleEl.innerText.trim() : ''
                        });
                    }
                });
                return results;
            }""")
            
            print(f"  DOM links found: {len(dom_links)}")
            for item in dom_links:
                if item['url'] not in seen:
                    seen.add(item['url'])
                    results.append({
                        'titel': item['title'][:100] or 'Rightmove Property',
                        'quelle': 'Rightmove',
                        'url': item['url'],
                        'preis': parse_price(item['price']),
                        'zimmer': None, 'grundstueck': None, 'wohnflaeche': None,
                        'ort': 'Mallorca'
                    })
            
            # Try multiple pages
            for pg_idx in range(1, 20):
                pg_url = f"https://www.rightmove.co.uk/overseas-property/in-Mallorca.html?index={pg_idx * 24}"
                page.goto(pg_url, wait_until='domcontentloaded', timeout=20000)
                time.sleep(3)
                
                pg_links = page.evaluate("""() => {
                    const results = [];
                    const seen = new Set();
                    document.querySelectorAll('a[href*="/properties/"]').forEach(a => {
                        const href = a.href;
                        if (!seen.has(href) && /\\/properties\\/\\d+/.test(href)) {
                            seen.add(href);
                            results.push(href);
                        }
                    });
                    return results;
                }""")
                
                new_found = sum(1 for u in pg_links if u not in seen)
                for u in pg_links:
                    if u not in seen:
                        seen.add(u)
                        results.append({
                            'titel': 'Rightmove Property',
                            'quelle': 'Rightmove',
                            'url': u,
                            'preis': None, 'zimmer': None,
                            'grundstueck': None, 'wohnflaeche': None,
                            'ort': 'Mallorca'
                        })
                
                print(f"  Page {pg_idx+1}: {new_found} new links, total {len(results)}")
                if new_found == 0: break
    
    except Exception as e:
        print(f"  Error: {e}")
    
    print(f"  Rightmove total: {len(results)}")
    return results


# =============================================
# Fotocasa - Use API endpoint interception
# =============================================
def scrape_fotocasa(page):
    print("\n=== Fotocasa (Final) ===")
    results = []
    seen = set()
    
    api_data = []
    def on_response(r):
        url = r.url
        if 'fotocasa' in url and r.status == 200:
            ct = r.headers.get('content-type', '')
            if 'json' in ct:
                try:
                    data = r.json()
                    if isinstance(data, dict):
                        if any(k in data for k in ['realEstates', 'items', 'content', 'properties']):
                            api_data.append({'url': url, 'data': data})
                            print(f"  API: {url[:100]} keys: {list(data.keys())[:5]}")
                except: pass
    
    page.on('response', on_response)
    
    try:
        page.goto("https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l",
                  wait_until='domcontentloaded', timeout=30000)
        time.sleep(5)
        
        # Click accept cookies
        try:
            btns = page.query_selector_all('button')
            for btn in btns:
                text = btn.inner_text() if btn else ''
                if text and re.search(r'aceptar|accept|ok', text, re.I):
                    btn.click()
                    time.sleep(1)
                    break
        except: pass
        
        # Scroll to trigger API calls
        for _ in range(4):
            page.evaluate("window.scrollBy(0, 800)")
            time.sleep(1.5)
        
        print(f"  API data captured: {len(api_data)}")
        
        # Parse API data
        for xd in api_data:
            data = xd['data']
            items_list = (data.get('realEstates') or data.get('items') or 
                          data.get('content') or data.get('properties', []))
            
            if isinstance(items_list, list):
                for item in items_list:
                    if not isinstance(item, dict): continue
                    
                    url = item.get('url', '') or item.get('link', '') or item.get('detailUrl', '')
                    if url and not url.startswith('http'):
                        url = 'https://www.fotocasa.es' + url
                    if not url: continue
                    if url in seen: continue
                    seen.add(url)
                    
                    price = item.get('price', {})
                    if isinstance(price, dict):
                        price = price.get('value', price.get('amount', price.get('displayPrice', None)))
                    
                    rooms = item.get('rooms', item.get('bedrooms', None))
                    surface = item.get('surface', item.get('squareMeters', item.get('area', None)))
                    
                    address = item.get('address', {}) or {}
                    ort = 'Mallorca'
                    if isinstance(address, dict):
                        ort = (address.get('municipalityName', '') or address.get('city', '') or 
                               address.get('town', '') or 'Mallorca')
                    
                    title = item.get('title', '') or item.get('description', '') or 'Fotocasa Property'
                    
                    results.append({
                        'titel': str(title)[:100],
                        'quelle': 'Fotocasa',
                        'url': url,
                        'preis': float(price) if price else None,
                        'zimmer': int(rooms) if rooms else None,
                        'grundstueck': None,
                        'wohnflaeche': float(surface) if surface else None,
                        'ort': str(ort)
                    })
        
        # Multiple pages
        for pg in range(2, 20):
            pg_url = f"https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l/{pg}"
            try:
                page.goto(pg_url, wait_until='domcontentloaded', timeout=20000)
                time.sleep(3)
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(2)
                
                # Process any new API data
                prev_count = len(results)
                for xd in api_data:
                    # Already processed above
                    pass
                
                # Get new API data from this page
                new_api = [x for x in api_data if pg_url.split('/')[-1] in x.get('url', '') or True]
                
                # Try DOM
                dom_data = page.evaluate("""() => {
                    const seen = new Set();
                    const results = [];
                    document.querySelectorAll('a[href*="/comprar/"]').forEach(a => {
                        const href = a.href;
                        if (/\\/comprar\\/.+\\d{5,}/.test(href) && !seen.has(href)) {
                            seen.add(href);
                            const card = a.closest('article') || a.closest('[class*="card"]') || a.parentElement;
                            const priceEl = card ? card.querySelector('[class*="price"], [class*="Price"]') : null;
                            const roomsEl = card ? card.querySelector('[class*="rooms"], [class*="Rooms"]') : null;
                            const titleEl = card ? card.querySelector('h2, h3, [class*="title"]') : null;
                            results.push({
                                url: href,
                                price: priceEl ? priceEl.innerText : '',
                                rooms: roomsEl ? roomsEl.innerText : '',
                                title: titleEl ? titleEl.innerText.trim() : ''
                            });
                        }
                    });
                    return results;
                }""")
                
                new_found = 0
                for item in dom_data:
                    if item['url'] not in seen:
                        seen.add(item['url'])
                        new_found += 1
                        results.append({
                            'titel': item['title'][:100] or 'Fotocasa Property',
                            'quelle': 'Fotocasa',
                            'url': item['url'],
                            'preis': parse_price(item['price']),
                            'zimmer': int(re.search(r'(\d+)', item['rooms']).group(1)) if re.search(r'(\d+)', item['rooms']) else None,
                            'grundstueck': None, 'wohnflaeche': None, 'ort': 'Mallorca'
                        })
                
                print(f"  Page {pg}: {new_found} new DOM links, total {len(results)}")
                if new_found == 0: break
                    
            except Exception as e:
                print(f"  Page {pg} error: {e}")
                break
    
    except Exception as e:
        print(f"  Error: {e}")
    
    print(f"  Fotocasa total: {len(results)}")
    return results


# =============================================
# MAIN
# =============================================
def main():
    print("Loading Excel...")
    wb, ws, existing_urls = load_existing_data()
    print(f"Existing: {len(existing_urls)}")
    
    summary = {}
    total_added = 0
    
    # Yaencontre (requests-based)
    try:
        r = scrape_yaencontre()
        added = save_objects(wb, ws, r, existing_urls) if r else 0
        summary['Yaencontré'] = {'scraped': len(r), 'added': added}
        total_added += added
        print(f"  ✓ Yaencontré: {added} added")
    except Exception as e:
        summary['Yaencontré'] = {'scraped': 0, 'added': 0, 'error': str(e)[:100]}
        print(f"  ✗ Yaencontré: {e}")
    
    # Playwright scrapers
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-dev-shm-usage'])
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
            locale='en-GB',
            viewport={'width': 1280, 'height': 900}
        )
        page = context.new_page()
        
        for name, fn in [('A Place in the Sun', scrape_aplaceinthesun), ('Rightmove', scrape_rightmove), ('Fotocasa', scrape_fotocasa)]:
            try:
                r = fn(page)
                added = save_objects(wb, ws, r, existing_urls) if r else 0
                summary[name] = {'scraped': len(r), 'added': added}
                total_added += added
                print(f"\n  ✓ {name}: {len(r)} scraped, {added} added")
            except Exception as e:
                summary[name] = {'scraped': 0, 'added': 0, 'error': str(e)[:100]}
                print(f"\n  ✗ {name}: {e}")
        
        browser.close()
    
    print(f"\n{'='*50}")
    print("FINAL PASS SUMMARY")
    for name, s in summary.items():
        print(f"  {name}: {s.get('scraped',0)} → {s.get('added',0)}")
        if 'error' in s: print(f"    ERR: {s['error']}")
    print(f"\nTotal new: {total_added}")
    
    with open('/Users/robin/.openclaw/workspace/mallorca-projekt/phase1_final_results.json', 'w') as f:
        json.dump({'summary': summary, 'total_added': total_added}, f, indent=2)

if __name__ == '__main__':
    main()
