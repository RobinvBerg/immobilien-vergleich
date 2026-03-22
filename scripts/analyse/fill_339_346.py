#!/usr/bin/env python3
"""Fill missing fields for Ordnungsnummern 339-346."""

import openpyxl

XLSX = '/Users/robin/.openclaw/workspace/mallorca-projekt/data/mallorca-kandidaten-v2.xlsx'

# Pre-fetched Google Maps distances [km, min]
DISTANCES = {
    339: {"flughafen": (48.4, 41), "daia": (59.4, 61), "andratx": (73.0, 58), "ses_salines": (49.4, 52)},
    340: {"flughafen": (45.7, 39), "daia": (78.6, 73), "andratx": (85.8, 67), "ses_salines": (9.9, 15)},
    341: {"flughafen": (33.1, 27), "daia": (44.1, 48), "andratx": (57.7, 44), "ses_salines": (76.4, 61)},
    342: {"flughafen": (35.3, 33), "daia": (46.3, 53), "andratx": (59.9, 50), "ses_salines": (48.8, 55)},
    343: {"flughafen": (45.0, 40), "daia": (77.9, 73), "andratx": (85.1, 68), "ses_salines": (27.2, 31)},
    344: {"flughafen": (23.7, 21), "daia": (34.7, 41), "andratx": (48.3, 38), "ses_salines": (67.0, 55)},
    345: {"flughafen": (16.1, 20), "daia": (27.8, 37), "andratx": (37.7, 34), "ses_salines": (59.3, 54)},
    346: {"flughafen": (28.5, 27), "daia": (57.4, 58), "andratx": (64.6, 52), "ses_salines": (31.6, 33)},
}

# Vision analysis results (from image analysis)
VISION = {
    339: {"charme": 4, "renovierung": 75, "reno_begruendung": "Guter Renovierungszustand mit authentischen Holzbalkendecken und Marmorboden, leicht datierter Einrichtungsstil.", "gaestehaueser": 0},
    340: {"charme": 5, "renovierung": 15, "reno_begruendung": "Finca-Ruine mit eingestürztem Dach und verfallenen Natursteinmauern – nahezu vollständiger Renovierungsbedarf.", "gaestehaueser": 2},
    341: {"charme": 3, "renovierung": 20, "reno_begruendung": "Finca-Projekt / Rohbau – kein fertiggestelltes Gebäude erkennbar, vollständige Renovierung/Fertigstellung erforderlich.", "gaestehaueser": 1},
    342: {"charme": 4, "renovierung": 95, "reno_begruendung": "Moderne Luxusfinca, neuwertig gebaut mit hochwertiger Ausstattung und beheiztem Pool.", "gaestehaueser": 0},
    343: {"charme": 3, "renovierung": 95, "reno_begruendung": "Kompletter Neubau im traditionellen Stil mit Natursteinfassade, makelloser Zustand.", "gaestehaueser": 0},
    344: {"charme": 5, "renovierung": 90, "reno_begruendung": "Herrschaftliche Finca mit Steinbögen und blauen Fensterläden, hochwertig renoviert mit modernem Pool und Solarpanels.", "gaestehaueser": 1},
    345: {"charme": 4, "renovierung": 35, "reno_begruendung": "Historische Possessió mit Wehrturm, deutlicher Renovierungsbedarf an Fassade, Pool und Außenanlagen.", "gaestehaueser": 1},
    346: {"charme": 5, "renovierung": 10, "reno_begruendung": "Monumentales historisches Landgut nahe Ruinenzustand – bröckelnde Fassade, instabile Dächer, kein moderner Standard erkennbar.", "gaestehaueser": 2},
}

# Object data (existing + estimated bebaut where missing)
OBJECTS = {
    339: {"name": "Llubí", "grundstueck": 22033, "bebaut": 22033, "zimmer": 8, "preis": 3995000, "bewirtschaftung": 3, "vermietlizenz": 0},
    340: {"name": "Santanyí", "grundstueck": 947109, "bebaut": 1200, "zimmer": 15, "preis": 10000000, "bewirtschaftung": 1, "vermietlizenz": 0},
    341: {"name": "Alaró", "grundstueck": 125576, "bebaut": 350, "zimmer": 5, "preis": 2100000, "bewirtschaftung": 2, "vermietlizenz": 0},
    342: {"name": "Biniali", "grundstueck": 15000, "bebaut": 400, "zimmer": 5, "preis": 4250000, "bewirtschaftung": 4, "vermietlizenz": 0},
    343: {"name": "Felanitx", "grundstueck": 15883, "bebaut": 350, "zimmer": 5, "preis": 3500000, "bewirtschaftung": 4, "vermietlizenz": 0},
    344: {"name": "Santa Maria", "grundstueck": 30100, "bebaut": 500, "zimmer": 7, "preis": 6500000, "bewirtschaftung": 3, "vermietlizenz": 0},
    345: {"name": "Establiments", "grundstueck": 85637, "bebaut": 450, "zimmer": 7, "preis": 3900000, "bewirtschaftung": 2, "vermietlizenz": 0},
    346: {"name": "Algaida", "grundstueck": 728851, "bebaut": 900, "zimmer": 10, "preis": 5000000, "bewirtschaftung": 2, "vermietlizenz": 0},
}


def calc_erreichbarkeit(flughafen_min, daia_min, ses_min, andratx_min):
    def score_for(mins, ideal, akzeptabel, dealbreaker):
        if mins is None:
            return 50
        if mins <= ideal:
            return 100
        elif mins <= akzeptabel:
            return round(100 - 40 * (mins - ideal) / (akzeptabel - ideal))
        elif mins <= dealbreaker:
            return round(60 - 60 * (mins - akzeptabel) / (dealbreaker - akzeptabel))
        else:
            return 0

    s_flughafen = score_for(flughafen_min, 15, 25, 40)
    s_daia = score_for(daia_min, 20, 40, 70)
    s_ses = score_for(ses_min, 15, 30, 45)
    s_andratx = score_for(andratx_min, 25, 40, 60)
    return round(s_flughafen * 0.30 + s_daia * 0.30 + s_ses * 0.30 + s_andratx * 0.10)


def calc_score(obj, charme, renovierung, bewirtschaftung, vermietlizenz, erreichbarkeit, gaestehaueser):
    zimmer = obj["zimmer"]
    grundstueck = obj["grundstueck"]
    bebaut = obj["bebaut"]
    preis = obj["preis"]

    zimmer_score = min(100, max(0, (zimmer - 5) / (8 - 5) * 100))

    eur_m2 = preis / bebaut if bebaut else preis / grundstueck
    preis_score = max(0, min(100, 100 - (eur_m2 - 5000) / 150))

    gast_score = (gaestehaueser or 0) * 50

    erreich_score = erreichbarkeit or 0

    garten = max(0, grundstueck - bebaut)
    garten_score = min(100, max(0, (garten - 1000) / (5000 - 1000) * 100))

    vmiet_score = vermietlizenz or 0

    bewirt_score = ((bewirtschaftung or 3) - 1) / 4 * 100

    charme_score = ((charme or 3) - 1) / 4 * 100

    reno_score = renovierung or 60

    score = (
        zimmer_score * 0.20 +
        preis_score * 0.15 +
        gast_score * 0.15 +
        erreich_score * 0.15 +
        garten_score * 0.10 +
        vmiet_score * 0.05 +
        bewirt_score * 0.05 +
        charme_score * 0.10 +
        reno_score * 0.05
    )
    return round(score, 2)


# Load workbook
wb = openpyxl.load_workbook(XLSX)
ws = wb.active

# Find rows for 339-346
row_map = {}
for row in ws.iter_rows(min_row=2):
    nr = row[0].value
    if nr in range(339, 347):
        row_map[nr] = row[0].row

print(f"Found rows: {row_map}")

# Column map
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
col = {h: i + 1 for i, h in enumerate(headers)}
print(f"Key columns: {[(k, col[k]) for k in ['Charme/Ästhetik (1-5)', 'Entfernung Flughafen (km)', 'Score (0-100)', 'Gästehäuser (0/1/2)']]}")

for nr in range(339, 347):
    row_idx = row_map.get(nr)
    if not row_idx:
        print(f"Row {nr} not found!")
        continue

    obj = OBJECTS[nr]
    dists = DISTANCES[nr]
    vision = VISION[nr]

    flughafen_km, flughafen_min = dists["flughafen"]
    daia_km, daia_min = dists["daia"]
    andratx_km, andratx_min = dists["andratx"]
    ses_km, ses_min = dists["ses_salines"]

    # Distances
    ws.cell(row_idx, col["Entfernung Flughafen (km)"]).value = flughafen_km
    ws.cell(row_idx, col["Entfernung Flughafen (min)"]).value = flughafen_min
    ws.cell(row_idx, col["Entfernung Daia Haus (km)"]).value = daia_km
    ws.cell(row_idx, col["Entfernung Daia Haus (min)"]).value = daia_min
    ws.cell(row_idx, col["Entfernung Andratx (km)"]).value = andratx_km
    ws.cell(row_idx, col["Entfernung Andratx (min)"]).value = andratx_min
    ws.cell(row_idx, col["Entfernung Ses Salines (km)"]).value = ses_km
    ws.cell(row_idx, col["Entfernung Ses Salines (min)"]).value = ses_min

    # Bebaute Fläche (only if not already set)
    existing_bebaut = ws.cell(row_idx, col["Bebaute Fläche (m²)"]).value
    if not existing_bebaut:
        ws.cell(row_idx, col["Bebaute Fläche (m²)"]).value = obj["bebaut"]

    # Garten
    bebaut_actual = ws.cell(row_idx, col["Bebaute Fläche (m²)"]).value or obj["bebaut"]
    garten = max(0, obj["grundstueck"] - bebaut_actual)
    ws.cell(row_idx, col["Garten zu bewirtschaften (m²)"]).value = garten

    # €/m²
    if bebaut_actual:
        ws.cell(row_idx, col["€/m² (bebaut)"]).value = round(obj["preis"] / bebaut_actual)
    ws.cell(row_idx, col["€/m² (Grundstück)"]).value = round(obj["preis"] / obj["grundstueck"])

    # Vision data
    charme = vision["charme"]
    renovierung = vision["renovierung"]
    gaestehaueser = vision["gaestehaueser"]
    ws.cell(row_idx, col["Charme/Ästhetik (1-5)"]).value = charme
    ws.cell(row_idx, col["Renovierung (0-100)"]).value = renovierung
    ws.cell(row_idx, col["Reno-Score (0-100)"]).value = renovierung
    ws.cell(row_idx, col["Reno-Begründung"]).value = vision["reno_begruendung"]
    ws.cell(row_idx, col["Gästehäuser (0/1/2)"]).value = gaestehaueser

    # Bewirtschaftung & Vermietlizenz
    bewirtschaftung = obj["bewirtschaftung"]
    vermietlizenz = obj["vermietlizenz"]
    ws.cell(row_idx, col["Bewirtschaftung (1-5, 5=pflegeleicht)"]).value = bewirtschaftung
    ws.cell(row_idx, col["Vermietlizenz (100/50/0)"]).value = vermietlizenz

    # Erreichbarkeit
    erreichbarkeit = calc_erreichbarkeit(flughafen_min, daia_min, ses_min, andratx_min)
    ws.cell(row_idx, col["Erreichbarkeit (0-100)"]).value = erreichbarkeit

    # Score
    score = calc_score(obj, charme, renovierung, bewirtschaftung, vermietlizenz, erreichbarkeit, gaestehaueser)
    ws.cell(row_idx, col["Score (0-100)"]).value = score

    print(f"{nr} ({obj['name']}): Erreichbarkeit={erreichbarkeit}, Score={score}, Garten={garten}")

wb.save(XLSX)
print("\nDone! Saved to Excel.")
