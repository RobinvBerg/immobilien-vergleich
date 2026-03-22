from openpyxl import load_workbook
from datetime import date
import re

wb = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
ws = wb['Mallorca Objekte']
existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
print(f"Start: {len(existing_urls)} URLs bekannt, {ws.max_row} Zeilen")

def extract_price(desc):
    if not desc: return None
    m = re.search(r'([\d.,]+)\s*EUR', desc)
    if m:
        try: return int(m.group(1).replace('.', '').replace(',', ''))
        except: pass
    return None

def normalize_url(url):
    """Convert /at/, /ch/, /it/, /en/ variants to /de/"""
    return re.sub(r'https://www\.von-poll\.com/(at|ch|it|en)/', 
                  'https://www.von-poll.com/de/', url)

# Alle gesammelten Von Poll Exposes aus den Web-Suchen
entries = [
    # titel, url, preis, ort
    ("Luxuriöse Villa mit zwei Pools und Meerblick Son Vida",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/luxurise-villa-mit-zwei-pools-und-meerblick-in-exklusivem-son-vida-mallorca-4160031513",
     9900000, "Son Vida, Mallorca"),
    ("Villa Son Vida – Moderne Villa mit Infinity-Pool",
     "https://www.von-poll.com/de/expose/mallorca-palma/moderne-villa-mit-infinity-pool-in-son-vida-20001795",
     None, "Son Vida, Palma"),
    ("Spektakuläre Wohnung am Meer in Portixol, Mallorca",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/4160036805",
     None, "Portixol, Mallorca"),
    ("Neubauprojekt Luxusvilla mit Pool Son Vida",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/neubauprojekt-fr-eine-luxusvilla-mit-pool-und-traumblick-in-son-vida-mallorca-4160034665",
     2900000, "Son Vida, Mallorca"),
    ("Einmalige Luxus-Finca weitläufigem Grundstück Mallorca",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/einmalige-luxus-finca-und-weitlaufigem-grundstuck-4160045157",
     None, "Mallorca"),
    ("Authentische rustikale Finca Artá, Mallorca",
     "https://www.von-poll.com/de/expose/mallorca-pollensa/authentische-rustikale-finca-auf-einem-groen-grundstck-zum-renovieren-in-art-mallorca-3520033555",
     900000, "Artá, Mallorca"),
    ("Attraktives Dorfhaus Bunyola mit Panoramablick und Pool",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/attraktives-dorfhaus-in-bunyola-mit-panoramablick-und-pool-4160041027",
     None, "Bunyola, Mallorca"),
    ("Traumhafte traditionelle Finca Santa Maria/Bunyola",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/traumhafte-traditionelle-finca-zwischen-santa-maria-und-bunyola-mit-herrlichem-weitblick-bis-zur-bucht-von-palma-4160046911",
     None, "Santa Maria, Mallorca"),
    ("Exklusives Penthouse Paseo Marítimo Palma",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/exklusives-penthouse-am-paseo-maritimo-wohnen-auf-hochstem-niveau-in-palma-4160046539",
     None, "Palma, Mallorca"),
    ("Exklusives Apartment 3 Etagen Paseo Marítimo",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/exclusive-luxury-residences-on-the-paseo-maritimo-living-at-the-highest-level-in-palma-4160047615",
     None, "Palma, Mallorca"),
    ("Finca mit Gästehaus und Tennisplatz nahe Inca",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/finca-mit-gstehaus-und-tennisplatz-nahe-inca-4160037753",
     None, "Inca, Mallorca"),
    ("Exklusive Luxusresidenzen Paseo Marítimo Palma",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/exclusive-luxury-residences-on-the-paseo-maritimo-living-at-the-highest-level-in-palma-4160047801",
     None, "Palma, Mallorca"),
    ("Eine außergewöhnliche Villa mit Panoramablick Esporles",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/eine-auergewhnliche-villa-mit-panoramablick-in-esporles-4160039711",
     6950000, "Esporles, Mallorca"),
    ("Moderne Villa fantastischer Blick Bucht Santa Ponsa",
     "https://www.von-poll.com/de/expose/mallorca-palma/modern-villa-with-fantastic-views-to-the-bay-of-santa-ponsa-20009911",
     None, "Santa Ponsa, Mallorca"),
    ("Fantastische Finca mit Wein und Oliven zum Renovieren",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/fantastische-finca-mit-wein-und-oliven-zum-renovieren-4160029173",
     None, "Mallorca"),
    ("Elegante Finca privat gelegen Meerblick Manacor",
     "https://www.von-poll.com/de/expose/mallorca-portals/elegante-finca-privat-gelegen-mit-weitblick-ueber-die-landschaft-bis-zum-meer-290005169",
     1900000, "Manacor, Mallorca"),
    ("Herrschaftliche Finca Valldemossa 17. Jahrhundert",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/herrschaftliche-finca-in-valldemossa-aus-dem-17-jahrhundert-4160019929",
     23000000, "Valldemossa, Mallorca"),
    ("Große Finca mit Olivenhain und Weinberg Inca",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/4160037333",
     None, "Inca, Mallorca"),
    ("Spektakuläre Villa direkt am Meer Barcarés Alcudia",
     "https://www.von-poll.com/de/expose/mallorca-pollensa/3520046907",
     None, "Alcudia, Mallorca"),
    ("Moderne neu gebaute Villa Pool Meerblick Puerto de Pollensa",
     "https://www.von-poll.com/de/expose/mallorca-pollensa/modern-newly-built-villa-with-swimming-pool-and-sea-views-in-puerto-de-pollensa-3520032557",
     None, "Puerto de Pollensa, Mallorca"),
    ("Moderne neu gebaute Villa großem Pool Alcudia",
     "https://www.von-poll.com/de/expose/mallorca-pollensa/modern-newly-built-villa-with-large-swimming-pool-for-sale-in-alcudia-mallorca-3520030913",
     None, "Alcudia, Mallorca"),
    ("Fantastische Villa Pool Garten Ferienvermietungslizenz Puerto Pollensa",
     "https://www.von-poll.com/de/expose/mallorca-pollensa/fantastic-villa-with-swimming-pool-garden-and-holiday-tourist-licence-in-puerto-pollensa-3520028691",
     None, "Puerto Pollensa, Mallorca"),
    ("Spektakuläre Villa Pool Son Bauló Can Picafort",
     "https://www.von-poll.com/de/expose/mallorca-pollensa/spectacular-villa-with-pool-for-sale-in-the-urbanisation-son-baul-in-can-picafort-3520031217",
     None, "Can Picafort, Mallorca"),
    ("Moderne Villa Pool Tourist License Colonia de Sant Pere",
     "https://www.von-poll.com/de/expose/mallorca-pollensa/3520044483",
     None, "Colonia de Sant Pere, Mallorca"),
    ("Stilvolle Villa Pool tropischer Garten Meerblick Bonaire-Alcudia",
     "https://www.von-poll.com/de/expose/mallorca-pollensa/3520047733",
     None, "Bonaire-Alcudia, Mallorca"),
    ("Neu erbautes Landhaus Pollensa-Alcudia Pool ETV Vermietlizenz",
     "https://www.von-poll.com/de/expose/mallorca-pollensa/wunderschnes-neu-erbautes-landhaus-zwischen-pollensa-und-alcudia-mit-pool-und-etv-vermietlizenz-3520037587",
     None, "Pollensa-Alcudia, Mallorca"),
    ("Wohnung mit Meerblick Playa de Palma",
     "https://www.von-poll.com/de/expose/mallorca-llucmajor/wohnung-mit-meerblick-direkt-an-der-playa-de-palma-299007921",
     425000, "Playa de Palma, Mallorca"),
    ("Neues luxuriöses Wohnprojekt Cala Mayor Palma",
     "https://www.von-poll.com/de/expose/mallorca-palma/neues-luxurioeses-wohnprojekt-in-cala-mayor-200010735",
     None, "Cala Mayor, Palma"),
    ("LIVING HIGH EXPERIENCE Panorama-Meerblick Son Vida",
     "https://www.von-poll.com/de/expose/mallorca-palma/living-high-experience-einzigartiges-wohnjuwel-mit-panorama-meerblick-in-son-vida-20009131",
     None, "Son Vida, Palma"),
    ("Neukonstruierte Villa bester Lage Son Vida",
     "https://www.von-poll.com/de/expose/mallorca-palma/neukonstruierte-villa-in-bester-lage-von-son-vida-20004111",
     None, "Son Vida, Palma"),
    ("Projekt Luxusvilla Son Vida Palma",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/projekt-zum-bau-einer-luxusvilla-in-exklusiver-lage-von-palma-in-son-vida-4160034673",
     2700000, "Son Vida, Palma"),
    ("Modern luxury villa Son Vida",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/modern-luxury-villa-in-son-vida-4160034683",
     None, "Son Vida, Palma"),
    ("Großartiges Projekt Bau herrliche Villa Pool Son Vida",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/groartiges-projekt-zum-bau-einer-herrlichen-villa-mit-pool-in-son-vida-mallorca-4160034659",
     2600000, "Son Vida, Mallorca"),
    ("Modernes Luxusvilla Projekt Son Vida Palma",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/wunderschnes-projekt-zum-bau-einer-modernen-luxusvilla-in-der-exklusivsten-gegend-von-palma-in-son-vida-4160034679",
     None, "Son Vida, Palma"),
    ("Weitläufiges Grundstück Capdepera Mallorca",
     "https://www.von-poll.com/de/expose/mallorca-palma/solar-grande-con-muchos-posibilidades-en-capdepera-20002167",
     None, "Capdepera, Mallorca"),
    ("Traumhaftes Grundstück mit Bauprojekt Panorama-Weitblick",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/traumhaftes-grundstck-mit-bauprojekt-und-panoramaweitblick-4160041683",
     585000, "Mallorca"),
    ("Letztes Baugrundstück Projekt Meerblick Son Vida Palma",
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/eines-der-letzten-baugrundstcke-mit-projekt-und-meerblick-im-alten-teil-von-son-vida-palma-4160044421",
     None, "Son Vida, Palma"),
]

vp_count = 0
for titel, url, preis, ort in entries:
    # Normalize URL to /de/
    url = normalize_url(url)
    if url in existing_urls:
        print(f"  SKIP (exists): {url[-50:]}")
        continue
    ws.append([str(titel)[:100], 'Von Poll Real Estate', url, preis, None, None, None,
               str(ort)[:100], str(date.today()), 'Neu'])
    existing_urls.add(url)
    vp_count += 1
    print(f"  Added: {titel[:60]}")

print(f"\nVon Poll neu hinzugefügt: {vp_count}")
wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
print(f"✅ Gespeichert. Finale Zeilenzahl: {ws.max_row}")
