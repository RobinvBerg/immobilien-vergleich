import requests
import json, time, re
from openpyxl import load_workbook
from datetime import date

wb = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
ws = wb['Mallorca Objekte']
existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
print(f"Start: {len(existing_urls)} URLs bekannt, {ws.max_row} Zeilen")

# Brave API Key (aus OpenClaw workspace)
BRAVE_KEY = "BSAFHaKRaG_Xq0vt1C73eo7sFLqIpz4"  # Robin's key aus workspace

# Fallback: Suche via OpenClaw's web_search – wir nutzen direkt requests ans Brave API
# Da der Key ungültig ist, nutzen wir die web_fetch Methode zum Scrapen der Von Poll Suchseite

# Stattdessen: Von Poll Listing-Seiten direkt crawlen mit erweiterten Headers
session = requests.Session()
session.headers.update({
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Language': 'de-DE,de;q=0.9,en;q=0.8',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Cache-Control': 'max-age=0',
})

vp_count = 0
all_found_urls = {}

# Von Poll search results via their search/listing pages
listing_pages = [
    'https://www.von-poll.com/de/immobilien/mallorca',
    'https://www.von-poll.com/de/haus-kaufen/mallorca',
    'https://www.von-poll.com/de/finca-kaufen/mallorca',
    'https://www.von-poll.com/de/wohnung-kaufen/mallorca',
    'https://www.von-poll.com/de/villa-kaufen/mallorca',
    'https://www.von-poll.com/de/immobilien-mallorca',
]

from bs4 import BeautifulSoup

for page_url in listing_pages:
    try:
        r = session.get(page_url, timeout=30)
        print(f"URL: {page_url[:70]} -> HTTP {r.status_code}")
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')
            # Suche nach Expose-Links
            links = soup.find_all('a', href=True)
            expose_links = [l for l in links if '/expose/' in l.get('href', '')]
            print(f"  Expose-Links: {len(expose_links)}")
            for link in expose_links:
                url = link.get('href', '')
                if url and not url.startswith('http'):
                    url = 'https://www.von-poll.com' + url
                if url not in existing_urls and url not in all_found_urls:
                    titel = link.get_text(strip=True)[:100] or 'Von Poll Objekt'
                    all_found_urls[url] = titel
        time.sleep(2)
    except Exception as e:
        print(f"Error {page_url}: {e}")

print(f"\nGefunden via direktes Crawling: {len(all_found_urls)}")

# Bekannte URLs aus den Web-Suchen (hart einkodiert aus vorherigen Suchen)
known_results = [
    ("Luxuriöse Villa mit zwei Pools und Meerblick in exklusivem Son Vida, Mallorca", 
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/luxurise-villa-mit-zwei-pools-und-meerblick-in-exklusivem-son-vida-mallorca-4160031513",
     "9900000"),
    ("Villa in Son Vida zu verkaufen – Moderne Villa mit Infinity-Pool", 
     "https://www.von-poll.com/de/expose/mallorca-palma/moderne-villa-mit-infinity-pool-in-son-vida-20001795",
     None),
    ("Spektakuläre Wohnung am Meer in Portixol, Mallorca", 
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/4160036805",
     None),
    ("Neubauprojekt für eine Luxusvilla mit Pool und Traumblick in Son Vida, Mallorca", 
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/neubauprojekt-fr-eine-luxusvilla-mit-pool-und-traumblick-in-son-vida-mallorca-4160034665",
     None),
    ("Einmalige Luxus-Finca und weitläufigem Grundstück", 
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/einmalige-luxus-finca-und-weitlaufigem-grundstuck-4160045157",
     None),
    ("Authentische rustikale Finca in Artá, Mallorca", 
     "https://www.von-poll.com/de/expose/mallorca-pollensa/authentische-rustikale-finca-auf-einem-groen-grundstck-zum-renovieren-in-art-mallorca-3520033555",
     None),
    ("Attraktives Dorfhaus in Bunyola mit Panoramablick und Pool", 
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/attraktives-dorfhaus-in-bunyola-mit-panoramablick-und-pool-4160041027",
     None),
    ("Traumhafte traditionelle Finca zwischen Santa Maria und Bunyola", 
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/traumhafte-traditionelle-finca-zwischen-santa-maria-und-bunyola-mit-herrlichem-weitblick-bis-zur-bucht-von-palma-4160046911",
     None),
    ("Exklusives Penthouse am Paseo Marítimo – Palma", 
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/exklusives-penthouse-am-paseo-maritimo-wohnen-auf-hochstem-niveau-in-palma-4160046539",
     None),
    ("Exklusives Apartment über 3 Etagen – Palma", 
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/exclusive-luxury-residences-on-the-paseo-maritimo-living-at-the-highest-level-in-palma-4160047615",
     None),
    ("Finca mit Gästehaus und Tennisplatz nahe Inca", 
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/finca-mit-gstehaus-und-tennisplatz-nahe-inca-4160037753",
     None),
    ("Exklusive Luxusresidenzen am Paseo Marítimo – Palma", 
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/exclusive-luxury-residences-on-the-paseo-maritimo-living-at-the-highest-level-in-palma-4160047801",
     None),
    ("Eine außergewöhnliche Villa mit Panoramablick in Esporles", 
     "https://www.von-poll.com/de/expose/mallorca-santa-maria/eine-auergewhnliche-villa-mit-panoramablick-in-esporles-4160039711",
     "6950000"),
    ("Moderne Villa mit fantastischem Blick auf die Bucht von Santa Ponsa", 
     "https://www.von-poll.com/de/expose/mallorca-palma/modern-villa-with-fantastic-views-to-the-bay-of-santa-ponsa-20009911",
     None),
]

# Merge with found URLs
for titel, url, preis in known_results:
    if url not in all_found_urls and url not in existing_urls:
        all_found_urls[url] = (titel, preis)

print(f"Gesamt potenzielle Von Poll URLs: {len(all_found_urls)}")

# Hinzufügen zur Excel
for url, info in all_found_urls.items():
    if url in existing_urls:
        continue
    if isinstance(info, tuple):
        titel, preis_str = info
        preis = None
        if preis_str:
            try: preis = int(preis_str)
            except: preis = None
    else:
        titel = info or 'Von Poll Objekt'
        preis = None
    
    ws.append([str(titel)[:100], 'Von Poll Real Estate', url, preis, None, None, None,
               'Mallorca', str(date.today()), 'Neu'])
    existing_urls.add(url)
    vp_count += 1

print(f"\nVon Poll hinzugefügt: {vp_count}")

wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
print(f"✅ Gespeichert. Finale Zeilenzahl: {ws.max_row}")
