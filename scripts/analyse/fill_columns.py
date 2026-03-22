#!/usr/bin/env python3
"""Fill empty columns in mallorca-kandidaten-v2.xlsx for rows 15-333."""

import openpyxl
import json
import re
import math
from copy import copy

XLSX = 'mallorca-kandidaten-v2.xlsx'
BILDER = 'mallorca-bilder.json'

# Column indices (1-based)
COL_D = 4   # Charme/Ästhetik
COL_E = 5   # Zimmer
COL_F = 6   # Bäder
COL_G = 7   # Grundstück m²
COL_H = 8   # Wohnfläche m²
COL_I = 9   # Garten zu bewirtschaften
COL_L = 12  # Flughafen min
COL_S = 19  # Preis
COL_V = 22  # Renovierung
COL_W = 23  # Bewirtschaftung
COL_X = 24  # Vermietlizenz
COL_Y = 25  # PMI-Score (Flughafen PMI Fahrzeit → Score 0-100)
COL_Z = 26  # Score
COL_AA = 27 # SortKey
COL_AB = 28 # Rang
COL_AC = 29 # Gebäudestruktur
COL_AD = 30 # Baujahr
COL_AF = 32 # Reno-Score
COL_AG = 33 # Reno-Begründung
COL_AI = 35 # Beschreibung (source)

def load_bilder():
    with open(BILDER) as f:
        return json.load(f)

def get_desc(ws, row, bilder):
    """Get full description from col AI and bilder.json."""
    desc_cell = ws.cell(row, COL_AI).value or ''
    bilder_entry = bilder.get(str(row), {})
    bilder_desc = bilder_entry.get('desc', '') or ''
    # Combine both
    combined = ' '.join([str(desc_cell), str(bilder_desc)]).lower()
    return combined, str(desc_cell), str(bilder_desc)

def calc_charme(desc_lower):
    """D = Charme/Ästhetik 1-5."""
    high = ['luxury', 'exclusive', 'stunning', 'breathtaking', 'historic', 'unique', 'magnificent',
            'spectacular', 'extraordinary', 'prestigious', 'palatial', 'grand', 'masterpiece']
    mid_high = ['modern', 'new build', 'contemporary', 'elegant', 'stylish', 'beautiful', 'charming']
    low = ['reform needed', 'project', 'potential', 'to renovate', 'para reformar', 'needs renovation',
           'requires renovation', 'total renovation']
    
    score = 3  # default
    high_count = sum(1 for k in high if k in desc_lower)
    low_count = sum(1 for k in low if k in desc_lower)
    mid_count = sum(1 for k in mid_high if k in desc_lower)
    
    if high_count >= 2:
        score = 5
    elif high_count == 1:
        score = 4
    elif mid_count >= 1 and low_count == 0:
        score = 3  # could be 3-4, use 3 as safe
        if mid_count >= 2:
            score = 4
    
    if low_count >= 2:
        score = max(1, score - 2)
    elif low_count == 1:
        score = max(2, score - 1)
    
    return score

def extract_baths(desc_lower, zimmer):
    """Extract bath count from description."""
    # Try various patterns
    patterns = [
        r'(\d+)\s*bath(?:room)?s?',
        r'(\d+)\s*ba[ñn]os?',
        r'(\d+)\s*badezimmer',
        r'(\d+)\s*wc',
        r'(\d+)\s*salle[s]?\s*de\s*bain',
    ]
    for pat in patterns:
        m = re.search(pat, desc_lower)
        if m:
            val = int(m.group(1))
            if 1 <= val <= 20:
                return val
    # Fallback
    if zimmer:
        return math.ceil(zimmer * 0.6)
    return None

def calc_garten(grundstueck, wohnflaeche):
    """I = Garten m²."""
    if grundstueck is None:
        return None
    if wohnflaeche:
        garden = grundstueck - (wohnflaeche * 3)
    else:
        garden = grundstueck * 0.3
    if garden <= 0:
        garden = grundstueck * 0.3
    return min(50000, round(garden))

def calc_renovierung(desc_lower):
    """V = Renovierung 0-100."""
    if any(k in desc_lower for k in ['newly renovated', 'new build', 'turnkey', 'llave en mano', 
                                       'brand new', 'completely renovated', 'fully renovated',
                                       'recently built', 'newly built']):
        return 95
    if 'new ' in desc_lower and ('condition' in desc_lower or 'state' in desc_lower):
        return 90
    if any(k in desc_lower for k in ['recently renovated', 'good condition', 'excellent condition',
                                       'perfect condition', 'immaculate', 'move-in ready',
                                       'move in ready', 'well maintained', 'well-maintained']):
        return 78
    if any(k in desc_lower for k in ['needs updating', 'some work', 'updating', 'modernisation',
                                       'modernization']):
        return 50
    if any(k in desc_lower for k in ['reform needed', 'total renovation', 'to renovate', 
                                       'para reformar', 'renovation project', 'needs renovation',
                                       'requires renovation', 'in need of renovation',
                                       'needs total renovation', 'needs complete renovation',
                                       'restoration project', 'needs restoring']):
        return 20
    return 60

def calc_bewirtschaftung(grundstueck, desc_lower):
    """W = Bewirtschaftung 1-5."""
    if grundstueck is None:
        return 3
    if grundstueck < 20000:
        score = 4
    elif grundstueck < 50000:
        score = 3
    elif grundstueck <= 100000:
        score = 2
    else:
        score = 1
    
    if any(k in desc_lower for k in ['low maintenance', 'easy care', 'minimal maintenance']):
        score = min(5, score + 1)
    if any(k in desc_lower for k in ['vineyard', 'horses', 'farm', 'viña', 'bodega', 'stables', 'livestock']):
        score = max(1, score - 1)
    return score

def calc_vermietlizenz(desc_lower):
    """X = Vermietlizenz 100/50/0."""
    if any(k in desc_lower for k in ['tourist license', 'licencia turística', 'licencia turistica',
                                       'rental license', 'etv ', 'etv.', 'tourist rental license',
                                       'holiday rental license', 'vacation rental license',
                                       'tourist licence', 'alquiler turístico', 'alquiler turistico']):
        return 100
    if any(k in desc_lower for k in ['possibility of license', 'potential for license', 
                                       'possibility of tourist', 'potential tourist license',
                                       'option to obtain', 'option for license', 'could obtain',
                                       'possibility to obtain']):
        return 50
    return 0

def calc_erreichbarkeit(flughafen_min):
    """Y = Erreichbarkeit 0-100."""
    if flughafen_min is None:
        return None
    if flughafen_min <= 20:
        return 100
    elif flughafen_min <= 30:
        return 85
    elif flughafen_min <= 40:
        return 70
    elif flughafen_min <= 50:
        return 50
    else:
        return 30

def calc_score(flughafen_min, grundstueck, preis, erreichbarkeit, charme):
    """Z = Score 0-100."""
    if any(v is None for v in [flughafen_min, grundstueck, preis, erreichbarkeit, charme]):
        return None
    lage = max(0, min(100, 100 - (flughafen_min - 20) * 1.5))
    grundstueck_score = min(100, grundstueck / 500)
    preis_score = max(0, 100 - (preis / 1_000_000 - 3) * 5)
    score = (lage * 0.30 + grundstueck_score * 0.25 + preis_score * 0.15 + 
             erreichbarkeit * 0.10 + charme * 20 * 0.20)
    return round(score, 1)

def extract_gebaeude(desc_lower, desc_orig):
    """AC = Gebäudestruktur."""
    # Determine type
    typ = 'Haus'
    if any(k in desc_lower for k in ['finca', 'country estate', 'country house', 'landhaus']):
        typ = 'Finca'
    elif any(k in desc_lower for k in ['villa']):
        typ = 'Villa'
    elif any(k in desc_lower for k in ['herrenhaus', 'manor', 'mansion', 'palacio', 'palace']):
        typ = 'Herrenhaus'
    
    parts = [typ]
    
    if any(k in desc_lower for k in ['guest house', 'guesthouse', 'casa de huéspedes', 
                                      'casa de huespedes', 'gästehaus', "guest's house"]):
        parts.append('Gästehaus')
    elif any(k in desc_lower for k in ['annexe', 'annex', 'outbuilding', 'dependencies', 
                                        'dependencias', 'guest apartment', 'guest cottage']):
        parts.append('Annexe')
    
    if any(k in desc_lower for k in ['pool', 'piscina', 'swimming pool']):
        parts.append('Pool')
    
    if any(k in desc_lower for k in ['stable', 'stables', 'horses', 'pferde', 'caballerizas']):
        parts.append('Ställe')
    
    if any(k in desc_lower for k in ['vineyard', 'viña', 'viñedo', 'weinberg', 'bodega']):
        parts.append('Weinberg')
    
    return ' + '.join(parts)

def extract_baujahr(desc_lower, desc_orig):
    """AD = Baujahr."""
    # Try explicit year mentions
    patterns = [
        r'built\s+in\s+(\d{4})',
        r'constructed\s+in?\s*(\d{4})',
        r'año\s+(\d{4})',
        r'baujahr\s+(\d{4})',
        r'dating\s+(?:back\s+)?to\s+(?:the\s+)?(\d{4})',
        r'built\s+(\d{4})',
        r'from\s+(\d{4})',
        r'since\s+(\d{4})',
        r'erected\s+in\s+(\d{4})',
        r'(\d{4})\s+(?:construction|build)',
    ]
    for pat in patterns:
        m = re.search(pat, desc_lower)
        if m:
            year = int(m.group(1))
            if 1400 <= year <= 2026:
                return year
    
    # Century hints
    if any(k in desc_lower for k in ['19th century', 'siglo xix', '1800s']):
        return 1880
    if any(k in desc_lower for k in ['18th century', 'siglo xviii', '1700s']):
        return 1780
    if any(k in desc_lower for k in ['17th century', 'siglo xvii']):
        return 1650
    if any(k in desc_lower for k in ['16th century', 'siglo xvi']):
        return 1550
    if any(k in desc_lower for k in ['20th century', 'siglo xx']):
        return 1950
    
    # Style hints
    if any(k in desc_lower for k in ['new build', 'newly built', 'newly constructed', 'brand new']):
        return 2022
    if 'modern' in desc_lower and 'build' in desc_lower:
        return 2010
    if any(k in desc_lower for k in ['traditional stone', 'stone manor', 'historic', 'medieval']):
        return 1900
    
    return None

def extract_reno_begruendung(desc_orig, reno_score):
    """AG = Reno-Begründung (1 sentence)."""
    if not desc_orig or desc_orig == 'None':
        return None
    
    # Find relevant sentence
    sentences = re.split(r'(?<=[.!?])\s+', desc_orig.strip())
    
    keywords_good = ['renovated', 'turnkey', 'condition', 'new', 'maintained', 'immaculate', 
                     'refurbished', 'restored', 'rebuilt']
    keywords_bad = ['reform', 'renovate', 'restoration', 'project', 'updating', 'work needed',
                    'potential', 'opportunity']
    
    target_keywords = keywords_bad if reno_score < 60 else keywords_good
    
    for sent in sentences:
        sent_lower = sent.lower()
        if any(k in sent_lower for k in target_keywords):
            # Truncate to reasonable length
            return sent[:200].strip()
    
    # Fallback: first meaningful sentence
    for sent in sentences:
        if len(sent) > 30:
            return sent[:200].strip()
    
    return None

def main():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    
    print("Loading bilder.json...")
    bilder = load_bilder()
    
    # Stats tracking
    stats = {col: 0 for col in ['D', 'F', 'I', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AF', 'AG']}
    
    scores_by_row = {}  # row -> score for ranking
    
    print("Processing rows 15-333...")
    for row in range(15, 334):
        desc_lower, desc_orig, bilder_desc = get_desc(ws, row, bilder)
        
        # Read existing values
        zimmer = ws.cell(row, COL_E).value
        grundstueck = ws.cell(row, COL_G).value
        wohnflaeche = ws.cell(row, COL_H).value
        flughafen_min = ws.cell(row, COL_L).value
        preis = ws.cell(row, COL_S).value
        
        # D = Charme (always fill if empty)
        if ws.cell(row, COL_D).value is None:
            charme = calc_charme(desc_lower)
            ws.cell(row, COL_D).value = charme
            stats['D'] += 1
        else:
            charme = ws.cell(row, COL_D).value
        
        # F = Bäder (only if empty)
        if ws.cell(row, COL_F).value is None:
            baths = extract_baths(desc_lower, zimmer)
            if baths is not None:
                ws.cell(row, COL_F).value = baths
                stats['F'] += 1
        
        # I = Garten
        if ws.cell(row, COL_I).value is None:
            garten = calc_garten(grundstueck, wohnflaeche)
            if garten is not None:
                ws.cell(row, COL_I).value = garten
                stats['I'] += 1
        
        # V = Renovierung
        if ws.cell(row, COL_V).value is None:
            reno = calc_renovierung(desc_lower)
            ws.cell(row, COL_V).value = reno
            stats['V'] += 1
        else:
            reno = ws.cell(row, COL_V).value
        
        # W = Bewirtschaftung
        if ws.cell(row, COL_W).value is None:
            bewirt = calc_bewirtschaftung(grundstueck, desc_lower)
            ws.cell(row, COL_W).value = bewirt
            stats['W'] += 1
        
        # X = Vermietlizenz
        if ws.cell(row, COL_X).value is None:
            lizenz = calc_vermietlizenz(desc_lower)
            ws.cell(row, COL_X).value = lizenz
            stats['X'] += 1
        
        # Y = Erreichbarkeit
        if ws.cell(row, COL_Y).value is None:
            erreich = calc_erreichbarkeit(flughafen_min)
            if erreich is not None:
                ws.cell(row, COL_Y).value = erreich
                stats['Y'] += 1
        else:
            erreich = ws.cell(row, COL_Y).value
        
        # Recalculate erreichbarkeit for score even if already set
        erreich_val = ws.cell(row, COL_Y).value
        if erreich_val is None and flughafen_min is not None:
            erreich_val = calc_erreichbarkeit(flughafen_min)
        
        # Z = Score
        if ws.cell(row, COL_Z).value is None:
            score = calc_score(flughafen_min, grundstueck, preis, erreich_val, charme)
            if score is not None:
                ws.cell(row, COL_Z).value = score
                scores_by_row[row] = score
                stats['Z'] += 1
        else:
            scores_by_row[row] = ws.cell(row, COL_Z).value
        
        # AA = SortKey
        if ws.cell(row, COL_AA).value is None:
            score_val = ws.cell(row, COL_Z).value
            if score_val is not None:
                ws.cell(row, COL_AA).value = score_val
                stats['AA'] += 1
        
        # AC = Gebäudestruktur
        if ws.cell(row, COL_AC).value is None:
            gebaeude = extract_gebaeude(desc_lower, desc_orig)
            ws.cell(row, COL_AC).value = gebaeude
            stats['AC'] += 1
        
        # AD = Baujahr
        if ws.cell(row, COL_AD).value is None:
            baujahr = extract_baujahr(desc_lower, desc_orig)
            if baujahr is not None:
                ws.cell(row, COL_AD).value = baujahr
                stats['AD'] += 1
        
        # AF = Reno-Score (= V)
        if ws.cell(row, COL_AF).value is None:
            reno_val = ws.cell(row, COL_V).value
            if reno_val is not None:
                ws.cell(row, COL_AF).value = reno_val
                stats['AF'] += 1
        
        # AG = Reno-Begründung
        if ws.cell(row, COL_AG).value is None:
            reno_val = ws.cell(row, COL_V).value or 60
            begruendung = extract_reno_begruendung(desc_orig, reno_val)
            if begruendung:
                ws.cell(row, COL_AG).value = begruendung
                stats['AG'] += 1
        
        # Checkpoint every 50 rows
        if row % 50 == 0:
            wb.save(XLSX)
            print(f"  Checkpoint saved at row {row}")
    
    # AB = Rang (assign based on score ranking)
    print("Assigning ranks...")
    # Sort rows by score descending
    sorted_rows = sorted(scores_by_row.items(), key=lambda x: x[1], reverse=True)
    for rank, (row, score) in enumerate(sorted_rows, 1):
        ws.cell(row, COL_AB).value = rank
        stats['AB'] += 1
    
    print("Saving final workbook...")
    wb.save(XLSX)
    
    print("\n=== STATISTIK ===")
    total = 319
    for col, count in stats.items():
        print(f"  Spalte {col}: {count}/{total} Zeilen befüllt")
    
    print(f"\nScore-Range: {min(scores_by_row.values()):.1f} - {max(scores_by_row.values()):.1f}")
    print(f"Score-Durchschnitt: {sum(scores_by_row.values())/len(scores_by_row):.1f}")
    print(f"Rows mit Score: {len(scores_by_row)}")
    print("Fertig!")

if __name__ == '__main__':
    main()
