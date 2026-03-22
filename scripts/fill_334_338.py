#!/usr/bin/env python3
"""Fill all missing fields for Ordnungsnummern 334-338."""

import openpyxl
import requests
import base64
import json
import os
import anthropic

XLSX = '/Users/robin/.openclaw/workspace/mallorca-projekt/data/mallorca-kandidaten-v2.xlsx'
BILDER_DIR = '/Users/robin/.openclaw/workspace/mallorca-projekt/bilder'
API_KEY = "AIzaSyAzplaFOGUrCygk3mrzS--wrHNIz0arKok"

# Objekte 334-338 mit Koordinaten
OBJECTS = {
    334: {"name": "Sencelles", "lat": 39.6497, "lng": 2.9008, "grundstueck": 14483, "bebaut": 414, "zimmer": 5, "baeder": 4, "preis": 4250000, "baujahr": 2021, "desc": "Die exklusive Finca wurde 2022 neu gebaut und befindet sich inmitten der Weinregion Biniali. Die großzügige Wohnfläche von 360 m² befindet sich auf einem 15.000 m² großen Grundstück. Beheizte Pool, Fitnessstudio."},
    335: {"name": "Santanyí", "lat": 39.3554, "lng": 3.1243, "grundstueck": 17700, "bebaut": 792, "zimmer": 9, "baeder": 5, "preis": 3500000, "baujahr": 1800, "desc": "Diese 300 Jahre alte, aufwendig sanierte Naturstein-Finca, die aus einem Haupthaus und mehreren Nebengebäuden besteht, beeindruckt sowohl durch ihre Ursprünglichkeit als auch ihren wunderbaren Charme."},
    336: {"name": "Felanitx", "lat": 39.4688, "lng": 3.1483, "grundstueck": 19533, "bebaut": 336, "zimmer": 6, "baeder": 3, "preis": 2860000, "baujahr": 2000, "desc": "Diese traditionelle Finca befindet sich in ruhiger, ländlicher Lage auf einem großzügigen Grundstück von 19.533 m² mit eigenem Weinberg und traumhaftem Blick auf den Berg San Salvador."},
    337: {"name": "Sant Llorenç", "lat": 39.6200, "lng": 3.2833, "grundstueck": 16239, "bebaut": 330, "zimmer": 6, "baeder": 4, "preis": 4650000, "baujahr": 2017, "desc": "2018 im Mallorquinischen Stil mit Natursteinfassade erbaut, erhebt sie sich über das eigene 16.239 m² Land. Panoramablick, modern ausgestattet."},
    338: {"name": "Montuïri", "lat": 39.5667, "lng": 3.0000, "grundstueck": 15033, "bebaut": 473, "zimmer": 5, "baeder": 4, "preis": 2890000, "baujahr": 2019, "desc": "Diese beeindruckende Finca vereint modernen Komfort perfekt mit dem zeitlosen Charme mediterranen Lebens. Das Haupthaus erstreckt sich über 311 m² auf einer Ebene, ergänzt durch einen 70 m² großen Keller."},
}

# Referenzpunkte
REFS = {
    "flughafen": "39.5517,2.7388",
    "daia":      "39.7456,2.6489",
    "andratx":   "39.5747,2.3818",
    "ses_salines":"39.3444,3.0503",
}

def get_distances(lat, lng):
    origin = f"{lat},{lng}"
    dest_str = "|".join(REFS.values())
    url = "https://maps.googleapis.com/maps/api/distancematrix/json"
    params = {"origins": origin, "destinations": dest_str, "mode": "driving", "key": API_KEY}
    r = requests.get(url, params=params, timeout=15)
    data = r.json()
    if data["status"] != "OK":
        raise Exception(f"API error: {data['status']}")
    result = {}
    row_data = data["rows"][0]["elements"]
    for i, key in enumerate(REFS.keys()):
        el = row_data[i]
        if el["status"] == "OK":
            km = round(el["distance"]["value"] / 1000, 1)
            mins = round(el["duration"]["value"] / 60)
            result[key] = (km, mins)
        else:
            result[key] = (None, None)
    return result

def analyse_image_with_claude(nr, desc):
    """Use Claude vision to analyze image for Charme and Renovierung."""
    img_path = f"{BILDER_DIR}/{nr}_main.jpg"
    if not os.path.exists(img_path):
        print(f"  Image not found: {img_path}")
        return None, None, None, None
    
    with open(img_path, "rb") as f:
        img_data = base64.standard_b64encode(f.read()).decode("utf-8")
    
    client = anthropic.Anthropic()
    prompt = f"""Analysiere dieses Bild einer Mallorca-Finca/Immobilie.

Objektbeschreibung: {desc}

Bitte bewerte folgendes auf Basis des Bildes und der Beschreibung:

1. Charme/Ästhetik (1-5 Punkte):
   - 5 = außergewöhnlich charmant, historisch wertvoll, beeindruckend
   - 4 = sehr attraktiv, hochwertig
   - 3 = solide, ansprechend
   - 2 = eher nüchtern, wenig Charme
   - 1 = renovierungsbedürftig, wenig Charme

2. Renovierungszustand (0-100):
   - 90-100 = Neubau oder vollständig renoviert, einzugsbereit
   - 70-89 = guter Zustand, kaum Renovierung nötig
   - 50-69 = solider Zustand, leichte Modernisierungen sinnvoll
   - 30-49 = teilweise renovierungsbedürftig
   - 0-29 = stark renovierungsbedürftig

3. Reno-Begründung (1 kurzer Satz auf Deutsch)

4. Gästehäuser (0, 1 oder 2) - basierend auf Beschreibung und Bild

Antworte NUR als JSON:
{{"charme": <1-5>, "renovierung": <0-100>, "reno_begruendung": "<text>", "gaestehaueser": <0-2>}}"""

    msg = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=300,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": img_data}},
                {"type": "text", "text": prompt}
            ]
        }]
    )
    
    try:
        text = msg.content[0].text.strip()
        # Extract JSON
        start = text.find('{')
        end = text.rfind('}') + 1
        result = json.loads(text[start:end])
        return result.get("charme"), result.get("renovierung"), result.get("reno_begruendung"), result.get("gaestehaueser", 0)
    except Exception as e:
        print(f"  JSON parse error: {e}, text: {text[:200]}")
        return None, None, None, None

def calc_erreichbarkeit(flughafen_min, daia_min, ses_salines_min, andratx_min):
    """Calculate Erreichbarkeit 0-100 based on Einstellungen weights."""
    # Weights from Einstellungen: Flughafen=30%, Daia=30%, Ses Salines=30%, Andratx=10%
    # Ideal/Akzeptabel/Dealbreaker from Einstellungen sheet
    
    def score_for(mins, ideal, akzeptabel, dealbreaker):
        if mins is None:
            return 50
        if mins <= ideal:
            return 100
        elif mins <= akzeptabel:
            # Linear from 100 to 60
            return round(100 - 40 * (mins - ideal) / (akzeptabel - ideal))
        elif mins <= dealbreaker:
            # Linear from 60 to 0
            return round(60 - 60 * (mins - akzeptabel) / (dealbreaker - akzeptabel))
        else:
            return 0
    
    s_flughafen = score_for(flughafen_min, 15, 25, 40)
    s_daia = score_for(daia_min, 20, 40, 70)
    s_ses = score_for(ses_salines_min, 15, 30, 45)
    s_andratx = score_for(andratx_min, 25, 40, 60)
    
    erreich = round(s_flughafen * 0.30 + s_daia * 0.30 + s_ses * 0.30 + s_andratx * 0.10)
    return erreich

def calc_score(obj, charme, renovierung, bewirtschaftung, vermietlizenz, erreichbarkeit, gaestehaueser):
    """Calculate overall Score 0-100 based on Einstellungen weights."""
    zimmer = obj["zimmer"]
    grundstueck = obj["grundstueck"]
    bebaut = obj["bebaut"]
    preis = obj["preis"]
    
    # Gewichte from Einstellungen:
    # Zimmer & Platz: 20, Preis-Leistung (€/m²): 15, Gästehaus: 15
    # Erreichbarkeit: 15, Grundstück/Garten-Verhältnis: 10
    # Vermietlizenz: 5, Bewirtschaftung: 5, Charme: 10, Renovierung: 5
    
    # 1. Zimmer & Platz (0-100): ideal=8, min=5
    zimmer_score = min(100, max(0, (zimmer - 5) / (8 - 5) * 100)) if zimmer else 50
    zimmer_score = min(100, zimmer_score)
    
    # 2. Preis-Leistung €/m² (bebaut): lower is better
    eur_m2 = preis / bebaut if bebaut else preis / grundstueck
    # Reference: ~5000 = good, ~15000 = bad for Mallorca fincas
    preis_score = max(0, min(100, 100 - (eur_m2 - 5000) / 150))
    
    # 3. Gästehaus (0/1/2): 0=0, 1=50, 2=100
    gast_score = (gaestehaueser or 0) * 50
    
    # 4. Erreichbarkeit (already 0-100)
    erreich_score = erreichbarkeit or 0
    
    # 5. Grundstück/Garten-Verhältnis: ideal=15000m², min=3000m²
    garten = grundstueck - bebaut
    garten_score = min(100, max(0, (garten - 1000) / (5000 - 1000) * 100))
    
    # 6. Vermietlizenz: 0=0, 50=50, 100=100
    vmietliz_score = vermietlizenz or 0
    
    # 7. Bewirtschaftung (1-5): normalize to 0-100
    bewirt_score = ((bewirtschaftung or 3) - 1) / 4 * 100
    
    # 8. Charme (1-5): normalize to 0-100
    charme_score = ((charme or 3) - 1) / 4 * 100
    
    # 9. Renovierung (0-100)
    reno_score = renovierung or 60
    
    score = (
        zimmer_score * 0.20 +
        preis_score * 0.15 +
        gast_score * 0.15 +
        erreich_score * 0.15 +
        garten_score * 0.10 +
        vmietliz_score * 0.05 +
        bewirt_score * 0.05 +
        charme_score * 0.10 +
        reno_score * 0.05
    )
    return round(score, 1)

def calc_bewirtschaftung(grundstueck, desc_lower):
    if grundstueck < 20000:
        score = 4
    elif grundstueck < 50000:
        score = 3
    else:
        score = 2
    if any(k in desc_lower for k in ['weinberg', 'vineyard']):
        score = max(1, score - 1)
    return score

def calc_vermietlizenz(desc_lower):
    if any(k in desc_lower for k in ['tourist', 'touristic', 'vermiet', 'rental license', 'licencia']):
        return 100
    if any(k in desc_lower for k in ['möglich', 'possible', 'potential']):
        return 50
    return 0

# Load workbook
print("Loading workbook...")
wb = openpyxl.load_workbook(XLSX)
ws = wb.active

# Find rows for 334-338
row_map = {}
for row in ws.iter_rows(min_row=2):
    nr = row[0].value
    if nr in [334, 335, 336, 337, 338]:
        row_map[nr] = row[0].row

print(f"Found rows: {row_map}")

results = {}

for nr, obj in OBJECTS.items():
    print(f"\n=== Processing {nr} ({obj['name']}) ===")
    row_idx = row_map[nr]
    
    # 1. Google Maps distances
    print(f"  Fetching distances...")
    dists = get_distances(obj["lat"], obj["lng"])
    print(f"  Distances: {dists}")
    
    flughafen_km, flughafen_min = dists["flughafen"]
    daia_km, daia_min = dists["daia"]
    andratx_km, andratx_min = dists["andratx"]
    ses_km, ses_min = dists["ses_salines"]
    
    # Write distances
    ws.cell(row_idx, 11).value = flughafen_km
    ws.cell(row_idx, 12).value = flughafen_min
    ws.cell(row_idx, 13).value = daia_km
    ws.cell(row_idx, 14).value = daia_min
    ws.cell(row_idx, 15).value = andratx_km
    ws.cell(row_idx, 16).value = andratx_min
    ws.cell(row_idx, 17).value = ses_km
    ws.cell(row_idx, 18).value = ses_min
    
    # 2. Garten
    garten = obj["grundstueck"] - obj["bebaut"]
    ws.cell(row_idx, 9).value = garten
    print(f"  Garten: {garten} m²")
    
    # 3. €/m²
    eur_m2_bebaut = round(obj["preis"] / obj["bebaut"])
    eur_m2_grundstueck = round(obj["preis"] / obj["grundstueck"])
    ws.cell(row_idx, 20).value = eur_m2_bebaut
    ws.cell(row_idx, 21).value = eur_m2_grundstueck
    
    # 4. Vision analysis
    print(f"  Analyzing image...")
    charme, renovierung, reno_begruendung, gaestehaueser = analyse_image_with_claude(nr, obj["desc"])
    print(f"  Charme={charme}, Renovierung={renovierung}, Gästehäuser={gaestehaueser}")
    print(f"  Reno-Begründung: {reno_begruendung}")
    
    if charme is None:
        charme = 3
    if renovierung is None:
        renovierung = 60
    if gaestehaueser is None:
        gaestehaueser = 0
    
    ws.cell(row_idx, 4).value = charme   # Charme/Ästhetik
    ws.cell(row_idx, 22).value = renovierung  # Renovierung
    ws.cell(row_idx, 32).value = renovierung  # Reno-Score
    ws.cell(row_idx, 33).value = reno_begruendung  # Reno-Begründung
    ws.cell(row_idx, 40).value = gaestehaueser  # Gästehäuser
    
    # 5. Bewirtschaftung
    desc_lower = obj["desc"].lower()
    bewirtschaftung = calc_bewirtschaftung(obj["grundstueck"], desc_lower)
    ws.cell(row_idx, 23).value = bewirtschaftung
    print(f"  Bewirtschaftung: {bewirtschaftung}")
    
    # 6. Vermietlizenz
    vermietlizenz = calc_vermietlizenz(desc_lower)
    ws.cell(row_idx, 24).value = vermietlizenz
    print(f"  Vermietlizenz: {vermietlizenz}")
    
    # 7. Erreichbarkeit
    erreichbarkeit = calc_erreichbarkeit(flughafen_min, daia_min, ses_min, andratx_min)
    ws.cell(row_idx, 25).value = erreichbarkeit
    print(f"  Erreichbarkeit: {erreichbarkeit}")
    
    # 8. Score
    score = calc_score(obj, charme, renovierung, bewirtschaftung, vermietlizenz, erreichbarkeit, gaestehaueser)
    ws.cell(row_idx, 26).value = score
    ws.cell(row_idx, 27).value = score  # SortKey
    print(f"  Score: {score}")
    
    results[nr] = {
        "flughafen_min": flughafen_min,
        "daia_min": daia_min,
        "andratx_min": andratx_min,
        "ses_min": ses_min,
        "charme": charme,
        "renovierung": renovierung,
        "gaestehaueser": gaestehaueser,
        "bewirtschaftung": bewirtschaftung,
        "vermietlizenz": vermietlizenz,
        "erreichbarkeit": erreichbarkeit,
        "score": score,
        "garten": garten,
    }

print("\nSaving workbook...")
wb.save(XLSX)
print("Saved!")

print("\n=== SUMMARY ===")
for nr, r in results.items():
    print(f"\nNr. {nr} ({OBJECTS[nr]['name']}):")
    print(f"  Flughafen: {r['flughafen_min']} min")
    print(f"  Daia: {r['daia_min']} min")
    print(f"  Andratx: {r['andratx_min']} min")
    print(f"  Ses Salines: {r['ses_min']} min")
    print(f"  Garten: {r['garten']} m²")
    print(f"  Charme: {r['charme']}/5")
    print(f"  Renovierung: {r['renovierung']}")
    print(f"  Gästehäuser: {r['gaestehaueser']}")
    print(f"  Bewirtschaftung: {r['bewirtschaftung']}/5")
    print(f"  Vermietlizenz: {r['vermietlizenz']}")
    print(f"  Erreichbarkeit: {r['erreichbarkeit']}")
    print(f"  Score: {r['score']}")
