#!/usr/bin/env python3
"""
Balearic Properties — Mallorca Scraper (structure-aware)

Listing structure per article.card-property:
  - article.card-default.card-property  ← container
  - h3.card-title a[href]               ← title + URL
  - p.card-price                         ← "888,000 €"
  - ul.list-unstyled li                  ← "3 Beds", "2 Baths", "151 m² Built", "0 m² " (plot)
  - span.card-ref / span (sibling)       ← ref / location
"""

import requests
from bs4 import BeautifulSoup
import time
import re
from openpyxl import load_workbook
from datetime import date

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
BASE_URL = 'https://www.balearic-properties.com/property-for-sale/mallorca.html'
TOTAL_PAGES = 88

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Referer': 'https://www.balearic-properties.com/',
    'Connection': 'keep-alive',
}


def parse_price(text):
    if not text:
        return None
    digits = re.sub(r'[^\d]', '', text)
    return int(digits) if digits else None


def parse_m2(text):
    if not text:
        return None
    m = re.search(r'([\d,\.]+)\s*m', text)
    if m:
        val = m.group(1).replace('.', '').replace(',', '')
        try:
            return int(val)
        except:
            return None
    return None


def parse_page(soup):
    listings = []
    cards = soup.select('article.card-property')
    for card in cards:
        obj = {}

        # Title + URL
        title_el = card.select_one('h3.card-title a, h2.card-title a')
        if title_el:
            obj['titel'] = title_el.get_text(strip=True)
            href = title_el.get('href', '')
            obj['url'] = ('https://www.balearic-properties.com' + href
                          if href.startswith('/') else href)
        else:
            obj['titel'] = ''
            obj['url'] = '—'

        # Price
        price_el = card.select_one('p.card-price, .card-price')
        obj['preis'] = parse_price(price_el.get_text() if price_el else None)

        # Beds, Baths, Built m², Plot m²
        obj['zimmer'] = None
        obj['bader'] = None
        obj['wohnflaeche'] = None
        obj['grundstueck'] = None

        specs_ul = card.select_one('ul.list-unstyled')
        if specs_ul:
            for li in specs_ul.select('li'):
                txt = li.get_text(strip=True)
                if not txt:
                    continue
                tl = txt.lower()
                if 'bed' in tl:
                    m = re.search(r'(\d+)', txt)
                    if m:
                        obj['zimmer'] = int(m.group(1))
                elif 'bath' in tl:
                    m = re.search(r'(\d+)', txt)
                    if m:
                        obj['bader'] = int(m.group(1))
                elif 'built' in tl or 'm²' in txt:
                    # First m² is built area
                    if obj['wohnflaeche'] is None:
                        obj['wohnflaeche'] = parse_m2(txt)
                    else:
                        obj['grundstueck'] = parse_m2(txt)
                # Some pages show plot separately
                if 'plot' in tl or 'land' in tl:
                    obj['grundstueck'] = parse_m2(txt)

        # Location — in the ref/loc paragraph
        ref_loc = card.select_one('p.card-ref-loc')
        if ref_loc:
            # Structure: <span class="card-ref">REF</span> / <span>Location</span>
            spans = ref_loc.select('span')
            if len(spans) >= 2:
                obj['ort'] = spans[-1].get_text(strip=True)
            elif len(spans) == 1:
                obj['ort'] = spans[0].get_text(strip=True)
            else:
                obj['ort'] = ref_loc.get_text(strip=True)
        else:
            obj['ort'] = ''

        listings.append(obj)
    return listings


def load_existing_urls():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    return set(
        str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True)
        if r[2] and str(r[2]).strip() not in ('—', 'None', '')
    )


def save_batch(batch, existing_urls):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    saved = 0
    for obj in batch:
        url = obj.get('url', '—')
        if url != '—' and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', ''),
            'Balearic Properties',
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            str(date.today()),
            'Neu'
        ])
        if url != '—':
            existing_urls.add(url)
        saved += 1
    wb.save(EXCEL_PATH)
    return saved


def main():
    existing_urls = load_existing_urls()
    print(f"Existing URLs in Excel: {len(existing_urls)}")

    session = requests.Session()
    session.headers.update(HEADERS)

    total_saved = 0
    total_parsed = 0
    batch = []
    consecutive_errors = 0

    for page_num in range(1, TOTAL_PAGES + 1):
        url = BASE_URL if page_num == 1 else f'{BASE_URL}?page={page_num}'
        print(f"[{page_num:02d}/{TOTAL_PAGES}] {url}", end='  ')

        try:
            r = session.get(url, timeout=30)
            print(f"HTTP {r.status_code} | {len(r.text):,} chars", end='  ')

            if r.status_code != 200:
                print(f"→ SKIP")
                consecutive_errors += 1
                if consecutive_errors >= 5:
                    print("5 consecutive errors — aborting.")
                    break
                time.sleep(2)
                continue

            consecutive_errors = 0
            soup = BeautifulSoup(r.text, 'html.parser')
            listings = parse_page(soup)
            print(f"| {len(listings)} listings", end='  ')
            total_parsed += len(listings)
            batch.extend(listings)

            # Save every 10 pages or on last page
            if page_num % 10 == 0 or page_num == TOTAL_PAGES:
                saved = save_batch(batch, existing_urls)
                total_saved += saved
                print(f"→ saved {saved} (total: {total_saved})", end='')
                batch = []

            print()

        except Exception as e:
            print(f"→ ERROR: {e}")
            consecutive_errors += 1
            if consecutive_errors >= 5:
                print("5 consecutive errors — aborting.")
                break
            time.sleep(3)
            continue

        # Rate limiting: 0.8s between requests
        if page_num < TOTAL_PAGES:
            time.sleep(0.8)

    # Flush any remaining batch
    if batch:
        saved = save_batch(batch, existing_urls)
        total_saved += saved
        print(f"\nFinal batch flushed: {saved}")

    print(f"\n{'='*60}")
    print(f"COMPLETE.")
    print(f"  Pages scraped:      {TOTAL_PAGES}")
    print(f"  Total parsed:       {total_parsed}")
    print(f"  New saved to Excel: {total_saved}")


if __name__ == '__main__':
    main()
