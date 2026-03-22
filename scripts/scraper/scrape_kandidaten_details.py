#!/usr/bin/env python3
"""
Scraper für mallorca-kandidaten-v2.xlsx — Zeilen 15-333
Holt: Bild-URL, Bäder, Beschreibung, Baujahr, Makler
Speichert in mallorca-bilder.json + Excel-Update
"""

import json, time, re, os, sys
from urllib.parse import urlparse
import requests
from bs4 import BeautifulSoup
import openpyxl

# Spalten-Mapping (1-basiert)
COL_URL = 3       # C
COL_BATHS = 6     # F
COL_YEAR = 30     # AD
COL_DESC = 35     # AI
COL_MAKLER = 36   # AJ

XLSX_FILE = 'mallorca-kandidaten-v2.xlsx'
JSON_FILE = 'mallorca-bilder.json'
CHECKPOINT_EVERY = 25

BASE_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Encoding': 'gzip, deflate',  # NO brotli — requests kann br nicht dekomprimieren
    'Connection': 'keep-alive',
}

session = requests.Session()
session.headers.update(BASE_HEADERS)

def fetch(url, lang='en-GB', retries=2):
    headers = {'Accept-Language': lang}
    for attempt in range(retries + 1):
        try:
            r = session.get(url, headers=headers, timeout=15)
            if r.status_code in (403, 429, 503):
                print(f"  [SKIP] {r.status_code} for {url[:60]}")
                return None
            if r.status_code == 200:
                return r
            print(f"  [WARN] Status {r.status_code} for {url[:60]}")
            return None
        except Exception as e:
            if attempt < retries:
                time.sleep(2)
            else:
                print(f"  [ERR] {e} for {url[:60]}")
                return None

def clean_text(t):
    if not t:
        return None
    return re.sub(r'\s+', ' ', t).strip() or None

# ─── Source-spezifische Extraktion ────────────────────────────────────────────

def scrape_engelvoelkers(url):
    """EV nutzt __NEXT_DATA__ JSON"""
    r = fetch(url, lang='de-DE')
    if not r:
        return {}
    soup = BeautifulSoup(r.text, 'html.parser')
    nd = soup.find('script', id='__NEXT_DATA__')
    result = {}
    if nd:
        try:
            data = json.loads(nd.string)
            s = json.dumps(data)
            # Image (og:image als Fallback)
            og = soup.find('meta', property='og:image')
            if og:
                result['img_url'] = og.get('content')
            # Bathrooms: suche nach "bathrooms":{"min":X,"max":Y} oder "bathrooms":X
            m = re.search(r'"bathrooms"\s*:\s*\{[^}]*"max"\s*:\s*(\d+)', s)
            if not m:
                m = re.search(r'"bathrooms"\s*:\s*(\d+)', s)
            if m:
                result['baths'] = int(m.group(1))
            # Year
            m = re.search(r'"constructionYear"\s*:\s*\{[^}]*"max"\s*:\s*(\d+)', s)
            if not m:
                m = re.search(r'"constructionYear"\s*:\s*(\d{4})', s)
            if m:
                y = int(m.group(1))
                if 1800 <= y <= 2025:
                    result['year'] = y
            # Description (first long German text)
            m = re.search(r'"description"\s*:\s*"([^"]{50,})"', s)
            if m:
                result['desc'] = m.group(1).encode('utf-8').decode('unicode_escape') if '\\u' in m.group(1) else m.group(1)
            # Fallback description via HTML
            if not result.get('desc'):
                for div in soup.find_all(['div','p'], class_=re.compile(r'desc|text|content', re.I)):
                    t = clean_text(div.get_text())
                    if t and len(t) > 80:
                        result['desc'] = t[:2000]
                        break
            # Makler: "name" near "agencyName"
            m = re.search(r'"agencyName"\s*:\s*"([^"]+)"', s)
            if not m:
                m = re.search(r'"agentName"\s*:\s*"([^"]+)"', s)
            if m:
                result['makler'] = m.group(1)
        except Exception as e:
            print(f"  [EV ERR] {e}")
    if not result.get('img_url'):
        og = soup.find('meta', property='og:image')
        if og:
            result['img_url'] = og.get('content')
    return result


def scrape_balearic(url):
    """Balearic Properties — reguläres HTML"""
    r = fetch(url, lang='en-GB')
    if not r:
        return {}
    soup = BeautifulSoup(r.text, 'html.parser')
    result = {}
    # Image
    og = soup.find('meta', property='og:image')
    if og:
        result['img_url'] = og.get('content')
    # Bathrooms: <span class="display-3">N</span> direkt vor Bathrooms
    for el in soup.find_all('span', class_='lead'):
        if 'bathroom' in el.get_text().lower():
            prev = el.find_previous_sibling('span')
            if prev and prev.get_text().strip().isdigit():
                result['baths'] = int(prev.get_text().strip())
            break
    # Description
    for sel in [('div', {'class': re.compile(r'property.*desc|description', re.I)}),
                ('div', {'id': re.compile(r'desc', re.I)}),
                ('div', {'class': re.compile(r'property-detail', re.I)})]:
        el = soup.find(*sel)
        if el:
            t = clean_text(el.get_text())
            if t and len(t) > 50:
                result['desc'] = t[:2000]
                break
    if not result.get('desc'):
        # Find meta description
        meta = soup.find('meta', {'name': 'description'})
        if meta and meta.get('content'):
            result['desc'] = meta['content'][:2000]
    # Year: look for "built" or "year" with 4 digits
    m = re.search(r'(built|constructed|year\s+built)\s*[:\-]?\s*(\d{4})', r.text, re.I)
    if m:
        y = int(m.group(2))
        if 1800 <= y <= 2025:
            result['year'] = y
    # Makler
    result['makler'] = 'Balearic Properties'
    return result


def scrape_luxury_estates(url):
    """luxury-estates-mallorca.com"""
    r = fetch(url, lang='en-GB')
    if not r:
        return {}
    soup = BeautifulSoup(r.text, 'html.parser')
    result = {}
    # Image from JSON-LD Product
    for tag in soup.find_all('script', type='application/ld+json'):
        try:
            d = json.loads(tag.string)
            if d.get('@type') == 'Product':
                imgs = d.get('image', [])
                if imgs:
                    img = imgs[0] if isinstance(imgs[0], str) else None
                    if img:
                        if img.startswith('/'):
                            img = 'https://www.luxury-estates-mallorca.com' + img
                        result['img_url'] = img
                break
        except:
            pass
    if not result.get('img_url'):
        og = soup.find('meta', property='og:image')
        if og:
            result['img_url'] = og.get('content')
    # Baths: "1Bathroom" or "NBathrooms" pattern
    m = re.search(r'(\d+)\s*Bathroom', r.text, re.I)
    if m:
        result['baths'] = int(m.group(1))
    # Year
    m = re.search(r'(built|year|constructed).{0,20}(\d{4})', r.text, re.I)
    if m:
        y = int(m.group(2))
        if 1800 <= y <= 2025:
            result['year'] = y
    # Description from meta
    meta = soup.find('meta', {'name': 'description'})
    if meta and meta.get('content') and len(meta['content']) > 30:
        result['desc'] = meta['content']
    # Try og:description
    if not result.get('desc'):
        og_d = soup.find('meta', property='og:description')
        if og_d and og_d.get('content'):
            result['desc'] = og_d['content']
    result['makler'] = 'Luxury Estates Mallorca'
    return result


def scrape_kyero(url):
    """Kyero — JSON-LD SingleFamilyResidence"""
    r = fetch(url, lang='en-GB')
    if not r:
        return {}
    soup = BeautifulSoup(r.text, 'html.parser')
    result = {}
    # Image
    og = soup.find('meta', property='og:image')
    if og:
        result['img_url'] = og.get('content')
    # JSON-LD
    for tag in soup.find_all('script', type='application/ld+json'):
        try:
            d = json.loads(tag.string)
            if d.get('@type') in ('SingleFamilyResidence', 'Residence', 'House', 'Apartment'):
                if 'numberOfBathroomsTotal' in d:
                    result['baths'] = int(d['numberOfBathroomsTotal'])
                if 'description' in d:
                    result['desc'] = d['description'][:2000]
                if 'yearBuilt' in d:
                    y = int(d['yearBuilt'])
                    if 1800 <= y <= 2025:
                        result['year'] = y
                break
        except:
            pass
    # Makler from page
    for el in soup.find_all(['div','span'], class_=re.compile(r'agent|makler|agency', re.I)):
        t = clean_text(el.get_text())
        if t and len(t) < 80:
            result['makler'] = t
            break
    return result


def scrape_sandberg(url):
    """sandberg-estates.com"""
    r = fetch(url, lang='en-GB')
    if not r:
        return {}
    soup = BeautifulSoup(r.text, 'html.parser')
    result = {}
    # Image
    og = soup.find('meta', property='og:image')
    if og:
        result['img_url'] = og.get('content')
    # Baths from data-caption pattern
    m = re.search(r'(\d+)\s*Bathrooms?', r.text, re.I)
    if m:
        result['baths'] = int(m.group(1))
    # Description
    meta = soup.find('meta', {'name': 'description'})
    if meta and meta.get('content') and len(meta['content']) > 30:
        result['desc'] = meta['content']
    # Year
    m = re.search(r'built.{0,20}(\d{4})|(\d{4}).{0,10}built', r.text, re.I)
    if m:
        y = int(m.group(1) or m.group(2))
        if 1800 <= y <= 2025:
            result['year'] = y
    result['makler'] = 'Sandberg Estates'
    return result


def scrape_generic(url):
    """Fallback für unbekannte Quellen"""
    r = fetch(url, lang='en-GB')
    if not r:
        return {}
    soup = BeautifulSoup(r.text, 'html.parser')
    result = {}
    # Image
    og = soup.find('meta', property='og:image')
    if og and og.get('content'):
        result['img_url'] = og['content']
    # Baths
    m = re.search(r'(\d+)\s*[Bb]athroom', r.text)
    if m:
        result['baths'] = int(m.group(1))
    # Description
    meta = soup.find('meta', {'name': 'description'}) or soup.find('meta', property='og:description')
    if meta and meta.get('content') and len(meta.get('content', '')) > 30:
        result['desc'] = meta['content']
    # Year
    m = re.search(r'built.{0,20}((?:19|20)\d{2})', r.text, re.I)
    if m:
        result['year'] = int(m.group(1))
    return result


# ─── LivingBlue via Playwright ────────────────────────────────────────────────

def scrape_livingblue_batch(rows_urls):
    """Nutzt Playwright um LivingBlue-Seiten zu scrapen"""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [WARN] Playwright nicht verfügbar für LivingBlue")
        return {}

    results = {}
    print(f"\n  [Playwright] Starte LivingBlue batch ({len(rows_urls)} URLs)...")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
            locale='de-DE',
            viewport={'width': 1280, 'height': 800},
        )

        for row, url in rows_urls:
            try:
                api_data = {}

                def handle_response(response):
                    u = response.url
                    if 'egorealestate.com' in u and 'json' in response.headers.get('content-type', ''):
                        try:
                            d = response.json()
                            if isinstance(d, dict) and ('images' in d or 'bathrooms' in str(d).lower()):
                                api_data.update(d)
                        except:
                            pass

                page = ctx.new_page()
                page.on("response", handle_response)
                page.goto(url, wait_until='domcontentloaded', timeout=25000)
                page.wait_for_timeout(3000)

                html = page.content()
                soup = BeautifulSoup(html, 'html.parser')

                result = {}

                # Try image from page
                imgs = page.query_selector_all('img[src*="egorealestate"]')
                if not imgs:
                    imgs = page.query_selector_all('img[data-src*="egorealestate"], .swiper-slide img, .gallery img, .slider img')
                if imgs:
                    for img in imgs:
                        src = img.get_attribute('src') or img.get_attribute('data-src') or ''
                        if src and 'egorealestate' in src and 'ORIGINAL' in src:
                            result['img_url'] = 'https:' + src if src.startswith('//') else src
                            break

                # Try og:image
                if not result.get('img_url'):
                    og = soup.find('meta', property='og:image')
                    if og and og.get('content'):
                        result['img_url'] = og['content']

                # Look for structured data in page
                full_text = html

                # Baths
                m = re.search(r'(\d+)\s*(?:Bad|Bäder|Bathroom|bath)', full_text, re.I)
                if m:
                    result['baths'] = int(m.group(1))

                # Description
                for sel in ['[class*="description"]', '[class*="desc"]', '[class*="text"]']:
                    els = page.query_selector_all(sel)
                    for el in els:
                        t = el.inner_text().strip()
                        if len(t) > 80:
                            result['desc'] = t[:2000]
                            break
                    if result.get('desc'):
                        break

                if not result.get('desc'):
                    meta = soup.find('meta', {'name': 'description'}) or soup.find('meta', property='og:description')
                    if meta and meta.get('content') and len(meta.get('content','')) > 30:
                        result['desc'] = meta['content']

                # Year
                m = re.search(r'(?:Baujahr|built|año)[:\s]*(\d{4})', full_text, re.I)
                if m:
                    y = int(m.group(1))
                    if 1800 <= y <= 2025:
                        result['year'] = y

                # Makler
                result['makler'] = 'Living Blue Mallorca'

                results[row] = result
                print(f"  [LB] Row {row}: img={bool(result.get('img_url'))}, baths={result.get('baths')}, desc={bool(result.get('desc'))}")

                page.close()
                time.sleep(1.5)

            except Exception as e:
                print(f"  [LB ERR] Row {row}: {e}")
                try:
                    page.close()
                except:
                    pass

        browser.close()

    return results


# ─── Main ─────────────────────────────────────────────────────────────────────

def get_scraper(url):
    domain = urlparse(url).netloc.replace('www.', '')
    if 'engelvoelkers' in domain:
        return scrape_engelvoelkers
    elif 'balearic-properties' in domain:
        return scrape_balearic
    elif 'luxury-estates-mallorca' in domain:
        return scrape_luxury_estates
    elif 'kyero' in domain:
        return scrape_kyero
    elif 'sandberg' in domain:
        return scrape_sandberg
    else:
        return scrape_generic


def main():
    print("=== Mallorca Kandidaten Detail Scraper ===")
    print(f"Excel: {XLSX_FILE}")
    print(f"JSON:  {JSON_FILE}")

    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active

    # Load existing JSON
    if os.path.exists(JSON_FILE):
        with open(JSON_FILE) as f:
            bilder = json.load(f)
        print(f"Loaded existing JSON: {len(bilder)} entries")
    else:
        bilder = {}

    # Collect all rows
    all_rows = []
    livingblue_rows = []

    for row in range(15, 334):
        url = ws.cell(row, COL_URL).value or ''
        if not url or url == '—':
            continue
        if 'livingblue' in url:
            livingblue_rows.append((row, url))
        else:
            all_rows.append((row, url))

    print(f"Non-LivingBlue rows: {len(all_rows)}")
    print(f"LivingBlue rows: {len(livingblue_rows)}")

    # Stats
    stats = {
        'total': len(all_rows) + len(livingblue_rows),
        'processed': 0,
        'skipped': 0,
        'img': 0,
        'baths': 0,
        'desc': 0,
        'year': 0,
        'by_source': {}
    }

    # ─── Process non-LivingBlue ───
    for i, (row, url) in enumerate(all_rows):
        row_key = str(row)
        domain = urlparse(url).netloc.replace('www.', '')

        # Skip if already done
        if row_key in bilder and bilder[row_key].get('img_url'):
            print(f"  [SKIP-cached] Row {row} ({domain})")
            stats['processed'] += 1
            continue

        scraper = get_scraper(url)
        print(f"[{i+1}/{len(all_rows)}] Row {row} ({domain}) → {url[:60]}")

        try:
            result = scraper(url)
        except Exception as e:
            print(f"  [ERR] {e}")
            result = {}

        bilder[row_key] = {
            'url': url,
            'img_url': result.get('img_url'),
            'baths': result.get('baths'),
            'desc': result.get('desc'),
            'year': result.get('year'),
            'makler': result.get('makler'),
        }

        # Update stats
        stats['processed'] += 1
        src = domain.split('.')[0]
        stats['by_source'][src] = stats['by_source'].get(src, 0) + 1
        if result.get('img_url'): stats['img'] += 1
        if result.get('baths'): stats['baths'] += 1
        if result.get('desc'): stats['desc'] += 1
        if result.get('year'): stats['year'] += 1

        # Update Excel
        if result.get('baths') and not ws.cell(row, COL_BATHS).value:
            ws.cell(row, COL_BATHS).value = result['baths']
        if result.get('year') and not ws.cell(row, COL_YEAR).value:
            ws.cell(row, COL_YEAR).value = result['year']
        if result.get('desc') and not ws.cell(row, COL_DESC).value:
            ws.cell(row, COL_DESC).value = result['desc'][:500]
        if result.get('makler') and not ws.cell(row, COL_MAKLER).value:
            ws.cell(row, COL_MAKLER).value = result['makler']

        # Checkpoint
        if (i + 1) % CHECKPOINT_EVERY == 0:
            wb.save(XLSX_FILE)
            with open(JSON_FILE, 'w', encoding='utf-8') as f:
                json.dump(bilder, f, ensure_ascii=False, indent=2)
            print(f"  [CHECKPOINT] Saved after row {row}")

        # Delay
        time.sleep(1.2)

    # ─── Process LivingBlue via Playwright ───
    # Filter out already-cached
    lb_todo = [(r, u) for r, u in livingblue_rows if not bilder.get(str(r), {}).get('img_url')]
    print(f"\n=== LivingBlue: {len(lb_todo)} URLs to scrape (skipping {len(livingblue_rows)-len(lb_todo)} cached) ===")

    if lb_todo:
        lb_results = scrape_livingblue_batch(lb_todo)
        for row, result in lb_results.items():
            row_key = str(row)
            bilder[row_key] = {
                'url': dict(livingblue_rows).get(row, ''),
                'img_url': result.get('img_url'),
                'baths': result.get('baths'),
                'desc': result.get('desc'),
                'year': result.get('year'),
                'makler': result.get('makler', 'Living Blue Mallorca'),
            }
            stats['processed'] += 1
            if result.get('img_url'): stats['img'] += 1
            if result.get('baths'): stats['baths'] += 1
            if result.get('desc'): stats['desc'] += 1
            if result.get('year'): stats['year'] += 1

            # Update Excel
            url = dict(livingblue_rows).get(row, '')
            if result.get('baths') and not ws.cell(row, COL_BATHS).value:
                ws.cell(row, COL_BATHS).value = result['baths']
            if result.get('year') and not ws.cell(row, COL_YEAR).value:
                ws.cell(row, COL_YEAR).value = result['year']
            if result.get('desc') and not ws.cell(row, COL_DESC).value:
                ws.cell(row, COL_DESC).value = result['desc'][:500]
            if result.get('makler') and not ws.cell(row, COL_MAKLER).value:
                ws.cell(row, COL_MAKLER).value = result['makler']

    # ─── Final save ───
    wb.save(XLSX_FILE)
    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(bilder, f, ensure_ascii=False, indent=2)

    print("\n=== STATISTIK ===")
    print(f"Gesamt verarbeitet: {stats['processed']}")
    print(f"Mit Bild-URL:       {stats['img']}")
    print(f"Mit Bäder:          {stats['baths']}")
    print(f"Mit Beschreibung:   {stats['desc']}")
    print(f"Mit Baujahr:        {stats['year']}")
    print(f"Nach Quelle:        {stats['by_source']}")
    print(f"\nJSON gespeichert:   {JSON_FILE}")
    print(f"Excel gespeichert:  {XLSX_FILE}")


if __name__ == '__main__':
    main()
