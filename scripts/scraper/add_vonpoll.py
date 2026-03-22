#!/usr/bin/env python3
"""Add 35 Von Poll properties to mallorca-kandidaten-v2.xlsx"""

import requests
import time
import openpyxl

EXCEL_PATH = "/Users/robin/.openclaw/workspace/mallorca-projekt/data/mallorca-kandidaten-v2.xlsx"
GMAPS_KEY = "AIzaSyAzplaFOGUrCygk3mrzS--wrHNIz0arKok"

REFS = {
    "Flughafen": "39.5517,2.7388",
    "Deia": "39.7456,2.6489",
    "Andratx": "39.5747,2.3818",
    "SesSalines": "39.3444,3.0503",
}

# (nr, name, url, zimmer, grundstuck, bebaut, preis, location_str, anzeigename)
OBJECTS = [
    (366, "Finca Panorama Santa Maria", "https://www.von-poll.com/de/expose/mallorca-santa-maria/traumhafte-finca-mit-panorama-und-meerblick-in-bester-lage-4160045287", 8, 14210, 700.0, 16200000, "Santa Maria del Cami, Mallorca, Spain", "Santa Maria — Finca Panorama; Meerblick"),
    (367, "Mediterrane Finca Santa Maria", "https://www.von-poll.com/de/expose/mallorca-santa-maria/exklusiver-mediterraner-ruckzugsort-mit-zeitgenossischer-eleganz-4160047549", 11, 16911, 580.0, 11500000, "Santa Maria del Cami, Mallorca, Spain", "Santa Maria — Mediterrane Finca; Zeitgenössische Eleganz"),
    (368, "Panorama Finca Tramuntana", "https://www.von-poll.com/de/expose/mallorca-santa-maria/fantastische-finca-mit-panoramablick-in-die-tramuntana-in-der-nahe-von-santa-maria-del-cami-4160040127", 7, 15886, 598.0, 8750000, "Santa Maria del Cami, Mallorca, Spain", "Santa Maria — Finca Tramuntana; Panoramablick"),
    (369, "Neubau Finca Bunyola", "https://www.von-poll.com/de/expose/mallorca-santa-maria/luxuriose-neubaufinca-mit-pool-und-panoramablick-in-bunyola-4160036281", 7, 33280, 487.0, 8500000, "Bunyola, Mallorca, Spain", "Bunyola — Neubau Finca; Panoramablick Pool"),
    (370, "Luxusfinca Alaró Santa Maria", "https://www.von-poll.com/de/expose/mallorca-santa-maria/spektakulare-luxusfinca-mit-grossem-pool-zwischen-alaro-und-santa-maria-4160038693", 8, 17300, 663.0, 6800000, "Alaro, Mallorca, Spain", "Alaró — Luxusfinca; Großer Pool"),
    (371, "Neubau Finca Campos Es Trenc", "https://www.von-poll.com/de/expose/mallorca-santa-maria/grosszugige-neubau-finca-in-der-nahe-des-es-trenc-strandes-bei-campos-mit-pool-und-eigenem-weinanbau-4160043131", 8, 45820, 485.0, 6670000, "Campos, Mallorca, Spain", "Campos — Neubau Finca; Weinanbau Es Trenc"),
    (372, "Weinberg Finca Montuïri", "https://www.von-poll.com/de/expose/mallorca-santa-maria/exklusive-neubaufinca-inmitten-eines-weinbergs-mit-pool-in-montuiri-4160045567", 7, 16132, 718.0, 5900000, "Montuiri, Mallorca, Spain", "Montuïri — Finca im Weinberg; Pool"),
    (373, "Villa Son Vida Golf", "https://www.von-poll.com/de/expose/mallorca-santa-maria/luxuriose-villa-in-son-vida-in-einer-privilegierten-lage-mit-blick-auf-den-nahe-gelegenen-golfplatz-4160039419", 9, 2000, 600.0, 4300000, "Son Vida, Palma, Mallorca, Spain", "Son Vida — Villa Golf; Privilegierte Lage"),
    (374, "Finca Projekt Llubí", "https://www.von-poll.com/de/expose/mallorca-santa-maria/projekt-zum-bau-einer-finca-mit-pool-in-der-landlichen-umgebung-von-llubi-4160049307", 8, 22033, 426.0, 3995000, "Llubi, Mallorca, Spain", "Llubí — Finca Projekt; Ländliche Idylle"),
    (375, "Finca Inca Costitx", "https://www.von-poll.com/de/expose/mallorca-santa-maria/finca-mit-pool-in-der-landlichen-idylle-zwischen-inca-und-costitx-4160044625", 8, 19347, 397.0, 3995000, "Costitx, Mallorca, Spain", "Costitx — Finca Pool; Inca Landidylle"),
    (376, "Finca Sa Cabaneta Marratxí", "https://www.von-poll.com/de/expose/mallorca-santa-maria/traumhafte-neu-gebaute-finca-mit-pool-in-sa-cabaneta-gemeinde-marratxi-4160044321", 7, 14159, 568.0, 3995000, "Sa Cabaneta, Marratxi, Mallorca, Spain", "Marratxí — Neubau Finca; Sa Cabaneta Pool"),
    (377, "Finca ETV Großes Grundstück", "https://www.von-poll.com/de/expose/mallorca-santa-maria/weitlaufige-finca-mit-etv-und-grossem-grundstuck-4160044995", 9, 13128, 551.0, 3950000, "Santa Maria del Cami, Mallorca, Spain", "Santa Maria — Finca ETV; Großes Grundstück"),
    (378, "Neubau Villa Son Gual", "https://www.von-poll.com/de/expose/mallorca-santa-maria/moderne-neubau-villa-mit-pool-in-son-gual-4160044109", 8, 2118, 474.0, 3900000, "Son Gual, Palma, Mallorca, Spain", "Son Gual — Neubau Villa; Moderner Pool"),
    (379, "Finca Moscari Selva Vermietlizenz", "https://www.von-poll.com/de/expose/mallorca-santa-maria/moderne-finca-mit-ferienvermietungslizenz-pool-und-traumhaftem-panoramablick-in-moscari-selva-mallorca-4160025903", 7, 16000, 390.0, 2900000, "Moscari, Selva, Mallorca, Spain", "Selva — Finca Vermietlizenz; Panoramablick"),
    (380, "Luxusvilla Projekt Son Vida", "https://www.von-poll.com/de/expose/mallorca-santa-maria/wunderschones-projekt-zum-bau-einer-modernen-luxusvilla-in-der-exklusivsten-gegend-von-palma-in-son-vida-4160034679", 8, 2452, 553.0, 2900000, "Son Vida, Palma, Mallorca, Spain", "Son Vida — Luxusvilla Projekt; Exklusivste Lage"),
    (381, "Neubauprojekt Son Vida Traumblick", "https://www.von-poll.com/de/expose/mallorca-santa-maria/neubauprojekt-fur-eine-luxusvilla-mit-pool-und-traumblick-in-son-vida-mallorca-4160034665", 8, 2960, 659.0, 2900000, "Son Vida, Palma, Mallorca, Spain", "Son Vida — Neubauprojekt; Traumblick Pool"),
    (382, "Luxusvilla Projekt Son Vida 2", "https://www.von-poll.com/de/expose/mallorca-santa-maria/projekt-zum-bau-einer-luxusvilla-in-exklusiver-lage-von-palma-in-son-vida-4160034673", 8, 2380, 664.0, 2700000, "Son Vida, Palma, Mallorca, Spain", "Son Vida — Luxusvilla Projekt; Exklusive Lage"),
    (383, "Villa Projekt Son Vida 3", "https://www.von-poll.com/de/expose/mallorca-santa-maria/grossartiges-projekt-zum-bau-einer-herrlichen-villa-mit-pool-in-son-vida-mallorca-4160034659", 7, 2017, 664.0, 2600000, "Son Vida, Palma, Mallorca, Spain", "Son Vida — Villa Projekt; Herrlicher Pool"),
    (384, "Doppelprojekt Orient Panorama", "https://www.von-poll.com/de/expose/mallorca-santa-maria/zwei-einzigartige-projekte-in-orient-mit-panoramablick-4160044053", 12, 577091, 0.0, 2400000, "Orient, Bunyola, Mallorca, Spain", "Orient — Doppelprojekt; Panoramablick"),
    (385, "Finca Hotel Tramuntana Weingut", "https://www.von-poll.com/de/expose/mallorca-pollensa/secret-marketing-exklusives-finca-mit-hotel-lizenz-und-biologischem-weingut-in-der-serra-de-tramuntana-3520045049", 27, 720000, 1900.0, 16000000, "Soller, Mallorca, Spain", "Tramuntana — Finca Hotel; Bio-Weingut Lizenz"),
    (386, "Historisches Anwesen Pollença 1920", "https://www.von-poll.com/de/expose/mallorca-pollensa/einzigartiges-historisches-luxusanwesen-von-1920-mit-meerblick-an-der-bucht-von-pollensa-vor-zwei-traumstranden-gelegen-3520036561", 20, 18000, 1454.0, 15500000, "Pollensa, Mallorca, Spain", "Pollença — Historisches Anwesen 1920; Meerblick"),
    (387, "Strandvilla Mal Pas Alcúdia", "https://www.von-poll.com/de/expose/mallorca-pollensa/atemberaubende-luxusvilla-in-erster-linie-mit-direktem-strandzugang-in-mal-pas-alcudia-an-der-bucht-von-pollensa-3520033675", 7, 2874, 1000.0, 14000000, "Mal Pas, Alcudia, Mallorca, Spain", "Alcúdia — Strandvilla Mal Pas; Direkter Strandzugang"),
    (388, "Landgut Aixartell Pollença", "https://www.von-poll.com/de/expose/mallorca-pollensa/herrliches-landgut-mit-pool-und-panoramablick-im-wunderschonen-tal-von-aixartell-pollensa-3520035683", 10, 408190, 2179.0, 9950000, "Pollensa, Mallorca, Spain", "Pollença — Landgut Aixartell; Panoramablick"),
    (389, "Villa Mal Pas Alcúdia Premium", "https://www.von-poll.com/de/expose/mallorca-pollensa/spektakulare-villa-in-bester-lage-von-alcudia-in-mal-pas-3520045101", 8, 4217, 488.0, 8995000, "Mal Pas, Alcudia, Mallorca, Spain", "Alcúdia — Villa Mal Pas; Beste Lage"),
    (390, "Meereslinie Villa Sa Torre Llucmajor", "https://www.von-poll.com/de/expose/mallorca-pollensa/luxusvilla-in-erster-meereslinie-an-der-kuste-von-sa-torre-llucmajor-3520047335", 7, 1300, 750.0, 8250000, "Sa Torre, Llucmajor, Mallorca, Spain", "Llucmajor — Villa Sa Torre; Erste Meereslinie"),
    (391, "Jagdanwesen Artà Hotel Projekt", "https://www.von-poll.com/de/expose/mallorca-pollensa/traumhaftes-historisches-jagdanwesen-in-arta-in-bei-colonia-sant-pere-mit-projekt-fur-ein-luxuslandhotel-mit-spa-3520037285", 19, 5442649, 1630.0, 7800000, "Arta, Mallorca, Spain", "Artà — Jagdanwesen; Landhotel Projekt"),
    (392, "Neubau Villa Alcúdia Bonaire", "https://www.von-poll.com/de/expose/mallorca-pollensa/spektakulare-neu-gebaute-villa-in-einer-privilegierten-wohngegend-von-alcudia-in-paraiso-de-bonaire-3520049259", 7, 2000, 470.0, 4500000, "Bonaire, Alcudia, Mallorca, Spain", "Alcúdia — Neubau Villa Bonaire; Privilegierte Lage"),
    (393, "Luxusvilla Canyamel Meerblick", "https://www.von-poll.com/de/expose/mallorca-pollensa/spektakulare-luxusvilla-mit-unverbaubarem-meerblick-in-canyamel-3520044205", 8, 887, 545.0, 4500000, "Canyamel, Capdepera, Mallorca, Spain", "Canyamel — Luxusvilla; Unverbaubarer Meerblick"),
    (394, "Immobilie Muro Alcúdia Bucht", "https://www.von-poll.com/de/expose/mallorca-pollensa/secret-marketing-interessante-immobilie-mit-spektakularem-blick-auf-die-bucht-von-alcudia-in-muro-3520048123", 7, 26900, 800.0, 4200000, "Muro, Mallorca, Spain", "Muro — Immobilie Alcúdia Bucht; Spektakulärer Blick"),
    (395, "Landhaus Colònia Sant Pere Meer", "https://www.von-poll.com/de/expose/mallorca-pollensa/wunderschones-landhaus-in-der-nahe-von-colonia-sant-pere-mit-panoramablick-auf-das-meer-3520042991", 9, 14300, 366.0, 3800000, "Colonia de Sant Pere, Arta, Mallorca, Spain", "Colònia Sant Pere — Landhaus; Meerblick"),
    (396, "Neubau Finca Alcúdia Meerblick", "https://www.von-poll.com/de/expose/mallorca-pollensa/fabelhafte-neu-gebaute-finca-mit-meerblick-in-alcudia-mallorca-3520042635", 8, 26456, 439.0, 3650000, "Alcudia, Mallorca, Spain", "Alcúdia — Neubau Finca; Meerblick"),
    (397, "Villa Pollença Berge Aussicht", "https://www.von-poll.com/de/expose/mallorca-pollensa/wunderschone-villa-in-den-bergen-von-pollensa-mit-spektakularer-aussicht-3520046987", 7, 1513, 618.0, 3500000, "Pollensa, Mallorca, Spain", "Pollença — Villa Berge; Spektakuläre Aussicht"),
    (398, "Herrenhaus Pollença Panorama", "https://www.von-poll.com/de/expose/mallorca-pollensa/herrliches-mallorquinischer-herrenhaus-mit-pool-und-fantastischem-panoramablick-in-pollensa-3520038669", 14, 76000, 1000.0, 3500000, "Pollensa, Mallorca, Spain", "Pollença — Mallorquinisches Herrenhaus; Panoramablick"),
    (399, "Villa Portocolom Meer Gästewohnung", "https://www.von-poll.com/de/expose/mallorca-pollensa/moderne-villa-mit-gasteapartment-direkt-am-meer-in-portocolom-3520044169", 9, 2242, 427.0, 3200000, "Portocolom, Felanitx, Mallorca, Spain", "Portocolom — Villa am Meer; Gästeapartment"),
    (400, "Historische Finca Sant Llorenç Hotel", "https://www.von-poll.com/de/expose/mallorca-pollensa/fantastische-historische-finca-in-sant-llorenc-des-cardassar-mit-der-moglichkeit-einer-lizenz-fur-ein-kleines-landhotel-3520028253", 27, 212511, 1345.0, 2750000, "Sant Llorenc des Cardassar, Mallorca, Spain", "Sant Llorenç — Historische Finca; Landhotel Lizenz"),
]

def get_distances(origin_address):
    """Query Google Maps Distance Matrix for all 4 reference points"""
    destinations = "|".join(REFS.values())
    url = "https://maps.googleapis.com/maps/api/distancematrix/json"
    params = {
        "origins": origin_address,
        "destinations": destinations,
        "mode": "driving",
        "key": GMAPS_KEY,
        "units": "metric",
        "language": "de",
    }
    resp = requests.get(url, params=params, timeout=10)
    data = resp.json()
    
    results = {}
    if data.get("status") == "OK":
        elements = data["rows"][0]["elements"]
        ref_keys = list(REFS.keys())
        for i, key in enumerate(ref_keys):
            el = elements[i]
            if el.get("status") == "OK":
                km = round(el["distance"]["value"] / 1000, 1)
                mins = round(el["duration"]["value"] / 60)
                results[key] = (km, mins)
            else:
                results[key] = (None, None)
    else:
        print(f"  ERROR: {data.get('status')} for {origin_address}")
        for key in REFS:
            results[key] = (None, None)
    return results

def calc_erreichbarkeit(distances):
    mins = [v[1] for v in distances.values() if v[1] is not None]
    if not mins:
        return None
    max_min = max(mins)
    score = 100 - max(0, (max_min - 30) * 1.5)
    return round(max(0, min(100, score)), 1)

# Load workbook
print("Loading workbook...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Mallorca Kandidaten"]

print(f"Current rows: {ws.max_row}")

for obj in OBJECTS:
    nr, name, url, zimmer, grundstuck, bebaut, preis, location, anzeigename = obj
    
    print(f"Processing {nr}: {name} ({location})")
    
    # Get distances
    distances = get_distances(location)
    erreichbarkeit = calc_erreichbarkeit(distances)
    
    print(f"  Distances: {distances}, Erreichbarkeit: {erreichbarkeit}")
    
    # Row number (nr 366 = row 367 if row 1 is header, etc.)
    # Find next empty row
    row_num = nr + 1  # assuming row 1 = header, row 2 = entry #1... 
    # But let's just append after current last row
    # Actually: nr=366 means it should be at excel row 367 (header at row 1)
    
    row = [None] * 40
    row[0] = nr          # Col 1: Ordnungsnummer
    row[1] = name        # Col 2: Name
    row[2] = url         # Col 3: URL
    row[4] = zimmer      # Col 5: Zimmer
    row[6] = int(grundstuck)  # Col 7: Grundstück
    row[7] = float(bebaut) if bebaut else None  # Col 8: Bebaut
    row[18] = preis      # Col 19: Preis
    
    # Distances
    d = distances
    row[10] = d["Flughafen"][0]   # Col 11: Flughafen km
    row[11] = d["Flughafen"][1]   # Col 12: Flughafen min
    row[12] = d["Deia"][0]        # Col 13: Deia km
    row[13] = d["Deia"][1]        # Col 14: Deia min
    row[14] = d["Andratx"][0]     # Col 15: Andratx km
    row[15] = d["Andratx"][1]     # Col 16: Andratx min
    row[16] = d["SesSalines"][0]  # Col 17: Ses Salines km
    row[17] = d["SesSalines"][1]  # Col 18: Ses Salines min
    
    row[24] = erreichbarkeit  # Col 25: Erreichbarkeit
    row[35] = "Von Poll"      # Col 36: Makler
    row[37] = "active"        # Col 38: Link Status
    row[38] = anzeigename     # Col 39: Anzeigename
    
    ws.append(row)
    time.sleep(0.1)  # small delay between API calls

print(f"\nFinal row count: {ws.max_row}")
print("Saving...")
wb.save(EXCEL_PATH)
print("Done!")
