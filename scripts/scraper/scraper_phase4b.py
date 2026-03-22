#!/usr/bin/env python3
"""Phase 4b Scraper - Makler Quellen mit Playwright"""

import sys
import json
import time
import re
from datetime import date
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'

def save_to_excel(new_objects, source_name):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
    
    new_count = 0
    for obj in new_objects:
        url = obj.get('url', '—')
        if url != '—' and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', ''),
            obj.get('quelle', source_name),
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            str(date.today()),
            'Neu'
        ])
        if url != '—':
            existing_urls.add(url)
        new_count += 1
    
    wb.save(EXCEL_PATH)
    print(f"  ✅ {new_count} neue Objekte aus {source_name} gespeichert")
    return new_count


def parse_price(text):
    if not text:
        return None
    nums = re.findall(r'[\d\.]+', text.replace(',', '.'))
    for n in nums:
        try:
            val = float(n.replace('.', ''))
            if val > 1000:
                return int(val)
        except:
            pass
    return None


def parse_int(text):
    if not text:
        return None
    nums = re.findall(r'\d+', str(text))
    return int(nums[0]) if nums else None


def run_playwright_scraper(fn):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-setuid-sandbox'])
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            viewport={'width': 1280, 'height': 800},
            locale='de-DE',
            extra_http_headers={'Accept-Language': 'de-DE,de;q=0.9,en;q=0.8'}
        )
        page = context.new_page()
        Stealth().apply_stealth_sync(page)
        
        try:
            result = fn(page)
        finally:
            browser.close()
        
        return result


# ============================================================
# SOURCE 4: Von Poll Real Estate
# ============================================================
def scrape_von_poll():
    print("\n🏠 SOURCE 4: Von Poll Real Estate")
    objects = []
    
    def scrape(page):
        url = 'https://www.von-poll.com/de/immobilien/spanien/balearen/mallorca'
        try:
            page.goto(url, wait_until='domcontentloaded', timeout=30000)
            time.sleep(3)
            
            # Accept cookies if present
            try:
                page.click('button:has-text("Akzeptieren")', timeout=3000)
                time.sleep(1)
            except:
                pass
            
            # Look for property cards
            page.wait_for_selector('.property-item, .expose-item, [class*="property"], [class*="listing"], article', timeout=10000)
            
            # Get all property links and data
            items = page.query_selector_all('.property-item, .expose-item, article.listing, [class*="PropertyItem"], [class*="property-card"]')
            print(f"  Elemente gefunden: {len(items)}")
            
            if not items:
                # Try to get all links
                links = page.query_selector_all('a[href*="/expose/"], a[href*="/immobilie/"]')
                print(f"  Links gefunden: {len(links)}")
                
                for link in links[:50]:
                    href = link.get_attribute('href')
                    if href:
                        if not href.startswith('http'):
                            href = 'https://www.von-poll.com' + href
                        objects.append({
                            'titel': link.inner_text()[:100],
                            'quelle': 'Von Poll Real Estate',
                            'url': href,
                            'preis': None,
                            'zimmer': None,
                            'grundstueck': None,
                            'wohnflaeche': None,
                            'ort': 'Mallorca',
                        })
            else:
                for item in items[:50]:
                    try:
                        title_el = item.query_selector('h2, h3, .title, .expose-title')
                        price_el = item.query_selector('.price, .preis, [class*="price"]')
                        link_el = item.query_selector('a')
                        
                        href = link_el.get_attribute('href') if link_el else None
                        if href and not href.startswith('http'):
                            href = 'https://www.von-poll.com' + href
                        
                        objects.append({
                            'titel': title_el.inner_text()[:100] if title_el else '',
                            'quelle': 'Von Poll Real Estate',
                            'url': href or '—',
                            'preis': parse_price(price_el.inner_text() if price_el else ''),
                            'zimmer': None,
                            'grundstueck': None,
                            'wohnflaeche': None,
                            'ort': 'Mallorca',
                        })
                    except Exception as e:
                        pass
            
            # Try pagination
            for pg in range(2, 6):
                try:
                    next_url = f'{url}?page={pg}'
                    page.goto(next_url, wait_until='domcontentloaded', timeout=20000)
                    time.sleep(2)
                    more = page.query_selector_all('a[href*="/expose/"], a[href*="/immobilie/"]')
                    before = len(objects)
                    for link in more[:50]:
                        href = link.get_attribute('href')
                        if href:
                            if not href.startswith('http'):
                                href = 'https://www.von-poll.com' + href
                            if href not in [o['url'] for o in objects]:
                                objects.append({
                                    'titel': link.inner_text()[:100],
                                    'quelle': 'Von Poll Real Estate',
                                    'url': href,
                                    'preis': None,
                                    'zimmer': None,
                                    'grundstueck': None,
                                    'wohnflaeche': None,
                                    'ort': 'Mallorca',
                                })
                    if len(objects) == before:
                        break
                except:
                    break
        
        except Exception as e:
            print(f"  Error: {e}")
        
        return objects
    
    result = run_playwright_scraper(scrape)
    print(f"  Gesammelt: {len(result)} Objekte")
    if result:
        return save_to_excel(result, 'Von Poll Real Estate')
    return 0


# ============================================================
# SOURCE 5: Knight Frank
# ============================================================
def scrape_knight_frank():
    print("\n🏠 SOURCE 5: Knight Frank")
    objects = []
    
    import requests
    # Try Knight Frank API first
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0',
            'Accept': 'application/json',
            'Content-Type': 'application/json',
        }
        
        # Knight Frank search API
        payload = {
            "searchParams": {
                "CountryIds": [68],  # Spain
                "RegionIds": [],
                "PropertyTypes": [],
                "MinPrice": None,
                "MaxPrice": None,
                "MinBedrooms": None,
                "SortOrder": "newest",
                "Page": 1,
                "PageSize": 24,
                "SearchQuery": "mallorca"
            }
        }
        
        resp = requests.post(
            'https://www.knightfrank.com/api/property/search',
            json=payload, headers=headers, timeout=15
        )
        print(f"  KF API status: {resp.status_code}")
        if resp.status_code == 200:
            data = resp.json()
            print(f"  Keys: {list(data.keys()) if isinstance(data, dict) else 'list'}")
    except Exception as e:
        print(f"  KF API error: {e}")
    
    def scrape(page):
        urls_to_try = [
            'https://www.knightfrank.com/residential-property-for-sale/spain/balearic-islands/mallorca',
            'https://www.knightfrank.com/international-residential-property-for-sale/spain/mallorca',
        ]
        
        for url in urls_to_try:
            try:
                page.goto(url, wait_until='domcontentloaded', timeout=30000)
                time.sleep(3)
                
                # Cookie consent
                try:
                    page.click('[id*="accept"], button:has-text("Accept")', timeout=3000)
                    time.sleep(1)
                except:
                    pass
                
                content = page.content()
                
                # Try to find property listings
                cards = page.query_selector_all('[class*="PropertyCard"], [class*="property-card"], [class*="listing-item"], article')
                print(f"  Cards auf {url}: {len(cards)}")
                
                for card in cards[:30]:
                    try:
                        link = card.query_selector('a')
                        href = link.get_attribute('href') if link else None
                        if href and not href.startswith('http'):
                            href = 'https://www.knightfrank.com' + href
                        
                        title_el = card.query_selector('h2, h3, [class*="title"]')
                        price_el = card.query_selector('[class*="price"], [class*="Price"]')
                        beds_el = card.query_selector('[class*="bed"], [data-beds]')
                        area_el = card.query_selector('[class*="area"], [class*="size"]')
                        loc_el = card.query_selector('[class*="location"], [class*="address"]')
                        
                        objects.append({
                            'titel': title_el.inner_text()[:100] if title_el else '',
                            'quelle': 'Knight Frank',
                            'url': href or '—',
                            'preis': parse_price(price_el.inner_text() if price_el else ''),
                            'zimmer': parse_int(beds_el.inner_text() if beds_el else ''),
                            'grundstueck': None,
                            'wohnflaeche': parse_int(area_el.inner_text() if area_el else ''),
                            'ort': loc_el.inner_text()[:50] if loc_el else 'Mallorca',
                        })
                    except:
                        pass
                
                if objects:
                    break
                    
            except Exception as e:
                print(f"  Error on {url}: {e}")
        
        return objects
    
    result = run_playwright_scraper(scrape)
    print(f"  Gesammelt: {len(result)} Objekte")
    if result:
        return save_to_excel(result, 'Knight Frank')
    return 0


# ============================================================
# SOURCE 6: Savills
# ============================================================
def scrape_savills():
    print("\n🏠 SOURCE 6: Savills")
    objects = []
    
    import requests
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': 'application/json',
        'Content-Type': 'application/json',
        'x-api-key': '',
    }
    
    # Try Savills API
    try:
        # Savills search API endpoint
        url = 'https://search.savills.com/es/en/list?SearchList=IsInRegion_EUR5002&SaleType=ForSale&PropertyType=RES'
        resp = requests.get(url, headers=headers, timeout=15)
        print(f"  Savills page status: {resp.status_code}")
    except Exception as e:
        print(f"  Savills error: {e}")
    
    def scrape(page):
        try:
            url = 'https://search.savills.com/es/en/list?SearchList=IsInRegion_EUR5002&SaleType=ForSale&PropertyType=RES'
            page.goto(url, wait_until='domcontentloaded', timeout=30000)
            time.sleep(4)
            
            # Accept cookies
            try:
                page.click('button:has-text("Accept"), button:has-text("Aceptar"), #onetrust-accept-btn-handler', timeout=4000)
                time.sleep(2)
            except:
                pass
            
            # Wait for listings
            page.wait_for_selector('[class*="PropertyCard"], [class*="property-card"], [class*="ListingCard"]', timeout=15000)
            time.sleep(2)
            
            cards = page.query_selector_all('[class*="PropertyCard"], [class*="property-card"], [class*="ListingCard"]')
            print(f"  Savills cards: {len(cards)}")
            
            for card in cards[:50]:
                try:
                    link = card.query_selector('a')
                    href = link.get_attribute('href') if link else None
                    if href and not href.startswith('http'):
                        href = 'https://search.savills.com' + href
                    
                    title_el = card.query_selector('h2, h3, [class*="Title"], [class*="title"]')
                    price_el = card.query_selector('[class*="Price"], [class*="price"]')
                    beds_el = card.query_selector('[class*="Bed"], [class*="bed"]')
                    area_el = card.query_selector('[class*="Area"], [class*="area"], [class*="size"]')
                    loc_el = card.query_selector('[class*="Location"], [class*="location"], [class*="Address"]')
                    
                    objects.append({
                        'titel': title_el.inner_text()[:100] if title_el else '',
                        'quelle': 'Savills',
                        'url': href or '—',
                        'preis': parse_price(price_el.inner_text() if price_el else ''),
                        'zimmer': parse_int(beds_el.inner_text() if beds_el else ''),
                        'grundstueck': None,
                        'wohnflaeche': parse_int(area_el.inner_text() if area_el else ''),
                        'ort': loc_el.inner_text()[:50] if loc_el else 'Mallorca',
                    })
                except:
                    pass
            
            # Try next pages
            for pg in range(2, 5):
                try:
                    next_btn = page.query_selector('[aria-label="Next page"], [class*="next"], button:has-text("Next")')
                    if next_btn:
                        next_btn.click()
                        time.sleep(3)
                        more_cards = page.query_selector_all('[class*="PropertyCard"], [class*="property-card"], [class*="ListingCard"]')
                        before = len(objects)
                        for card in more_cards:
                            try:
                                link = card.query_selector('a')
                                href = link.get_attribute('href') if link else None
                                if href and not href.startswith('http'):
                                    href = 'https://search.savills.com' + href
                                if href not in [o['url'] for o in objects]:
                                    title_el = card.query_selector('h2, h3, [class*="Title"]')
                                    price_el = card.query_selector('[class*="Price"]')
                                    objects.append({
                                        'titel': title_el.inner_text()[:100] if title_el else '',
                                        'quelle': 'Savills',
                                        'url': href or '—',
                                        'preis': parse_price(price_el.inner_text() if price_el else ''),
                                        'zimmer': None,
                                        'grundstueck': None,
                                        'wohnflaeche': None,
                                        'ort': 'Mallorca',
                                    })
                            except:
                                pass
                        if len(objects) == before:
                            break
                    else:
                        break
                except:
                    break
        
        except Exception as e:
            print(f"  Savills playwright error: {e}")
            import traceback; traceback.print_exc()
        
        return objects
    
    result = run_playwright_scraper(scrape)
    print(f"  Gesammelt: {len(result)} Objekte")
    if result:
        return save_to_excel(result, 'Savills')
    return 0


# ============================================================
# SOURCE 7: Balearic Properties
# ============================================================
def scrape_balearic_properties():
    print("\n🏠 SOURCE 7: Balearic Properties")
    objects = []
    
    def scrape(page):
        try:
            page.goto('https://www.balearicproperties.com/properties-for-sale/mallorca', 
                     wait_until='domcontentloaded', timeout=30000)
            time.sleep(3)
            
            # Cookie consent
            try:
                page.click('button:has-text("Accept"), button:has-text("OK"), .accept-cookies', timeout=3000)
                time.sleep(1)
            except:
                pass
            
            # Get all property links/cards
            cards = page.query_selector_all('[class*="property"], [class*="listing"], article, .item, li.property')
            links = page.query_selector_all('a[href*="/property/"], a[href*="/properties/"]')
            
            print(f"  Cards: {len(cards)}, Links: {len(links)}")
            
            seen = set()
            for link in links[:100]:
                try:
                    href = link.get_attribute('href')
                    if not href or href in seen:
                        continue
                    seen.add(href)
                    if not href.startswith('http'):
                        href = 'https://www.balearicproperties.com' + href
                    
                    parent = link
                    # Try to get price from parent
                    price_text = ''
                    for _ in range(5):
                        parent_el = page.evaluate('el => el.parentElement', parent)
                        if parent_el:
                            txt = link.inner_text()
                            if '€' in txt or 'EUR' in txt:
                                price_text = txt
                                break
                    
                    objects.append({
                        'titel': link.inner_text()[:100],
                        'quelle': 'Balearic Properties',
                        'url': href,
                        'preis': parse_price(price_text),
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
                except:
                    pass
            
            # If no links, try to get all text content
            if not objects:
                content = page.content()
                print(f"  Page content length: {len(content)}")
                # Extract URLs from content
                urls_found = re.findall(r'href=["\']([^"\']*(?:property|listing|immobilien)[^"\']*)["\']', content)
                for u in urls_found[:30]:
                    if not u.startswith('http'):
                        u = 'https://www.balearicproperties.com' + u
                    objects.append({
                        'titel': 'Balearic Properties Objekt',
                        'quelle': 'Balearic Properties',
                        'url': u,
                        'preis': None,
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
        
        except Exception as e:
            print(f"  Error: {e}")
        
        return objects
    
    result = run_playwright_scraper(scrape)
    print(f"  Gesammelt: {len(result)} Objekte")
    if result:
        return save_to_excel(result, 'Balearic Properties')
    return 0


# ============================================================
# SOURCE 8: Reiderstad Invest
# ============================================================
def scrape_reiderstad():
    print("\n🏠 SOURCE 8: Reiderstad Invest")
    objects = []
    
    def scrape(page):
        try:
            # First check the main site
            page.goto('https://www.reiderstadinvest.com', wait_until='domcontentloaded', timeout=20000)
            time.sleep(2)
            
            # Find property links
            links = page.query_selector_all('a[href*="property"], a[href*="objekt"], a[href*="fastighet"]')
            nav_links = page.query_selector_all('nav a, .menu a, header a')
            
            print(f"  Property links: {len(links)}, Nav: {len(nav_links)}")
            
            # Get all links to find properties page
            all_links = page.query_selector_all('a')
            for link in all_links:
                try:
                    href = link.get_attribute('href') or ''
                    text = link.inner_text().lower()
                    if any(kw in text for kw in ['mallorca', 'property', 'properties', 'sale', 'till salu', 'köp', 'buy']):
                        if href and not href.startswith('#'):
                            print(f"  Found: {text[:50]} -> {href[:80]}")
                except:
                    pass
            
            # Try common property paths
            for path in ['/en/properties', '/properties', '/mallorca', '/for-sale', '/buy']:
                try:
                    url = 'https://www.reiderstadinvest.com' + path
                    page.goto(url, wait_until='domcontentloaded', timeout=15000)
                    time.sleep(2)
                    
                    cards = page.query_selector_all('[class*="property"], article, .card, li.listing')
                    if cards:
                        print(f"  Cards found at {path}: {len(cards)}")
                        for card in cards[:30]:
                            try:
                                link = card.query_selector('a')
                                href = link.get_attribute('href') if link else None
                                if href and not href.startswith('http'):
                                    href = 'https://www.reiderstadinvest.com' + href
                                title_el = card.query_selector('h2, h3, .title')
                                price_el = card.query_selector('[class*="price"], .price')
                                objects.append({
                                    'titel': title_el.inner_text()[:100] if title_el else '',
                                    'quelle': 'Reiderstad Invest',
                                    'url': href or '—',
                                    'preis': parse_price(price_el.inner_text() if price_el else ''),
                                    'zimmer': None,
                                    'grundstueck': None,
                                    'wohnflaeche': None,
                                    'ort': 'Mallorca',
                                })
                            except:
                                pass
                        if objects:
                            break
                except:
                    pass
        
        except Exception as e:
            print(f"  Error: {e}")
        
        return objects
    
    result = run_playwright_scraper(scrape)
    print(f"  Gesammelt: {len(result)} Objekte")
    if result:
        return save_to_excel(result, 'Reiderstad Invest')
    return 0


# ============================================================
# SOURCE 9: The Agency RE
# ============================================================
def scrape_the_agency():
    print("\n🏠 SOURCE 9: The Agency RE")
    objects = []
    
    import requests
    try:
        # Try their API
        headers = {
            'User-Agent': 'Mozilla/5.0',
            'Accept': 'application/json',
        }
        resp = requests.get(
            'https://www.theagencyre.com/api/properties?type=sale&location=mallorca&limit=50',
            headers=headers, timeout=10
        )
        print(f"  API status: {resp.status_code}")
        if resp.status_code == 200:
            try:
                data = resp.json()
                print(f"  Keys: {list(data.keys()) if isinstance(data, dict) else type(data)}")
            except:
                pass
    except Exception as e:
        print(f"  API error: {e}")
    
    def scrape(page):
        try:
            page.goto('https://www.theagencyre.com/search?type=sale&location=mallorca',
                     wait_until='domcontentloaded', timeout=30000)
            time.sleep(4)
            
            # Cookie consent
            try:
                page.click('button:has-text("Accept"), [id*="accept"], .accept', timeout=3000)
                time.sleep(1)
            except:
                pass
            
            cards = page.query_selector_all('[class*="PropertyCard"], [class*="property-card"], [class*="listing"], article')
            links = page.query_selector_all('a[href*="/property/"], a[href*="/listing/"]')
            
            print(f"  Cards: {len(cards)}, Links: {len(links)}")
            
            seen = set()
            for link in links[:50]:
                try:
                    href = link.get_attribute('href')
                    if not href or href in seen:
                        continue
                    seen.add(href)
                    if not href.startswith('http'):
                        href = 'https://www.theagencyre.com' + href
                    objects.append({
                        'titel': link.inner_text()[:100],
                        'quelle': 'The Agency RE',
                        'url': href,
                        'preis': None,
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
                except:
                    pass
            
            for card in cards[:30]:
                try:
                    link = card.query_selector('a')
                    href = link.get_attribute('href') if link else None
                    if not href or href in seen:
                        continue
                    seen.add(href)
                    if not href.startswith('http'):
                        href = 'https://www.theagencyre.com' + href
                    
                    title_el = card.query_selector('h2, h3, [class*="title"]')
                    price_el = card.query_selector('[class*="price"]')
                    beds_el = card.query_selector('[class*="bed"]')
                    
                    objects.append({
                        'titel': title_el.inner_text()[:100] if title_el else '',
                        'quelle': 'The Agency RE',
                        'url': href or '—',
                        'preis': parse_price(price_el.inner_text() if price_el else ''),
                        'zimmer': parse_int(beds_el.inner_text() if beds_el else ''),
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
                except:
                    pass
        
        except Exception as e:
            print(f"  Error: {e}")
        
        return objects
    
    result = run_playwright_scraper(scrape)
    print(f"  Gesammelt: {len(result)} Objekte")
    if result:
        return save_to_excel(result, 'The Agency RE')
    return 0


# ============================================================
# SOURCE 10: Mallorca Finest
# ============================================================
def scrape_mallorca_finest():
    print("\n🏠 SOURCE 10: Mallorca Finest")
    objects = []
    
    def scrape(page):
        try:
            page.goto('https://www.mallorcafinest.com/properties',
                     wait_until='domcontentloaded', timeout=30000)
            time.sleep(3)
            
            # Cookie
            try:
                page.click('button:has-text("Accept"), .accept, #accept-cookies', timeout=3000)
                time.sleep(1)
            except:
                pass
            
            cards = page.query_selector_all('[class*="property"], [class*="listing"], article, .item, li.property-item')
            links = page.query_selector_all('a[href*="/property"], a[href*="/properties/"], a[href*="/buy/"]')
            
            print(f"  Cards: {len(cards)}, Links: {len(links)}")
            
            seen = set()
            for link in links[:100]:
                try:
                    href = link.get_attribute('href')
                    if not href or href in seen or href == 'https://www.mallorcafinest.com/properties':
                        continue
                    seen.add(href)
                    if not href.startswith('http'):
                        href = 'https://www.mallorcafinest.com' + href
                    text = link.inner_text().strip()
                    if len(text) > 3:
                        objects.append({
                            'titel': text[:100],
                            'quelle': 'Mallorca Finest',
                            'url': href,
                            'preis': None,
                            'zimmer': None,
                            'grundstueck': None,
                            'wohnflaeche': None,
                            'ort': 'Mallorca',
                        })
                except:
                    pass
            
            # Process cards for better data
            for card in cards[:50]:
                try:
                    link = card.query_selector('a')
                    href = link.get_attribute('href') if link else None
                    if not href:
                        continue
                    if not href.startswith('http'):
                        href = 'https://www.mallorcafinest.com' + href
                    
                    title_el = card.query_selector('h2, h3, h4, .title, [class*="title"]')
                    price_el = card.query_selector('[class*="price"], .price')
                    beds_el = card.query_selector('[class*="bed"], [class*="room"]')
                    area_el = card.query_selector('[class*="area"], [class*="size"], [class*="sqm"]')
                    loc_el = card.query_selector('[class*="location"], [class*="area-name"], [class*="place"]')
                    
                    if href not in seen:
                        seen.add(href)
                        objects.append({
                            'titel': title_el.inner_text()[:100] if title_el else '',
                            'quelle': 'Mallorca Finest',
                            'url': href,
                            'preis': parse_price(price_el.inner_text() if price_el else ''),
                            'zimmer': parse_int(beds_el.inner_text() if beds_el else ''),
                            'grundstueck': None,
                            'wohnflaeche': parse_int(area_el.inner_text() if area_el else ''),
                            'ort': loc_el.inner_text()[:50] if loc_el else 'Mallorca',
                        })
                except:
                    pass
        
        except Exception as e:
            print(f"  Error: {e}")
        
        return objects
    
    result = run_playwright_scraper(scrape)
    print(f"  Gesammelt: {len(result)} Objekte")
    if result:
        return save_to_excel(result, 'Mallorca Finest')
    return 0


# ============================================================
# SOURCE 11: Pollensa Properties
# ============================================================
def scrape_pollensa_properties():
    print("\n🏠 SOURCE 11: Pollensa Properties")
    objects = []
    
    def scrape(page):
        try:
            page.goto('https://www.pollensaproperties.com/properties-for-sale',
                     wait_until='domcontentloaded', timeout=30000)
            time.sleep(3)
            
            # Cookie
            try:
                page.click('button:has-text("Accept"), .accept', timeout=3000)
                time.sleep(1)
            except:
                pass
            
            cards = page.query_selector_all('[class*="property"], article, .property-item, li.listing')
            links = page.query_selector_all('a[href*="/property/"], a[href*="/properties/"]')
            
            print(f"  Cards: {len(cards)}, Links: {len(links)}")
            
            seen = set()
            
            def add_from_card(card):
                try:
                    link = card.query_selector('a')
                    href = link.get_attribute('href') if link else None
                    if not href or href in seen:
                        return
                    seen.add(href)
                    if not href.startswith('http'):
                        href = 'https://www.pollensaproperties.com' + href
                    
                    title_el = card.query_selector('h2, h3, h4, .title')
                    price_el = card.query_selector('[class*="price"], .price')
                    beds_el = card.query_selector('[class*="bed"], [class*="bedroom"]')
                    area_el = card.query_selector('[class*="area"], [class*="size"]')
                    
                    objects.append({
                        'titel': title_el.inner_text()[:100] if title_el else '',
                        'quelle': 'Pollensa Properties',
                        'url': href,
                        'preis': parse_price(price_el.inner_text() if price_el else ''),
                        'zimmer': parse_int(beds_el.inner_text() if beds_el else ''),
                        'grundstueck': None,
                        'wohnflaeche': parse_int(area_el.inner_text() if area_el else ''),
                        'ort': 'Pollença / Mallorca',
                    })
                except:
                    pass
            
            for card in cards[:50]:
                add_from_card(card)
            
            for link in links[:50]:
                try:
                    href = link.get_attribute('href')
                    if not href or href in seen:
                        continue
                    seen.add(href)
                    if not href.startswith('http'):
                        href = 'https://www.pollensaproperties.com' + href
                    objects.append({
                        'titel': link.inner_text()[:100],
                        'quelle': 'Pollensa Properties',
                        'url': href,
                        'preis': None,
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Pollença / Mallorca',
                    })
                except:
                    pass
            
            # Pagination
            for pg in range(2, 8):
                try:
                    next_url = f'https://www.pollensaproperties.com/properties-for-sale?page={pg}'
                    page.goto(next_url, wait_until='domcontentloaded', timeout=15000)
                    time.sleep(2)
                    before = len(objects)
                    new_cards = page.query_selector_all('[class*="property"], article, .property-item')
                    new_links = page.query_selector_all('a[href*="/property/"], a[href*="/properties/"]')
                    for card in new_cards:
                        add_from_card(card)
                    for link in new_links:
                        try:
                            href = link.get_attribute('href')
                            if not href or href in seen:
                                continue
                            seen.add(href)
                            if not href.startswith('http'):
                                href = 'https://www.pollensaproperties.com' + href
                            objects.append({
                                'titel': link.inner_text()[:100],
                                'quelle': 'Pollensa Properties',
                                'url': href,
                                'preis': None, 'zimmer': None, 'grundstueck': None,
                                'wohnflaeche': None, 'ort': 'Pollença / Mallorca',
                            })
                        except:
                            pass
                    if len(objects) == before:
                        break
                except:
                    break
        
        except Exception as e:
            print(f"  Error: {e}")
        
        return objects
    
    result = run_playwright_scraper(scrape)
    print(f"  Gesammelt: {len(result)} Objekte")
    if result:
        return save_to_excel(result, 'Pollensa Properties')
    return 0


# ============================================================
# SOURCE 12: Vives Pons Developer
# ============================================================
def scrape_vives_pons():
    print("\n🏠 SOURCE 12: Vives Pons")
    objects = []
    
    def scrape(page):
        try:
            page.goto('https://www.vivespons.com/en/properties',
                     wait_until='domcontentloaded', timeout=30000)
            time.sleep(3)
            
            # Cookie
            try:
                page.click('button:has-text("Accept"), .accept', timeout=3000)
                time.sleep(1)
            except:
                pass
            
            cards = page.query_selector_all('[class*="property"], [class*="development"], article, .project-item, .listing')
            links = page.query_selector_all('a[href*="/property/"], a[href*="/properties/"], a[href*="/project/"]')
            
            print(f"  Cards: {len(cards)}, Links: {len(links)}")
            
            seen = set()
            for card in cards[:50]:
                try:
                    link = card.query_selector('a')
                    href = link.get_attribute('href') if link else None
                    if not href or href in seen:
                        continue
                    seen.add(href)
                    if not href.startswith('http'):
                        href = 'https://www.vivespons.com' + href
                    
                    title_el = card.query_selector('h2, h3, h4, .title, [class*="title"]')
                    price_el = card.query_selector('[class*="price"], .price')
                    area_el = card.query_selector('[class*="area"], [class*="size"]')
                    loc_el = card.query_selector('[class*="location"], [class*="area"], [class*="zone"]')
                    
                    objects.append({
                        'titel': title_el.inner_text()[:100] if title_el else '',
                        'quelle': 'Vives Pons',
                        'url': href,
                        'preis': parse_price(price_el.inner_text() if price_el else ''),
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': parse_int(area_el.inner_text() if area_el else ''),
                        'ort': loc_el.inner_text()[:50] if loc_el else 'Mallorca',
                    })
                except:
                    pass
            
            for link in links[:50]:
                try:
                    href = link.get_attribute('href')
                    if not href or href in seen:
                        continue
                    seen.add(href)
                    if not href.startswith('http'):
                        href = 'https://www.vivespons.com' + href
                    objects.append({
                        'titel': link.inner_text()[:100],
                        'quelle': 'Vives Pons',
                        'url': href,
                        'preis': None, 'zimmer': None, 'grundstueck': None,
                        'wohnflaeche': None, 'ort': 'Mallorca',
                    })
                except:
                    pass
        
        except Exception as e:
            print(f"  Error: {e}")
        
        return objects
    
    result = run_playwright_scraper(scrape)
    print(f"  Gesammelt: {len(result)} Objekte")
    if result:
        return save_to_excel(result, 'Vives Pons')
    return 0


# ============================================================
# Main
# ============================================================
if __name__ == '__main__':
    results = {}
    
    results['Von Poll RE'] = scrape_von_poll()
    results['Knight Frank'] = scrape_knight_frank()
    results['Savills'] = scrape_savills()
    results['Balearic Properties'] = scrape_balearic_properties()
    results['Reiderstad Invest'] = scrape_reiderstad()
    results['The Agency RE'] = scrape_the_agency()
    results['Mallorca Finest'] = scrape_mallorca_finest()
    results['Pollensa Properties'] = scrape_pollensa_properties()
    results['Vives Pons'] = scrape_vives_pons()
    
    print("\n=== ERGEBNIS PLAYWRIGHT-QUELLEN ===")
    for src, count in results.items():
        status = "✅" if count > 0 else "❌"
        print(f"  {status} {src}: {count} Objekte")
