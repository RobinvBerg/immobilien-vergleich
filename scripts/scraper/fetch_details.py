#!/usr/bin/env python3
"""
fetch_details.py — Grundstücksgröße (m²) für blinde Objekte nachladen.
Reihenfolge: Kleine Quellen → Rightmove → Fotocasa → Idealista
Progress-Tracking in fetchdetails_progress.json (Resume-fähig)
"""

import sys, time, json, re, requests
from pathlib import Path
from collections import Counter
import openpyxl
from bs4 import BeautifulSoup

sys.stdout.reconfigure(line_buffering=True)

EXCEL_PATH    = Path("/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx")
PROGRESS_PATH = Path("/Users/robin/.openclaw/workspace/mallorca-projekt/fetchdetails_progress.json")
SAVE_EVERY    = 50

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
HEADERS = {
    "User-Agent": UA,
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ── Progress tracking ────────────────────────────────────────────────────────
def load_progress():
    if PROGRESS_PATH.exists():
        return json.loads(PROGRESS_PATH.read_text())
    return {}  # url -> {"plot": int_or_null, "done": true}

def save_progress(prog):
    PROGRESS_PATH.write_text(json.dumps(prog, ensure_ascii=False, indent=2))

# ── Workbook ─────────────────────────────────────────────────────────────────
print("📂 Loading workbook...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

def save_wb():
    wb.save(EXCEL_PATH)
    print("  💾 Gespeichert", flush=True)

# ── Collect blind rows ───────────────────────────────────────────────────────
blind_rows = []
for r in range(2, ws.max_row + 1):
    titel      = str(ws.cell(r, 1).value or '')
    quelle     = str(ws.cell(r, 2).value or '')
    url        = str(ws.cell(r, 3).value or '')
    preis_raw  = ws.cell(r, 4).value
    zimmer_raw = ws.cell(r, 5).value
    grundstueck= ws.cell(r, 6).value
    ort        = str(ws.cell(r, 8).value or '')
    try:    preis = float(str(preis_raw).replace('.','').replace(',','.')) if preis_raw else 0
    except: preis = 0
    try:    zimmer = int(zimmer_raw) if zimmer_raw else 0
    except: zimmer = 0
    if (zimmer >= 5 and preis >= 2900000
            and 'Nordost' not in titel and 'Nordost' not in ort
            and (grundstueck is None or grundstueck == '' or grundstueck == 0)
            and url.startswith('http')):
        blind_rows.append({'row': r, 'quelle': quelle, 'url': url})

print(f"🔍 Blinde Objekte: {len(blind_rows)}")
src_count = Counter(b['quelle'] for b in blind_rows)
print("Quellen:", dict(src_count))

progress = load_progress()
already_done = sum(1 for b in blind_rows if b['url'] in progress)
print(f"📋 Bereits verarbeitet (Resume): {already_done}/{len(blind_rows)}")

# ── Helper: extract plot from HTML text ─────────────────────────────────────
PLOT_PATTERNS = [
    # Spanish
    r'(?:parcela|solar|terreno)\s*[:\-–]?\s*([\d\.]+)\s*m[²2]',
    r'([\d\.]+)\s*m[²2]\s*(?:de\s+)?(?:parcela|solar|terreno)',
    # English
    r'[Pp]lot\s*(?:size|area|:|-|–)?\s*([\d,\.]+)\s*(?:sq\s*m|m[²2])',
    r'([\d,\.]+)\s*(?:sq\s*m|m[²2])\s*[Pp]lot',
    r'[Ll]and\s*(?:area|size|:|-|–)?\s*([\d,\.]+)\s*(?:sq\s*m|m[²2])',
    # Generic m² with large values
    r'[Gg]round(?:s)?\s*[:\-–]?\s*([\d,\.]+)\s*m[²2]',
]

def extract_plot_from_text(text):
    for pat in PLOT_PATTERNS:
        m = re.search(pat, text)
        if m:
            raw = m.group(1).replace(',', '').replace('.', '') if '.' in m.group(1) and ',' not in m.group(1) else m.group(1).replace(',', '')
            try:
                val = int(float(raw))
                if 50 <= val <= 10_000_000:
                    return val
            except:
                pass
    return None

def extract_plot_from_json(html_or_text):
    """Search JSON keys for plot area."""
    for key in ('plotArea', 'plot_area', 'landArea', 'land_area', 'lotSize',
                'sueloM2', 'parcela', 'surfacePlot', 'landSurface', 'plotSurface'):
        m = re.search(rf'"{key}"\s*:\s*(\d+)', html_or_text, re.I)
        if m:
            val = int(m.group(1))
            if 50 <= val <= 10_000_000:
                return val
    return None

def extract_plot_from_jsonld(soup):
    for script in soup.find_all('script', type='application/ld+json'):
        try:
            data = json.loads(script.string)
            if isinstance(data, list): data = data[0]
            for key in ('plotArea', 'landArea', 'lotSize'):
                if key in data:
                    val = int(float(data[key]))
                    if 50 <= val <= 10_000_000:
                        return val
            fs = data.get('floorSize', {})
            if isinstance(fs, dict) and fs.get('value'):
                val = int(float(fs['value']))
                if val > 200:  # likely plot, not living area
                    return val
        except:
            pass
    return None

# ══════════════════════════════════════════════════════════
# RIGHTMOVE (requests)
# ══════════════════════════════════════════════════════════
def fetch_rightmove(url):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        if resp.status_code != 200:
            return None
        soup = BeautifulSoup(resp.text, 'html.parser')
        text = soup.get_text(' ', strip=True)

        # JSON-LD first
        val = extract_plot_from_jsonld(soup)
        if val: return val

        # Embedded JSON
        val = extract_plot_from_json(resp.text)
        if val: return val

        # Text patterns
        val = extract_plot_from_text(text)
        if val: return val

        # Rightmove-specific: key features table
        for li in soup.find_all(['li', 'dt', 'dd', 'span']):
            t = li.get_text(' ', strip=True)
            val = extract_plot_from_text(t)
            if val: return val

        return None
    except Exception as e:
        return None

# ══════════════════════════════════════════════════════════
# GENERIC small sources (requests)
# ══════════════════════════════════════════════════════════
def fetch_generic(url):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        if resp.status_code != 200:
            return None
        soup = BeautifulSoup(resp.text, 'html.parser')
        text = soup.get_text(' ', strip=True)
        val = extract_plot_from_jsonld(soup)
        if val: return val
        val = extract_plot_from_json(resp.text)
        if val: return val
        val = extract_plot_from_text(text)
        if val: return val
        return None
    except:
        return None

# ══════════════════════════════════════════════════════════
# PLAYWRIGHT (Fotocasa + Idealista)
# ══════════════════════════════════════════════════════════
def make_browser():
    from playwright.sync_api import sync_playwright
    from playwright_stealth import stealth_sync
    pw = sync_playwright().start()
    browser = pw.chromium.launch(
        headless=True,
        args=[
            '--no-sandbox',
            '--disable-blink-features=AutomationControlled',
            '--disable-dev-shm-usage',
        ]
    )
    ctx = browser.new_context(
        user_agent=UA,
        locale='es-ES',
        viewport={'width': 1280, 'height': 900},
        extra_http_headers={
            'Accept-Language': 'es-ES,es;q=0.9,en;q=0.8',
        }
    )
    page = ctx.new_page()
    stealth_sync(page)
    return pw, browser, ctx, page

def fetch_with_playwright(page, url, source='generic', retry=True):
    try:
        page.goto(url, wait_until='domcontentloaded', timeout=30000)
        time.sleep(2)
        html = page.content()

        if 'captcha' in html.lower() or 'robot' in html.lower() or len(html) < 3000:
            if retry:
                print(f"    ⏳ Block detected, waiting 30s...")
                time.sleep(30)
                page.goto(url, wait_until='domcontentloaded', timeout=30000)
                time.sleep(3)
                html = page.content()
            else:
                return None

        soup = BeautifulSoup(html, 'html.parser')
        text = soup.get_text(' ', strip=True)

        # Source-specific extraction
        if source == 'fotocasa':
            # __NEXT_DATA__
            nd = soup.find('script', id='__NEXT_DATA__')
            if nd and nd.string:
                try:
                    raw = nd.string
                    val = extract_plot_from_json(raw)
                    if val: return val
                except:
                    pass

        elif source == 'idealista':
            # Idealista embeds data in script tags
            for script in soup.find_all('script'):
                if script.string and ('plotArea' in script.string or
                                       'superficie' in script.string.lower() or
                                       'parcela' in script.string.lower()):
                    val = extract_plot_from_json(script.string)
                    if val: return val

            # utag_data
            m = re.search(r'utag_data\s*=\s*({.*?});', html, re.S)
            if m:
                try:
                    ud = json.loads(m.group(1))
                    for k in ('plot_area', 'plotArea', 'land_area'):
                        if k in ud:
                            return int(float(ud[k]))
                except:
                    pass

        # Universal fallbacks
        val = extract_plot_from_jsonld(soup)
        if val: return val
        val = extract_plot_from_json(html)
        if val: return val
        val = extract_plot_from_text(text)
        if val: return val
        return None

    except Exception as e:
        return None

# ══════════════════════════════════════════════════════════
# MAIN PROCESSING
# ══════════════════════════════════════════════════════════
# Split by source
idealista_rows = [b for b in blind_rows if b['quelle'] == 'Idealista']
fotocasa_rows  = [b for b in blind_rows if b['quelle'] == 'Fotocasa']
rightmove_rows = [b for b in blind_rows if b['quelle'] == 'Rightmove']
pisos_rows     = [b for b in blind_rows if b['quelle'] == 'Pisos.com']
other_rows     = [b for b in blind_rows if b['quelle'] not in
                  ('Idealista', 'Fotocasa', 'Rightmove', 'Pisos.com')]

stats = {'found': 0, 'skipped': 0, 'failed': 0}

def process_batch(rows, fetch_fn, label, delay=1.5, playwright_source=None, page=None):
    pending = [b for b in rows if b['url'] not in progress]
    print(f"\n{'='*60}\n🏠 {label}: {len(rows)} total | {len(pending)} pending")
    if not pending:
        print("  ✅ Alle bereits verarbeitet")
        return

    local_count = 0
    for i, b in enumerate(pending):
        url = b['url']
        row = b['row']

        if playwright_source and page is not None:
            plot = fetch_with_playwright(page, url, source=playwright_source)
        else:
            plot = fetch_fn(url)

        progress[url] = {'plot': plot, 'done': True}

        if plot and plot > 0:
            ws.cell(row, 6, plot)
            stats['found'] += 1
            local_count += 1
        else:
            stats['failed'] += 1

        total_done = sum(1 for b2 in blind_rows if b2['url'] in progress)
        if (i + 1) % 10 == 0 or i == 0:
            print(f"  [{i+1}/{len(pending)}] +{local_count} Grundstücke | gesamt gefunden={stats['found']}")

        if (i + 1) % SAVE_EVERY == 0:
            save_wb()
            save_progress(progress)

        time.sleep(delay)

    save_wb()
    save_progress(progress)
    print(f"  ✅ {label} fertig: {local_count} neue Grundstücke")

# ── 1. Andere kleine Quellen (requests, kein Block-Risiko) ──────────────────
process_batch(other_rows,     fetch_generic,    "ANDERE (EV, Savills, ThinkSpain…)", delay=1.5)
process_batch(pisos_rows,     fetch_generic,    "PISOS.COM",                          delay=1.5)
process_batch(rightmove_rows, fetch_rightmove,  "RIGHTMOVE",                          delay=1.5)

# ── 2. Fotocasa via Playwright ───────────────────────────────────────────────
fotocasa_pending = [b for b in fotocasa_rows if b['url'] not in progress]
if fotocasa_pending:
    print(f"\n{'='*60}\n🏠 FOTOCASA ({len(fotocasa_rows)} total | {len(fotocasa_pending)} pending) via Playwright")
    pw, browser, ctx, page = make_browser()
    try:
        process_batch(fotocasa_rows, None, "FOTOCASA", delay=2.0,
                      playwright_source='fotocasa', page=page)
    finally:
        browser.close()
        pw.stop()
else:
    print(f"\n{'='*60}\n🏠 FOTOCASA — alle bereits verarbeitet ✅")

# ── 3. Idealista via Playwright + Stealth ───────────────────────────────────
idealista_pending = [b for b in idealista_rows if b['url'] not in progress]
if idealista_pending:
    print(f"\n{'='*60}\n🏠 IDEALISTA ({len(idealista_rows)} total | {len(idealista_pending)} pending) via Playwright+Stealth")
    pw, browser, ctx, page = make_browser()
    try:
        process_batch(idealista_rows, None, "IDEALISTA", delay=3.0,
                      playwright_source='idealista', page=page)
    finally:
        browser.close()
        pw.stop()
else:
    print(f"\n{'='*60}\n🏠 IDEALISTA — alle bereits verarbeitet ✅")

# ── Final summary ─────────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print(f"✅ FERTIG")
print(f"  📐 Grundstücksgröße gefunden:  {stats['found']}")
print(f"  ❌ Nicht gefunden/geblockt:     {stats['failed']}")

# Count new pool ≥7000m²
pool_count = 0
wb2 = openpyxl.load_workbook(EXCEL_PATH)
ws2 = wb2.active
for r in range(2, ws2.max_row + 1):
    try:
        preis  = float(str(ws2.cell(r,4).value or 0).replace('.','').replace(',','.'))
        zimmer = int(ws2.cell(r,5).value or 0)
        gs     = float(ws2.cell(r,6).value or 0)
        titel  = str(ws2.cell(r,1).value or '')
        ort    = str(ws2.cell(r,8).value or '')
        if zimmer>=5 and preis>=2900000 and gs>=7000 and 'Nordost' not in titel and 'Nordost' not in ort:
            pool_count += 1
    except:
        pass
print(f"  🏆 Neuer Pool (≥7.000m², ≥5Z, ≥2,9M€): {pool_count}")
