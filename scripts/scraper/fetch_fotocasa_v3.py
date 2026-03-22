#!/usr/bin/env python3
"""
fetch_fotocasa_v3.py
Fetches plot size (Terreno/Parcela) for all Fotocasa URLs in Excel using Playwright.
Re-processes all Fotocasa rows, even if previously done.
"""

import re
import time
import json
import sys
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

EXCEL_PATH = Path("/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx")
PROGRESS_PATH = Path("/Users/robin/.openclaw/workspace/mallorca-projekt/fetchdetails_progress.json")

PATTERNS = [
    r'[Tt]erreno[:\s]+([0-9][0-9.]+)',
    r'superficieParcela[":\s]+([0-9]+)',
    r'[Pp]arcela de ([0-9][0-9.]+)',
    r'[Ss]olar[:\s]+([0-9][0-9.]+)',
]

UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36'


def extract_plot(html):
    for pat in PATTERNS:
        m = re.search(pat, html)
        if m:
            raw = m.group(1).replace('.', '').replace(',', '')
            try:
                val = int(raw)
                if val >= 50:
                    return val
            except ValueError:
                continue
    return None


def find_columns(ws):
    """Find URL and plot column indices (1-based) from header row."""
    url_col = None
    plot_col = None
    for cell in ws[1]:
        if cell.value and 'url' in str(cell.value).lower():
            url_col = cell.column
        if cell.value and ('plot' in str(cell.value).lower() or 'terreno' in str(cell.value).lower() or 'grundst' in str(cell.value).lower()):
            plot_col = cell.column
    return url_col, plot_col


def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    url_col, plot_col = find_columns(ws)
    if url_col is None:
        print("ERROR: Could not find URL column in Excel header!")
        # Print headers for debugging
        headers = [c.value for c in ws[1]]
        print(f"Headers: {headers}")
        sys.exit(1)

    print(f"URL column: {url_col}, Plot column: {plot_col}")
    if plot_col is None:
        print("WARNING: No plot column found — will only print results, not save to Excel")

    # Collect all Fotocasa rows
    fotocasa_rows = []
    for row in ws.iter_rows(min_row=2):
        url_cell = row[url_col - 1]
        url = url_cell.value
        if url and 'fotocasa' in str(url).lower():
            fotocasa_rows.append(row)

    total = len(fotocasa_rows)
    print(f"Found {total} Fotocasa rows to process\n")

    # Load progress
    progress = {}
    if PROGRESS_PATH.exists():
        with open(PROGRESS_PATH) as f:
            progress = json.load(f)

    results = {"success": 0, "failed": 0, "updated": 0}
    save_counter = 0

    with sync_playwright() as p:
        for idx, row in enumerate(fotocasa_rows, 1):
            url = row[url_col - 1].value
            print(f"[{idx}/{total}] {url}", end=" → ", flush=True)

            plot = None
            try:
                browser = p.chromium.launch(headless=True)
                ctx = browser.new_context(
                    locale='es-ES',
                    user_agent=UA
                )
                page = ctx.new_page()
                page.goto(url, timeout=30000)
                time.sleep(1)  # let JS render
                html = page.content()
                browser.close()

                plot = extract_plot(html)
            except Exception as e:
                print(f"ERROR: {e}")
                results["failed"] += 1
                progress[url] = {"plot": None, "error": str(e)}
                try:
                    browser.close()
                except:
                    pass
                time.sleep(2)
                continue

            if plot:
                print(f"plot={plot}m²")
                results["success"] += 1
            else:
                print("plot=None")
                results["failed"] += 1

            progress[url] = {"plot": plot}

            # Write to Excel
            if plot_col and plot:
                row[plot_col - 1].value = plot
                results["updated"] += 1

            save_counter += 1
            if save_counter >= 50:
                wb.save(EXCEL_PATH)
                with open(PROGRESS_PATH, 'w') as f:
                    json.dump(progress, f, indent=2)
                print(f"  [Saved after {idx} rows]")
                save_counter = 0

            time.sleep(2)

    # Final save
    wb.save(EXCEL_PATH)
    with open(PROGRESS_PATH, 'w') as f:
        json.dump(progress, f, indent=2)

    print("\n=== DONE ===")
    print(f"Total Fotocasa rows: {total}")
    print(f"Successfully fetched plot: {results['success']}")
    print(f"Failed / no plot found:    {results['failed']}")
    print(f"Excel rows updated:        {results['updated']}")

    # Summary pool
    print("\n=== POOL (URLs with plot found) ===")
    for url, data in progress.items():
        if 'fotocasa' in url and data.get('plot'):
            print(f"  {data['plot']}m² — {url}")


if __name__ == "__main__":
    main()
