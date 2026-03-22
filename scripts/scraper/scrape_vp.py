#!/usr/bin/env python3
"""
scrape_vp.py — Von Poll Mallorca Finca Scraper
================================================
TEAM-PROZESS:
    1. Kira öffnet: https://www.von-poll.com/de/finca-kaufen/mallorca
    2. Robin filtert rechts: Zimmer≥5, Preis≥2.000.000, scrollt bis alle geladen
    3. Robin führt in DevTools Console aus:
       [...new Set([...document.querySelectorAll('a[href*="expose"]')].map(e=>e.href))].join('\\n')
    4. URLs in Datei speichern → python scrape_vp.py --urls /tmp/vp_urls.txt

USAGE:
    python scrape_vp.py --open                        # Suchseite öffnen
    python scrape_vp.py --urls /tmp/vp_urls.txt       # URLs aus Datei scrapen
    python scrape_vp.py --url https://...             # Einzelne URL

FELDER (vollautomatisch):
    Zimmer, Bäder, Wohnfläche, Grundstück, Preis, Baujahr, Haustyp, Zustand,
    ETV/Vermietlizenz, Gästehäuser, Location, Beschreibung, Makler-Ref,
    Hauptbild, Distanzen (4x km+min), Erreichbarkeit, €/m², Reno-Score,
    Gebäudestruktur, Anzeigename,
    Charme (Vision-Modell auf Hauptbild),
    Bewirtschaftung (geschätzt aus Grundstücksgröße + Typ)
"""

import argparse, re, os, sys, time, json, requests, io, base64
import openpyxl
from PIL import Image

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXCEL_PATH   = os.path.join(PROJECT_ROOT, 'data', 'mallorca-kandidaten-v2.xlsx')
BILDER_PATH  = os.path.join(PROJECT_ROOT, 'bilder')
MAPS_KEY     = 'AIzaSyAzplaFOGUrCygk3mrzS--wrHNIz0arKok'
ANTHROPIC_KEY = os.environ.get('ANTHROPIC_API_KEY', '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')

REFERENCE_POINTS = {
    'Flughafen':  '39.5517,2.7388',
    'Deia':       '39.7456,2.6489',
    'Andratx':    '39.5747,2.3818',
    'SesSalines': '39.3444,3.0503',
}

ETV_KEYWORDS   = ['ETV', 'Ferienvermiet', 'Tourismuslizenz', 'Hotellizenz',
                  'hotel-lizenz', 'Vermietlizenz', 'Mietlizenz', 'Ferienlizenz',
                  'licencia turística', 'alquiler vacacional']
GAESTE_KEYWORDS = ['Gästehaus', 'Gästeapartment', 'Nebengebäude', 'Einliegerwohnung',
                   'Gästewohnung', 'Annexgebäude', 'Gästevilla', 'Gästesuite',
                   'casa de huéspedes', 'annexe', 'guest house']


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_existing_urls(ws):
    urls = set()
    for row in range(2, ws.max_row + 1):
        url = ws.cell(row=row, column=3).value
        if url:
            urls.add(url.strip())
    return urls


def get_next_nr(ws):
    max_nr = 0
    for row in range(2, ws.max_row + 1):
        nr = ws.cell(row=row, column=1).value
        if nr and isinstance(nr, int):
            max_nr = max(max_nr, nr)
    return max_nr + 1


def get_distances(location_str):
    results = {}
    origins = requests.utils.quote(location_str)
    for name, dest in REFERENCE_POINTS.items():
        try:
            url = (f'https://maps.googleapis.com/maps/api/distancematrix/json'
                   f'?origins={origins}&destinations={dest}&mode=driving&key={MAPS_KEY}')
            r = requests.get(url, timeout=10).json()
            el = r['rows'][0]['elements'][0]
            if el['status'] == 'OK':
                results[name] = {
                    'km':  round(el['distance']['value'] / 1000, 1),
                    'min': round(el['duration']['value'] / 60)
                }
        except:
            pass
    return results


def erreichbarkeit(dist):
    if not dist: return None
    max_min = max(v['min'] for v in dist.values() if 'min' in v)
    return max(0, round(100 - max(0, (max_min - 30) * 1.5)))


def reno_score(zustand, baujahr):
    z = (zustand or '').lower()
    if 'erstbezug' in z or 'neuwertig' in z: return 100
    if 'projektiert' in z:                    return 90
    if 'saniert' in z or 'modernisiert' in z: return 75
    if 'gepflegt' in z:                       return 60
    if 'renovierungsbedürftig' in z:          return 20
    if baujahr and baujahr >= 2015:           return 95
    if baujahr and baujahr >= 2000:           return 70
    if baujahr and baujahr >= 1980:           return 50
    return 55


def bewirtschaftung(grundstueck, haustyp, beschreibung):
    """
    Schätzt Bewirtschaftungsaufwand (1=sehr aufwändig, 5=pflegeleicht).
    Kriterien: Grundstücksgröße, Typ, Hinweise auf Weinberg/Olivenhain/Pferde.
    """
    desc = (beschreibung or '').lower()
    g = grundstueck or 0
    aufwaendig = any(k in desc for k in ['weinberg', 'olivenhain', 'pferd', 'landwirtschaft', 'weinanbau', 'obstgarten'])
    if aufwaendig:           return 1
    if g > 200000:           return 1
    if g > 50000:            return 2
    if g > 20000:            return 3
    if g > 5000:             return 4
    return 5


def vision_analyse(img_path, beschreibung=''):
    """
    Claude Haiku Vision auf Hauptbild + Beschreibungstext.
    Liefert: charme (1-5), charme_grund, gebaeude_struktur, anzeigename
    """
    try:
        import anthropic as _anthropic
        client = _anthropic.Anthropic(api_key=ANTHROPIC_KEY)
        with open(img_path, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode()
        prompt = (
            "You are a luxury real estate expert for Mallorca fincas. Analyze this property image.\n\n"
            f"Listing description: {beschreibung[:300]}\n\n"
            "Answer exactly in this format:\n"
            "CHARME: X | GRUND: <one sentence in German — X must be 1-5 where 1=no appeal, 5=dream property>\n"
            "GEBAEUDE: <one sentence in German about building structure, style, materials, epoch, special features like guesthouse/tower/courtyard>\n"
            "ANZEIGENAME: <creative name in German format: Ort — Kurztitel; Highlight>"
        )
        msg = client.messages.create(
            model='claude-haiku-4-5',
            max_tokens=200,
            messages=[{'role': 'user', 'content': [
                {'type': 'image', 'source': {'type': 'base64', 'media_type': 'image/jpeg', 'data': b64}},
                {'type': 'text', 'text': prompt}
            ]}]
        )
        text = msg.content[0].text

        m = re.search(r'CHARME:\s*(\d)', text)
        charme = int(m.group(1)) if m else None

        m = re.search(r'GRUND:\s*(.+?)(?:\n|$)', text)
        charme_grund = m.group(1).strip() if m else ''

        m = re.search(r'GEBAEUDE:\s*(.+?)(?:\n|$)', text)
        gebaeude = m.group(1).strip() if m else ''

        m = re.search(r'ANZEIGENAME:\s*(.+?)(?:\n|$)', text)
        anzeigename = m.group(1).strip() if m else ''

        return charme, charme_grund, gebaeude, anzeigename
    except Exception as e:
        return None, str(e), None, None


def gebaeude_struktur_fallback(data):
    """Fallback wenn Vision nicht verfügbar: aus Kennzahlen zusammenbauen."""
    parts = []
    haustyp = data.get('haustyp', 'Objekt')
    baujahr = data.get('baujahr')
    zustand = data.get('zustand', '')
    bebaut  = data.get('wohnflaeche', 0) or 0
    grund   = data.get('grundstueck', 0) or 0

    if baujahr and baujahr >= 2018:    parts.append(f'Neubau-{haustyp} {baujahr}')
    elif baujahr and baujahr >= 2000:  parts.append(f'Moderner {haustyp} {baujahr}')
    elif baujahr and baujahr <= 1950:  parts.append(f'Histor. {haustyp} ~{baujahr}')
    elif baujahr:                      parts.append(f'{haustyp} {baujahr}')
    else:                              parts.append(haustyp)

    if zustand: parts.append(zustand)
    if bebaut:  parts.append(f'{int(bebaut)} m² Wfl.')
    if grund:
        parts.append(f'{grund:,.0f} m² Grund'.replace(',', '.'))
    if data.get('etv'):    parts.append('ETV-Lizenz')
    if data.get('gaeste'): parts.append('Gästehaus')
    return ', '.join(parts)


def extract_description(text):
    """Zweites Vorkommen von 'Objektbeschreibung' = echter Freitext."""
    idx1 = text.find('Objektbeschreibung')
    if idx1 < 0: return ''
    idx2 = text.find('Objektbeschreibung', idx1 + 20)
    if idx2 < 0: return ''
    after = text[idx2 + len('Objektbeschreibung'):].strip()
    m = re.match(r'\n+(.+?)(?:\n\n|\Z)', after, re.DOTALL)
    if m: return m.group(1).strip().replace('\n', ' ')[:500]
    return after[:500]


def download_image(img_url, path, cookies=None):
    try:
        r = requests.get(img_url, cookies=cookies or {}, timeout=15,
                         headers={'User-Agent': 'Mozilla/5.0'})
        img = Image.open(io.BytesIO(r.content)).convert('RGB')
        img.save(path, 'JPEG')
        return True
    except Exception as e:
        print(f'  Bild-Fehler: {e}')
        return False


# ---------------------------------------------------------------------------
# Scraper
# ---------------------------------------------------------------------------

def scrape_expose(page, url):
    page.goto(url, timeout=35000, wait_until='domcontentloaded')
    time.sleep(5)
    text = page.inner_text('body')
    data = {'url': url}

    data['ref'] = url.split('-')[-1]

    m = re.search(r'(\d+)\s*\nZimmer', text)
    if m: data['zimmer'] = int(m.group(1))

    m = re.search(r'Badezimmer\s*\n\s*(\d+)', text)
    if m: data['baeder'] = int(m.group(1))

    m = re.search(r'ca\.\s*([\d\.]+)\s*m²\s*\nWohnfläche', text)
    if m: data['wohnflaeche'] = float(m.group(1).replace('.', '').replace(',', '.'))

    m = re.search(r'ca\.\s*([\d\.]+)\s*m²\s*\nGrundstück', text)
    if m: data['grundstueck'] = int(m.group(1).replace('.', ''))

    m = re.search(r'([\d\.]+)\s*EUR\s*\nKaufpreis', text)
    if m: data['preis'] = int(m.group(1).replace('.', ''))

    m = re.search(r'Baujahr\s*\n\s*(\d{4})', text)
    if m: data['baujahr'] = int(m.group(1))

    m = re.search(r'Haustyp\s*\n\s*(.+)', text)
    if m: data['haustyp'] = m.group(1).strip()

    m = re.search(r'Zustand der Immobilie\s*\n\s*(.+)', text)
    if m: data['zustand'] = m.group(1).strip()

    m = re.search(r'(\d{5})\s+([A-ZÄÖÜ][^\n(]+)\s*\(ESP\)', text)
    if m: data['location'] = m.group(2).strip().title()

    data['beschreibung'] = extract_description(text)
    data['etv']   = any(k.lower() in text.lower() for k in ETV_KEYWORDS)
    data['gaeste'] = any(k.lower() in text.lower() for k in GAESTE_KEYWORDS)

    try:
        og = page.get_attribute('meta[property="og:image"]', 'content')
        if og: data['img_url'] = og
    except:
        pass

    return data


# ---------------------------------------------------------------------------
# Excel Writer
# ---------------------------------------------------------------------------

def write_to_excel(ws, nr, data, dist, charme=None, charme_grund=None, gebaeude=None, anzeigename=None):
    row = None
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=1).value == nr:
            row = r
            break
    if not row:
        row = ws.max_row + 1

    preis  = data.get('preis', 0) or 0
    bebaut = data.get('wohnflaeche', 0) or 0
    grund  = data.get('grundstueck', 0) or 0

    ws.cell(row=row, column=1).value  = nr
    ws.cell(row=row, column=3).value  = data['url']
    ws.cell(row=row, column=4).value  = charme  # Charme
    ws.cell(row=row, column=5).value  = data.get('zimmer')
    ws.cell(row=row, column=6).value  = data.get('baeder')
    ws.cell(row=row, column=7).value  = data.get('grundstueck')
    ws.cell(row=row, column=8).value  = data.get('wohnflaeche')
    # Garten = Grundstück - Bebaut
    if grund and bebaut: ws.cell(row=row, column=9).value = int(grund - bebaut)
    ws.cell(row=row, column=10).value = data.get('location')
    ws.cell(row=row, column=19).value = data.get('preis')
    ws.cell(row=row, column=20).value = round(preis / bebaut) if bebaut else None
    ws.cell(row=row, column=21).value = round(preis / grund)  if grund  else None
    ws.cell(row=row, column=23).value = bewirtschaftung(grund, data.get('haustyp'), data.get('beschreibung'))
    ws.cell(row=row, column=24).value = 100 if data.get('etv') else None
    # Gebäudestruktur: Vision-Text bevorzugen, sonst Fallback
    ws.cell(row=row, column=29).value = gebaeude if gebaeude else gebaeude_struktur_fallback(data)
    ws.cell(row=row, column=30).value = data.get('baujahr')
    ws.cell(row=row, column=32).value = reno_score(data.get('zustand'), data.get('baujahr'))
    ws.cell(row=row, column=33).value = f"Zustand: {data.get('zustand', 'unbekannt')}"
    ws.cell(row=row, column=34).value = charme_grund  # Kommentar = Vision-Begründung
    ws.cell(row=row, column=35).value = data.get('beschreibung')
    ws.cell(row=row, column=36).value = 'Von Poll'
    ws.cell(row=row, column=37).value = data.get('ref')
    ws.cell(row=row, column=38).value = 'active'
    ws.cell(row=row, column=40).value = 1 if data.get('gaeste') else None

    # Distanzen
    if dist:
        cols = {'Flughafen': (11, 12), 'Deia': (13, 14), 'Andratx': (15, 16), 'SesSalines': (17, 18)}
        for name, (ckm, cmin) in cols.items():
            if name in dist:
                ws.cell(row=row, column=ckm).value = dist[name]['km']
                ws.cell(row=row, column=cmin).value = dist[name]['min']
        ws.cell(row=row, column=25).value = erreichbarkeit(dist)

    # Anzeigename: Vision bevorzugen, sonst Fallback-Template
    if anzeigename:
        ws.cell(row=row, column=39).value = anzeigename
    else:
        loc     = data.get('location', 'Mallorca')
        haustyp = data.get('haustyp', 'Objekt')
        ws.cell(row=row, column=39).value = f"{loc} — {haustyp}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Von Poll Mallorca Scraper')
    parser.add_argument('--open',    action='store_true', help='Suchseite öffnen')
    parser.add_argument('--urls',    help='Datei mit URLs (eine pro Zeile)')
    parser.add_argument('--url',     nargs='+', help='Einzelne URLs')
    parser.add_argument('--no-vision', action='store_true', help='Charme-Vision überspringen')
    args = parser.parse_args()

    from camoufox.sync_api import Camoufox

    if args.open:
        print('Öffne Von Poll Suchseite...')
        print('Console-Snippet:')
        print('[...new Set([...document.querySelectorAll(\'a[href*="expose"]\')].map(e=>e.href))].join("\\n")')
        with Camoufox(headless=False) as browser:
            page = browser.new_page()
            page.goto('https://www.von-poll.com/de/finca-kaufen/mallorca')
            input('Enter zum Beenden...')
        return

    url_list = []
    if args.urls:
        with open(args.urls) as f:
            url_list = [l.strip() for l in f if l.strip().startswith('http')]
    if args.url:
        url_list.extend(args.url)

    if not url_list:
        print('Keine URLs. Nutze --open, --urls oder --url')
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Kandidaten']
    existing  = get_existing_urls(ws)
    new_urls  = [u for u in url_list if u not in existing]
    print(f'{len(url_list)} URLs → {len(new_urls)} neu')

    if not new_urls:
        print('Nichts Neues.')
        return

    with Camoufox(headless=True) as browser:
        page     = browser.new_page()
        next_nr  = get_next_nr(ws)

        for i, url in enumerate(new_urls):
            print(f'\n[{i+1}/{len(new_urls)}] {url.split("/")[-1][:70]}')
            try:
                data  = scrape_expose(page, url)
                preis = data.get('preis', 0) or 0
                zimmer = data.get('zimmer', 0) or 0

                if preis < 2000000 or zimmer < 5:
                    print(f'  SKIP: {preis:,}€ / {zimmer} Zi')
                    continue

                # Distanzen
                loc  = data.get('location', '')
                dist = get_distances(f'{loc}, Mallorca, Spain') if loc else {}

                # Bild laden
                img_path = os.path.join(BILDER_PATH, f'{next_nr}_main.jpg')
                if data.get('img_url') and not os.path.exists(img_path):
                    cookies = {c['name']: c['value'] for c in page.context.cookies()}
                    download_image(data['img_url'], img_path, cookies)

                # Vision: Charme + Gebäudestruktur
                charme, charme_grund, gebaeude, anzeigename = None, None, None, None
                if not args.no_vision and os.path.exists(img_path) and ANTHROPIC_KEY:
                    charme, charme_grund, gebaeude, anzeigename = vision_analyse(img_path, data.get('beschreibung', ''))
                    print(f'  Charme: {charme} — {charme_grund}')
                    print(f'  Gebäude: {gebaeude}')
                    print(f'  Anzeigename: {anzeigename}')
                elif not ANTHROPIC_KEY:
                    print('  Vision: übersprungen (kein OPENAI_API_KEY)')

                write_to_excel(ws, next_nr, data, dist, charme, charme_grund, gebaeude, anzeigename)
                wb.save(EXCEL_PATH)

                print(f'  ✅ Nr.{next_nr} | {preis:,}€ | {zimmer}Zi | {loc} | Bew={bewirtschaftung(data.get("grundstueck",0), data.get("haustyp"), data.get("beschreibung"))}')
                next_nr += 1

            except Exception as e:
                print(f'  ERROR: {e}')

    print(f'\n✅ Fertig.')


if __name__ == '__main__':
    main()
