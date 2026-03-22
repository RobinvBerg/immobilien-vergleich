#!/usr/bin/env python3
"""
Scrape 35 Von Poll Exposé pages, extract fields, download main images,
rate Charme+Renovierung via oracle CLI, write to Excel.
"""

import re
import os
import sys
import json
import time
import subprocess
import requests
import openpyxl
from pathlib import Path
from PIL import Image
from io import BytesIO

EXCEL_PATH = Path("/Users/robin/.openclaw/workspace/mallorca-projekt/data/mallorca-kandidaten-v2.xlsx")
BILDER_DIR = Path("/Users/robin/.openclaw/workspace/mallorca-projekt/bilder")
ORACLE_BIN = "/opt/homebrew/bin/oracle"
ORACLE_AVAILABLE = os.path.exists(ORACLE_BIN)

OBJECTS = {
    366: "https://www.von-poll.com/de/expose/mallorca-santa-maria/traumhafte-finca-mit-panorama-und-meerblick-in-bester-lage-4160045287",
    367: "https://www.von-poll.com/de/expose/mallorca-santa-maria/exklusiver-mediterraner-ruckzugsort-mit-zeitgenossischer-eleganz-4160047549",
    368: "https://www.von-poll.com/de/expose/mallorca-santa-maria/fantastische-finca-mit-panoramablick-in-die-tramuntana-in-der-nahe-von-santa-maria-del-cami-4160040127",
    369: "https://www.von-poll.com/de/expose/mallorca-santa-maria/luxuriose-neubaufinca-mit-pool-und-panoramablick-in-bunyola-4160036281",
    370: "https://www.von-poll.com/de/expose/mallorca-santa-maria/spektakulare-luxusfinca-mit-grossem-pool-zwischen-alaro-und-santa-maria-4160038693",
    371: "https://www.von-poll.com/de/expose/mallorca-santa-maria/grosszugige-neubau-finca-in-der-nahe-des-es-trenc-strandes-bei-campos-mit-pool-und-eigenem-weinanbau-4160043131",
    372: "https://www.von-poll.com/de/expose/mallorca-santa-maria/exklusive-neubaufinca-inmitten-eines-weinbergs-mit-pool-in-montuiri-4160045567",
    373: "https://www.von-poll.com/de/expose/mallorca-santa-maria/luxuriose-villa-in-son-vida-in-einer-privilegierten-lage-mit-blick-auf-den-nahe-gelegenen-golfplatz-4160039419",
    374: "https://www.von-poll.com/de/expose/mallorca-santa-maria/projekt-zum-bau-einer-finca-mit-pool-in-der-landlichen-umgebung-von-llubi-4160049307",
    375: "https://www.von-poll.com/de/expose/mallorca-santa-maria/finca-mit-pool-in-der-landlichen-idylle-zwischen-inca-und-costitx-4160044625",
    376: "https://www.von-poll.com/de/expose/mallorca-santa-maria/traumhafte-neu-gebaute-finca-mit-pool-in-sa-cabaneta-gemeinde-marratxi-4160044321",
    377: "https://www.von-poll.com/de/expose/mallorca-santa-maria/weitlaufige-finca-mit-etv-und-grossem-grundstuck-4160044995",
    378: "https://www.von-poll.com/de/expose/mallorca-santa-maria/moderne-neubau-villa-mit-pool-in-son-gual-4160044109",
    379: "https://www.von-poll.com/de/expose/mallorca-santa-maria/moderne-finca-mit-ferienvermietungslizenz-pool-und-traumhaftem-panoramablick-in-moscari-selva-mallorca-4160025903",
    380: "https://www.von-poll.com/de/expose/mallorca-santa-maria/wunderschones-projekt-zum-bau-einer-modernen-luxusvilla-in-der-exklusivsten-gegend-von-palma-in-son-vida-4160034679",
    381: "https://www.von-poll.com/de/expose/mallorca-santa-maria/neubauprojekt-fur-eine-luxusvilla-mit-pool-und-traumblick-in-son-vida-mallorca-4160034665",
    382: "https://www.von-poll.com/de/expose/mallorca-santa-maria/projekt-zum-bau-einer-luxusvilla-in-exklusiver-lage-von-palma-in-son-vida-4160034673",
    383: "https://www.von-poll.com/de/expose/mallorca-santa-maria/grossartiges-projekt-zum-bau-einer-herrlichen-villa-mit-pool-in-son-vida-mallorca-4160034659",
    384: "https://www.von-poll.com/de/expose/mallorca-santa-maria/zwei-einzigartige-projekte-in-orient-mit-panoramablick-4160044053",
    385: "https://www.von-poll.com/de/expose/mallorca-pollensa/secret-marketing-exklusives-finca-mit-hotel-lizenz-und-biologischem-weingut-in-der-serra-de-tramuntana-3520045049",
    386: "https://www.von-poll.com/de/expose/mallorca-pollensa/einzigartiges-historisches-luxusanwesen-von-1920-mit-meerblick-an-der-bucht-von-pollensa-vor-zwei-traumstranden-gelegen-3520036561",
    387: "https://www.von-poll.com/de/expose/mallorca-pollensa/atemberaubende-luxusvilla-in-erster-linie-mit-direktem-strandzugang-in-mal-pas-alcudia-an-der-bucht-von-pollensa-3520033675",
    388: "https://www.von-poll.com/de/expose/mallorca-pollensa/herrliches-landgut-mit-pool-und-panoramablick-im-wunderschonen-tal-von-aixartell-pollensa-3520035683",
    389: "https://www.von-poll.com/de/expose/mallorca-pollensa/spektakulare-villa-in-bester-lage-von-alcudia-in-mal-pas-3520045101",
    390: "https://www.von-poll.com/de/expose/mallorca-pollensa/luxusvilla-in-erster-meereslinie-an-der-kuste-von-sa-torre-llucmajor-3520047335",
    391: "https://www.von-poll.com/de/expose/mallorca-pollensa/traumhaftes-historisches-jagdanwesen-in-arta-in-bei-colonia-sant-pere-mit-projekt-fur-ein-luxuslandhotel-mit-spa-3520037285",
    392: "https://www.von-poll.com/de/expose/mallorca-pollensa/spektakulare-neu-gebaute-villa-in-einer-privilegierten-wohngegend-von-alcudia-in-paraiso-de-bonaire-3520049259",
    393: "https://www.von-poll.com/de/expose/mallorca-pollensa/spektakulare-luxusvilla-mit-unverbaubarem-meerblick-in-canyamel-3520044205",
    394: "https://www.von-poll.com/de/expose/mallorca-pollensa/secret-marketing-interessante-immobilie-mit-spektakularem-blick-auf-die-bucht-von-alcudia-in-muro-3520048123",
    395: "https://www.von-poll.com/de/expose/mallorca-pollensa/wunderschones-landhaus-in-der-nahe-von-colonia-sant-pere-mit-panoramablick-auf-das-meer-3520042991",
    396: "https://www.von-poll.com/de/expose/mallorca-pollensa/fabelhafte-neu-gebaute-finca-mit-meerblick-in-alcudia-mallorca-3520042635",
    397: "https://www.von-poll.com/de/expose/mallorca-pollensa/wunderschone-villa-in-den-bergen-von-pollensa-mit-spektakularer-aussicht-3520046987",
    398: "https://www.von-poll.com/de/expose/mallorca-pollensa/herrliches-mallorquinischer-herrenhaus-mit-pool-und-fantastischem-panoramablick-in-pollensa-3520038669",
    399: "https://www.von-poll.com/de/expose/mallorca-pollensa/moderne-villa-mit-gasteapartment-direkt-am-meer-in-portocolom-3520044169",
    400: "https://www.von-poll.com/de/expose/mallorca-pollensa/fantastische-historische-finca-in-sant-llorenc-des-cardassar-mit-der-moglichkeit-einer-lizenz-fur-ein-kleines-landhotel-3520028253",
}

# Column indices (1-based)
COL_CHARME = 4
COL_BAEDER = 6
COL_RENOVIERUNG = 22
COL_BEWIRTSCHAFTUNG = 23
COL_VERMIETLIZENZ = 24
COL_GEBAEUDE = 29
COL_BAUJAHR = 30
COL_LETZTE_RENO = 31
COL_GAESTEHAUS = 40


def derive_struktur(title, text):
    """Derive building structure from title and text."""
    title_lower = (title or "").lower()
    url_lower = title_lower  # URL slug also helpful
    
    keywords = [
        ("Jagdanwesen", ["jagdanwesen", "jagd"]),
        ("Herrenhaus", ["herrenhaus", "herrenhaus"]),
        ("Landhaus", ["landhaus", "landgut"]),
        ("Villa", ["villa"]),
        ("Finca", ["finca"]),
        ("Neubau", ["neubau", "neubauprojekt", "projekt zum bau"]),
        ("Altbau", ["altbau", "historisch", "1920"]),
    ]
    for name, kws in keywords:
        for kw in kws:
            if kw in title_lower:
                return name
    # Check text too
    text_lower = (text or "").lower()
    for name, kws in keywords:
        for kw in kws:
            if kw in text_lower:
                return name
    return None


def extract_fields(text, title, url):
    """Extract all fields from page text."""
    result = {}
    
    # Bäder
    m = re.search(r'(\d+)\s*Badezimmer', text, re.IGNORECASE)
    if not m:
        m = re.search(r'(\d+)\s*Bad(?:ezimmer)?(?:\s|,|\.)', text, re.IGNORECASE)
    if m:
        result['baeder'] = int(m.group(1))
    
    # Baujahr
    m = re.search(r'Baujahr[:\s]+(\d{4})', text, re.IGNORECASE)
    if not m:
        m = re.search(r'erbaut[:\s]+(\d{4})', text, re.IGNORECASE)
    if not m:
        m = re.search(r'von\s+(\d{4})\b', title or "", re.IGNORECASE)
    if m:
        yr = int(m.group(1))
        if 1800 <= yr <= 2030:
            result['baujahr'] = yr
    
    # Gebäudestruktur
    struk = derive_struktur(title, text)
    if struk:
        result['gebaeude'] = struk
    
    # Vermietlizenz
    if re.search(r'\bETV\b', text) or 'ferienvermietungslizenz' in text.lower() or 'hotel-lizenz' in text.lower() or 'hotellizenz' in text.lower():
        result['vermietlizenz'] = 100
    
    # Gästehäuser
    if (re.search(r'(\d+)\s*Gästeh', text, re.IGNORECASE) or
        'nebengebäude' in text.lower() or
        'gästeapartment' in text.lower() or
        'gästehaus' in text.lower() or
        'guest' in text.lower() or
        'gasteapartment' in url.lower()):
        result['gaestehaus'] = 1
    
    # LetzteReno: wenn Baujahr > 2015 → Baujahr (Neubau)
    if 'baujahr' in result and result['baujahr'] > 2015:
        result['letzte_reno'] = result['baujahr']
    
    return result


def download_image(nr, img_url, session_cookies=None):
    """Download main image to bilder/{nr}_main.jpg"""
    img_path = BILDER_DIR / f"{nr}_main.jpg"
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Referer': 'https://www.von-poll.com/',
    }
    
    try:
        r = requests.get(img_url, headers=headers, timeout=30, cookies=session_cookies)
        r.raise_for_status()
        
        content_type = r.headers.get('content-type', '')
        data = r.content
        
        # If webp, convert
        if 'webp' in content_type or img_url.lower().endswith('.webp'):
            img = Image.open(BytesIO(data))
            img = img.convert('RGB')
            img.save(img_path, 'JPEG', quality=90)
        else:
            with open(img_path, 'wb') as f:
                f.write(data)
        
        print(f"  → Bild gespeichert: {img_path.name}")
        return str(img_path)
    except Exception as e:
        print(f"  ✗ Bild-Download fehlgeschlagen: {e}")
        return None


def rate_with_oracle(img_path):
    """Rate image with oracle CLI. Returns (charme, renovierung)."""
    if not ORACLE_AVAILABLE:
        print("  oracle nicht verfügbar → Fallback: Charme=3, Renovierung=70")
        return 3, 70
    
    prompt = 'Bewerte dieses Immobilienfoto. Gib NUR JSON zurück: {"charme": <1-5>, "renovierung": <0-100>}. Charme: 1=sehr schlecht/Ruine, 2=sanierungsbedürftig, 3=okay, 4=gut renoviert, 5=traumhaft neu/luxus. Renovierung: 0=Ruine, 50=normal, 100=Neubau/perfekt.'
    
    try:
        result = subprocess.run(
            [ORACLE_BIN, "-e", prompt, "--file", img_path],
            capture_output=True, text=True, timeout=60
        )
        output = result.stdout + result.stderr
        
        # Parse JSON from output
        m = re.search(r'\{[^}]*"charme"[^}]*\}', output, re.DOTALL)
        if m:
            data = json.loads(m.group(0))
            charme = int(data.get('charme', 3))
            renovierung = int(data.get('renovierung', 70))
            print(f"  → oracle: Charme={charme}, Renovierung={renovierung}")
            return charme, renovierung
    except Exception as e:
        print(f"  oracle Fehler: {e}")
    
    return 3, 70


def scrape_object(nr, url):
    """Scrape one object. Returns dict of extracted data."""
    from camoufox.sync_api import Camoufox
    from bs4 import BeautifulSoup
    
    print(f"\n[{nr}] {url}")
    
    data = {}
    img_url = None
    page_text = ""
    title = url.split('/')[-1]  # URL slug as fallback title
    
    try:
        with Camoufox(headless=True) as browser:
            page = browser.new_page()
            page.goto(url, timeout=30000)
            time.sleep(5)
            
            content = page.content()
            
            # Get cookies for image download
            cookies = {c['name']: c['value'] for c in page.context.cookies()}
            
            soup = BeautifulSoup(content, 'html.parser')
            
            # Get title
            h1 = soup.find('h1')
            if h1:
                title = h1.get_text(strip=True)
            
            # Get page text
            page_text = soup.get_text(separator=' ', strip=True)
            
            # Find og:image for main image
            og_img = soup.find('meta', property='og:image')
            if og_img and og_img.get('content'):
                img_url = og_img['content']
            else:
                # Find largest img
                best_img = None
                best_size = 0
                for img in soup.find_all('img', src=True):
                    src = img['src']
                    w = int(img.get('width', 0) or 0)
                    h = int(img.get('height', 0) or 0)
                    size = w * h
                    if size > best_size and ('jpg' in src.lower() or 'jpeg' in src.lower() or 'webp' in src.lower()):
                        best_size = size
                        best_img = src
                if best_img:
                    img_url = best_img
                    if img_url.startswith('//'):
                        img_url = 'https:' + img_url
                    elif img_url.startswith('/'):
                        img_url = 'https://www.von-poll.com' + img_url
            
            data['cookies'] = cookies
            
    except Exception as e:
        print(f"  ✗ Camoufox Fehler: {e}")
        # Try web_fetch fallback
        try:
            import urllib.request
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=20) as resp:
                content = resp.read().decode('utf-8', errors='ignore')
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(content, 'html.parser')
            h1 = soup.find('h1')
            if h1:
                title = h1.get_text(strip=True)
            page_text = soup.get_text(separator=' ', strip=True)
            og_img = soup.find('meta', property='og:image')
            if og_img and og_img.get('content'):
                img_url = og_img['content']
        except Exception as e2:
            print(f"  ✗ Fallback auch fehlgeschlagen: {e2}")
    
    # Extract fields
    fields = extract_fields(page_text, title, url)
    print(f"  Extrahiert: {fields}")
    
    # Download image
    img_path = None
    if img_url:
        print(f"  Bild-URL: {img_url[:80]}...")
        img_path = download_image(nr, img_url, data.get('cookies'))
    else:
        # Check if already exists
        existing = BILDER_DIR / f"{nr}_main.jpg"
        if existing.exists():
            img_path = str(existing)
            print(f"  → Bestehendes Bild gefunden: {existing.name}")
    
    # Rate with oracle
    charme, renovierung = 3, 70
    if img_path and os.path.exists(img_path):
        charme, renovierung = rate_with_oracle(img_path)
    else:
        print(f"  Kein Bild → Fallback: Charme=3, Renovierung=70")
    
    return {
        'charme': charme,
        'renovierung': renovierung,
        'baeder': fields.get('baeder'),
        'gebaeude': fields.get('gebaeude'),
        'baujahr': fields.get('baujahr'),
        'letzte_reno': fields.get('letzte_reno'),
        'vermietlizenz': fields.get('vermietlizenz'),
        'gaestehaus': fields.get('gaestehaus'),
    }


def write_to_excel(wb, ws, nr, data):
    """Find row by nr in col1, write data."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value == nr:
            row_idx = row[0].row
            
            def set_cell(col, val):
                if val is not None:
                    ws.cell(row=row_idx, column=col, value=val)
            
            set_cell(COL_CHARME, data.get('charme'))
            set_cell(COL_BAEDER, data.get('baeder'))
            set_cell(COL_RENOVIERUNG, data.get('renovierung'))
            set_cell(COL_GEBAEUDE, data.get('gebaeude'))
            set_cell(COL_BAUJAHR, data.get('baujahr'))
            set_cell(COL_LETZTE_RENO, data.get('letzte_reno'))
            set_cell(COL_VERMIETLIZENZ, data.get('vermietlizenz'))
            set_cell(COL_GAESTEHAUS, data.get('gaestehaus'))
            
            print(f"  ✓ Excel-Zeile {row_idx} geschrieben")
            return True
    print(f"  ✗ Nr {nr} nicht in Excel gefunden!")
    return False


def main():
    print(f"Oracle verfügbar: {ORACLE_AVAILABLE}")
    print(f"Bilder-Ordner: {BILDER_DIR}")
    BILDER_DIR.mkdir(exist_ok=True)
    
    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Kandidaten']
    
    # Stats
    stats = {
        'charme': 0, 'renovierung': 0, 'baeder': 0, 'gebaeude': 0,
        'baujahr': 0, 'letzte_reno': 0, 'vermietlizenz': 0, 'gaestehaus': 0
    }
    
    total = len(OBJECTS)
    for i, (nr, url) in enumerate(OBJECTS.items(), 1):
        print(f"\n{'='*60}")
        print(f"Objekt {i}/{total}: Nr {nr}")
        
        try:
            data = scrape_object(nr, url)
            
            # Update stats
            for key in stats:
                if data.get(key) is not None:
                    stats[key] += 1
            
            # Write to Excel
            write_to_excel(wb, ws, nr, data)
            
            # Save after each object
            wb.save(EXCEL_PATH)
            print(f"  ✓ Excel gespeichert")
            
        except Exception as e:
            print(f"  ✗ FEHLER bei {nr}: {e}")
            import traceback
            traceback.print_exc()
    
    # Final summary
    print(f"\n{'='*60}")
    print("ZUSAMMENFASSUNG:")
    print(f"  Gesamt Objekte: {total}")
    for field, count in stats.items():
        print(f"  {field}: {count}/{total} gefüllt")
    
    print("\nFertig!")


if __name__ == '__main__':
    main()
