#!/usr/bin/env python3
"""
Mallorca Property Scraper - Phase 1
Direct scraping for portals without Apify actors.
"""

import time
import re
import json
import random
from datetime import date
from urllib.parse import urljoin, quote
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
TODAY = str(date.today())

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xhtml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
}

def get_page(url, timeout=15, delay=2):
    """Fetch page with delay."""
    time.sleep(delay + random.uniform(0.5, 1.5))
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return BeautifulSoup(r.text, 'html.parser')
    except Exception as e:
        print(f"  ERROR fetching {url}: {e}")
        return None

def load_existing_data():
    """Load existing URLs from Excel."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set()
    existing_combos = set()  # (price_bucket, rooms, city) for fuzzy dedup
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            existing_urls.add(str(row[2]).strip())
        # Build combo key for fuzzy dedup
        try:
            price = float(row[3]) if row[3] else 0
            rooms = int(row[4]) if row[4] else 0
            city = str(row[7]).lower().strip() if row[7] else ''
            if price > 0:
                price_bucket = round(price / 1000)  # round to nearest 1k
                existing_combos.add((price_bucket, rooms, city))
        except:
            pass
    return wb, ws, existing_urls, existing_combos

def is_duplicate(obj, existing_urls, existing_combos):
    """Check if property already exists."""
    if obj.get('url') and str(obj['url']).strip() in existing_urls:
        return True
    # Fuzzy check
    try:
        price = float(obj.get('preis', 0)) if obj.get('preis') else 0
        rooms = int(obj.get('zimmer', 0)) if obj.get('zimmer') else 0
        city = str(obj.get('ort', '')).lower().strip()
        if price > 0:
            price_bucket = round(price / 1000)
            if (price_bucket, rooms, city) in existing_combos:
                return True
            # ±5% range
            for delta in range(-1, 2):
                adj = round(price * (1 + delta * 0.05) / 1000)
                if (adj, rooms, city) in existing_combos:
                    return True
    except:
        pass
    return False

def save_objects(wb, ws, new_objects, existing_urls, existing_combos):
    """Append new objects to Excel."""
    added = 0
    for obj in new_objects:
        if is_duplicate(obj, existing_urls, existing_combos):
            continue
        ws.append([
            obj.get('titel', ''),
            obj.get('quelle', ''),
            obj.get('url', ''),
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            TODAY,
            'Neu'
        ])
        if obj.get('url'):
            existing_urls.add(str(obj['url']).strip())
        added += 1
    wb.save(EXCEL_PATH)
    return added


# ============================================================
# SCRAPER: ThinkSpain
# ============================================================
def scrape_thinkspain():
    """Scrape ThinkSpain for Mallorca properties."""
    print("\n=== ThinkSpain ===")
    results = []
    base = "https://www.thinkspain.com"
    
    urls = [
        "https://www.thinkspain.com/property-for-sale/3/mallorca-balearic-islands?beds=5",
        "https://www.thinkspain.com/property-for-sale/3/mallorca-balearic-islands?beds=4",
        "https://www.thinkspain.com/property-for-sale/3/mallorca-balearic-islands",
    ]
    
    seen_urls = set()
    for url in urls:
        for page in range(1, 6):
            page_url = url + (f"&page={page}" if page > 1 else "")
            soup = get_page(page_url)
            if not soup:
                break
            
            # Find property listings
            listings = soup.find_all('article', class_=re.compile(r'property|listing', re.I))
            if not listings:
                listings = soup.find_all('div', class_=re.compile(r'property-item|listing-item', re.I))
            if not listings:
                # Try generic approach
                listings = soup.select('a[href*="/property/"]')
            
            if not listings:
                print(f"  No listings found on page {page}, stopping")
                break
            
            for item in listings:
                try:
                    # Get link
                    link = item.get('href') if item.name == 'a' else None
                    if not link:
                        a = item.find('a', href=re.compile(r'/property/'))
                        link = a['href'] if a else None
                    if not link:
                        continue
                    full_url = urljoin(base, link)
                    if full_url in seen_urls:
                        continue
                    seen_urls.add(full_url)
                    
                    # Title
                    title_el = item.find(['h2', 'h3', 'h4']) or item.find(class_=re.compile(r'title', re.I))
                    title = title_el.get_text(strip=True) if title_el else ''
                    
                    # Price
                    price_el = item.find(class_=re.compile(r'price', re.I)) or item.find(string=re.compile(r'€|\d+,\d+'))
                    price = None
                    if price_el:
                        price_text = price_el.get_text() if hasattr(price_el, 'get_text') else str(price_el)
                        nums = re.findall(r'[\d,\.]+', price_text.replace('.', '').replace(',', ''))
                        if nums:
                            try:
                                price = float(nums[0])
                            except:
                                pass
                    
                    # Rooms
                    rooms_el = item.find(string=re.compile(r'\d+\s*bed|\d+\s*hab', re.I))
                    rooms = None
                    if rooms_el:
                        m = re.search(r'(\d+)', str(rooms_el))
                        if m:
                            rooms = int(m.group(1))
                    
                    results.append({
                        'titel': title or 'ThinkSpain Property',
                        'quelle': 'ThinkSpain',
                        'url': full_url,
                        'preis': price,
                        'zimmer': rooms,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
                except Exception as e:
                    continue
            
            print(f"  Page {page}: {len(results)} total so far")
            if len(listings) < 10:
                break
    
    print(f"  ThinkSpain total: {len(results)}")
    return results


# ============================================================
# SCRAPER: A Place in the Sun
# ============================================================
def scrape_aplaceinthesun():
    """Scrape A Place in the Sun for Mallorca."""
    print("\n=== A Place in the Sun ===")
    results = []
    base = "https://www.aplaceinthesun.com"
    
    urls = [
        "https://www.aplaceinthesun.com/property-for-sale/spain/mallorca/",
        "https://www.aplaceinthesun.com/property-for-sale/spain/mallorca/?bedrooms=5",
        "https://www.aplaceinthesun.com/property-for-sale/spain/mallorca/?bedrooms=4",
    ]
    
    seen_urls = set()
    for url in urls:
        for page in range(1, 8):
            page_url = url + (f"page/{page}/" if page > 1 else "")
            soup = get_page(page_url, delay=2)
            if not soup:
                break
            
            # Find property cards
            cards = soup.find_all('div', class_=re.compile(r'property-card|listing|property-item', re.I))
            if not cards:
                cards = soup.select('article[class*="property"], div[class*="property-list"]')
            if not cards:
                links = soup.select('a[href*="/property-for-sale/spain/mallorca/"]')
                cards = [l.parent for l in links if l.get('href', '').count('/') > 5]
            
            if not cards:
                print(f"  No cards on page {page}")
                break
            
            found_new = 0
            for card in cards:
                try:
                    a = card.find('a', href=True)
                    if not a:
                        continue
                    href = a['href']
                    if 'mallorca' not in href.lower() and 'spain' not in href.lower():
                        continue
                    full_url = urljoin(base, href)
                    if full_url in seen_urls:
                        continue
                    seen_urls.add(full_url)
                    found_new += 1
                    
                    title_el = card.find(['h2', 'h3']) or a
                    title = title_el.get_text(strip=True) if title_el else ''
                    
                    price_el = card.find(class_=re.compile(r'price', re.I))
                    price = None
                    if price_el:
                        nums = re.findall(r'\d+', price_el.get_text().replace(',', '').replace('.', ''))
                        if nums:
                            try:
                                price = float(''.join(nums[:2]))
                            except:
                                pass
                    
                    beds_el = card.find(string=re.compile(r'\d+\s*bed', re.I))
                    beds = None
                    if beds_el:
                        m = re.search(r'(\d+)', str(beds_el))
                        beds = int(m.group(1)) if m else None
                    
                    results.append({
                        'titel': title or 'APTS Property',
                        'quelle': 'A Place in the Sun',
                        'url': full_url,
                        'preis': price,
                        'zimmer': beds,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
                except:
                    continue
            
            print(f"  Page {page}: {len(results)} total, {found_new} new")
            if found_new == 0:
                break
    
    print(f"  A Place in the Sun total: {len(results)}")
    return results


# ============================================================
# SCRAPER: abc-mallorca.de
# ============================================================
def scrape_abcmallorca():
    """Scrape abc-mallorca.de."""
    print("\n=== abc-mallorca ===")
    results = []
    base = "https://www.abc-mallorca.de"
    
    search_urls = [
        "https://www.abc-mallorca.de/immobilien/?type=buy&propertytype=villa,finca,house,land&min_bedrooms=4",
        "https://www.abc-mallorca.de/immobilien/?type=buy",
        "https://www.abc-mallorca.de/real-estate/?type=buy",
    ]
    
    seen_urls = set()
    for url in search_urls:
        for page in range(1, 8):
            page_url = url + (f"&paged={page}" if page > 1 else "")
            soup = get_page(page_url, delay=2)
            if not soup:
                break
            
            # Find property links
            cards = soup.find_all('div', class_=re.compile(r'property|listing|immob', re.I))
            if not cards:
                cards = soup.find_all('article')
            
            found_new = 0
            for card in cards:
                try:
                    a = card.find('a', href=True)
                    if not a:
                        continue
                    href = a['href']
                    full_url = urljoin(base, href)
                    if full_url in seen_urls or len(full_url) < 30:
                        continue
                    # Only property detail pages
                    if not any(x in full_url for x in ['immobil', 'real-estate', 'property']):
                        continue
                    seen_urls.add(full_url)
                    found_new += 1
                    
                    title_el = card.find(['h2', 'h3', 'h4'])
                    title = title_el.get_text(strip=True) if title_el else a.get_text(strip=True)
                    
                    price_el = card.find(class_=re.compile(r'price', re.I))
                    price = None
                    if price_el:
                        price_text = price_el.get_text()
                        nums = re.findall(r'[\d]+', price_text.replace('.', '').replace(',', ''))
                        if nums:
                            try:
                                price = float(nums[0]) if len(nums[0]) > 3 else None
                            except:
                                pass
                    
                    results.append({
                        'titel': title or 'abc-mallorca Property',
                        'quelle': 'abc-mallorca',
                        'url': full_url,
                        'preis': price,
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
                except:
                    continue
            
            print(f"  Page {page}: {len(results)} total, {found_new} new")
            if found_new == 0:
                break
    
    print(f"  abc-mallorca total: {len(results)}")
    return results


# ============================================================
# SCRAPER: Habitaclia
# ============================================================
def scrape_habitaclia():
    """Scrape Habitaclia for Mallorca."""
    print("\n=== Habitaclia ===")
    results = []
    base = "https://www.habitaclia.com"
    
    urls = [
        "https://www.habitaclia.com/comprar-casas-en-mallorca.htm",
        "https://www.habitaclia.com/comprar-villas-en-mallorca.htm",
        "https://www.habitaclia.com/comprar-finca-en-mallorca.htm",
    ]
    
    seen_urls = set()
    for url in urls:
        for page in range(1, 8):
            page_url = url + (f"?page={page}" if page > 1 else "")
            soup = get_page(page_url, delay=2)
            if not soup:
                break
            
            # Try multiple selectors
            cards = soup.find_all('article') or soup.find_all('div', class_=re.compile(r'property|listing|list-item', re.I))
            
            found_new = 0
            for card in cards:
                try:
                    a = card.find('a', href=re.compile(r'habitaclia\.com|/comprar|/venta'))
                    if not a:
                        a = card.find('a', href=True)
                    if not a:
                        continue
                    href = a['href']
                    full_url = urljoin(base, href)
                    if full_url in seen_urls:
                        continue
                    # Must be a property page
                    if not re.search(r'habitaclia\.com/.+\.htm', full_url):
                        continue
                    seen_urls.add(full_url)
                    found_new += 1
                    
                    title_el = card.find(['h2', 'h3']) or a
                    title = title_el.get_text(strip=True) if title_el else ''
                    
                    # Price
                    price_el = card.find(class_=re.compile(r'price|precio', re.I))
                    price = None
                    if price_el:
                        nums = re.findall(r'\d+', price_el.get_text().replace('.', '').replace(',', ''))
                        if nums:
                            try:
                                price = float(nums[0]) if len(nums[0]) > 3 else None
                            except:
                                pass
                    
                    # Rooms
                    rooms_el = card.find(string=re.compile(r'\d+\s*hab|\d+\s*dorm', re.I))
                    rooms = None
                    if rooms_el:
                        m = re.search(r'(\d+)', str(rooms_el))
                        rooms = int(m.group(1)) if m else None
                    
                    # Location
                    loc_el = card.find(class_=re.compile(r'location|zona|ciudad', re.I))
                    loc = loc_el.get_text(strip=True) if loc_el else 'Mallorca'
                    
                    results.append({
                        'titel': title or 'Habitaclia Property',
                        'quelle': 'Habitaclia',
                        'url': full_url,
                        'preis': price,
                        'zimmer': rooms,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': loc,
                    })
                except:
                    continue
            
            print(f"  Page {page}: {len(results)} total, {found_new} new")
            if found_new == 0:
                break
    
    print(f"  Habitaclia total: {len(results)}")
    return results


# ============================================================
# SCRAPER: Green-Acres
# ============================================================
def scrape_greenacres():
    """Scrape Green-Acres for Mallorca."""
    print("\n=== Green-Acres ===")
    results = []
    base = "https://www.green-acres.es"
    
    urls = [
        "https://www.green-acres.es/es/propiedades/venta/espana/islas-baleares/mallorca/",
        "https://www.green-acres.es/en/properties/buy/spain/balearic-islands/majorca/",
        "https://www.green-acres.es/de/immobilien/kauf/spanien/balearische-inseln/mallorca/",
    ]
    
    seen_urls = set()
    for url in urls:
        for page in range(1, 6):
            page_url = url + (f"?p={page}" if page > 1 else "")
            soup = get_page(page_url, delay=2)
            if not soup:
                break
            
            # Property links
            links = soup.find_all('a', href=re.compile(r'/propiedad|/property|/immobilie', re.I))
            if not links:
                cards = soup.find_all(['article', 'div'], class_=re.compile(r'property|annonce|listing'))
                links = [c.find('a') for c in cards if c.find('a')]
            
            found_new = 0
            for a in links:
                if not a or not a.get('href'):
                    continue
                try:
                    full_url = urljoin(base, a['href'])
                    if full_url in seen_urls:
                        continue
                    if 'green-acres' not in full_url:
                        continue
                    seen_urls.add(full_url)
                    found_new += 1
                    
                    parent = a.parent or a
                    title = a.get_text(strip=True) or 'Green-Acres Property'
                    
                    # Find price in parent context
                    price_el = parent.find(string=re.compile(r'€|\d+\.\d+')) if parent else None
                    price = None
                    if price_el:
                        nums = re.findall(r'\d+', str(price_el).replace('.', '').replace(',', ''))
                        if nums and len(nums[0]) > 3:
                            try:
                                price = float(nums[0])
                            except:
                                pass
                    
                    results.append({
                        'titel': title,
                        'quelle': 'Green-Acres',
                        'url': full_url,
                        'preis': price,
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
                except:
                    continue
            
            print(f"  Page {page}: {len(results)} total, {found_new} new")
            if found_new == 0:
                break
    
    print(f"  Green-Acres total: {len(results)}")
    return results


# ============================================================
# SCRAPER: Yaencontre
# ============================================================
def scrape_yaencontre():
    """Scrape Yaencontre for Mallorca."""
    print("\n=== Yaencontré ===")
    results = []
    base = "https://www.yaencontre.com"
    
    urls = [
        "https://www.yaencontre.com/casas/venta/mallorca",
        "https://www.yaencontre.com/villas/venta/mallorca",
        "https://www.yaencontre.com/fincas/venta/mallorca",
    ]
    
    seen_urls = set()
    for url in urls:
        for page in range(1, 6):
            page_url = url + (f"/{page}" if page > 1 else "")
            soup = get_page(page_url, delay=2)
            if not soup:
                break
            
            # Find listings
            cards = soup.find_all(['article', 'div'], class_=re.compile(r'property|listing|result|item', re.I))
            if not cards:
                cards = soup.select('a[href*="/pisos/"], a[href*="/casas/"], a[href*="/villa/"]')
            
            found_new = 0
            for card in cards:
                try:
                    if card.name == 'a':
                        a = card
                    else:
                        a = card.find('a', href=True)
                    if not a:
                        continue
                    href = a.get('href', '')
                    full_url = urljoin(base, href)
                    if full_url in seen_urls:
                        continue
                    # Must be property page
                    if 'yaencontre.com' not in full_url:
                        continue
                    seen_urls.add(full_url)
                    found_new += 1
                    
                    title_el = card.find(['h2', 'h3']) if card.name != 'a' else None
                    title = title_el.get_text(strip=True) if title_el else a.get_text(strip=True) or 'Yaencontre Property'
                    
                    price_el = card.find(class_=re.compile(r'price|precio', re.I)) if card.name != 'a' else None
                    price = None
                    if price_el:
                        nums = re.findall(r'\d+', price_el.get_text().replace('.', '').replace(',', ''))
                        if nums and len(nums[0]) > 3:
                            try:
                                price = float(nums[0])
                            except:
                                pass
                    
                    results.append({
                        'titel': title,
                        'quelle': 'Yaencontré',
                        'url': full_url,
                        'preis': price,
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
                except:
                    continue
            
            print(f"  Page {page}: {len(results)} total, {found_new} new")
            if found_new == 0:
                break
    
    print(f"  Yaencontré total: {len(results)}")
    return results


# ============================================================
# MAIN
# ============================================================
def main():
    print("Loading existing Excel data...")
    wb, ws, existing_urls, existing_combos = load_existing_data()
    print(f"Loaded {len(existing_urls)} existing URLs")
    
    total_added = 0
    results_summary = {}
    
    scrapers = [
        ('ThinkSpain', scrape_thinkspain),
        ('A Place in the Sun', scrape_aplaceinthesun),
        ('abc-mallorca', scrape_abcmallorca),
        ('Habitaclia', scrape_habitaclia),
        ('Green-Acres', scrape_greenacres),
        ('Yaencontré', scrape_yaencontre),
    ]
    
    for name, scraper_fn in scrapers:
        try:
            results = scraper_fn()
            if results:
                added = save_objects(wb, ws, results, existing_urls, existing_combos)
                results_summary[name] = {'scraped': len(results), 'added': added}
                total_added += added
                print(f"  ✓ {name}: {len(results)} scraped, {added} new added to Excel")
            else:
                results_summary[name] = {'scraped': 0, 'added': 0}
                print(f"  ✗ {name}: 0 results")
        except Exception as e:
            print(f"  ✗ {name}: FAILED - {e}")
            results_summary[name] = {'scraped': 0, 'added': 0, 'error': str(e)}
    
    print(f"\n{'='*50}")
    print(f"DIRECT SCRAPING SUMMARY")
    print(f"{'='*50}")
    for name, stats in results_summary.items():
        print(f"  {name}: {stats.get('scraped',0)} scraped → {stats.get('added',0)} added")
        if 'error' in stats:
            print(f"    ERROR: {stats['error']}")
    print(f"\nTotal new objects added: {total_added}")
    
    # Save summary to file for the Apify follow-up
    with open('/Users/robin/.openclaw/workspace/mallorca-projekt/phase1_direct_results.json', 'w') as f:
        json.dump({'summary': results_summary, 'total_added': total_added}, f, indent=2)

if __name__ == '__main__':
    main()
