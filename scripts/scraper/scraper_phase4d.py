#!/usr/bin/env python3
"""Phase 4d - Gezieltes Scraping für die schwierigen Quellen"""

import sys, json, time, re, requests
from datetime import date
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
APIFY_TOKEN = 'apify_api_feD2KhARHjtuV9CrSwOReYgoePFSF44nsDL6'

def save_to_excel(new_objects, source_name):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
    new_count = 0
    for obj in new_objects:
        url = obj.get('url', '—')
        if url != '—' and url in existing_urls:
            continue
        ws.append([obj.get('titel',''), obj.get('quelle', source_name), url,
                   obj.get('preis'), obj.get('zimmer'), obj.get('grundstueck'),
                   obj.get('wohnflaeche'), obj.get('ort',''), str(date.today()), 'Neu'])
        if url != '—': existing_urls.add(url)
        new_count += 1
    wb.save(EXCEL_PATH)
    print(f"  ✅ {new_count} neue Objekte aus {source_name} gespeichert")
    return new_count

def parse_price(text):
    if not text: return None
    text = str(text).replace('\xa0', '').replace(' ', '').replace(',', '.')
    nums = re.findall(r'\d+(?:\.\d+)?', text)
    for n in nums:
        try:
            val = float(n.replace('.', '', n.count('.')-1)) if n.count('.') > 1 else float(n)
            val2 = int(float(n.replace('.', '')))
            if val2 > 10000: return val2
        except: pass
    return None

def parse_int(text):
    if not text: return None
    nums = re.findall(r'\d+', str(text))
    return int(nums[0]) if nums else None

def make_browser():
    p = sync_playwright().start()
    browser = p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-blink-features=AutomationControlled'])
    context = browser.new_context(
        user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        viewport={'width': 1280, 'height': 900},
    )
    page = context.new_page()
    Stealth().apply_stealth_sync(page)
    return p, browser, page


# ============================================================
# REIDERSTAD - visit mallorca properties page directly
# ============================================================
def scrape_reiderstad():
    print("\n🏠 Reiderstad Invest - Gezielte URL")
    objects = []
    
    p, browser, page = make_browser()
    try:
        page.goto('https://reiderstadinvest.com/listing/?form=Buy&destinationid=74', 
                 wait_until='domcontentloaded', timeout=25000)
        time.sleep(3)
        
        # Also try direct mallorca page
        urls_to_try = [
            'https://reiderstadinvest.com/mallorca/properties/',
            'https://reiderstadinvest.com/listing/?form=Buy&destinationid=74',
            'https://reiderstadinvest.com/listing/?form=Buy',
        ]
        
        for url in urls_to_try:
            page.goto(url, wait_until='domcontentloaded', timeout=25000)
            time.sleep(3)
            
            content = page.content()
            # Look for Mallorca properties
            if 'mallorca' in content.lower() or 'palma' in content.lower():
                # Extract all property links
                links = page.query_selector_all('a[href*="mallorca"], a[href*="palma"], a[href*="/buy/"]')
                cards = page.query_selector_all('[class*="property"], [class*="listing"], article, .card, li')
                print(f"  {url}: links={len(links)}, cards={len(cards)}")
                
                seen = set()
                for link in links:
                    try:
                        href = link.get_attribute('href') or ''
                        if not href or href in seen or '#' in href: continue
                        seen.add(href)
                        if not href.startswith('http'): href = 'https://reiderstadinvest.com' + href
                        
                        # Get card context
                        text = link.inner_text().strip()
                        parent_text = page.evaluate("""(el) => {
                            let p = el.parentElement;
                            for (let i = 0; i < 4; i++) {
                                if (p && p.innerText && p.innerText.length > 10) break;
                                p = p ? p.parentElement : null;
                            }
                            return p ? p.innerText : '';
                        }""", link)
                        
                        price = parse_price(parent_text)
                        rooms = parse_int(re.search(r'(\d+)\s*bedroom', parent_text, re.I).group(1) if re.search(r'(\d+)\s*bedroom', parent_text, re.I) else '')
                        area = parse_int(re.search(r'(\d+)\s*m2', parent_text, re.I).group(1) if re.search(r'(\d+)\s*m2', parent_text, re.I) else '')
                        
                        objects.append({
                            'titel': text[:100] or 'Reiderstad Objekt',
                            'quelle': 'Reiderstad Invest',
                            'url': href,
                            'preis': price,
                            'zimmer': rooms,
                            'grundstueck': None,
                            'wohnflaeche': area,
                            'ort': 'Mallorca',
                        })
                    except Exception as e:
                        pass
                
                if objects:
                    break
    except Exception as e:
        print(f"  Error: {e}")
        import traceback; traceback.print_exc()
    finally:
        browser.close()
        p.stop()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Reiderstad Invest')
    return 0


# ============================================================
# Von Poll - Try with correct URL and selectors
# ============================================================
def scrape_von_poll():
    print("\n🏠 Von Poll Real Estate - Verbessert")
    objects = []
    
    p, browser, page = make_browser()
    api_data = []
    
    def handle_response(resp):
        try:
            if 'von-poll' in resp.url and resp.status == 200:
                ct = resp.headers.get('content-type', '')
                if 'json' in ct:
                    body = resp.body()
                    if body and len(body) > 100:
                        api_data.append({'url': resp.url, 'body': body})
                        print(f"  API: {resp.url[:80]}")
        except: pass
    
    page.on('response', handle_response)
    
    try:
        # Von Poll URL for Mallorca
        page.goto('https://www.von-poll.com/de/immobilien/spanien/balearen/mallorca', 
                 wait_until='networkidle', timeout=30000)
        time.sleep(3)
        
        # Try consent
        try:
            page.click('#cmpwelcomebtnyes, [id*="accept"], button[class*="accept"]', timeout=3000)
            time.sleep(1)
        except: pass
        
        # Get page source and look at structure
        html = page.content()
        print(f"  HTML length: {len(html)}")
        print(f"  Has 'immobilie': {'immobilie' in html.lower()}")
        print(f"  API responses: {len(api_data)}")
        
        # Process API data if any
        for api in api_data:
            try:
                data = json.loads(api['body'])
                items = None
                if isinstance(data, dict):
                    for k in ['items', 'results', 'properties', 'listings', 'data', 'objects']:
                        if k in data and isinstance(data[k], list):
                            items = data[k]
                            print(f"  API items under '{k}': {len(items)}")
                            break
                if items:
                    for item in items[:50]:
                        url_val = item.get('url', item.get('link', '—'))
                        if url_val and not url_val.startswith('http') and url_val != '—':
                            url_val = 'https://www.von-poll.com' + url_val
                        objects.append({
                            'titel': item.get('title', item.get('address', str(item.get('id', '')))),
                            'quelle': 'Von Poll Real Estate',
                            'url': url_val,
                            'preis': item.get('price', item.get('priceValue')),
                            'zimmer': item.get('rooms', item.get('bedrooms')),
                            'grundstueck': item.get('plotArea'),
                            'wohnflaeche': item.get('area', item.get('livingArea')),
                            'ort': item.get('location', item.get('city', 'Mallorca')),
                        })
            except: pass
        
        if not objects:
            # HTML fallback - look for any links to property pages
            links = page.query_selector_all('a[href*="expose"], a[href*="/objekt"], a[href*="/property"]')
            all_links = page.query_selector_all('a')
            
            print(f"  Expose links: {len(links)}, All links: {len(all_links)}")
            
            # Look for links with expose IDs (Von Poll typically uses /expose/12345)
            seen = set()
            for link in all_links:
                try:
                    href = link.get_attribute('href') or ''
                    if not href or href in seen: continue
                    if re.search(r'/expose/\d+|/property/\d+|/objekt/', href):
                        seen.add(href)
                        if not href.startswith('http'): href = 'https://www.von-poll.com' + href
                        text = link.inner_text().strip()
                        objects.append({
                            'titel': text[:100] or 'Von Poll Objekt',
                            'quelle': 'Von Poll Real Estate',
                            'url': href,
                            'preis': None, 'zimmer': None, 'grundstueck': None,
                            'wohnflaeche': None, 'ort': 'Mallorca',
                        })
                except: pass
            
            # Also try looking for data in JSON embedded in page
            json_matches = re.findall(r'window\.__INITIAL_STATE__\s*=\s*(\{.+?\});', html, re.DOTALL)
            json_matches += re.findall(r'window\.__NUXT__\s*=\s*(\{.+?\});', html, re.DOTALL)
            json_matches += re.findall(r'"properties"\s*:\s*(\[.+?\])', html, re.DOTALL)
            
            print(f"  JSON matches: {len(json_matches)}")
            for jm in json_matches[:3]:
                try:
                    data = json.loads(jm)
                    print(f"  Parsed: {type(data)}, keys: {list(data.keys()) if isinstance(data, dict) else 'list'}")
                except Exception as e:
                    print(f"  JSON parse error: {e}")
        
        # Print page title for debugging
        print(f"  Page title: {page.title()}")
        
    except Exception as e:
        print(f"  Error: {e}")
        import traceback; traceback.print_exc()
    finally:
        browser.close()
        p.stop()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Von Poll Real Estate')
    return 0


# ============================================================
# Knight Frank - Try correct URL
# ============================================================
def scrape_knight_frank():
    print("\n🏠 Knight Frank - Verbessert")
    objects = []
    
    # Try their API
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Content-Type': 'application/json',
        'Origin': 'https://www.knightfrank.com',
        'Referer': 'https://www.knightfrank.com/',
    }
    
    # KF uses a specific search endpoint
    try:
        # Try different API patterns
        endpoints = [
            ('GET', 'https://www.knightfrank.com/api/properties/search?location=mallorca&listingType=sale'),
            ('GET', 'https://www.knightfrank.com/api/search?query=mallorca&type=residential'),
            ('POST', 'https://www.knightfrank.com/api/property/search'),
        ]
        
        for method, url in endpoints:
            try:
                if method == 'GET':
                    resp = requests.get(url, headers=headers, timeout=10)
                else:
                    resp = requests.post(url, json={"location": "mallorca", "listingType": "sale"}, headers=headers, timeout=10)
                print(f"  {method} {url[:60]}: {resp.status_code}")
                if resp.status_code == 200:
                    print(f"  Response: {resp.text[:200]}")
            except Exception as e:
                print(f"  {url[:50]}: {e}")
    except Exception as e:
        print(f"  API error: {e}")
    
    p, browser, page = make_browser()
    api_data = []
    
    def handle_response(resp):
        try:
            if 'knightfrank' in resp.url and resp.status == 200:
                ct = resp.headers.get('content-type', '')
                if 'json' in ct:
                    body = resp.body()
                    if body and len(body) > 200:
                        api_data.append({'url': resp.url, 'body': body})
                        print(f"  KF API: {resp.url[:100]}")
        except: pass
    
    page.on('response', handle_response)
    
    try:
        urls = [
            'https://www.knightfrank.com/residential-property-for-sale/spain/mallorca',
            'https://www.knightfrank.com/residential-property-for-sale/spain/balearic-islands',
            'https://www.knightfrank.com/international-residential-property-for-sale/spain',
        ]
        
        for url in urls:
            page.goto(url, wait_until='networkidle', timeout=30000)
            time.sleep(4)
            
            # Cookie
            try:
                page.click('#onetrust-accept-btn-handler, button:has-text("Accept All"), button:has-text("Accept")', timeout=3000)
                time.sleep(2)
            except: pass
            
            html = page.content()
            print(f"  {url}: {len(html)} bytes, API calls: {len(api_data)}")
            print(f"  Title: {page.title()}")
            
            # Look for property cards with any selector
            all_elements = page.query_selector_all('article, [class*="card"], [class*="Card"], [class*="property"], [class*="Property"]')
            print(f"  Elements: {len(all_elements)}")
            
            for api in api_data:
                try:
                    data = json.loads(api['body'])
                    if isinstance(data, dict):
                        for k in ['properties', 'listings', 'results', 'items', 'data']:
                            if k in data and isinstance(data[k], list) and data[k]:
                                print(f"  KF data under '{k}': {len(data[k])}")
                                for item in data[k][:5]:
                                    print(f"    Sample: {json.dumps(item, ensure_ascii=False)[:200]}")
                                for item in data[k]:
                                    url_val = item.get('url', item.get('propertyUrl', item.get('link', '—')))
                                    if url_val and not url_val.startswith('http') and url_val != '—':
                                        url_val = 'https://www.knightfrank.com' + url_val
                                    objects.append({
                                        'titel': item.get('title', item.get('address', item.get('displayAddress', ''))),
                                        'quelle': 'Knight Frank',
                                        'url': url_val,
                                        'preis': item.get('price', item.get('priceValue', item.get('askingPrice'))),
                                        'zimmer': item.get('bedrooms', item.get('beds')),
                                        'grundstueck': item.get('landArea', item.get('plotSize')),
                                        'wohnflaeche': item.get('floorArea', item.get('area')),
                                        'ort': item.get('location', item.get('city', item.get('region', 'Mallorca'))),
                                    })
                                break
                except Exception as e:
                    print(f"  Parse error: {e}")
            
            if objects:
                break
        
    except Exception as e:
        print(f"  Error: {e}")
        import traceback; traceback.print_exc()
    finally:
        browser.close()
        p.stop()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Knight Frank')
    return 0


# ============================================================
# Savills - Use Apify or better scraping
# ============================================================
def scrape_savills():
    print("\n🏠 Savills - Apify oder direkt")
    objects = []
    
    # Try Apify free actor for Savills
    try:
        search_url = f"https://api.apify.com/v2/store?token={APIFY_TOKEN}&search=savills&limit=5"
        resp = requests.get(search_url, timeout=10)
        actors = resp.json().get('data', {}).get('items', [])
        for a in actors:
            print(f"  Actor: {a.get('username')}/{a.get('name')} - {a.get('title')} - ${a.get('stats', {}).get('totalRuns', 0)} runs")
    except Exception as e:
        print(f"  Apify search error: {e}")
    
    p, browser, page = make_browser()
    api_data = []
    
    def handle_response(resp):
        try:
            if 'savills' in resp.url and resp.status == 200:
                ct = resp.headers.get('content-type', '')
                if 'json' in ct or 'application' in ct:
                    body = resp.body()
                    if body and len(body) > 200:
                        api_data.append({'url': resp.url, 'body': body})
                        print(f"  Savills API: {resp.url[:100]} ({len(body)}b)")
        except: pass
    
    page.on('response', handle_response)
    
    try:
        urls = [
            'https://search.savills.com/es/en/list?SearchList=IsInRegion_EUR5002&SaleType=ForSale&PropertyType=RES',
            'https://search.savills.com/list?LocationIdentifiers=EUR5002&SaleType=ForSale&PropertyType=RES',
        ]
        
        for url in urls:
            page.goto(url, wait_until='networkidle', timeout=30000)
            time.sleep(5)
            
            # Cookie
            try:
                page.click('#onetrust-accept-btn-handler, button:has-text("Accept All"), button:has-text("Aceptar")', timeout=3000)
                time.sleep(2)
            except: pass
            
            html = page.content()
            print(f"  {url[:60]}: {len(html)}b, APIs: {len(api_data)}")
            print(f"  Title: {page.title()}")
            
            # Look for any cards/listings
            all_items = page.query_selector_all('article, [class*="Card"], [class*="listing"], [class*="property"]')
            links = page.query_selector_all('a[href*="/property/"], a[href*="/properties/"]')
            print(f"  Items: {len(all_items)}, Links: {len(links)}")
            
            # Process APIs
            for api in api_data:
                try:
                    data = json.loads(api['body'])
                    if isinstance(data, dict):
                        for k in ['Results', 'results', 'listings', 'properties', 'data', 'items', 'SearchResults']:
                            if k in data and isinstance(data[k], list) and data[k]:
                                print(f"  Savills data under '{k}': {len(data[k])}")
                                print(f"  Sample: {json.dumps(data[k][0], ensure_ascii=False)[:300]}")
                                for item in data[k]:
                                    url_val = item.get('url', item.get('Url', item.get('Link', item.get('DetailUrl', '—'))))
                                    if url_val and not url_val.startswith('http') and url_val != '—':
                                        url_val = 'https://search.savills.com' + url_val
                                    
                                    price_raw = (item.get('Price', item.get('price', {})))
                                    price = price_raw.get('Value', price_raw.get('value', price_raw)) if isinstance(price_raw, dict) else price_raw
                                    
                                    objects.append({
                                        'titel': item.get('Title', item.get('title', item.get('Address', item.get('address', '')))),
                                        'quelle': 'Savills',
                                        'url': url_val,
                                        'preis': price,
                                        'zimmer': item.get('Bedrooms', item.get('bedrooms', item.get('Beds'))),
                                        'grundstueck': item.get('PlotSize', item.get('LandArea')),
                                        'wohnflaeche': item.get('FloorArea', item.get('Area', item.get('Size'))),
                                        'ort': item.get('Location', item.get('City', item.get('Region', 'Mallorca'))),
                                    })
                                break
                except Exception as e:
                    print(f"  Parse error: {e}")
            
            if objects:
                break
    
    except Exception as e:
        print(f"  Error: {e}")
        import traceback; traceback.print_exc()
    finally:
        browser.close()
        p.stop()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Savills')
    return 0


# ============================================================
# Fotocasa - Request capture with network intercept
# ============================================================
def scrape_fotocasa():
    print("\n🏠 Fotocasa - Netzwerk-Intercept")
    objects = []
    
    p, browser, page = make_browser()
    captured = []
    
    def handle_response(resp):
        try:
            url = resp.url
            if resp.status == 200 and 'fotocasa' in url:
                ct = resp.headers.get('content-type', '')
                if 'json' in ct:
                    body = resp.body()
                    if body and len(body) > 500:
                        captured.append({'url': url, 'body': body})
                        print(f"  Captured: {url[:100]} ({len(body)}b)")
        except: pass
    
    page.on('response', handle_response)
    
    try:
        page.goto('https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l', 
                 wait_until='networkidle', timeout=30000)
        time.sleep(4)
        
        # Accept didomi/cookie consent
        for selector in ['#didomi-notice-agree-button', 'button[aria-label*="Accept"]', 
                         'button:has-text("Aceptar")', 'button:has-text("Acepto todo")', '.fc-cta-consent']:
            try:
                page.click(selector, timeout=2000)
                time.sleep(2)
                print(f"  Cookie accepted via: {selector}")
                break
            except: pass
        
        time.sleep(3)
        print(f"  Total captured: {len(captured)}")
        
        # Try to get data from captured requests
        for c in captured:
            try:
                data = json.loads(c['body'])
                listings = None
                
                if isinstance(data, dict):
                    # Look for the real estate list
                    for k in ['realEstates', 'results', 'items', 'listings', 'data']:
                        if k in data:
                            v = data[k]
                            if isinstance(v, list) and len(v) > 0:
                                listings = v
                                print(f"  Found '{k}': {len(v)} items")
                                print(f"  Sample: {json.dumps(v[0], ensure_ascii=False)[:400]}")
                                break
                
                if listings:
                    for item in listings:
                        # Price extraction
                        price = None
                        if 'transactions' in item and item['transactions']:
                            tx = item['transactions'][0]
                            vals = tx.get('value', [])
                            price = vals[0] if vals else tx.get('price')
                        elif 'price' in item:
                            price = item['price']
                        
                        # Features
                        features = {}
                        for f in item.get('features', []):
                            features[f.get('key', '')] = f.get('value', [None])[0] if isinstance(f.get('value'), list) else f.get('value')
                        
                        # URL
                        detail = item.get('detail', {})
                        url_val = detail.get('es', detail.get('url', '')) if isinstance(detail, dict) else ''
                        if url_val and not url_val.startswith('http'):
                            url_val = 'https://www.fotocasa.es' + url_val
                        
                        # Location
                        addr = item.get('address', {})
                        municipality = addr.get('municipality', '') if isinstance(addr, dict) else ''
                        ubication = addr.get('ubication', '') if isinstance(addr, dict) else ''
                        
                        objects.append({
                            'titel': ubication or municipality,
                            'quelle': 'Fotocasa',
                            'url': url_val or '—',
                            'preis': price,
                            'zimmer': features.get('roomsNumber'),
                            'grundstueck': features.get('plotArea'),
                            'wohnflaeche': features.get('constructedArea', features.get('surface')),
                            'ort': municipality,
                        })
            except Exception as e:
                pass
        
        # Pagination via network requests
        if objects:
            total_pages_est = 10
            for pg in range(2, total_pages_est + 1):
                captured_before = len(captured)
                next_url = f'https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l?page={pg}'
                page.goto(next_url, wait_until='networkidle', timeout=20000)
                time.sleep(3)
                
                new_captures = captured[captured_before:]
                added = 0
                for c in new_captures:
                    try:
                        data = json.loads(c['body'])
                        for k in ['realEstates', 'results', 'items']:
                            if k in data and isinstance(data[k], list) and data[k]:
                                for item in data[k]:
                                    price = None
                                    if 'transactions' in item and item['transactions']:
                                        vals = item['transactions'][0].get('value', [])
                                        price = vals[0] if vals else None
                                    detail = item.get('detail', {})
                                    url_val = (detail.get('es', '') if isinstance(detail, dict) else '') 
                                    if url_val and not url_val.startswith('http'):
                                        url_val = 'https://www.fotocasa.es' + url_val
                                    features = {f.get('key',''): (f.get('value',[None])[0] if isinstance(f.get('value'),list) else f.get('value')) for f in item.get('features',[])}
                                    addr = item.get('address', {})
                                    objects.append({
                                        'titel': (addr.get('ubication','') if isinstance(addr,dict) else ''),
                                        'quelle': 'Fotocasa', 'url': url_val or '—',
                                        'preis': price, 'zimmer': features.get('roomsNumber'),
                                        'grundstueck': features.get('plotArea'),
                                        'wohnflaeche': features.get('constructedArea', features.get('surface')),
                                        'ort': (addr.get('municipality','') if isinstance(addr,dict) else ''),
                                    })
                                    added += 1
                                break
                    except: pass
                print(f"  Page {pg}: +{added}")
                if added == 0:
                    break
        
        # Fallback HTML if still nothing
        if not objects:
            print("  HTML fallback...")
            cards = page.query_selector_all('[class*="re-Card"], [class*="Card"], article')
            print(f"  Cards: {len(cards)}")
    
    except Exception as e:
        print(f"  Error: {e}")
        import traceback; traceback.print_exc()
    finally:
        browser.close()
        p.stop()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Fotocasa')
    return 0


# ============================================================
# Rightmove - Direct Playwright with correct URLs
# ============================================================
def scrape_rightmove():
    print("\n🏠 Rightmove - Playwright")
    objects = []
    
    # Try free Apify actor first
    try:
        run_input = {
            "startUrls": [{"url": "https://www.rightmove.co.uk/overseas-property/in-Mallorca.html"}],
            "maxItems": 100,
        }
        resp = requests.post(
            f"https://api.apify.com/v2/acts/dhrumil~rightmove-scraper/runs?token={APIFY_TOKEN}",
            json=run_input, timeout=30
        )
        print(f"  Apify dhrumil/rightmove: {resp.status_code}")
        if resp.status_code in (200, 201):
            run_id = resp.json().get('data', {}).get('id')
            print(f"  Run ID: {run_id}")
            deadline = time.time() + 120
            while time.time() < deadline:
                time.sleep(8)
                sr = requests.get(f"https://api.apify.com/v2/actor-runs/{run_id}?token={APIFY_TOKEN}", timeout=10)
                status = sr.json().get('data', {}).get('status', '')
                print(f"  Status: {status}")
                if status == 'SUCCEEDED':
                    ds_id = sr.json()['data']['defaultDatasetId']
                    items = requests.get(f"https://api.apify.com/v2/datasets/{ds_id}/items?token={APIFY_TOKEN}&limit=500", timeout=30).json()
                    print(f"  Apify items: {len(items)}")
                    if items:
                        print(f"  Sample: {json.dumps(items[0], ensure_ascii=False)[:300]}")
                        for item in items:
                            url_val = item.get('url', item.get('propertyUrl', item.get('link', '—')))
                            if url_val and not url_val.startswith('http') and url_val != '—':
                                url_val = 'https://www.rightmove.co.uk' + url_val
                            objects.append({
                                'titel': item.get('title', item.get('address', item.get('displayAddress', ''))),
                                'quelle': 'Rightmove',
                                'url': url_val,
                                'preis': item.get('price', item.get('priceValue', item.get('amount'))),
                                'zimmer': item.get('bedrooms', item.get('beds')),
                                'grundstueck': item.get('plotArea'),
                                'wohnflaeche': item.get('floorArea', item.get('area')),
                                'ort': item.get('location', item.get('city', 'Mallorca')),
                            })
                        if objects:
                            return save_to_excel(objects, 'Rightmove')
                    break
                elif status in ('FAILED', 'ABORTED', 'TIMED-OUT'):
                    print(f"  Run {status}: {sr.json().get('data', {}).get('statusMessage', '')}")
                    break
        elif resp.status_code == 403:
            print(f"  403 - actor needs subscription: {resp.json().get('error', {}).get('message', '')}")
    except Exception as e:
        print(f"  Apify error: {e}")
    
    # Playwright fallback
    p, browser, page = make_browser()
    api_data = []
    
    def handle_response(resp):
        try:
            url = resp.url
            if 'rightmove' in url and resp.status == 200:
                ct = resp.headers.get('content-type', '')
                if 'json' in ct:
                    body = resp.body()
                    if body and len(body) > 200:
                        api_data.append({'url': url, 'body': body})
                        print(f"  RM API: {url[:100]}")
        except: pass
    
    page.on('response', handle_response)
    
    try:
        page.goto('https://www.rightmove.co.uk/overseas-property/in-Mallorca.html', 
                 wait_until='networkidle', timeout=30000)
        time.sleep(4)
        
        html = page.content()
        print(f"  RM HTML: {len(html)}b, APIs: {len(api_data)}")
        
        # Try to find property links
        links = page.query_selector_all('a[href*="overseas-property"], a[href*="/properties/"]')
        cards = page.query_selector_all('[class*="propertyCard"], [class*="property-card"], [data-test*="property"]')
        print(f"  Links: {len(links)}, Cards: {len(cards)}")
        
        # Process any JSON data from APIs
        for api in api_data:
            try:
                data = json.loads(api['body'])
                for k in ['properties', 'listings', 'results', 'searchResults']:
                    if k in data and isinstance(data[k], list) and data[k]:
                        print(f"  RM data '{k}': {len(data[k])}")
                        for item in data[k]:
                            url_val = item.get('url', item.get('propertyUrl', '—'))
                            if url_val and not url_val.startswith('http') and url_val != '—':
                                url_val = 'https://www.rightmove.co.uk' + url_val
                            objects.append({
                                'titel': item.get('displayAddress', item.get('title', item.get('address', ''))),
                                'quelle': 'Rightmove',
                                'url': url_val,
                                'preis': item.get('price', {}).get('amount') if isinstance(item.get('price'), dict) else item.get('price'),
                                'zimmer': item.get('bedrooms'),
                                'grundstueck': None,
                                'wohnflaeche': None,
                                'ort': item.get('displayAddress', 'Mallorca').split(',')[-1].strip(),
                            })
                        break
            except: pass
        
        # Scrape HTML cards
        seen = set()
        for card in cards:
            try:
                link = card.query_selector('a')
                href = link.get_attribute('href') if link else None
                if not href or href in seen: continue
                seen.add(href)
                if not href.startswith('http'): href = 'https://www.rightmove.co.uk' + href
                
                price_el = card.query_selector('[data-test="property-price"], [class*="price"], .property-price')
                title_el = card.query_selector('[data-test="property-heading"], h2, h3, .property-h3')
                beds_el = card.query_selector('[class*="beds"], [data-beds]')
                
                objects.append({
                    'titel': title_el.inner_text()[:100] if title_el else '',
                    'quelle': 'Rightmove',
                    'url': href,
                    'preis': parse_price(price_el.inner_text() if price_el else ''),
                    'zimmer': parse_int(beds_el.inner_text() if beds_el else ''),
                    'grundstueck': None, 'wohnflaeche': None, 'ort': 'Mallorca',
                })
            except: pass
        
    except Exception as e:
        print(f"  Playwright error: {e}")
        import traceback; traceback.print_exc()
    finally:
        browser.close()
        p.stop()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Rightmove')
    return 0


# ============================================================
# Properstar - proper search  
# ============================================================
def scrape_properstar():
    print("\n🏠 Properstar - Verbessert")
    objects = []
    
    p, browser, page = make_browser()
    captured = []
    
    def handle_response(resp):
        try:
            url = resp.url
            if resp.status == 200 and ('properstar' in url or 'listing' in url):
                ct = resp.headers.get('content-type', '')
                if 'json' in ct:
                    body = resp.body()
                    if body and len(body) > 300:
                        captured.append({'url': url, 'body': body})
                        print(f"  Properstar API: {url[:100]}")
        except: pass
    
    page.on('response', handle_response)
    
    try:
        # Try different Properstar URLs
        urls = [
            'https://www.properstar.com/spain/buy?location=mallorca',
            'https://www.properstar.com/es/venta?country=ES&region=Balearic+Islands',
            'https://www.properstar.co.uk/spain/buy?location=mallorca',
        ]
        
        for url in urls:
            page.goto(url, wait_until='networkidle', timeout=25000)
            time.sleep(4)
            
            html = page.content()
            print(f"  {url[:60]}: {len(html)}b, APIs: {len(captured)}")
            
            # Cookie
            try:
                page.click('button:has-text("Accept"), button:has-text("OK"), [id*="accept"]', timeout=2000)
                time.sleep(1)
            except: pass
            
            for c in captured:
                try:
                    data = json.loads(c['body'])
                    listings = None
                    if isinstance(data, dict):
                        for k in ['listings', 'results', 'properties', 'items', 'data', 'Listings']:
                            if k in data and isinstance(data[k], list) and data[k]:
                                listings = data[k]
                                print(f"  Properstar '{k}': {len(listings)}")
                                break
                    elif isinstance(data, list) and data:
                        listings = data
                    
                    if listings:
                        for item in listings:
                            url_v = item.get('url', item.get('Url', item.get('link', item.get('detailUrl', '—'))))
                            if url_v and not url_v.startswith('http') and url_v != '—':
                                url_v = 'https://www.properstar.com' + url_v
                            objects.append({
                                'titel': item.get('title', item.get('Title', item.get('address', item.get('Address', '')))),
                                'quelle': 'Properstar',
                                'url': url_v,
                                'preis': item.get('price', item.get('Price', item.get('salePrice'))),
                                'zimmer': item.get('bedrooms', item.get('Bedrooms', item.get('rooms'))),
                                'grundstueck': item.get('landArea', item.get('LandArea', item.get('plotSize'))),
                                'wohnflaeche': item.get('livingArea', item.get('LivingArea', item.get('area'))),
                                'ort': item.get('city', item.get('City', item.get('location', 'Mallorca'))),
                            })
                except Exception as e:
                    print(f"  Parse: {e}")
            
            if objects:
                break
    
    except Exception as e:
        print(f"  Error: {e}")
    finally:
        browser.close()
        p.stop()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Properstar')
    return 0


# ============================================================
# Vives Pons - better scraping
# ============================================================
def scrape_vives_pons():
    print("\n🏠 Vives Pons - Verbessert")
    objects = []
    
    p, browser, page = make_browser()
    captured = []
    
    def handle_response(resp):
        try:
            if 'vivespons' in resp.url and resp.status == 200:
                ct = resp.headers.get('content-type', '')
                if 'json' in ct:
                    body = resp.body()
                    if body and len(body) > 100:
                        captured.append({'url': resp.url, 'body': body})
                        print(f"  VP API: {resp.url[:80]}")
        except: pass
    
    page.on('response', handle_response)
    
    try:
        page.goto('https://www.vivespons.com/en/properties', wait_until='networkidle', timeout=25000)
        time.sleep(3)
        
        try:
            page.click('button:has-text("Accept"), [id*="accept"]', timeout=2000)
            time.sleep(1)
        except: pass
        
        html = page.content()
        print(f"  HTML: {len(html)}b, APIs: {len(captured)}")
        
        # Look for all property links
        all_links = page.query_selector_all('a')
        seen = set()
        for link in all_links:
            try:
                href = link.get_attribute('href') or ''
                if not href or href in seen or href.startswith('#'): continue
                seen.add(href)
                if not href.startswith('http'): href = 'https://www.vivespons.com' + href
                
                # Filter for property pages
                if re.search(r'/property/|/properties/\w|/en/properties/.', href):
                    text = link.inner_text().strip()
                    objects.append({
                        'titel': text[:100] or 'Vives Pons Objekt',
                        'quelle': 'Vives Pons',
                        'url': href,
                        'preis': None, 'zimmer': None, 'grundstueck': None,
                        'wohnflaeche': None, 'ort': 'Mallorca',
                    })
            except: pass
        
        # Try pagination
        for pg in range(2, 5):
            page.goto(f'https://www.vivespons.com/en/properties?page={pg}', wait_until='networkidle', timeout=15000)
            time.sleep(2)
            before = len(objects)
            for link in page.query_selector_all('a'):
                try:
                    href = link.get_attribute('href') or ''
                    if not href or href in seen: continue
                    seen.add(href)
                    if not href.startswith('http'): href = 'https://www.vivespons.com' + href
                    if re.search(r'/property/|/properties/\w|/en/properties/.', href):
                        objects.append({'titel': link.inner_text()[:100], 'quelle': 'Vives Pons', 'url': href,
                                       'preis': None, 'zimmer': None, 'grundstueck': None, 'wohnflaeche': None, 'ort': 'Mallorca'})
                except: pass
            if len(objects) == before: break
        
        print(f"  Links found: {len(objects)}")
        
        # Also process API data
        for c in captured:
            try:
                data = json.loads(c['body'])
                if isinstance(data, dict):
                    for k in ['properties', 'results', 'items', 'data']:
                        if k in data and isinstance(data[k], list):
                            print(f"  VP API '{k}': {len(data[k])}")
                            for item in data[k]:
                                url_v = item.get('url', item.get('link', '—'))
                                if url_v and not url_v.startswith('http') and url_v != '—':
                                    url_v = 'https://www.vivespons.com' + url_v
                                if url_v not in seen:
                                    seen.add(url_v)
                                    objects.append({
                                        'titel': item.get('title', item.get('name', '')),
                                        'quelle': 'Vives Pons',
                                        'url': url_v,
                                        'preis': item.get('price'), 'zimmer': item.get('bedrooms'),
                                        'grundstueck': item.get('plotArea'), 'wohnflaeche': item.get('area'),
                                        'ort': item.get('location', 'Mallorca'),
                                    })
                            break
            except: pass
    
    except Exception as e:
        print(f"  Error: {e}")
    finally:
        browser.close()
        p.stop()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'Vives Pons')
    return 0


# ============================================================
# The Agency RE - with different approach
# ============================================================
def scrape_the_agency():
    print("\n🏠 The Agency RE")
    objects = []
    
    p, browser, page = make_browser()
    captured = []
    
    def handle_response(resp):
        try:
            if 'theagencyre' in resp.url and resp.status == 200:
                ct = resp.headers.get('content-type', '')
                if 'json' in ct:
                    body = resp.body()
                    if body and len(body) > 100:
                        captured.append({'url': resp.url, 'body': body})
                        print(f"  Agency API: {resp.url[:100]}")
        except: pass
    
    page.on('response', handle_response)
    
    try:
        page.goto('https://www.theagencyre.com/international/spain/mallorca', 
                 wait_until='domcontentloaded', timeout=25000)
        time.sleep(4)
        
        html = page.content()
        print(f"  HTML: {len(html)}b")
        
        # Find property links
        links = page.query_selector_all('a[href*="/international/"], a[href*="/listing/"]')
        all_links = page.query_selector_all('a')
        print(f"  Property links: {len(links)}, All: {len(all_links)}")
        
        seen = set()
        for link in all_links:
            try:
                href = link.get_attribute('href') or ''
                if not href or href in seen: continue
                seen.add(href)
                if not href.startswith('http'): href = 'https://www.theagencyre.com' + href
                if re.search(r'/listing/|/property/|/international/spain/mallorca/', href):
                    text = link.inner_text().strip()
                    if len(text) > 3:
                        objects.append({
                            'titel': text[:100],
                            'quelle': 'The Agency RE',
                            'url': href,
                            'preis': None, 'zimmer': None, 'grundstueck': None,
                            'wohnflaeche': None, 'ort': 'Mallorca',
                        })
            except: pass
    
    except Exception as e:
        print(f"  Error: {e}")
        import traceback; traceback.print_exc()
    finally:
        browser.close()
        p.stop()
    
    print(f"  Gesammelt: {len(objects)}")
    if objects:
        return save_to_excel(objects, 'The Agency RE')
    return 0


if __name__ == '__main__':
    results = {}
    
    results['Reiderstad'] = scrape_reiderstad()
    results['Von Poll'] = scrape_von_poll()
    results['Knight Frank'] = scrape_knight_frank()
    results['Savills'] = scrape_savills()
    results['Fotocasa'] = scrape_fotocasa()
    results['Rightmove'] = scrape_rightmove()
    results['Properstar'] = scrape_properstar()
    results['Vives Pons'] = scrape_vives_pons()
    results['The Agency RE'] = scrape_the_agency()
    
    print("\n=== FINALE ERGEBNISSE ===")
    total = 0
    for src, count in results.items():
        status = "✅" if count > 0 else "❌"
        print(f"  {status} {src}: {count} Objekte")
        total += count
    print(f"\n  Total neu: {total}")
