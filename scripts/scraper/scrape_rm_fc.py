#!/usr/bin/env python3
"""Direct scraper for Rightmove + Fotocasa plot sizes.
- Rightmove: window.PAGE_MODEL description text → regex for plot size
- Fotocasa: HTML text pattern for superficie de terreno/parcela
- Handles gzip encoding
"""

import json, re, time, random, sys, gzip
import urllib.request
import openpyxl

EXCEL = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
PROGRESS = '/Users/robin/.openclaw/workspace/mallorca-projekt/fetchdetails_progress.json'
SAVE_EVERY = 50
DELAY = 2.0

RM_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml',
    'Accept-Language': 'en-GB,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate',
    'Connection': 'keep-alive',
}

FC_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml',
    'Accept-Language': 'es-ES,es;q=0.9',
    'Accept-Encoding': 'gzip, deflate',
    'Connection': 'keep-alive',
}

def fetch(url, hdrs=RM_HEADERS, timeout=20):
    req = urllib.request.Request(url, headers=hdrs)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            raw = r.read()
            enc = r.headers.get('Content-Encoding', '')
            if 'gzip' in enc:
                raw = gzip.decompress(raw)
            try:
                return raw.decode('utf-8')
            except:
                return raw.decode('latin-1', errors='replace')
    except Exception as e:
        return None

def parse_num(s):
    """Parse a number string like '60,000' or '60.000' → 60000"""
    s = str(s).strip().replace(',','').replace('.','').replace('\xa0','').replace(' ','')
    try:
        return int(s)
    except:
        return None

# Rightmove patterns for plot size in description
RM_PLOT_PATTERNS = [
    # "plot of approx. 60,000 m²"
    r'plot[^<]{0,80}?(\d[\d,\.]{1,9})\s*m[²2]',
    r'(\d[\d,\.]{1,9})\s*m[²2]\s*plot',
    r'land\s*area[^<]{0,50}?(\d[\d,\.]{1,9})\s*m[²2]',
    r'(\d[\d,\.]{1,9})\s*m[²2]\s*(?:of\s+)?(?:private\s+)?land',
    r'plot\s+size[^<]{0,50}?(\d[\d,\.]{1,9})\s*m[²2]',
    r'plot[^\d<]{0,80}(\d[\d,\.]+)\s*m[²2]',
    r'(\d[\d,\.]{1,9})\s*m[²2]\s*(?:total\s+)?(?:plot|terrain)',
    # Hectares
    r'plot[^<]{0,80}?(\d[\d,\.]{0,4})\s*ha\b',
    r'(\d[\d,\.]{0,4})\s*ha\s+(?:plot|of\s+land)',
    r'(\d[\d,\.]{0,4})\s*hectare',
]

def extract_rightmove_plot(html):
    if not html:
        return None
    # Get description from PAGE_MODEL
    desc = ''
    m = re.search(r'window\.PAGE_MODEL\s*=\s*(\{)', html)
    if m:
        try:
            start = m.start(1)
            txt = html[start:]
            depth = 0
            for i, c in enumerate(txt):
                if c == '{': depth += 1
                elif c == '}':
                    depth -= 1
                    if depth == 0:
                        txt = txt[:i+1]; break
            data = json.loads(txt)
            pd = data.get('propertyData', {})
            desc = pd.get('text', {}).get('description', '')
            # Also check keyFeatures
            kf = ' '.join(pd.get('keyFeatures', []))
            desc = desc + ' ' + kf
            # Check sizings — but these are usually building size, not plot
            # We can still log them
        except:
            pass
    
    # Try plot patterns on description first
    search_text = desc + ' ' + html
    for pat in RM_PLOT_PATTERNS:
        m = re.search(pat, search_text, re.IGNORECASE)
        if m:
            grp = m.group(1).strip()
            if 'ha' in pat and 'ha' in m.group(0).lower():
                # convert hectares to m²
                try:
                    return int(float(grp.replace(',','.')) * 10000)
                except:
                    pass
            else:
                val = parse_num(grp)
                if val and 100 <= val <= 10000000:
                    return val
    return None

# Fotocasa patterns
FC_PLOT_PATTERNS = [
    r'Superficie\s+(?:de\s+)?(?:terreno|parcela)[^<\d]{0,30}?(\d[\d\.]{0,7})',
    r'(\d[\d\.]{0,7})\s*m[²2]\s*(?:de\s+)?(?:terreno|parcela)',
    r'terreno[^<\d]{0,20}?(\d[\d\.]{0,7})\s*m',
    r'parcela[^<\d]{0,20}?(\d[\d\.]{0,7})\s*m',
]

def extract_fotocasa_plot(html):
    if not html:
        return None
    for pat in FC_PLOT_PATTERNS:
        m = re.search(pat, html, re.IGNORECASE)
        if m:
            val = parse_num(m.group(1))
            if val and 100 <= val <= 10000000:
                return val
    return None

def load_progress():
    try:
        with open(PROGRESS) as f:
            return json.load(f)
    except:
        return {}

def save_progress(p):
    with open(PROGRESS, 'w') as f:
        json.dump(p, f, indent=2)

def get_pending_urls(progress):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    url_idx = headers.index('URL')
    plot_idx = headers.index('Grundstück (m²)')
    rm_urls, fc_urls = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[url_idx]
        plot = row[plot_idx]
        if not url or plot:
            continue
        url = str(url).strip()
        if url in progress and progress[url].get('done'):
            continue
        if 'rightmove' in url:
            rm_urls.append(url)
        elif 'fotocasa' in url:
            fc_urls.append(url)
    return rm_urls, fc_urls

def update_excel(progress):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    url_idx = headers.index('URL')
    plot_idx = headers.index('Grundstück (m²)')
    updated = 0
    for row in ws.iter_rows(min_row=2):
        url = row[url_idx].value
        if not url: continue
        url = str(url).strip()
        if url in progress and progress[url].get('plot'):
            if not row[plot_idx].value:
                row[plot_idx].value = progress[url]['plot']
                updated += 1
    wb.save(EXCEL)
    return updated

def compute_pool(progress):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    url_idx = headers.index('URL')
    plot_idx = headers.index('Grundstück (m²)')
    price_idx = headers.index('Preis (€)')
    rooms_idx = headers.index('Zimmer')
    loc_idx = headers.index('Ort / Gemeinde')

    pool = []
    nordost_kw = ['alcudia','pollenca','pollença','artà','arta','cala ratjada','capdepera',
                  'son servera','sa coma','cala millor','porto cristo','manacor','sant llorenç']
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = str(row[url_idx]).strip() if row[url_idx] else ''
        plot = row[plot_idx]
        price = row[price_idx]
        rooms = row[rooms_idx]
        loc = str(row[loc_idx]).lower() if row[loc_idx] else ''
        if not plot and url in progress:
            plot = progress[url].get('plot')
        if not plot or not price or not rooms:
            continue
        try:
            plot = int(plot)
            price = float(str(price).replace('€','').replace('.','').replace(',','.').strip())
            rooms = int(rooms)
        except:
            continue
        if plot < 7000: continue
        if price < 2900000: continue
        if rooms < 5: continue
        if any(kw in loc for kw in nordost_kw): continue
        pool.append({'url': url, 'plot': plot, 'price': price, 'rooms': rooms, 'loc': row[loc_idx]})
    return pool

def main():
    progress = load_progress()
    rm_urls, fc_urls = get_pending_urls(progress)
    print(f"Pending: {len(rm_urls)} Rightmove, {len(fc_urls)} Fotocasa", flush=True)
    
    total_found = 0
    counter = 0
    all_urls = [('rm', u) for u in rm_urls] + [('fc', u) for u in fc_urls]
    
    for kind, url in all_urls:
        if url in progress and progress[url].get('done'):
            continue
        
        hdrs = RM_HEADERS if kind == 'rm' else FC_HEADERS
        html = fetch(url, hdrs=hdrs)
        if kind == 'rm':
            plot = extract_rightmove_plot(html)
        else:
            plot = extract_fotocasa_plot(html)
        
        progress[url] = {'plot': plot, 'done': True}
        if plot:
            total_found += 1
            print(f"  ✓ {kind} plot={plot}m² | ...{url[-55:]}", flush=True)
        
        counter += 1
        if counter % SAVE_EVERY == 0:
            save_progress(progress)
            print(f"[{counter}/{len(all_urls)}] saved. Found: {total_found}", flush=True)
        
        time.sleep(DELAY + random.uniform(0, 0.5))
    
    save_progress(progress)
    updated = update_excel(progress)
    print(f"\n=== DONE === Scraped: {counter} | Plot found: {total_found} | Excel updated: {updated}", flush=True)
    
    pool = compute_pool(progress)
    print(f"\n=== POOL (≥5Z, ≥2.9M€, kein Nordost, ≥7000m²): {len(pool)} Objekte ===")
    for p in sorted(pool, key=lambda x: -x['plot'])[:30]:
        print(f"  {p['rooms']}Z | {p['plot']}m² | €{p['price']/1e6:.2f}M | {p['loc']} | {p['url'][-50:]}")

if __name__ == '__main__':
    main()
