import requests
from bs4 import BeautifulSoup
import json, time, re
from openpyxl import load_workbook
from datetime import date

headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'}

wb = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
ws = wb['Mallorca Objekte']
existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')

count = 0

# === Versuch 1: Apify Generic Scraper mit Von Poll ===
APIFY_TOKEN = "apify_api_feD2KhARHjtuV9CrSwOReYgoePFSF44nsDL6"

# Suche nach Generic Web Scraper Actor
r = requests.get('https://api.apify.com/v2/store?search=web+scraper&limit=10',
                 headers={'Authorization': f'Bearer {APIFY_TOKEN}'})
actors = r.json().get('data', {}).get('items', [])
for a in actors:
    print(f"{a.get('id')} | {a.get('name')}")

# Versuche apify/web-scraper
actor_id = 'apify~web-scraper'
run_input = {
    "startUrls": [{"url": "https://www.von-poll.com/de/immobilien/spanien/balearen/mallorca"}],
    "pageFunction": """async function pageFunction(context) {
        const { $, request, log } = context;
        const listings = [];
        $('.expose-item, .property-item, article, [class*="expose"]').each((i, el) => {
            const $el = $(el);
            const title = $el.find('h2, h3, .title').first().text().trim();
            const price = $el.find('.price, [class*="price"]').first().text().trim();
            const url = $el.find('a').first().attr('href');
            if (title || price) listings.push({ title, price, url });
        });
        return listings;
    }""",
    "maxPagesPerCrawl": 20,
}

run_id = None
dataset_id = None
status = None

try:
    r = requests.post(
        f'https://api.apify.com/v2/acts/{actor_id}/runs',
        json=run_input,
        headers={'Authorization': f'Bearer {APIFY_TOKEN}', 'Content-Type': 'application/json'}
    )
    print(f"Apify run start: {r.status_code} | {r.text[:500]}")
    if r.status_code in [200, 201]:
        run_data = r.json().get('data', {})
        run_id = run_data.get('id')
        dataset_id = run_data.get('defaultDatasetId')
        print(f"Run ID: {run_id}")
        
        # Warte auf Ergebnis
        for _ in range(30):
            time.sleep(10)
            status_r = requests.get(
                f'https://api.apify.com/v2/acts/{actor_id}/runs/{run_id}',
                headers={'Authorization': f'Bearer {APIFY_TOKEN}'}
            )
            status = status_r.json().get('data', {}).get('status', '')
            print(f"Status: {status}")
            if status in ['SUCCEEDED', 'FAILED', 'ABORTED']:
                break
        
        if status == 'SUCCEEDED':
            results_r = requests.get(
                f'https://api.apify.com/v2/datasets/{dataset_id}/items?limit=500',
                headers={'Authorization': f'Bearer {APIFY_TOKEN}'}
            )
            items = results_r.json()
            print(f"Results: {len(items)} items")
            for item in items:
                if isinstance(item, list):
                    for obj in item:
                        url = obj.get('url','')
                        if url and not url.startswith('http'):
                            url = 'https://www.von-poll.com' + url
                        if url and url not in existing_urls:
                            ws.append([obj.get('title','Von Poll'), 'Von Poll Real Estate', url,
                                       None, None, None, None, 'Mallorca', str(date.today()), 'Neu'])
                            existing_urls.add(url)
                            count += 1
                elif isinstance(item, dict):
                    url = item.get('url','')
                    if url and not url.startswith('http'):
                        url = 'https://www.von-poll.com' + url
                    if url and url not in existing_urls:
                        ws.append([item.get('title','Von Poll'), 'Von Poll Real Estate', url,
                                   None, None, None, None, 'Mallorca', str(date.today()), 'Neu'])
                        existing_urls.add(url)
                        count += 1
            print(f"Apify Objekte: {count}")
except Exception as e:
    print(f"Apify error: {e}")

# === Versuch 2: Brave Search nach Von Poll Mallorca Listings ===
print("\n=== Brave Suche ===")
search_queries = [
    'site:von-poll.com mallorca villa',
    'site:von-poll.com balearen immobilien',
    'von-poll.com immobilien mallorca kaufen',
    'von poll mallorca villa kaufen site:von-poll.com',
]

for query in search_queries:
    try:
        r = requests.get(
            'https://api.search.brave.com/res/v1/web/search',
            params={'q': query, 'count': 20},
            headers={
                'Accept': 'application/json',
                'Accept-Encoding': 'gzip',
                'X-Subscription-Token': 'BSA5KWR8-QpVt9-S1fSCZEjAY2bdC0M'
            },
            timeout=15
        )
        print(f"Brave status: {r.status_code} for '{query}'")
        if r.status_code == 200:
            results = r.json().get('web', {}).get('results', [])
            print(f"  {len(results)} results")
            for res in results:
                url = res.get('url', '')
                if 'von-poll.com' in url and '/immobilien/' in url and url not in existing_urls:
                    titel = res.get('title', 'Von Poll Objekt')
                    ws.append([titel[:100], 'Von Poll Real Estate', url, None, None, None, None, 'Mallorca', str(date.today()), 'Neu'])
                    existing_urls.add(url)
                    count += 1
                    print(f"  Added: {url}")
        else:
            print(f"  Response: {r.text[:200]}")
    except Exception as e:
        print(f"Search error: {e}")
    time.sleep(1)

wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
print(f"\n✅ Von Poll: {count} Objekte gespeichert")
