#!/usr/bin/env python3
import openpyxl
import json
import re
from typing import Dict, Optional

def clean_text(text: str) -> str:
    """Clean text from HTML tags and normalize whitespace"""
    if not text:
        return ""
    # Remove HTML tags
    text = re.sub(r'<[^>]+>', ' ', text)
    # Normalize whitespace
    text = ' '.join(text.split())
    return text.strip()

def extract_number(text: str) -> Optional[float]:
    """Extract first number from text"""
    if not text:
        return None
    # Look for numbers with possible decimal separators
    match = re.search(r'([\d.,]+)', clean_text(text))
    if match:
        num_str = match.group(1).replace(',', '.')
        # Handle cases like "1.000.000" (German number format)
        if num_str.count('.') > 1:
            parts = num_str.split('.')
            num_str = ''.join(parts[:-1]) + '.' + parts[-1]
        try:
            return float(num_str)
        except ValueError:
            return None
    return None

def extract_property_data(content: str, url: str) -> Dict:
    """Extract property data from webpage content"""
    data = {
        'zimmer': None,
        'baeder': None,
        'grundstueck_m2': None,
        'bebaute_flaeche_m2': None,
        'location': None,
        'preis_euro': None
    }
    
    content_lower = content.lower()
    
    # Extract rooms/bedrooms
    room_patterns = [
        r'(\d+)\s*schlafzimmer',
        r'(\d+)\s*zimmer',
        r'(\d+)\s*bedroom',
        r'(\d+)\s*room',
        r'bedrooms?\s*:?\s*(\d+)',
        r'rooms?\s*:?\s*(\d+)',
        r'habitaciones?\s*:?\s*(\d+)',
    ]
    
    for pattern in room_patterns:
        match = re.search(pattern, content_lower)
        if match and not data['zimmer']:
            data['zimmer'] = int(match.group(1))
            break
    
    # Extract bathrooms
    bath_patterns = [
        r'(\d+)\s*badezimmer',
        r'(\d+)\s*bad',
        r'(\d+)\s*bathroom',
        r'bathrooms?\s*:?\s*(\d+)',
        r'baños?\s*:?\s*(\d+)',
    ]
    
    for pattern in bath_patterns:
        match = re.search(pattern, content_lower)
        if match and not data['baeder']:
            data['baeder'] = int(match.group(1))
            break
    
    # Extract plot size (Grundstücksgröße)
    plot_patterns = [
        r'grundstück\s*:?\s*([\d.,]+)\s*m',
        r'plot\s*:?\s*([\d.,]+)\s*m',
        r'terreno\s*:?\s*([\d.,]+)\s*m',
        r'parcela\s*:?\s*([\d.,]+)\s*m',
        r'land\s*:?\s*([\d.,]+)\s*m',
        r'([\d.,]+)\s*m²?\s*grundstück',
        r'([\d.,]+)\s*m²?\s*plot',
        r'([\d.,]+)\s*m²?\s*land',
    ]
    
    for pattern in plot_patterns:
        match = re.search(pattern, content_lower)
        if match and not data['grundstueck_m2']:
            num = extract_number(match.group(1))
            if num and num > 100:  # Reasonable minimum for plot size
                data['grundstueck_m2'] = num
                break
    
    # Extract built area (Bebaute Fläche)
    built_patterns = [
        r'wohnfläche\s*:?\s*([\d.,]+)\s*m',
        r'built\s*:?\s*([\d.,]+)\s*m',
        r'construida\s*:?\s*([\d.,]+)\s*m',
        r'living\s*area\s*:?\s*([\d.,]+)\s*m',
        r'floor\s*area\s*:?\s*([\d.,]+)\s*m',
        r'([\d.,]+)\s*m²?\s*wohnfläche',
        r'([\d.,]+)\s*m²?\s*built',
        r'([\d.,]+)\s*m²?\s*living',
    ]
    
    for pattern in built_patterns:
        match = re.search(pattern, content_lower)
        if match and not data['bebaute_flaeche_m2']:
            num = extract_number(match.group(1))
            if num and num > 50:  # Reasonable minimum for built area
                data['bebaute_flaeche_m2'] = num
                break
    
    # Extract location
    location_patterns = [
        r'(?:location|ort|lugar|ubicación)\s*:?\s*([^\n,;]{5,50})',
        r'(?:in|en)\s+([A-ZÄÖÜ][a-zäöü\s-]{3,30})',
        # Specific Mallorca locations
        r'(palma|alcudia|pollensa|port\s*d[\'e]\s*pollensa|cala\s*\w+|puerto\s*\w+|santa\s*\w+|son\s*\w+)',
    ]
    
    for pattern in location_patterns:
        match = re.search(pattern, content_lower)
        if match and not data['location']:
            loc = clean_text(match.group(1))
            if len(loc) > 2:
                data['location'] = loc.title()
                break
    
    # Extract price
    price_patterns = [
        r'preis\s*:?\s*([\d.,]+)(?:\s*€|\s*euro)',
        r'price\s*:?\s*([\d.,]+)(?:\s*€|\s*euro)',
        r'precio\s*:?\s*([\d.,]+)(?:\s*€|\s*euro)',
        r'€\s*([\d.,]+)',
        r'([\d.,]+)\s*€',
        r'([\d.,]+)\s*euro',
    ]
    
    for pattern in price_patterns:
        matches = re.findall(pattern, content_lower)
        for match in matches:
            num = extract_number(match)
            if num and num > 10000:  # Reasonable minimum price
                data['preis_euro'] = num
                break
        if data['preis_euro']:
            break
    
    return data

def analyze_excel():
    """Analyze the Excel file structure"""
    try:
        workbook = openpyxl.load_workbook('/Users/robin/.openclaw/media/inbound/file_2---d7f26368-0190-4df2-948d-5eeb41356988.xlsx')
        sheet = workbook.active
        
        print("Excel file analysis:")
        print(f"Sheet name: {sheet.title}")
        print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")
        
        # Check headers and URLs
        print("\nHeaders (Row 1):")
        for col in range(1, min(sheet.max_column + 1, 20)):  # First 20 columns
            cell = sheet.cell(row=1, column=col)
            if cell.value:
                print(f"Column {col} ({chr(64+col)}): {cell.value}")
        
        print("\nURLs found in column C:")
        urls = []
        for row in range(2, 14):  # Rows 2-13
            cell = sheet.cell(row=row, column=3)  # Column C
            if cell.value:
                urls.append(cell.value)
                print(f"Row {row}: {cell.value}")
        
        return urls
    except Exception as e:
        print(f"Error analyzing Excel: {e}")
        return []

if __name__ == "__main__":
    urls = analyze_excel()
    print(f"\nFound {len(urls)} URLs to process")