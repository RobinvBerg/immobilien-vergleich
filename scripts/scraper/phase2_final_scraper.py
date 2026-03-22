#!/usr/bin/env python3
"""
Phase 2 Final: Comprehensive Mallorca Real Estate Scraper
Uses both requests and Playwright based on what works for each site
"""
import sys, json, re, time, traceback
from datetime import date
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
HEADERS = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36', 'Accept-Language': 'de,en;q=0.9'}

all_results = []
summary = {}

def clean_price(txt):
    if not txt: return None
    txt = str(txt).replace('\xa0','').replace('\u202f','').replace(' ','').replace('.','').replace(',','').replace('€','').replace('EUR','').replace('$','')
    m = re.search(r'(\d{4,})', txt)
    if m:
        v = int(m.group(1))
        return v
    return None

def clean_int(txt):
    if not txt: return None
    txt = str(txt).replace('.','').replace(',','').replace('\xa0','').replace(' ','').replace('m²','').replace('m2','').replace('sqm','')
    m = re.search(r'(\d+)', txt)
    return int(m.group(1)) if m else None

def req(url, timeout=15):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        return r
    except Exception as e:
        print(f"    req error {url[:60]}: {e}")
        return None

def save_results(new_objects, label=""):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set(str(row[2]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[2])
    n = 0
    for obj in new_objects:
        url = obj.get('url','—')
        if url and url != '—' and url in existing_urls: continue
        ws.append([obj.get('titel','?'), obj.get('quelle',''), url, obj.get('preis'), obj.get('zimmer'), obj.get('grundstueck'), obj.get('wohnflaeche'), obj.get('ort',''), str(date.today()), 'Neu'])
        if url and url != '—': existing_urls.add(url)
        n += 1
    wb.save(EXCEL_PATH)
    if label:
        print(f"    💾 Saved {n} new ({label})")
    return n

# ─────────────────────────────────────────────────────────────
# 1. PORTA MALLORQUINA (portamallorquina.com)
# ─────────────────────────────────────────────────────────────
def scrape_portamallorquina():
    print("\n[1] Porta Mallorquina...")
    results = []
    # Their main listing page - they use Wordpress with custom post types
    for url in ["https://www.portamallorquina.com/kaufen/", "https://www.portamallorquina.com/kaufen/?per_page=100"]:
        r = req(url)
        if not r or r.status_code != 200:
            continue
        soup = BeautifulSoup(r.text, 'html.parser')
        # Their items are lazy-loaded, but links should be in HTML
        # Find all article/property links
        links = soup.select('a[href*="/expose/"], a[href*="/property/"], a[href*="/immobilie/"], a[href*="/objekt/"]')
        print(f"  Found {len(links)} expose links at {url}")
        for link in links[:80]:
            href = link['href']
            if not href.startswith('http'):
                href = 'https://www.portamallorquina.com' + href
            title = link.get_text(strip=True)
            if not title:
                parent = link.parent
                for _ in range(3):
                    if parent:
                        t = parent.select_one('h2,h3,h4,.title')
                        if t:
                            title = t.get_text(strip=True)
                            break
                        parent = parent.parent
            if href not in [o.get('url') for o in results]:
                results.append({'titel': title or 'Porta Mallorquina', 'quelle': 'Porta Mallorquina', 'url': href})
        if results: break
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 2. UNIQUE MALLORCA (uniquemallorca.com/property-sales/)
# ─────────────────────────────────────────────────────────────
def scrape_unique_mallorca():
    print("\n[2] Unique Mallorca...")
    results = []
    page = 1
    while page <= 20:
        url = f"https://www.uniquemallorca.com/property-sales/?page={page}" if page > 1 else "https://www.uniquemallorca.com/property-sales/"
        r = req(url)
        if not r or r.status_code != 200:
            break
        soup = BeautifulSoup(r.text, 'html.parser')
        articles = soup.select('article')
        if not articles:
            break
        print(f"  Page {page}: {len(articles)} articles")
        for art in articles:
            # Title from line-clamp div or first link
            title_el = art.select_one('.line-clamp, h2, h3, h4')
            title = title_el.get_text(strip=True) if title_el else ''
            # Price
            price_el = art.select_one('[class*="price"], [class*="Price"]')
            price = clean_price(price_el.get_text() if price_el else None)
            # Link
            link_el = art.select_one('a[href]')
            href = link_el['href'] if link_el else url
            if href and not href.startswith('http'):
                href = 'https://www.uniquemallorca.com' + href
            # Features: small-features contains sizes and bedrooms
            feat = art.select_one('.small-features')
            feat_text = feat.get_text(separator=' ', strip=True) if feat else ''
            # Extract: "2300 m² 4256 m² 79" or similar
            areas = re.findall(r'(\d+)\s*m[²2]', feat_text)
            bedrooms = re.search(r'(\d+)\s*(?:bed|bedroom|schlafzimmer)', feat_text, re.I)
            plot = areas[1] if len(areas) >= 2 else (areas[0] if areas else None)
            built = areas[0] if len(areas) >= 2 else None
            # Location
            loc_el = art.select_one('.subtitle, [class*="location"], [class*="region"]')
            loc = ''
            if loc_el:
                loc_text = loc_el.get_text(strip=True)
                # "Property for sale in Santanyi" → "Santanyi"
                m = re.search(r'(?:in|in)\s+(.+?)(?:\s*-|\s*\||\s*$)', loc_text, re.I)
                loc = m.group(1).strip() if m else loc_text[:50]
            results.append({
                'titel': title or 'Unique Mallorca Property',
                'quelle': 'Unique Mallorca',
                'url': href,
                'preis': price,
                'grundstueck': clean_int(plot),
                'wohnflaeche': clean_int(built),
                'ort': loc,
            })
        # Check for next page
        next_btn = soup.select_one('a.next, a[rel="next"], .pagination a[href*="page"]')
        if not next_btn:
            break
        page += 1
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 3. LUCAS FOX (lucasfox.com/property/spain/balearic-islands/mallorca.html)
# ─────────────────────────────────────────────────────────────
def scrape_lucas_fox():
    print("\n[3] Lucas Fox...")
    results = []
    page = 1
    while page <= 20:
        url = f"https://www.lucasfox.com/property/spain/balearic-islands/mallorca.html?page={page}"
        r = req(url)
        if not r or r.status_code != 200:
            break
        soup = BeautifulSoup(r.text, 'html.parser')
        cards = soup.select('.property-card')
        if not cards:
            break
        print(f"  Page {page}: {len(cards)} cards")
        for card in cards:
            # Extract from the text content
            text = card.get_text(separator='|', strip=True)
            # Price: usually "€4,200,000"
            price_m = re.search(r'€[\d,\.]+', text)
            price = clean_price(price_m.group(0) if price_m else None)
            # Link
            link = card.select_one('a[href]')
            href = link['href'] if link else ''
            if href and not href.startswith('http'):
                href = 'https://www.lucasfox.com' + href
            # Title
            title_el = card.select_one('h2, h3, h4, .property-title, .title, [class*="title"]')
            if not title_el:
                # Get from link text or first text segment
                parts = [p for p in text.split('|') if len(p) > 10 and '€' not in p]
                title = parts[0][:100] if parts else 'Lucas Fox Property'
            else:
                title = title_el.get_text(strip=True)
            # Location
            loc_el = card.select_one('[class*="location"], [class*="area"], [class*="region"]')
            loc = loc_el.get_text(strip=True)[:60] if loc_el else ''
            # Features (beds, area, plot)
            beds_el = card.select_one('[class*="bed"], [class*="room"]')
            area_el = card.select_one('[class*="area"], [class*="m2"], [class*="sqm"]')
            # From text: "580m² country house"
            area_m = re.search(r'(\d+)\s*m²\s*(?:country|villa|house|finca|apartment)?', text, re.I)
            results.append({
                'titel': title,
                'quelle': 'Lucas Fox',
                'url': href,
                'preis': price,
                'wohnflaeche': clean_int(area_m.group(1)) if area_m else None,
                'ort': loc,
            })
        page += 1
        if len(cards) < 6:  # last page
            break
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 4. KENSINGTON INTERNATIONAL (kensington-international.com)
# ─────────────────────────────────────────────────────────────
def scrape_kensington():
    print("\n[4] Kensington International...")
    results = []
    base = "https://kensington-international.com"
    
    # They have area-based pages
    areas = [
        "/en/es/mallorca/properties-palma-de-mallorca",
        "/en/es/mallorca/properties-soller",
        "/en/es/mallorca/properties-andratx",
        "/en/es/mallorca/properties-calvia",
        "/en/es/mallorca/properties-southwest",
        "/en/es/mallorca/properties-northwest",
        "/en/es/mallorca/",
    ]
    for area_path in areas:
        url = base + area_path
        r = req(url)
        if not r or r.status_code != 200:
            continue
        soup = BeautifulSoup(r.text, 'html.parser')
        t = soup.select_one('title')
        print(f"  {t.text[:50] if t else 'no title'} at {url}")
        
        for sel in ['.property-card', '.listing-item', 'article', '.property', '[class*="property-card"]', '.estate']:
            cards = soup.select(sel)
            if len(cards) > 1:
                print(f"    {len(cards)} cards with '{sel}'")
                for card in cards[:50]:
                    title_el = card.select_one('h2, h3, h4, .title')
                    price_el = card.select_one('[class*="price"]')
                    link_el = card.select_one('a[href]')
                    if title_el or price_el:
                        href = link_el['href'] if link_el else ''
                        if href and not href.startswith('http'):
                            href = base + href
                        results.append({
                            'titel': title_el.get_text(strip=True) if title_el else 'Kensington',
                            'quelle': 'Kensington Finest',
                            'url': href or url,
                            'preis': clean_price(price_el.get_text() if price_el else None),
                        })
                break
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 5. MALLORCA DREAM HOMES (mallorcadreamhomes.com/ventas)
# ─────────────────────────────────────────────────────────────
def scrape_mallorca_dream():
    print("\n[5] Mallorca Dream Homes...")
    results = []
    base = "https://www.mallorcadreamhomes.com"
    r = req(base + "/ventas")
    if not r:
        return results
    soup = BeautifulSoup(r.text, 'html.parser')
    # Find all REF- links
    seen = set()
    for link in soup.select('a[href]'):
        href = link.get('href', '')
        if 'REF-' in href.upper():
            # Make absolute
            if not href.startswith('http'):
                href = base + '/' + href.lstrip('./')
            if href in seen:
                continue
            seen.add(href)
            # Get title from sibling/parent text
            text = link.get_text(strip=True)
            parent = link.parent
            title = text
            for _ in range(3):
                if parent and len(parent.get_text(strip=True)) > 20:
                    t = parent.select_one('h2,h3,h4,.title,p')
                    if t and len(t.get_text(strip=True)) > 5:
                        title = t.get_text(strip=True)
                        break
                if parent:
                    parent = parent.parent
            # Get price from siblings
            price = None
            if link.parent:
                price_el = link.parent.select_one('[class*="price"]')
                if not price_el and link.parent.parent:
                    price_el = link.parent.parent.select_one('[class*="price"]')
                if price_el:
                    price = clean_price(price_el.get_text())
                else:
                    # Try to find price in text
                    full_text = ''
                    p = link.parent
                    for _ in range(4):
                        if p:
                            full_text += p.get_text(' ', strip=True)
                            p = p.parent
                    pm = re.search(r'€\s*[\d\.,]+', full_text)
                    if pm:
                        price = clean_price(pm.group(0))
            if not title or title == href:
                title = href.split('/')[-1].replace('-', ' ')
            results.append({
                'titel': title[:100],
                'quelle': 'Mallorca Dream Homes',
                'url': href,
                'preis': price,
                'ort': 'Sóller',
            })
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 6. PRIVATE PROPERTY MALLORCA (privatepropertymallorca.com)
# ─────────────────────────────────────────────────────────────
def scrape_private_property():
    print("\n[6] Private Property Mallorca...")
    results = []
    base = "https://www.privatepropertymallorca.com"
    r = req(base + "/immo-suche/")
    if not r:
        return results
    soup = BeautifulSoup(r.text, 'html.parser')
    print(f"  Page length: {len(r.text)}")
    
    # Their listing uses Wordpress with property plugin
    for sel in ['.property-item', '.listing-item', 'article', '.property', '.property-card', '.oi-item', '[class*="property"]', '.wp-block-column']:
        cards = soup.select(sel)
        if len(cards) > 2:
            print(f"  {len(cards)} cards with '{sel}'")
            for card in cards[:50]:
                title_el = card.select_one('h2, h3, h4, .title, a')
                price_el = card.select_one('[class*="price"]')
                link_el = card.select_one('a[href]')
                if title_el and len(title_el.get_text(strip=True)) > 5:
                    href = link_el['href'] if link_el else ''
                    if href and not href.startswith('http'):
                        href = base + href
                    results.append({
                        'titel': title_el.get_text(strip=True)[:100],
                        'quelle': 'Private Property Mallorca',
                        'url': href or base,
                        'preis': clean_price(price_el.get_text() if price_el else None),
                    })
            break
    
    if not results:
        # Try scraping individual region pages
        region_urls = [
            base + "/immobilien/kaufen/",
            base + "/regions/andratx/",
            base + "/regions/son-vida/",
            base + "/regions/soller/",
        ]
        for rurl in region_urls:
            r2 = req(rurl)
            if not r2: continue
            soup2 = BeautifulSoup(r2.text, 'html.parser')
            for sel in ['article', '.property', '.listing-item', '[class*="property"]']:
                cards = soup2.select(sel)
                if len(cards) > 2:
                    for card in cards[:30]:
                        title_el = card.select_one('h2, h3, h4, .title')
                        price_el = card.select_one('[class*="price"]')
                        link_el = card.select_one('a[href]')
                        if title_el:
                            href = link_el['href'] if link_el else ''
                            if href and not href.startswith('http'):
                                href = base + href
                            results.append({
                                'titel': title_el.get_text(strip=True)[:100],
                                'quelle': 'Private Property Mallorca',
                                'url': href or rurl,
                                'preis': clean_price(price_el.get_text() if price_el else None),
                            })
                    break
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 7. SANDBERG ESTATES (sandberg-estates.com) - needs Playwright
# ─────────────────────────────────────────────────────────────
def scrape_sandberg():
    print("\n[7] Sandberg Estates...")
    results = []
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(user_agent=HEADERS['User-Agent'])
            page = context.new_page()
            page.goto("https://sandberg-estates.com/properties/for-sale/", timeout=30000, wait_until='domcontentloaded')
            page.wait_for_timeout(3000)
            html = page.content()
            browser.close()
        
        soup = BeautifulSoup(html, 'html.parser')
        t = soup.select_one('title')
        print(f"  Page: {t.text[:60] if t else 'no title'}")
        
        for sel in ['.property-item', 'article', '.property', '.property-card', '[class*="property"]', '.listing']:
            cards = soup.select(sel)
            if len(cards) > 2:
                print(f"  {len(cards)} cards with '{sel}'")
                for card in cards[:60]:
                    title_el = card.select_one('h2, h3, h4, .title')
                    price_el = card.select_one('[class*="price"]')
                    link_el = card.select_one('a[href]')
                    if title_el:
                        href = link_el['href'] if link_el else ''
                        if href and not href.startswith('http'):
                            href = 'https://sandberg-estates.com' + href
                        results.append({
                            'titel': title_el.get_text(strip=True)[:100],
                            'quelle': 'Sandberg Estates',
                            'url': href,
                            'preis': clean_price(price_el.get_text() if price_el else None),
                        })
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 8. MINKNER & BONITZ - Playwright needed (JS-rendered)
# ─────────────────────────────────────────────────────────────
def scrape_minkner():
    print("\n[8] Minkner & Bonitz...")
    results = []
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(user_agent=HEADERS['User-Agent'])
            page = context.new_page()
            page.goto("https://www.minkner.com/immobilien/kaufen/", timeout=30000, wait_until='domcontentloaded')
            page.wait_for_timeout(4000)
            html = page.content()
            browser.close()
        
        soup = BeautifulSoup(html, 'html.parser')
        t = soup.select_one('title')
        print(f"  Page: {t.text[:60] if t else 'no title'}, len={len(html)}")
        
        # They use lazy-loaded div.item - check if now populated
        items = soup.select('div.item')
        print(f"  div.item: {len(items)}")
        if items:
            for item in items[:3]:
                print(f"    Text: {item.get_text(separator='|',strip=True)[:100]}")
        
        for sel in ['.expose-item', '.property-item', 'article', '.property', 'div.item', '.listing', '.expose']:
            cards = soup.select(sel)
            if len(cards) > 2:
                sample = cards[0].get_text(separator='|', strip=True)
                if len(sample) > 20:  # has content
                    print(f"  {len(cards)} cards with '{sel}'")
                    print(f"  Sample: {sample[:120]}")
                    for card in cards[:80]:
                        title_el = card.select_one('h2, h3, h4, .title, .expose-title, [class*="title"]')
                        price_el = card.select_one('[class*="price"], [class*="preis"], .kaufpreis')
                        link_el = card.select_one('a[href]')
                        if title_el or (link_el and link_el['href'].count('/') > 3):
                            href = link_el['href'] if link_el else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.minkner.com' + href
                            results.append({
                                'titel': title_el.get_text(strip=True)[:100] if title_el else 'Minkner Objekt',
                                'quelle': 'Minkner & Bonitz',
                                'url': href,
                                'preis': clean_price(price_el.get_text() if price_el else None),
                            })
                    break
        
        # If still no results, extract from expose links
        if not results:
            expose_links = soup.select('a[href*="/expose/"]')
            print(f"  Expose links: {len(expose_links)}")
            seen = set()
            for link in expose_links:
                href = link['href']
                if not href.startswith('http'):
                    href = 'https://www.minkner.com' + href
                if href in seen: continue
                seen.add(href)
                title = link.get_text(strip=True) or 'Minkner Objekt'
                # Look for price nearby
                parent = link.parent
                price = None
                for _ in range(4):
                    if parent:
                        pe = parent.select_one('[class*="preis"], [class*="price"]')
                        if pe:
                            price = clean_price(pe.get_text())
                            break
                        parent = parent.parent
                results.append({'titel': title[:100], 'quelle': 'Minkner & Bonitz', 'url': href, 'preis': price})
                
    except Exception as e:
        print(f"  Error: {e}")
        traceback.print_exc()
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 9. KNIGHT FRANK - Playwright  
# ─────────────────────────────────────────────────────────────
def scrape_knight_frank():
    print("\n[9] Knight Frank...")
    results = []
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(user_agent=HEADERS['User-Agent'])
            page = context.new_page()
            page.goto("https://www.knightfrank.com/properties/residential/for-sale/spain-balearic-islands/all-types/5-and-more-beds", timeout=30000)
            page.wait_for_timeout(4000)
            html = page.content()
            browser.close()
        
        soup = BeautifulSoup(html, 'html.parser')
        t = soup.select_one('title')
        print(f"  Page: {t.text[:60] if t else 'no title'}, len={len(html)}")
        
        for sel in ['.listing-card', '.property-listing', '.property-card', 'article', '.property', '[class*="listing-card"]']:
            cards = soup.select(sel)
            if len(cards) > 1:
                sample = cards[0].get_text(separator='|', strip=True)
                if len(sample) > 20:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:50]:
                        title_el = card.select_one('h2, h3, h4, .title')
                        price_el = card.select_one('[class*="price"]')
                        link_el = card.select_one('a[href]')
                        if title_el:
                            href = link_el['href'] if link_el else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.knightfrank.com' + href
                            results.append({
                                'titel': title_el.get_text(strip=True)[:100],
                                'quelle': 'Knight Frank',
                                'url': href,
                                'preis': clean_price(price_el.get_text() if price_el else None),
                            })
                    break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 10. SOTHEBY'S - Playwright
# ─────────────────────────────────────────────────────────────
def scrape_sothebys():
    print("\n[10] Mallorca Sotheby's...")
    results = []
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(user_agent=HEADERS['User-Agent'])
            page = context.new_page()
            page.goto("https://www.sothebysrealty.com/eng/sales/mallorca-sp", timeout=30000)
            page.wait_for_timeout(5000)
            html = page.content()
            browser.close()
        
        soup = BeautifulSoup(html, 'html.parser')
        t = soup.select_one('title')
        print(f"  Page: {t.text[:60] if t else 'no title'}, len={len(html)}")
        
        for sel in ['.listing-card', '.property-card', 'article', '.property', '[class*="listing"]', '.card']:
            cards = soup.select(sel)
            if len(cards) > 1:
                sample = cards[0].get_text(separator='|', strip=True)
                if len(sample) > 20:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:50]:
                        title_el = card.select_one('h2, h3, h4, .title, [class*="title"]')
                        price_el = card.select_one('[class*="price"]')
                        link_el = card.select_one('a[href]')
                        if title_el:
                            href = link_el['href'] if link_el else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.sothebysrealty.com' + href
                            results.append({
                                'titel': title_el.get_text(strip=True)[:100],
                                'quelle': "Mallorca Sotheby's",
                                'url': href,
                                'preis': clean_price(price_el.get_text() if price_el else None),
                            })
                    break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 11. FINCALLORCA - Playwright (Cloudflare)
# ─────────────────────────────────────────────────────────────
def scrape_fincallorca():
    print("\n[11] Fincallorca...")
    results = []
    try:
        from playwright.sync_api import sync_playwright
        try:
            from playwright_stealth import Stealth
            stealth_available = True
        except:
            stealth_available = False
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(user_agent=HEADERS['User-Agent'])
            page = context.new_page()
            if stealth_available:
                try:
                    Stealth().use_sync(page)
                except:
                    pass
            page.goto("https://www.fincallorca.com/", timeout=30000, wait_until='domcontentloaded')
            page.wait_for_timeout(3000)
            html = page.content()
            final_url = page.url
            browser.close()
        
        soup = BeautifulSoup(html, 'html.parser')
        t = soup.select_one('title')
        print(f"  Page: {t.text[:60] if t else 'no title'} at {final_url}")
        
        # Find Mallorca buy links
        for link in soup.select('a[href]'):
            href = link.get('href', '')
            text = link.get_text(strip=True)
            if any(kw in href.lower() or kw in text.lower() for kw in ['kaufen', 'buy', 'sale', 'mallorca', 'immobil']):
                print(f"  Link: {href}: {text[:40]}")
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 12. VON POLL - Playwright
# ─────────────────────────────────────────────────────────────
def scrape_vonpoll():
    print("\n[12] Von Poll...")
    results = []
    try:
        from playwright.sync_api import sync_playwright
        try:
            from playwright_stealth import Stealth
            stealth_available = True
        except:
            stealth_available = False
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(user_agent=HEADERS['User-Agent'])
            page = context.new_page()
            if stealth_available:
                try:
                    Stealth().use_sync(page)
                except:
                    pass
            page.goto("https://www.von-poll.com/de/immobilien/spanien/balearen/mallorca", timeout=40000, wait_until='domcontentloaded')
            page.wait_for_timeout(5000)
            html = page.content()
            final_url = page.url
            browser.close()
        
        soup = BeautifulSoup(html, 'html.parser')
        t = soup.select_one('title')
        print(f"  Page: {t.text[:60] if t else 'no title'} ({len(html)} chars) at {final_url}")
        
        for sel in ['.expose-item', '.expose', '.property-item', 'article', '.property', '.listing-item']:
            cards = soup.select(sel)
            if len(cards) > 2:
                sample = cards[0].get_text(separator='|', strip=True)
                if len(sample) > 20:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:60]:
                        title_el = card.select_one('h2, h3, h4, [class*="title"]')
                        price_el = card.select_one('[class*="price"], [class*="preis"]')
                        link_el = card.select_one('a[href]')
                        if title_el:
                            href = link_el['href'] if link_el else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.von-poll.com' + href
                            results.append({
                                'titel': title_el.get_text(strip=True)[:100],
                                'quelle': 'Von Poll Real Estate',
                                'url': href,
                                'preis': clean_price(price_el.get_text() if price_el else None),
                            })
                    break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 13. DAHLER COMPANY (already worked, now with pagination)
# ─────────────────────────────────────────────────────────────
def scrape_dahler():
    print("\n[13] DAHLER Company...")
    results = []
    base = "https://www.dahlercompany.com"
    for page_n in range(1, 5):
        url = f"{base}/de/mallorca/immobilie-kaufen?page={page_n}" if page_n > 1 else f"{base}/de/mallorca/immobilie-kaufen"
        r = req(url)
        if not r or r.status_code != 200:
            break
        soup = BeautifulSoup(r.text, 'html.parser')
        
        for sel in ['article', '.property-card', '.expose', '.listing-item', '.property', '[class*="expose"]']:
            cards = soup.select(sel)
            if len(cards) > 0:
                for card in cards[:30]:
                    title_el = card.select_one('h2, h3, h4, .title')
                    price_el = card.select_one('[class*="price"]')
                    link_el = card.select_one('a[href]')
                    if title_el:
                        href = link_el['href'] if link_el else ''
                        if href and not href.startswith('http'):
                            href = base + href
                        results.append({
                            'titel': title_el.get_text(strip=True)[:100],
                            'quelle': 'DAHLER Company',
                            'url': href or url,
                            'preis': clean_price(price_el.get_text() if price_el else None),
                        })
                break
        if len(results) == 0:
            break
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 14. FINE & COUNTRY - Playwright
# ─────────────────────────────────────────────────────────────
def scrape_fine_country():
    print("\n[14] Fine & Country...")
    results = []
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(user_agent=HEADERS['User-Agent'])
            page = context.new_page()
            # Try to navigate and click through
            page.goto("https://www.fineandcountry.com/", timeout=20000, wait_until='domcontentloaded')
            page.wait_for_timeout(2000)
            html = page.content()
            soup = BeautifulSoup(html, 'html.parser')
            # Find property search/listing links
            for link in soup.select('a[href]'):
                href = link.get('href', '')
                text = link.get_text(strip=True)
                if any(kw in href.lower() or kw in text.lower() for kw in ['search', 'property', 'spain', 'mallorca', 'buy']):
                    print(f"  F&C link: {href}: {text[:40]}")
            browser.close()
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 15. COLDWELL BANKER (better scraper)
# ─────────────────────────────────────────────────────────────
def scrape_coldwell():
    print("\n[15] Coldwell Banker...")
    results = []
    
    # Their current URL structure
    for url in [
        "https://www.coldwellbanker.es/propiedades/?pais=ES&region=Islas+Baleares&habitaciones=5",
        "https://www.coldwellbanker.es/propiedades/?pais=ES&zona=Mallorca&habitaciones=5",
        "https://www.coldwellbanker.es/propiedades/?pais=ES&zona=Mallorca",
    ]:
        r = req(url)
        if not r: continue
        soup = BeautifulSoup(r.text, 'html.parser')
        t = soup.select_one('title')
        print(f"  {r.status_code}: {t.text[:50] if t else 'no title'}")
        for sel in ['.property', '.property-card', 'article', '.listing', '[class*="property"]']:
            cards = soup.select(sel)
            if len(cards) > 2:
                print(f"  {len(cards)} cards with '{sel}'")
                for card in cards[:50]:
                    title_el = card.select_one('h2, h3, h4, .title')
                    price_el = card.select_one('[class*="price"], [class*="precio"]')
                    link_el = card.select_one('a[href]')
                    loc_el = card.select_one('[class*="location"], [class*="zona"]')
                    if title_el:
                        href = link_el['href'] if link_el else ''
                        if href and not href.startswith('http'):
                            href = 'https://www.coldwellbanker.es' + href
                        results.append({
                            'titel': title_el.get_text(strip=True)[:100],
                            'quelle': 'Coldwell Banker',
                            'url': href or url,
                            'preis': clean_price(price_el.get_text() if price_el else None),
                            'ort': loc_el.get_text(strip=True)[:50] if loc_el else '',
                        })
                break
        if results: break
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 16. REMAX ES - Playwright
# ─────────────────────────────────────────────────────────────
def scrape_remax():
    print("\n[16] Re/Max...")
    results = []
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(user_agent=HEADERS['User-Agent'])
            page = context.new_page()
            page.goto("https://www.remax.es/", timeout=20000, wait_until='domcontentloaded')
            page.wait_for_timeout(2000)
            html = page.content()
            soup = BeautifulSoup(html, 'html.parser')
            # Find search/listing links
            for link in soup.select('a[href]'):
                href = link.get('href', '')
                text = link.get_text(strip=True)
                if any(kw in href.lower() or kw in text.lower() for kw in ['mallorca', 'baleares', 'comprar', 'buscar']):
                    print(f"  Link: {href[:60]}: {text[:40]}")
            # Try direct search
            page.goto("https://www.remax.es/Comprar?zona=mallorca", timeout=20000, wait_until='domcontentloaded')
            page.wait_for_timeout(3000)
            html = page.content()
            browser.close()
        
        soup = BeautifulSoup(html, 'html.parser')
        t = soup.select_one('title')
        print(f"  Page: {t.text[:60] if t else 'no title'}")
        
        for sel in ['.property-item', '.listing-item', 'article', '.property', '.propiedad']:
            cards = soup.select(sel)
            if len(cards) > 1:
                sample = cards[0].get_text(strip=True)
                if len(sample) > 20:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:50]:
                        title_el = card.select_one('h2, h3, h4, .title')
                        price_el = card.select_one('[class*="price"], [class*="precio"]')
                        link_el = card.select_one('a[href]')
                        if title_el:
                            href = link_el['href'] if link_el else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.remax.es' + href
                            results.append({
                                'titel': title_el.get_text(strip=True)[:100],
                                'quelle': 'Re/Max Mallorca',
                                'url': href,
                                'preis': clean_price(price_el.get_text() if price_el else None),
                            })
                    break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 17. ENGEL & VÖLKERS (EV) - already working, more pages
# ─────────────────────────────────────────────────────────────
def scrape_ev():
    print("\n[17] Engel & Völkers...")
    results = []
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(user_agent=HEADERS['User-Agent'])
            page = context.new_page()
            page.goto("https://www.engelvoelkers.com/de/search/?q=&domainId=21&businessArea=residential&mode=buy&country=ESP&categories=villa,finca&rooms=5", timeout=30000, wait_until='domcontentloaded')
            page.wait_for_timeout(4000)
            html = page.content()
            browser.close()
        
        soup = BeautifulSoup(html, 'html.parser')
        t = soup.select_one('title')
        print(f"  Page: {t.text[:60] if t else 'no title'}, len={len(html)}")
        
        for sel in ['article', '.property-card', '.ev-property-card', '.listing', '[class*="ev-property"]', '[class*="property"]']:
            cards = soup.select(sel)
            if len(cards) > 2:
                sample = cards[0].get_text(separator='|', strip=True)
                if len(sample) > 20:
                    print(f"  {len(cards)} cards with '{sel}'")
                    for card in cards[:60]:
                        title_el = card.select_one('h2, h3, h4, .title, [class*="title"]')
                        price_el = card.select_one('[class*="price"]')
                        link_el = card.select_one('a[href]')
                        loc_el = card.select_one('[class*="location"], [class*="region"], [class*="area"]')
                        if title_el:
                            href = link_el['href'] if link_el else ''
                            if href and not href.startswith('http'):
                                href = 'https://www.engelvoelkers.com' + href
                            results.append({
                                'titel': title_el.get_text(strip=True)[:100],
                                'quelle': 'Engel & Völkers',
                                'url': href,
                                'preis': clean_price(price_el.get_text() if price_el else None),
                                'ort': loc_el.get_text(strip=True)[:50] if loc_el else '',
                            })
                    break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 18. MALLORCA FINEST (SSL issue - try with requests ignore SSL)
# ─────────────────────────────────────────────────────────────
def scrape_mallorca_finest():
    print("\n[18] Mallorca Finest...")
    results = []
    try:
        import urllib3
        urllib3.disable_warnings()
        r = requests.get("https://mallorcafinest.com/", headers=HEADERS, timeout=15, verify=False)
        soup = BeautifulSoup(r.text, 'html.parser')
        t = soup.select_one('title')
        print(f"  Page: {r.status_code} - {t.text[:60] if t else 'no title'}")
        for link in soup.select('a[href]'):
            href = link.get('href', '')
            text = link.get_text(strip=True)
            if any(kw in href.lower() or kw in text.lower() for kw in ['kaufen', 'buy', 'sale', 'property']):
                print(f"  Link: {href}: {text[:40]}")
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# RUN ALL
# ─────────────────────────────────────────────────────────────
scrapers = [
    ("portamallorquina", scrape_portamallorquina),
    ("unique_mallorca", scrape_unique_mallorca),
    ("lucas_fox", scrape_lucas_fox),
    ("kensington", scrape_kensington),
    ("mallorca_dream", scrape_mallorca_dream),
    ("private_property", scrape_private_property),
    ("sandberg", scrape_sandberg),
    ("minkner", scrape_minkner),
    ("knight_frank", scrape_knight_frank),
    ("sothebys", scrape_sothebys),
    ("fincallorca", scrape_fincallorca),
    ("vonpoll", scrape_vonpoll),
    ("dahler", scrape_dahler),
    ("fine_country", scrape_fine_country),
    ("coldwell", scrape_coldwell),
    ("remax", scrape_remax),
    ("ev", scrape_ev),
    ("mallorca_finest", scrape_mallorca_finest),
]

total_saved = 0
for name, scraper in scrapers:
    try:
        items = scraper()
        all_results.extend(items)
        summary[name] = len(items)
        # Save incrementally
        if items:
            n = save_results(items, name)
            total_saved += n
    except Exception as e:
        print(f"  FATAL {name}: {e}")
        summary[name] = 0

# Save raw JSON
with open('/Users/robin/.openclaw/workspace/mallorca-projekt/phase2_final_raw.json', 'w') as f:
    json.dump(all_results, f, ensure_ascii=False, indent=2)

print(f"\n\n{'='*50}")
print(f"TOTAL RAW: {len(all_results)} Objekte")
print(f"TOTAL SAVED: {total_saved} neue Objekte in Excel")
print(f"\n--- Summary by source ---")
for name, count in sorted(summary.items(), key=lambda x: -x[1]):
    print(f"  {name}: {count}")
