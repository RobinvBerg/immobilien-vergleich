import requests
from bs4 import BeautifulSoup
import json, time, re
from openpyxl import load_workbook
from datetime import date

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.5',
}

# Load workbook
wb = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
ws = wb['Mallorca Objekte']
existing_urls = set(str(row[2]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[2] and row[2] != '—')

count = 0

savills_urls = [
    'https://www.savills.com/search?country=spain&region=mallorca&propertyType=residential&transactionType=sale&page=1',
    'https://www.savills.com/search#type=residential&market=sales&country=ESP&region=Mallorca&currency=EUR&page=1',
    'https://search.savills.com/es/en/list?SearchList=IsInRegion_EUR5002&SaleType=ForSale&PropertyType=RES&ResidentialActivity=Primary',
]

for url in savills_urls:
    try:
        r = requests.get(url, headers=headers, timeout=20)
        print(f"{url}: {r.status_code} | {len(r.text)} chars")
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')
            next_data = soup.find('script', {'id': '__NEXT_DATA__'})
            if next_data:
                data = json.loads(next_data.string)
                print(f"  __NEXT_DATA__ found! Keys: {list(data.keys())}")
                props = []
                def find_props(obj, depth=0):
                    if depth > 10: return
                    if isinstance(obj, list) and len(obj) > 0:
                        if isinstance(obj[0], dict) and any(k in obj[0] for k in ['price', 'bedrooms', 'id', 'address']):
                            props.extend(obj)
                    if isinstance(obj, dict):
                        for v in obj.values():
                            find_props(v, depth+1)
                find_props(data)
                print(f"  Found {len(props)} potential property records")
                if props:
                    print(f"  Sample: {json.dumps(props[0], indent=2)[:500]}")
                    for item in props:
                        url_prop = item.get('url', item.get('link', '—'))
                        if url_prop and not url_prop.startswith('http'):
                            url_prop = 'https://www.savills.com' + url_prop
                        if url_prop in existing_urls: continue
                        preis = item.get('price', item.get('Price'))
                        if isinstance(preis, dict): preis = preis.get('value', preis.get('amount'))
                        zimmer = item.get('bedrooms', item.get('rooms'))
                        ort = item.get('city', item.get('area', ''))
                        if isinstance(ort, dict): ort = ort.get('name', '')
                        if not ort and isinstance(item.get('address'), dict):
                            ort = item['address'].get('city', '')
                        titel = str(item.get('title', item.get('name', 'Savills listing')))[:100]
                        ws.append([titel, 'Savills', url_prop, preis, zimmer, None,
                                   item.get('surface'), str(ort), str(date.today()), 'Neu'])
                        if url_prop != '—': existing_urls.add(url_prop)
                        count += 1
                    wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
                    print(f"Savills: {count} gespeichert")
                    break
            else:
                print("  No __NEXT_DATA__ found")
                # Try to find any JSON with properties
                scripts = soup.find_all('script', type='application/json')
                print(f"  application/json scripts: {len(scripts)}")
                for s in scripts[:3]:
                    print(f"  Script content[:200]: {s.string[:200] if s.string else 'empty'}")
                
                # Try Savills API
                print("  Trying Savills API directly...")
                api_urls = [
                    'https://www.savills.com/api/search/search?Country=ESP&SearchList=IsInRegion_EUR5002&SaleType=ForSale',
                    'https://search.savills.com/api/search?country=ESP&area=Mallorca&saleType=ForSale',
                ]
                for api_url in api_urls:
                    try:
                        ar = requests.get(api_url, headers={**headers, 'Accept': 'application/json'}, timeout=15)
                        print(f"  API {api_url}: {ar.status_code} | {ar.text[:300]}")
                    except Exception as ae:
                        print(f"  API error: {ae}")
    except Exception as e:
        print(f"Error for {url}: {e}")
    time.sleep(1)

print(f"\n✅ Savills gesamt: {count} Objekte gespeichert")
