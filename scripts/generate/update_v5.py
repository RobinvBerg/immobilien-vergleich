#!/usr/bin/env python3
"""Update Excel + build HTML v5 for Mallorca project."""
import json, re, shutil

# ── Load reno data ──
with open('/Users/robin/.openclaw/workspace/mallorca-projekt/debug/reno_scores.json') as f:
    reno_data = json.load(f)

# ── TEIL 1: Excel update ──
import openpyxl
wb = openpyxl.load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/data/mallorca-objekte.xlsx')
ws = wb.active

# Headers in row 1
ws.cell(row=1, column=29, value="Baujahr (geschätzt)")
ws.cell(row=1, column=30, value="Letzte Renovierung (geschätzt)")
ws.cell(row=1, column=31, value="Reno-Score (0-100)")
ws.cell(row=1, column=32, value="Reno-Begründung")

for i, obj in enumerate(reno_data):
    row = i + 2  # rows 2-14
    ws.cell(row=row, column=29, value=obj['baujahr_est'])
    ws.cell(row=row, column=30, value=obj['last_reno_est'])
    ws.cell(row=row, column=31, value=obj['new_reno'])
    ws.cell(row=row, column=32, value=obj['reasoning'])

wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/data/mallorca-objekte.xlsx')
print("✅ Excel updated")

# ── TEIL 2: HTML v5 ──
src = '/Users/robin/.openclaw/workspace/mallorca-projekt/html/mallorca-ranking-v4.html'
dst = '/Users/robin/.openclaw/workspace/mallorca-projekt/html/mallorca-ranking-v5.html'
shutil.copy2(src, dst)

with open(dst, 'r') as f:
    html = f.read()

# 1. Replace reno values and add baujahr/lastReno to each object in DATA array
for obj in reno_data:
    old_reno = obj['old_reno']
    new_reno = obj['new_reno']
    baujahr = obj['baujahr_est']
    last_reno = obj['last_reno_est']

# Build replacement map: for each object, find "reno":X and replace + add fields
# The objects have "reno":4 or "reno":3 etc., but multiple objects share same value
# Better approach: find each object by a unique identifier like location or name snippet

# Strategy: find each "vermiet":XXX,"reno":Y pattern (vermiet is right before reno in the data)
# Actually looking at the data: ...charme:5,reno:4,bewirt:3,vermiet:100
# Let me check the actual order in the properties

# From the grep output, each object ends with: charme:X,reno:Y,bewirt:Z,vermiet:W
# So the pattern for each object is unique by the combination

# Better: replace sequentially - find each occurrence of the object pattern
# The objects appear in order in the DATA array. Let me replace them in order.

# Map of (old_reno, vermiet) pairs might not be unique. Let me use a different approach:
# Find each object by its unique name/location and replace reno value.

# Actually, simplest: the reno values appear in exact order in the file.
# Let me find all "reno":X occurrences (not in base64 data) and replace them in order.

# The DATA array has exactly 13 objects. Let me find the DATA block and work within it.

# Find the start and end of the DATA array
data_start = html.find('const DATA = [')
if data_start == -1:
    # Try alternate pattern
    data_start = html.find('const DATA=[')

# Find all reno occurrences within the DATA block
# Each object has pattern like: "reno":4  or reno:4
# Let's find the end of DATA array
data_end = html.find('];', data_start) + 2

data_block = html[data_start:data_end]

# Replace each reno value in order and add baujahr/lastReno
new_data_block = data_block
for i, obj in enumerate(reno_data):
    # Find the i-th occurrence of reno:X pattern
    # Pattern: "reno":X or reno:X  
    pass

# Different approach: use regex to find all reno:N patterns in the data block
# and replace them one by one in order
reno_pattern = re.compile(r'("reno"|reno)\s*:\s*(\d+)')
matches = list(reno_pattern.finditer(data_block))

print(f"Found {len(matches)} reno entries in DATA block")

# Replace from end to start to preserve indices
for i in range(len(matches) - 1, -1, -1):
    m = matches[i]
    obj = reno_data[i]
    old_text = m.group(0)
    key_part = m.group(1)  # "reno" or reno
    # New reno value
    new_reno_text = f'{key_part}:{obj["new_reno"]}'
    # Also add baujahr and lastReno before reno
    baujahr_text = f'baujahr:"{obj["baujahr_est"]}",lastReno:"{obj["last_reno_est"]}",'
    new_text = baujahr_text + new_reno_text
    
    start_idx = m.start()
    end_idx = m.end()
    new_data_block = new_data_block[:start_idx] + new_text + new_data_block[end_idx:]

html = html[:data_start] + new_data_block + html[data_end:]

# 2. Fix score calculation: reno is now 0-100, no normalization needed
# Old: const sRe = (o.reno/5)*100;
# New: const sRe = o.reno;
html = html.replace('const sRe = (o.reno/5)*100;', 'const sRe = o.reno;  // 0-100 scale, no normalization')

# 3. Update card rendering - reno badge
# Old: <span class="badge badge-info">Reno ${o.reno}/5</span>
# New: show baujahr + reno score
html = html.replace(
    '<span class="badge badge-info">Reno ${o.reno}/5</span>',
    '<span class="badge badge-info">🔧 ${o.reno}/100</span>'
)

# 4. Add Baujahr display to cards - add after location line
html = html.replace(
    '<div class="card-location">📍 ${o.location}</div>',
    '<div class="card-location">📍 ${o.location}</div>\n        <div class="card-location" style="font-size:0.85em;opacity:0.8">🏗️ Baujahr: ${o.baujahr||"?"} · Reno: ${o.lastReno||"?"}</div>'
)

# 5. Update sidebar slider label if needed
# The slider label is already generic: 'Renovierung' - that's fine for 0-100

with open(dst, 'w') as f:
    f.write(html)

print("✅ HTML v5 created")

# Verify
with open(dst, 'r') as f:
    content = f.read()

# Check reno values
for obj in reno_data:
    if f'reno:{obj["new_reno"]}' in content:
        print(f"  ✓ reno:{obj['new_reno']} found")
    else:
        print(f"  ✗ reno:{obj['new_reno']} NOT found")

if 'const sRe = o.reno;' in content:
    print("  ✓ Score calculation updated")
else:
    print("  ✗ Score calculation NOT updated")

if '🔧 ${o.reno}/100' in content:
    print("  ✓ Reno badge updated")
else:
    print("  ✗ Reno badge NOT updated")

if 'Baujahr' in content:
    print("  ✓ Baujahr display added")
else:
    print("  ✗ Baujahr display NOT added")
