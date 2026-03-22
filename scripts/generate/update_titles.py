#!/usr/bin/env python3
"""Update Excel and HTML with new property titles and descriptions."""
import sys
sys.modules['numpy'] = None
import openpyxl
import re

# Mapping: old HTML/Excel name -> (new title, description)
# Excel names from "Objekte" sheet column B
EXCEL_MAPPING = {
    "EV Exposé Binissalem": ("Binissalem — Platz für alle; 8 Zimmer, endlose Gärten", "Der Allrounder — genug Platz für große Abende unter freiem Himmel."),
    "EV – Finca mit Vermietlizenz nahe Es Trenc": ("Campos — Finca mit Lizenz; 14 Hektar mallorquinischer Traum", "Landleben, das sich selbst trägt — hier arbeitet die Finca für dich."),
    "Idealista – Gästehaus, viele Räume": ("Establiments — Charme-Refugium; wo Stein und Seele verschmelzen", "Wo Stein und Seele verschmelzen — ein Ort, der Geschichten erzählt."),
    "Villa Bunyola mit Lizenz - DS Poligon 8 Rustega": ("Bunyola — Berglage mit Lizenz; Tramuntana vor der Tür", "Panorama, Ruhe, Ankommen — 20 Minuten zu allem, was zählt."),
    "EV – Estate mit Weinberg & Olivenhain": ("Sencelles — 13 Hektar Paradies; dein eigenes Landgut", "Dein eigenes Königreich — Herrenhaus mit Geschichte, Platz ohne Ende."),
    "Idealista – Kreativprojekt": ("Sa Ràpita — Vision 1800; 5 Hektar warten auf deine Handschrift", "Wer hier baut, baut Legende — für Visionäre mit Geduld."),
    "Idealista – gute Substanz, Renovierung": ("Campos — Zum Selbermachen; ehrliche Finca, ehrlicher Preis", "Ehrlich, bodenständig, deins — für alle, die lieber selbst gestalten."),
    "EV – Contemporary Design Landscape": ("Moscari — Architekten-Traum; Design trifft Serra de Tramuntana", "Wohnen wie im Designmagazin — mitten in den Bergen."),
    "Santa Maria – teils unfertig": ("Santa Maria — Potenzial pur; Rohling in Traumlage", "Die Lage schreit nach Zukunft — wer's sieht, gewinnt."),
    "Idealista – Nähe Es Trenc": ("Sa Ràpita — Strandgold; aufwachen, Es Trenc, fertig", "Aufwachen, Salz auf der Haut, Sand unter den Füßen."),
    "EV – Unique Designer House": ("Ses Salines — Neubau Deluxe; einziehen und leben", "Koffer abstellen, ankommen, leben — nichts mehr tun müssen."),
    "Casa o chalet independiente en venta en Ds Poligon 8 Rustega, Bunyola (Idealista)": ("Establiments — Raum ohne Ende; 762m² pures Volumen", "Die Große — wer Wände verschieben kann, hat hier ein Monster."),
    "Idealista – modern, kleineres Grundstück": ("Palmanyola — Klein aber Wow; Designvilla, null Kompromisse", "Klein im Grundstück, groß im Auftritt — null Kompromisse."),
    # Also the old EV Exposé Bunyola
    "Engel & Völkers Exposé Bunyola": ("Bunyola — Berglage mit Lizenz; Tramuntana vor der Tür", "Panorama, Ruhe, Ankommen — 20 Minuten zu allem, was zählt."),
}

# HTML name mapping (these are the names actually in the HTML)
HTML_MAPPING = {
    "EV Exposé Binissalem": ("Binissalem — Platz für alle; 8 Zimmer, endlose Gärten", "Der Allrounder — genug Platz für große Abende unter freiem Himmel."),
    "Finca Campos (Lizenz)": ("Campos — Finca mit Lizenz; 14 Hektar mallorquinischer Traum", "Landleben, das sich selbst trägt — hier arbeitet die Finca für dich."),
    "Gästehaus Establiments": ("Establiments — Charme-Refugium; wo Stein und Seele verschmelzen", "Wo Stein und Seele verschmelzen — ein Ort, der Geschichten erzählt."),
    "Villa Bunyola (Lizenz)": ("Bunyola — Berglage mit Lizenz; Tramuntana vor der Tür", "Panorama, Ruhe, Ankommen — 20 Minuten zu allem, was zählt."),
    "Estate Sencelles": ("Sencelles — 13 Hektar Paradies; dein eigenes Landgut", "Dein eigenes Königreich — Herrenhaus mit Geschichte, Platz ohne Ende."),
    "Kreativprojekt Sa Ràpita": ("Sa Ràpita — Vision 1800; 5 Hektar warten auf deine Handschrift", "Wer hier baut, baut Legende — für Visionäre mit Geduld."),
    "Gute Substanz Campos": ("Campos — Zum Selbermachen; ehrliche Finca, ehrlicher Preis", "Ehrlich, bodenständig, deins — für alle, die lieber selbst gestalten."),
    "Contemporary Moscari": ("Moscari — Architekten-Traum; Design trifft Serra de Tramuntana", "Wohnen wie im Designmagazin — mitten in den Bergen."),
    "Santa Maria unfertig": ("Santa Maria — Potenzial pur; Rohling in Traumlage", "Die Lage schreit nach Zukunft — wer's sieht, gewinnt."),
    "Nähe Es Trenc": ("Sa Ràpita — Strandgold; aufwachen, Es Trenc, fertig", "Aufwachen, Salz auf der Haut, Sand unter den Füßen."),
    "Designer House Ses Salines": ("Ses Salines — Neubau Deluxe; einziehen und leben", "Koffer abstellen, ankommen, leben — nichts mehr tun müssen."),
    "Casa o chalet Establiments": ("Establiments — Raum ohne Ende; 762m² pures Volumen", "Die Große — wer Wände verschieben kann, hat hier ein Monster."),
    "Modern Palmanyola": ("Palmanyola — Klein aber Wow; Designvilla, null Kompromisse", "Klein im Grundstück, groß im Auftritt — null Kompromisse."),
}

# ---- Step 1: Update Excel ----
print("=== Updating Excel ===")
import os as _os
_ROOT = _os.path.dirname(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
wb = openpyxl.load_workbook(_os.path.join(_ROOT, 'data', 'mallorca-objekte.xlsx'))
ws = wb['Objekte']

# Find last used column for header
max_col = ws.max_column
# Add Beschreibung header in next free column after existing data
# Find AG = column 33
desc_col = 33
ws.cell(1, desc_col, "Beschreibung")

renamed = 0
for row in range(2, ws.max_row + 1):
    old_name = ws.cell(row, 2).value
    if old_name and old_name in EXCEL_MAPPING:
        new_title, desc = EXCEL_MAPPING[old_name]
        ws.cell(row, 2, new_title)
        ws.cell(row, desc_col, desc)
        print(f"  Excel row {row}: '{old_name}' -> '{new_title}'")
        renamed += 1

print(f"  Renamed {renamed} properties in Objekte sheet")

# Rangliste sheet has formulas referencing Objekte!B, so it auto-updates
# But let's check if there are hardcoded values
ws_rank = wb['Rangliste']
for row in range(3, 20):
    val = ws_rank.cell(row, 2).value
    if val and isinstance(val, str) and not val.startswith('='):
        if val in EXCEL_MAPPING:
            ws_rank.cell(row, 2, EXCEL_MAPPING[val][0])
            print(f"  Rangliste row {row}: hardcoded '{val}' -> updated")

wb.save(_os.path.join(_ROOT, 'data', 'mallorca-objekte.xlsx'))
print("  Excel saved.\n")

# ---- Step 2: Update HTML ----
print("=== Updating HTML ===")

# Read the full file (it's big but we need to do replacements)
with open(_os.path.join(_ROOT, 'html', 'mallorca-ranking-v5.html'), 'r', encoding='utf-8') as f:
    html = f.read()

print(f"  HTML size: {len(html)} bytes")

# Replace names in JS data and add desc field
for old_name, (new_title, desc) in HTML_MAPPING.items():
    # Replace "name":"OLD" with "name":"NEW","desc":"DESC"
    old_pattern = f'"name":"{old_name}"'
    new_pattern = f'"name":"{new_title}","desc":"{desc}"'
    if old_pattern in html:
        html = html.replace(old_pattern, new_pattern)
        print(f"  Replaced: {old_name}")
    else:
        print(f"  WARNING: Not found in HTML: {old_name}")

# Now add the description display in the card template
# Look for where the name is rendered and add desc below it
# We need to find the render function - let's search for a pattern
# Typically something like: obj.name or item.name in a template

# Let's find the card rendering pattern
if 'obj.name' in html or 'p.name' in html or 'item.name' in html or '.name' in html:
    # Find where name is displayed in the card
    # Common patterns: ${p.name} or similar
    # Let's check what variable name is used
    import re
    name_refs = re.findall(r'\$\{(\w+)\.name\}', html)
    if name_refs:
        var = name_refs[0]
        print(f"  Found template variable: {var}.name")
        
        # Add desc display after name display
        # Find the name display and add desc after its container
        # Look for the pattern like: ${p.name}</...>
        # Add: <div style="font-style:italic;font-size:0.85em;color:#666;margin-top:2px">${p.desc || ''}</div>
        
        desc_html = f'<div style="font-style:italic;font-size:0.85em;color:#888;margin-top:4px">${{{var}.desc || ""}}</div>'
        
        # Find where name appears in template and add desc after the closing tag
        # Pattern: ${var.name}</something>
        pattern = re.compile(rf'(\$\{{{var}\.name\}})(</[^>]+>)')
        match = pattern.search(html)
        if match:
            old_text = match.group(0)
            new_text = match.group(1) + match.group(2) + desc_html
            html = html.replace(old_text, new_text, 1)
            print(f"  Added desc display after name")
        else:
            print(f"  Could not find name template pattern to add desc display")
    else:
        print("  Could not find template variable for name")

with open(_os.path.join(_ROOT, 'html', 'mallorca-ranking-v5.html'), 'w', encoding='utf-8') as f:
    f.write(html)

print(f"  HTML saved. New size: {len(html)} bytes")
print("\nDone!")
