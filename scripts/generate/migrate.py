#!/usr/bin/env python3
"""
Architektur-Migration: Mallorca_Markt_Gesamt → mallorca-objekte-v2
"""

from openpyxl import load_workbook
import requests as req
import time

# === Schritt 1: Master laden ===
print("=== Schritt 1: Master laden ===")
wb_master = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/data/mallorca-objekte-v2.xlsx')
ws_master = wb_master['Objekte']

existing_urls = set()
max_ordnung = 0
for row in ws_master.iter_rows(min_row=2, values_only=True):
    if row[2]:  # URL
        existing_urls.add(str(row[2]).strip())
    if row[0] and isinstance(row[0], (int, float)):
        max_ordnung = max(max_ordnung, int(row[0]))

print(f"Bestehende Objekte: {ws_master.max_row - 1}")
print(f"Höchste Ordnungsnummer: {max_ordnung}")
print(f"Bekannte URLs: {len(existing_urls)}")

# === Schritt 2: Neue Objekte aus Raw Data laden ===
print("\n=== Schritt 2: Raw Data laden ===")
wb_raw = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/data/Mallorca_Markt_Gesamt.xlsx')
ws_raw = wb_raw['Mallorca Objekte']

new_objects = []
skipped_no_url = 0
skipped_duplicate = 0

for row in ws_raw.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    
    # Spalten: Titel | Quelle | URL | Preis (€) | Zimmer | Grundstück (m²) | Wohnfläche (m²) | Ort / Gemeinde | Gefunden am | Status
    titel = row[0] if len(row) > 0 else None
    quelle = row[1] if len(row) > 1 else None
    url = row[2] if len(row) > 2 else None
    preis = row[3] if len(row) > 3 else None
    zimmer = row[4] if len(row) > 4 else None
    grundstueck = row[5] if len(row) > 5 else None
    wohnflaeche = row[6] if len(row) > 6 else None
    ort = row[7] if len(row) > 7 else None
    datum = row[8] if len(row) > 8 else None
    status = row[9] if len(row) > 9 else None
    
    url_clean = str(url).strip() if url else '—'
    
    if url_clean == '—' or url_clean == 'None':
        skipped_no_url += 1
        continue
    
    if url_clean in existing_urls:
        skipped_duplicate += 1
        continue
    
    new_objects.append({
        'name': titel or '',
        'url': url_clean,
        'preis': preis,
        'zimmer': zimmer,
        'grundstueck': grundstueck,
        'wohnflaeche': wohnflaeche,
        'ort': ort or '',
        'quelle': quelle or ''
    })

print(f"Neue Objekte zum Hinzufügen: {len(new_objects)}")
print(f"Übersprungen (kein URL): {skipped_no_url}")
print(f"Übersprungen (Duplikat): {skipped_duplicate}")

# === Schritt 3: Berechnbare Felder ===
def calc_eur_per_m2_bebaut(preis, wohnflaeche):
    try:
        if preis and wohnflaeche and float(wohnflaeche) > 0:
            return round(float(preis) / float(wohnflaeche))
    except: pass
    return None

def calc_eur_per_m2_grundstueck(preis, grundstueck):
    try:
        if preis and grundstueck and float(grundstueck) > 0:
            return round(float(preis) / float(grundstueck))
    except: pass
    return None

# === Schritt 4: Google Maps Distanzen ===
def get_distances(origin_address):
    destinations = [
        ('Flughafen', '39.5517,2.7388'),
        ('Daia', '39.3558,3.0597'),
        ('Andratx', '39.5731,2.3987'),
        ('Ses Salines', '39.3497,3.0469'),
    ]
    
    results = {}
    for name, dest in destinations:
        try:
            url = 'https://maps.googleapis.com/maps/api/distancematrix/json'
            params = {
                'origins': f'{origin_address}, Mallorca, Spain',
                'destinations': dest,
                'mode': 'driving',
                'key': 'AIzaSyAzplaFOGUrCygk3mrzS--wrHNIz0arKok'
            }
            r = req.get(url, params=params, timeout=10)
            data = r.json()
            if data['status'] == 'OK' and data['rows'][0]['elements'][0]['status'] == 'OK':
                el = data['rows'][0]['elements'][0]
                results[name] = {
                    'km': round(el['distance']['value'] / 1000, 1),
                    'min': round(el['duration']['value'] / 60)
                }
        except Exception as e:
            print(f"  Distance error for {name}: {e}")
        time.sleep(0.1)
    
    return results

# === Schritt 5: In Master-Excel schreiben ===
print("\n=== Schritt 5: Schreibe in Master ===")
ordnung = max_ordnung + 1
added = 0
dist_calculated = 0

# Für dedup: Ort-basierte Distanz-Cache (gleicher Ort → gleiche Distanzen)
dist_cache = {}

for obj in new_objects:
    ort = obj['ort']
    preis = obj['preis']
    wohnflaeche = obj['wohnflaeche']
    grundstueck = obj['grundstueck']
    zimmer = obj['zimmer']
    
    # Distanzen nur für qualifizierte Objekte (max 500)
    dist = {}
    try:
        p = float(preis) if preis else 0
        z = float(zimmer) if zimmer else 0
        if ort and (p >= 1000000 or z >= 5) and dist_calculated < 500:
            if ort in dist_cache:
                dist = dist_cache[ort]
            else:
                dist = get_distances(ort)
                dist_cache[ort] = dist
                dist_calculated += 1
                print(f"  Distanz für '{ort}': {dist}")
    except Exception as e:
        print(f"  Distanz-Fehler: {e}")
    
    eur_bebaut = calc_eur_per_m2_bebaut(preis, wohnflaeche)
    eur_grund = calc_eur_per_m2_grundstueck(preis, grundstueck)
    
    row = [
        ordnung,                                    # Ordnungsnummer
        obj['name'],                                # Name
        obj['url'],                                 # Link Objekt (URL)
        None,                                       # Charme/Ästhetik (manuell)
        zimmer,                                     # Zimmer
        None,                                       # Bäder (manuell)
        grundstueck,                                # Grundstücksgröße (m²)
        wohnflaeche,                                # Bebaute Fläche (m²)
        None,                                       # Garten zu bewirtschaften (manuell)
        ort,                                        # Location
        dist.get('Flughafen', {}).get('km'),        # Entfernung Flughafen (km)
        dist.get('Flughafen', {}).get('min'),       # Entfernung Flughafen (min)
        dist.get('Daia', {}).get('km'),             # Entfernung Daia (km)
        dist.get('Daia', {}).get('min'),            # Entfernung Daia (min)
        dist.get('Andratx', {}).get('km'),          # Entfernung Andratx (km)
        dist.get('Andratx', {}).get('min'),         # Entfernung Andratx (min)
        dist.get('Ses Salines', {}).get('km'),      # Entfernung Ses Salines (km)
        dist.get('Ses Salines', {}).get('min'),     # Entfernung Ses Salines (min)
        preis,                                      # Preis (€)
        eur_bebaut,                                 # €/m² (bebaut)
        eur_grund,                                  # €/m² (Grundstück)
        None,                                       # Renovierung (manuell)
        None,                                       # Bewirtschaftung (manuell)
        None,                                       # Vermietlizenz (manuell)
        None,                                       # Erreichbarkeit (manuell)
        None,                                       # Score (wird per Formel berechnet)
        None,                                       # SortKey
        None,                                       # Rang
        None,                                       # Gebäudestruktur
        None,                                       # Baujahr
        None,                                       # Letzte Renovierung
        None,                                       # Reno-Score
        None,                                       # Reno-Begründung
        f"Quelle: {obj['quelle']}",                 # Kommentar (Quelle vermerken)
        None,                                       # Beschreibung
    ]
    
    ws_master.append(row)
    existing_urls.add(obj['url'])
    ordnung += 1
    added += 1
    
    if added % 100 == 0:
        wb_master.save('/Users/robin/.openclaw/workspace/mallorca-projekt/data/mallorca-objekte-v2.xlsx')
        print(f"  Zwischengespeichert: {added} Objekte")

wb_master.save('/Users/robin/.openclaw/workspace/mallorca-projekt/data/mallorca-objekte-v2.xlsx')

total_rows = ws_master.max_row - 1
print(f"\n✅ FERTIG!")
print(f"  Neu hinzugefügt: {added} Objekte")
print(f"  Distanzen berechnet (unique Orte): {dist_calculated}")
print(f"  Gesamt in mallorca-objekte-v2.xlsx: {total_rows} Objekte")
