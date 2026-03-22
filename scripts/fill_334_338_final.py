#!/usr/bin/env python3
"""Fill all missing fields for 334-338 - final version with pre-analyzed vision data."""

import openpyxl

XLSX = '/Users/robin/.openclaw/workspace/mallorca-projekt/data/mallorca-kandidaten-v2.xlsx'

# Pre-fetched Google Maps distances (from API call)
DISTANCES = {
    334: {"flughafen": (34.0, 31), "daia": (44.9, 52), "andratx": (58.5, 48), "ses_salines": (47.4, 53)},
    # Need to fetch 335-338 - will compute below
}

# Pre-analyzed vision data
VISION = {
    334: {"charme": 4, "renovierung": 96, "reno_begruendung": "Neubau von 2022 mit hochwertiger Ausstattung, modern und neuwertig.", "gaestehaueser": 0},
    335: {"charme": 5, "renovierung": 65, "reno_begruendung": "300 Jahre alte Natursteinfassade mit authentischer Patina, gepflegt aber sichtbare Altersspuren deuten auf weiteren Modernisierungsbedarf hin.", "gaestehaueser": 2},
    336: {"charme": 4, "renovierung": 70, "reno_begruendung": "Traditionelle Finca in malerischer Weinberg-Lage mit Pool, solider Zustand aber Fassade zeigt leichten Renovierungsbedarf.", "gaestehaueser": 1},
    337: {"charme": 5, "renovierung": 95, "reno_begruendung": "2018 erbaut im gehobenen mallorquinischen Stil mit makelloser Natursteinfassade – praktisch neuwertig.", "gaestehaueser": 1},
    338: {"charme": 3, "renovierung": 95, "reno_begruendung": "Moderner Neubau von 2019 mit klaren Linien und neuwertigem Zustand, jedoch fehlt der typische Finca-Charakter.", "gaestehaueser": 0},
}

# Object data
OBJECTS = {
    334: {"name": "Sencelles", "lat": 39.6497, "lng": 2.9008, "grundstueck": 14483, "bebaut": 414, "zimmer": 5, "baeder": 4, "preis": 4250000, "baujahr": 2021},
    335: {"name": "Santanyí",  "lat": 39.3554, "lng": 3.1243, "grundstueck": 17700, "bebaut": 792, "zimmer": 9, "baeder": 5, "preis": 3500000, "baujahr": 1800},
    336: {"name": "Felanitx",  "lat": 39.4688, "lng": 3.1483, "grundstueck": 19533, "bebaut": 336, "zimmer": 6, "baeder": 3, "preis": 2860000, "baujahr": 2000},
    337: {"name": "Sant Llorenç", "lat": 39.6200, "lng": 3.2833, "grundstueck": 16239, "bebaut": 330, "zimmer": 6, "baeder": 4, "preis": 4650000, "baujahr": 2017},
    338: {"name": "Montuïri",  "lat": 39.5667, "lng": 3.0000, "grundstueck": 15033, "bebaut": 473, "zimmer": 5, "baeder": 4, "preis": 2890000, "baujahr": 2019},
}

import requests

API_KEY = "AIzaSyAzplaFOGUrCygk3mrzS--wrHNIz0arKok"
REFS = {
    "flughafen": "39.5517,2.7388",
    "daia":      "39.7456,2.6489",
    "andratx":   "39.5747,2.3818",
    "ses_salines":"39.3444,3.0503",
}

def get_distances(lat, lng):
    origin = f"{lat},{lng}"
    dest_str = "|".join(REFS.values())
    url = "https://maps.googleapis.com/maps/api/distancematrix/json"
    params = {"origins": origin, "destinations": dest_str, "mode": "driving", "key": API_KEY}
    r = requests.get(url, params=params, timeout=15)
    data = r.json()
    result = {}
    row_data = data["rows"][0]["elements"]
    for i, key in enumerate(REFS.keys()):
        el = row_data[i]
        if el["status"] == "OK":
            km = round(el["distance"]["value"] / 1000, 1)
            mins = round(el["duration"]["value"] / 60)
            result[key] = (km, mins)
        else:
            result[key] = (None, None)
    return result

def calc_erreichbarkeit(flughafen_min, daia_min, ses_min, andratx_min):
    def score_for(mins, ideal, akzeptabel, dealbreaker):
        if mins is None: return 50
        if mins <= ideal: return 100
        elif mins <= akzeptabel:
            return round(100 - 40 * (mins - ideal) / (akzeptabel - ideal))
        elif mins <= dealbreaker:
            return round(60 - 60 * (mins - akzeptabel) / (dealbreaker - akzeptabel))
        else: return 0
    
    s_flughafen = score_for(flughafen_min, 15, 25, 40)
    s_daia = score_for(daia_min, 20, 40, 70)
    s_ses = score_for(ses_min, 15, 30, 45)
    s_andratx = score_for(andratx_min, 25, 40, 60)
    return round(s_flughafen * 0.30 + s_daia * 0.30 + s_ses * 0.30 + s_andratx * 0.10)

def calc_score(obj, charme, renovierung, bewirtschaftung, vermietlizenz, erreichbarkeit, gaestehaueser):
    zimmer = obj["zimmer"]
    grundstueck = obj["grundstueck"]
    bebaut = obj["bebaut"]
    preis = obj["preis"]
    
    # Zimmer & Platz (20%): ideal=8, min=5
    zimmer_score = min(100, max(0, (zimmer - 5) / (8 - 5) * 100))
    
    # Preis-Leistung €/m² bebaut (15%): ~5000=gut, ~15000=schlecht
    eur_m2 = preis / bebaut
    preis_score = max(0, min(100, 100 - (eur_m2 - 5000) / 150))
    
    # Gästehaus (15%): 0=0, 1=50, 2=100
    gast_score = (gaestehaueser or 0) * 50
    
    # Erreichbarkeit (15%): already 0-100
    erreich_score = erreichbarkeit or 0
    
    # Grundstück/Garten (10%)
    garten = grundstueck - bebaut
    garten_score = min(100, max(0, (garten - 1000) / (5000 - 1000) * 100))
    
    # Vermietlizenz (5%): 0/50/100
    vmiet_score = vermietlizenz or 0
    
    # Bewirtschaftung (5%): 1-5 → 0-100
    bewirt_score = ((bewirtschaftung or 3) - 1) / 4 * 100
    
    # Charme (10%): 1-5 → 0-100
    charme_score = ((charme or 3) - 1) / 4 * 100
    
    # Renovierung (5%): 0-100
    reno_score = renovierung or 60
    
    score = (
        zimmer_score * 0.20 +
        preis_score * 0.15 +
        gast_score * 0.15 +
        erreich_score * 0.15 +
        garten_score * 0.10 +
        vmiet_score * 0.05 +
        bewirt_score * 0.05 +
        charme_score * 0.10 +
        reno_score * 0.05
    )
    return round(score, 1)

# Load workbook
wb = openpyxl.load_workbook(XLSX)
ws = wb.active

# Find rows
row_map = {}
for row in ws.iter_rows(min_row=2):
    nr = row[0].value
    if nr in [334, 335, 336, 337, 338]:
        row_map[nr] = row[0].row

print(f"Row map: {row_map}")

all_distances = {}
for nr, obj in OBJECTS.items():
    print(f"Fetching distances for {nr} ({obj['name']})...")
    all_distances[nr] = get_distances(obj["lat"], obj["lng"])
    print(f"  -> {all_distances[nr]}")

print("\nFilling Excel...")

for nr, obj in OBJECTS.items():
    row_idx = row_map[nr]
    dists = all_distances[nr]
    vision = VISION[nr]
    
    flughafen_km, flughafen_min = dists["flughafen"]
    daia_km, daia_min = dists["daia"]
    andratx_km, andratx_min = dists["andratx"]
    ses_km, ses_min = dists["ses_salines"]
    
    # Distances
    ws.cell(row_idx, 11).value = flughafen_km
    ws.cell(row_idx, 12).value = flughafen_min
    ws.cell(row_idx, 13).value = daia_km
    ws.cell(row_idx, 14).value = daia_min
    ws.cell(row_idx, 15).value = andratx_km
    ws.cell(row_idx, 16).value = andratx_min
    ws.cell(row_idx, 17).value = ses_km
    ws.cell(row_idx, 18).value = ses_min
    
    # Garten
    garten = obj["grundstueck"] - obj["bebaut"]
    ws.cell(row_idx, 9).value = garten
    
    # €/m²
    ws.cell(row_idx, 20).value = round(obj["preis"] / obj["bebaut"])
    ws.cell(row_idx, 21).value = round(obj["preis"] / obj["grundstueck"])
    
    # Vision
    charme = vision["charme"]
    renovierung = vision["renovierung"]
    gaestehaueser = vision["gaestehaueser"]
    ws.cell(row_idx, 4).value = charme
    ws.cell(row_idx, 22).value = renovierung
    ws.cell(row_idx, 32).value = renovierung
    ws.cell(row_idx, 33).value = vision["reno_begruendung"]
    ws.cell(row_idx, 40).value = gaestehaueser
    
    # Bewirtschaftung (1-5): based on plot size + features
    grundstueck = obj["grundstueck"]
    if grundstueck < 20000:
        bewirtschaftung = 4
    elif grundstueck < 50000:
        bewirtschaftung = 3
    else:
        bewirtschaftung = 2
    # Weinberg = harder
    if nr == 336:
        bewirtschaftung = max(1, bewirtschaftung - 1)  # Weinberg
    ws.cell(row_idx, 23).value = bewirtschaftung
    
    # Vermietlizenz (0/50/100): default 0 unless mentioned
    # None of these descriptions mention ETV license explicitly
    vermietlizenz = 0
    ws.cell(row_idx, 24).value = vermietlizenz
    
    # Erreichbarkeit
    erreichbarkeit = calc_erreichbarkeit(flughafen_min, daia_min, ses_min, andratx_min)
    ws.cell(row_idx, 25).value = erreichbarkeit
    
    # Score
    score = calc_score(obj, charme, renovierung, bewirtschaftung, vermietlizenz, erreichbarkeit, gaestehaueser)
    ws.cell(row_idx, 26).value = score
    ws.cell(row_idx, 27).value = score  # SortKey
    
    print(f"Nr.{nr} ({obj['name']}): Flughafen={flughafen_min}min, Daia={daia_min}min, "
          f"Andratx={andratx_min}min, SesSal={ses_min}min | "
          f"Garten={garten}m² | Charme={charme}, Reno={renovierung}, Gästehaus={gaestehaueser} | "
          f"Bewirt={bewirtschaftung}, VermietLiz={vermietlizenz}, Erreich={erreichbarkeit} | Score={score}")

wb.save(XLSX)
print("\n✅ Saved!")
