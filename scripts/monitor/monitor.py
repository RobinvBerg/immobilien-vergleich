#!/usr/bin/env python3
"""Mallorca Market Monitor — Automatische Neuanlage-Erkennung"""

import requests, json, re, time, argparse, logging, sys
from bs4 import BeautifulSoup
from datetime import datetime
from pathlib import Path
import openpyxl
from urllib.parse import urljoin, urlparse
import subprocess

_SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = _SCRIPT_DIR.parent.parent
BASE = PROJECT_ROOT  # kept for compatibility
GESAMT_FILE  = PROJECT_ROOT / 'data'  / 'Mallorca_Markt_Gesamt.xlsx'
STATE_FILE   = PROJECT_ROOT / 'debug' / 'monitor_state.json'
RESULTS_FILE = PROJECT_ROOT / 'debug' / 'monitor_results.json'
LOG_FILE     = PROJECT_ROOT / 'debug' / 'monitor.log'

logging.basicConfig(level=logging.INFO, format='%(asctime)s | %(levelname)s | %(message)s',
                   handlers=[logging.FileHandler(LOG_FILE), logging.StreamHandler()])
LOG = logging.getLogger(__name__)

# ─── Credentials ─────────────────────────────────────────
PROXY = "http://sp1e6lma32:pxjc5K6_LBg3Is6vzo@gate.decodo.com:10001"
UNBLOCK = "http://U0000364062:PW_1047072161848b0d67b68ff1b160986e6@unblock.decodo.com:60000"
APIFY_TOKEN = "apify_api_feD2KhARHjtuV9CrSwOReYgoePFSF44nsDL6"
HEADERS = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121'}

# ─── Load existing URLs from Gesamtliste ────────────────
def load_existing_urls():
    """Return set of normalized URLs from Mallorca_Markt_Gesamt.xlsx"""
    try:
        wb = openpyxl.load_workbook(GESAMT_FILE, data_only=True)
        ws = wb.active
        urls = set()
        for r in range(2, ws.max_row + 1):
            url = ws.cell(r, 3).value  # Spalte C = URL
            if url:
                urls.add(normalize_url(str(url)))
        LOG.info(f"✓ Loaded {len(urls)} existing URLs from Gesamtliste")
        return urls
    except Exception as e:
        LOG.error(f"✗ Error loading Gesamtliste: {e}")
        return set()

def normalize_url(url):
    """Normalize URL for comparison"""
    url = str(url).strip().lower().rstrip('/')
    return url

def is_new(url, existing_urls):
    """Check if URL is new"""
    norm = normalize_url(url)
    return norm not in existing_urls

# ─── Load/Save State ─────────────────────────────────────
def load_state():
    """Load monitor_state.json — trackt letzte URLs pro Quelle"""
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text())
        except:
            return {}
    return {}

def save_state(state, dry_run=False):
    """Save state — verhindert Re-Reporting"""
    if not dry_run:
        STATE_FILE.write_text(json.dumps(state, indent=2))

# ─── Filter Criteria ─────────────────────────────────────
def passes_filter(item):
    """Check if item meets criteria: 5+ rooms, 2.8M-20M budget"""
    zimmer = item.get('zimmer')
    preis = item.get('preis')
    
    if zimmer and zimmer < 5:
        return False
    if preis:
        if preis < 2_800_000 or preis > 20_000_000:
            return False
    return True

# ─── Sources ─────────────────────────────────────────────
def scrape_balearic_properties(existing_urls, state):
    """Balearic Properties"""
    LOG.info("[1] Balearic Properties...")
    new = []
    s = requests.Session(); s.headers.update(HEADERS)
    try:
        last_urls = set(state.get('Balearic Properties', []))
        for pg in range(1, 100):
            url = f"https://www.balearic-properties.com/property-for-sale/mallorca.html?page={pg}"
            r = s.get(url, timeout=15)
            if r.status_code != 200: break
            soup = BeautifulSoup(r.text, 'html.parser')
            containers = soup.select('article') or soup.select('[class*="property"]')
            if not containers: break
            
            found_new = False
            for c in containers:
                a = c.find('a', href=lambda h: h and '/property/id/' in h)
                if not a: continue
                href = a['href']
                full_url = href if href.startswith('http') else f'https://www.balearic-properties.com{href}'
                norm_url = normalize_url(full_url)
                
                if norm_url in last_urls or not is_new(full_url, existing_urls):
                    continue
                
                # Parse details
                beds_el = c.find(text=re.compile(r'(\d+)\s*[Bb]ed'))
                zimmer = 0
                if beds_el:
                    m = re.search(r'(\d+)', str(beds_el))
                    if m: zimmer = int(m.group(1))
                
                title_el = c.find(['h2','h3','h4'])
                title = title_el.get_text(strip=True) if title_el else ''
                
                if passes_filter({'zimmer': zimmer}):
                    new.append({'url': full_url, 'source': 'Balearic Properties', 'title': title, 'zimmer': zimmer})
                    last_urls.add(norm_url)
                    found_new = True
            
            if not found_new and pg > 2: break
            time.sleep(0.3)
        
        state['Balearic Properties'] = list(last_urls)
        LOG.info(f"  → {len(new)} neu")
    except Exception as e:
        LOG.error(f"  ERROR: {e}")
    return new, state

def scrape_living_blue(existing_urls, state):
    """Living Blue Mallorca"""
    LOG.info("[2] Living Blue...")
    new = []
    try:
        s = requests.Session(); s.headers.update(HEADERS)
        last_urls = set(state.get('Living Blue', []))
        
        for pg in range(1, 50):
            url = f"https://www.livingblue-mallorca.com/de-de/immobilien?pag={pg}"
            r = s.get(url, timeout=15)
            if r.status_code != 200: break
            soup = BeautifulSoup(r.text, 'html.parser')
            links = soup.find_all('a', href=re.compile(r'/de-de/immobilien/[^/]+$'))
            if not links and pg > 1: break
            
            for a in links:
                href = a['href']
                full_url = f'https://www.livingblue-mallorca.com{href}' if href.startswith('/') else href
                norm_url = normalize_url(full_url)
                
                if norm_url in last_urls or not is_new(full_url, existing_urls):
                    continue
                
                title = a.get_text(strip=True) or ''
                new.append({'url': full_url, 'source': 'Living Blue', 'title': title})
                last_urls.add(norm_url)
            
            time.sleep(0.3)
        
        state['Living Blue'] = list(last_urls)
        LOG.info(f"  → {len(new)} neu")
    except Exception as e:
        LOG.error(f"  ERROR: {e}")
    return new, state

def scrape_engelvoelkers(existing_urls, state):
    """Engel & Völkers mit Proxy"""
    LOG.info("[3] Engel & Völkers...")
    new = []
    try:
        s = requests.Session(); s.proxies = {'http': PROXY, 'https': PROXY}; s.headers.update(HEADERS)
        last_urls = set(state.get('Engel & Völkers', []))
        
        for start in range(0, 600, 24):
            url = f"https://www.engelvoelkers.com/de/suche/?facets=rgn%3Amallorca%3B&_boolFilters=buy%3A&startIndex={start}&pageSize=24"
            r = s.get(url, timeout=15)
            if r.status_code != 200: break
            soup = BeautifulSoup(r.text, 'html.parser')
            links = soup.find_all('a', href=re.compile(r'expose'))
            if not links and start > 0: break
            
            for a in links:
                href = a['href']
                full_url = href if href.startswith('http') else f'https://www.engelvoelkers.com{href}'
                norm_url = normalize_url(full_url)
                
                if norm_url in last_urls or not is_new(full_url, existing_urls):
                    continue
                
                title = a.get_text(strip=True)
                new.append({'url': full_url, 'source': 'Engel & Völkers', 'title': title})
                last_urls.add(norm_url)
            
            time.sleep(0.4)
        
        state['Engel & Völkers'] = list(last_urls)
        LOG.info(f"  → {len(new)} neu")
    except Exception as e:
        LOG.error(f"  ERROR: {e}")
    return new, state

def scrape_kyero(existing_urls, state):
    """Kyero"""
    LOG.info("[4] Kyero...")
    new = []
    try:
        s = requests.Session(); s.headers.update(HEADERS)
        last_urls = set(state.get('Kyero', []))
        
        for pg in range(1, 50):
            url = f"https://www.kyero.com/en/property-for-sale/mallorca?min_beds=5&page={pg}"
            r = s.get(url, timeout=15)
            if r.status_code != 200: break
            soup = BeautifulSoup(r.text, 'html.parser')
            links = soup.find_all('a', href=re.compile(r'/property/.+mallorca'))
            if not links and pg > 1: break
            
            for a in links:
                href = a['href']
                full_url = urljoin('https://www.kyero.com', href)
                norm_url = normalize_url(full_url)
                
                if norm_url in last_urls or not is_new(full_url, existing_urls):
                    continue
                
                title = a.get_text(strip=True)
                new.append({'url': full_url, 'source': 'Kyero', 'title': title})
                last_urls.add(norm_url)
            
            time.sleep(0.3)
        
        state['Kyero'] = list(last_urls)
        LOG.info(f"  → {len(new)} neu")
    except Exception as e:
        LOG.error(f"  ERROR: {e}")
    return new, state

def scrape_luxury_estates(existing_urls, state):
    """Luxury Estates Mallorca"""
    LOG.info("[5] Luxury Estates...")
    new = []
    try:
        s = requests.Session(); s.headers.update(HEADERS)
        last_urls = set(state.get('Luxury Estates', []))
        
        for pg in range(1, 30):
            url = f"https://www.luxury-estates-mallorca.com/en/properties?page={pg}"
            r = s.get(url, timeout=15)
            if r.status_code != 200: break
            soup = BeautifulSoup(r.text, 'html.parser')
            links = soup.find_all('a', href=re.compile(r'property|offer'))
            if not links and pg > 1: break
            
            for a in links:
                href = a.get('href','')
                if not href or '#' in href: continue
                full_url = urljoin('https://www.luxury-estates-mallorca.com', href)
                norm_url = normalize_url(full_url)
                
                if norm_url in last_urls or not is_new(full_url, existing_urls):
                    continue
                
                title = a.get_text(strip=True)
                new.append({'url': full_url, 'source': 'Luxury Estates', 'title': title})
                last_urls.add(norm_url)
            
            time.sleep(0.3)
        
        state['Luxury Estates'] = list(last_urls)
        LOG.info(f"  → {len(new)} neu")
    except Exception as e:
        LOG.error(f"  ERROR: {e}")
    return new, state

def scrape_idealista(existing_urls, state):
    """Idealista via Apify"""
    LOG.info("[6] Idealista (Apify)...")
    new = []
    try:
        last_urls = set(state.get('Idealista', []))
        inp = {"locationName":"Illes Balears","country":"es","propertyType":"homes","operation":"sale","minRooms":5,"maxPages":5}
        r = requests.post(f"https://api.apify.com/v2/acts/memo23~idealista-scraper/runs?token={APIFY_TOKEN}", json=inp, timeout=30)
        
        if r.status_code in (200, 201):
            run_id = r.json()['data']['id']
            for _ in range(24):
                time.sleep(5)
                sr = requests.get(f"https://api.apify.com/v2/acts/memo23~idealista-scraper/runs/{run_id}?token={APIFY_TOKEN}")
                st = sr.json()['data']['status']
                
                if st == 'SUCCEEDED':
                    did = sr.json()['data']['defaultDatasetId']
                    items = requests.get(f"https://api.apify.com/v2/datasets/{did}/items?token={APIFY_TOKEN}&limit=500").json()
                    for item in items:
                        url = item.get('url','') or item.get('propertyUrl','')
                        if not url: continue
                        norm_url = normalize_url(url)
                        
                        if norm_url in last_urls or not is_new(url, existing_urls):
                            continue
                        
                        new.append({
                            'url': url, 'source': 'Idealista',
                            'title': item.get('title',''),
                            'zimmer': item.get('rooms'),
                            'flaeche': item.get('size'),
                            'preis': item.get('price')
                        })
                        last_urls.add(norm_url)
                    break
                
                if st in ('FAILED','ABORTED'):
                    break
        
        state['Idealista'] = list(last_urls)
        LOG.info(f"  → {len(new)} neu")
    except Exception as e:
        LOG.error(f"  ERROR: {e}")
    return new, state

# ─── Main ────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true', help='Nicht speichern')
    parser.add_argument('--notify', choices=['telegram', 'none'], default='none', help='Benachrichtigung')
    parser.add_argument('--update-excel', action='store_true', help='An Gesamtliste anhängen')
    args = parser.parse_args()
    
    LOG.info("="*60)
    LOG.info("START — Mallorca Market Monitor")
    
    existing_urls = load_existing_urls()
    state = load_state()
    
    all_new = []
    
    # Run scrapers
    new, state = scrape_balearic_properties(existing_urls, state)
    all_new.extend(new)
    
    new, state = scrape_living_blue(existing_urls, state)
    all_new.extend(new)
    
    new, state = scrape_engelvoelkers(existing_urls, state)
    all_new.extend(new)
    
    new, state = scrape_kyero(existing_urls, state)
    all_new.extend(new)
    
    new, state = scrape_luxury_estates(existing_urls, state)
    all_new.extend(new)
    
    new, state = scrape_idealista(existing_urls, state)
    all_new.extend(new)
    
    # Save results
    LOG.info(f"\n{'='*60}")
    LOG.info(f"GESAMT: {len(all_new)} neue Objekte")
    
    if all_new:
        RESULTS_FILE.write_text(json.dumps(all_new, indent=2, ensure_ascii=False))
        LOG.info(f"Saved: {RESULTS_FILE}")
        
        # Telegram notification
        if args.notify == 'telegram' and not args.dry_run:
            msg = f"🏝️ Mallorca Monitor — {len(all_new)} neue Objekte\n\n"
            by_source = {}
            for item in all_new:
                src = item.get('source')
                by_source[src] = by_source.get(src,0) + 1
            for src, cnt in sorted(by_source.items()):
                msg += f"• {src}: {cnt}\n"
            msg += f"\nDetails: {RESULTS_FILE}"
            try:
                subprocess.run(['openclaw', 'message', 'send', '--accountId', 'zweiter-bot', 
                               '--target', '803179451', '--message', msg], check=False)
                LOG.info("✓ Telegram notification sent")
            except Exception as e:
                LOG.error(f"✗ Telegram error: {e}")
    
    # Save state
    save_state(state, dry_run=args.dry_run)
    
    LOG.info(f"END — {'DRY-RUN' if args.dry_run else 'LIVE'}")
    LOG.info("="*60)

if __name__ == '__main__':
    main()
