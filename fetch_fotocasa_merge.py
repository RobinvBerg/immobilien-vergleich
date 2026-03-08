#!/usr/bin/env python3
"""
fetch_fotocasa_merge.py
Merges all fetchdetails_progress_0..4.json into Excel column Grundstück (m²).
Run after all workers finish.
"""

import json
from pathlib import Path
import openpyxl

EXCEL_PATH = Path("/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx")
BASE_PROG  = Path("/Users/robin/.openclaw/workspace/mallorca-projekt")

print("Loading Excel …")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

url_col = plot_col = None
for cell in ws[1]:
    v = str(cell.value).lower() if cell.value else ''
    if 'url' in v:     url_col  = cell.column
    if 'grundst' in v or 'plot' in v or 'terreno' in v:
        plot_col = cell.column

print(f"URL col: {url_col}, Plot col: {plot_col}")

# Merge all progress files
merged = {}
for i in range(5):
    p = BASE_PROG / f"fetchdetails_progress_{i}.json"
    if p.exists():
        with open(p) as f:
            data = json.load(f)
        merged.update(data)
        plots = sum(1 for v in data.values() if v.get('plot'))
        print(f"  Worker {i}: {len(data)} URLs, {plots} plots found")
    else:
        print(f"  Worker {i}: file not found — skipped")

total_plots = sum(1 for v in merged.values() if v.get('plot'))
print(f"\nMerged total: {len(merged)} URLs, {total_plots} with plot")

# Write to Excel
updated = 0
for row in ws.iter_rows(min_row=2):
    url = row[url_col - 1].value
    if url and url in merged:
        plot = merged[url].get('plot')
        if plot and plot_col:
            row[plot_col - 1].value = plot
            updated += 1

wb.save(EXCEL_PATH)
print(f"Excel updated: {updated} rows written. Saved.")

# Save merged progress
merged_path = BASE_PROG / "fetchdetails_progress.json"
with open(merged_path, 'w') as f:
    json.dump(merged, f, indent=2)
print(f"Merged progress saved: {merged_path}")

# Pool output
print("\n=== POOL (URLs with plot) ===")
pool = [(v['plot'], url) for url, v in merged.items() if v.get('plot') and 'fotocasa' in url]
pool.sort(reverse=True)
for plot, url in pool[:50]:
    print(f"  {plot}m² — {url}")
if len(pool) > 50:
    print(f"  … and {len(pool)-50} more")
