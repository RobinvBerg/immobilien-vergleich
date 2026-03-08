#!/usr/bin/env python3
"""
Scraper für: The Agency RE Mallorca + Pollentia Properties
(Mallorca Finest / mallorcafinest.com ist defunct/SSL-Fehler, kein Immobilien-Anbieter)
"""
import requests
from bs4 import BeautifulSoup
import time, re
from openpyxl import load_workbook
from datetime import date

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.5',
}

EXCEL = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
wb = load_workbook(EXCEL)
ws = wb['Mallorca Objekte']
existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
print(f"Loaded Excel: {len(existing_urls)} existing URLs")

def save_obj(titel, quelle, url, preis=None, zimmer=None, grundstueck=None, flaeche=None, ort=''):
    if url and url in existing_urls: return False
    ws.append([str(titel)[:100], quelle, url or '—', preis, zimmer, grundstueck, flaeche, ort, str(date.today()), 'Neu'])
    if url: existing_urls.add(url)
    return True

def parse_price(text):
    """Extract price from text, return integer or None"""
    text = text.replace(',', '').replace('.', '').replace('\xa0', '').replace(' ', '')
    # Look for numbers next to € sign
    match = re.search(r'€(\d{4,})', text)
    if not match:
        match = re.search(r'(\d{4,})€', text)
    if match:
        try:
            return int(match.group(1))
        except:
            pass
    return None

def parse_int(text):
    """Extract first reasonable integer from text"""
    m = re.search(r'(\d+(?:[.,]\d+)?)', text.replace('\xa0', ''))
    if m:
        try:
            return int(float(m.group(1).replace(',', '.')))
        except:
            pass
    return None

results = {}

# ============================================================
# 1. THE AGENCY RE MALLORCA
# https://theagencyre-mallorca.com/properties/
# ============================================================
print("\n=== The Agency RE Mallorca ===")
count_agency = 0
base_url = 'https://theagencyre-mallorca.com'

try:
    # First page to get total pages
    r = requests.get(f'{base_url}/properties/', headers=headers, timeout=20)
    print(f"Page 1: {r.status_code} | {len(r.text)} chars")
    soup = BeautifulSoup(r.text, 'html.parser')
    
    # Find last page number
    last_page = 1
    pagination = soup.find_all('a', href=re.compile(r'/properties/page/(\d+)/'))
    for a in pagination:
        m = re.search(r'/page/(\d+)/', a['href'])
        if m:
            last_page = max(last_page, int(m.group(1)))
    print(f"Total pages: {last_page}")
    
    def process_agency_page(soup_page):
        count = 0
        # Properties are in containers with a link to /property/
        items = soup_page.find_all('article') or soup_page.find_all('div', class_=re.compile('property|listing', re.I))
        if not items:
            # Fallback: find all property links
            prop_links = list(set([
                a['href'] for a in soup_page.find_all('a', href=True)
                if '/property/' in a['href'] and 'theagencyre-mallorca.com' in a['href']
            ]))
            for link in prop_links:
                titel = 'The Agency RE Mallorca listing'
                if save_obj(titel, 'The Agency RE Mallorca', link, ort='Mallorca'):
                    count += 1
            return count
        
        for item in items:
            a = item.find('a', href=lambda h: h and '/property/' in h)
            if not a: continue
            link = a['href']
            if not link.startswith('http'):
                link = base_url + link
            
            # Title
            h = item.find(['h2', 'h3', 'h4', 'h5'])
            titel = h.get_text(strip=True) if h else 'The Agency RE Mallorca listing'
            
            # Price
            item_text = item.get_text(separator=' ', strip=True)
            preis = parse_price(item_text)
            
            # Zimmer / m²
            zimmer = None
            flaeche = None
            ort_item = 'Mallorca'
            
            # Look for bedroom count (icon pattern: "3" near bed icon)
            bed_m = re.search(r'(\d)\s*(?:bed|zimmer|schlafzimmer)', item_text, re.I)
            if bed_m:
                zimmer = int(bed_m.group(1))
            
            # m²
            sqm_m = re.search(r'(\d+)\s*(?:㎡|m²|sqm|sq\.?m)', item_text, re.I)
            if sqm_m:
                flaeche = int(sqm_m.group(1))
            
            # Location
            loc_el = item.find(class_=re.compile('location|address|city', re.I))
            if loc_el:
                ort_item = loc_el.get_text(strip=True)[:50] or 'Mallorca'
            
            if save_obj(titel, 'The Agency RE Mallorca', link, preis=preis, zimmer=zimmer, flaeche=flaeche, ort=ort_item):
                count += 1
        return count
    
    count_agency += process_agency_page(soup)
    
    # Remaining pages
    for page_num in range(2, last_page + 1):
        time.sleep(0.5)
        url = f'{base_url}/properties/page/{page_num}/'
        try:
            r = requests.get(url, headers=headers, timeout=20)
            if r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')
                n = process_agency_page(soup)
                count_agency += n
                print(f"  Page {page_num}/{last_page}: +{n} (total {count_agency})")
            else:
                print(f"  Page {page_num}: HTTP {r.status_code}")
        except Exception as e:
            print(f"  Page {page_num}: Error {e}")
    
    results['The Agency RE Mallorca'] = count_agency
    print(f"The Agency RE Mallorca: {count_agency} new objects")

except Exception as e:
    print(f"Error: {e}")
    import traceback; traceback.print_exc()
    results['The Agency RE Mallorca'] = count_agency

# ============================================================
# 2. POLLENTIA PROPERTIES
# https://www.pollentiaproperties.com/en/properties?page=X
# ============================================================
print("\n=== Pollentia Properties ===")
count_pollentia = 0

try:
    # First: get total from /en/results
    r = requests.get('https://www.pollentiaproperties.com/en/results', headers=headers, timeout=15)
    soup = BeautifulSoup(r.text, 'html.parser')
    
    # Find last page
    last_page = 1
    for a in soup.find_all('a', href=re.compile(r'/en/properties\?page=(\d+)')):
        m = re.search(r'page=(\d+)', a['href'])
        if m:
            last_page = max(last_page, int(m.group(1)))
    print(f"Total pages: {last_page}")
    
    def process_pollentia_page(soup_page):
        count = 0
        prop_links = list(set([
            a['href'] for a in soup_page.find_all('a', href=True)
            if '/property/id/' in a['href']
        ]))
        
        for link in prop_links:
            full_url = 'https://www.pollentiaproperties.com' + link if not link.startswith('http') else link
            
            # Slug gives info
            slug = link.split('/')[-1]
            
            # Title from slug
            # e.g. "962544-townhouse-for-sale-pollensa-pool-terrace-garage"
            parts = slug.split('-')
            if parts and parts[0].isdigit():
                parts = parts[1:]  # remove ID
            titel = ' '.join(parts).title()[:100]
            
            # Location from slug
            ort_item = 'Mallorca'
            loc_keywords = ['pollensa', 'palma', 'arta', 'alcudia', 'soller', 'andratx', 'calvia', 'manacor', 'felanitx', 'santanyi']
            for kw in loc_keywords:
                if kw in slug.lower():
                    ort_item = kw.title()
                    break
            
            # Try to get price from container
            a_el = soup_page.find('a', href=link)
            preis = None
            zimmer = None
            flaeche = None
            if a_el:
                parent = a_el.find_parent(['article', 'div', 'li', 'td'])
                if parent:
                    txt = parent.get_text(separator=' ', strip=True)
                    preis = parse_price(txt)
                    bed_m = re.search(r'(\d)\s*(?:hab|bed|room|zimmer)', txt, re.I)
                    if bed_m: zimmer = int(bed_m.group(1))
                    sqm_m = re.search(r'(\d+)\s*m[²2]', txt)
                    if sqm_m: flaeche = int(sqm_m.group(1))
            
            if save_obj(titel, 'Pollentia Properties', full_url, preis=preis, zimmer=zimmer, flaeche=flaeche, ort=ort_item):
                count += 1
        return count
    
    count_pollentia += process_pollentia_page(soup)
    print(f"Page 1 (from /en/results): +{count_pollentia}")
    
    # Pages 2 to last_page
    for page_num in range(2, min(last_page + 1, 50)):  # cap at 50 pages = ~550 properties
        time.sleep(0.4)
        url = f'https://www.pollentiaproperties.com/en/properties?page={page_num}'
        try:
            r = requests.get(url, headers=headers, timeout=20)
            if r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')
                n = process_pollentia_page(soup)
                count_pollentia += n
                if page_num % 5 == 0 or n > 0:
                    print(f"  Page {page_num}/{last_page}: +{n} (total {count_pollentia})")
            else:
                print(f"  Page {page_num}: HTTP {r.status_code}")
        except Exception as e:
            print(f"  Page {page_num}: Error {e}")
    
    results['Pollentia Properties'] = count_pollentia
    print(f"Pollentia Properties: {count_pollentia} new objects")

except Exception as e:
    print(f"Error: {e}")
    import traceback; traceback.print_exc()
    results['Pollentia Properties'] = count_pollentia

# ============================================================
# 3. MALLORCA FINEST (mallorcafinest.com)
# SSL error + domain ist kein Immobilien-Anbieter → skip
# ============================================================
print("\n=== Mallorca Finest ===")
print("SKIP: mallorcafinest.com SSL-Fehler (TLSV1_UNRECOGNIZED_NAME)")
print("      mallorcasfinest.com = Lifestyle-Magazin, keine Immobilien")
results['Mallorca Finest'] = 0

# ============================================================
# SAVE
# ============================================================
wb.save(EXCEL)
print(f"\n=== ERGEBNIS ===")
for src, count in results.items():
    print(f"  {src}: {count} Objekte")
print(f"  Gesamt neu: {sum(results.values())}")
