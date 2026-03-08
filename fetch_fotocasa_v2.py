#!/usr/bin/env python3
"""
fetch_fotocasa_v2.py — Fotocasa Grundstücksdaten via direktes HTTP (kein Proxy, keine Playwright)
Spanische IP direkt → SSR-Seiten mit vollem HTML. Korrigiertes Terreno-Pattern.
"""

import json, re, time, sys
sys.stdout.reconfigure(line_buffering=True)
import openpyxl
import requests

EXCEL_PATH = "/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx"
PROGRESS_PATH = "/Users/robin/.openclaw/workspace/mallorca-projekt/fetchdetails_progress.json"

URL_COL  = 3   # Column C
PLOT_COL = 6   # Column F = Grundstück (m²)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Language': 'es-ES,es;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Cache-Control': 'max-age=0',
}

PATTERNS = [
    r'([0-9][0-9.]+)\s*m\xb2\s+de\s+terreno',       # "28.000 m² de terreno"
    r'([0-9][0-9.]+)\s*m\xb2\s+de\s+parcela',        # "28.000 m² de parcela"
    r'([0-9][0-9.]+)\s*m\xb2\s+terreno',              # "564 m² terreno"
    r'terreno\s+de\s+([0-9][0-9.]+)\s*m',             # "terreno de 1.310 m"
    r'parcela[,\s]+(?:de\s+)?([0-9][0-9.]+)\s*m',     # "parcela de 396 m"
    r'[Tt]erreno[\":\s]+([0-9][0-9.]+)',               # "Terreno: 73.715"
    r'superficieParcela[\":\s]+([0-9]+)',               # JSON field
    r'plotSurface[\":\s]+([0-9]+)',
    r'landSurface[\":\s]+([0-9]+)',
    r'superficie_parcela[\":\s]+([0-9]+)',
]

def extract_plot(html):
    for pat in PATTERNS:
        for m in re.finditer(pat, html, re.IGNORECASE):
            raw = m.group(1).replace('.', '').replace(',', '')
            try:
                val = int(raw)
                if val >= 100:
                    return val
            except ValueError:
                continue
    return None

UA_LIST = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0',
]

def make_session():
    """Create a fresh session and warm it up with homepage visit."""
    s = requests.Session()
    s.headers.update(HEADERS)
    try:
        s.get('https://www.fotocasa.es/', timeout=10)
        time.sleep(3)
    except Exception:
        pass
    return s

def fetch_plot(url, session):
    try:
        r = session.get(url, headers={'Referer': 'https://www.fotocasa.es/es/comprar/terrenos/mallorca/todas-las-zonas/l'}, timeout=15)
        if r.status_code == 200 and len(r.text) > 50000:
            return extract_plot(r.text), len(r.text)
        return None, f"HTTP {r.status_code} len={len(r.text)}"
    except Exception as e:
        return None, str(e)

# Load Excel
print("Lade Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

# Load progress
try:
    with open(PROGRESS_PATH) as f:
        progress = json.load(f)
except Exception:
    progress = {}

# Collect only terreno URLs not yet with valid data
to_process = []
for row in ws.iter_rows(min_row=2, values_only=False):
    url_cell = row[URL_COL - 1]
    plot_cell = row[PLOT_COL - 1]
    if not url_cell.value:
        continue
    url = str(url_cell.value).strip()
    # Only terreno URLs (these are actual land listings)
    if '/terreno' not in url.lower():
        continue
    # Skip if already has valid plot value
    existing = plot_cell.value
    if existing and isinstance(existing, (int, float)) and existing > 100:
        print(f"Row {url_cell.row}: bereits {existing} m² → skip")
        continue
    to_process.append((url_cell.row, url))

print(f"Terreno URLs to process: {len(to_process)}")
print("=" * 60)

import random

SESSION_REFRESH_EVERY = 40  # Re-warm session every N requests

results_found = {}
results_null  = []
session = None
consecutive_blocked = 0

for i, (row_num, url) in enumerate(to_process):
    # Refresh session every N requests or at start
    if i % SESSION_REFRESH_EVERY == 0 or consecutive_blocked >= 3:
        print(f"\n[Session refresh at {i+1}/{len(to_process)}...]", flush=True)
        session = make_session()
        consecutive_blocked = 0
        print("Session ready.", flush=True)

    print(f"\n[{i+1}/{len(to_process)}] Row {row_num}: {url[:80]}")

    plot, meta = fetch_plot(url, session)
    print(f"  {meta}, plot={plot}")

    if plot:
        ws.cell(row=row_num, column=PLOT_COL).value = plot
        results_found[url] = plot
        consecutive_blocked = 0
    else:
        results_null.append(url)
        if isinstance(meta, str) and '13080' in meta:
            consecutive_blocked += 1

    progress[url] = {'plot': plot, 'done': True}

    # Checkpoint every 10 rows
    if (i + 1) % 10 == 0 or (i + 1) == len(to_process):
        wb.save(EXCEL_PATH)
        with open(PROGRESS_PATH, 'w') as f:
            json.dump(progress, f, indent=2)
        print(f"  [Checkpoint — {i+1}/{len(to_process)}]")

    if i < len(to_process) - 1:
        delay = random.uniform(5, 9)
        print(f"  [sleep {delay:.1f}s]")
        time.sleep(delay)

print("\n" + "=" * 60)
print("FINALER POOL — Fotocasa Terreno mit Grundstücksgröße:")
print("=" * 60)
print(f"Gefunden:       {len(results_found)}")
print(f"Nicht gefunden: {len(results_null)}")
print()
for url, plot in sorted(results_found.items(), key=lambda x: -x[1]):
    gemeinde = url.split('/')[6] if len(url.split('/')) > 6 else '?'
    print(f"  {plot:>8} m²  |  {gemeinde:<30}  |  {url[-30:]}")

print("\n✓ Fertig.")
