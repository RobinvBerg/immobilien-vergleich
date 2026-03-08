#!/usr/bin/env python3
"""
Phase 2: Makler-Websites Scraper
Scrapes 23 real estate agencies for Mallorca properties with 5+ rooms
"""

import sys
import json
import time
import re
import traceback
from datetime import date
from openpyxl import load_workbook

# Try requests+BS4 first, playwright for JS-heavy sites
import requests
from bs4 import BeautifulSoup

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Language': 'de,en;q=0.9',
}

all_results = []

def fetch(url, timeout=20):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return r
    except Exception as e:
        print(f"  ✗ fetch error {url}: {e}")
        return None

def clean_price(txt):
    if not txt:
        return None
    txt = str(txt).replace('\xa0', '').replace(' ', '').replace('.', '').replace(',', '')
    m = re.search(r'(\d{4,})', txt)
    if m:
        return int(m.group(1))
    return None

def clean_int(txt):
    if not txt:
        return None
    m = re.search(r'(\d+)', str(txt).replace('.', '').replace(',', ''))
    if m:
        return int(m.group(1))
    return None

def save_to_excel(new_objects):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set(str(row[2]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[2])
    
    new_count = 0
    for obj in new_objects:
        url = obj.get('url', '—')
        if url and url != '—' and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', 'Unbekannt'),
            obj.get('quelle', ''),
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            str(date.today()),
            'Neu'
        ])
        if url and url != '—':
            existing_urls.add(url)
        new_count += 1
    
    wb.save(EXCEL_PATH)
    return new_count

# ─────────────────────────────────────────────────────────────
# 1. PORTA MALLORQUINA
# ─────────────────────────────────────────────────────────────
def scrape_porta_mallorquina():
    print("\n[1] Porta Mallorquina...")
    results = []
    try:
        # They have a JSON API
        base = "https://www.porta-mallorquina.com/api/properties"
        params = {
            'filter[location]': 'mallorca',
            'filter[rooms_min]': 5,
            'filter[property_type]': '',
            'page': 1,
            'per_page': 50
        }
        # Try their search page
        urls_to_try = [
            "https://www.porta-mallorquina.com/immobilien/kaufen/mallorca/?rooms_min=5&has_plot=1",
            "https://www.porta-mallorquina.com/immobilien/kaufen/mallorca/?rooms=5",
            "https://www.porta-mallorquina.com/de/kaufen?zimmer_min=5",
        ]
        for url in urls_to_try:
            r = fetch(url)
            if r and r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')
                # Look for property cards
                cards = soup.select('.property-card, .listing-item, article.property, .estate-item, .object-item')
                if not cards:
                    cards = soup.select('[class*="property"], [class*="listing"], [class*="estate"]')
                print(f"  Found {len(cards)} cards at {url}")
                for card in cards[:50]:
                    title = card.select_one('h2, h3, .title, [class*="title"]')
                    price = card.select_one('[class*="price"], .price')
                    link = card.select_one('a[href]')
                    if title or price:
                        obj = {
                            'titel': title.get_text(strip=True) if title else 'Porta Mallorquina Objekt',
                            'quelle': 'Porta Mallorquina',
                            'url': link['href'] if link else url,
                            'preis': clean_price(price.get_text() if price else None),
                        }
                        results.append(obj)
                if results:
                    break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 2. UNIQUE MALLORCA
# ─────────────────────────────────────────────────────────────
def scrape_unique_mallorca():
    print("\n[2] Unique Mallorca...")
    results = []
    try:
        urls = [
            "https://www.uniquemallorca.com/properties/?min_rooms=5",
            "https://www.uniquemallorca.com/buy/",
            "https://www.uniquemallorca.com/properties/",
        ]
        for url in urls:
            r = fetch(url)
            if not r or r.status_code != 200:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property-item, .listing, .property-card, article')
            if not cards:
                # Try JSON-LD
                scripts = soup.select('script[type="application/ld+json"]')
                for s in scripts:
                    try:
                        data = json.loads(s.string)
                        if isinstance(data, list):
                            for item in data:
                                if 'name' in item:
                                    results.append({'titel': item.get('name'), 'quelle': 'Unique Mallorca', 'url': item.get('url',''), 'preis': None})
                    except:
                        pass
            for card in cards[:50]:
                title = card.select_one('h2, h3, .property-title, [class*="title"]')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                rooms_el = card.select_one('[class*="room"], [class*="bed"]')
                if title:
                    obj = {
                        'titel': title.get_text(strip=True),
                        'quelle': 'Unique Mallorca',
                        'url': link['href'] if link else url,
                        'preis': clean_price(price.get_text() if price else None),
                        'zimmer': clean_int(rooms_el.get_text() if rooms_el else None),
                    }
                    results.append(obj)
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 3. BALEARIC PROPERTIES
# ─────────────────────────────────────────────────────────────
def scrape_balearic_properties():
    print("\n[3] Balearic Properties...")
    results = []
    try:
        urls = [
            "https://www.balearicproperties.com/buy/?bedrooms=5&island=Mallorca",
            "https://www.balearicproperties.com/properties/?type=buy&bedrooms_min=5",
            "https://www.balearicproperties.com/buy/",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .listing-item, .property-card, article')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.balearicproperties.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Balearic Properties',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 4. FINCALLORCA
# ─────────────────────────────────────────────────────────────
def scrape_fincallorca():
    print("\n[4] Fincallorca...")
    results = []
    try:
        urls = [
            "https://www.fincallorca.com/en/mallorca/buy/?rooms=5",
            "https://www.fincallorca.com/en/mallorca/buy/",
            "https://www.fincallorca.com/de/mallorca/kaufen/",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property-item, .listing, article.property, .property')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:60]:
                title = card.select_one('h2, h3, .property-name, [class*="title"]')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                beds = card.select_one('[class*="bed"], [class*="room"]')
                plot = card.select_one('[class*="plot"], [class*="grundst"]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.fincallorca.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Fincallorca',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                        'zimmer': clean_int(beds.get_text() if beds else None),
                        'grundstueck': clean_int(plot.get_text() if plot else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 5. MALLORCA FINEST
# ─────────────────────────────────────────────────────────────
def scrape_mallorca_finest():
    print("\n[5] Mallorca Finest...")
    results = []
    try:
        urls = [
            "https://www.mallorcafinest.com/properties/?type=sale&rooms=5",
            "https://www.mallorcafinest.com/buy/",
            "https://www.mallorcafinest.com/properties/",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .property-item, article, .listing')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.mallorcafinest.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Mallorca Finest',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 6. SANDBERG ESTATES
# ─────────────────────────────────────────────────────────────
def scrape_sandberg():
    print("\n[6] Sandberg Estates...")
    results = []
    try:
        urls = [
            "https://www.sandberg-estates.com/properties/?location=mallorca&rooms_min=5",
            "https://www.sandberg-estates.com/buy/",
            "https://www.sandberg-estates.com/properties/",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .property-card, article, .listing-item')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.sandberg-estates.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Sandberg Estates',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 7. PRIVATE PROPERTY MALLORCA
# ─────────────────────────────────────────────────────────────
def scrape_private_property():
    print("\n[7] Private Property Mallorca...")
    results = []
    try:
        urls = [
            "https://www.privatepropertymallorca.com/properties/?rooms=5",
            "https://www.privatepropertymallorca.com/buy/",
            "https://www.privatepropertymallorca.com/",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, article, .listing, .property-card')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.privatepropertymallorca.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Private Property Mallorca',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 8. MALLORCA DREAM HOMES
# ─────────────────────────────────────────────────────────────
def scrape_mallorca_dream():
    print("\n[8] Mallorca Dream Homes...")
    results = []
    try:
        urls = [
            "https://www.mallorcadreamhomes.com/properties/?bedrooms=5",
            "https://www.mallorcadreamhomes.com/buy/",
            "https://www.mallorcadreamhomes.com/properties/",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .property-card, article, .listing')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.mallorcadreamhomes.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Mallorca Dream Homes',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 9. PRESTIGE PROPERTY GROUP
# ─────────────────────────────────────────────────────────────
def scrape_prestige():
    print("\n[9] Prestige Property Group...")
    results = []
    try:
        urls = [
            "https://www.prestigepropertygroup.com/mallorca/buy/?bedrooms=5",
            "https://www.prestigepropertygroup.com/mallorca/",
            "https://www.prestigepropertygroup.com/search/?location=mallorca&rooms_min=5",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .property-item, .listing-item, article')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.prestigepropertygroup.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Prestige Property Group',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 10. POLLENSA PROPERTIES
# ─────────────────────────────────────────────────────────────
def scrape_pollensa():
    print("\n[10] Pollensa Properties...")
    results = []
    try:
        urls = [
            "https://www.pollensaproperties.com/buy/?bedrooms_min=5",
            "https://www.pollensaproperties.com/properties/",
            "https://www.pollensaproperties.com/",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, article, .listing, .property-card')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.pollensaproperties.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Pollensa Properties',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 11. MINKNER & BONITZ (already in Excel but try)
# ─────────────────────────────────────────────────────────────
def scrape_minkner():
    print("\n[11] Minkner & Bonitz...")
    results = []
    try:
        urls = [
            "https://www.minkner.com/kaufen/?zimmer_min=5",
            "https://www.minkner.com/immobilien/kaufen/",
            "https://www.minkner.com/de/kaufen/",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .object, .estate-item, article, .listing')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"], [class*="preis"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.minkner.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Minkner & Bonitz',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 12. KENSINGTON FINEST
# ─────────────────────────────────────────────────────────────
def scrape_kensington():
    print("\n[12] Kensington Finest...")
    results = []
    try:
        urls = [
            "https://www.kensington.com/de/immobilien/kaufen/spanien/mallorca/?zimmer_min=5",
            "https://www.kensington.com/de/immobilien/kaufen/spanien/mallorca/",
            "https://www.kensington.com/immobilien/mallorca/",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .listing-item, .property-card, article.estate')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title, [class*="title"]')
                price = card.select_one('[class*="price"], [class*="preis"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.kensington.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Kensington Finest',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 13. DAHLER COMPANY
# ─────────────────────────────────────────────────────────────
def scrape_dahler():
    print("\n[13] DAHLER Company...")
    results = []
    try:
        urls = [
            "https://www.dahlercompany.com/immobilien/kaufen/mallorca/?zimmer_min=5",
            "https://www.dahlercompany.com/immobilien/kaufen/mallorca/",
            "https://www.dahlercompany.com/de/immobilien/?location=mallorca",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .estate, .listing, article, [class*="property"]')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.dahlercompany.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'DAHLER Company',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 14. FINE & COUNTRY
# ─────────────────────────────────────────────────────────────
def scrape_fine_country():
    print("\n[14] Fine & Country...")
    results = []
    try:
        urls = [
            "https://www.fineandcountry.com/search-properties?location=Mallorca&rooms_min=5",
            "https://www.fineandcountry.com/es/properties-for-sale/mallorca",
            "https://www.fineandcountry.com/search-properties?country=ES&location=mallorca",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .property-card, article, .listing-item')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title, .property-title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.fineandcountry.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Fine & Country',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 15. LUCAS FOX
# ─────────────────────────────────────────────────────────────
def scrape_lucas_fox():
    print("\n[15] Lucas Fox...")
    results = []
    try:
        # Lucas Fox has an API
        urls = [
            "https://www.lucasfox.com/api/properties?location=mallorca&bedrooms_min=5&for_sale=1",
            "https://www.lucasfox.com/mallorca-property-for-sale/?bedrooms=5",
            "https://www.lucasfox.com/mallorca-property-for-sale/",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            # Try JSON first
            try:
                data = r.json()
                props = data.get('properties', data.get('results', data.get('data', [])))
                if isinstance(props, list) and props:
                    for p in props[:60]:
                        results.append({
                            'titel': p.get('title', p.get('name', 'Lucas Fox')),
                            'quelle': 'Lucas Fox',
                            'url': p.get('url', p.get('link', '')),
                            'preis': clean_price(p.get('price', p.get('sale_price', ''))),
                            'zimmer': clean_int(p.get('bedrooms', p.get('rooms', ''))),
                            'grundstueck': clean_int(p.get('plot_area', p.get('plot', ''))),
                            'wohnflaeche': clean_int(p.get('built_area', p.get('area', ''))),
                            'ort': p.get('location', p.get('area', '')),
                        })
                    break
            except:
                pass
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .property-card, article, .listing')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.lucasfox.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Lucas Fox',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 16. KNIGHT FRANK
# ─────────────────────────────────────────────────────────────
def scrape_knight_frank():
    print("\n[16] Knight Frank...")
    results = []
    try:
        urls = [
            "https://www.knightfrank.com/search/results.aspx?SearchType=Buy&Location=Mallorca&Bedrooms=5",
            "https://www.knightfrank.com/international-residential/search/?country=ES&area=Mallorca&bedrooms=5",
            "https://www.knightfrank.com/search/?location=mallorca&bedrooms_min=5&transaction=sale",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .listing-item, article, [class*="property-card"]')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.knightfrank.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Knight Frank',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 17. SAVILLS
# ─────────────────────────────────────────────────────────────
def scrape_savills():
    print("\n[17] Savills...")
    results = []
    try:
        urls = [
            "https://www.savills.es/find-a-property/residential-property-for-sale/spain/mallorca/bedrooms-5-",
            "https://www.savills.com/find-a-property/residential-property-for-sale/spain/balearic-islands/mallorca",
            "https://www.savills.es/find-a-property/residential-property-for-sale/spain/mallorca",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .property-card, article, .listing-item, [class*="property-list"]')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.savills.es' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Savills',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 18. COLDWELL BANKER
# ─────────────────────────────────────────────────────────────
def scrape_coldwell():
    print("\n[18] Coldwell Banker...")
    results = []
    try:
        urls = [
            "https://www.coldwellbanker.es/propiedades/?location=mallorca&bedrooms=5",
            "https://www.coldwellbanker.es/mallorca/",
            "https://www.coldwellbanker.es/en/mallorca/properties-for-sale/",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .property-card, article, .listing')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.coldwellbanker.es' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Coldwell Banker',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 19. VON POLL
# ─────────────────────────────────────────────────────────────
def scrape_vonpoll():
    print("\n[19] Von Poll Real Estate...")
    results = []
    try:
        urls = [
            "https://www.von-poll.com/de/immobilien/spanien/balearen/mallorca?zimmer_min=5",
            "https://www.von-poll.com/de/immobilien/spanien/balearen/mallorca",
            "https://www.von-poll.com/de/immobilien?region=mallorca",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.estate, .property, .listing-item, article, [class*="expose"]')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title, [class*="title"]')
                price = card.select_one('[class*="price"], [class*="preis"]')
                link = card.select_one('a[href]')
                rooms = card.select_one('[class*="zimmer"], [class*="room"]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.von-poll.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Von Poll Real Estate',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                        'zimmer': clean_int(rooms.get_text() if rooms else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 20. ENGEL & VÖLKERS
# ─────────────────────────────────────────────────────────────
def scrape_ev():
    print("\n[20] Engel & Völkers...")
    results = []
    try:
        urls = [
            "https://www.engelvoelkers.com/de/search/?q=&domainId=21&businessArea=residential&i=0&mode=buy&zip=&city=&country=ESP&categories=villa%2Cfinca&rooms=5.0",
            "https://www.engelvoelkers.com/de-de/search/?q=mallorca&mode=buy&categories=villa,finca&rooms=5",
            "https://www.engelvoelkers.com/de/spanien/balearen/mallorca/?mode=buy&rooms=5",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.ev-property-card, .property-item, article[class*="property"], .listing')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:60]:
                title = card.select_one('h2, h3, .ev-property-card__title, .title')
                price = card.select_one('[class*="price"], .ev-property-card__price')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.engelvoelkers.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Engel & Völkers',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 21. SOTHEBY'S MALLORCA
# ─────────────────────────────────────────────────────────────
def scrape_sothebys():
    print("\n[21] Mallorca Sotheby's...")
    results = []
    try:
        urls = [
            "https://www.sothebysrealty.com/eng/sales/mallorca-sp?bedrooms=5",
            "https://www.sothebysrealty.com/eng/sales/country/sp/region/mallorca",
            "https://www.sothebysrealty.com/eng/sales/mallorca-sp",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .listing-item, article, .property-card, [class*="listing"]')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title, .listing-title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.sothebysrealty.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': "Mallorca Sotheby's",
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 22. THE AGENCY RE
# ─────────────────────────────────────────────────────────────
def scrape_the_agency():
    print("\n[22] The Agency RE...")
    results = []
    try:
        urls = [
            "https://www.theagencyre.com/search/?location=mallorca&bedrooms=5",
            "https://www.theagencyre.com/mallorca/",
            "https://www.theagencyre.com/search/?location=mallorca",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .property-card, article, .listing')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.theagencyre.com' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'The Agency RE',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# 23. RE/MAX MALLORCA
# ─────────────────────────────────────────────────────────────
def scrape_remax():
    print("\n[23] Re/Max Mallorca...")
    results = []
    try:
        urls = [
            "https://www.remax.es/Comprar/mallorca?habitaciones=5",
            "https://www.remax.es/Comprar/Islas+Baleares/Mallorca",
            "https://www.remax.es/Buscar?operacion=Comprar&provincia=Islas+Baleares&municipio=Mallorca",
        ]
        for url in urls:
            r = fetch(url)
            if not r:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.select('.property, .listing, .property-card, article, [class*="property"]')
            print(f"  {len(cards)} cards at {url}")
            for card in cards[:50]:
                title = card.select_one('h2, h3, .title')
                price = card.select_one('[class*="price"], [class*="precio"]')
                link = card.select_one('a[href]')
                if title:
                    href = link['href'] if link else ''
                    if href and not href.startswith('http'):
                        href = 'https://www.remax.es' + href
                    results.append({
                        'titel': title.get_text(strip=True),
                        'quelle': 'Re/Max Mallorca',
                        'url': href or url,
                        'preis': clean_price(price.get_text() if price else None),
                    })
            if results:
                break
    except Exception as e:
        print(f"  Error: {e}")
    print(f"  → {len(results)} Objekte")
    return results

# ─────────────────────────────────────────────────────────────
# RUN ALL
# ─────────────────────────────────────────────────────────────
scrapers = [
    scrape_porta_mallorquina,
    scrape_unique_mallorca,
    scrape_balearic_properties,
    scrape_fincallorca,
    scrape_mallorca_finest,
    scrape_sandberg,
    scrape_private_property,
    scrape_mallorca_dream,
    scrape_prestige,
    scrape_pollensa,
    scrape_minkner,
    scrape_kensington,
    scrape_dahler,
    scrape_fine_country,
    scrape_lucas_fox,
    scrape_knight_frank,
    scrape_savills,
    scrape_coldwell,
    scrape_vonpoll,
    scrape_ev,
    scrape_sothebys,
    scrape_the_agency,
    scrape_remax,
]

summary = {}
for scraper in scrapers:
    name = scraper.__name__.replace('scrape_', '')
    try:
        items = scraper()
        all_results.extend(items)
        summary[name] = len(items)
    except Exception as e:
        print(f"  FATAL: {e}")
        summary[name] = 0

print(f"\n\n=== TOTAL RAW: {len(all_results)} ===")

# Save to JSON for inspection
with open('/Users/robin/.openclaw/workspace/mallorca-projekt/phase2_raw_results.json', 'w') as f:
    json.dump(all_results, f, ensure_ascii=False, indent=2)
print("Raw results saved to phase2_raw_results.json")

# Save to Excel
saved = save_to_excel(all_results)
print(f"\n=== SAVED TO EXCEL: {saved} neue Objekte ===")

print("\n--- Summary by source ---")
for name, count in sorted(summary.items(), key=lambda x: -x[1]):
    print(f"  {name}: {count}")
