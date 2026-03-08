#!/usr/bin/env python3
"""Fill mallorca-kandidaten-v2.xlsx from row 15 onward with auto-computed values."""

import openpyxl
import re
import time
import requests

EXCEL_PATH = "mallorca-kandidaten-v2.xlsx"
API_KEY = "AIzaSyAzplaFOGUrCygk3mrzS--wrHNIz0arKok"

DESTINATIONS = {
    "flughafen": "Aeropuerto de Palma de Mallorca, Spain",
    "deia":      "Deia, Mallorca, Spain",
    "andratx":   "Andratx, Mallorca, Spain",
    "ses_salines": "Ses Salines, Mallorca, Spain",
}

# Column indices (1-based)
COL = {
    'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6,
    'G': 7, 'H': 8, 'I': 9, 'J': 10,
    'K': 11, 'L': 12, 'M': 13, 'N': 14,
    'O': 15, 'P': 16, 'Q': 17, 'R': 18,
    'S': 19, 'T': 20, 'U': 21,
    'V': 22, 'W': 23, 'X': 24, 'Y': 25,
    'Z': 26, 'AA': 27, 'AB': 28, 'AC': 29,
    'AD': 30, 'AE': 31, 'AF': 32, 'AG': 33,
    'AH': 34, 'AI': 35, 'AJ': 36, 'AK': 37, 'AL': 38,
}

def get_distances(origin_location):
    """Call Google Maps Distance Matrix API for all 4 destinations."""
    if not origin_location or str(origin_location).strip() in ('', 'Balearic Islands', 'Spain', 'Mallorca'):
        return None
    
    origin = f"{origin_location}, Mallorca, Spain"
    destinations_list = list(DESTINATIONS.values())
    dest_str = "|".join(destinations_list)
    
    url = "https://maps.googleapis.com/maps/api/distancematrix/json"
    params = {
        "origins": origin,
        "destinations": dest_str,
        "mode": "driving",
        "key": API_KEY,
        "language": "de",
    }
    
    try:
        r = requests.get(url, params=params, timeout=10)
        data = r.json()
        if data.get("status") != "OK":
            print(f"  API error for '{origin}': {data.get('status')}")
            return None
        
        results = {}
        elements = data["rows"][0]["elements"]
        dest_keys = list(DESTINATIONS.keys())
        
        for i, key in enumerate(dest_keys):
            el = elements[i]
            if el.get("status") == "OK":
                dist_km = round(el["distance"]["value"] / 1000, 1)
                dur_min = round(el["duration"]["value"] / 60)
                results[key] = (dist_km, dur_min)
            else:
                results[key] = None
        
        return results
    except Exception as e:
        print(f"  Exception for '{origin}': {e}")
        return None


def extract_makler_ref(url):
    if not url:
        return None
    # Idealista
    m = re.search(r'/inmueble/(\d+)/', str(url))
    if m:
        return m.group(1)
    # EV (Engel & Völkers)
    m = re.search(r'/exposes/([a-f0-9-]+)', str(url))
    if m:
        return m.group(1)
    # Rightmove
    m = re.search(r'/property-for-sale/(\d+)', str(url))
    if m:
        return m.group(1)
    return None


def erreichbarkeit_score(min_flughafen):
    if min_flughafen is None:
        return None
    m = int(min_flughafen)
    if m <= 20:
        return 100
    elif m <= 30:
        return 85
    elif m <= 40:
        return 70
    elif m <= 50:
        return 50
    else:
        return 30


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    max_row = ws.max_row
    print(f"Total rows: {max_row}, processing rows 15-{max_row}")
    
    stats = {
        'distances': 0,
        'eur_m2_T': 0,
        'eur_m2_U': 0,
        'makler_ref': 0,
        'erreichbarkeit': 0,
    }
    
    # Cache distances per location
    dist_cache = {}
    
    for row in range(15, max_row + 1):
        def g(col_key):
            return ws.cell(row, COL[col_key]).value
        def s(col_key, val):
            ws.cell(row, COL[col_key]).value = val
        
        location = g('J')
        url = g('C')
        preis = g('S')
        wohnflaeche = g('H')
        grundstueck = g('G')
        
        # 1. Distances
        needs_dist = (g('K') is None or g('L') is None or g('M') is None or 
                      g('N') is None or g('O') is None or g('P') is None or
                      g('Q') is None or g('R') is None)
        
        if needs_dist and location:
            loc_key = str(location).strip()
            if loc_key not in dist_cache:
                print(f"  Row {row}: Fetching distances for '{loc_key}'")
                dist_cache[loc_key] = get_distances(loc_key)
                time.sleep(0.2)
            
            dists = dist_cache.get(loc_key)
            if dists:
                if g('K') is None and dists.get('flughafen'):
                    s('K', dists['flughafen'][0]); stats['distances'] += 1
                if g('L') is None and dists.get('flughafen'):
                    s('L', dists['flughafen'][1]); stats['distances'] += 1
                if g('M') is None and dists.get('deia'):
                    s('M', dists['deia'][0]); stats['distances'] += 1
                if g('N') is None and dists.get('deia'):
                    s('N', dists['deia'][1]); stats['distances'] += 1
                if g('O') is None and dists.get('andratx'):
                    s('O', dists['andratx'][0]); stats['distances'] += 1
                if g('P') is None and dists.get('andratx'):
                    s('P', dists['andratx'][1]); stats['distances'] += 1
                if g('Q') is None and dists.get('ses_salines'):
                    s('Q', dists['ses_salines'][0]); stats['distances'] += 1
                if g('R') is None and dists.get('ses_salines'):
                    s('R', dists['ses_salines'][1]); stats['distances'] += 1
        
        # 2. €/m²
        if g('T') is None and preis and wohnflaeche and float(wohnflaeche) > 0:
            s('T', round(float(preis) / float(wohnflaeche)))
            stats['eur_m2_T'] += 1
        if g('U') is None and preis and grundstueck and float(grundstueck) > 0:
            s('U', round(float(preis) / float(grundstueck)))
            stats['eur_m2_U'] += 1
        
        # 3. Makler-Ref
        if g('AK') is None and url:
            ref = extract_makler_ref(url)
            if ref:
                s('AK', ref)
                stats['makler_ref'] += 1
        
        # 4. Erreichbarkeit Score (Y) — after we have L filled
        if g('Y') is None:
            min_flug = g('L')
            score = erreichbarkeit_score(min_flug)
            if score is not None:
                s('Y', score)
                stats['erreichbarkeit'] += 1
        
        # Checkpoint every 50 rows
        if (row - 14) % 50 == 0:
            wb.save(EXCEL_PATH)
            print(f"  Checkpoint: saved at row {row}")
    
    wb.save(EXCEL_PATH)
    print(f"\nDone! Stats:")
    for k, v in stats.items():
        print(f"  {k}: {v} fields filled")
    total = sum(stats.values())
    print(f"  TOTAL: {total} fields filled")


if __name__ == "__main__":
    main()
