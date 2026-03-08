#!/usr/bin/env python3
"""
Fills missing Location + Entfernungen for 22 entries via Google Maps API.
Extracts location from property name, geocodes, calculates distances.
"""
import requests, json, re, time
import openpyxl
from pathlib import Path

BASE = Path(__file__).parent
MAPS_KEY = 'AIzaSyAzplaFOGUrCygk3mrzS--wrHNIz0arKok'

# Reference points
REFS = {
    'Flughafen': 'Palma de Mallorca Airport, Spain',
    'Daia':      'Daia, Mallorca, Spain',
    'Andratx':   'Andratx, Mallorca, Spain',
    'Ses Salines': 'Ses Salines, Mallorca, Spain',
}

# Known locations from property names
KNOWN_LOCS = {
    16:  'Mallorca, Spain',           # panoramic sea views - generic
    26:  'Mallorca, Spain',           # winery - generic
    39:  'Mallorca, Spain',           # equestrian center - generic
    47:  'Mallorca, Spain',           # country resort - generic
    53:  'Mallorca, Spain',           # vineyard bodega - generic
    56:  'Santa Margalida, Mallorca, Spain',
    66:  'Northeast Mallorca, Spain',
    67:  'Mallorca, Spain',
    71:  'Son Macia, Mallorca, Spain',
    101: 'Es Carritxo, Mallorca, Spain',
    126: 'Es Trenc, Mallorca, Spain',
    134: 'Mallorca, Spain',
    146: 'Mallorca, Spain',
    152: 'Felanitx, Mallorca, Spain',
    178: 'Marratxi, Mallorca, Spain',
    185: 'Mallorca, Spain',
    193: 'Mallorca, Spain',
    195: 'Mallorca, Spain',
    227: 'Mallorca, Spain',
    244: 'Montuiri, Mallorca, Spain',
    248: 'Mallorca, Spain',
    252: 'Mallorca, Spain',
    265: 'Binissalem, Mallorca, Spain',
    293: 'Campos, Mallorca, Spain',
    303: 'Santa Maria del Cami, Mallorca, Spain',
    310: 'Genova, Mallorca, Spain',
    329: 'Mallorca, Spain',
}

# Display names (short)
DISPLAY_LOCS = {
    56: 'Santa Margalida', 71: 'Son Macia', 101: 'Es Carritxó',
    126: 'Es Trenc (Süd)', 152: 'Felanitx', 178: 'Marratxí',
    244: 'Montuïri', 265: 'Binissalem', 293: 'Campos',
    303: 'Santa Maria del Camí', 310: 'Genova',
}

def geocode(address):
    url = 'https://maps.googleapis.com/maps/api/geocode/json'
    r = requests.get(url, params={'address': address, 'key': MAPS_KEY}, timeout=10)
    data = r.json()
    if data['status'] == 'OK':
        loc = data['results'][0]['geometry']['location']
        return loc['lat'], loc['lng']
    return None, None

def get_distances(origin_lat, origin_lng, destinations):
    """Get distances and durations via Distance Matrix API."""
    dest_str = '|'.join(destinations)
    origin_str = f"{origin_lat},{origin_lng}"
    url = 'https://maps.googleapis.com/maps/api/distancematrix/json'
    r = requests.get(url, params={
        'origins': origin_str,
        'destinations': dest_str,
        'key': MAPS_KEY,
        'mode': 'driving',
        'units': 'metric'
    }, timeout=15)
    data = r.json()
    results = []
    if data['status'] == 'OK':
        for elem in data['rows'][0]['elements']:
            if elem['status'] == 'OK':
                km = round(elem['distance']['value'] / 1000, 1)
                mins = round(elem['duration']['value'] / 60)
                results.append((km, mins))
            else:
                results.append((None, None))
    return results

def main():
    wb = openpyxl.load_workbook(BASE / 'mallorca-kandidaten-v2.xlsx')
    ws = wb.active
    headers = [c.value for c in ws[1]]

    # Column indices
    loc_idx = headers.index('Location')
    cols = {
        'Flughafen_km':  headers.index('Entfernung Flughafen (km)'),
        'Flughafen_min': headers.index('Entfernung Flughafen (min)'),
        'Daia_km':       headers.index('Entfernung Daia Haus (km)'),
        'Daia_min':      headers.index('Entfernung Daia Haus (min)'),
        'Andratx_km':    headers.index('Entfernung Andratx (km)'),
        'Andratx_min':   headers.index('Entfernung Andratx (min)'),
        'Ses_km':        headers.index('Entfernung Ses Salines (km)'),
        'Ses_min':       headers.index('Entfernung Ses Salines (min)'),
        'Erreichbarkeit': headers.index('Erreichbarkeit (0-100)'),
    }

    # Geocode reference points once
    print("Geocoding Referenzpunkte...", flush=True)
    ref_coords = {}
    for name, addr in REFS.items():
        lat, lng = geocode(addr)
        ref_coords[name] = (lat, lng)
        print(f"  {name}: {lat:.4f}, {lng:.4f}")
    dest_list = [f"{ref_coords[k][0]},{ref_coords[k][1]}" for k in ['Flughafen','Daia','Andratx','Ses Salines']]

    # Process each missing entry
    updated = 0
    for row in ws.iter_rows(min_row=2):
        nr = row[headers.index('Ordnungsnummer')].value
        if not nr or int(nr) not in KNOWN_LOCS: continue
        nr = int(nr)

        dist_val = row[cols['Flughafen_km']].value
        if dist_val: continue  # already filled

        address = KNOWN_LOCS[nr]
        display = DISPLAY_LOCS.get(nr, address.split(',')[0])

        print(f"Nr.{nr:3d} {display:<25} geocoding...", end=' ', flush=True)
        lat, lng = geocode(address)
        if not lat:
            print("❌ geocode fehlgeschlagen")
            continue

        dists = get_distances(lat, lng, dest_list)
        if len(dists) < 4 or not dists[0][0]:
            print("❌ Entfernungen fehlgeschlagen")
            continue

        # Write location
        if not row[loc_idx].value or row[loc_idx].value == 'Mallorca':
            row[loc_idx].value = display

        # Write distances
        row[cols['Flughafen_km']].value = dists[0][0]
        row[cols['Flughafen_min']].value = dists[0][1]
        row[cols['Daia_km']].value = dists[1][0]
        row[cols['Daia_min']].value = dists[1][1]
        row[cols['Andratx_km']].value = dists[2][0]
        row[cols['Andratx_min']].value = dists[2][1]
        row[cols['Ses_km']].value = dists[3][0]
        row[cols['Ses_min']].value = dists[3][1]

        # Erreichbarkeit: based on airport distance (max 90min = 0, 0min = 100)
        airport_min = dists[0][1] or 60
        erreichbarkeit = max(0, round(100 - (airport_min / 90 * 100)))
        if not row[cols['Erreichbarkeit']].value:
            row[cols['Erreichbarkeit']].value = erreichbarkeit

        print(f"✅ Flughafen {dists[0][0]}km/{dists[0][1]}min | Daia {dists[1][1]}min")
        updated += 1
        time.sleep(0.2)

    wb.save(BASE / 'mallorca-kandidaten-v2.xlsx')
    print(f"\n✅ Fertig — {updated} Einträge aktualisiert")

main()
