"""
Savills Mallorca scraper - Full data extraction
Uses SearchByUrl API with Id_mallorca+Category_Area
Also parses the page HTML for server-side rendered data
"""
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth
import json, time, re, requests
from openpyxl import load_workbook
from datetime import date

# First: Direct API approach
print("=== Savills API Direct Approach ===\n")

session = requests.Session()
session.headers.update({
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json, text/plain, */*',
    'Content-Type': 'application/json',
    'Origin': 'https://search.savills.com',
    'Referer': 'https://search.savills.com/es/en/list/property-for-sale/spain/mallorca',
})

all_properties = []
seen_ids = set()

def extract_property(prop):
    """Extract key property data from Savills API response"""
    ext_id = prop.get('ExternalPropertyID', '')
    desc = prop.get('MetaInformation', {}).get('Description', '')
    
    # Price
    price_info = prop.get('PriceInformation', {})
    price = price_info.get('Price', None) if price_info else None
    currency = price_info.get('Currency', 'EUR') if price_info else 'EUR'
    
    # Size
    res_size = prop.get('ResidentialFloorArea', None)
    
    # Bedrooms
    bedrooms = prop.get('Bedrooms', None)
    bathrooms = prop.get('Bathrooms', None)
    
    # Location
    address = prop.get('AddressText', '')
    city = prop.get('CityTownText', '')
    
    # URL
    canonical = prop.get('MetaInformation', {}).get('CanonicalUrl', '')
    url = f"https://search.savills.com/{canonical}" if canonical else '—'
    
    # Title from description
    title = desc[:100] if desc else f"Savills Mallorca {ext_id}"
    
    return {
        'titel': title[:100],
        'quelle': 'Savills',
        'url': url,
        'preis': price,
        'zimmer': bedrooms,
        'grundstueck': None,
        'wohnflaeche': res_size,
        'ort': f"{city} {address}".strip() or 'Mallorca',
        'extern_id': ext_id
    }

# Try multiple search parameters for Mallorca
search_params = [
    '?SearchList=Id_mallorca+Category_Area&Tenure=GRS_T_B&Currency=EUR&Category=GRS_CAT_RES',
    '?SearchList=Id_1257+Category_Area&Tenure=GRS_T_B&Currency=EUR&Category=GRS_CAT_RES',
    '?SearchList=Id_mallorca+Category_Area&Tenure=GRS_T_B&Currency=EUR&Category=GRS_CAT_RES&SortOrder=SO_PCDD',
    '?SearchList=Id_mallorca+Category_Island&Tenure=GRS_T_B&Currency=EUR&Category=GRS_CAT_RES',
]

best_props = []
best_params = None

for params in search_params:
    try:
        r = session.post(
            'https://livev6-searchapi.savills.com/Data/SearchByUrl',
            json={'url': params},
            timeout=20
        )
        if r.status_code == 200:
            data = r.json()
            results = data.get('Results', {})
            ctx = results.get('SearchContext', {})
            criteria = ctx.get('Criteria', {})
            props = results.get('Properties', [])
            total = ctx.get('TotalResults', len(props))
            search_list = criteria.get('SearchList', [])
            print(f"Params: {params[:80]}")
            print(f"  SearchList: {search_list}")
            print(f"  Properties: {len(props)} | Total: {total}")
            if props:
                first = props[0]
                print(f"  First: {first.get('ExternalPropertyID')} | {first.get('MetaInformation',{}).get('Description','')[:60]}")
            if len(props) > len(best_props):
                best_props = props
                best_params = params
    except Exception as e:
        print(f"Error for {params[:50]}: {e}")

print(f"\nBest params: {best_params}")
print(f"Best count: {len(best_props)}")

# Also try the page-based approach (HTML parsing)
print("\n=== Page HTML approach ===")

all_api_data = []

def handle_response(response):
    if 'livev6-searchapi.savills.com/Data/SearchByUrl' in response.url:
        try:
            data = response.json()
            results = data.get('Results', {})
            props = results.get('Properties', [])
            ctx = results.get('SearchContext', {})
            criteria = ctx.get('Criteria', {})
            search_list = criteria.get('SearchList', [])
            print(f"SearchByUrl: {len(props)} props | SearchList: {search_list}")
            all_api_data.extend(props)
        except Exception as e:
            print(f"API parse error: {e}")

with sync_playwright() as p:
    browser = p.chromium.launch(headless=True)
    context = browser.new_context(
        user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    )
    page = context.new_page()
    Stealth().apply_stealth_sync(page)
    page.on('response', handle_response)
    
    url = 'https://search.savills.com/es/en/list/property-for-sale/spain/mallorca'
    print(f"Loading: {url}")
    page.goto(url, wait_until='domcontentloaded', timeout=60000)
    time.sleep(15)  # Give time for all API calls
    
    # Scroll to trigger more
    for y in [500, 1000, 1500, 2000]:
        page.evaluate(f"window.scrollTo(0, {y})")
        time.sleep(2)
    
    time.sleep(5)
    
    # Parse page content for property data
    content = page.content()
    
    # Extract all ExternalPropertyIDs visible in source
    ext_ids = re.findall(r'"ExternalPropertyID":\s*"([^"]+)"', content)
    print(f"\nExternalPropertyIDs in page source: {ext_ids[:20]}")
    
    # Extract full property JSON blocks
    # The page embeds properties in script tags or data attributes
    prop_pattern = re.compile(r'\{[^{}]*"ExternalPropertyID"[^{}]*\}', re.DOTALL)
    prop_matches = prop_pattern.findall(content)
    print(f"Property JSON fragments in source: {len(prop_matches)}")
    
    # Save raw page
    with open('/tmp/savills_mallorca_page.html', 'w') as f:
        f.write(content)
    print("Page saved to /tmp/savills_mallorca_page.html")
    
    browser.close()

# Try to extract full property data from saved HTML
print("\n=== Extracting structured property data from HTML ===")
try:
    with open('/tmp/savills_mallorca_page.html') as f:
        html_content = f.read()
    
    # Find embedded JSON data (next.js style or similar)
    # Look for the big JSON object containing properties
    json_pattern = re.compile(r'window\.__.*?=\s*(\{.+?\})\s*;?\s*</script>', re.DOTALL)
    json_matches = json_pattern.findall(html_content)
    print(f"JSON objects in page: {len(json_matches)}")
    
    for i, jm in enumerate(json_matches[:5]):
        try:
            obj = json.loads(jm)
            print(f"  JSON {i}: keys={list(obj.keys())[:5]}, size={len(jm)}")
        except:
            pass
    
    # Also look for specific property data pattern
    # Savills embeds data in <script type="application/ld+json"> tags
    ld_pattern = re.compile(r'<script type="application/ld\+json">(.+?)</script>', re.DOTALL)
    ld_matches = ld_pattern.findall(html_content)
    print(f"\nLD+JSON blocks: {len(ld_matches)}")
    for i, lm in enumerate(ld_matches[:3]):
        try:
            obj = json.loads(lm)
            print(f"  LD+JSON {i}: type={obj.get('@type')} | {str(obj)[:200]}")
        except Exception as e:
            print(f"  LD+JSON {i}: parse error {e} | {lm[:100]}")
except Exception as e:
    print(f"Error: {e}")

# Combine all found properties
print("\n=== Combining all found Savills properties ===")

# From API data captured via interceptor
for prop in all_api_data:
    ext_id = prop.get('ExternalPropertyID', '')
    if ext_id and ext_id not in seen_ids and 'MAO' in ext_id:
        seen_ids.add(ext_id)
        obj = extract_property(prop)
        all_properties.append(obj)
        print(f"  Added from API intercept: {ext_id}")

# From direct API calls  
for prop in best_props:
    ext_id = prop.get('ExternalPropertyID', '')
    if ext_id and ext_id not in seen_ids:
        seen_ids.add(ext_id)
        obj = extract_property(prop)
        all_properties.append(obj)
        print(f"  Added from direct API: {ext_id}")

print(f"\nTotal Savills Mallorca properties: {len(all_properties)}")

# Save full JSON
with open('/tmp/savills_properties.json', 'w') as f:
    json.dump(all_properties, f, indent=2, default=str)
print("Saved to /tmp/savills_properties.json")

# Save to Excel
if all_properties:
    wb = load_workbook('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
    ws = wb['Mallorca Objekte']
    existing_urls = set(str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[2] != '—')
    
    added = 0
    for obj in all_properties:
        url = obj.get('url', '—')
        if url != '—' and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', ''), 
            obj.get('quelle', ''), 
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            str(date.today()),
            'Neu'
        ])
        if url != '—':
            existing_urls.add(url)
        added += 1
    
    wb.save('/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx')
    print(f"\nSaved {added} new Savills properties to Excel!")
else:
    print("No properties to save.")
