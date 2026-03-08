#!/usr/bin/env python3
"""Scannt alle LivingBlue Übersichtsseiten, sammelt property data + Bilder"""
import json, time, sys
from playwright.sync_api import sync_playwright
import openpyxl

# Unsere 113 LivingBlue Einträge laden
wb = openpyxl.load_workbook('mallorca-kandidaten-v2.xlsx', read_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
lb_entries = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row): continue
    r = dict(zip(headers, row))
    if r.get('Makler') and 'living' in str(r.get('Makler','')).lower():
        lb_entries.append(r)

print(f"Matching gegen {len(lb_entries)} LivingBlue Einträge")

all_cards = []  # alle gefundenen Cards von der Übersicht

with sync_playwright() as p:
    browser = p.chromium.launch(headless=True)
    page = browser.new_page()
    
    for pg in range(1, 89):
        url = f"https://www.livingblue-mallorca.com/de-de/immobilien?pag={pg}"
        try:
            page.goto(url, wait_until='domcontentloaded', timeout=20000)
            page.wait_for_timeout(2000)
        except:
            print(f"Seite {pg} Timeout, weiter...")
            continue
        
        cards = page.evaluate("""() => {
            const results = [];
            const infoLinks = document.querySelectorAll('a[href*="/immobilie/"]');
            const seen = new Set();
            infoLinks.forEach(link => {
                if(seen.has(link.href)) return;
                seen.add(link.href);
                const container = link.closest('li') || link.closest('div.item') || link.parentElement?.parentElement?.parentElement;
                const img = container?.querySelector('img[src*="images.egorealestate"]');
                const text = container?.textContent || '';
                const priceMatch = text.match(/[\\d\\.]+\\s*€|€\\s*[\\d\\.]+/);
                const numId = link.href.match(/\\/(\\d+)$/)?.[1];
                if(img && numId) {
                    results.push({
                        numId,
                        href: link.href,
                        img: img.src,
                        text: text.replace(/\\s+/g,' ').trim().substring(0,200)
                    });
                }
            });
            return results;
        }""")
        
        all_cards.extend(cards)
        print(f"Seite {pg}/88 — {len(cards)} Cards — Total: {len(all_cards)}", flush=True)
    
    browser.close()

with open('lb_all_cards.json', 'w') as f:
    json.dump(all_cards, f, indent=2)

print(f"\nGesamt gesammelt: {len(all_cards)} Cards")
print("Gespeichert: lb_all_cards.json")
