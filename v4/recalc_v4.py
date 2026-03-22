#!/usr/bin/env python3
"""
recalc_v4.py — Neuberechnung Y (Erreichbarkeit) + Z (Score) für alle 469 Einträge
nach den Einstellungen (aktuell) in mallorca-kandidaten-v4.xlsx

Einstellungen (aktuell):
  GEWICHTE:
    Zimmer & Platz:      20%
    Preis-Leistung:      10%
    Gästehaus:           15%
    Erreichbarkeit:       5%
    Grundstück/Garten:   10%
    Vermietlizenz:        5%
    Bewirtschaftung:     10%
    Charme / Ästhetik:   10%
    Renovierung:         15%

  ERREICHBARKEIT intern (Y):
    Flughafen PMI:  Ideal≤15, Akzept≤25, Dealbreaker>60, Gewicht 33.33%
    Daia:           Ideal≤20, Akzept≤40, Dealbreaker>70, Gewicht 33.33%
    Ses Salines:    Ideal≤15, Akzept≤30, Dealbreaker>60, Gewicht 33.33%
    Andratx:        gestrichen (0%)
"""

import openpyxl

XLSX = '../data/mallorca-kandidaten-v4.xlsx'

# Spalten (1-basiert)
COL_NR       = 1
COL_NAME     = 2
COL_CHARME   = 4   # D
COL_ZIMMER   = 5   # E
COL_GS       = 7   # G  Grundstück m²
COL_WF       = 8   # H  Wohnfläche m²
COL_GARTEN   = 9   # I  Garten m²
COL_PMI_MIN  = 12  # L  Flughafen min
COL_DAIA_MIN = 14  # N  Daia min
COL_AND_MIN  = 16  # P  Andratx min
COL_SES_MIN  = 18  # R  Ses Salines min
COL_PREIS    = 19  # S
COL_EPM2     = 20  # T  €/m² bebaut
COL_RENO     = 22  # V  Renovierung
COL_BEWIRT   = 23  # W  Bewirtschaftung
COL_LIZENZ   = 24  # X  Vermietlizenz
COL_Y        = 25  # Y  Erreichbarkeit 0-100
COL_Z        = 26  # Z  Score 0-100
COL_AA       = 27  # AA SortKey
COL_AB       = 28  # AB Rang
COL_GAESTEHAUS = 40 # AN Gästehäuser


# --- Erreichbarkeit (Y) ---
def score_destination(minuten, ideal, akzept, dealbreaker):
    """Gibt Score 0-100 für eine Destination zurück."""
    if minuten is None:
        return None
    if minuten <= ideal:
        return 100
    elif minuten <= akzept:
        # Linear von 100 → 70
        return round(100 - 30 * (minuten - ideal) / (akzept - ideal), 1)
    elif minuten <= dealbreaker:
        # Linear von 70 → 20
        return round(70 - 50 * (minuten - akzept) / (dealbreaker - akzept), 1)
    else:
        return 0  # Dealbreaker überschritten


def calc_erreichbarkeit(pmi_min, daia_min, ses_min):
    """Y = gewichteter Durchschnitt der 3 Destinationen (33.33% je)."""
    scores = []
    weights = []

    configs = [
        (pmi_min,  15, 25, 60,  33.33),  # PMI
        (daia_min, 20, 40, 70,  33.33),  # Daia
        (ses_min,  15, 30, 60,  33.33),  # Ses Salines
    ]

    for minuten, ideal, akzept, dealbreaker, gewicht in configs:
        if minuten is not None:
            s = score_destination(minuten, ideal, akzept, dealbreaker)
            scores.append(s * gewicht)
            weights.append(gewicht)

    if not weights:
        return None

    y = sum(scores) / sum(weights) * 100 / 100
    return round(y, 1)


# --- Hilfsfunktionen Scoring ---
def score_zimmer(zimmer, minimum=5, ideal=8):
    """0-100 basierend auf Zimmeranzahl."""
    if zimmer is None:
        return 50  # Neutral wenn unbekannt
    if zimmer >= ideal:
        return 100
    elif zimmer >= minimum:
        return round(50 + 50 * (zimmer - minimum) / (ideal - minimum), 1)
    else:
        return round(50 * zimmer / minimum, 1)


def score_preis_leistung(epm2, preis):
    """0-100: Option B — 50% Gesamtpreis-Kurve + 50% €/m² Effizienz."""
    # --- Gesamtpreis-Kurve ---
    # ≤4.3M = 100, 4.3-5M = 100→85, 5-6M = 85→0, >6M = 0
    if preis is None or preis <= 0:
        s_preis = 50
    elif preis <= 4_300_000:
        s_preis = 100
    elif preis <= 5_000_000:
        s_preis = round(100 - 15 * (preis - 4_300_000) / 700_000, 1)
    elif preis <= 6_000_000:
        s_preis = round(85 - 85 * (preis - 5_000_000) / 1_000_000, 1)
    else:
        s_preis = 0

    # --- €/m² Effizienz ---
    # ≤3000 = 100, bis 15000 = 0
    if epm2 is None or epm2 <= 0:
        s_epm2 = 50
    elif epm2 <= 3000:
        s_epm2 = 100
    elif epm2 <= 15000:
        s_epm2 = round(100 - 100 * (epm2 - 3000) / 12000, 1)
    else:
        s_epm2 = 0

    return round(s_preis * 0.5 + s_epm2 * 0.5, 1)


def score_gaestehaus(gh):
    """0-100 basierend auf Anzahl Gästehäuser."""
    if gh is None or gh == 0:
        return 0
    elif gh == 1:
        return 70
    else:
        return 100


def score_grundstueck_garten(grundstueck, garten):
    """0-100: Kombination aus Grundstücksgröße und Garten."""
    if grundstueck is None:
        return 50
    # Grundstück: Ideal ≥15000m²
    gs_score = min(100, round(grundstueck / 150, 1))

    # Gartenverhältnis: Garten/Grundstück (pflegeleichter = besser)
    if garten is not None and grundstueck > 0:
        ratio = garten / grundstueck
        # Ideal: Garten ≤60% des Grundstücks (Rest bebaut/Terrasse)
        if ratio <= 0.6:
            ratio_score = 100
        else:
            ratio_score = max(0, round(100 - 80 * (ratio - 0.6) / 0.4, 1))
        return round((gs_score * 0.6 + ratio_score * 0.4), 1)
    return gs_score


def score_vermietlizenz(lizenz):
    """0-100: 100=hat ETV, 50=möglich, 0=nein."""
    if lizenz is None:
        return 0
    return min(100, max(0, lizenz))  # Wert direkt (0, 50 oder 100)


def score_bewirtschaftung(bewirt):
    """0-100 aus Bewirtschaftung 1-5."""
    if bewirt is None:
        return 50
    return round((bewirt - 1) / 4 * 100, 1)


def score_charme(charme):
    """0-100 aus Charme 1-5."""
    if charme is None:
        return 50
    return round((charme - 1) / 4 * 100, 1)


def score_renovierung(reno):
    """0-100 direkt (Renovierung ist bereits 0-100)."""
    if reno is None:
        return 50
    return min(100, max(0, reno))


# --- Z-Score Gesamtberechnung ---
GEWICHTE = {
    'zimmer':       0.20,
    'preis_leis':   0.10,
    'gaestehaus':   0.15,
    'erreichbar':   0.05,
    'grundstueck':  0.10,
    'vermietliz':   0.05,
    'bewirtschaft': 0.10,
    'charme':       0.10,
    'renovierung':  0.15,
}


def calc_z_score(zimmer, epm2, gaestehaus, y, grundstueck, garten, lizenz, bewirt, charme, reno, preis=None):
    """Z = gewichteter Score 0-100."""
    s_zimmer    = score_zimmer(zimmer)
    s_preis     = score_preis_leistung(epm2, preis)
    s_gh        = score_gaestehaus(gaestehaus)
    s_erreich   = y if y is not None else 50
    s_gs        = score_grundstueck_garten(grundstueck, garten)
    s_lizenz    = score_vermietlizenz(lizenz)
    s_bewirt    = score_bewirtschaftung(bewirt)
    s_charme    = score_charme(charme)
    s_reno      = score_renovierung(reno)

    z = (
        s_zimmer    * GEWICHTE['zimmer'] +
        s_preis     * GEWICHTE['preis_leis'] +
        s_gh        * GEWICHTE['gaestehaus'] +
        s_erreich   * GEWICHTE['erreichbar'] +
        s_gs        * GEWICHTE['grundstueck'] +
        s_lizenz    * GEWICHTE['vermietliz'] +
        s_bewirt    * GEWICHTE['bewirtschaft'] +
        s_charme    * GEWICHTE['charme'] +
        s_reno      * GEWICHTE['renovierung']
    )
    return round(z, 2)


# --- Hauptprogramm ---
def main():
    print("Lade Excel...")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb['Mallorca Kandidaten']

    print("Berechne Y und Z für alle Einträge...")
    scores = {}  # row → Z-Score

    for row in range(2, 471):  # Zeile 2–470 (469 Objekte)
        nr = ws.cell(row, COL_NR).value
        if nr is None:
            continue

        # Eingangswerte lesen
        zimmer    = ws.cell(row, COL_ZIMMER).value
        grundst   = ws.cell(row, COL_GS).value
        garten    = ws.cell(row, COL_GARTEN).value
        pmi_min   = ws.cell(row, COL_PMI_MIN).value
        daia_min  = ws.cell(row, COL_DAIA_MIN).value
        ses_min   = ws.cell(row, COL_SES_MIN).value
        preis     = ws.cell(row, COL_PREIS).value
        epm2      = ws.cell(row, COL_EPM2).value
        reno      = ws.cell(row, COL_RENO).value
        bewirt    = ws.cell(row, COL_BEWIRT).value
        lizenz    = ws.cell(row, COL_LIZENZ).value
        charme    = ws.cell(row, COL_CHARME).value
        gh        = ws.cell(row, COL_GAESTEHAUS).value

        # Y berechnen
        y = calc_erreichbarkeit(pmi_min, daia_min, ses_min)
        ws.cell(row, COL_Y).value = y

        # Z berechnen
        z = calc_z_score(zimmer, epm2, gh, y, grundst, garten, lizenz, bewirt, charme, reno, preis)
        # Preis-Dealbreaker: >6 Mio = -30 Punkte hart
        if preis and preis > 6_000_000:
            z = max(0, round(z - 30, 2))
        ws.cell(row, COL_Z).value = z
        scores[row] = z

        if row % 100 == 0:
            print(f"  ...{row} Zeilen verarbeitet")

    # Rang (AB) neu berechnen
    print("Berechne Rang...")
    sorted_rows = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    for rang, (row, _) in enumerate(sorted_rows, 1):
        ws.cell(row, COL_AB).value = rang
        ws.cell(row, COL_AA).value = ws.cell(row, COL_Z).value  # SortKey = Z

    print("Speichere...")
    wb.save(XLSX)

    # Ausgabe Top 20
    print("\n=== TOP 20 nach neuem Ranking ===")
    top20 = sorted_rows[:20]
    for rang, (row, z) in enumerate(top20, 1):
        name = ws.cell(row, COL_NAME).value or '?'
        print(f"  #{rang:2d}  Score {z:.1f}  —  {name[:50]}")

    print(f"\nScore-Range: {min(scores.values()):.1f} – {max(scores.values()):.1f}")
    print(f"Durchschnitt: {sum(scores.values())/len(scores):.1f}")
    print(f"Objekte: {len(scores)}")
    print("\n✅ Fertig!")


if __name__ == '__main__':
    main()
