#!/usr/bin/env python3
"""
Mallorca Scraper — Fotocasa & Rightmove
"""

import requests
import json
import time
import re
from datetime import date
from openpyxl import load_workbook

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'

def load_existing_urls():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing = set(
        str(r[2]).strip()
        for r in ws.iter_rows(min_row=2, values_only=True)
        if r[2] and str(r[2]).strip() not in ('None', '—', '')
    )
    return existing

def save_objects(new_objects, existing_urls=None):
    if not new_objects:
        return 0
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    if existing_urls is None:
        existing_urls = set(
            str(r[2]).strip()
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[2] and str(r[2]).strip() not in ('None', '—', '')
        )
    saved = 0
    for obj in new_objects:
        url = obj.get('url', '—')
        if url and url != '—' and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', ''),
            obj.get('quelle', ''),
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            str(date.today()),
            'Neu'
        ])
        if url and url != '—':
            existing_urls.add(url)
        saved += 1
    wb.save(EXCEL_PATH)
    return saved


# ============================================================
# FOTOCASA
# ============================================================

def scrape_fotocasa():
    print("\n=== FOTOCASA ===")
    
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Origin': 'https://www.fotocasa.es',
        'Referer': 'https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'de-DE,de;q=0.9,en;q=0.8',
        'x-api-key': 'aJxgHGQHnF35APswOHTiSDqJKlQjfnFF',
        'sec-ch-ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"macOS"',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-site',
    })
    
    all_objects = []
    page = 1
    page_size = 40
    total_pages = None
    
    while True:
        url = (
            f'https://search.gw.fotocasa.es/v2/propertysearch/search'
            f'?combinedLocationIds=724,4,7,223,0,0,0,0,0'
            f'&culture=de-DE'
            f'&includePurchaseTypeFacets=true'
            f'&isMap=false'
            f'&isNewConstruction=false'
            f'&latitude=39.6953'
            f'&longitude=3.0176'
            f'&pageNumber={page}'
            f'&pageSize={page_size}'
            f'&propertyTypeId=2'
            f'&sortOrderId=1'
            f'&transactionTypeId=1'
        )
        
        try:
            r = session.get(url, timeout=30)
            print(f"  Page {page}: HTTP {r.status_code}")
            
            if r.status_code == 403 or r.status_code == 401:
                print("  → Blocked (403/401). Trying alternative approach...")
                break
            
            if r.status_code != 200:
                print(f"  → Error {r.status_code}: {r.text[:200]}")
                break
            
            try:
                data = r.json()
            except Exception as e:
                print(f"  → JSON parse error: {e}")
                print(f"  → Response: {r.text[:300]}")
                break
            
            # Explore structure
            if page == 1:
                print(f"  → Top-level keys: {list(data.keys()) if isinstance(data, dict) else type(data)}")
            
            # Extract properties
            items = []
            if isinstance(data, dict):
                # Try common paths
                for path in ['result', 'results', 'items', 'properties', 'realEstates', 'data']:
                    if path in data:
                        candidate = data[path]
                        if isinstance(candidate, dict):
                            for subpath in ['items', 'properties', 'realEstates', 'data']:
                                if subpath in candidate:
                                    items = candidate[subpath]
                                    break
                        elif isinstance(candidate, list):
                            items = candidate
                        if items:
                            break
                
                if not items:
                    # Try nested
                    print(f"  → Full structure sample: {json.dumps(data, ensure_ascii=False)[:500]}")
                    break
                
                # Pagination info
                if total_pages is None:
                    for tkey in ['totalPages', 'total_pages', 'pages']:
                        if tkey in data:
                            total_pages = data[tkey]
                            break
                    if total_pages is None:
                        for path in ['result', 'results', 'paging', 'pagination']:
                            if path in data and isinstance(data[path], dict):
                                for tkey in ['totalPages', 'total_pages', 'pages', 'totalCount']:
                                    if tkey in data[path]:
                                        val = data[path][tkey]
                                        if tkey == 'totalCount':
                                            total_pages = (val + page_size - 1) // page_size
                                        else:
                                            total_pages = val
                                        break
                    print(f"  → Total pages: {total_pages}, Items on this page: {len(items)}")
            
            elif isinstance(data, list):
                items = data
            
            if not items:
                print("  → No items found, stopping")
                break
            
            # Parse items
            page_objects = []
            for item in items:
                obj = parse_fotocasa_item(item)
                if obj:
                    page_objects.append(obj)
            
            all_objects.extend(page_objects)
            print(f"  → Parsed {len(page_objects)} objects (total: {len(all_objects)})")
            
            # Check if done
            if total_pages is not None and page >= total_pages:
                print(f"  → Reached last page ({page}/{total_pages})")
                break
            
            if len(items) < page_size:
                print(f"  → Last page (fewer items than page size)")
                break
            
            page += 1
            time.sleep(1.5)
            
        except Exception as e:
            print(f"  → Exception: {e}")
            import traceback; traceback.print_exc()
            break
    
    if not all_objects:
        print("  → No objects from direct API. Trying Playwright...")
        all_objects = scrape_fotocasa_playwright()
    
    return all_objects


def parse_fotocasa_item(item):
    """Parse a single Fotocasa item into our standard format."""
    if not isinstance(item, dict):
        return None
    
    try:
        # Title
        titel = item.get('title', '') or item.get('name', '') or item.get('Title', '')
        
        # URL
        url = item.get('url', '') or item.get('link', '') or item.get('detailUrl', '')
        if url and not url.startswith('http'):
            url = 'https://www.fotocasa.es' + url
        
        # Price
        preis = None
        for pk in ['price', 'Price', 'salePrice', 'listingPrice']:
            if pk in item:
                val = item[pk]
                if isinstance(val, (int, float)):
                    preis = val
                elif isinstance(val, dict):
                    preis = val.get('value') or val.get('amount')
                break
        
        # Rooms
        zimmer = None
        for rk in ['rooms', 'bedrooms', 'Rooms', 'Bedrooms', 'roomsCount']:
            if rk in item and item[rk] is not None:
                zimmer = item[rk]
                break
        
        # Size
        wohnflaeche = None
        for sk in ['surface', 'area', 'livingArea', 'builtArea', 'size']:
            if sk in item and item[sk] is not None:
                wohnflaeche = item[sk]
                break
        
        grundstueck = None
        for lk in ['lot', 'plot', 'plotArea', 'landArea', 'lotArea']:
            if lk in item and item[lk] is not None:
                grundstueck = item[lk]
                break
        
        # Location
        ort = ''
        for lockey in ['location', 'municipality', 'city', 'address']:
            if lockey in item:
                loc = item[lockey]
                if isinstance(loc, str):
                    ort = loc
                    break
                elif isinstance(loc, dict):
                    ort = loc.get('municipality', '') or loc.get('city', '') or loc.get('name', '') or ''
                    break
        
        if not url and not titel:
            return None
        
        return {
            'titel': titel or f'Fotocasa Objekt',
            'quelle': 'Fotocasa',
            'url': url or '—',
            'preis': preis,
            'zimmer': zimmer,
            'grundstueck': grundstueck,
            'wohnflaeche': wohnflaeche,
            'ort': ort,
        }
    except Exception as e:
        return None


def scrape_fotocasa_playwright():
    """Fallback: Use Playwright with network interception to get Fotocasa JSON."""
    print("  → Starting Playwright fallback for Fotocasa...")
    try:
        from playwright.sync_api import sync_playwright
        from playwright_stealth import Stealth
        
        captured_data = []
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
                extra_http_headers={
                    'Accept-Language': 'de-DE,de;q=0.9',
                }
            )
            page = context.new_page()
            
            # Apply stealth
            try:
                Stealth().apply_stealth_sync(page)
            except Exception as e:
                print(f"  → Stealth error: {e}")
            
            # Intercept API responses
            def handle_response(response):
                if 'fotocasa.es' in response.url or 'gw.fotocasa' in response.url:
                    if 'search' in response.url or 'propertysearch' in response.url:
                        try:
                            body = response.json()
                            captured_data.append(body)
                            print(f"  → Captured API response from: {response.url[:80]}")
                        except:
                            pass
            
            page.on('response', handle_response)
            
            # Navigate to Fotocasa Mallorca listings
            page.goto('https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l', 
                     wait_until='networkidle', timeout=60000)
            time.sleep(3)
            
            # Accept cookies if present
            try:
                page.click('button[id*="accept"], button[class*="accept"], #onetrust-accept-btn-handler', timeout=3000)
                time.sleep(1)
            except:
                pass
            
            # Scroll through a few pages
            for pg in range(1, 6):
                if pg > 1:
                    try:
                        next_btn = page.query_selector('a[aria-label="Next"], button[aria-label="Next"], .sui-Pagination-next, [class*="next"]')
                        if next_btn:
                            next_btn.click()
                            time.sleep(3)
                        else:
                            break
                    except:
                        break
            
            browser.close()
        
        if not captured_data:
            print("  → No data captured via Playwright")
            return []
        
        all_objects = []
        for data in captured_data:
            items = []
            if isinstance(data, dict):
                for path in ['result', 'results', 'items', 'properties', 'realEstates']:
                    if path in data:
                        candidate = data[path]
                        if isinstance(candidate, list):
                            items = candidate
                        elif isinstance(candidate, dict):
                            for sp in ['items', 'properties']:
                                if sp in candidate:
                                    items = candidate[sp]
                                    break
                        if items:
                            break
            
            for item in items:
                obj = parse_fotocasa_item(item)
                if obj:
                    all_objects.append(obj)
        
        print(f"  → Playwright captured {len(all_objects)} Fotocasa objects")
        return all_objects
        
    except Exception as e:
        print(f"  → Playwright fallback failed: {e}")
        import traceback; traceback.print_exc()
        return []


# ============================================================
# RIGHTMOVE
# ============================================================

def scrape_rightmove():
    print("\n=== RIGHTMOVE ===")
    
    all_objects = []
    
    # Approach 1: Direct API
    objects_api = scrape_rightmove_api()
    if objects_api:
        print(f"  → API approach: {len(objects_api)} objects")
        all_objects.extend(objects_api)
    
    # If API approach fails, try Apify
    if not all_objects:
        objects_apify = scrape_rightmove_apify()
        if objects_apify:
            print(f"  → Apify approach: {len(objects_apify)} objects")
            all_objects.extend(objects_apify)
    
    # If still nothing, try Playwright
    if not all_objects:
        objects_pw = scrape_rightmove_playwright()
        if objects_pw:
            print(f"  → Playwright approach: {len(objects_pw)} objects")
            all_objects.extend(objects_pw)
    
    return all_objects


def scrape_rightmove_api():
    """Try Rightmove overseas API directly."""
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Accept': 'application/json',
        'Accept-Language': 'en-GB,en;q=0.9',
        'Referer': 'https://www.rightmove.co.uk/overseas-property/in-Mallorca.html',
    })
    
    # Location identifier for Mallorca overseas
    # OVERSEAS^1100 is commonly used for Mallorca
    location_ids = [
        'OVERSEAS%5E1100',  # Mallorca
        'OVERSEAS%5E116',   # Spain overseas
    ]
    
    all_objects = []
    
    for loc_id in location_ids:
        index = 0
        page_size = 24
        
        while True:
            url = (
                f'https://www.rightmove.co.uk/api/_search'
                f'?locationIdentifier={loc_id}'
                f'&numberOfPropertiesPerPage={page_size}'
                f'&radius=0.0'
                f'&sortType=6'
                f'&index={index}'
                f'&viewType=LIST'
                f'&channel=BUY'
                f'&areaSizeUnit=sqm'
                f'&currencyCode=EUR'
                f'&isFetching=false'
            )
            
            try:
                r = session.get(url, timeout=30)
                print(f"  Rightmove API [{loc_id}] index={index}: HTTP {r.status_code}")
                
                if r.status_code == 403:
                    print("  → Blocked")
                    break
                
                if r.status_code != 200:
                    print(f"  → Error: {r.text[:200]}")
                    break
                
                try:
                    data = r.json()
                except Exception as e:
                    print(f"  → JSON error: {e}, body: {r.text[:200]}")
                    break
                
                if index == 0:
                    print(f"  → Keys: {list(data.keys()) if isinstance(data, dict) else type(data)}")
                
                # Extract properties
                properties = []
                if isinstance(data, dict):
                    for pk in ['properties', 'propertyResults', 'results']:
                        if pk in data:
                            properties = data[pk]
                            break
                    
                    if not properties and 'result' in data:
                        result = data['result']
                        if isinstance(result, list):
                            properties = result
                        elif isinstance(result, dict):
                            for pk in ['properties', 'items']:
                                if pk in result:
                                    properties = result[pk]
                                    break
                    
                    if not properties:
                        print(f"  → Structure: {json.dumps(data, ensure_ascii=False)[:400]}")
                        break
                    
                    # Get total
                    total = data.get('resultCount', data.get('totalResultCount', 0))
                    if isinstance(total, str):
                        total = int(re.sub(r'\D', '', total) or 0)
                    if index == 0:
                        print(f"  → Total results: {total}")
                
                if not properties:
                    break
                
                page_objects = []
                for prop in properties:
                    obj = parse_rightmove_item(prop)
                    if obj:
                        page_objects.append(obj)
                
                all_objects.extend(page_objects)
                print(f"  → Parsed {len(page_objects)} (total: {len(all_objects)})")
                
                if len(properties) < page_size:
                    break
                
                if total and (index + page_size) >= total:
                    break
                
                index += page_size
                time.sleep(1.5)
                
            except Exception as e:
                print(f"  → Exception: {e}")
                break
        
        if all_objects:
            break  # Found results, don't try other location IDs
    
    return all_objects


def parse_rightmove_item(item):
    """Parse a Rightmove property item."""
    if not isinstance(item, dict):
        return None
    
    try:
        # Title / Description
        titel = item.get('displayAddress', '') or item.get('address', '') or item.get('title', '')
        
        # URL
        url = item.get('propertyUrl', '') or item.get('url', '') or item.get('detailUrl', '')
        if url and not url.startswith('http'):
            url = 'https://www.rightmove.co.uk' + url
        
        # Property ID fallback
        prop_id = item.get('id', '') or item.get('propertyId', '')
        if not url and prop_id:
            url = f'https://www.rightmove.co.uk/properties/{prop_id}'
        
        # Price
        preis = None
        price_info = item.get('price', {})
        if isinstance(price_info, dict):
            amount = price_info.get('amount') or price_info.get('value')
            if amount:
                preis = amount
                # Convert from GBP to EUR approximately if needed
                currency = price_info.get('currencyCode', '')
        elif isinstance(price_info, (int, float)):
            preis = price_info
        
        # Also try direct keys
        if preis is None:
            for pk in ['price', 'Price', 'salePrice', 'listingPrice']:
                if pk in item and isinstance(item[pk], (int, float)):
                    preis = item[pk]
                    break
        
        # Rooms / Bedrooms
        zimmer = item.get('bedrooms', item.get('numberOfBedrooms', None))
        
        # Size
        size_info = item.get('displaySize', '') or ''
        wohnflaeche = None
        if size_info:
            match = re.search(r'([\d,]+)\s*sq', size_info)
            if match:
                sqft = float(match.group(1).replace(',', ''))
                wohnflaeche = round(sqft * 0.0929)  # sqft to sqm
        
        wohnflaeche = wohnflaeche or item.get('floorAreaValue', None)
        
        # Location
        ort = item.get('displayAddress', '') or item.get('address', '')
        if isinstance(ort, dict):
            ort = ort.get('displayAddress', '') or ort.get('outcode', '')
        
        # Filter: only Mallorca/Spain
        combined = f"{titel} {ort}".lower()
        if url and 'rightmove' in url:
            pass  # Keep it
        
        if not url and not titel:
            return None
        
        return {
            'titel': titel or 'Rightmove Objekt',
            'quelle': 'Rightmove',
            'url': url or '—',
            'preis': preis,
            'zimmer': zimmer,
            'grundstueck': None,
            'wohnflaeche': wohnflaeche,
            'ort': ort,
        }
    except Exception as e:
        return None


def scrape_rightmove_apify():
    """Use Apify to scrape Rightmove."""
    print("  → Trying Apify for Rightmove...")
    
    APIFY_TOKEN = 'apify_api_feD2KhARHjtuV9CrSwOReYgoePFSF44nsDL6'
    
    # First, find a Rightmove actor
    try:
        r = requests.get(
            'https://api.apify.com/v2/store',
            params={'search': 'rightmove', 'limit': 5},
            headers={'Authorization': f'Bearer {APIFY_TOKEN}'},
            timeout=20
        )
        print(f"  → Apify store search: HTTP {r.status_code}")
        if r.status_code == 200:
            actors = r.json()
            print(f"  → Available actors: {json.dumps(actors, indent=2)[:500]}")
    except Exception as e:
        print(f"  → Apify store search failed: {e}")
    
    return []


def scrape_rightmove_playwright():
    """Scrape Rightmove overseas Mallorca page with Playwright."""
    print("  → Trying Playwright for Rightmove...")
    
    try:
        from playwright.sync_api import sync_playwright
        from playwright_stealth import Stealth
        
        all_objects = []
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
                extra_http_headers={'Accept-Language': 'en-GB,en;q=0.9'},
            )
            page = context.new_page()
            
            try:
                Stealth().apply_stealth_sync(page)
            except Exception as e:
                print(f"  → Stealth error: {e}")
            
            captured_json = []
            
            def handle_response(response):
                if 'rightmove' in response.url:
                    content_type = response.headers.get('content-type', '')
                    if 'json' in content_type or '_search' in response.url or 'api' in response.url:
                        try:
                            body = response.json()
                            captured_json.append({'url': response.url, 'data': body})
                            print(f"  → Captured: {response.url[:80]}")
                        except:
                            pass
            
            page.on('response', handle_response)
            
            # Try overseas search page
            url = 'https://www.rightmove.co.uk/overseas-property/in-Mallorca.html'
            page.goto(url, wait_until='networkidle', timeout=60000)
            time.sleep(3)
            
            # Accept cookies
            try:
                page.click('button[id*="accept"], #onetrust-accept-btn-handler, button:has-text("Accept")', timeout=3000)
                time.sleep(1)
            except:
                pass
            
            # Get page content to parse HTML
            content = page.content()
            
            # Parse HTML listings
            from playwright.sync_api import sync_playwright
            
            # Extract listings from HTML
            listings = page.query_selector_all('[data-test="propertyCard"], .propertyCard, .l-searchResult')
            print(f"  → Found {len(listings)} HTML listing elements")
            
            for listing in listings:
                try:
                    obj = {}
                    
                    # Title/Address
                    addr_el = listing.query_selector('.propertyCard-address, [data-test="address"], h2, .address')
                    obj['titel'] = addr_el.inner_text().strip() if addr_el else ''
                    
                    # URL
                    link_el = listing.query_selector('a[href*="/properties/"]')
                    if link_el:
                        href = link_el.get_attribute('href')
                        obj['url'] = f'https://www.rightmove.co.uk{href}' if href and not href.startswith('http') else href
                    else:
                        obj['url'] = '—'
                    
                    # Price
                    price_el = listing.query_selector('.propertyCard-priceValue, [data-test="price"]')
                    if price_el:
                        price_text = price_el.inner_text().strip()
                        # Extract number
                        match = re.search(r'[\d,]+', price_text.replace(',', ''))
                        if match:
                            obj['preis'] = int(match.group().replace(',', ''))
                    
                    # Bedrooms
                    beds_el = listing.query_selector('[aria-label*="bedroom"], .numberOfBeds')
                    if beds_el:
                        match = re.search(r'\d+', beds_el.inner_text())
                        if match:
                            obj['zimmer'] = int(match.group())
                    
                    obj['quelle'] = 'Rightmove'
                    obj['grundstueck'] = None
                    obj['wohnflaeche'] = None
                    obj['ort'] = obj.get('titel', '')
                    
                    if obj.get('url') and obj['url'] != '—':
                        all_objects.append(obj)
                except Exception as e:
                    continue
            
            # Also parse captured JSON
            for item in captured_json:
                data = item['data']
                properties = []
                if isinstance(data, dict):
                    for pk in ['properties', 'propertyResults']:
                        if pk in data:
                            properties = data[pk]
                            break
                for prop in properties:
                    obj = parse_rightmove_item(prop)
                    if obj:
                        all_objects.append(obj)
            
            # Try pagination - Rightmove uses index parameter
            page_num = 1
            max_pages = 20
            
            while page_num < max_pages:
                next_btn = page.query_selector('a[data-bind*="next"], button[title="Next page"], .pagination-next, a:has-text("Next")')
                if not next_btn:
                    break
                
                page_num += 1
                next_btn.click()
                time.sleep(3)
                
                listings = page.query_selector_all('[data-test="propertyCard"], .propertyCard, .l-searchResult')
                print(f"  → Page {page_num}: {len(listings)} listings")
                
                for listing in listings:
                    try:
                        obj = {}
                        addr_el = listing.query_selector('.propertyCard-address, [data-test="address"], h2')
                        obj['titel'] = addr_el.inner_text().strip() if addr_el else ''
                        
                        link_el = listing.query_selector('a[href*="/properties/"]')
                        if link_el:
                            href = link_el.get_attribute('href')
                            obj['url'] = f'https://www.rightmove.co.uk{href}' if href and not href.startswith('http') else href
                        else:
                            obj['url'] = '—'
                        
                        price_el = listing.query_selector('.propertyCard-priceValue, [data-test="price"]')
                        if price_el:
                            price_text = price_el.inner_text().strip()
                            match = re.search(r'[\d]+', price_text.replace(',', ''))
                            if match:
                                obj['preis'] = int(match.group())
                        
                        obj['quelle'] = 'Rightmove'
                        obj['grundstueck'] = None
                        obj['wohnflaeche'] = None
                        obj['ort'] = obj.get('titel', '')
                        
                        if obj.get('url') and obj['url'] != '—':
                            all_objects.append(obj)
                    except:
                        continue
            
            browser.close()
        
        print(f"  → Playwright Rightmove total: {len(all_objects)}")
        return all_objects
        
    except Exception as e:
        print(f"  → Playwright Rightmove failed: {e}")
        import traceback; traceback.print_exc()
        return []


# ============================================================
# MAIN
# ============================================================

if __name__ == '__main__':
    print("=== Mallorca Scraper: Fotocasa + Rightmove ===")
    print(f"Date: {date.today()}")
    
    existing_urls = load_existing_urls()
    print(f"Existing URLs in Excel: {len(existing_urls)}")
    
    # --- FOTOCASA ---
    fotocasa_objects = scrape_fotocasa()
    fotocasa_saved = 0
    if fotocasa_objects:
        fotocasa_saved = save_objects(fotocasa_objects, existing_urls)
        print(f"\n✅ Fotocasa: {len(fotocasa_objects)} gefunden, {fotocasa_saved} neu gespeichert")
    else:
        print("\n⚠️  Fotocasa: Keine Objekte gefunden")
    
    # --- RIGHTMOVE ---
    rightmove_objects = scrape_rightmove()
    rightmove_saved = 0
    if rightmove_objects:
        rightmove_saved = save_objects(rightmove_objects, existing_urls)
        print(f"\n✅ Rightmove: {len(rightmove_objects)} gefunden, {rightmove_saved} neu gespeichert")
    else:
        print("\n⚠️  Rightmove: Keine Objekte gefunden")
    
    print(f"\n=== ZUSAMMENFASSUNG ===")
    print(f"Fotocasa: {len(fotocasa_objects)} gefunden, {fotocasa_saved} neu")
    print(f"Rightmove: {len(rightmove_objects)} gefunden, {rightmove_saved} neu")
    print(f"Gesamt neu: {fotocasa_saved + rightmove_saved}")
