#!/usr/bin/env python3
"""
Phase 3 Final Scraper - alle Bankportale und Developer
"""

import requests
from bs4 import BeautifulSoup
import re
import time
import json
from datetime import date
from openpyxl import load_workbook

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept-Language': 'es-ES,es;q=0.9',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}

def parse_price(text):
    for pattern in [
        r'([\d]{1,3}(?:\.[\d]{3})+),(\d{2})\s*€',
        r'([\d]+),(\d{2})\s*€',
        r'€\s*([\d]{1,3}(?:[,\.][\d]{3})+)',
    ]:
        m = re.search(pattern, text)
        if m:
            try:
                if m.lastindex >= 2:
                    return float(m.group(1).replace('.', '') + '.' + m.group(2))
                else:
                    return float(m.group(1).replace(',', '').replace('.', ''))
            except:
                pass
    return None

def parse_zimmer(text):
    for pattern in [r'(\d+)\s*(?:dormitorios?|habitaciones?|dorm\.)', r'(\d+)\s*hab\.']:
        m = re.search(pattern, text, re.I)
        if m:
            try:
                z = int(m.group(1))
                if 1 <= z <= 20:
                    return z
            except:
                pass
    return None

def parse_flaeche(text):
    m = re.search(r'(\d{2,4}(?:[,\.]\d+)?)\s*m[²2]', text)
    if m:
        try:
            return float(m.group(1).replace(',', '.'))
        except:
            pass
    return None

def parse_ort(text):
    places = ['Palma', 'Calvià', 'Andratx', 'Pollença', 'Sóller', 'Deià', 
              'Valldemossa', 'Alcúdia', 'Artà', 'Manacor', 'Inca', 'Llucmajor', 
              'Santanyí', 'Felanitx', 'Muro', 'Petra', 'Santa Margalida', 'Marratxí',
              'Binissalem', 'Campanet', 'Selva', 'Alaró', 'Esporles', 'Bunyola']
    for place in places:
        if place.lower() in text.lower():
            return place
    return 'Mallorca'

# ===================================================================
# SERVIHABITAT
# ===================================================================
def scrape_servihabitat():
    objects = []
    seen = set()
    
    base_urls = [
        'https://www.servihabitat.com/es/venta/vivienda/illesbalears-islademallorca',
        'https://www.servihabitat.com/es/venta/vivienda/illesbalears',
    ]
    
    for base in base_urls:
        for page in range(1, 10):
            url = base if page == 1 else f"{base}?pagina={page}"
            try:
                r = requests.get(url, headers=HEADERS, timeout=30)
                if r.status_code != 200:
                    break
                soup = BeautifulSoup(r.text, 'html.parser')
                
                # Servihabitat property links haben 8-stellige IDs
                links = soup.find_all('a', href=re.compile(r'/es/venta/vivienda[^?]+/\d{7,9}$'))
                
                page_new = 0
                for link in links:
                    href = link.get('href', '')
                    if not href.startswith('http'):
                        href = 'https://www.servihabitat.com' + href
                    
                    if href in seen:
                        continue
                    seen.add(href)
                    page_new += 1
                    
                    parent = link.find_parent(['article', 'li', 'div'])
                    parent_text = parent.get_text(' ', strip=True) if parent else ''
                    
                    # Titel aus dem Link-Text oder Parent
                    titel = link.get_text(strip=True)
                    if len(titel) < 5:
                        # Versuche Überschriften im Parent
                        h = parent.find(['h1','h2','h3','h4','span']) if parent else None
                        if h:
                            titel = h.get_text(strip=True)
                    
                    # Ort aus URL
                    ort_match = re.search(r'islademallorca-([^/]+)/\d', href)
                    ort = ort_match.group(1).replace('-', ' ').title() if ort_match else parse_ort(parent_text)
                    
                    # Wenn der parent_text 'manacor', 'inca' etc hat
                    p_ort = parse_ort(parent_text)
                    if p_ort != 'Mallorca':
                        ort = p_ort
                    
                    preis = parse_price(parent_text)
                    flaeche = parse_flaeche(parent_text)
                    zimmer = parse_zimmer(parent_text)
                    
                    objects.append({
                        'titel': (titel or f'Servihabitat Inmueble')[:80],
                        'quelle': 'Servihabitat',
                        'url': href,
                        'preis': preis,
                        'zimmer': zimmer,
                        'grundstueck': None,
                        'wohnflaeche': flaeche,
                        'ort': ort,
                    })
                
                print(f"  Servihabitat {url[-30:]}: {page_new} neue ({len(objects)} gesamt)")
                
                if page_new == 0:
                    break
                    
                time.sleep(1.5)
                
            except Exception as e:
                print(f"  Servihabitat Fehler: {e}")
                break
    
    return objects

# ===================================================================
# HAYA REAL ESTATE (= Solvia)
# ===================================================================
def scrape_haya_solvia():
    objects = []
    seen = set()
    
    # Haya ist gleich Solvia
    base_urls = [
        ('Haya Real Estate', 'https://www.haya.es/comprar/viviendas/baleares/'),
        ('Solvia', 'https://www.solvia.es/es/comprar/viviendas/balears-illes'),
    ]
    
    for quelle, base_url in base_urls:
        domain = base_url.split('/')[2]
        base_domain = f"https://{domain}"
        
        for page in range(1, 8):
            url = base_url if page == 1 else f"{base_url}?page={page}"
            try:
                r = requests.get(url, headers=HEADERS, timeout=30, allow_redirects=True)
                if r.status_code != 200:
                    break
                
                soup = BeautifulSoup(r.text, 'html.parser')
                
                # Property-Links finden
                # Haya/Solvia: Links zu /comprar/viviendas/<location>/<id>
                links = soup.find_all('a', href=re.compile(r'/(?:comprar|es/comprar)/(?:viviendas?|piso|casa|chalet)/[^?]+/\d+'))
                
                # Alternativ: alle Links zu property detail pages
                if not links:
                    links = soup.find_all('a', href=re.compile(r'property|inmueble|detail|/\d{6,10}'))
                
                page_new = 0
                for link in links:
                    href = link.get('href', '')
                    if not href.startswith('http'):
                        href = base_domain + href
                    
                    if href in seen:
                        continue
                    seen.add(href)
                    page_new += 1
                    
                    parent = link.find_parent(['article', 'li', 'div', 'section'])
                    parent_text = parent.get_text(' ', strip=True) if parent else ''
                    
                    titel = link.get_text(strip=True)
                    if len(titel) < 5:
                        h = parent.find(['h2','h3','h4']) if parent else None
                        titel = h.get_text(strip=True) if h else f'{quelle} Inmueble'
                    
                    objects.append({
                        'titel': titel[:80],
                        'quelle': quelle,
                        'url': href,
                        'preis': parse_price(parent_text),
                        'zimmer': parse_zimmer(parent_text),
                        'grundstueck': None,
                        'wohnflaeche': parse_flaeche(parent_text),
                        'ort': parse_ort(parent_text),
                    })
                
                print(f"  {quelle} p{page}: {page_new} neue ({len(objects)} gesamt)")
                
                # Prüfe ob letzter Link eine nächste Seite hat
                next_link = soup.find('a', string=re.compile(r'siguiente|next|›', re.I))
                if not next_link or page_new == 0:
                    break
                    
                time.sleep(1.5)
                
            except Exception as e:
                print(f"  {quelle} Fehler: {e}")
                break
    
    return objects

# ===================================================================
# TAYLOR WIMPEY ESPAÑA
# ===================================================================
def scrape_taylor_wimpey():
    objects = []
    
    urls = [
        'https://www.taylorwimpeyspain.com/en/locations/mallorca/',
        'https://www.taylorwimpeyspain.com/en/',
    ]
    
    seen = set()
    
    for url in urls:
        try:
            r = requests.get(url, headers=HEADERS, timeout=30)
            print(f"  Taylor Wimpey {url[-40:]}: {r.status_code}")
            if r.status_code != 200:
                continue
            
            soup = BeautifulSoup(r.text, 'html.parser')
            
            # Development-Links
            links = soup.find_all('a', href=True)
            for link in links:
                href = link.get('href', '')
                if not href.startswith('http'):
                    href = 'https://www.taylorwimpeyspain.com' + href
                
                # Nur Mallorca-Developments
                if 'mallorca' not in href.lower():
                    continue
                
                # Keine Duplikate
                if href in seen:
                    continue
                
                # Nur Development-Detail-Seiten
                if not re.search(r'/en/\w+.*/', href):
                    continue
                    
                seen.add(href)
                
                parent = link.find_parent(['article', 'div', 'li', 'section'])
                parent_text = parent.get_text(' ', strip=True) if parent else ''
                
                titel = link.get_text(strip=True)
                if len(titel) < 3:
                    continue
                
                preis = None
                preis_match = re.search(r'(?:from|desde)[^0-9]*([\d,\.]+)', parent_text, re.I)
                if preis_match:
                    try:
                        p_str = preis_match.group(1).replace(',', '')
                        preis = float(p_str)
                        if preis < 100000:
                            preis *= 1000
                    except:
                        pass
                
                zimmer = None
                bed_match = re.search(r'(\d+)\s*(?:bed|dorm)', parent_text, re.I)
                if bed_match:
                    zimmer = int(bed_match.group(1))
                
                objects.append({
                    'titel': titel[:80],
                    'quelle': 'Taylor Wimpey España',
                    'url': href,
                    'preis': preis,
                    'zimmer': zimmer,
                    'grundstueck': None,
                    'wohnflaeche': parse_flaeche(parent_text),
                    'ort': parse_ort(parent_text),
                })
            
            print(f"  Taylor Wimpey Mallorca: {len(objects)} Objekte")
            
            if objects:
                break
                
        except Exception as e:
            print(f"  Taylor Wimpey Fehler: {e}")
    
    # Playwright wenn leer
    if not objects:
        try:
            from playwright.sync_api import sync_playwright
            from playwright_stealth import Stealth
            
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()
                Stealth().apply_stealth_sync(page)
                
                page.goto('https://www.taylorwimpeyspain.com/en/locations/mallorca/', timeout=30000, wait_until='domcontentloaded')
                time.sleep(3)
                
                content = page.content()
                soup = BeautifulSoup(content, 'html.parser')
                
                # Alle Links nach Mallorca-Entwicklungen
                all_links = soup.find_all('a', href=re.compile(r'mallorca', re.I))
                print(f"  PW TW Links: {len(all_links)}")
                
                for link in all_links[:20]:
                    href = link.get('href', '')
                    if not href.startswith('http'):
                        href = 'https://www.taylorwimpeyspain.com' + href
                    
                    if href in seen:
                        continue
                    seen.add(href)
                    
                    parent = link.find_parent(['article', 'div'])
                    parent_text = parent.get_text(' ', strip=True) if parent else ''
                    
                    titel = link.get_text(strip=True)
                    if len(titel) < 3:
                        continue
                    
                    objects.append({
                        'titel': titel[:80],
                        'quelle': 'Taylor Wimpey España',
                        'url': href,
                        'preis': None,
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
                
                browser.close()
        except Exception as e:
            print(f"  TW Playwright Fehler: {e}")
    
    return objects

# ===================================================================
# VIVES PONS
# ===================================================================
def scrape_vivespons():
    objects = []
    
    urls_to_try = [
        'https://www.vivespons.com/en/properties/',
        'https://www.vivespons.com/es/propiedades/',
        'https://vivespons.com/',
    ]
    
    for url in urls_to_try:
        try:
            r = requests.get(url, headers=HEADERS, timeout=20)
            print(f"  VivesPons {url[-40:]}: {r.status_code}")
            if r.status_code != 200:
                continue
            
            soup = BeautifulSoup(r.text, 'html.parser')
            
            cards = soup.find_all(['article', 'div'], class_=re.compile(r'property|card|listing|development', re.I))
            
            for card in cards[:30]:
                link_tag = card.find('a')
                if not link_tag:
                    continue
                href = link_tag.get('href', '')
                if not href.startswith('http'):
                    href = 'https://www.vivespons.com' + href
                
                card_text = card.get_text(' ', strip=True)
                titel = link_tag.get_text(strip=True)[:80] or card_text[:60]
                if len(titel) < 3:
                    continue
                
                objects.append({
                    'titel': titel,
                    'quelle': 'Vives Pons',
                    'url': href,
                    'preis': parse_price(card_text),
                    'zimmer': parse_zimmer(card_text),
                    'grundstueck': None,
                    'wohnflaeche': parse_flaeche(card_text),
                    'ort': parse_ort(card_text),
                })
            
            print(f"  VivesPons: {len(objects)} Objekte")
            if objects:
                break
                
        except Exception as e:
            print(f"  VivesPons Fehler bei {url}: {e}")
    
    return objects

# ===================================================================
# BARRAU GESTIÓ
# ===================================================================
def scrape_barrau():
    objects = []
    
    for url in ['https://www.barrau.com/en/', 'https://www.barrau.com/']:
        try:
            r = requests.get(url, headers=HEADERS, timeout=20)
            print(f"  Barrau {url}: {r.status_code}")
            if r.status_code != 200:
                continue
            
            soup = BeautifulSoup(r.text, 'html.parser')
            
            # Schaue was auf der Seite ist
            text_sample = soup.get_text(' ', strip=True)[:500]
            print(f"  Barrau Sample: {text_sample}")
            
            links = soup.find_all('a', href=True)
            for link in links:
                href = link.get('href', '')
                if any(kw in href.lower() for kw in ['property', 'inmueble', 'sale', 'venta', 'project']):
                    print(f"  Link: {href[:60]}")
            
            break
        except Exception as e:
            print(f"  Barrau Fehler: {e}")
    
    return objects

# ===================================================================
# WALLAPOP (mit verschiedenen Approaches)
# ===================================================================
def scrape_wallapop():
    objects = []
    
    # Versuche Wallapop mit mehr Headers
    searches = ['mallorca casa', 'mallorca finca', 'mallorca villa', 'mallorca chalet']
    
    for keyword in searches:
        kw_encoded = keyword.replace(' ', '+')
        api_url = f'https://api.wallapop.com/api/v3/general/search?keywords={kw_encoded}&category_ids=200&filters_source=default_filters'
        
        walla_headers = {
            'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148',
            'Accept': 'application/json',
            'Accept-Language': 'es-ES,es;q=0.9',
            'Origin': 'https://es.wallapop.com',
            'Referer': 'https://es.wallapop.com/',
            'X-DeviceOS': '2',  # iOS
        }
        
        try:
            r = requests.get(api_url, headers=walla_headers, timeout=20)
            print(f"  Wallapop '{keyword}': {r.status_code}")
            
            if r.status_code == 200:
                data = r.json()
                # Wallapop Struktur
                items = data.get('search_objects', [])
                if not items and isinstance(data, dict):
                    for key in data.keys():
                        if isinstance(data[key], list) and data[key]:
                            items = data[key]
                            break
                
                print(f"  Items: {len(items)}")
                
                for item in items[:20]:
                    if not isinstance(item, dict):
                        continue
                    
                    preis_data = item.get('price', {})
                    if isinstance(preis_data, dict):
                        preis = float(preis_data.get('amount', 0))
                    else:
                        preis = float(preis_data or 0) if preis_data else None
                    
                    if preis and preis < 50000:
                        continue
                    
                    slug = item.get('web_slug', item.get('url_slug', ''))
                    item_url = f"https://es.wallapop.com/item/{slug}" if slug else ''
                    
                    loc = item.get('location', {})
                    ort = loc.get('city', 'Mallorca') if isinstance(loc, dict) else 'Mallorca'
                    
                    objects.append({
                        'titel': item.get('title', '')[:80],
                        'quelle': 'Wallapop',
                        'url': item_url,
                        'preis': preis,
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': str(ort)[:60],
                    })
                    
                if objects:
                    break
                    
        except Exception as e:
            print(f"  Wallapop Fehler: {e}")
        
        time.sleep(1)
    
    return objects

# ===================================================================
# SAVE TO EXCEL
# ===================================================================
def save_to_excel(new_objects, source_name=""):
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    
    existing_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            existing_urls.add(str(row[2]).strip())
    
    new_count = 0
    skip_count = 0
    
    for obj in new_objects:
        url = obj.get('url', '')
        
        if url and url in existing_urls:
            skip_count += 1
            continue
        
        ws.append([
            obj.get('titel') or '—',
            obj.get('quelle') or source_name or '—',
            url or '—',
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', '—'),
            str(date.today()),
            'Neu',
        ])
        
        if url:
            existing_urls.add(url)
        new_count += 1
    
    wb.save(EXCEL_PATH)
    return new_count, skip_count

# ===================================================================
# MAIN
# ===================================================================
def main():
    results = {}
    
    # === SERVIHABITAT ===
    print("\n=== SERVIHABITAT ===")
    servi = scrape_servihabitat()
    print(f"Servihabitat: {len(servi)} gefunden")
    if servi:
        n, s = save_to_excel(servi)
        print(f"  Gespeichert: {n} neu, {s} Duplikate")
        results['Servihabitat'] = n
    
    # === HAYA / SOLVIA ===
    print("\n=== HAYA / SOLVIA ===")
    haya = scrape_haya_solvia()
    print(f"Haya/Solvia: {len(haya)} gefunden")
    if haya:
        n, s = save_to_excel(haya)
        print(f"  Gespeichert: {n} neu, {s} Duplikate")
        results['Haya/Solvia'] = n
    
    # === TAYLOR WIMPEY ===
    print("\n=== TAYLOR WIMPEY ===")
    tw = scrape_taylor_wimpey()
    print(f"Taylor Wimpey: {len(tw)} gefunden")
    if tw:
        n, s = save_to_excel(tw)
        print(f"  Gespeichert: {n} neu, {s} Duplikate")
        results['Taylor Wimpey'] = n
    
    # === VIVES PONS ===
    print("\n=== VIVES PONS ===")
    vp = scrape_vivespons()
    print(f"Vives Pons: {len(vp)} gefunden")
    if vp:
        n, s = save_to_excel(vp)
        print(f"  Gespeichert: {n} neu, {s} Duplikate")
        results['Vives Pons'] = n
    
    # === BARRAU ===
    print("\n=== BARRAU ===")
    br = scrape_barrau()
    print(f"Barrau: {len(br)} gefunden")
    
    # === WALLAPOP ===
    print("\n=== WALLAPOP ===")
    wl = scrape_wallapop()
    print(f"Wallapop: {len(wl)} gefunden")
    if wl:
        n, s = save_to_excel(wl)
        print(f"  Gespeichert: {n} neu, {s} Duplikate")
        results['Wallapop'] = n
    
    # === ZUSAMMENFASSUNG ===
    print("\n" + "="*60)
    print("ZUSAMMENFASSUNG")
    print("="*60)
    total = 0
    for src, n in results.items():
        print(f"  {src}: {n} neue Objekte")
        total += n
    print(f"\nGesamt neue Objekte gespeichert: {total}")
    
    return total

if __name__ == '__main__':
    main()
