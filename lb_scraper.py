#!/usr/bin/env python3
"""
LivingBlue Scraper: Sammelt alle Property-IDs + Bild-URLs von der Übersichtsseite
und matched sie gegen unsere 113 UUIDs aus mallorca-kandidaten-v2.xlsx
"""
import json, time, subprocess, sys
from playwright.sync_api import sync_playwright

# Unsere 113 LivingBlue UUIDs aus der Liste
import openpyxl
wb = openpyxl.load_workbook('mallorca-kandidaten-v2.xlsx', read_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
lb_map = {}  # uuid -> ordnungsnummer
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row): continue
    r = dict(zip(headers, row))
    if r.get('Makler') and 'living' in str(r.get('Makler','')).lower():
        url = r.get('Link Objekt (URL)', '')
        if url:
            uuid = url.rstrip('/').split('/')[-1]
            lb_map[uuid] = r['Ordnungsnummer']

print(f"Suche nach {len(lb_map)} LivingBlue UUIDs")
found = {}  # uuid -> img_url

with sync_playwright() as p:
    browser = p.chromium.launch(headless=True)
    page = browser.new_page()
    
    for pg in range(1, 89):  # 88 Seiten
        url = f"https://www.livingblue-mallorca.com/de-de/immobilien?pag={pg}"
        page.goto(url, wait_until='networkidle', timeout=30000)
        time.sleep(1)
        
        # INFO-Links + Bilder zusammen extrahieren
        data = page.evaluate("""() => {
            const results = [];
            const infoLinks = document.querySelectorAll('a[href*="/immobilie/"]');
            infoLinks.forEach(link => {
                // Finde das nächste Bild im gleichen Container
                const container = link.closest('li, div.item, div.property, article') || link.parentElement?.parentElement;
                const img = container?.querySelector('img[src*="images.egorealestate"]');
                if(img) {
                    // Extrahiere UUID aus data-attribute oder aus dem Link-Container
                    const uuidLink = container?.querySelector('a[href*="immobilien/"][href*="-11"]');
                    results.push({
                        infoHref: link.href,
                        imgSrc: img.src,
                        uuid: uuidLink?.href?.match(/immobilien\\/([0-9a-f-]{36})/)?.[1] || null
                    });
                }
            });
            return results;
        }""")
        
        if data:
            for item in data:
                if item.get('uuid') and item['uuid'] in lb_map:
                    found[item['uuid']] = item['imgSrc']
                    nr = lb_map[item['uuid']]
                    print(f"✅ Nr.{nr} gefunden: {item['imgSrc'][:60]}")
        
        print(f"Seite {pg}/88 — {len(found)}/{len(lb_map)} gefunden", flush=True)
        
        if len(found) == len(lb_map):
            print("Alle gefunden!")
            break
    
    browser.close()

# Speichern
with open('lb_uuid_to_img.json', 'w') as f:
    json.dump({'found': found, 'lb_map': lb_map}, f, indent=2)

print(f"\nErgebnis: {len(found)}/{len(lb_map)} LivingBlue Bilder gefunden")
print("Gespeichert: lb_uuid_to_img.json")
