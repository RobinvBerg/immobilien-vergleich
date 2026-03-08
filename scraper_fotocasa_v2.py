#!/usr/bin/env python3
"""
Fotocasa Mallorca Scraper - v2
Uses __INITIAL_PROPS__ embedded in HTML pages
"""

import requests
import re
import json
import time
from datetime import date
from openpyxl import load_workbook

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
BASE_URL = 'https://www.fotocasa.es/es/comprar/viviendas/mallorca/todas-las-zonas/l'


def load_existing_urls():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    return set(
        str(r[2]).strip()
        for r in ws.iter_rows(min_row=2, values_only=True)
        if r[2] and str(r[2]).strip() not in ('None', '—', '')
    )


def save_objects(objects):
    if not objects:
        return 0
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing = set(
        str(r[2]).strip()
        for r in ws.iter_rows(min_row=2, values_only=True)
        if r[2] and str(r[2]).strip() not in ('None', '—', '')
    )
    saved = 0
    for obj in objects:
        url = obj.get('url', '—')
        if url and url != '—' and url in existing:
            continue
        ws.append([
            obj.get('titel', ''),
            obj.get('quelle', ''),
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            str(date.today()),
            'Neu'
        ])
        if url and url != '—':
            existing.add(url)
        saved += 1
    wb.save(EXCEL_PATH)
    return saved


def extract_props(html):
    """Extract __INITIAL_PROPS__ from HTML."""
    m = re.search(r'window\.__INITIAL_PROPS__\s*=\s*JSON\.parse\("(.{100,}?)"\);\s*\n', html, re.DOTALL)
    if not m:
        return None
    try:
        raw = m.group(1)
        return json.loads(json.loads('"' + raw + '"'))
    except Exception as e:
        print(f"  JSON parse error: {e}")
        return None


def parse_item(item):
    """Parse a resultsV2.items entry."""
    if not isinstance(item, dict):
        return None
    
    detail_url = item.get('detailUrl', '')
    url = f'https://www.fotocasa.es{detail_url}' if detail_url else '—'
    
    price_info = item.get('price', {})
    preis = price_info.get('amount') if isinstance(price_info, dict) else None
    
    features = item.get('features', {})
    zimmer = features.get('rooms')
    wohnflaeche = features.get('surface')
    
    location = item.get('location', {})
    ort = location.get('locality', '') or location.get('municipality', '') or ''
    ort = ort.strip()
    
    # Title from address/location
    addr = location.get('address', '') or ''
    titel = f"{addr} – {ort}" if addr and ort else (addr or ort or 'Fotocasa Mallorca')
    
    if not url or url == 'https://www.fotocasa.es':
        return None
    
    return {
        'titel': titel[:120],
        'quelle': 'Fotocasa',
        'url': url,
        'preis': preis,
        'zimmer': zimmer,
        'grundstueck': None,
        'wohnflaeche': wohnflaeche,
        'ort': ort,
    }


def scrape_fotocasa():
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'es-ES,es;q=0.9',
    })
    
    all_objects = []
    page = 1
    total_items = None
    page_size = 30
    
    print("Starting Fotocasa scraper...")
    
    while True:
        url = BASE_URL if page == 1 else f'{BASE_URL}?pageNumber={page}'
        
        try:
            r = session.get(url, timeout=30)
            if r.status_code != 200:
                print(f"  Page {page}: HTTP {r.status_code} - stopping")
                break
            
            data = extract_props(r.text)
            if not data:
                print(f"  Page {page}: Could not extract __INITIAL_PROPS__")
                break
            
            # Navigate to resultsV2.items
            initial_search = data.get('initialSearch', {})
            result = initial_search.get('result', {})
            results_v2 = result.get('resultsV2', {})
            items = results_v2.get('items', [])
            
            # Also try realEstates
            if not items:
                items = result.get('realEstates', [])
            
            if total_items is None:
                total_items = results_v2.get('totalItems') or result.get('count', 0)
                page_size_info = results_v2.get('page', {}).get('size', 30)
                total_pages = (total_items + page_size_info - 1) // page_size_info if total_items else None
                print(f"  Total items: {total_items}, page size: {page_size_info}, total pages: ~{total_pages}")
            
            if not items:
                print(f"  Page {page}: No items found - stopping")
                break
            
            page_objects = []
            for item in items:
                obj = parse_item(item)
                if obj:
                    page_objects.append(obj)
            
            all_objects.extend(page_objects)
            print(f"  Page {page}: {len(items)} items, {len(page_objects)} parsed (total: {len(all_objects)})")
            
            # Check if we're done
            if total_items and len(all_objects) >= total_items:
                print(f"  Reached total ({total_items}), stopping")
                break
            
            if len(items) < page_size:
                print(f"  Last page (only {len(items)} items)")
                break
            
            page += 1
            time.sleep(1.2)
            
        except KeyboardInterrupt:
            print("  Interrupted!")
            break
        except Exception as e:
            print(f"  Page {page} error: {e}")
            import traceback; traceback.print_exc()
            break
    
    return all_objects


if __name__ == '__main__':
    print(f"=== Fotocasa Mallorca Scraper v2 ===")
    print(f"Date: {date.today()}")
    
    existing = load_existing_urls()
    print(f"Existing URLs: {len(existing)}")
    
    objects = scrape_fotocasa()
    print(f"\nTotal parsed: {len(objects)}")
    
    if objects:
        saved = save_objects(objects)
        print(f"Saved {saved} new objects to Excel")
    else:
        print("No objects to save")
