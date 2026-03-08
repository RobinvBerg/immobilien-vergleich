
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import statistics
from collections import Counter, defaultdict

GESAMT = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
QUELLEN = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Immobilien_Quellen.xlsx'

C_BG_DARK = "1A1A2E"
C_BG_MID  = "16213E"
C_SUBHDR  = "0D1B2A"
C_TEAL    = "00B4D8"
C_GOLD    = "E2B96F"
C_WHITE   = "FFFFFF"
C_LIGHT   = "E0E0E0"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color=C_WHITE, bold=False, size=11):
    return Font(color=color, bold=bold, size=size)

wb = load_workbook(GESAMT)
ws_obj = wb['Mallorca Objekte']
rows = list(ws_obj.iter_rows(min_row=2, values_only=True))
total = len(rows)

prices, rooms_list, wfl_list, grd_list, orte_list, quellen_list = [], [], [], [], [], []
ort_preise = defaultdict(list)

for r in rows:
    titel, quelle, url, preis, zimmer, grundstueck, wohnflaeche, ort, gefunden, status = r
    if preis is not None:
        try:
            p = float(preis)
            prices.append(p)
            if ort:
                ort_preise[str(ort).strip()].append(p)
        except: pass
    if zimmer is not None:
        try: rooms_list.append(float(zimmer))
        except: pass
    if wohnflaeche is not None:
        try: wfl_list.append(float(wohnflaeche))
        except: pass
    if grundstueck is not None:
        try: grd_list.append(float(grundstueck))
        except: pass
    if ort: orte_list.append(str(ort).strip())
    if quelle: quellen_list.append(str(quelle).strip())

n_preis  = len(prices)
n_zimmer = len(rooms_list)
n_wfl    = len(wfl_list)
n_grd    = len(grd_list)

median_preis = statistics.median(prices) if prices else 0
mean_preis   = statistics.mean(prices)   if prices else 0
min_preis    = min(prices) if prices else 0
max_preis    = max(prices) if prices else 0

zimmer_cnt = Counter()
for z in rooms_list:
    key = int(z) if z <= 9 else 10
    zimmer_cnt[key] += 1

ort_count = Counter(orte_list)
top20_orte = ort_count.most_common(20)
top10_quellen = Counter(quellen_list).most_common(10)

seg = {"<1M":0, "1-2M":0, "2-3M":0, "3-5M":0, "5M+":0}
for p in prices:
    if p < 1_000_000: seg["<1M"] += 1
    elif p < 2_000_000: seg["1-2M"] += 1
    elif p < 3_000_000: seg["2-3M"] += 1
    elif p < 5_000_000: seg["3-5M"] += 1
    else: seg["5M+"] += 1

# ─── Marktübersicht ──────────────────────────────────────────────────────────
if 'Marktübersicht' in wb.sheetnames:
    del wb['Marktübersicht']
ws = wb.create_sheet('Marktübersicht', 1)
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = "0F3460"

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 34
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 2

def bg(ws, row, cols="BCDEF"):
    for col in cols:
        ws[f'{col}{row}'].fill = fill(C_BG_DARK)

def section_hdr(ws, row, label):
    ws.merge_cells(f'B{row}:F{row}')
    c = ws[f'B{row}']
    c.value = f"  {label}"
    c.fill = fill(C_BG_MID)
    c.font = Font(color=C_TEAL, bold=True, size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24
    return row + 1

def col_hdr(ws, row, labels):
    cols = ['B','C','D','E','F']
    for i, lbl in enumerate(labels):
        if i >= len(cols): break
        c = ws[f'{cols[i]}{row}']
        c.value = lbl
        c.fill = fill(C_SUBHDR)
        c.font = Font(color=C_TEAL, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col in cols: ws[f'{col}{row}'].fill = fill(C_SUBHDR)
    ws.row_dimensions[row].height = 18
    return row + 1

def drow(ws, row, label, *vals):
    cols = ['B','C','D','E','F']
    items = [label] + list(vals)
    for i, v in enumerate(items):
        if i >= len(cols): break
        c = ws[f'{cols[i]}{row}']
        c.value = v
        c.fill = fill(C_BG_DARK)
        c.font = Font(color=C_GOLD if i > 0 else C_LIGHT, size=10)
        c.alignment = Alignment(horizontal="center" if i > 0 else "left", vertical="center")
    for col in cols: ws[f'{col}{row}'].fill = fill(C_BG_DARK)
    ws.row_dimensions[row].height = 20
    return row + 1

def spacer(ws, row):
    for col in 'BCDEF':
        ws[f'{col}{row}'].fill = fill(C_BG_DARK)
    ws.row_dimensions[row].height = 8
    return row + 1

r = 1
# Title
ws.merge_cells(f'B{r}:F{r}')
c = ws[f'B{r}']
c.value = "🏝  MALLORCA IMMOBILIENMARKT  ·  MARKTÜBERSICHT"
c.fill = fill(C_BG_DARK)
c.font = Font(color=C_TEAL, bold=True, size=15)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 40
r += 1

ws.merge_cells(f'B{r}:F{r}')
c = ws[f'B{r}']
c.value = "Datenstand: 03.03.2026  |  61 Quellen  |  19.009 Objekte"
c.fill = fill(C_BG_DARK)
c.font = Font(color=C_LIGHT, size=10)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 18
r += 1
r = spacer(ws, r)

# Datenvollständigkeit
r = section_hdr(ws, r, "DATENVOLLSTÄNDIGKEIT")
r = col_hdr(ws, r, ["Merkmal", "Anzahl", "Anteil", "", ""])
r = drow(ws, r, "Gesamtzahl Objekte", f"{total:,}".replace(","," "), "100 %")
r = drow(ws, r, "Mit Preis", f"{n_preis:,}".replace(","," "), f"{n_preis/total*100:.1f} %")
r = drow(ws, r, "Mit Zimmerzahl", f"{n_zimmer:,}".replace(","," "), f"{n_zimmer/total*100:.1f} %")
r = drow(ws, r, "Mit Wohnfläche", f"{n_wfl:,}".replace(","," "), f"{n_wfl/total*100:.1f} %")
r = drow(ws, r, "Mit Grundstücksfläche", f"{n_grd:,}".replace(","," "), f"{n_grd/total*100:.1f} %")
r = spacer(ws, r)

# Preisstatistik
r = section_hdr(ws, r, "PREIS-STATISTIK  (nur Objekte mit Preis)")
r = col_hdr(ws, r, ["Kennzahl", "Wert (€)", "", "", ""])
r = drow(ws, r, "Median-Preis", f"€ {median_preis:,.0f}".replace(","," "))
r = drow(ws, r, "Mittelwert", f"€ {mean_preis:,.0f}".replace(","," "))
r = drow(ws, r, "Minimum", f"€ {min_preis:,.0f}".replace(","," "))
r = drow(ws, r, "Maximum", f"€ {max_preis:,.0f}".replace(","," "))
r = spacer(ws, r)

# Preissegmente
seg_colors = {"<1M":"2ECC71","1-2M":"3498DB","2-3M":"9B59B6","3-5M":"E67E22","5M+":"E74C3C"}
r = section_hdr(ws, r, "PREISSEGMENTE")
r = col_hdr(ws, r, ["Segment", "Anzahl", "Anteil", "", ""])
for seg_name, seg_val in seg.items():
    pct = f"{seg_val/n_preis*100:.1f} %" if n_preis else "—"
    for col in 'BCDEF': ws[f'{col}{r}'].fill = fill(C_BG_DARK)
    ws[f'B{r}'].value = seg_name
    ws[f'B{r}'].fill = fill(C_BG_DARK)
    ws[f'B{r}'].font = Font(color=seg_colors[seg_name], bold=True, size=10)
    ws[f'B{r}'].alignment = Alignment(horizontal="left", vertical="center")
    ws[f'C{r}'].value = f"{seg_val:,}".replace(","," ")
    ws[f'C{r}'].fill = fill(C_BG_DARK)
    ws[f'C{r}'].font = Font(color=C_GOLD, size=10)
    ws[f'C{r}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'D{r}'].value = pct
    ws[f'D{r}'].fill = fill(C_BG_DARK)
    ws[f'D{r}'].font = Font(color=C_LIGHT, size=10)
    ws[f'D{r}'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1
r = spacer(ws, r)

# Zimmer-Verteilung
r = section_hdr(ws, r, "ZIMMER-VERTEILUNG")
r = col_hdr(ws, r, ["Zimmer", "Anzahl", "Anteil", "", ""])
for z in sorted(zimmer_cnt.keys()):
    label = f"{z}+ Zimmer" if z == 10 else f"{z} Zimmer"
    cnt = zimmer_cnt[z]
    pct = f"{cnt/n_zimmer*100:.1f} %" if n_zimmer else "—"
    for col in 'BCDEF': ws[f'{col}{r}'].fill = fill(C_BG_DARK)
    ws[f'B{r}'].value = label; ws[f'B{r}'].fill = fill(C_BG_DARK)
    ws[f'B{r}'].font = Font(color=C_LIGHT, size=10)
    ws[f'B{r}'].alignment = Alignment(horizontal="left", vertical="center")
    ws[f'C{r}'].value = cnt; ws[f'C{r}'].fill = fill(C_BG_DARK)
    ws[f'C{r}'].font = Font(color=C_GOLD, size=10)
    ws[f'C{r}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'D{r}'].value = pct; ws[f'D{r}'].fill = fill(C_BG_DARK)
    ws[f'D{r}'].font = Font(color=C_LIGHT, size=10)
    ws[f'D{r}'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1
r = spacer(ws, r)

# Top 20 Orte
r = section_hdr(ws, r, "TOP 20 ORTE  (nach Anzahl Objekte)")
r = col_hdr(ws, r, ["Ort / Gemeinde", "Anzahl", "Median-Preis", "", ""])
medals = ["🥇","🥈","🥉"]
for i, (ort, cnt) in enumerate(top20_orte):
    med = statistics.median(ort_preise[ort]) if ort_preise.get(ort) else None
    med_str = f"€ {med:,.0f}".replace(","," ") if med else "—"
    prefix = medals[i] if i < 3 else f"{i+1:2d}."
    for col in 'BCDEF': ws[f'{col}{r}'].fill = fill(C_BG_DARK)
    ws[f'B{r}'].value = f"{prefix} {ort}"; ws[f'B{r}'].fill = fill(C_BG_DARK)
    ws[f'B{r}'].font = Font(color=C_WHITE if i < 3 else C_LIGHT, size=10, bold=(i<3))
    ws[f'B{r}'].alignment = Alignment(horizontal="left", vertical="center")
    ws[f'C{r}'].value = cnt; ws[f'C{r}'].fill = fill(C_BG_DARK)
    ws[f'C{r}'].font = Font(color=C_GOLD, size=10)
    ws[f'C{r}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'D{r}'].value = med_str; ws[f'D{r}'].fill = fill(C_BG_DARK)
    ws[f'D{r}'].font = Font(color=C_TEAL, size=10)
    ws[f'D{r}'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1
r = spacer(ws, r)

# Top 10 Quellen
r = section_hdr(ws, r, "TOP 10 QUELLEN  (nach Anzahl Objekte)")
r = col_hdr(ws, r, ["Quelle", "Anzahl", "Anteil", "", ""])
for src, cnt in top10_quellen:
    pct = f"{cnt/total*100:.1f} %"
    for col in 'BCDEF': ws[f'{col}{r}'].fill = fill(C_BG_DARK)
    ws[f'B{r}'].value = src; ws[f'B{r}'].fill = fill(C_BG_DARK)
    ws[f'B{r}'].font = Font(color=C_LIGHT, size=10)
    ws[f'B{r}'].alignment = Alignment(horizontal="left", vertical="center")
    ws[f'C{r}'].value = cnt; ws[f'C{r}'].fill = fill(C_BG_DARK)
    ws[f'C{r}'].font = Font(color=C_GOLD, size=10)
    ws[f'C{r}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'D{r}'].value = pct; ws[f'D{r}'].fill = fill(C_BG_DARK)
    ws[f'D{r}'].font = Font(color=C_LIGHT, size=10)
    ws[f'D{r}'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

# Fill A and G columns dark
for row_i in range(1, r+2):
    ws[f'A{row_i}'].fill = fill(C_BG_DARK)
    ws[f'G{row_i}'].fill = fill(C_BG_DARK)

# ─── Quellen sheet ───────────────────────────────────────────────────────────
wb_q = load_workbook(QUELLEN)
ws_q = wb_q.active

if 'Quellen' in wb.sheetnames:
    del wb['Quellen']
ws_new = wb.create_sheet('Quellen')
ws_new.sheet_properties.tabColor = "0D3B66"

all_q_rows = list(ws_q.iter_rows(values_only=True))
max_col = ws_q.max_column

for r_idx, row_data in enumerate(all_q_rows, start=1):
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws_new.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.fill = fill(C_SUBHDR)
            cell.font = Font(color=C_TEAL, bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.fill = fill("111827")
            cell.font = Font(color=C_LIGHT, size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")

# Auto column widths
for col_cells in ws_new.columns:
    max_len = 0
    col_letter = get_column_letter(col_cells[0].column)
    for cell in col_cells:
        try:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        except: pass
    ws_new.column_dimensions[col_letter].width = min(max_len + 3, 45)

ws_new.freeze_panes = 'A2'
ws_new.row_dimensions[1].height = 22

wb.save(GESAMT)
print("✅ Gespeichert.")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Marktübersicht: {r} Zeilen aufgebaut")
print(f"   Quellen: {len(all_q_rows)} Zeilen, {max_col} Spalten")
print(f"\nStatistiken:")
print(f"   Gesamt: {total:,} Objekte")
print(f"   Median-Preis: € {median_preis:,.0f}")
print(f"   Preis vorhanden: {n_preis:,} ({n_preis/total*100:.1f}%)")
print(f"   Top-Ort: {top20_orte[0][0]} ({top20_orte[0][1]} Objekte)")
