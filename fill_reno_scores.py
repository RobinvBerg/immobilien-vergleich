#!/usr/bin/env python3
"""
Fills missing Reno-Score (0-100) for properties using GPT-4o Vision.
Analyzes bilder/X_main.jpg and estimates renovation state.
"""
import base64, json, os, sys
from pathlib import Path
import openpyxl
import openai

BASE = Path(__file__).parent
OPENAI_KEY = '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'

client = openai.OpenAI(api_key=OPENAI_KEY)

PROMPT = """Du bist Immobilienexperte für Mallorca.
Analysiere dieses Bild einer Immobilie und schätze den Renovierungszustand.

Reno-Score (0-100):
- 0-20: Rohbau / Ruine / totale Sanierung nötig
- 21-40: Stark renovierungsbedürftig
- 41-60: Teilrenoviert / gemischter Zustand
- 61-80: Gut renoviert / zeitgemäß
- 81-100: Neuwertig / top modernisiert / Neubau

Antworte NUR im Format:
SCORE: <Zahl>
GRUND: <1 Satz Begründung>"""

def analyze_image(img_path: Path) -> tuple[int, str]:
    img_bytes = img_path.read_bytes()
    b64 = base64.b64encode(img_bytes).decode()
    
    ext = img_path.suffix.lower()
    mime = 'image/jpeg' if ext in ['.jpg','.jpeg'] else 'image/png' if ext == '.png' else 'image/webp'
    
    resp = client.chat.completions.create(
        model='gpt-4o-mini',
        max_tokens=100,
        messages=[{
            'role': 'user',
            'content': [
                {'type': 'text', 'text': PROMPT},
                {'type': 'image_url', 'image_url': {'url': f'data:{mime};base64,{b64}', 'detail': 'low'}}
            ]
        }]
    )
    
    text = resp.choices[0].message.content.strip()
    score = 50
    grund = text
    
    for line in text.split('\n'):
        if line.startswith('SCORE:'):
            try: score = int(line.split(':')[1].strip())
            except: pass
        elif line.startswith('GRUND:'):
            grund = line.split(':', 1)[1].strip()
    
    return score, grund

def main():
    wb = openpyxl.load_workbook(BASE / 'mallorca-kandidaten-v2.xlsx')
    ws = wb.active
    headers = [c.value for c in ws[1]]
    
    reno_idx = headers.index('Reno-Score (0-100)')
    reno_beg_idx = headers.index('Reno-Begründung')
    
    missing = []
    for row in ws.iter_rows(min_row=2):
        nr = row[headers.index('Ordnungsnummer')].value
        if not nr: continue
        reno = row[reno_idx].value
        komm = str(row[headers.index('Kommentar')].value or '')
        if '⚠️ delisted' in komm: continue
        if not reno or reno == '':
            missing.append((row, int(nr)))
    
    print(f"Fehlende Reno-Scores: {len(missing)}", flush=True)
    
    for i, (row, nr) in enumerate(missing):
        img = BASE / f'bilder/{nr}_main.jpg'
        if not img.exists():
            print(f"[{i+1}/{len(missing)}] Nr.{nr:3d} ⬜ kein Bild")
            continue
        
        try:
            score, grund = analyze_image(img)
            row[reno_idx].value = score
            row[reno_beg_idx].value = grund
            print(f"[{i+1}/{len(missing)}] Nr.{nr:3d} ✅ Score={score:3d} | {grund[:60]}")
        except Exception as e:
            print(f"[{i+1}/{len(missing)}] Nr.{nr:3d} ⚠️  {e}")
        
        # Save every 10
        if (i+1) % 10 == 0:
            wb.save(BASE / 'mallorca-kandidaten-v2.xlsx')
            print(f"  → Zwischenstand gespeichert", flush=True)
    
    wb.save(BASE / 'mallorca-kandidaten-v2.xlsx')
    print(f"\n✅ Fertig — {len(missing)} Reno-Scores eingetragen")

main()
