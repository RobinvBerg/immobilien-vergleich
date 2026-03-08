#!/usr/bin/env python3
"""
Mallorca Playwright Scraper - Phase 1
Uses browser automation for JS-heavy real estate portals.
"""

import time
import re
import json
import asyncio
from datetime import date
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

EXCEL_PATH = '/Users/robin/.openclaw/workspace/mallorca-projekt/Mallorca_Markt_Gesamt.xlsx'
TODAY = str(date.today())
RESULTS_FILE = '/Users/robin/.openclaw/workspace/mallorca-projekt/phase1_playwright_results.json'

def load_existing_data():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Mallorca Objekte']
    existing_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            existing_urls.add(str(row[2]).strip())
    return wb, ws, existing_urls

def save_objects(wb, ws, new_objects, existing_urls):
    added = 0
    for obj in new_objects:
        url = str(obj.get('url', '')).strip()
        if url and url in existing_urls:
            continue
        ws.append([
            obj.get('titel', ''),
            obj.get('quelle', ''),
            url,
            obj.get('preis'),
            obj.get('zimmer'),
            obj.get('grundstueck'),
            obj.get('wohnflaeche'),
            obj.get('ort', ''),
            TODAY,
            'Neu'
        ])
        if url:
            existing_urls.add(url)
        added += 1
    wb.save(EXCEL_PATH)
    return added

def parse_price(text):
    if not text:
        return None
    # Remove non-numeric except dots and commas
    text = str(text).replace('\xa0', '').replace(' ', '')
    # Remove currency symbols
    text = re.sub(r'[€$£]', '', text)
    # Try to find price number
    # Handle formats like "1.500.000" or "1,500,000"
    # Remove thousand separators (. or , when followed by 3 digits)
    text = re.sub(r'[.,](?=\d{3})', '', text)
    # Remove remaining commas/dots that might be decimal
    text = text.split(',')[0].split('.')[0]
    nums = re.findall(r'\d+', text)
    if nums:
        try:
            val = float(''.join(nums[:1]))
            if val > 10000:  # Reasonable property price
                return val
        except:
            pass
    return None

def parse_int(text):
    if not text:
        return None
    m = re.search(r'(\d+)', str(text))
    return int(m.group(1)) if m else None


# ============================================================
# ThinkSpain Scraper
# ============================================================
def scrape_thinkspain(page):
    print("\n=== ThinkSpain ===")
    results = []
    base = "https://www.thinkspain.com"
    
    search_urls = [
        "https://www.thinkspain.com/property-for-sale/majorca/villas",
        "https://www.thinkspain.com/property-for-sale/majorca/fincas-country-houses",
        "https://www.thinkspain.com/property-for-sale/majorca/houses",
    ]
    
    seen_urls = set()
    
    for base_url in search_urls:
        for pg in range(1, 6):
            url = base_url + (f"?page={pg}" if pg > 1 else "")
            try:
                page.goto(url, wait_until='domcontentloaded', timeout=20000)
                time.sleep(2)
                
                # Get page content
                content = page.content()
                
                # Extract property links
                links = page.query_selector_all('a[href*="/property-for-sale/"]')
                
                found_new = 0
                for a in links:
                    try:
                        href = a.get_attribute('href')
                        if not href:
                            continue
                        # Only property detail pages (has numeric ID)
                        if not re.search(r'/property-for-sale/[^/]+/[^/]+-\d+', href):
                            continue
                        full_url = href if href.startswith('http') else base + href
                        if full_url in seen_urls:
                            continue
                        seen_urls.add(full_url)
                        found_new += 1
                        
                        # Try to get parent card for details
                        parent = a.evaluate_handle('el => el.closest("article") || el.closest(".property-card") || el.closest("[class*=property]") || el.parentElement')
                        
                        title = a.inner_text().strip() or 'ThinkSpain Property'
                        price = None
                        rooms = None
                        ort = 'Mallorca'
                        
                        # Try to get more details from parent
                        try:
                            parent_text = parent.inner_text()
                            # Extract price
                            price_match = re.search(r'€\s*([\d\.,]+)', parent_text)
                            if price_match:
                                price = parse_price(price_match.group(1))
                            # Extract beds
                            bed_match = re.search(r'(\d+)\s*bed|(\d+)\s*Bed', parent_text)
                            if bed_match:
                                rooms = int(bed_match.group(1) or bed_match.group(2))
                        except:
                            pass
                        
                        results.append({
                            'titel': title,
                            'quelle': 'ThinkSpain',
                            'url': full_url,
                            'preis': price,
                            'zimmer': rooms,
                            'grundstueck': None,
                            'wohnflaeche': None,
                            'ort': ort,
                        })
                    except Exception as e:
                        continue
                
                print(f"  {base_url.split('/')[-1]} page {pg}: {found_new} new, {len(results)} total")
                if found_new == 0:
                    break
                    
            except PlaywrightTimeout:
                print(f"  Timeout on page {pg}")
                break
            except Exception as e:
                print(f"  Error: {e}")
                break
    
    print(f"  ThinkSpain total: {len(results)}")
    return results


# ============================================================
# Green-Acres Scraper
# ============================================================
def scrape_greenacres(page):
    print("\n=== Green-Acres ===")
    results = []
    base = "https://www.green-acres.es"
    
    search_urls = [
        "https://www.green-acres.es/property-for-sale/majorca",
        "https://www.green-acres.es/house/majorca",
        "https://www.green-acres.es/villa/majorca",
        "https://www.green-acres.es/country-house/majorca",
    ]
    
    seen_urls = set()
    
    for base_url in search_urls:
        for pg in range(1, 6):
            url = base_url + (f"?page={pg}" if pg > 1 else "")
            try:
                page.goto(url, wait_until='networkidle', timeout=25000)
                time.sleep(2)
                
                # Try to find property cards
                # Green-Acres uses various selectors
                cards = page.query_selector_all('article, [class*="PropertyCard"], [class*="property-card"], [class*="listing"]')
                
                if not cards:
                    # Try to find links to property detail pages
                    links = page.query_selector_all('a[href*="/property/"], a[href*="/house/"], a[href*="/villa/"], a[href*="/country-house/"]')
                    cards = links
                
                found_new = 0
                for card in cards:
                    try:
                        if card.get_attribute('href'):
                            a = card
                        else:
                            a = card.query_selector('a[href]')
                        
                        if not a:
                            continue
                        href = a.get_attribute('href')
                        if not href or len(href) < 15:
                            continue
                        full_url = href if href.startswith('http') else base + href
                        if full_url in seen_urls or 'green-acres' not in full_url:
                            continue
                        seen_urls.add(full_url)
                        found_new += 1
                        
                        # Get card text for details
                        try:
                            card_text = card.inner_text()
                        except:
                            card_text = ''
                        
                        title_el = None
                        try:
                            title_el = card.query_selector('h2, h3, [class*="title"]')
                        except:
                            pass
                        title = title_el.inner_text().strip() if title_el else a.inner_text().strip() or 'Green-Acres Property'
                        
                        price = None
                        price_match = re.search(r'€\s*([\d\.\,]+)', card_text)
                        if price_match:
                            price = parse_price(price_match.group(1))
                        
                        rooms = None
                        bed_match = re.search(r'(\d+)\s*(?:bed|rooms?|chamb|hab|Zimmer)', card_text, re.I)
                        if bed_match:
                            rooms = int(bed_match.group(1))
                        
                        results.append({
                            'titel': title[:100],
                            'quelle': 'Green-Acres',
                            'url': full_url,
                            'preis': price,
                            'zimmer': rooms,
                            'grundstueck': None,
                            'wohnflaeche': None,
                            'ort': 'Mallorca',
                        })
                    except Exception as e:
                        continue
                
                print(f"  {base_url.split('/')[-2]}/{base_url.split('/')[-1]} page {pg}: {found_new} new, {len(results)} total")
                if found_new == 0:
                    break
                    
            except PlaywrightTimeout:
                print(f"  Timeout")
                break
            except Exception as e:
                print(f"  Error: {e}")
                break
    
    print(f"  Green-Acres total: {len(results)}")
    return results


# ============================================================
# A Place in the Sun Scraper
# ============================================================
def scrape_aplaceinthesun(page):
    print("\n=== A Place in the Sun ===")
    results = []
    base = "https://www.aplaceinthesun.com"
    
    try:
        page.goto("https://www.aplaceinthesun.com/property/spain/balearic-islands/mallorca", 
                  wait_until='networkidle', timeout=25000)
        time.sleep(3)
        
        # Scroll to load more
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(2)
        
        seen_urls = set()
        
        # Find all property links
        for attempt in range(5):
            links = page.query_selector_all('a[href*="/property/spain/balearic-islands/mallorca/"]')
            
            for a in links:
                try:
                    href = a.get_attribute('href')
                    if not href:
                        continue
                    full_url = href if href.startswith('http') else base + href
                    # Must be a property detail page (longer URL)
                    if len(href) < 50 or full_url in seen_urls:
                        continue
                    if not re.search(r'/property/spain/balearic-islands/mallorca/.{5,}', href):
                        continue
                    seen_urls.add(full_url)
                    
                    parent = a.evaluate_handle('el => el.closest("[class*=card]") || el.closest("article") || el.parentElement')
                    
                    title = ''
                    price = None
                    rooms = None
                    
                    try:
                        title_el = parent.query_selector('h2, h3, [class*=title]')
                        title = title_el.inner_text().strip() if title_el else ''
                    except:
                        pass
                    
                    try:
                        card_text = parent.inner_text()
                        price_m = re.search(r'€\s*([\d\.,]+)', card_text)
                        if price_m:
                            price = parse_price(price_m.group(1))
                        bed_m = re.search(r'(\d+)\s*bed', card_text, re.I)
                        if bed_m:
                            rooms = int(bed_m.group(1))
                    except:
                        pass
                    
                    results.append({
                        'titel': title or 'APTS Property',
                        'quelle': 'A Place in the Sun',
                        'url': full_url,
                        'preis': price,
                        'zimmer': rooms,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
                except:
                    continue
            
            # Try to go to next page
            next_btn = page.query_selector('a[aria-label*="next"], button[aria-label*="next"], [class*="next"]')
            if next_btn:
                try:
                    next_btn.click()
                    time.sleep(3)
                except:
                    break
            else:
                break
            
            print(f"  Attempt {attempt+1}: {len(results)} total")
    
    except Exception as e:
        print(f"  Error: {e}")
    
    print(f"  A Place in the Sun total: {len(results)}")
    return results


# ============================================================
# Habitaclia Scraper  
# ============================================================
def scrape_habitaclia(page):
    print("\n=== Habitaclia ===")
    results = []
    base = "https://english.habitaclia.com"
    
    search_urls = [
        "https://english.habitaclia.com/buy-home-in-mallorca/buscador.htm",
        "https://english.habitaclia.com/homes-province-mallorca.htm",
    ]
    
    seen_urls = set()
    
    for base_url in search_urls:
        try:
            page.goto(base_url, wait_until='networkidle', timeout=25000)
            time.sleep(2)
            
            # Find property links - habitaclia uses .htm property pages
            links = page.query_selector_all('a[href*="habitaclia.com"]')
            if not links:
                links = page.query_selector_all('a[href$=".htm"]')
            
            found = 0
            for a in links:
                try:
                    href = a.get_attribute('href') or ''
                    if not href or 'buscador' in href or len(href) < 30:
                        continue
                    full_url = href if href.startswith('http') else base + href
                    if full_url in seen_urls:
                        continue
                    if not re.search(r'habitaclia\.com/.+\.htm', full_url):
                        continue
                    seen_urls.add(full_url)
                    found += 1
                    
                    parent = a.evaluate_handle('el => el.closest("article") || el.closest("[class*=list]") || el.closest("[class*=card]") || el.parentElement')
                    title = a.inner_text().strip() or 'Habitaclia Property'
                    price = None
                    rooms = None
                    
                    try:
                        card_text = parent.inner_text()
                        price_m = re.search(r'€\s*([\d\.,]+)', card_text)
                        if price_m:
                            price = parse_price(price_m.group(1))
                        bed_m = re.search(r'(\d+)\s*(?:rooms?|hab)', card_text, re.I)
                        if bed_m:
                            rooms = int(bed_m.group(1))
                    except:
                        pass
                    
                    results.append({
                        'titel': title[:100],
                        'quelle': 'Habitaclia',
                        'url': full_url,
                        'preis': price,
                        'zimmer': rooms,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
                except:
                    continue
            
            print(f"  {base_url.split('/')[-1]}: {found} new, {len(results)} total")
            
            # Try pagination
            for pg in range(2, 8):
                try:
                    next_url = base_url.replace('.htm', f'-{pg}.htm') if 'buscador' not in base_url else base_url + f"?pagina={pg}"
                    page.goto(next_url, wait_until='domcontentloaded', timeout=15000)
                    time.sleep(1.5)
                    
                    links = page.query_selector_all('a[href*="habitaclia.com"]')
                    found_pg = 0
                    for a in links:
                        try:
                            href = a.get_attribute('href') or ''
                            full_url = href if href.startswith('http') else base + href
                            if full_url in seen_urls or not re.search(r'habitaclia\.com/.+\.htm', full_url):
                                continue
                            seen_urls.add(full_url)
                            found_pg += 1
                            title = a.inner_text().strip() or 'Habitaclia Property'
                            results.append({
                                'titel': title[:100],
                                'quelle': 'Habitaclia',
                                'url': full_url,
                                'preis': None,
                                'zimmer': None,
                                'grundstueck': None,
                                'wohnflaeche': None,
                                'ort': 'Mallorca',
                            })
                        except:
                            continue
                    
                    print(f"  Page {pg}: {found_pg} new")
                    if found_pg == 0:
                        break
                except:
                    break
                    
        except Exception as e:
            print(f"  Error on {base_url}: {e}")
    
    print(f"  Habitaclia total: {len(results)}")
    return results


# ============================================================
# Yaencontre Scraper
# ============================================================
def scrape_yaencontre(page):
    print("\n=== Yaencontré ===")
    results = []
    base = "https://www.yaencontre.com"
    
    search_urls = [
        "https://www.yaencontre.com/venta/casas/costa-mallorca",
        "https://www.yaencontre.com/venta/chalets/costa-mallorca",
    ]
    
    seen_urls = set()
    
    for base_url in search_urls:
        for pg in range(1, 8):
            url = base_url + (f"/{pg}" if pg > 1 else "")
            try:
                page.goto(url, wait_until='networkidle', timeout=25000)
                time.sleep(2)
                
                # Find property cards/links
                links = page.query_selector_all('a[href*="/pisos/"], a[href*="/casas/"], a[href*="/chalet"], a[href*="/villa/"]')
                if not links:
                    links = page.query_selector_all('[class*="property"] a, [class*="listing"] a, article a')
                
                found_new = 0
                for a in links:
                    try:
                        href = a.get_attribute('href') or ''
                        full_url = href if href.startswith('http') else base + href
                        if full_url in seen_urls or 'yaencontre' not in full_url:
                            continue
                        if len(href) < 20:
                            continue
                        seen_urls.add(full_url)
                        found_new += 1
                        
                        title = a.inner_text().strip() or 'Yaencontre Property'
                        
                        parent = a.evaluate_handle('el => el.closest("article") || el.closest("[class*=card]") || el.parentElement')
                        price = None
                        rooms = None
                        
                        try:
                            card_text = parent.inner_text()
                            price_m = re.search(r'([\d\.,]+)\s*€', card_text) or re.search(r'€\s*([\d\.,]+)', card_text)
                            if price_m:
                                price = parse_price(price_m.group(1))
                            bed_m = re.search(r'(\d+)\s*hab', card_text, re.I)
                            if bed_m:
                                rooms = int(bed_m.group(1))
                        except:
                            pass
                        
                        results.append({
                            'titel': title[:100],
                            'quelle': 'Yaencontré',
                            'url': full_url,
                            'preis': price,
                            'zimmer': rooms,
                            'grundstueck': None,
                            'wohnflaeche': None,
                            'ort': 'Mallorca',
                        })
                    except:
                        continue
                
                print(f"  {base_url.split('/')[-1]} page {pg}: {found_new} new, {len(results)} total")
                if found_new == 0:
                    break
                    
            except PlaywrightTimeout:
                print(f"  Timeout on page {pg}")
                break
            except Exception as e:
                print(f"  Error: {e}")
                break
    
    print(f"  Yaencontré total: {len(results)}")
    return results


# ============================================================
# abc-mallorca Immobilien Scraper
# ============================================================
def scrape_abcmallorca(page):
    print("\n=== abc-mallorca Immobilien ===")
    results = []
    base = "https://immobilien.abc-mallorca.de"
    
    try:
        page.goto("https://immobilien.abc-mallorca.de/", wait_until='networkidle', timeout=25000)
        time.sleep(2)
        
        # Find property listing links
        links = page.query_selector_all('a[href]')
        seen_urls = set()
        
        property_urls = []
        for a in links:
            href = a.get_attribute('href') or ''
            if 'immobilien.abc-mallorca.de' in href or href.startswith('/'):
                full_url = href if href.startswith('http') else base + href
                # Look for property detail pages
                if re.search(r'/(kaufen|mieten|kauf|buy|property|villa|finca|haus)/', full_url, re.I):
                    property_urls.append(full_url)
        
        print(f"  Found {len(property_urls)} potential property URLs")
        
        # Also try the search/listing page
        listing_urls = [
            "https://immobilien.abc-mallorca.de/kaufen/",
            "https://immobilien.abc-mallorca.de/buy/",
        ]
        
        for listing_url in listing_urls:
            try:
                page.goto(listing_url, wait_until='networkidle', timeout=20000)
                time.sleep(2)
                
                links = page.query_selector_all('article a, [class*=property] a, [class*=listing] a')
                for a in links:
                    href = a.get_attribute('href') or ''
                    if len(href) < 15:
                        continue
                    full_url = href if href.startswith('http') else base + href
                    if full_url not in seen_urls and 'abc-mallorca' in full_url:
                        seen_urls.add(full_url)
                        title = a.inner_text().strip()
                        if title:
                            results.append({
                                'titel': title[:100],
                                'quelle': 'abc-mallorca',
                                'url': full_url,
                                'preis': None,
                                'zimmer': None,
                                'grundstueck': None,
                                'wohnflaeche': None,
                                'ort': 'Mallorca',
                            })
            except Exception as e:
                print(f"  Error on {listing_url}: {e}")
        
        # If no results found, try main page differently
        if not results:
            page.goto("https://immobilien.abc-mallorca.de/", wait_until='networkidle', timeout=20000)
            time.sleep(2)
            
            # Get all links that look like properties
            all_links = page.query_selector_all('a[href]')
            for a in all_links:
                href = a.get_attribute('href') or ''
                full_url = href if href.startswith('http') else base + href
                text = a.inner_text().strip()
                if len(href) > 30 and text and len(text) > 10 and full_url not in seen_urls and 'abc-mallorca' in full_url:
                    seen_urls.add(full_url)
                    results.append({
                        'titel': text[:100],
                        'quelle': 'abc-mallorca',
                        'url': full_url,
                        'preis': None,
                        'zimmer': None,
                        'grundstueck': None,
                        'wohnflaeche': None,
                        'ort': 'Mallorca',
                    })
    
    except Exception as e:
        print(f"  Error: {e}")
    
    print(f"  abc-mallorca total: {len(results)}")
    return results


# ============================================================
# Rightmove Direct Playwright Scraper
# ============================================================
def scrape_rightmove(page):
    print("\n=== Rightmove ===")
    results = []
    base = "https://www.rightmove.co.uk"
    
    search_urls = [
        "https://www.rightmove.co.uk/overseas-property/in-Mallorca.html",
        "https://www.rightmove.co.uk/overseas-property/in-Majorca.html",
    ]
    
    seen_urls = set()
    
    for base_url in search_urls:
        try:
            page.goto(base_url, wait_until='networkidle', timeout=30000)
            time.sleep(3)
            
            # Accept cookies if needed
            try:
                cookie_btn = page.query_selector('button[id*="onetrust-accept"], button:has-text("Accept all"), button:has-text("Accept All Cookies")')
                if cookie_btn:
                    cookie_btn.click()
                    time.sleep(1)
            except:
                pass
            
            for pg in range(1, 10):
                # Find property cards
                cards = page.query_selector_all('.l-searchResult, .propertyCard, [class*="propertyCard"], article[id*="property"]')
                
                if not cards:
                    # Try generic property links
                    cards = page.query_selector_all('a[href*="/properties/"]')
                
                found_new = 0
                for card in cards:
                    try:
                        if card.tag_name == 'a':
                            a = card
                        else:
                            a = card.query_selector('a[href*="/properties/"]')
                        
                        if not a:
                            continue
                        href = a.get_attribute('href') or ''
                        if not href or '/properties/' not in href:
                            continue
                        full_url = href if href.startswith('http') else base + href
                        # Remove query params for dedup
                        clean_url = full_url.split('#')[0].split('?')[0]
                        if clean_url in seen_urls:
                            continue
                        seen_urls.add(clean_url)
                        found_new += 1
                        
                        # Get details
                        title = ''
                        price = None
                        rooms = None
                        ort = 'Mallorca'
                        
                        try:
                            card_text = card.inner_text()
                            # Title from heading
                            h = card.query_selector('h2, h3, [class*="title"], [class*="price"]')
                            if h:
                                title = h.inner_text().strip()
                            
                            # Price
                            price_m = re.search(r'€\s*([\d\.,]+)', card_text) or re.search(r'([\d\.,]+)\s*€', card_text)
                            if price_m:
                                price = parse_price(price_m.group(1))
                            
                            # Beds
                            bed_m = re.search(r'(\d+)\s*bed', card_text, re.I)
                            if bed_m:
                                rooms = int(bed_m.group(1))
                            
                            # Location
                            loc_m = re.search(r'Mallorca|Majorca|Palma|Alcudia|Pollensa|Soller|Deia|Andratx', card_text)
                            if loc_m:
                                ort = loc_m.group(0)
                        except:
                            pass
                        
                        results.append({
                            'titel': title or 'Rightmove Mallorca',
                            'quelle': 'Rightmove',
                            'url': full_url,
                            'preis': price,
                            'zimmer': rooms,
                            'grundstueck': None,
                            'wohnflaeche': None,
                            'ort': ort,
                        })
                    except:
                        continue
                
                print(f"  Page {pg}: {found_new} new, {len(results)} total")
                
                # Next page
                next_btn = page.query_selector('a[data-test="pagination-next"], .pagination-next, a[aria-label*="Next"]')
                if next_btn and found_new > 0:
                    next_btn.click()
                    time.sleep(3)
                else:
                    break
                    
        except Exception as e:
            print(f"  Error on {base_url}: {e}")
    
    print(f"  Rightmove total: {len(results)}")
    return results


# ============================================================
# Fotocasa API Scraper (use their internal API)
# ============================================================
def scrape_fotocasa(page):
    print("\n=== Fotocasa (Playwright) ===")
    results = []
    base = "https://www.fotocasa.es"
    
    search_urls = [
        "https://www.fotocasa.es/es/comprar/casas/mallorca/todas-las-zonas/l",
        "https://www.fotocasa.es/es/comprar/viviendas/mallorca/todas-las-zonas/l",
        "https://www.fotocasa.es/es/comprar/fincas-rusticas/mallorca/todas-las-zonas/l",
    ]
    
    seen_urls = set()
    
    # Intercept API calls
    api_responses = []
    
    def handle_response(response):
        if 'fotocasa.es' in response.url and ('search' in response.url.lower() or 'properties' in response.url.lower()):
            try:
                data = response.json()
                if isinstance(data, dict) and 'realEstates' in data:
                    api_responses.append(data['realEstates'])
            except:
                pass
    
    page.on('response', handle_response)
    
    for base_url in search_urls:
        try:
            page.goto(base_url, wait_until='networkidle', timeout=25000)
            time.sleep(3)
            
            # Accept cookies
            try:
                btn = page.query_selector('button[id*="accept"], button:has-text("Aceptar")')
                if btn:
                    btn.click()
                    time.sleep(1)
            except:
                pass
            
            for pg in range(1, 8):
                url = base_url + (f"/{pg}" if pg > 1 else "")
                if pg > 1:
                    page.goto(url, wait_until='networkidle', timeout=20000)
                    time.sleep(2)
                
                # Try to find property links
                links = page.query_selector_all('a[href*="/comprar/"]')
                
                found_new = 0
                for a in links:
                    try:
                        href = a.get_attribute('href') or ''
                        # Property detail URLs have format /es/comprar/.../ID
                        if not re.search(r'/comprar/[^/]+/[^/]+/[^/]+/\d+', href):
                            continue
                        full_url = href if href.startswith('http') else base + href
                        if full_url in seen_urls:
                            continue
                        seen_urls.add(full_url)
                        found_new += 1
                        
                        parent = a.evaluate_handle('el => el.closest("article") || el.closest("[class*=card]") || el.parentElement')
                        title = a.inner_text().strip() or 'Fotocasa Property'
                        price = None
                        rooms = None
                        
                        try:
                            card_text = parent.inner_text()
                            price_m = re.search(r'([\d\.,]+)\s*€', card_text) or re.search(r'€\s*([\d\.,]+)', card_text)
                            if price_m:
                                price = parse_price(price_m.group(1))
                            bed_m = re.search(r'(\d+)\s*hab', card_text, re.I)
                            if bed_m:
                                rooms = int(bed_m.group(1))
                        except:
                            pass
                        
                        results.append({
                            'titel': title[:100],
                            'quelle': 'Fotocasa',
                            'url': full_url,
                            'preis': price,
                            'zimmer': rooms,
                            'grundstueck': None,
                            'wohnflaeche': None,
                            'ort': 'Mallorca',
                        })
                    except:
                        continue
                
                print(f"  {base_url.split('/')[-3]} page {pg}: {found_new} new, {len(results)} total")
                if found_new == 0:
                    break
                    
        except Exception as e:
            print(f"  Error: {e}")
    
    print(f"  Fotocasa total: {len(results)}")
    return results


# ============================================================
# MAIN
# ============================================================
def main():
    print("Loading existing Excel data...")
    wb, ws, existing_urls = load_existing_data()
    print(f"Loaded {len(existing_urls)} existing URLs")
    
    all_results = {}
    total_added = 0
    
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=['--no-sandbox', '--disable-setuid-sandbox', '--disable-dev-shm-usage']
        )
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
            locale='de-DE',
            timezone_id='Europe/Berlin',
            viewport={'width': 1280, 'height': 900}
        )
        page = context.new_page()
        
        scrapers = [
            ('ThinkSpain', scrape_thinkspain),
            ('Green-Acres', scrape_greenacres),
            ('A Place in the Sun', scrape_aplaceinthesun),
            ('Habitaclia', scrape_habitaclia),
            ('Yaencontré', scrape_yaencontre),
            ('abc-mallorca', scrape_abcmallorca),
            ('Rightmove', scrape_rightmove),
            ('Fotocasa', scrape_fotocasa),
        ]
        
        for name, scraper_fn in scrapers:
            try:
                results = scraper_fn(page)
                if results:
                    added = save_objects(wb, ws, results, existing_urls)
                    all_results[name] = {'scraped': len(results), 'added': added}
                    total_added += added
                    print(f"\n  ✓ {name}: {len(results)} scraped, {added} new added")
                else:
                    all_results[name] = {'scraped': 0, 'added': 0}
                    print(f"\n  ✗ {name}: 0 results")
            except Exception as e:
                print(f"\n  ✗ {name}: FAILED - {e}")
                all_results[name] = {'scraped': 0, 'added': 0, 'error': str(e)}
        
        browser.close()
    
    print(f"\n{'='*50}")
    print("PLAYWRIGHT SCRAPING SUMMARY")
    print(f"{'='*50}")
    for name, stats in all_results.items():
        print(f"  {name}: {stats.get('scraped',0)} scraped → {stats.get('added',0)} added")
        if 'error' in stats:
            print(f"    ERROR: {stats['error'][:100]}")
    print(f"\nTotal new objects added: {total_added}")
    
    with open(RESULTS_FILE, 'w') as f:
        json.dump({'summary': all_results, 'total_added': total_added, 'date': TODAY}, f, indent=2)
    
    return all_results, total_added

if __name__ == '__main__':
    main()
